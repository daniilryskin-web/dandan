"""
Тест: режим «по бренду» пишет итоговый XLSX в ТОЧНО ТАКОЙ ЖЕ структуре, как
result.xlsx (режим проверки по запросу) — те же листы и те же столбцы листов
«results»/«Подробности», и плашка «Оригинал» нормализована в оригинал/не указано.

Проверяем конвертацию BrandResult -> main_v39.ResultRow + запись отчёта
(main_v39.ResultStore), не запуская реальный сбор/сеть.

Запуск:  python3 tests/test_brand_unified_report.py
"""
import asyncio
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

RESULTS = []


def check(name, cond):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}")


# Ожидаемые листы и столбцы технического листа result.xlsx (схема ResultRow).
EXPECTED_SHEETS = ["Сводка", "results", "review", "summary", "Подробности"]


def main():
    import main_brand as mb
    import main_v39 as mv

    # Эталон столбцов «results» = поля ResultRow.
    expected_cols = list(mv.ResultRow.__dataclass_fields__.keys())

    # Готовим пару BrandResult (с реестром и без), без сети.
    br1 = mb.BrandResult(
        brand_query="adidas", nm_id=111, product_name="Носки adidas 3 пары",
        brand="adidas", subject="11", price_rub=999.0, seller_name="ИП Иванов",
        supplier_id="123", is_original="ДА", status=mb.STATUS_LINK_COLLECTED,
        registry_urls="https://pub.fsa.gov.ru/rds/declaration/view/123/common",
        registry_hosts="pub.fsa.gov.ru", registry_record_ids="123",
        registry_doc_number="ЕАЭС N RU Д-XX.01", registry_status="Действует",
        registry_product_full="Изделия чулочно-носочные трикотажные",
        registry_manufacturer="Фабрика", product_url="https://www.wildberries.ru/catalog/111/detail.aspx",
        details="ok", worker="w1", checked_at=mb.now_iso(),
    )
    br2 = mb.BrandResult(
        brand_query="adidas", nm_id=222, product_name="Футболка adidas",
        brand="adidas", subject="177", price_rub=0.0, seller_name="",
        supplier_id="", is_original="НЕТ", status=mb.STATUS_NO_DOCS,
        product_url="https://www.wildberries.ru/catalog/222/detail.aspx",
        details="no docs", worker="w2", checked_at=mb.now_iso(),
    )

    # Конвертация без сети: originality по флагу бренд-движнка (card.json не зовём).
    rr1 = mb._brand_result_to_result_row(br1, mv, "оригинал", "OK", 80.0)
    rr2 = mb._brand_result_to_result_row(br2, mv, "не указано", br2.status, 0.0)
    check("конвертация: тип main_v39.ResultRow", type(rr1).__name__ == "ResultRow")
    check("маппинг реестра: certificate_number", rr1.certificate_number == "ЕАЭС N RU Д-XX.01")
    check("маппинг: certificate_product_name", rr1.certificate_product_name.startswith("Изделия чулочно"))
    check("маппинг: registry_host из первого", rr1.registry_host == "pub.fsa.gov.ru")
    check("originality нормализована", rr1.is_original == "оригинал" and rr2.is_original == "не указано")

    # Пишем отчёт через ResultStore (как result.xlsx) и проверяем структуру файла.
    with tempfile.TemporaryDirectory() as d:
        out = Path(d) / "brand_result.xlsx"

        async def _write():
            rs = mv.ResultStore(out, csv_path=None, expiry_warning_days=30, make_report_xlsx=True)
            await rs.add(rr1)
            await rs.add(rr2)
            await rs.save()

        asyncio.run(_write())
        check("XLSX создан", out.exists())

        import openpyxl
        wb = openpyxl.load_workbook(out, read_only=True)
        for sh in EXPECTED_SHEETS:
            check(f"лист «{sh}» присутствует", sh in wb.sheetnames)

        ws = wb["results"]
        hdr = [str(c) if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
        check("столбцы листа «results» = поля ResultRow (как в result.xlsx)", hdr == expected_cols)

        # Лист «Подробности» — русские заголовки, как в result.xlsx.
        wsd = wb["Подробности"]
        dhdr = [str(c) if c is not None else "" for c in next(wsd.iter_rows(values_only=True))]
        for need in ["Технический статус", "Плашка 'Оригинал'", "Реестр (хост)", "Название в реестре"]:
            check(f"«Подробности» содержит столбец «{need}»", need in dhdr)


if __name__ == "__main__":
    main()
    p, t = sum(RESULTS), len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {p}/{t} прошло")
    sys.exit(0 if p == t else 1)
