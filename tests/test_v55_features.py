"""
Тест (v54.5): доработки №5/№6/№7/№8.

  №5 «Причины»   — _classify_reason_v39 + лист «Причины» (агрегат проблемных вердиктов);
  №6 «Продавцы»  — лист «Продавцы» (агрегат по продавцам, % проблемных);
  №7 «Изменения» — снимок прошлого прогона (<output>.prev.json) + лист диффа;
  №8 CardCache   — глобальный кэш карточек 1-го этапа по nm_id (SQLite) + CLI-аргументы.

Запуск:  python3 tests/test_v55_features.py
"""
import sys
import tempfile
import time
import types
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.modules.setdefault("webview", types.ModuleType("webview"))

import main_v39 as m  # noqa: E402
from openpyxl import Workbook  # noqa: E402

RESULTS = []


def check(name, cond):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}")


def mkrow(**kw):
    base = dict(query="q", nm_id=1, product_name="Товар", brand="", subject="",
                product_url="", status="OK")
    base.update(kw)
    return m.ResultRow(**base)


def _sheet_rows(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_classify():
    f = m._classify_reason_v39
    check("№5 классификатор: недействующий документ",
          "не действует" in f(mkrow(status="НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ", details="doc_status=Прекращён")).lower())
    check("№5 классификатор: возраст",
          "возраст" in f(mkrow(status="НЕСООТВЕТСТВИЕ",
                               details="conflicts=карточка детская, а сертификат явно для взрослых")).lower())
    check("№5 классификатор: категории",
          "категори" in f(mkrow(status="НЕСООТВЕТСТВИЕ",
                                details="conflicts=категория карточки ['footwear'] не совпадает с категорией сертификата ['food']")).lower())
    check("№5 классификатор: другая группа",
          "групп" in f(mkrow(status="НЕСООТВЕТСТВИЕ",
                             details="conflicts=сертификат относится к другой группе: ['toys']")).lower())
    check("№5 классификатор: низкое совпадение",
          "низк" in f(mkrow(status="ПРОВЕРИТЬ ВРУЧНУЮ",
                            details="Низкое совпадение без доказанного конфликта; ...")).lower())
    check("№5 классификатор: нет названия",
          "не удалось" in f(mkrow(status="НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА", details="...")).lower())
    check("№5 статус-наборы: OK не проблемный",
          "OK" in m.VERDICT_STATUSES_V39 and "OK" not in m.PROBLEM_STATUSES_V39)
    check("№5 статус-наборы: НЕСООТВЕТСТВИЕ проблемный",
          "НЕСООТВЕТСТВИЕ" in m.PROBLEM_STATUSES_V39)


def test_reasons_sheet():
    rows = [
        mkrow(status="OK"),
        mkrow(status="НЕСООТВЕТСТВИЕ", details="conflicts=сертификат относится к другой группе: ['toys']"),
        mkrow(status="НЕСООТВЕТСТВИЕ", details="conflicts=сертификат относится к другой группе: ['food']"),
        mkrow(status="ПРОВЕРИТЬ ВРУЧНУЮ", details="Низкое совпадение без доказанного конфликта"),
        mkrow(status="НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ", details="doc_status=Прекращён"),
    ]
    wb = Workbook()
    m._build_reasons_sheet_v39(wb, rows)
    check("№5 лист «Причины» создан", "Причины" in wb.sheetnames)
    flat = [str(c) for r in _sheet_rows(wb["Причины"]) for c in r]
    # 4 проблемных строки, "другой группе" встретилось дважды
    check("№5 «Причины»: всего проблемных = 4", any("4" == str(x) for x in flat))
    check("№5 «Причины»: есть строка про другую группу",
          any("другой товарной группе" in x for x in flat))


def test_sellers_sheet():
    rows = [
        mkrow(status="OK", seller_name="ИП Хороший", supplier_id="111"),
        mkrow(status="OK", seller_name="ИП Хороший", supplier_id="111"),
        mkrow(status="НЕСООТВЕТСТВИЕ", seller_name="ООО Плохой", supplier_id="222"),
        mkrow(status="ПРОВЕРИТЬ ВРУЧНУЮ", seller_name="ООО Плохой", supplier_id="222"),
        mkrow(status="OK", seller_name="ООО Плохой", supplier_id="222"),
    ]
    wb = Workbook()
    m._build_sellers_sheet_v39(wb, rows)
    check("№6 лист «Продавцы» создан", "Продавцы" in wb.sheetnames)
    data = _sheet_rows(wb["Продавцы"])
    # первая строка данных — самый проблемный продавец (ООО Плохой: 2 проблемных)
    check("№6 «Продавцы»: проблемный продавец вверху", data[1][0] == "ООО Плохой")
    check("№6 «Продавцы»: посчитан % проблемных",
          isinstance(data[1][-1], (int, float)) and abs(data[1][-1] - round(2 * 100.0 / 3, 1)) < 0.2)
    check("№6 «Продавцы»: хороший продавец 0% проблемных",
          any(r[0] == "ИП Хороший" and r[-1] == 0 for r in data[1:]))


def test_diff_snapshot_cycle():
    with tempfile.TemporaryDirectory() as td:
        xlsx = Path(td) / "result.xlsx"
        store = m.ResultStore(xlsx, None, expiry_warning_days=30, make_report_xlsx=True)
        check("№7 первый прогон: снимок пуст", store._prev_snapshot == {})
        # сохраняем снимок «прошлого прогона»
        prev_rows = [
            mkrow(nm_id=10, status="OK", registry_url="u10", document_status="Действует"),
            mkrow(nm_id=11, status="ПРОВЕРИТЬ ВРУЧНУЮ", registry_url="u11", document_status="Действует"),
            mkrow(nm_id=12, status="OK", registry_url="u12", document_status="Действует"),
        ]
        store._write_prev_snapshot(prev_rows)
        check("№7 снимок записан на диск", store._prev_path.exists())

        # links-only набор НЕ должен затирать снимок (нет вердиктов)
        store._write_prev_snapshot([mkrow(nm_id=99, status=m.STATUS_LINK_COLLECTED, registry_url="x")])
        import json
        with store._prev_path.open(encoding="utf-8") as f:
            after = json.load(f)
        check("№7 links-only прогон не затирает снимок вердиктов", "10|u10" in after)

        # новый прогон читает снимок
        store2 = m.ResultStore(xlsx, None)
        check("№7 второй прогон видит снимок прошлого", "10|u10" in store2._prev_snapshot)

        # текущие результаты: 11 стал OK, 10 стал НЕСООТВЕТСТВИЕ, 13 новый, 12 исчез
        cur_rows = [
            mkrow(nm_id=10, status="НЕСООТВЕТСТВИЕ", registry_url="u10", document_status="Действует"),
            mkrow(nm_id=11, status="OK", registry_url="u11", document_status="Действует"),
            mkrow(nm_id=13, status="OK", registry_url="u13", document_status="Действует"),
        ]
        wb = Workbook()
        m._build_diff_sheet_v39(wb, cur_rows, store2._prev_snapshot)
        flat = "\n".join(str(c) for r in _sheet_rows(wb["Изменения"]) for c in r if c is not None)
        check("№7 дифф: 10 стал НЕСООТВЕТСТВИЕ", "Стало НЕСООТВЕТСТВИЕ" in flat and "10" in flat)
        check("№7 дифф: 11 проблема устранена", "устранена" in flat)
        check("№7 дифф: 13 новая", "Новая" in flat and "13" in flat)
        check("№7 дифф: 12 исчезла", "Исчезла" in flat and "12" in flat)


def test_card_cache():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "card_cache.sqlite"
        c = m.CardCache(p, ttl_days=7.0)
        c.put(12345, m.STATUS_LINK_COLLECTED, "https://pub.fsa.gov.ru/x/baseInfo",
              "pub.fsa.gov.ru", "x", "ЕАЭС RU С-1", "Да", "http_fast_path; cert_json_ok")
        c.commit()
        got = c.get(12345)
        check("№8 CardCache: get возвращает запись", got is not None)
        if got:
            check("№8 CardCache: ссылка сохранена",
                  got["registry_url"].endswith("baseInfo") and got["status"] == m.STATUS_LINK_COLLECTED)
            check("№8 CardCache: плашка docs_verified сохранена", got["docs_verified"] == "Да")
        check("№8 CardCache: неизвестный nm → None", c.get(999) is None)
        c.close()

        # персистентность между «запусками»
        c2 = m.CardCache(p, ttl_days=7.0)
        check("№8 CardCache: после переоткрытия запись на месте", c2.get(12345) is not None)
        c2.close()

        # несвежая запись → None (карточка перепроверится)
        c3 = m.CardCache(p, ttl_days=1e-9)
        time.sleep(0.01)
        check("№8 CardCache: запись старше TTL → None", c3.get(12345) is None)
        c3.close()

        # TTL=0 → бессрочно
        c4 = m.CardCache(p, ttl_days=0)
        check("№8 CardCache: TTL=0 → запись всегда свежая", c4.get(12345) is not None)
        c4.close()

        # смена версии кэша сбрасывает старые записи
        import sqlite3
        con = sqlite3.connect(str(p))
        con.execute("UPDATE meta SET v='OLD' WHERE k='cache_version'")
        con.commit()
        con.close()
        c5 = m.CardCache(p, ttl_days=7.0)
        check("№8 CardCache: смена версии сбрасывает кэш", c5.get(12345) is None)
        c5.close()


def test_cli_and_integration():
    ap = m.build_parser()
    a = ap.parse_args(["--query", "x"])
    check("№8 --card-cache по умолчанию True", a.card_cache is True)
    check("№8 --card-cache-ttl-days по умолчанию 7", a.card_cache_ttl_days == 7.0)
    check("№8 --card-cache-file по умолчанию пуст", a.card_cache_file == "")
    a2 = ap.parse_args(["--query", "x", "--card-cache", "false"])
    check("№8 --card-cache false выключает кэш", a2.card_cache is False)

    src = (Path(__file__).resolve().parents[1] / "main_v39.py").read_text(encoding="utf-8")
    check("№8 кэш читается перед prefetch (seed)", "_card_cache.get(c.nm_id)" in src)
    check("№8 кэш пополняется после prefetch", "_card_cache.put(" in src and "Кэш карточек пополнен" in src)
    check("№8 кэшируется только STATUS_LINK_COLLECTED",
          "got.get(\"status\") != STATUS_LINK_COLLECTED" in src)
    check("№5/6/7 листы строятся в _save_xlsx",
          "_build_reasons_sheet_v39(wb, rows)" in src
          and "_build_sellers_sheet_v39(wb, rows)" in src
          and "_build_diff_sheet_v39(wb, rows" in src)
    check("№7 снимок пишется только на финале с вердиктами",
          "if final and write_xlsx:" in src and "_write_prev_snapshot(rows)" in src)


def main():
    test_classify()
    test_reasons_sheet()
    test_sellers_sheet()
    test_diff_snapshot_cycle()
    test_card_cache()
    test_cli_and_integration()


if __name__ == "__main__":
    main()
    p, t = sum(RESULTS), len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {p}/{t} прошло")
    sys.exit(0 if p == t else 1)
