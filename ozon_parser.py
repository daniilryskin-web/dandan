#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ozon parser - v27 production

Парсер Ozon на базе мобильного API (api.ozon.ru/composer-api.bx/page/json/v2).
CLI-совместим с main_v39.py: принимает те же аргументы, выдаёт тот же формат
CSV/XLSX (поля совместимы с ResultRow из main_v39.py).

Источники:
  - https://github.com/JTJag/ozon-sellers-parser
  - https://gist.github.com/DxDiagDx/710ac65e117bdd45d4dbb3c64d07849c
  - /home/user/workspace/research_apis.md (раздел 1 — Ozon composer-api)
  - /home/user/workspace/references/ozon-sellers-parser/src/routes.ts
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import datetime as dt
import json
import logging
import math
import os
import random
import re
import sys
import time
import traceback
import urllib.parse
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union
from urllib.parse import urlparse

# ---------------------------------------------------------------------------
# Optional heavy dependencies — degraded operation if missing
# ---------------------------------------------------------------------------
try:
    import aiohttp
    AIOHTTP_OK = True
except ImportError:
    AIOHTTP_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_OK = True
except ImportError:
    async_playwright = None
    PlaywrightTimeoutError = TimeoutError
    PLAYWRIGHT_OK = False

try:
    from bs4 import BeautifulSoup
    BS4_OK = True
except ImportError:
    BeautifulSoup = None
    BS4_OK = False

try:
    from curl_cffi import requests as _curl_requests
    CURL_CFFI_OK = True
except ImportError:
    _curl_requests = None
    CURL_CFFI_OK = False

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger("ozon_parser")

# ---------------------------------------------------------------------------
# Constants — API
# ---------------------------------------------------------------------------
OZON_API_BASE = "https://api.ozon.ru/composer-api.bx/page/json/v2"
OZON_WEB_BASE = "https://www.ozon.ru"
APP_VERSION = "2026-06-07-v27.9-ozon"

# Подтверждённые заголовки из JTJag/ozon-sellers-parser + research_apis.md
# Версия 18.47.0 задокументирована в research_apis.md (раздел 1.3)
OZON_HEADERS: Dict[str, str] = {
    "User-Agent": "OzonAndroid/18.47.0 (Android 12; Pixel 6)",
    "x-o3-app-name": "ru.ozon.app.android",
    "x-o3-app-version": "18.47.0",
    "x-o3-device-type": "mobile",
    "x-o3-sample-trace": "false",
    "Accept": "application/json; charset=UTF-8",
    "Accept-Language": "ru-RU,ru;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Content-Type": "application/json",
}

# Альтернативный набор заголовков (из routes.ts-подтверждённого source JTJag)
OZON_HEADERS_ALT: Dict[str, str] = {
    "User-Agent": "ozonapp_android/16.28.1+2401",
    "X-O3-App-Name": "ozonapp_android",
    "X-O3-App-Version": "16.28.1(2401)",
    "X-O3-Device-Type": "mobile",
    "Mobile-Lat": "0",
    "X-O3-Sample-Trace": "false",
    "Accept": "application/json; charset=utf-8",
    "Accept-Encoding": "gzip, deflate",
}

# Ротация User-Agent для снижения fingerprint-риска
OZON_USER_AGENTS: List[str] = [
    "OzonAndroid/18.47.0 (Android 12; Pixel 6)",
    "OzonAndroid/18.40.0 (Android 13; Samsung Galaxy S23)",
    "OzonAndroid/17.15.0 (Android 11; Xiaomi Redmi Note 10)",
    "ozonapp_android/16.28.1+2401",
    "OzonAndroid/18.10.0 (Android 12; OnePlus 9)",
]

# Реестровые хосты — идентично ALLOWED_REGISTRY_HOSTS из main_v39.py
ALLOWED_REGISTRY_HOSTS: Set[str] = {
    "pub.fsa.gov.ru",
    "fsa.gov.ru",
    "swis.trade.kg",
    "trade.kg",
    "belgiss.by",
    "www.belgiss.by",
    "tsouz.belgiss.by",
    "portal.eaeunion.org",
    "eaeunion.org",
    "www.eaeunion.org",
}

BLOCKED_HOST_PARTS: Tuple[str, ...] = (
    "wildberries",
    "wb.ru",
    "ozon.ru",
    "google",
    "yandex",
    "facebook",
    "vk.com",
    "t.me",
    "telegram",
    "doubleclick",
    "adservice",
)

HTTP_URL_RE = re.compile(r"https?://[^\s\"'<>)\]]+", re.IGNORECASE)

# Статусы — совместимы с main_v39.py
STATUS_LINK_COLLECTED = "ССЫЛКА НА РЕЕСТР СОБРАНА"
STATUS_NO_DOCS = "НЕТ ДОКУМЕНТОВ"
STATUS_NO_REGISTRY_LINK = "НЕТ ССЫЛКИ НА РЕЕСТР"
STATUS_TIMEOUT = "ТАЙМАУТ"
STATUS_ERROR = "ОШИБКА"
STATUS_CF_BLOCK = "БЛОКИРОВКА_CLOUDFLARE"
STATUS_NOT_FOUND = "НЕ НАЙДЕН"

# Expiry risk labels — идентичны main_v39.py
EXPIRY_RISK_OK = "Действует"
EXPIRY_RISK_SOON = "Скоро истекает"
EXPIRY_RISK_EXPIRED = "Истёк"
EXPIRY_RISK_UNKNOWN = "Срок не известен"

# ---------------------------------------------------------------------------
# Data classes — совместимы с ResultRow из main_v39.py
# ---------------------------------------------------------------------------

@dataclass
class OzonCard:
    """Карточка товара Ozon (промежуточный объект первого этапа)."""
    sku: str                        # Артикул Ozon (числовой ID)
    product_name: str = ""
    brand: str = ""
    brand_url: str = ""
    category: str = ""              # Хлебные крошки → последний уровень
    source_query: str = ""
    product_url: str = ""           # URL на ozon.ru/product/...
    slug: str = ""                  # Slug из URL карточки
    price_rub: float = 0.0          # Цена до скидки
    sale_price_rub: float = 0.0     # Цена со скидкой
    seller_name: str = ""
    seller_id: str = ""
    seller_url: str = ""
    rating: float = 0.0
    feedbacks: int = 0
    in_stock: int = 0
    images: List[str] = field(default_factory=list)
    characteristics: List[Dict[str, Any]] = field(default_factory=list)
    documents: List[Dict[str, Any]] = field(default_factory=list)  # webProductDocuments
    is_original: str = ""           # "Да" / "Нет" / ""
    original_mark_info: str = ""
    description: str = ""


@dataclass
class OzonResultRow:
    """
    Финальная строка результата — поля ПОЛНОСТЬЮ совместимы с ResultRow из main_v39.py.
    Имена полей идентичны, добавлены только ozon-специфичные extras.
    """
    # ---- ядро совместимости с ResultRow main_v39.py ----
    query: str = ""
    nm_id: int = 0                  # Для Ozon = числовой SKU; для совместимости int
    product_name: str = ""
    brand: str = ""
    subject: str = ""               # = category
    product_url: str = ""
    status: str = ""
    price_rub: float = 0.0
    sale_price_rub: float = 0.0
    seller_name: str = ""
    supplier_id: str = ""           # = seller_id
    rating: float = 0.0
    feedbacks: int = 0
    is_original: str = ""
    registry_url: str = ""
    registry_host: str = ""
    registry_record_id: str = ""
    certificate_number: str = ""
    document_type: str = ""
    document_status: str = ""
    certificate_product_name: str = ""
    document_date_start: str = ""
    document_date_end: str = ""
    applicant_name: str = ""
    applicant_inn: str = ""
    manufacturer_name: str = ""
    tnved: str = ""
    scheme: str = ""
    technical_regulation: str = ""
    score: float = 0.0
    details: str = ""
    worker: str = ""
    checked_at: str = ""
    # ---- Ozon-специфичные дополнительные поля ----
    ozon_sku: str = ""
    ozon_seller_url: str = ""
    ozon_brand_url: str = ""
    ozon_docs_raw: str = ""         # JSON-список документов из webProductDocuments
    ozon_characteristics: str = ""  # JSON-список характеристик
    ozon_original_mark_info: str = ""


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def now_iso() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def str_to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).lower() in ("1", "true", "yes", "y", "on")


def _safe_float(val: Any, default: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _safe_int(val: Any, default: int = 0) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _deep_get(obj: Any, *keys: str, default: Any = None) -> Any:
    """Безопасный глубокий доступ по цепочке ключей."""
    cur = obj
    for k in keys:
        if isinstance(cur, dict):
            cur = cur.get(k)
        elif isinstance(cur, list) and isinstance(k, int):
            try:
                cur = cur[k]
            except IndexError:
                return default
        else:
            return default
        if cur is None:
            return default
    return cur


def _extract_sku_from_url(url: str) -> str:
    """
    Извлекает SKU (числовой артикул) из URL вида:
      /product/nazvanie-tovara-SKU-12345678/
      https://www.ozon.ru/product/nazvanie-tovara-12345678/
    """
    m = re.search(r"-(\d{7,12})(?:/|$|\?)", url)
    if m:
        return m.group(1)
    # Fallback: последовательность цифр в конце пути
    m = re.search(r"(\d{6,12})(?:/|$|\?)", url)
    if m:
        return m.group(1)
    return ""


def _build_product_url(slug: str, sku: str) -> str:
    """Строит URL карточки товара на ozon.ru."""
    if slug:
        return f"https://www.ozon.ru/product/{slug}/"
    if sku:
        return f"https://www.ozon.ru/product/{sku}/"
    return ""


def _build_product_path(slug: str, sku: str) -> str:
    """Строит path для composer-api (параметр url=...)."""
    if slug:
        path = slug if slug.startswith("/") else f"/product/{slug}/"
        if not path.endswith("/"):
            path += "/"
        return path
    if sku:
        return f"/product/{sku}/"
    return ""


def _parse_price(val: Any) -> float:
    """Парсит строку цены вида '1 234 ₽' или '1234.56' → float."""
    if val is None:
        return 0.0
    s = str(val).replace("\u00a0", "").replace(" ", "").replace("₽", "").replace(",", ".").strip()
    m = re.search(r"[\d.]+", s)
    if m:
        return _safe_float(m.group())
    return 0.0


def _is_registry_url(url: str) -> bool:
    """Проверяет, является ли URL ссылкой на реестр сертификации."""
    try:
        parsed = urlparse(url)
        host = parsed.netloc.lower().lstrip("www.")
        # Проверка по точному совпадению хоста
        if host in ALLOWED_REGISTRY_HOSTS:
            return True
        # Проверка по частичному совпадению (субдомены)
        for allowed in ALLOWED_REGISTRY_HOSTS:
            if host.endswith("." + allowed) or host == allowed:
                return True
    except Exception:
        pass
    return False


def _extract_registry_urls_from_text(text: str) -> List[str]:
    """Извлекает все реестровые URLs из произвольного текста."""
    found = HTTP_URL_RE.findall(text)
    result = []
    for u in found:
        # Очистка от trailing punctuation
        u = u.rstrip(".,;:\"'")
        if _is_registry_url(u):
            result.append(u)
    return list(dict.fromkeys(result))  # dedupe сохраняя порядок


def _extract_registry_record_id(url: str) -> str:
    """
    Извлекает ID записи из реестрового URL:
    pub.fsa.gov.ru/rss/certificate/view/12345/... → 12345
    swis.trade.kg/Doc/uuid → uuid
    """
    m = re.search(r"/view/(\d+)/", url)
    if m:
        return m.group(1)
    m = re.search(r"/(\d{5,})", url)
    if m:
        return m.group(1)
    # SWIS UUID
    m = re.search(r"/Doc/([0-9a-fA-F-]{36})", url)
    if m:
        return m.group(1)
    return ""


def _parse_date_loose(s: Any) -> Optional[dt.date]:
    """Парсит дату в форматах DD.MM.YYYY или YYYY-MM-DD."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def _compute_expiry(row: "OzonResultRow", warning_days: int,
                    today: Optional[dt.date] = None) -> Tuple[Optional[int], str]:
    today = today or dt.date.today()
    end_date = _parse_date_loose(getattr(row, "document_date_end", ""))
    if end_date is None:
        return None, EXPIRY_RISK_UNKNOWN
    days_left = (end_date - today).days
    if days_left < 0:
        return days_left, EXPIRY_RISK_EXPIRED
    if warning_days > 0 and days_left <= warning_days:
        return days_left, EXPIRY_RISK_SOON
    return days_left, EXPIRY_RISK_OK


# ---------------------------------------------------------------------------
# Widget state dispatcher — парсинг widgetStates
# ---------------------------------------------------------------------------

def extract_widget_data(widget_key: str, widget_json: Dict[str, Any]) -> Dict[str, Any]:
    """
    Диспетчер по ключу виджета. Возвращает нормализованный dict с извлечёнными
    данными или пустой dict если виджет не распознан.

    Подстроки в ключах (согласно research_apis.md раздел 1.4):
      webSale             → цена, ID, название (cellTrackingInfo.product.*)
      webProductHeading   → title, rating, reviewCount
      webProductBrand     → brand.name, brand.url
      webSeller           → seller.name, seller.id, seller.url
      webProductCharacteristics → characteristics[]
      webProductDocuments → documents[] (КЛЮЧЕВОЙ виджет для сертификатов)
      webOriginalProduct  → hasOriginalMark, originalMarkInfo
      webGallery          → images[]
      webAddToCart        → offers[].{price, stockCount}
      webBreadCrumbs      → breadcrumbs[]
      webDescription      → description
    """
    result: Dict[str, Any] = {}
    k = widget_key.lower()

    # ------------------------------------------------------------------
    # webSale — основные данные о товаре и цене
    # ------------------------------------------------------------------
    if "websale" in k:
        # Первичный источник для цены и ID
        cti = _deep_get(widget_json, "cellTrackingInfo", "product") or {}
        if cti:
            result["sku"] = str(cti.get("id", "") or cti.get("sku", "") or "")
            result["product_name"] = str(cti.get("title", "") or cti.get("name", "") or "")
            # Цена может быть в разных полях — пробуем последовательно
            price_raw = (cti.get("price") or cti.get("originalPrice") or
                         cti.get("priceForCard") or 0)
            final_raw = (cti.get("finalPrice") or cti.get("price") or
                         cti.get("cardPrice") or 0)
            result["price_rub"] = _parse_price(price_raw)
            result["sale_price_rub"] = _parse_price(final_raw)

        # Альтернативные пути для цены (эвристика)
        if not result.get("sale_price_rub"):
            # webSale может содержать price прямо в корне
            for price_key in ("price", "finalPrice", "cardPrice", "salePrice", "currentPrice"):
                v = widget_json.get(price_key)
                if v:
                    result["sale_price_rub"] = _parse_price(v)
                    break

        # Ещё один вариант: items[0] или product прямо в корне
        if not result.get("product_name"):
            items = widget_json.get("items") or []
            if items and isinstance(items, list) and isinstance(items[0], dict):
                first = items[0]
                result["product_name"] = (first.get("title") or first.get("name") or "")
                if not result.get("sku"):
                    result["sku"] = str(first.get("id") or first.get("sku") or "")

    # ------------------------------------------------------------------
    # webProductHeading — заголовок карточки
    # ------------------------------------------------------------------
    elif "webproductheading" in k:
        result["product_name"] = str(widget_json.get("title") or
                                     widget_json.get("name") or "")
        result["rating"] = _safe_float(widget_json.get("rating") or
                                       widget_json.get("reviewRating") or 0)
        result["feedbacks"] = _safe_int(widget_json.get("reviewCount") or
                                        widget_json.get("feedbacks") or
                                        widget_json.get("totalReviewCount") or 0)

    # ------------------------------------------------------------------
    # webProductBrand — бренд
    # ------------------------------------------------------------------
    elif "webproductbrand" in k:
        brand = widget_json.get("brand") or widget_json
        if isinstance(brand, dict):
            result["brand"] = str(brand.get("name") or brand.get("brandName") or "")
            result["brand_url"] = str(brand.get("url") or brand.get("link") or "")
        else:
            # Иногда бренд — просто строка
            result["brand"] = str(brand or "")

    # ------------------------------------------------------------------
    # webSeller — продавец
    # ------------------------------------------------------------------
    elif "webseller" in k:
        seller = widget_json.get("seller") or widget_json
        if isinstance(seller, dict):
            result["seller_name"] = str(seller.get("name") or
                                        seller.get("title") or
                                        seller.get("shopName") or "")
            result["seller_id"] = str(seller.get("id") or
                                      seller.get("sellerId") or "")
            result["seller_url"] = str(seller.get("url") or
                                       seller.get("link") or "")
        # Иногда seller лежит прямо как name/url в корне
        if not result.get("seller_name"):
            result["seller_name"] = str(widget_json.get("name") or
                                        widget_json.get("shopName") or "")
        if not result.get("seller_id"):
            result["seller_id"] = str(widget_json.get("id") or
                                      widget_json.get("sellerId") or "")

    # ------------------------------------------------------------------
    # webProductCharacteristics — характеристики
    # ------------------------------------------------------------------
    elif "webproductcharacteristics" in k:
        chars = widget_json.get("characteristics") or []
        if not chars:
            # Эвристика: ищем любое поле с массивом характеристик
            for k_alt in ("attributes", "features", "specs", "params"):
                v = widget_json.get(k_alt)
                if isinstance(v, list):
                    chars = v
                    break
        result["characteristics"] = chars if isinstance(chars, list) else []

    # ------------------------------------------------------------------
    # webProductDocuments — ГЛАВНЫЙ виджет для сертификатов
    # ------------------------------------------------------------------
    elif "webproductdocuments" in k:
        docs = widget_json.get("documents") or []
        if not docs:
            # Эвристика: ищем вложенные массивы
            for k_alt in ("items", "files", "attachments", "certs", "certificates"):
                v = widget_json.get(k_alt)
                if isinstance(v, list):
                    docs = v
                    break
        result["documents"] = docs if isinstance(docs, list) else []

    # ------------------------------------------------------------------
    # webOriginalProduct — плашка «Оригинал»
    # ------------------------------------------------------------------
    elif "weboriginalproduct" in k or "weboriginal" in k:
        has_orig = widget_json.get("hasOriginalMark")
        if has_orig is None:
            # Эвристика: ищем похожие ключи
            for alt in ("isOriginal", "originalMark", "original", "hasOriginal"):
                v = widget_json.get(alt)
                if v is not None:
                    has_orig = v
                    break
        result["is_original"] = "Да" if has_orig else ("Нет" if has_orig is False else "")
        orig_info = widget_json.get("originalMarkInfo") or widget_json.get("originalInfo") or {}
        if isinstance(orig_info, dict):
            result["original_mark_info"] = str(orig_info.get("text") or
                                               orig_info.get("description") or "")
        elif isinstance(orig_info, str):
            result["original_mark_info"] = orig_info

    # ------------------------------------------------------------------
    # webGallery — изображения
    # ------------------------------------------------------------------
    elif "webgallery" in k:
        imgs = widget_json.get("images") or widget_json.get("photos") or []
        if isinstance(imgs, list):
            result["images"] = [
                str(i.get("src") or i.get("srcMobile") or i.get("url") or i)
                for i in imgs
                if isinstance(i, (dict, str))
            ]

    # ------------------------------------------------------------------
    # webAddToCart — наличие и цена
    # ------------------------------------------------------------------
    elif "webaddtocart" in k:
        offers = widget_json.get("offers") or []
        if isinstance(offers, list) and offers:
            first = offers[0] if isinstance(offers[0], dict) else {}
            price_v = first.get("price") or first.get("cardPrice") or 0
            if price_v and not result.get("sale_price_rub"):
                result["offer_price"] = _parse_price(price_v)
            stock = first.get("stockCount") or first.get("quantity") or first.get("stock") or 0
            result["in_stock"] = _safe_int(stock)

    # ------------------------------------------------------------------
    # webBreadCrumbs — категория/категории
    # ------------------------------------------------------------------
    elif "webbreadcrumbs" in k or "breadcrumb" in k:
        crumbs = (widget_json.get("breadcrumbs") or
                  widget_json.get("items") or
                  widget_json.get("breadcrumb") or [])
        if isinstance(crumbs, list) and crumbs:
            # Берём последний (самый конкретный) уровень
            last = crumbs[-1] if isinstance(crumbs[-1], dict) else {}
            result["category"] = str(last.get("title") or last.get("name") or
                                     last.get("text") or "")
            # Все уровни через " > "
            titles = []
            for c in crumbs:
                if isinstance(c, dict):
                    t = c.get("title") or c.get("name") or c.get("text") or ""
                    if t:
                        titles.append(str(t))
            result["category_full"] = " > ".join(titles)

    # ------------------------------------------------------------------
    # webDescription — описание
    # ------------------------------------------------------------------
    elif "webdescription" in k:
        desc = widget_json.get("description") or widget_json.get("text") or ""
        if isinstance(desc, list):
            desc = " ".join(str(d) for d in desc)
        result["description"] = str(desc)[:4000]

    # ------------------------------------------------------------------
    # Неизвестный виджет — логируем при отладке
    # ------------------------------------------------------------------
    else:
        # Попытка эвристически извлечь полезные данные из неизвестного виджета
        pass

    return result


def extract_documents_from_widgets(widget_states: Dict[str, str]) -> List[Dict[str, Any]]:
    """
    Итерирует по всем widgetStates и собирает документы/сертификаты из
    виджетов типа webProductDocuments, а также из любых других виджетов,
    содержащих реестровые URLs.

    Возвращает список документов вида:
      {"title": str, "url": str, "type": str, "source_widget": str}
    """
    documents: List[Dict[str, Any]] = []
    registry_urls_found: Set[str] = set()

    for widget_key, widget_raw in widget_states.items():
        try:
            widget = json.loads(widget_raw)
        except (json.JSONDecodeError, TypeError):
            continue

        k = widget_key.lower()

        # Прямые документы из webProductDocuments
        if "webproductdocuments" in k:
            docs = widget.get("documents") or widget.get("items") or []
            if not isinstance(docs, list):
                docs = []
            for doc in docs:
                if not isinstance(doc, dict):
                    continue
                doc_url = str(doc.get("url") or doc.get("link") or doc.get("href") or "")
                doc_title = str(doc.get("title") or doc.get("name") or doc.get("fileName") or "")
                doc_type = str(doc.get("type") or doc.get("docType") or
                               doc.get("documentType") or "")
                entry = {
                    "title": doc_title,
                    "url": doc_url,
                    "type": doc_type,
                    "source_widget": widget_key,
                }
                documents.append(entry)
                if doc_url:
                    registry_urls_found.add(doc_url)

        # Дополнительно: ищем реестровые URLs в любом виджете
        raw_text = widget_raw if isinstance(widget_raw, str) else json.dumps(widget, ensure_ascii=False)
        for url in _extract_registry_urls_from_text(raw_text):
            if url not in registry_urls_found:
                registry_urls_found.add(url)
                documents.append({
                    "title": f"Реестровый документ (найден в {widget_key[:40]})",
                    "url": url,
                    "type": "extracted",
                    "source_widget": widget_key,
                })

    return documents


def extract_original_mark(widget_states: Dict[str, str]) -> Tuple[str, str]:
    """
    Извлекает признак оригинальности из widgetStates.
    Возвращает (is_original_str, original_mark_info_str).
    """
    for widget_key, widget_raw in widget_states.items():
        k = widget_key.lower()
        if "weboriginal" not in k and "original" not in k:
            continue
        try:
            widget = json.loads(widget_raw)
        except (json.JSONDecodeError, TypeError):
            continue
        data = extract_widget_data(widget_key, widget)
        is_orig = data.get("is_original", "")
        info = data.get("original_mark_info", "")
        if is_orig:
            return is_orig, info
    return "", ""


def parse_widget_states(widget_states: Dict[str, str]) -> OzonCard:
    """
    Проходит по всем widgetStates и собирает данные в OzonCard.
    Порядок приоритетов: webProductHeading > webSale для названия,
    webSeller для продавца и т.д.
    """
    card = OzonCard(sku="")

    for widget_key, widget_raw in widget_states.items():
        try:
            widget = json.loads(widget_raw)
        except (json.JSONDecodeError, TypeError):
            continue

        data = extract_widget_data(widget_key, widget)
        if not data:
            continue

        k = widget_key.lower()

        # Название товара (webProductHeading имеет приоритет над webSale)
        if "webproductheading" in k:
            if data.get("product_name"):
                card.product_name = data["product_name"]
            if data.get("rating"):
                card.rating = data["rating"]
            if data.get("feedbacks"):
                card.feedbacks = data["feedbacks"]

        elif "websale" in k:
            if not card.product_name and data.get("product_name"):
                card.product_name = data["product_name"]
            if not card.sku and data.get("sku"):
                card.sku = data["sku"]
            if data.get("price_rub"):
                card.price_rub = data["price_rub"]
            if data.get("sale_price_rub"):
                card.sale_price_rub = data["sale_price_rub"]

        elif "webproductbrand" in k:
            if data.get("brand"):
                card.brand = data["brand"]
            if data.get("brand_url"):
                card.brand_url = data["brand_url"]

        elif "webseller" in k:
            if data.get("seller_name"):
                card.seller_name = data["seller_name"]
            if data.get("seller_id"):
                card.seller_id = data["seller_id"]
            if data.get("seller_url"):
                card.seller_url = data["seller_url"]

        elif "webproductcharacteristics" in k:
            if data.get("characteristics"):
                card.characteristics = data["characteristics"]

        elif "webproductdocuments" in k:
            if data.get("documents"):
                card.documents = data["documents"]

        elif "weboriginal" in k or "original" in k:
            if data.get("is_original") is not None:
                if not card.is_original:
                    card.is_original = data.get("is_original", "")
            if data.get("original_mark_info"):
                card.original_mark_info = data["original_mark_info"]

        elif "webgallery" in k:
            if data.get("images") and not card.images:
                card.images = data["images"]

        elif "webaddtocart" in k:
            if data.get("in_stock"):
                card.in_stock = data["in_stock"]
            # Может уточнить цену
            if not card.sale_price_rub and data.get("offer_price"):
                card.sale_price_rub = data["offer_price"]

        elif "webbreadcrumbs" in k or "breadcrumb" in k:
            if data.get("category"):
                card.category = data["category"]

        elif "webdescription" in k:
            if data.get("description") and not card.description:
                card.description = data["description"]

    return card


# ---------------------------------------------------------------------------
# OzonClient — HTTP-клиент для Ozon composer-api
# ---------------------------------------------------------------------------

class OzonClient:
    """
    Асинхронный клиент для api.ozon.ru/composer-api.bx/page/json/v2.

    Согласно research_apis.md:
    - api.ozon.ru НЕ блокирует запросы Cloudflare
    - 19 параллельных потоков обработали 46 000+ страниц без блокировок
    - Рекомендованная задержка: 200–500 мс между запросами
    """

    def __init__(
        self,
        workers: int = 10,
        delay_min_ms: int = 200,
        delay_max_ms: int = 500,
        timeout_sec: float = 15.0,
        retries: int = 3,
    ):
        self.workers = max(5, min(30, workers))
        self.delay_min = delay_min_ms / 1000.0
        self.delay_max = delay_max_ms / 1000.0
        self.timeout = aiohttp.ClientTimeout(total=timeout_sec, connect=8.0)
        self.retries = retries
        self._semaphore: Optional[asyncio.Semaphore] = None
        self._session: Optional[aiohttp.ClientSession] = None
        self._ua_index: int = 0
        # Счётчики диагностики: позволяют после прогона понять, был ли CF-блок
        self.stats_403: int = 0
        self.stats_307: int = 0
        self.stats_curl_cffi_success: int = 0
        self.stats_curl_cffi_fail: int = 0
        self.stats_total_requests: int = 0
        # curl_cffi сессия (создаётся лениво на первом 403)
        self._curl_session: Optional[Any] = None

    def _next_headers(self) -> Dict[str, str]:
        """Ротация User-Agent для снижения fingerprint."""
        h = dict(OZON_HEADERS)
        h["User-Agent"] = OZON_USER_AGENTS[self._ua_index % len(OZON_USER_AGENTS)]
        self._ua_index += 1
        return h

    async def __aenter__(self) -> "OzonClient":
        self._semaphore = asyncio.Semaphore(self.workers)
        connector = aiohttp.TCPConnector(
            limit=self.workers * 3,
            limit_per_host=self.workers,
            ttl_dns_cache=300,
            ssl=False,  # api.ozon.ru не требует строгой проверки TLS
        )
        self._session = aiohttp.ClientSession(
            connector=connector,
            timeout=self.timeout,
            headers=self._next_headers(),
        )
        return self

    async def __aexit__(self, *args: Any) -> None:
        if self._session:
            await self._session.close()
            self._session = None

    async def _delay(self) -> None:
        """Рандомизированная задержка между запросами."""
        await asyncio.sleep(random.uniform(self.delay_min, self.delay_max))

    def _curl_cffi_get_sync(self, url_path: str) -> Optional[Dict[str, Any]]:
        """
        Синхронный запрос через curl_cffi с TLS-impersonate (chrome120).
        Используется в fallback когда aiohttp получает 403 от Cloudflare.
        api.ozon.ru с конца 2025 блокирует стандартный Python TLS — только impersonate помогает.

        Возвращает dict с widgetStates или None.
        """
        if not CURL_CFFI_OK:
            return None
        try:
            if self._curl_session is None:
                self._curl_session = _curl_requests.Session()
            sess = self._curl_session
            headers = self._next_headers()
            # Убираем Content-Type для GET (лишний)
            headers.pop("Content-Type", None)
            # Пробуем несколько impersonate-вариантов
            for imp in ("chrome120", "chrome110", "chrome"):
                try:
                    r = sess.get(
                        OZON_API_BASE,
                        params={"url": url_path},
                        headers=headers,
                        timeout=15.0,
                        impersonate=imp,
                        allow_redirects=True,
                    )
                    sc = int(getattr(r, "status_code", 0) or 0)
                    if sc == 200:
                        txt = getattr(r, "text", "") or ""
                        if not txt:
                            continue
                        try:
                            return r.json()
                        except Exception:
                            try:
                                return json.loads(txt)
                            except Exception:
                                continue
                    elif sc in (403, 429):
                        # Пробуем следующий impersonate
                        continue
                    else:
                        # Иные ошибки — выходим
                        return None
                except Exception as e:
                    log.debug("curl_cffi (%s) ошибка для %s: %s", imp, url_path, e)
                    continue
            return None
        except Exception as e:
            log.debug("curl_cffi fatal для %s: %s", url_path, e)
            return None

    async def _fetch_via_curl_cffi(self, url_path: str) -> Optional[Dict[str, Any]]:
        """Асинхронная обёртка curl_cffi в executor."""
        if not CURL_CFFI_OK:
            return None
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._curl_cffi_get_sync, url_path)

    async def fetch_json(self, url_path: str) -> Optional[Dict[str, Any]]:
        """
        GET https://api.ozon.ru/composer-api.bx/page/json/v2?url=<url_path>
        с экспоненциальной задержкой на retry (1с/2с/4с).

        Если получен 403/CF-challenge → возвращает None (вызывающий код
        использует Playwright fallback).
        """
        if not self._session or not self._semaphore:
            raise RuntimeError("OzonClient не запущен — используйте async with")

        params = {"url": url_path}
        last_exc: Optional[Exception] = None
        self.stats_total_requests += 1

        async with self._semaphore:
            for attempt in range(self.retries):
                try:
                    await self._delay()
                    async with self._session.get(
                        OZON_API_BASE,
                        params=params,
                        headers=self._next_headers(),
                    ) as resp:
                        status = resp.status
                        if status == 403:
                            self.stats_403 += 1
                            log.debug("403 на %s — вероятно CF. Пробуем alt-заголовки.", url_path)
                            # Попытка с ALT-заголовками (без TLS-fingerprint)
                            async with self._session.get(
                                OZON_API_BASE,
                                params=params,
                                headers=OZON_HEADERS_ALT,
                            ) as resp2:
                                if resp2.status == 200:
                                    return await resp2.json(content_type=None)
                            # CF всё равно блокирует — пробуем curl_cffi (TLS-impersonate Chrome).
                            # Это реальный fallback, который работает в 2025–2026.
                            if CURL_CFFI_OK:
                                log.debug("403 от CF — пробуем curl_cffi fallback")
                                cf_data = await self._fetch_via_curl_cffi(url_path)
                                if cf_data is not None:
                                    self.stats_curl_cffi_success += 1
                                    return cf_data
                                self.stats_curl_cffi_fail += 1
                            return None  # CF блокировка
                        if status == 307:
                            self.stats_307 += 1
                        if status == 404:
                            return None
                        if status == 429:
                            wait = 2 ** attempt * 2
                            log.warning("429 на %s — ждём %ds", url_path, wait)
                            await asyncio.sleep(wait)
                            continue
                        resp.raise_for_status()
                        data = await resp.json(content_type=None)
                        return data
                except aiohttp.ClientResponseError as e:
                    last_exc = e
                    if e.status in (400, 401, 403, 404):
                        return None
                    wait = 2 ** attempt
                    log.debug("HTTP %d на %s — retry %d/%d через %ds",
                               e.status, url_path, attempt + 1, self.retries, wait)
                    await asyncio.sleep(wait)
                except asyncio.TimeoutError:
                    last_exc = asyncio.TimeoutError(f"Timeout: {url_path}")
                    wait = 2 ** attempt
                    log.debug("Timeout на %s — retry %d/%d через %ds",
                               url_path, attempt + 1, self.retries, wait)
                    await asyncio.sleep(wait)
                except aiohttp.ClientError as e:
                    last_exc = e
                    wait = 2 ** attempt
                    log.debug("ClientError на %s: %s — retry %d/%d",
                               url_path, e, attempt + 1, self.retries)
                    await asyncio.sleep(wait)
                except Exception as e:
                    last_exc = e
                    log.warning("Неожиданная ошибка при fetch %s: %s", url_path, e)
                    break

        if last_exc:
            log.debug("Все попытки для %s исчерпаны: %s", url_path, last_exc)
        return None

    async def fetch_widget_states(self, url_path: str) -> Optional[Dict[str, str]]:
        """
        Получает widgetStates из ответа composer-api.
        Возвращает dict {widget_key: widget_json_string} или None при ошибке.
        """
        data = await self.fetch_json(url_path)
        if data is None:
            return None
        widget_states = data.get("widgetStates")
        if not isinstance(widget_states, dict):
            log.debug("widgetStates не найден в ответе для %s. Ключи: %s",
                      url_path, list(data.keys())[:10])
            return None
        return widget_states

    async def search(
        self,
        query: str,
        page: int = 1,
    ) -> Tuple[List[Dict[str, Any]], Optional[str]]:
        """
        Поиск товаров Ozon.
        URL: /search/?text={query}&page={page}

        Возвращает (список товаров, next_page_url или None).
        Каждый товар — dict с полями: sku, title, url, price, finalPrice, ...
        """
        encoded_query = urllib.parse.quote(query)
        url_path = f"/search/?text={encoded_query}&page={page}"
        data = await self.fetch_json(url_path)
        if data is None:
            return [], None

        products: List[Dict[str, Any]] = []
        widget_states = data.get("widgetStates") or {}

        # Извлекаем товары из search widgetStates
        # Основной виджет — webSearchResultsV2 или аналогичный
        for widget_key, widget_raw in widget_states.items():
            k = widget_key.lower()
            # Поиск по распространённым именам виджетов поиска
            if not any(x in k for x in (
                "searchresult", "searchproduct", "catalogresult",
                "productlist", "searchlist", "productitems",
                "searchitemslist", "catalog",
            )):
                continue
            try:
                widget = json.loads(widget_raw)
            except (json.JSONDecodeError, TypeError):
                continue
            items = _extract_search_items(widget)
            if items:
                products.extend(items)

        # Fallback: если специализированные виджеты не нашли — ищем в webSale виджетах
        if not products:
            for widget_key, widget_raw in widget_states.items():
                k = widget_key.lower()
                if "websale" not in k:
                    continue
                try:
                    widget = json.loads(widget_raw)
                except (json.JSONDecodeError, TypeError):
                    continue
                item = _extract_item_from_websale(widget)
                if item:
                    products.append(item)

        # Также пробуем layout для получения списка SKU
        if not products:
            products = _extract_from_layout(data)

        # Следующая страница
        next_page = data.get("nextPage")
        if not next_page:
            # Иногда nextPage зашит в shared или layoutTrackingInfo
            shared = data.get("shared") or {}
            next_page = shared.get("nextPage")

        return products, next_page

    async def parse_card(self, url_path: str) -> Optional[OzonCard]:
        """
        Парсит карточку товара по path вида /product/nazvanie-12345678/
        Возвращает OzonCard или None при ошибке.
        """
        widget_states = await self.fetch_widget_states(url_path)
        if widget_states is None:
            return None

        card = parse_widget_states(widget_states)

        # Также извлекаем документы отдельно (полное сканирование виджетов)
        if not card.documents:
            card.documents = extract_documents_from_widgets(widget_states)

        # SKU из URL если не найден в виджетах
        if not card.sku:
            card.sku = _extract_sku_from_url(url_path)

        # Product URL
        if not card.product_url:
            if card.slug:
                card.product_url = _build_product_url(card.slug, card.sku)
            else:
                card.product_url = _build_product_url("", card.sku)

        return card


# ---------------------------------------------------------------------------
# Search item extractors — вспомогательные функции для парсинга поисковых результатов
# ---------------------------------------------------------------------------

def _extract_search_items(widget: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Пытается извлечь список товаров из поискового виджета.
    Перебирает несколько известных структур.
    """
    items: List[Dict[str, Any]] = []

    # Структура 1: items[]
    raw_items = widget.get("items") or widget.get("products") or widget.get("results") or []
    if isinstance(raw_items, list):
        for item in raw_items:
            parsed = _normalize_search_item(item)
            if parsed:
                items.append(parsed)

    # Структура 2: searchResultsV2 / data.items
    data = widget.get("data") or {}
    if isinstance(data, dict):
        sub = data.get("items") or data.get("products") or []
        if isinstance(sub, list):
            for item in sub:
                parsed = _normalize_search_item(item)
                if parsed:
                    items.append(parsed)

    return items


def _normalize_search_item(item: Any) -> Optional[Dict[str, Any]]:
    """
    Нормализует один элемент из поискового результата.
    Возвращает dict с гарантированными полями или None.
    """
    if not isinstance(item, dict):
        return None

    # SKU: ищем в нескольких полях
    sku = str(
        item.get("id") or item.get("sku") or item.get("skuId") or
        item.get("itemId") or item.get("productId") or ""
    )
    if not sku:
        # Может быть вложен в tracking info
        cti = item.get("cellTrackingInfo") or item.get("trackingInfo") or {}
        if isinstance(cti, dict):
            prod = cti.get("product") or {}
            sku = str(prod.get("id") or prod.get("sku") or "")

    # Название
    title = str(
        item.get("title") or item.get("name") or
        item.get("productName") or item.get("displayName") or ""
    )
    if not title:
        cti = item.get("cellTrackingInfo") or {}
        if isinstance(cti, dict):
            prod = cti.get("product") or {}
            title = str(prod.get("title") or prod.get("name") or "")

    # URL
    url = str(item.get("url") or item.get("link") or item.get("href") or "")
    if url and not url.startswith("http"):
        url = "https://www.ozon.ru" + url

    # Цена
    price_raw = (item.get("price") or item.get("originalPrice") or
                 item.get("priceForCard") or 0)
    final_raw = (item.get("finalPrice") or item.get("cardPrice") or
                 item.get("salePrice") or price_raw)

    # Slug из URL
    slug = ""
    if url:
        m = re.search(r"/product/([^/?#]+)", url)
        if m:
            slug = m.group(1)

    return {
        "sku": sku,
        "title": title,
        "url": url,
        "slug": slug,
        "price": _parse_price(price_raw),
        "finalPrice": _parse_price(final_raw),
        "brand": str(item.get("brand") or item.get("brandName") or ""),
    }


def _extract_item_from_websale(widget: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Извлекает товар из webSale виджета поиска."""
    cti = _deep_get(widget, "cellTrackingInfo", "product") or {}
    sku = str(cti.get("id") or cti.get("sku") or "")
    title = str(cti.get("title") or cti.get("name") or "")
    if not sku and not title:
        return None
    url_path = str(widget.get("url") or cti.get("url") or "")
    if url_path and not url_path.startswith("http"):
        url_path = "https://www.ozon.ru" + url_path
    slug = ""
    if url_path:
        m = re.search(r"/product/([^/?#]+)", url_path)
        if m:
            slug = m.group(1)
    return {
        "sku": sku,
        "title": title,
        "url": url_path,
        "slug": slug,
        "price": _parse_price(cti.get("price") or widget.get("price")),
        "finalPrice": _parse_price(cti.get("finalPrice") or widget.get("finalPrice")),
        "brand": str(cti.get("brandName") or widget.get("brand") or ""),
    }


def _extract_from_layout(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Fallback: пытается найти ссылки на товары в data.layout.
    Возвращает список минимальных записей {sku, url, slug}.
    """
    items = []
    layout = data.get("layout") or []
    if not isinstance(layout, list):
        return []
    widget_states = data.get("widgetStates") or {}
    for comp in layout:
        if not isinstance(comp, dict):
            continue
        # Ищем stateId для получения данных
        state_id = comp.get("stateId")
        if state_id and state_id in widget_states:
            try:
                state = json.loads(widget_states[state_id])
                sub = _extract_search_items(state)
                items.extend(sub)
            except Exception:
                pass
    return items


# ---------------------------------------------------------------------------
# Playwright fallback — для CF-заблокированных карточек
# ---------------------------------------------------------------------------

async def fetch_card_playwright(
    product_url: str,
    headless: bool = True,
    timeout_ms: int = 30000,
) -> Optional[OzonCard]:
    """
    Playwright fallback для получения данных карточки через браузер.
    Используется когда api.ozon.ru вернул 403 или данные неполные.
    Открывает www.ozon.ru/product/... и извлекает widgetStates из XHR-ответов.
    """
    if not PLAYWRIGHT_OK:
        log.warning("Playwright не установлен — CF fallback недоступен")
        return None

    card_data: Optional[Dict[str, Any]] = None

    try:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=headless,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-blink-features=AutomationControlled",
                ],
            )
            ctx = await browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="ru-RU",
            )
            page = await ctx.new_page()

            # Перехватываем XHR с entrypoint-api
            captured: Dict[str, Any] = {}

            async def on_response(response: Any) -> None:
                try:
                    url_r = response.url
                    if "entrypoint-api.bx/page/json" in url_r or "composer-api.bx/page/json" in url_r:
                        j = await response.json()
                        if isinstance(j, dict) and "widgetStates" in j:
                            captured.update(j)
                except Exception:
                    pass

            page.on("response", on_response)

            try:
                await page.goto(product_url, timeout=timeout_ms, wait_until="domcontentloaded")
                await asyncio.sleep(2.5)
            except PlaywrightTimeoutError:
                log.debug("Playwright timeout для %s", product_url)

            if captured and "widgetStates" in captured:
                card_data = captured

            await browser.close()

    except Exception as e:
        log.warning("Playwright fallback error для %s: %s", product_url, e)

    if card_data:
        ws = card_data.get("widgetStates") or {}
        card = parse_widget_states(ws)
        if not card.documents:
            card.documents = extract_documents_from_widgets(ws)
        sku = _extract_sku_from_url(product_url)
        if not card.sku:
            card.sku = sku
        if not card.product_url:
            card.product_url = product_url
        return card

    return None


# ---------------------------------------------------------------------------
# Registry URL resolution from documents
# ---------------------------------------------------------------------------

def resolve_registry_url_from_documents(
    documents: List[Dict[str, Any]],
) -> Tuple[str, str, str]:
    """
    Из списка документов карточки выбирает лучший реестровый URL.
    Возвращает (registry_url, registry_host, registry_record_id).

    Приоритет:
    1. pub.fsa.gov.ru (ФСА — наиболее надёжный)
    2. swis.trade.kg (SWIS/TULPAR)
    3. belgiss.by / tsouz.belgiss.by
    4. portal.eaeunion.org
    5. Любой другой из ALLOWED_REGISTRY_HOSTS
    """
    PRIORITY_HOSTS = [
        "pub.fsa.gov.ru",
        "fsa.gov.ru",
        "swis.trade.kg",
        "tsouz.belgiss.by",
        "belgiss.by",
        "portal.eaeunion.org",
    ]

    candidates: Dict[str, List[str]] = {h: [] for h in PRIORITY_HOSTS}
    others: List[str] = []

    all_urls: List[str] = []
    for doc in documents:
        url = str(doc.get("url") or "")
        if url:
            all_urls.append(url)
        # Также ищем в тексте title
        for text_field in ("title", "description", "text"):
            t = str(doc.get(text_field) or "")
            all_urls.extend(_extract_registry_urls_from_text(t))

    for url in all_urls:
        if not url:
            continue
        try:
            host = urlparse(url).netloc.lower().lstrip("www.")
        except Exception:
            continue
        placed = False
        for ph in PRIORITY_HOSTS:
            if host == ph or host.endswith("." + ph):
                candidates[ph].append(url)
                placed = True
                break
        if not placed and _is_registry_url(url):
            others.append(url)

    # Выбираем лучший
    for ph in PRIORITY_HOSTS:
        if candidates[ph]:
            best = candidates[ph][0]
            host = urlparse(best).netloc.lower()
            rec_id = _extract_registry_record_id(best)
            return best, host, rec_id
    if others:
        best = others[0]
        host = urlparse(best).netloc.lower()
        rec_id = _extract_registry_record_id(best)
        return best, host, rec_id

    return "", "", ""


# ---------------------------------------------------------------------------
# Main collection flow
# ---------------------------------------------------------------------------

# Глобальная диагностика последнего прогона — читается в run() и пишется в xlsx
LAST_SEARCH_DIAGNOSTICS: Dict[str, Any] = {}


async def collect_search_results(
    query: str,
    limit: int = 100,
    workers: int = 10,
    delay_min_ms: int = 200,
    delay_max_ms: int = 500,
) -> List[Dict[str, Any]]:
    """
    Сбор товаров по поисковому запросу.
    Возвращает список dict-ов с полями: sku, title, url, slug, price, finalPrice, brand.

    Постранично обходит /search/?text=...&page=N до набора limit товаров
    или пока есть следующая страница.
    Сохраняет статистику HTTP-статусов в LAST_SEARCH_DIAGNOSTICS.
    """
    global LAST_SEARCH_DIAGNOSTICS
    LAST_SEARCH_DIAGNOSTICS = {
        "query": query, "limit": limit, "workers": workers,
        "started_at": now_iso(), "pages_tried": 0,
        "aiohttp_available": AIOHTTP_OK, "curl_cffi_available": CURL_CFFI_OK,
        "playwright_available": PLAYWRIGHT_OK,
    }

    if not AIOHTTP_OK:
        log.error("aiohttp не установлен. Установите: pip install aiohttp")
        LAST_SEARCH_DIAGNOSTICS["error"] = "aiohttp не установлен"
        return []

    all_items: List[Dict[str, Any]] = []
    page = 1
    seen_skus: Set[str] = set()

    async with OzonClient(
        workers=workers,
        delay_min_ms=delay_min_ms,
        delay_max_ms=delay_max_ms,
    ) as client:
        try:
            while len(all_items) < limit:
                log.info("Поиск '%s' — страница %d (собрано %d/%d)", query, page, len(all_items), limit)
                items, next_page = await client.search(query, page=page)
                LAST_SEARCH_DIAGNOSTICS["pages_tried"] = page

                if not items:
                    log.info("Поиск вернул 0 товаров на странице %d — стоп", page)
                    break

                for item in items:
                    sku = item.get("sku", "")
                    if sku and sku in seen_skus:
                        continue
                    if sku:
                        seen_skus.add(sku)
                    all_items.append(item)
                    if len(all_items) >= limit:
                        break

                if len(all_items) >= limit:
                    break
                if not next_page:
                    log.info("Нет следующей страницы — поиск завершён")
                    break
                page += 1
        finally:
            # Сохраняем статистику клиента для диагностики
            LAST_SEARCH_DIAGNOSTICS.update({
                "total_requests": client.stats_total_requests,
                "http_403": client.stats_403,
                "http_307": client.stats_307,
                "curl_cffi_success": client.stats_curl_cffi_success,
                "curl_cffi_fail": client.stats_curl_cffi_fail,
                "items_collected": len(all_items),
                "finished_at": now_iso(),
            })
            # Авто-диагноз: пытаемся сформулировать причину
            if not all_items:
                reasons = []
                if client.stats_403 > 0 and client.stats_curl_cffi_success == 0:
                    if not CURL_CFFI_OK:
                        reasons.append(
                            "Ozon API блокируется Cloudflare по TLS-фингерпринту. "
                            "Не установлен curl_cffi — установите ручно: pip install curl_cffi"
                        )
                    else:
                        reasons.append(
                            "Ozon API блокируется Cloudflare даже через curl_cffi. "
                            "Возможные причины: IP в чёрном списке (используйте VPN), "
                            "обновите curl_cffi: pip install -U curl_cffi, "
                            "либо добавьте Playwright fallback: pip install playwright && playwright install chromium"
                        )
                elif client.stats_total_requests == 0:
                    reasons.append("Не удалось отправить ни одного запроса — проверьте интернет")
                else:
                    reasons.append("Ответ API не содержит widgetStates — возможно изменилась структура widget'ов Ozon")
                LAST_SEARCH_DIAGNOSTICS["diagnosis"] = " | ".join(reasons)
                log.warning("Диагноз: %s", LAST_SEARCH_DIAGNOSTICS["diagnosis"])

    log.info("Собрано %d товаров для запроса '%s'", len(all_items), query)
    return all_items[:limit]


async def enrich_cards(
    search_items: List[Dict[str, Any]],
    query: str,
    workers: int = 10,
    headless: bool = True,
    delay_min_ms: int = 200,
    delay_max_ms: int = 500,
    progress_callback: Optional[Any] = None,
) -> List[OzonResultRow]:
    """
    Второй этап: для каждого товара из поиска загружает карточку и собирает
    данные о документах/сертификатах.

    Использует пул семафоров для контроля конкурентности.
    progress_callback(current, total) вызывается после каждой обработанной карточки.
    """
    if not AIOHTTP_OK:
        log.error("aiohttp не установлен")
        return []

    total = len(search_items)
    results: List[OzonResultRow] = []
    lock = asyncio.Lock()
    counter = [0]

    async with OzonClient(
        workers=workers,
        delay_min_ms=delay_min_ms,
        delay_max_ms=delay_max_ms,
    ) as client:
        async def process_one(item: Dict[str, Any]) -> OzonResultRow:
            sku = item.get("sku", "")
            url = item.get("url", "")
            slug = item.get("slug", "")
            title = item.get("title", "")
            brand = item.get("brand", "")

            # Строим path для API
            if slug:
                url_path = f"/product/{slug}/"
                product_url = f"https://www.ozon.ru/product/{slug}/"
            elif sku:
                url_path = f"/product/{sku}/"
                product_url = url or f"https://www.ozon.ru/product/{sku}/"
            else:
                # Нет ни slug, ни sku — минимальная запись
                row = OzonResultRow(
                    query=query,
                    nm_id=0,
                    product_name=title,
                    brand=brand,
                    status=STATUS_ERROR,
                    details="Нет SKU и slug для загрузки карточки",
                    checked_at=now_iso(),
                    worker="enrich_cards",
                )
                return row

            # Загружаем карточку
            card: Optional[OzonCard] = await client.parse_card(url_path)

            # CF Fallback через Playwright
            if card is None and PLAYWRIGHT_OK:
                log.info("API fallback → Playwright для %s", product_url)
                card = await fetch_card_playwright(product_url, headless=headless)

            if card is None:
                row = OzonResultRow(
                    query=query,
                    nm_id=_safe_int(sku),
                    ozon_sku=sku,
                    product_name=title,
                    brand=brand,
                    product_url=product_url,
                    status=STATUS_TIMEOUT,
                    details="Не удалось загрузить карточку",
                    checked_at=now_iso(),
                    worker="enrich_cards",
                )
                return row

            # Обновляем данные из search если карточка их не нашла
            if not card.product_name:
                card.product_name = title
            if not card.brand:
                card.brand = brand
            if not card.sku:
                card.sku = sku
            if not card.product_url:
                card.product_url = product_url

            # Ищем реестровые документы
            all_docs = card.documents
            if not all_docs:
                all_docs = extract_documents_from_widgets(
                    {}  # уже было вызвано внутри parse_card
                )

            # Дополнительно: ищем в характеристиках
            chars_text = json.dumps(card.characteristics, ensure_ascii=False)
            extra_registry = _extract_registry_urls_from_text(chars_text)
            for eu in extra_registry:
                all_docs.append({"url": eu, "title": "Из характеристик", "type": "extracted"})

            registry_url, registry_host, registry_record_id = \
                resolve_registry_url_from_documents(all_docs)

            if registry_url:
                status = STATUS_LINK_COLLECTED
            elif all_docs:
                status = STATUS_NO_REGISTRY_LINK
            else:
                status = STATUS_NO_DOCS

            # Строим строку результата
            row = OzonResultRow(
                query=query,
                nm_id=_safe_int(card.sku) if card.sku.isdigit() else 0,
                ozon_sku=card.sku,
                product_name=card.product_name,
                brand=card.brand,
                subject=card.category,
                product_url=card.product_url,
                status=status,
                price_rub=card.price_rub or _parse_price(item.get("price")),
                sale_price_rub=card.sale_price_rub or _parse_price(item.get("finalPrice")),
                seller_name=card.seller_name,
                supplier_id=card.seller_id,
                ozon_seller_url=card.seller_url,
                ozon_brand_url=card.brand_url,
                rating=card.rating,
                feedbacks=card.feedbacks,
                is_original=card.is_original,
                ozon_original_mark_info=card.original_mark_info,
                registry_url=registry_url,
                registry_host=registry_host,
                registry_record_id=registry_record_id,
                ozon_docs_raw=json.dumps(all_docs[:20], ensure_ascii=False)
                if all_docs else "",
                ozon_characteristics=json.dumps(card.characteristics[:50], ensure_ascii=False)
                if card.characteristics else "",
                checked_at=now_iso(),
                worker="enrich_cards",
            )

            return row

        tasks = []
        sem = asyncio.Semaphore(workers)

        async def guarded(item: Dict[str, Any]) -> OzonResultRow:
            async with sem:
                try:
                    row = await process_one(item)
                except Exception as e:
                    log.warning("Ошибка обработки товара %s: %s", item.get("sku", "?"), e)
                    row = OzonResultRow(
                        query=query,
                        nm_id=_safe_int(item.get("sku", "0")),
                        ozon_sku=str(item.get("sku", "")),
                        product_name=str(item.get("title", "")),
                        brand=str(item.get("brand", "")),
                        product_url=str(item.get("url", "")),
                        status=STATUS_ERROR,
                        details=str(e)[:500],
                        checked_at=now_iso(),
                        worker="enrich_cards",
                    )
                async with lock:
                    counter[0] += 1
                    n = counter[0]
                    pct = int(n * 100 / total) if total else 0
                    print(f"{n}/{total} {pct}%", flush=True)
                    if progress_callback:
                        try:
                            progress_callback(n, total)
                        except Exception:
                            pass
                return row

        tasks = [guarded(item) for item in search_items]
        results = list(await asyncio.gather(*tasks))

    return results


# ---------------------------------------------------------------------------
# Save functions — совместимы с main_v39.py
# ---------------------------------------------------------------------------

def _result_row_fields() -> List[str]:
    """Список полей OzonResultRow в порядке dataclass."""
    return list(OzonResultRow.__dataclass_fields__.keys())


def save_csv(rows: List[OzonResultRow], path: Path) -> None:
    """Сохраняет результаты в CSV (UTF-8 BOM, совместимо с main_v39.py)."""
    fields = _result_row_fields()
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with tmp.open("w", encoding="utf-8-sig", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=fields)
            wr.writeheader()
            for r in rows:
                wr.writerow(asdict(r))
        os.replace(tmp, path)
        log.info("CSV сохранён: %s (%d строк)", path, len(rows))
    except Exception as e:
        log.error("Ошибка сохранения CSV %s: %s", path, e)
        if tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass
        raise


def save_xlsx(
    rows: List[OzonResultRow],
    path: Path,
    expiry_warning_days: int = 30,
    make_report_xlsx: bool = True,
) -> None:
    """
    Сохраняет результаты в XLSX.
    Структура: лист results + лист review + лист summary + расширенные листы.
    Полностью совместима с форматом main_v39.py.
    """
    if not OPENPYXL_OK:
        log.error("openpyxl не установлен — XLSX не будет создан. pip install openpyxl")
        return

    wb = Workbook()
    fields = _result_row_fields()

    # ------------------------------------------------------------------ results
    ws = wb.active
    ws.title = "results"
    ws.append(fields)
    for r in rows:
        ws.append([getattr(r, f) for f in fields])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    # Column widths — соответствуют расположению полей OzonResultRow
    _set_column_widths(ws, fields, {
        "query": 18, "nm_id": 12, "product_name": 42, "brand": 22,
        "subject": 20, "product_url": 48, "status": 28,
        "price_rub": 12, "sale_price_rub": 12, "seller_name": 24,
        "registry_url": 60, "registry_host": 24, "registry_record_id": 20,
        "certificate_number": 30, "document_type": 16, "document_status": 14,
        "certificate_product_name": 55, "document_date_start": 14,
        "document_date_end": 14, "details": 60,
        "ozon_docs_raw": 80, "ozon_characteristics": 40,
    })

    # ------------------------------------------------------------------ review
    review = wb.create_sheet("review")
    review.append(fields)
    good_statuses = {STATUS_LINK_COLLECTED, "OK"}
    for r in rows:
        if r.status not in good_statuses:
            review.append([getattr(r, f) for f in fields])
    review.freeze_panes = "A2"
    review.auto_filter.ref = review.dimensions
    for cell in review[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE5CD")

    # ------------------------------------------------------------------ summary
    summary = wb.create_sheet("summary")
    counts: Dict[str, int] = {}
    for r in rows:
        counts[r.status] = counts.get(r.status, 0) + 1
    summary.append(["status", "count"])
    for k, v in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        summary.append([k, v])

    # ------------------------------------------------------------------ расширенные листы
    if make_report_xlsx:
        try:
            _build_summary_sheet(wb, rows, expiry_warning_days)
        except Exception as e:
            log.warning("build_summary_sheet failed: %s", e)
        try:
            _build_details_sheet(wb, rows, expiry_warning_days)
        except Exception as e:
            log.warning("build_details_sheet failed: %s", e)

    # ------------------------------------------------------------------ сохранение
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    last_err = None
    for attempt in range(4):
        try:
            os.replace(tmp, path)
            log.info("XLSX сохранён: %s (%d строк)", path, len(rows))
            return
        except PermissionError as e:
            last_err = e
            time.sleep(0.6)
    # Запасной файл
    try:
        alt = path.with_name(path.stem + "_live" + path.suffix)
        os.replace(tmp, alt)
        log.warning("XLSX файл занят (%s) — пишу в %s", path, alt)
    except Exception:
        log.error("Не удалось сохранить XLSX %s: %s", path, last_err)


def _set_column_widths(ws: Any, fields: List[str], widths: Dict[str, int]) -> None:
    """Устанавливает ширины столбцов по имени поля."""
    for i, f in enumerate(fields, 1):
        if f in widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[f]


def _build_summary_sheet(wb: Any, rows: List[OzonResultRow], warning_days: int) -> None:
    """Лист «Сводка» — совместим с main_v39.py."""
    ws = wb.create_sheet("Сводка", 0)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws.append(["Ozon Registry Checker — расширенный отчёт"])
    ws.append(["Дата формирования", now_iso()])
    ws.append(["Версия движка", APP_VERSION])
    ws.append(["Порог 'Скоро истекает', дней", warning_days])
    ws.append(["Всего товаров", len(rows)])
    ws.append([])

    ws.append(["Распределение по техническому статусу"])
    ws.append(["Статус", "Количество", "Доля, %"])
    status_counts: Dict[str, int] = {}
    for r in rows:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
    total = max(1, len(rows))
    for sn, cnt in sorted(status_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([sn, cnt, round(cnt * 100.0 / total, 1)])
    ws.append([])

    ws.append(["Плашка «Оригинал»"])
    ws.append(["Значение", "Количество", "Доля, %"])
    original_counts: Dict[str, int] = {}
    for r in rows:
        key = r.is_original or "—"
        original_counts[key] = original_counts.get(key, 0) + 1
    for key, cnt in sorted(original_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([key, cnt, round(cnt * 100.0 / total, 1)])
    ws.append([])

    ws.append(["Риски по сроку действия документа"])
    ws.append(["Категория", "Количество", "Доля, %"])
    risk_counts: Dict[str, int] = {
        EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
        EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0,
    }
    for r in rows:
        _, risk = _compute_expiry(r, warning_days)
        risk_counts[risk] = risk_counts.get(risk, 0) + 1
    for key in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
        ws.append([key, risk_counts[key], round(risk_counts[key] * 100.0 / total, 1)])

    ws.append([])
    ws.append(["Домены реестров"])
    ws.append(["Хост", "Количество", "Доля, %"])
    host_counts: Dict[str, int] = {}
    for r in rows:
        h = r.registry_host or "—"
        host_counts[h] = host_counts.get(h, 0) + 1
    for h, cnt in sorted(host_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([h, cnt, round(cnt * 100.0 / total, 1)])

    ws["A1"].font = Font(bold=True, size=14)
    for row_idx in (2, 3, 4, 5):
        ws[f"A{row_idx}"].font = Font(bold=True)
    # Заголовки таблиц
    for row_addr in (8, ):
        for col in ("A", "B", "C"):
            cell = ws[f"{col}{row_addr}"]
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12


def _build_details_sheet(wb: Any, rows: List[OzonResultRow], warning_days: int) -> None:
    """Лист «Подробности» — совместим с main_v39.py."""
    ws = wb.create_sheet("Подробности")
    base_fields = list(OzonResultRow.__dataclass_fields__.keys())
    # Приоритетный порядок (как в main_v39.py CORE_DETAILS_ORDER_V39)
    core_order = (
        "query", "nm_id", "product_name", "brand", "subject", "product_url",
        "status", "price_rub", "sale_price_rub", "seller_name", "supplier_id",
        "is_original", "registry_url", "document_type", "document_status",
        "certificate_number", "document_date_start", "document_date_end",
    )
    seen: Set[str] = set()
    ordered: List[str] = []
    for f in core_order:
        if f in base_fields and f not in seen:
            ordered.append(f)
            seen.add(f)
    for f in base_fields:
        if f not in seen:
            ordered.append(f)
            seen.add(f)
    fields = ordered + ["expiry_days_left", "expiry_risk"]

    ws.append(fields)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CFE2F3")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    expiry_fills = {
        EXPIRY_RISK_OK: PatternFill("solid", fgColor="D9EAD3"),
        EXPIRY_RISK_SOON: PatternFill("solid", fgColor="FFF2CC"),
        EXPIRY_RISK_EXPIRED: PatternFill("solid", fgColor="F4CCCC"),
        EXPIRY_RISK_UNKNOWN: PatternFill("solid", fgColor="EFEFEF"),
    }

    for r in rows:
        days_left, risk = _compute_expiry(r, warning_days)
        row_data = [getattr(r, f) if f in base_fields else "" for f in ordered]
        row_data += [days_left if days_left is not None else "", risk]
        ws.append(row_data)
        # Подсветка по риску
        fill = expiry_fills.get(risk, expiry_fills[EXPIRY_RISK_UNKNOWN])
        row_idx = ws.max_row
        for col_idx in range(1, len(fields) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    _set_column_widths(ws, fields, {
        "query": 18, "product_name": 42, "subject": 20,
        "product_url": 48, "status": 28, "registry_url": 60,
        "document_date_start": 14, "document_date_end": 14,
        "details": 55,
    })


def _write_run_log(
    rows: List[OzonResultRow],
    xlsx_path: Path,
    warning_days: int,
    started_at: float,
    log_path: Optional[Path] = None,
) -> Optional[Path]:
    """Текстовый лог прогона — совместим с main_v39.py."""
    try:
        if log_path is None:
            log_path = xlsx_path.with_name(xlsx_path.stem + "_run.log")
        finished = time.time()
        status_counts: Dict[str, int] = {}
        risk_counts: Dict[str, int] = {
            EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
            EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0,
        }
        for r in rows:
            status_counts[r.status] = status_counts.get(r.status, 0) + 1
            _, risk = _compute_expiry(r, warning_days)
            risk_counts[risk] = risk_counts.get(risk, 0) + 1

        lines: List[str] = []
        lines.append("Ozon Registry Checker — лог прогона")
        lines.append(f"Версия движка: {APP_VERSION}")
        lines.append(f"Старт:    {dt.datetime.fromtimestamp(started_at).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Финиш:    {dt.datetime.fromtimestamp(finished).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Длительность: {int(finished - started_at)} сек")
        lines.append(f"Итоговый файл: {xlsx_path}")
        lines.append(f"Порог 'Скоро истекает', дней: {warning_days}")
        lines.append(f"Всего строк: {len(rows)}")
        lines.append("")
        lines.append("Распределение по техническому статусу:")
        for k, v in sorted(status_counts.items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {k:<40} {v}")
        lines.append("")
        lines.append("Распределение по риску по сроку:")
        for k in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
            lines.append(f"  {k:<20} {risk_counts.get(k, 0)}")
        log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return log_path
    except Exception as e:
        log.warning("Не удалось записать лог прогона: %s", e)
        return None


# ---------------------------------------------------------------------------
# Диагностический xlsx — пишется когда Ozon вернул 0 товаров.
# UI видит файл, юзер открывает и видит объяснение причины.
# ---------------------------------------------------------------------------

def _write_diagnostic_xlsx(
    path: Path,
    query: str,
    started_at: float,
    diagnostics: Dict[str, Any],
) -> None:
    """
    Создаёт xlsx с двумя листами:
      - Сводка:    короткий итог (0 товаров, диагноз, время)
      - Диагностика: все счётчики HTTP, доступные библиотеки, рекомендации
    И пустой лист с обычными заголовками результатов — чтобы UI не фейлился.
    """
    if not OPENPYXL_OK:
        log.error("openpyxl не установлен — не могу создать xlsx")
        return

    elapsed = max(0.0, time.time() - started_at)
    wb = Workbook()

    # ---- Лист Сводка ----
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_sum.append(["Ozon Parser — диагностический отчёт", ""])
    ws_sum.append(["Поисковый запрос", query])
    ws_sum.append(["Собрано товаров", 0])
    ws_sum.append(["Длительность, с", round(elapsed, 1)])
    ws_sum.append(["Дата", now_iso()])
    diag = diagnostics.get("diagnosis") or "Поиск вернул 0 товаров — причина не определена"
    ws_sum.append(["Диагноз", diag])
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 110

    # ---- Лист Диагностика ----
    ws_d = wb.create_sheet("Диагностика")
    ws_d.append(["Показатель", "Значение"])
    for k in (
        "query", "limit", "workers", "started_at", "finished_at",
        "pages_tried", "total_requests", "http_403", "http_307",
        "curl_cffi_available", "curl_cffi_success", "curl_cffi_fail",
        "aiohttp_available", "playwright_available",
        "items_collected", "diagnosis", "error",
    ):
        if k in diagnostics:
            ws_d.append([k, str(diagnostics[k])])
    ws_d.append(["", ""])
    ws_d.append(["Что делать?", ""])
    ws_d.append(["Шаг 1", "Установить curl_cffi: pip install curl_cffi"])
    ws_d.append(["Шаг 2", "Если не помогло — обновить: pip install -U curl_cffi"])
    ws_d.append(["Шаг 3", "Если и это не помогает — включить VPN или установить Playwright: pip install playwright && playwright install chromium"])
    ws_d.column_dimensions["A"].width = 24
    ws_d.column_dimensions["B"].width = 100

    # ---- Пустой лист results с заголовками — нужен для UI ----
    ws_r = wb.create_sheet("results")
    ws_r.append(_result_row_fields())

    wb.save(path)


# ---------------------------------------------------------------------------
# Main async entry point
# ---------------------------------------------------------------------------

async def run(
    query: str,
    limit: int = 100,
    workers: int = 10,
    headless: bool = True,
    output: str = "ozon_result.xlsx",
    expiry_warning_days: int = 30,
    make_report_xlsx: bool = True,
    run_log: str = "",
) -> List[OzonResultRow]:
    """
    Полный цикл:
    1) Поиск товаров по запросу
    2) Загрузка карточек и документов
    3) Сохранение в CSV+XLSX

    Возвращает список OzonResultRow.
    """
    started_at = time.time()
    output_path = Path(output)

    log.info("=== Ozon Parser v27 ===")
    log.info("Запрос: '%s', лимит: %d, воркеры: %d", query, limit, workers)

    # --- Этап 1: сбор поиска ---
    print(f"[Ozon] Поиск: '{query}', лимит: {limit}")
    search_items = await collect_search_results(
        query=query,
        limit=limit,
        workers=workers,
    )

    if not search_items:
        log.warning("Поиск не дал результатов для '%s'", query)
        print("[Ozon] Нет результатов поиска.")
        # Создаём xlsx с листом Диагностика — чтобы UI не показывал «нет файла»
        if output_path.suffix.lower() in (".xlsx", ".xls"):
            try:
                _write_diagnostic_xlsx(
                    output_path, query, started_at,
                    diagnostics=dict(LAST_SEARCH_DIAGNOSTICS),
                )
                print(f"[Ozon] Создан диагностический XLSX: {output_path}")
            except Exception as e:
                log.error("Не удалось создать диагностический xlsx: %s", e)
        return []

    print(f"[Ozon] Найдено {len(search_items)} товаров, загружаю карточки...")

    # --- Этап 2: обогащение карточками ---
    results = await enrich_cards(
        search_items=search_items,
        query=query,
        workers=workers,
        headless=headless,
    )

    print(f"[Ozon] Обработано {len(results)} карточек. Сохраняю...")

    # --- Сохранение CSV ---
    csv_path = output_path.with_suffix(".csv")
    save_csv(results, csv_path)

    # --- Сохранение XLSX ---
    if output_path.suffix.lower() in (".xlsx", ".xls"):
        save_xlsx(
            results, output_path,
            expiry_warning_days=expiry_warning_days,
            make_report_xlsx=make_report_xlsx,
        )
    else:
        xlsx_path = output_path.with_suffix(".xlsx")
        save_xlsx(
            results, xlsx_path,
            expiry_warning_days=expiry_warning_days,
            make_report_xlsx=make_report_xlsx,
        )

    # --- Лог прогона ---
    if run_log:
        log_path = Path(run_log)
    else:
        log_path = None

    _write_run_log(
        results, output_path, expiry_warning_days, started_at, log_path
    )

    # --- Итоговая статистика ---
    status_counts: Dict[str, int] = {}
    for r in results:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
    print("[Ozon] Готово. Статистика:")
    for s, c in sorted(status_counts.items(), key=lambda x: (-x[1], x[0])):
        print(f"  {s:<40} {c}")

    return results


# ---------------------------------------------------------------------------
# CLI — argparse
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    """
    Строит парсер аргументов. Совместим с main_v39.py:
    принимает те же ключи --query, --limit, --workers, --headless,
    --output, --expiry-warning-days, --make-report-xlsx, --run-log.
    """
    ap = argparse.ArgumentParser(
        description="Ozon parser v27 — поиск товаров и проверка документов/сертификатов",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    # Основные аргументы (совместимы с main_v39.py)
    ap.add_argument(
        "--query", default="", required=False,
        help="Поисковый запрос Ozon"
    )
    ap.add_argument(
        "--limit", type=int, default=100,
        help="Максимальное количество товаров для обработки"
    )
    ap.add_argument(
        "--workers", type=int, default=10,
        help="Число параллельных воркеров (5–30)"
    )
    ap.add_argument(
        "--headless", type=str_to_bool, default=True,
        help="Playwright headless режим (True/False)"
    )
    ap.add_argument(
        "--output", default="ozon_result.xlsx",
        help="Путь к выходному файлу (.xlsx или .csv)"
    )
    ap.add_argument(
        "--expiry-warning-days", type=int, default=30,
        help="Порог в днях для флага 'Скоро истекает'"
    )
    ap.add_argument(
        "--make-report-xlsx", type=str_to_bool, default=True,
        help="Создавать расширенные листы отчёта в XLSX"
    )
    ap.add_argument(
        "--run-log", default="",
        help="Путь к текстовому логу прогона. Пусто — рядом с output как *_run.log"
    )
    # Дополнительные аргументы Ozon-парсера
    ap.add_argument(
        "--delay-min-ms", type=int, default=200,
        help="Минимальная задержка между запросами (мс)"
    )
    ap.add_argument(
        "--delay-max-ms", type=int, default=500,
        help="Максимальная задержка между запросами (мс)"
    )
    ap.add_argument(
        "--timeout-sec", type=float, default=15.0,
        help="Timeout HTTP-запроса в секундах"
    )
    ap.add_argument(
        "--geo", default="",
        help="Код геолокации Ozon (например: 0c5b2444-70a0-4932-980c-b4dc0d3f02b5 — Москва)"
    )
    ap.add_argument(
        "--verbose", action="store_true",
        help="Подробное логирование"
    )
    # v27.6: переключатель на новый Playwright-парсер (по умолчанию включён)
    ap.add_argument(
        "--use-playwright", type=str_to_bool, default=True,
        help="Использовать браузерный (Playwright) парсер вместо HTTP — обходит CloudFlare"
    )
    return ap


def main() -> None:
    """Точка входа CLI."""
    ap = build_arg_parser()
    args = ap.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Проверка зависимостей
    missing: List[str] = []
    if not AIOHTTP_OK:
        missing.append("aiohttp")
    if not OPENPYXL_OK:
        missing.append("openpyxl")
    if missing:
        print(f"[ERROR] Не установлены зависимости: {', '.join(missing)}")
        print(f"[ERROR] Установите: pip install {' '.join(missing)}")
        if not AIOHTTP_OK:
            sys.exit(1)

    if not args.query:
        print("[ERROR] Укажите поисковый запрос через --query")
        ap.print_help()
        sys.exit(1)

    # Нормализация workers
    workers = max(5, min(30, args.workers))

    # v27.6: диспетчер — если --use-playwright (по умолчанию ON),
    # передаём управление в ozon_parser_playwright.
    if getattr(args, "use_playwright", True):
        try:
            from ozon_parser_playwright import run_playwright_parser
        except ImportError as e:
            print(f"[ERROR] Не найден модуль ozon_parser_playwright: {e}")
            sys.exit(1)
        print("[Ozon] Режим: Playwright (браузерный) — v27.6")
        try:
            rc = asyncio.run(run_playwright_parser(
                query=args.query,
                limit=args.limit,
                workers=1,  # Ozon банит при параллели — всегда последовательно
                headless=args.headless,
                output=args.output,
                expiry_warning_days=args.expiry_warning_days,
                make_report_xlsx=args.make_report_xlsx,
                run_log=args.run_log,
                delay_min_ms=max(args.delay_min_ms, 800),
                delay_max_ms=max(args.delay_max_ms, 2500),
            ))
            sys.exit(rc)
        except KeyboardInterrupt:
            print("\n[Ozon] Прерван пользователем.")
            sys.exit(130)
        except Exception as e:
            log.error("Критическая ошибка Playwright-парсера: %s", e)
            if args.verbose:
                traceback.print_exc()
            sys.exit(1)

    # Legacy HTTP-путь (оставлен как fallback при --use-playwright=false)
    print("[Ozon] Режим: HTTP (legacy) — может быть заблокирован CloudFlare")
    try:
        asyncio.run(run(
            query=args.query,
            limit=args.limit,
            workers=workers,
            headless=args.headless,
            output=args.output,
            expiry_warning_days=args.expiry_warning_days,
            make_report_xlsx=args.make_report_xlsx,
            run_log=args.run_log,
        ))
    except KeyboardInterrupt:
        print("\n[Ozon] Прерван пользователем.")
    except Exception as e:
        log.error("Критическая ошибка: %s", e)
        if args.verbose:
            traceback.print_exc()
        sys.exit(1)


# ---------------------------------------------------------------------------
# Дополнительные публичные утилиты
# ---------------------------------------------------------------------------

async def parse_single_card(
    product_url: str,
    headless: bool = True,
) -> Optional[OzonResultRow]:
    """
    Парсит одну карточку товара по URL.
    Удобно для использования из других модулей.

    Пример:
        row = asyncio.run(parse_single_card("https://www.ozon.ru/product/nazvanie-12345678/"))
    """
    if not AIOHTTP_OK:
        log.error("aiohttp не установлен")
        return None

    # Строим path из URL
    parsed = urlparse(product_url)
    url_path = parsed.path
    if parsed.query:
        url_path += "?" + parsed.query

    sku = _extract_sku_from_url(product_url)

    async with OzonClient(workers=5) as client:
        card = await client.parse_card(url_path)

    if card is None and PLAYWRIGHT_OK:
        card = await fetch_card_playwright(product_url, headless=headless)

    if card is None:
        return None

    if not card.sku:
        card.sku = sku
    if not card.product_url:
        card.product_url = product_url

    all_docs = card.documents
    registry_url, registry_host, registry_record_id = \
        resolve_registry_url_from_documents(all_docs)

    if registry_url:
        status = STATUS_LINK_COLLECTED
    elif all_docs:
        status = STATUS_NO_REGISTRY_LINK
    else:
        status = STATUS_NO_DOCS

    return OzonResultRow(
        query="",
        nm_id=_safe_int(card.sku) if card.sku.isdigit() else 0,
        ozon_sku=card.sku,
        product_name=card.product_name,
        brand=card.brand,
        subject=card.category,
        product_url=card.product_url,
        status=status,
        price_rub=card.price_rub,
        sale_price_rub=card.sale_price_rub,
        seller_name=card.seller_name,
        supplier_id=card.seller_id,
        ozon_seller_url=card.seller_url,
        ozon_brand_url=card.brand_url,
        rating=card.rating,
        feedbacks=card.feedbacks,
        is_original=card.is_original,
        ozon_original_mark_info=card.original_mark_info,
        registry_url=registry_url,
        registry_host=registry_host,
        registry_record_id=registry_record_id,
        ozon_docs_raw=json.dumps(all_docs[:20], ensure_ascii=False) if all_docs else "",
        ozon_characteristics=json.dumps(card.characteristics[:50], ensure_ascii=False)
        if card.characteristics else "",
        checked_at=now_iso(),
        worker="parse_single_card",
    )


def get_allowed_registry_hosts() -> Set[str]:
    """Возвращает набор разрешённых реестровых хостов (идентично main_v39.py)."""
    return set(ALLOWED_REGISTRY_HOSTS)


async def search_ozon(
    query: str,
    limit: int = 50,
    workers: int = 5,
) -> List[Dict[str, Any]]:
    """
    Лёгкий поиск без загрузки карточек.
    Возвращает список {sku, title, url, price, finalPrice, brand}.
    Полезно для предварительного отбора.
    """
    return await collect_search_results(query, limit=limit, workers=workers)


# ---------------------------------------------------------------------------
# Smoke test / __main__
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    """
    Два режима:
    1) Если в аргументах есть --output или --run-log — это вызов из wb_checker.py (продакшн-режим).
       Идём в main() с полным run() — собираем товары и пишем xlsx.
    2) Иначе — smoke-тест (ручная проверка).
    """
    # Быстрая проверка аргументов: есть ли --output или --run-log (признак продакшн-вызова)
    if any(a in sys.argv for a in ("--output", "--run-log")):
        try:
            main()
            sys.exit(0)
        except SystemExit:
            raise
        except Exception as _e:
            print(f"[FATAL] Ozon parser упал: {_e}", file=sys.stderr)
            import traceback
            traceback.print_exc()
            sys.exit(2)

    # ---- Smoke test ветка (при запуске вручную без --output) ----
    import socket

    def _check_connectivity(host: str = "api.ozon.ru", port: int = 443, timeout: float = 3.0) -> bool:
        try:
            socket.setdefaulttimeout(timeout)
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.connect((host, port))
            return True
        except (socket.error, OSError):
            return False

    # Парсим аргументы
    _ap = build_arg_parser()
    # Если нет --query — устанавливаем дефолтный smoke-тест запрос
    _args, _unknown = _ap.parse_known_args()

    if not _args.query:
        _args.query = "кофемашина"
        _args.limit = 5
        print(f"[Smoke] Запрос не задан — используем тестовый: '{_args.query}', limit={_args.limit}")

    print("=" * 60)
    print(f"Ozon Parser v27 — Smoke Test")
    print(f"Запрос: '{_args.query}', limit: {_args.limit}")
    print("=" * 60)

    # Проверка зависимостей
    _deps_ok = True
    for _dep, _ok in [("aiohttp", AIOHTTP_OK), ("openpyxl", OPENPYXL_OK)]:
        status_str = "OK" if _ok else "MISSING"
        print(f"  {_dep:<20} {status_str}")
        if not _ok:
            _deps_ok = False
    for _dep, _ok in [("playwright", PLAYWRIGHT_OK), ("curl_cffi", CURL_CFFI_OK),
                       ("beautifulsoup4", BS4_OK)]:
        status_str = "OK" if _ok else "optional, not installed"
        print(f"  {_dep:<20} {status_str}")
    print()

    if not _deps_ok:
        print("[FATAL] Установите обязательные зависимости:")
        print("  pip install aiohttp openpyxl beautifulsoup4")
        print("Опционально:")
        print("  pip install playwright curl_cffi")
        print("  playwright install chromium")
        sys.exit(1)

    # Проверка интернета
    print(f"Проверка соединения с api.ozon.ru...", end=" ")
    if not _check_connectivity():
        print("НЕДОСТУПЕН")
        print("[WARNING] Нет доступа к api.ozon.ru. Проверьте интернет-соединение.")
        print("[INFO] Ozon Parser готов к работе при наличии сети.")
        sys.exit(0)
    print("OK")
    print()

    async def _smoke_test() -> None:
        print(f"[1/3] Поиск: '{_args.query}' (до {_args.limit} товаров)...")
        items = await collect_search_results(
            _args.query,
            limit=_args.limit,
            workers=max(5, min(10, _args.limit)),
        )
        print(f"  Найдено: {len(items)} товаров в поиске")

        if not items:
            print("[WARNING] Поиск вернул 0 результатов.")
            print("[INFO] Возможные причины:")
            print("  - Запрос слишком специфичен")
            print("  - Изменилась структура widgetStates Ozon")
            print("  - Временная недоступность API")
            return

        for i, item in enumerate(items[:3], 1):
            print(f"  {i}. [{item.get('sku','?')}] {item.get('title','?')[:60]}")
            print(f"     URL: {item.get('url','?')[:80]}")
            print(f"     Цена: {item.get('finalPrice', item.get('price', '?'))} руб.")

        print()
        print(f"[2/3] Загрузка карточек ({min(2, len(items))} шт. для smoke-теста)...")
        test_items = items[:min(2, len(items))]
        results = await enrich_cards(
            search_items=test_items,
            query=_args.query,
            workers=5,
            headless=_args.headless,
        )

        print(f"  Обработано: {len(results)} карточек")
        print()
        for r in results:
            print(f"  [{r.ozon_sku}] {r.product_name[:60]}")
            print(f"    Бренд:    {r.brand}")
            print(f"    Продавец: {r.seller_name}")
            print(f"    Статус:   {r.status}")
            print(f"    Реестр:   {r.registry_url or '(нет)'}")
            print(f"    Оригинал: {r.is_original or '(неизвестно)'}")
            if r.ozon_docs_raw:
                docs = json.loads(r.ozon_docs_raw)
                print(f"    Документов найдено: {len(docs)}")
                for d in docs[:3]:
                    print(f"      - {d.get('title','?')[:50]}: {d.get('url','?')[:80]}")
            print()

        print(f"[3/3] Smoke-тест завершён успешно.")
        print()

        # Итоговая статистика
        status_counts: Dict[str, int] = {}
        for r in results:
            status_counts[r.status] = status_counts.get(r.status, 0) + 1
        print("Статистика статусов:")
        for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
            print(f"  {s:<40} {c}")

        print()
        print("Для полного запуска с сохранением в файл:")
        print(f"  python ozon_parser.py --query \"{_args.query}\" --limit 100 --output result.xlsx")

    try:
        asyncio.run(_smoke_test())
    except KeyboardInterrupt:
        print("\n[Smoke] Прерван пользователем.")
    except Exception as _e:
        print(f"\n[Smoke] ОШИБКА: {_e}")
        import traceback as _tb
        _tb.print_exc()
        sys.exit(1)


# ===========================================================================
# ДОПОЛНИТЕЛЬНЫЕ МОДУЛИ (встроены в ozon_parser.py для самодостаточности)
# ===========================================================================


# ---------------------------------------------------------------------------
# Модуль 1: FSA HTTP-клиент (встроенный, аналог fsa_http_fetcher.py)
# ---------------------------------------------------------------------------

class OzonFSACache:
    """
    Кэш результатов запросов к реестрам ФСА.
    Предотвращает дублирующиеся запросы к одному и тому же документу.
    """
    def __init__(self, max_size: int = 10000):
        self._cache: Dict[str, Dict[str, str]] = {}
        self._max_size = max_size
        self._hits = 0
        self._misses = 0

    def get(self, key: str) -> Optional[Dict[str, str]]:
        val = self._cache.get(key)
        if val is not None:
            self._hits += 1
        else:
            self._misses += 1
        return val

    def put(self, key: str, val: Dict[str, str]) -> None:
        if len(self._cache) >= self._max_size:
            # FIFO-эвикция: удаляем первый добавленный
            oldest = next(iter(self._cache))
            del self._cache[oldest]
        self._cache[key] = dict(val)

    @property
    def size(self) -> int:
        return len(self._cache)

    def stats(self) -> str:
        total = self._hits + self._misses
        rate = self._hits * 100.0 / total if total else 0
        return f"FSA cache: {self.size} items, {self._hits} hits, {self._misses} misses ({rate:.1f}% hit rate)"


_FSA_CACHE_INSTANCE = OzonFSACache()


def extract_fsa_kind_id(url: str) -> Tuple[str, str]:
    """
    Разбирает FSA-URL и возвращает ('rss_certificate' | 'rds_declaration', '{id}').
    Поддерживает форматы:
      https://pub.fsa.gov.ru/rss/certificate/view/{id}/baseInfo
      https://pub.fsa.gov.ru/rds/declaration/view/{id}/common
      https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{id}
      https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{id}
    Возвращает ('', '') если URL не FSA.
    """
    try:
        u = urlparse(url)
        if "fsa.gov.ru" not in (u.netloc or "").lower():
            return "", ""
        path = u.path or ""
        # HTML view URLs
        m = re.search(r"/rss/certificate/(?:view|details|card|api)?/?(\d{3,})", path)
        if m:
            return "rss_certificate", m.group(1)
        m = re.search(r"/rds/declaration/(?:view|details|card|api)?/?(\d{3,})", path)
        if m:
            return "rds_declaration", m.group(1)
        # API URLs
        m = re.search(r"/api/v\d+/rss/(?:common/)?certificates?/(\d{3,})", path)
        if m:
            return "rss_certificate", m.group(1)
        m = re.search(r"/api/v\d+/rds/(?:common/)?declarations?/(\d{3,})", path)
        if m:
            return "rds_declaration", m.group(1)
    except Exception:
        pass
    return "", ""


def _fsa_api_url(kind: str, doc_id: str) -> str:
    """Строит API URL для получения данных документа ФСА."""
    if kind == "rss_certificate":
        return f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{doc_id}"
    if kind == "rds_declaration":
        return f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{doc_id}"
    return ""


def _fsa_referer(kind: str, doc_id: str) -> str:
    """Строит Referer URL для ФСА."""
    if kind == "rds_declaration":
        return f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common"
    return f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"


def _fsa_browser_headers_ozon(referer: str) -> Dict[str, str]:
    """Заголовки для запросов к ФСА, скопированные из HAR Chrome."""
    return {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": referer,
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
        ),
        "lkId": "",
        "orgId": "",
        "sec-ch-ua": '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }


def _fsa_unwrap_payload(obj: Any, max_depth: int = 5) -> Any:
    """ФСА иногда оборачивает payload в {data: ...} или {item: ...}."""
    cur = obj
    for _ in range(max_depth):
        if isinstance(cur, dict):
            for k in ("data", "item", "result", "content", "payload",
                      "declaration", "certificate", "document"):
                v = cur.get(k)
                if isinstance(v, (dict, list)):
                    cur = v
                    break
            else:
                return cur
        elif isinstance(cur, list) and len(cur) == 1 and isinstance(cur[0], dict):
            cur = cur[0]
        else:
            return cur
    return cur


def _fsa_walk_leaves(
    obj: Any, path: Tuple[str, ...] = ()
) -> List[Tuple[Tuple[str, ...], Any]]:
    """Глубокий обход JSON, возвращает все листовые значения с путями."""
    out: List[Tuple[Tuple[str, ...], Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            out.extend(_fsa_walk_leaves(v, path + (str(k),)))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            out.extend(_fsa_walk_leaves(v, path + (f"[{i}]",)))
    else:
        out.append((path, obj))
    return out


def _fsa_find_value(
    leaves: List[Tuple[Tuple[str, ...], Any]],
    include_any: Tuple[str, ...],
    include_all: Tuple[str, ...] = (),
    exclude_any: Tuple[str, ...] = (),
) -> str:
    """Найти первое значение, чей путь содержит ключевые слова."""
    inc_any = [s.lower() for s in include_any]
    inc_all = [s.lower() for s in include_all]
    exc_any = [s.lower() for s in exclude_any]
    for path, v in leaves:
        if v in (None, "", [], {}):
            continue
        pl = "/".join(path).lower()
        if any(x in pl for x in exc_any):
            continue
        if inc_all and not all(x in pl for x in inc_all):
            continue
        if inc_any and not any(x in pl for x in inc_any):
            continue
        if isinstance(v, (dict, list)):
            continue
        s = str(v).strip()
        if s and len(s) < 2000:
            return s
    return ""


def _fsa_format_date(s: Any) -> str:
    """ISO-дата → DD.MM.YYYY."""
    if not s:
        return ""
    s = str(s).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"
    return s[:10]


def parse_fsa_json_response(
    obj: Any, url: str, kind: str, doc_id: str
) -> Dict[str, str]:
    """
    Извлекает ключевые поля из JSON-ответа ФСА.
    Возвращает dict с полями совместимыми с OzonResultRow.
    """
    payload = _fsa_unwrap_payload(obj)
    leaves = _fsa_walk_leaves(payload)

    out: Dict[str, str] = {
        "document_type": "Сертификат" if kind == "rss_certificate" else "Декларация",
        "certificate_number": _fsa_find_value(
            leaves, ("number", "regnumber", "registrationnumber"),
            exclude_any=("blank", "applicant", "manufacturer")
        ),
        "document_status": _fsa_find_value(
            leaves, ("status",), exclude_any=("change", "history")
        ),
        "document_date_start": _fsa_format_date(
            _fsa_find_value(
                leaves,
                ("certregdate", "regdate", "datestart", "date_start", "issuedate"),
                exclude_any=("end", "expir")
            )
        ),
        "document_date_end": _fsa_format_date(
            _fsa_find_value(
                leaves,
                ("certenddate", "datetill", "dateend", "date_end", "expirationdate")
            )
        ),
        "applicant_name": (
            _fsa_find_value(leaves, ("applicant", "fullname"), include_all=("applicant",))
            or _fsa_find_value(leaves, ("applicantname",))
        ),
        "applicant_inn": (
            _fsa_find_value(leaves, ("applicantinn",), include_all=("applicant",))
            or _fsa_find_value(leaves, ("inn",), include_all=("applicant",))
        ),
        "manufacturer_name": (
            _fsa_find_value(leaves, ("manufacturer", "fullname"), include_all=("manufacturer",))
            or _fsa_find_value(leaves, ("manufacturername",))
        ),
        "certificate_product_name": _fsa_find_value(
            leaves, ("productname", "product_name", "product", "fullname"),
            exclude_any=("manufacturer", "applicant", "group", "type")
        ),
        "tnved": _fsa_find_value(leaves, ("tnved", "tncode", "tn_ved")),
        "scheme": _fsa_find_value(leaves, ("scheme",)),
        "technical_regulation": _fsa_find_value(
            leaves, ("techreg", "technicalregulation", "technical_regulation", "shortname"),
            include_all=("tech",)
        ),
    }
    # Убираем пустые поля
    return {k: v for k, v in out.items() if v}


def fetch_fsa_document_http(
    url: str,
    timeout_sec: float = 10.0,
    impersonate: str = "chrome",
) -> Optional[Dict[str, str]]:
    """
    Синхронный HTTP-запрос к ФСА через curl_cffi (имитация Chrome TLS).

    ФСА использует TLS fingerprint — обычный requests вернёт 403.
    curl_cffi обходит это через имитацию браузерного TLS.

    Возвращает dict с полями документа или None если:
    - URL не ФСА
    - curl_cffi не установлен
    - Запрос не удался

    Источник: research_apis.md раздел 3.7, раздел 5.3
    """
    if not CURL_CFFI_OK:
        log.debug("curl_cffi не установлен — ФСА HTTP недоступен. pip install curl_cffi")
        return None

    kind, doc_id = extract_fsa_kind_id(url)
    if not kind or not doc_id:
        return None

    cache_key = f"{kind}:{doc_id}"
    cached = _FSA_CACHE_INSTANCE.get(cache_key)
    if cached is not None:
        return dict(cached)

    referer = _fsa_referer(kind, doc_id)
    api_url = _fsa_api_url(kind, doc_id)
    if not api_url:
        return None

    # Пробуем несколько вариантов impersonate
    impersonates = [impersonate, "chrome120", "chrome110"]
    seen_imps: Set[str] = set()
    impersonates_dedup = [x for x in impersonates if not (x in seen_imps or seen_imps.add(x))]

    for imp in impersonates_dedup:
        try:
            sess = _curl_requests.Session()
            h = _fsa_browser_headers_ozon(referer)
            try:
                r = sess.get(api_url, headers=h, timeout=timeout_sec, impersonate=imp)
            except TypeError:
                r = sess.get(api_url, headers=h, timeout=timeout_sec)

            sc = int(getattr(r, "status_code", 0) or 0)
            if sc != 200:
                continue
            txt = getattr(r, "text", "") or ""
            if not txt:
                continue
            try:
                obj = r.json()
            except Exception:
                try:
                    obj = json.loads(txt)
                except Exception:
                    continue
            parsed = parse_fsa_json_response(obj, api_url, kind, doc_id)
            if parsed and parsed.get("certificate_number"):
                _FSA_CACHE_INSTANCE.put(cache_key, parsed)
                return parsed
        except Exception as e:
            log.debug("curl_cffi error для ФСА %s (%s): %s", api_url, imp, e)
            continue

    return None


async def fetch_fsa_document_async(
    url: str,
    session: Optional[Any] = None,
    timeout_sec: float = 10.0,
) -> Optional[Dict[str, str]]:
    """
    Асинхронный вариант fetch_fsa_document_http.
    Запускает синхронный curl_cffi в executor чтобы не блокировать event loop.
    """
    if not CURL_CFFI_OK:
        return None
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        None,
        lambda: fetch_fsa_document_http(url, timeout_sec=timeout_sec)
    )


async def enrich_with_fsa_data(
    rows: List[OzonResultRow],
    workers: int = 10,
    timeout_sec: float = 10.0,
) -> List[OzonResultRow]:
    """
    Второй мини-этап: для строк с registry_url на pub.fsa.gov.ru
    загружает детали документа (номер, статус, даты, заявитель и т.д.)
    через HTTP curl_cffi.

    Обновляет поля OzonResultRow на месте (возвращает тот же список).
    """
    if not CURL_CFFI_OK:
        log.info("curl_cffi не установлен — обогащение данными ФСА пропущено")
        return rows

    fsa_rows = [
        (i, r) for i, r in enumerate(rows)
        if r.registry_url and "fsa.gov.ru" in r.registry_url
        and not r.certificate_number  # Не перезаписываем уже заполненные
    ]

    if not fsa_rows:
        return rows

    log.info("Обогащение данными ФСА: %d строк из %d", len(fsa_rows), len(rows))
    sem = asyncio.Semaphore(workers)
    counter = [0]

    async def process_fsa_row(idx: int, row: OzonResultRow) -> None:
        async with sem:
            try:
                result = await fetch_fsa_document_async(row.registry_url, timeout_sec=timeout_sec)
                if result:
                    for field_name, value in result.items():
                        if hasattr(row, field_name) and value:
                            current = getattr(row, field_name, "")
                            if not current:
                                setattr(row, field_name, value)
                    # Обновляем статус если получили данные
                    if result.get("certificate_number") and row.status == STATUS_LINK_COLLECTED:
                        row.status = STATUS_LINK_COLLECTED + " (подтверждено ФСА)"
            except Exception as e:
                log.debug("FSA enrich error для %s: %s", row.registry_url, e)
            finally:
                counter[0] += 1
                if counter[0] % 10 == 0:
                    log.info("FSA обогащение: %d/%d", counter[0], len(fsa_rows))

    tasks = [process_fsa_row(i, r) for i, r in fsa_rows]
    await asyncio.gather(*tasks)

    return rows


# ---------------------------------------------------------------------------
# Модуль 2: SWIS/TULPAR HTML-парсер (встроенный)
# ---------------------------------------------------------------------------

def parse_swis_document_html(html_content: str) -> Dict[str, str]:
    """
    Парсит HTML-страницу документа SWIS/TULPAR (swis.trade.kg/Doc/{UUID}).

    SWIS — классический HTML-сайт без JSON API. Данные извлекаются из
    таблиц со строками вида:
      <td>Название поля</td><td>Значение</td>

    Источник: research_apis.md раздел 4.4
    """
    if not BS4_OK or BeautifulSoup is None:
        log.debug("BeautifulSoup не установлен — SWIS HTML-парсинг недоступен")
        # Упрощённый парсинг без BS4
        return _parse_swis_regex(html_content)

    result: Dict[str, str] = {}
    try:
        soup = BeautifulSoup(html_content, "html.parser")
        tables = soup.find_all("table")
        for table in tables:
            for row in table.find_all("tr"):
                cells = row.find_all("td")
                if len(cells) >= 2:
                    key = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if key and val:
                        result[key] = val
        # Специфические поля SWIS
        _extract_swis_special_fields(soup, result)
    except Exception as e:
        log.debug("SWIS HTML parse error: %s", e)
    return result


def _parse_swis_regex(html_content: str) -> Dict[str, str]:
    """Regex-парсинг SWIS без BeautifulSoup."""
    result: Dict[str, str] = {}
    # Паттерн: <td>Ключ</td><td>Значение</td>
    pattern = re.compile(
        r"<td[^>]*>\s*([^<]{3,200})\s*</td>\s*<td[^>]*>\s*([^<]{1,1000})\s*</td>",
        re.IGNORECASE | re.DOTALL
    )
    for m in pattern.finditer(html_content):
        key = re.sub(r"<[^>]+>", "", m.group(1)).strip()
        val = re.sub(r"<[^>]+>", "", m.group(2)).strip()
        if key and val and len(key) < 200:
            result[key] = val
    return result


def _extract_swis_special_fields(soup: Any, result: Dict[str, str]) -> None:
    """Извлекает специальные поля SWIS из структурированного HTML."""
    # Регистрационный номер — часто в h1 или заголовке таблицы
    for tag in ("h1", "h2", "h3", "title"):
        el = soup.find(tag)
        if el:
            text = el.get_text(strip=True)
            if re.search(r"ЕАЭС|РОСС|KG\d", text):
                result.setdefault("Регистрационный номер документа", text)
                break

    # Статус — часто в отдельном блоке
    status_candidates = [
        "Признак действия", "Статус", "Действие",
        "Действует", "Прекращена",
    ]
    for cand in status_candidates:
        if cand in result:
            result.setdefault("document_status_ru", result[cand])
            break


def normalize_swis_result(raw: Dict[str, str]) -> Dict[str, str]:
    """
    Нормализует сырые данные SWIS в формат полей OzonResultRow.
    Маппинг SWIS-полей → OzonResultRow-полей.
    """
    SWIS_FIELD_MAP = {
        "Регистрационный номер документа": "certificate_number",
        "Дата регистрации": "document_date_start",
        "Дата окончания действия": "document_date_end",
        "Признак действия": "document_status",
        "Полное наименование": "applicant_name",
        "ИНН": "applicant_inn",
        "Полное наименование организации-изготовителя": "manufacturer_name",
        "Полное наименование продукции и сведения для идентификации": "certificate_product_name",
        "Обозначение НПА/ТНПА": "technical_regulation",
        "Схема декларации": "scheme",
        "Код ТН ВЭД ТС": "tnved",
    }
    result: Dict[str, str] = {"document_type": "Декларация (SWIS)"}
    for swis_key, row_key in SWIS_FIELD_MAP.items():
        val = raw.get(swis_key, "")
        if val:
            result[row_key] = val
    # Сертификат или декларация
    reg_num = result.get("certificate_number", "")
    if reg_num:
        if "Д" in reg_num or "декл" in reg_num.lower():
            result["document_type"] = "Декларация (SWIS)"
        elif "С" in reg_num or "серт" in reg_num.lower():
            result["document_type"] = "Сертификат (SWIS)"
    return result


async def fetch_swis_document(
    url: str,
    session: Optional[aiohttp.ClientSession] = None,
    timeout_sec: float = 10.0,
) -> Optional[Dict[str, str]]:
    """
    Асинхронно загружает и парсит документ SWIS/TULPAR.
    swis.trade.kg не использует Cloudflare, достаточен стандартный User-Agent.

    Источник: research_apis.md раздел 4.3 — SWIS не использует CAPTCHA
    на страницах просмотра документов.
    """
    if not AIOHTTP_OK:
        return None

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/136.0.6478.127 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9",
    }

    close_session = False
    if session is None:
        connector = aiohttp.TCPConnector(ssl=False)
        timeout = aiohttp.ClientTimeout(total=timeout_sec)
        session = aiohttp.ClientSession(connector=connector, timeout=timeout)
        close_session = True

    try:
        async with session.get(url, headers=headers) as resp:
            if resp.status != 200:
                log.debug("SWIS вернул %d для %s", resp.status, url)
                return None
            html_content = await resp.text(encoding="utf-8", errors="replace")
        raw = parse_swis_document_html(html_content)
        return normalize_swis_result(raw) if raw else None
    except Exception as e:
        log.debug("SWIS fetch error для %s: %s", url, e)
        return None
    finally:
        if close_session:
            await session.close()


async def enrich_with_swis_data(
    rows: List[OzonResultRow],
    workers: int = 5,
    timeout_sec: float = 10.0,
) -> List[OzonResultRow]:
    """
    Мини-этап: для строк с registry_url на swis.trade.kg
    загружает данные документа через HTML-парсинг.
    """
    if not AIOHTTP_OK:
        return rows

    swis_rows = [
        (i, r) for i, r in enumerate(rows)
        if r.registry_url and "swis.trade.kg" in r.registry_url
        and not r.certificate_number
    ]

    if not swis_rows:
        return rows

    log.info("Обогащение данными SWIS: %d строк", len(swis_rows))

    connector = aiohttp.TCPConnector(limit=workers * 2, ssl=False)
    timeout = aiohttp.ClientTimeout(total=timeout_sec)
    sem = asyncio.Semaphore(workers)

    async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
        async def process_swis_row(idx: int, row: OzonResultRow) -> None:
            async with sem:
                await asyncio.sleep(random.uniform(0.5, 1.5))  # вежливая задержка
                result = await fetch_swis_document(row.registry_url, session=session)
                if result:
                    for field_name, value in result.items():
                        if hasattr(row, field_name) and value:
                            if not getattr(row, field_name, ""):
                                setattr(row, field_name, value)

        tasks = [process_swis_row(i, r) for i, r in swis_rows]
        await asyncio.gather(*tasks)

    return rows


# ---------------------------------------------------------------------------
# Модуль 3: Belgiss.by парсер (заглушка с HTML)
# ---------------------------------------------------------------------------

async def fetch_belgiss_document(
    url: str,
    timeout_sec: float = 10.0,
) -> Optional[Dict[str, str]]:
    """
    Загружает данные документа из belgiss.by или tsouz.belgiss.by.
    Оба сайта — классический HTML без JSON API.
    """
    if not AIOHTTP_OK:
        return None

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "ru-RU,ru;q=0.9",
    }

    try:
        connector = aiohttp.TCPConnector(ssl=False)
        timeout = aiohttp.ClientTimeout(total=timeout_sec)
        async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
            async with session.get(url, headers=headers) as resp:
                if resp.status != 200:
                    return None
                html_text = await resp.text(encoding="utf-8", errors="replace")

        # Извлечение данных — структура Belgiss аналогична SWIS (таблицы)
        raw = _parse_swis_regex(html_text)  # структура идентична
        if raw:
            result: Dict[str, str] = {"document_type": "Документ (Belgiss)"}
            for key, val in raw.items():
                result[key] = val
            # Ищем ключевые поля
            for k, v in raw.items():
                k_low = k.lower()
                if "номер" in k_low and not result.get("certificate_number"):
                    result["certificate_number"] = v
                elif "дата" in k_low and "начал" in k_low:
                    result["document_date_start"] = v
                elif "дата" in k_low and ("оконч" in k_low or "окончан" in k_low):
                    result["document_date_end"] = v
                elif "статус" in k_low or "действие" in k_low:
                    result["document_status"] = v
            return result
    except Exception as e:
        log.debug("Belgiss fetch error для %s: %s", url, e)
    return None


# ---------------------------------------------------------------------------
# Модуль 4: Универсальный registry enricher
# ---------------------------------------------------------------------------

async def enrich_registry_data(
    rows: List[OzonResultRow],
    workers: int = 10,
    timeout_sec: float = 10.0,
    fsa: bool = True,
    swis: bool = True,
    belgiss: bool = True,
) -> List[OzonResultRow]:
    """
    Полный цикл обогащения данными из реестров.
    Обрабатывает все реестровые URL параллельно по типам:
    - ФСА (pub.fsa.gov.ru) через curl_cffi
    - SWIS (swis.trade.kg) через HTML
    - Belgiss (belgiss.by, tsouz.belgiss.by) через HTML

    Обновляет поля OzonResultRow на месте и возвращает тот же список.
    """
    registry_rows = [r for r in rows if r.registry_url]
    if not registry_rows:
        log.info("Нет строк с registry_url — обогащение реестрами пропущено")
        return rows

    log.info("Обогащение реестрами: %d строк", len(registry_rows))

    tasks = []

    if fsa:
        tasks.append(enrich_with_fsa_data(rows, workers=workers, timeout_sec=timeout_sec))
    if swis:
        tasks.append(enrich_with_swis_data(rows, workers=min(workers, 5), timeout_sec=timeout_sec))

    if belgiss:
        belgiss_rows = [
            (i, r) for i, r in enumerate(rows)
            if r.registry_url and ("belgiss.by" in r.registry_url)
            and not r.certificate_number
        ]
        if belgiss_rows:
            sem = asyncio.Semaphore(min(workers, 5))

            async def _do_belgiss(idx: int, row: OzonResultRow) -> None:
                async with sem:
                    result = await fetch_belgiss_document(row.registry_url, timeout_sec=timeout_sec)
                    if result:
                        for fn, v in result.items():
                            if hasattr(row, fn) and v and not getattr(row, fn, ""):
                                setattr(row, fn, v)

            tasks.append(asyncio.gather(*[_do_belgiss(i, r) for i, r in belgiss_rows]))

    if tasks:
        await asyncio.gather(*tasks)

    return rows


# ---------------------------------------------------------------------------
# Модуль 5: Утилиты пакетной обработки
# ---------------------------------------------------------------------------

async def run_full_pipeline(
    query: str,
    limit: int = 100,
    workers: int = 10,
    headless: bool = True,
    output: str = "ozon_result.xlsx",
    expiry_warning_days: int = 30,
    make_report_xlsx: bool = True,
    run_log: str = "",
    enrich_registries: bool = True,
    fsa_enrich: bool = True,
    swis_enrich: bool = True,
    belgiss_enrich: bool = True,
    registry_timeout_sec: float = 10.0,
) -> List[OzonResultRow]:
    """
    Полный пайплайн с обогащением из реестров:
    1. Поиск товаров на Ozon
    2. Загрузка карточек с документами
    3. Обогащение данными из реестров ФСА / SWIS / Belgiss
    4. Сохранение CSV + XLSX

    Это расширенная версия функции run() с дополнительным шагом обогащения.
    """
    started_at = time.time()
    output_path = Path(output)

    log.info("=== Ozon Full Pipeline v27 ===")
    log.info("Запрос: '%s', лимит: %d, воркеры: %d", query, limit, workers)

    # Этап 1: поиск
    print(f"[Ozon Pipeline] Поиск: '{query}', лимит: {limit}")
    search_items = await collect_search_results(query=query, limit=limit, workers=workers)

    if not search_items:
        print("[Ozon Pipeline] Нет результатов поиска.")
        return []

    print(f"[Ozon Pipeline] Найдено {len(search_items)} товаров")

    # Этап 2: загрузка карточек
    results = await enrich_cards(
        search_items=search_items,
        query=query,
        workers=workers,
        headless=headless,
    )
    print(f"[Ozon Pipeline] Карточки обработаны: {len(results)}")

    # Этап 3: обогащение реестрами
    if enrich_registries:
        n_with_reg = sum(1 for r in results if r.registry_url)
        print(f"[Ozon Pipeline] Обогащение реестрами: {n_with_reg} строк с реестровыми ссылками")
        results = await enrich_registry_data(
            results,
            workers=workers,
            timeout_sec=registry_timeout_sec,
            fsa=fsa_enrich,
            swis=swis_enrich,
            belgiss=belgiss_enrich,
        )

    # Этап 4: сохранение
    csv_path = output_path.with_suffix(".csv")
    save_csv(results, csv_path)

    xlsx_path = output_path if output_path.suffix.lower() in (".xlsx", ".xls") \
        else output_path.with_suffix(".xlsx")
    save_xlsx(
        results, xlsx_path,
        expiry_warning_days=expiry_warning_days,
        make_report_xlsx=make_report_xlsx,
    )

    if run_log:
        log_path = Path(run_log)
    else:
        log_path = None
    _write_run_log(results, xlsx_path, expiry_warning_days, started_at, log_path)

    elapsed = int(time.time() - started_at)
    print(f"[Ozon Pipeline] Завершено за {elapsed}с. Файлы: {csv_path.name}, {xlsx_path.name}")

    return results


def load_results_from_csv(path: Union[str, Path]) -> List[OzonResultRow]:
    """
    Загружает результаты из CSV-файла, созданного ozon_parser или main_v39.
    Полезно для resume и постобработки.
    """
    path = Path(path)
    if not path.exists():
        log.warning("CSV файл не найден: %s", path)
        return []

    rows: List[OzonResultRow] = []
    known_fields = set(OzonResultRow.__dataclass_fields__.keys())

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for csv_row in reader:
                # Отфильтровываем только известные поля
                data: Dict[str, Any] = {}
                for field_name in known_fields:
                    val = csv_row.get(field_name, "")
                    field_def = OzonResultRow.__dataclass_fields__.get(field_name)
                    if field_def is None:
                        continue
                    # Конвертация типов
                    ft = field_def.type
                    if ft in (int, "int") or "int" in str(ft):
                        data[field_name] = _safe_int(val)
                    elif ft in (float, "float") or "float" in str(ft):
                        data[field_name] = _safe_float(val)
                    else:
                        data[field_name] = str(val or "")
                rows.append(OzonResultRow(**data))
        log.info("Загружено %d строк из %s", len(rows), path)
    except Exception as e:
        log.error("Ошибка загрузки CSV %s: %s", path, e)

    return rows


def merge_results(
    base: List[OzonResultRow],
    updates: List[OzonResultRow],
    key_field: str = "ozon_sku",
) -> List[OzonResultRow]:
    """
    Объединяет два списка результатов по key_field.
    Записи из updates перезаписывают записи из base при совпадении ключа.
    Полезно для resume-сценариев.
    """
    base_dict: Dict[str, OzonResultRow] = {}
    for r in base:
        k = getattr(r, key_field, "") or str(r.nm_id)
        if k:
            base_dict[k] = r

    for r in updates:
        k = getattr(r, key_field, "") or str(r.nm_id)
        if k:
            base_dict[k] = r

    return list(base_dict.values())


def filter_by_status(
    rows: List[OzonResultRow],
    statuses: Optional[List[str]] = None,
    exclude_statuses: Optional[List[str]] = None,
) -> List[OzonResultRow]:
    """
    Фильтрует результаты по статусу.

    Примеры:
        # Только с реестровой ссылкой:
        good = filter_by_status(rows, statuses=[STATUS_LINK_COLLECTED])

        # Без ошибок:
        clean = filter_by_status(rows, exclude_statuses=[STATUS_ERROR, STATUS_TIMEOUT])
    """
    result = rows
    if statuses is not None:
        statuses_set = set(statuses)
        result = [r for r in result if r.status in statuses_set]
    if exclude_statuses is not None:
        exclude_set = set(exclude_statuses)
        result = [r for r in result if r.status not in exclude_set]
    return result


def get_stats(rows: List[OzonResultRow]) -> Dict[str, Any]:
    """
    Вычисляет статистику по списку результатов.
    Возвращает dict с ключевыми метриками.
    """
    if not rows:
        return {"total": 0}

    total = len(rows)
    status_counts: Dict[str, int] = {}
    host_counts: Dict[str, int] = {}
    orig_counts: Dict[str, int] = {}
    risk_counts: Dict[str, int] = {
        EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
        EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0
    }

    for r in rows:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
        host = r.registry_host or "—"
        host_counts[host] = host_counts.get(host, 0) + 1
        orig = r.is_original or "—"
        orig_counts[orig] = orig_counts.get(orig, 0) + 1
        _, risk = _compute_expiry(r, 30)
        risk_counts[risk] = risk_counts.get(risk, 0) + 1

    with_registry = sum(1 for r in rows if r.registry_url)
    without_docs = sum(1 for r in rows if r.status == STATUS_NO_DOCS)
    errors = sum(1 for r in rows if r.status in (STATUS_ERROR, STATUS_TIMEOUT))

    return {
        "total": total,
        "with_registry_url": with_registry,
        "without_docs": without_docs,
        "errors": errors,
        "registry_coverage_pct": round(with_registry * 100.0 / total, 1),
        "status_distribution": dict(sorted(status_counts.items(), key=lambda x: -x[1])),
        "registry_hosts": dict(sorted(host_counts.items(), key=lambda x: -x[1])),
        "original_mark": dict(sorted(orig_counts.items(), key=lambda x: -x[1])),
        "expiry_risks": risk_counts,
    }


def print_stats(rows: List[OzonResultRow]) -> None:
    """Выводит статистику в консоль."""
    stats = get_stats(rows)
    total = stats["total"]
    if not total:
        print("[Stats] Нет данных")
        return

    print(f"\n{'=' * 60}")
    print(f"  Ozon Parser — Статистика ({total} товаров)")
    print(f"{'=' * 60}")
    print(f"  Реестровых ссылок:   {stats['with_registry_url']} ({stats['registry_coverage_pct']}%)")
    print(f"  Нет документов:      {stats['without_docs']}")
    print(f"  Ошибок/таймаутов:    {stats['errors']}")
    print(f"\n  Статусы:")
    for s, c in stats['status_distribution'].items():
        print(f"    {s:<44} {c}")
    print(f"\n  Реестры:")
    for h, c in stats['registry_hosts'].items():
        print(f"    {h:<30} {c}")
    print(f"\n  Риски по срокам:")
    for k in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
        print(f"    {k:<24} {stats['expiry_risks'].get(k, 0)}")
    print(f"{'=' * 60}\n")


# ---------------------------------------------------------------------------
# Модуль 6: Ozon product URL builder и slug utilities
# ---------------------------------------------------------------------------

def ozon_product_url_from_sku(sku: Union[str, int]) -> str:
    """Строит URL карточки по SKU. Без slug использует числовой ID."""
    return f"https://www.ozon.ru/product/{sku}/"


def ozon_product_path_from_sku(sku: Union[str, int]) -> str:
    """Строит path для composer-api по SKU."""
    return f"/product/{sku}/"


def ozon_search_path(query: str, page: int = 1) -> str:
    """Строит path поиска для composer-api."""
    encoded = urllib.parse.quote(query)
    return f"/search/?text={encoded}&page={page}"


def ozon_seller_path(seller_id: Union[str, int], page: int = 1) -> str:
    """Строит path страницы продавца для composer-api."""
    return f"/seller/{seller_id}/products/?page={page}"


def extract_sku_list_from_urls(urls: List[str]) -> List[str]:
    """
    Извлекает SKU из списка URLs товаров Ozon.
    Полезно для пакетной обработки входного списка.
    """
    skus = []
    for url in urls:
        sku = _extract_sku_from_url(url)
        if sku:
            skus.append(sku)
    return list(dict.fromkeys(skus))  # dedupe


async def parse_multiple_skus(
    skus: List[str],
    query: str = "",
    workers: int = 10,
    headless: bool = True,
) -> List[OzonResultRow]:
    """
    Парсит список SKU (артикулов) Ozon без этапа поиска.
    Полезно когда список артикулов уже известен.

    Пример:
        rows = asyncio.run(parse_multiple_skus(["397529235", "123456789"]))
    """
    # Создаём минимальные search_items из SKU
    search_items = [
        {
            "sku": str(sku),
            "title": "",
            "url": ozon_product_url_from_sku(sku),
            "slug": "",
            "price": 0.0,
            "finalPrice": 0.0,
            "brand": "",
        }
        for sku in skus
    ]
    return await enrich_cards(
        search_items=search_items,
        query=query,
        workers=workers,
        headless=headless,
    )


# ---------------------------------------------------------------------------
# Модуль 7: Валидация и нормализация данных
# ---------------------------------------------------------------------------

_CERT_NUMBER_PATTERNS: List[re.Pattern] = [
    # ЕАЭС / ТС сертификат
    re.compile(r"(ЕАЭС\s+[A-ZА-Яa-zа-я0-9./]+(?:\s+[A-Z0-9./]+)?)", re.IGNORECASE),
    # РОСС сертификат
    re.compile(r"(РОСС\s+[A-Z]{2}\.(?:[А-Яа-я]{3,6})?\d+(?:\.\w+)*)", re.IGNORECASE),
    # ТР ТС / ТР КР стандарт
    re.compile(r"(ТР\s+(?:ТС|КР)\s+\d{3}/\d{4})", re.IGNORECASE),
    # Декларация
    re.compile(r"(ЕАЭС\s+[A-Z]{2}\d+[./]\d+[./][Д]\d+)", re.IGNORECASE),
]


def extract_cert_numbers_from_text(text: str) -> List[str]:
    """
    Извлекает номера сертификатов/деклараций из произвольного текста.
    Полезно для поиска по описанию или характеристикам.
    """
    found: List[str] = []
    for pattern in _CERT_NUMBER_PATTERNS:
        for m in pattern.finditer(text):
            num = m.group(1).strip()
            if num and len(num) >= 5:
                found.append(num)
    return list(dict.fromkeys(found))


def validate_result_row(row: OzonResultRow) -> List[str]:
    """
    Валидирует OzonResultRow и возвращает список предупреждений.
    Полезно для QA перед финальным сохранением.
    """
    warnings: List[str] = []

    if not row.product_name:
        warnings.append("product_name пустое")
    if not row.ozon_sku and not row.nm_id:
        warnings.append("нет ни ozon_sku, ни nm_id")
    if row.registry_url and not _is_registry_url(row.registry_url):
        warnings.append(f"registry_url не является реестровым: {row.registry_url}")
    if row.status == STATUS_LINK_COLLECTED and not row.registry_url:
        warnings.append("статус ССЫЛКА СОБРАНА но registry_url пустой")
    if row.document_date_end and not _parse_date_loose(row.document_date_end):
        warnings.append(f"не удалось разобрать document_date_end: {row.document_date_end}")
    if row.price_rub < 0:
        warnings.append(f"отрицательная price_rub: {row.price_rub}")
    if row.rating and not (0 <= row.rating <= 5):
        warnings.append(f"рейтинг вне диапазона [0..5]: {row.rating}")

    return warnings


def validate_results(rows: List[OzonResultRow]) -> Dict[str, List[str]]:
    """
    Валидирует все строки. Возвращает dict {sku: [предупреждения]}.
    Строки без предупреждений не включаются в результат.
    """
    result: Dict[str, List[str]] = {}
    for row in rows:
        w = validate_result_row(row)
        if w:
            key = row.ozon_sku or str(row.nm_id) or row.product_name[:20]
            result[key] = w
    return result


# ---------------------------------------------------------------------------
# Модуль 8: CLI батч-режим — обработка списка запросов или SKU из файла
# ---------------------------------------------------------------------------

async def run_batch_queries(
    queries: List[str],
    limit_per_query: int = 50,
    workers: int = 10,
    headless: bool = True,
    output_dir: str = ".",
    expiry_warning_days: int = 30,
    make_report_xlsx: bool = True,
) -> Dict[str, List[OzonResultRow]]:
    """
    Пакетная обработка нескольких поисковых запросов.
    Для каждого запроса сохраняет отдельный CSV+XLSX.
    Также сохраняет сводный файл со всеми результатами.

    Возвращает dict {query: [OzonResultRow]}.
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    all_results: Dict[str, List[OzonResultRow]] = {}
    combined: List[OzonResultRow] = []
    started_at = time.time()

    for i, query in enumerate(queries, 1):
        print(f"\n[Batch {i}/{len(queries)}] Запрос: '{query}'")
        # Безопасное имя файла из запроса
        safe_query = re.sub(r"[^\w\s-]", "", query)[:40].strip().replace(" ", "_")
        out_file = output_path / f"ozon_{safe_query}.xlsx"

        try:
            results = await run(
                query=query,
                limit=limit_per_query,
                workers=workers,
                headless=headless,
                output=str(out_file),
                expiry_warning_days=expiry_warning_days,
                make_report_xlsx=make_report_xlsx,
            )
            all_results[query] = results
            combined.extend(results)
            print(f"[Batch {i}/{len(queries)}] Готово: {len(results)} строк → {out_file.name}")
        except Exception as e:
            log.error("[Batch] Ошибка для запроса '%s': %s", query, e)
            all_results[query] = []

        # Небольшая пауза между запросами
        if i < len(queries):
            await asyncio.sleep(random.uniform(1.0, 3.0))

    # Сводный файл
    if combined:
        combined_path = output_path / "ozon_combined.xlsx"
        save_csv(combined, combined_path.with_suffix(".csv"))
        save_xlsx(
            combined, combined_path,
            expiry_warning_days=expiry_warning_days,
            make_report_xlsx=make_report_xlsx,
        )
        elapsed = int(time.time() - started_at)
        print(f"\n[Batch] Итого: {len(combined)} строк из {len(queries)} запросов")
        print(f"[Batch] Сводный файл: {combined_path}")
        print(f"[Batch] Время: {elapsed}с")
        print_stats(combined)

    return all_results


async def run_batch_skus(
    skus: List[str],
    query: str = "batch",
    workers: int = 10,
    headless: bool = True,
    output: str = "ozon_result.xlsx",
    expiry_warning_days: int = 30,
    make_report_xlsx: bool = True,
) -> List[OzonResultRow]:
    """
    Пакетная обработка списка SKU (артикулов).
    Полезно когда список артикулов известен заранее (например, из другого парсера).
    """
    log.info("Пакетная обработка %d SKU", len(skus))
    results = await parse_multiple_skus(skus, query=query, workers=workers, headless=headless)
    output_path = Path(output)
    save_csv(results, output_path.with_suffix(".csv"))
    xlsx_path = output_path if output_path.suffix.lower() in (".xlsx", ".xls") \
        else output_path.with_suffix(".xlsx")
    save_xlsx(
        results, xlsx_path,
        expiry_warning_days=expiry_warning_days,
        make_report_xlsx=make_report_xlsx,
    )
    print_stats(results)
    return results


# ---------------------------------------------------------------------------
# Модуль 9: Ozon GEO (геолокация) утилиты
# ---------------------------------------------------------------------------

# Гео-коды из research_apis.md раздел 1.2
OZON_GEO_CODES: Dict[str, str] = {
    "Москва": "0c5b2444-70a0-4932-980c-b4dc0d3f02b5",
    "Санкт-Петербург": "c2deb16a-0330-4f05-821f-1d09c93331e6",
    "Новосибирск": "8dea00e3-9aab-4d8e-887c-ef2aaa546456",
}


def build_geo_cookie(city: str) -> Optional[str]:
    """
    Строит cookie select_location для задания геолокации.
    Поддерживаемые города: Москва, Санкт-Петербург, Новосибирск.
    Принимает также прямой UUID.
    """
    code = OZON_GEO_CODES.get(city)
    if code:
        return f"select_location={code}"
    # Проверка что передали UUID напрямую
    uuid_re = re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.I)
    if uuid_re.match(city.strip()):
        return f"select_location={city.strip()}"
    return None


def get_available_geo_cities() -> List[str]:
    """Возвращает список доступных городов для геолокации."""
    return list(OZON_GEO_CODES.keys())


# ---------------------------------------------------------------------------
# Модуль 10: Диагностика и health-check
# ---------------------------------------------------------------------------

async def health_check() -> Dict[str, Any]:
    """
    Проверяет доступность Ozon API и зависимостей.
    Возвращает dict с результатами проверок.
    """
    results: Dict[str, Any] = {
        "timestamp": now_iso(),
        "dependencies": {
            "aiohttp": AIOHTTP_OK,
            "openpyxl": OPENPYXL_OK,
            "playwright": PLAYWRIGHT_OK,
            "curl_cffi": CURL_CFFI_OK,
            "beautifulsoup4": BS4_OK,
        },
        "api_endpoints": {},
    }

    if not AIOHTTP_OK:
        results["status"] = "DEGRADED — aiohttp не установлен"
        return results

    # Проверка Ozon API
    try:
        connector = aiohttp.TCPConnector(ssl=False)
        timeout = aiohttp.ClientTimeout(total=8.0)
        async with aiohttp.ClientSession(connector=connector, timeout=timeout) as session:
            # Тест поиска
            params = {"url": "/search/?text=тест&page=1"}
            async with session.get(
                OZON_API_BASE, params=params, headers=OZON_HEADERS
            ) as resp:
                results["api_endpoints"]["ozon_mobile_api"] = {
                    "url": OZON_API_BASE,
                    "status_code": resp.status,
                    "ok": resp.status == 200,
                }
                if resp.status == 200:
                    data = await resp.json(content_type=None)
                    results["api_endpoints"]["ozon_mobile_api"]["has_widget_states"] = \
                        "widgetStates" in data
    except Exception as e:
        results["api_endpoints"]["ozon_mobile_api"] = {"error": str(e), "ok": False}

    # Проверка SWIS
    try:
        async with aiohttp.ClientSession(
            connector=aiohttp.TCPConnector(ssl=False),
            timeout=aiohttp.ClientTimeout(total=8.0)
        ) as session:
            async with session.get("https://swis.trade.kg/") as resp:
                results["api_endpoints"]["swis_trade_kg"] = {
                    "status_code": resp.status,
                    "ok": resp.status == 200,
                }
    except Exception as e:
        results["api_endpoints"]["swis_trade_kg"] = {"error": str(e), "ok": False}

    all_ok = all(
        ep.get("ok", False)
        for ep in results["api_endpoints"].values()
    )
    results["status"] = "OK" if all_ok else "PARTIAL"
    return results


def print_health_check(results: Dict[str, Any]) -> None:
    """Выводит результаты health-check в консоль."""
    print(f"\n{'=' * 55}")
    print(f"  Ozon Parser Health Check — {results.get('timestamp', '')}")
    print(f"{'=' * 55}")
    print("  Зависимости:")
    for dep, ok in results.get("dependencies", {}).items():
        print(f"    {dep:<20} {'✓' if ok else '✗ MISSING'}")
    print("  API Endpoints:")
    for name, info in results.get("api_endpoints", {}).items():
        if info.get("ok"):
            sc = info.get("status_code", "?")
            extra = " [widgetStates: OK]" if info.get("has_widget_states") else ""
            print(f"    {name:<30} HTTP {sc}{extra}")
        else:
            err = info.get("error", f"HTTP {info.get('status_code', '?')}")
            print(f"    {name:<30} ОШИБКА: {err}")
    print(f"  Статус: {results.get('status', 'UNKNOWN')}")
    print(f"{'=' * 55}\n")


# ---------------------------------------------------------------------------
# Версия и мета-информация
# ---------------------------------------------------------------------------

MODULE_VERSION = "27.0.0"
MODULE_DESCRIPTION = "Ozon parser — production ready, CLI-compatible with main_v39.py"
MODULE_DEPENDENCIES = [
    "aiohttp>=3.8.0",
    "openpyxl>=3.0.0",
    "playwright>=1.30.0 (optional, for Cloudflare fallback)",
    "curl_cffi>=0.5.0 (optional, for FSA HTTP)",
    "beautifulsoup4>=4.9.0 (optional, for SWIS HTML)",
]


def module_info() -> Dict[str, Any]:
    """Возвращает метаинформацию о модуле."""
    return {
        "version": MODULE_VERSION,
        "description": MODULE_DESCRIPTION,
        "app_version": APP_VERSION,
        "dependencies": MODULE_DEPENDENCIES,
        "available": {
            "aiohttp": AIOHTTP_OK,
            "openpyxl": OPENPYXL_OK,
            "playwright": PLAYWRIGHT_OK,
            "curl_cffi": CURL_CFFI_OK,
            "beautifulsoup4": BS4_OK,
        },
        "api_base": OZON_API_BASE,
        "allowed_registry_hosts": sorted(ALLOWED_REGISTRY_HOSTS),
        "status_codes": {
            "LINK_COLLECTED": STATUS_LINK_COLLECTED,
            "NO_DOCS": STATUS_NO_DOCS,
            "NO_REGISTRY_LINK": STATUS_NO_REGISTRY_LINK,
            "TIMEOUT": STATUS_TIMEOUT,
            "ERROR": STATUS_ERROR,
            "CF_BLOCK": STATUS_CF_BLOCK,
        },
    }

