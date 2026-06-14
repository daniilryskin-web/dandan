"""
Тест (issue v49): сохранение result.xlsx не должно биться при отмене save() на
середине записи (saver_task.cancel()) + параллельных сохранениях. Раньше отмена
освобождала asyncio-лок, а фоновый поток продолжал писать тот же .tmp → второй
save писал в него же параллельно → битый zip («invalid code lengths set»).

Проверяем:
  • много параллельных save() + отмена части из них → итоговый xlsx ВАЛИДЕН
    (openpyxl открывает, лист читается, нет усечения);
  • запись идёт через 1-поточный executor (сериализация) и уникальные .tmp.

Запуск:  python3 tests/test_save_integrity.py
"""
import asyncio
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import types
sys.modules.setdefault("webview", types.ModuleType("webview"))
import main_v39 as m  # noqa: E402

RESULTS = []


def check(name, cond):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}")


async def _scenario(td: Path) -> bool:
    store = m.ResultStore(td / "result.xlsx", None, make_report_xlsx=True)
    # наполняем достаточно строк, чтобы запись занимала заметное время
    for i in range(4000):
        await store.add(m.ResultRow(
            query="детская одежда", nm_id=i + 1,
            product_name=f"Кофточка для новорождённых {i}", brand="b",
            subject="Кофточки", product_url="http://x", status="OK",
            docs_verified="Да"))
    # запускаем пачку сохранений и ЧАСТЬ из них отменяем на середине —
    # имитируем saver_task.cancel() во время записи
    tasks = [asyncio.create_task(store.save()) for _ in range(6)]
    await asyncio.sleep(0.02)
    for t in tasks[:4]:
        t.cancel()
    await asyncio.gather(*tasks, return_exceptions=True)
    # финальное сохранение (как в конце прогона)
    await store.save(final=True)
    # даём фоновому executor'у добить возможные отменённые записи
    await asyncio.sleep(0.2)
    return (td / "result.xlsx").exists()


def main():
    src = (Path(__file__).resolve().parents[1] / "main_v39.py").read_text(encoding="utf-8")
    check("save через выделенный 1-поточный executor", "_save_executor" in src
          and "max_workers=1" in src and "run_in_executor(self._save_executor" in src)
    check("уникальные .tmp на запись", "_unique_tmp" in src)

    with tempfile.TemporaryDirectory() as d:
        td = Path(d)
        ok = asyncio.get_event_loop().run_until_complete(_scenario(td))
        check("result.xlsx создан", ok)
        # ГЛАВНОЕ: файл ВАЛИДНЫЙ (не битый zip)
        import openpyxl
        valid = False
        rowcount = 0
        try:
            wb = openpyxl.load_workbook(td / "result.xlsx", read_only=True)
            ws = wb["Подробности"] if "Подробности" in wb.sheetnames else wb[wb.sheetnames[0]]
            rowcount = sum(1 for _ in ws.iter_rows())
            wb.close()
            valid = True
        except Exception as e:
            print("   load error:", e)
        check("итоговый xlsx ОТКРЫВАЕТСЯ (не битый zip)", valid)
        check("в финальном файле все строки (≈4000)", rowcount >= 4000)
        # никаких .tmp не осталось
        leftovers = list(td.glob("*.tmp"))
        check("временные .tmp убраны", not leftovers)


if __name__ == "__main__":
    main()
    p, t = sum(RESULTS), len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {p}/{t} прошло")
    sys.exit(0 if p == t else 1)
