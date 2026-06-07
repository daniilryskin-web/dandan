"""
Офлайн-тест защиты записи в Excel (_xlsx_safe).

Спарсенный текст из карточек/реестров часто содержит управляющие символы,
на которых openpyxl падает с IllegalCharacterError и роняет ВЕСЬ отчёт
(симптом «данные сохраняются криво»). Проверяем, что _xlsx_safe это чинит.

Запуск:  python3 tests/test_xlsx_safe.py
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.cell.cell import IllegalCharacterError  # noqa: E402

import main_v39 as m  # noqa: E402

RESULTS = []


def check(name, cond, extra=""):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}  {extra}")


NASTY = "Носки\x00\x07 хлопок\nвторая строка\tтаб   лишние   пробелы  "

# 1) Без защиты openpyxl действительно падает на управляющих символах
raw_crashed = False
try:
    wb = Workbook()
    wb.active.cell(row=1, column=1, value=NASTY)
    wb.save(io.BytesIO())
except IllegalCharacterError:
    raw_crashed = True
check("raw openpyxl падает на control-символах (баг есть)", raw_crashed)

# 2) С _xlsx_safe значение сохраняется и читается обратно без мусора
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1, value=m._xlsx_safe(NASTY))
ws.cell(row=2, column=1, value=m._xlsx_safe(1234.5))   # число остаётся числом
ws.cell(row=3, column=1, value=m._xlsx_safe(0))
ws.cell(row=4, column=1, value=m._xlsx_safe(None))
ws.cell(row=5, column=1, value=m._xlsx_safe(True))
buf = io.BytesIO()
saved_ok = True
try:
    wb.save(buf)
except Exception as e:  # pragma: no cover
    saved_ok = False
    print("save error:", e)
check("после _xlsx_safe файл сохраняется без ошибок", saved_ok)

buf.seek(0)
s = load_workbook(buf).active
v1 = s.cell(1, 1).value or ""
check("control-символы вычищены", "\x00" not in v1 and "\x07" not in v1, repr(v1))
check("переводы строк/табы схлопнуты в пробел", "\n" not in v1 and "\t" not in v1, repr(v1))
check("кратные пробелы схлопнуты", "  " not in v1, repr(v1))
check("число осталось числом", s.cell(2, 1).value == 1234.5)
check("ноль сохранён как число 0", s.cell(3, 1).value == 0)
check("длина обрезается до лимита Excel", len(m._xlsx_safe("х" * 40000)) <= 32000)

if __name__ == "__main__":
    passed = sum(RESULTS)
    total = len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {passed}/{total} прошло")
    sys.exit(0 if passed == total else 1)
