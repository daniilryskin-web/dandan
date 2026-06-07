#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WB Brand Checker

Сценарий:
1) Пользователь вводит бренд, например adidas.
2) Программа собирает уникальные карточки WB по поисковой выдаче, фильтрует по бренду.
3) По каждой карточке открывает WB и собирает:
   - название товара;
   - бренд;
   - стоимость;
   - есть ли плашка «Оригинал»;
   - ссылка/ссылки на реестр документов, если есть.
4) Результаты сохраняются в CSV и Excel с autosave/resume.

Важно: ссылки на документы сначала берутся быстрым прямым способом из WB certificate.json:
https://basket-XX.wbbasket.ru/vol.../part.../<nm_id>/info/certificate.json
Если быстрый способ не сработал, программа может fallback-ом пройти строгий UI-путь:
Карточка -> «Характеристики и описание» -> «Документы проверены» -> «Смотреть на сайте».
Случайные внешние URL не принимаются.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import datetime as dt
import html
import json
import os
import random
import re
import sys
import time
import subprocess
import shutil
import tempfile
import urllib.parse
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import urlparse, parse_qs

import aiohttp
try:
    import requests as _requests
except Exception:
    _requests = None
try:
    from curl_cffi import requests as _curl_requests
except Exception:
    _curl_requests = None
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
except Exception:
    async_playwright = None
    PlaywrightTimeoutError = TimeoutError

# -----------------------------
# Constants
# -----------------------------

APP_VERSION = "2026-06-05-v25-reporting"
DEFAULT_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
FSA_DETAILS_LOCK = asyncio.Lock()
FSA_LAST_REQUEST_TS = 0.0

STATUS_LINK_COLLECTED = "ССЫЛКА НА РЕЕСТР СОБРАНА"
STATUS_JSON_NO_DOCS_BUTTON = "ССЫЛКА В JSON WB, КНОПКА ДОКУМЕНТОВ НЕ НАЙДЕНА"
STATUS_NO_DOCS = "НЕТ ДОКУМЕНТОВ"
STATUS_NO_REGISTRY_LINK = "НЕТ ССЫЛКИ НА РЕЕСТР"
STATUS_TIMEOUT = "ТАЙМАУТ"
STATUS_ERROR = "ОШИБКА"

ALLOWED_REGISTRY_HOSTS = {
    "pub.fsa.gov.ru",
    "fsa.gov.ru",
    "swis.trade.kg",
    "trade.kg",
    "belgiss.by",
    "www.belgiss.by",
    "tsouz.belgiss.by",
    "portal.eaeunion.org",
    "eaeunion.org",
    "www.eaeunion.org",
    "tech.eaeunion.org",
}

BLOCKED_HOST_PARTS = (
    "wildberries",
    "wb.ru",
    "wb-bank",
    "eapteka",
    "datatracker.ietf.org",
    "ietf.org",
    "google",
    "yandex",
    "facebook",
    "vk.com",
    "t.me",
    "telegram",
    "doubleclick",
    "adservice",
)

HTTP_URL_RE = re.compile(r"https?://[^\s\"'<>\\)\\]]+", re.IGNORECASE)
UUID_RE = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")
PUB_FSA_NO_SCHEME_RE = re.compile(r"(?:https?://)?pub\.fsa\.gov\.ru/[^\s\"'<>\)\]]+", re.IGNORECASE)
FSA_REL_PATH_RE = re.compile(r"/(?:rds/declaration|rss/certificate)/(?:view|details|card|api)[^\s\"'<>\)\]]*", re.IGNORECASE)

# -----------------------------
# Data
# -----------------------------

@dataclass
class BrandCard:
    nm_id: int
    product_name: str = ""
    brand: str = ""
    subject: str = ""
    price_rub: float = 0.0
    seller_name: str = ""
    supplier_id: str = ""
    source_query: str = ""
    product_url: str = ""

@dataclass
class BrandResult:
    brand_query: str
    nm_id: int
    product_name: str
    brand: str
    subject: str
    price_rub: float
    seller_name: str
    supplier_id: str
    is_original: str
    status: str
    registry_urls: str = ""
    registry_hosts: str = ""
    registry_record_ids: str = ""
    registry_doc_type: str = ""
    registry_doc_number: str = ""
    registry_blank_number: str = ""
    registry_status: str = ""
    registry_status_date: str = ""
    registry_status_basis: str = ""
    registry_date_start: str = ""
    registry_date_end: str = ""
    registry_applicant: str = ""
    registry_applicant_inn: str = ""
    registry_manufacturer: str = ""
    registry_product_group: str = ""
    registry_product_full: str = ""
    registry_tnved: str = ""
    registry_scheme: str = ""
    registry_technical_regulation: str = ""
    registry_evidence: str = ""
    registry_details_source: str = ""
    product_url: str = ""
    details: str = ""
    worker: str = ""
    checked_at: str = ""

# -----------------------------
# Utility
# -----------------------------

def now_iso() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def emit_progress(stage: str, done: int, total: int) -> None:
    """Машиночитаемый прогресс для GUI (см. main_v39.emit_progress)."""
    try:
        print(f"@@PROGRESS@@ stage={stage} done={int(done)} total={int(total)}",
              flush=True)
    except Exception:
        pass

def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("ё", "е").lower()).strip()

def norm_key(s: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", norm_text(s))


def clean_seller_name_py(s: Any) -> str:
    """Финальная защита: никогда не записывать служебную метку вместо имени продавца."""
    t = re.sub(r"\s+", " ", str(s or "")).strip().strip('"\'«»')
    t = re.sub(r"^продавец\s*[:\-–—]?\s*", "", t, flags=re.I).strip()
    t = re.sub(r"\s+(о продавце|перейти к продавцу|задать вопрос).*$", "", t, flags=re.I).strip()
    tl = norm_text(t)
    bad_exact = {
        "", "продавец", "о продавце", "перейти к продавцу", "задать вопрос", "seller",
        "товары продавца", "магазин продавца", "рейтинг продавца", "wildberries продавец",
        "склад wb", "wb",
    }
    if tl in bad_exact:
        return ""
    if len(t) < 2 or len(t) > 120:
        return ""
    if re.match(r"^(доставка|возврат|оплата|покупают|реклама|спонсорский|поставил|доставит|отзыв|отзывы|вопрос|оценк)", tl, flags=re.I):
        return ""
    if re.fullmatch(r"[0-9\s.,₽%+\-–—]+", t):
        return ""
    return t

def str_to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in {"1", "true", "yes", "y", "да", "on"}

def safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return default

def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(str(v).replace(",", "."))
    except Exception:
        return default

def product_url(nm_id: int) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"

def nm_id_from_url_or_text(s: str) -> int:
    m = re.search(r"(?:catalog/|^)(\d{5,12})(?:/|$)", str(s or ""))
    return safe_int(m.group(1)) if m else 0

def hostname(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().split(":")[0].lstrip("www.")
    except Exception:
        return ""

def clean_url(url: str) -> str:
    url = html.unescape(urllib.parse.unquote(str(url or "").strip()))
    return url.rstrip(").,;\"'<> ")

def extract_record_id(url: str) -> str:
    try:
        u = urlparse(url)
        m = UUID_RE.search(u.path)
        if m:
            return m.group(0)
        parts = [p for p in u.path.split("/") if p]
        for i, p in enumerate(parts):
            if p == "view" and i + 1 < len(parts):
                return parts[i + 1]
        nums = re.findall(r"\d{4,}", u.path)
        return nums[-1] if nums else ""
    except Exception:
        return ""

def is_allowed_registry_url(url: str) -> bool:
    if not url:
        return False
    try:
        decoded = urllib.parse.unquote(url.strip())
        parsed = urlparse(decoded)
        if parsed.scheme not in {"http", "https"}:
            return False
        h = parsed.netloc.lower().split(":")[0].lstrip("www.")
        full = (h + parsed.path).lower()
        if any(b in full for b in BLOCKED_HOST_PARTS):
            if "pub.fsa.gov.ru" not in h and "fsa.gov.ru" not in h:
                return False
        if h == "pub.fsa.gov.ru":
            p = parsed.path.lower()
            return any(x in p for x in ("/rds/declaration/", "/rss/certificate/", "/api/v1/rds/", "/api/v1/rss/"))
        if h == "fsa.gov.ru":
            return True
        if h == "swis.trade.kg":
            return parsed.path.lower().startswith("/doc/")
        if h == "trade.kg":
            return True
        if h in {"belgiss.by", "tsouz.belgiss.by", "portal.eaeunion.org", "eaeunion.org", "tech.eaeunion.org"}:
            return True
        return False
    except Exception:
        return False

def extract_urls_from_text(text: str, max_len: int = 200000) -> List[str]:
    if not text:
        return []
    t = str(text)[:max_len]
    out: List[str] = []
    for m in HTTP_URL_RE.finditer(t):
        out.append(clean_url(m.group(0)))
    expanded = list(out)
    for url in out:
        try:
            qs = parse_qs(urlparse(url).query)
            for vals in qs.values():
                for val in vals:
                    val = clean_url(val)
                    if val.startswith("http"):
                        expanded.append(val)
                        expanded.extend(extract_urls_from_text(val, 10000))
        except Exception:
            pass
    seen: Set[str] = set()
    res: List[str] = []
    for u in expanded:
        if u not in seen:
            seen.add(u)
            res.append(u)
    return res

def allowed_urls(urls: Iterable[str]) -> List[str]:
    seen: Set[str] = set()
    res: List[str] = []
    for u in urls:
        u = clean_url(u)
        if is_allowed_registry_url(u) and u not in seen:
            seen.add(u)
            res.append(u)
    return res


def extract_registry_urls_deep(obj: Any, max_depth: int = 8) -> List[str]:
    """Достаёт ссылки на реестр из произвольного JSON/строки.

    WB certificate.json может вернуть как прямую ссылку, так и URL внутри вложенного поля,
    иногда URL бывает закодированным или без схемы. Поэтому обычного HTTP_URL_RE недостаточно.
    """
    raw_parts: List[str] = []

    def add_string(s: Any):
        if s is None:
            return
        t = str(s)
        if not t:
            return
        raw_parts.append(t)
        # Несколько слоёв URL-decoding: WB/реестр иногда кладут ссылку в query-параметр.
        cur = t
        for _ in range(3):
            dec = urllib.parse.unquote(html.unescape(cur))
            if dec == cur:
                break
            raw_parts.append(dec)
            cur = dec

    def walk(x: Any, depth: int):
        if x is None or depth > max_depth:
            return
        if isinstance(x, dict):
            for k, v in x.items():
                add_string(k)
                if isinstance(v, (dict, list)):
                    walk(v, depth + 1)
                else:
                    add_string(v)
        elif isinstance(x, list):
            for v in x[:1000]:
                if isinstance(v, (dict, list)):
                    walk(v, depth + 1)
                else:
                    add_string(v)
        else:
            add_string(x)

    walk(obj, 0)

    candidates: List[str] = []
    for part in raw_parts:
        candidates.extend(extract_urls_from_text(part, 50000))
        for m in PUB_FSA_NO_SCHEME_RE.finditer(part):
            u = m.group(0)
            if not u.lower().startswith("http"):
                u = "https://" + u
            candidates.append(clean_url(u))
        for m in FSA_REL_PATH_RE.finditer(part):
            candidates.append(clean_url("https://pub.fsa.gov.ru" + m.group(0)))
    return allowed_urls(candidates)


def wb_volume_part(nm_id: int) -> Tuple[int, int]:
    return int(nm_id) // 100000, int(nm_id) // 1000


def wb_basket_by_volume(vol: int) -> int:
    """Приблизительная официальная шардировка WB media/static basket.

    Для 198242850 из HAR: vol=1982 -> basket-13. Если WB поменяет диапазоны,
    прямой сбор всё равно попробует соседние/все basket-хосты fallback-ом.
    """
    ranges = [
        (143, 1), (287, 2), (431, 3), (719, 4), (1007, 5),
        (1061, 6), (1115, 7), (1169, 8), (1313, 9), (1601, 10),
        (1655, 11), (1919, 12), (2045, 13), (2189, 14), (2405, 15),
        (2621, 16), (2837, 17), (3053, 18), (3269, 19), (3485, 20),
        (3701, 21), (3917, 22), (4133, 23), (4349, 24), (4565, 25),
        (4781, 26), (4997, 27), (5213, 28), (5429, 29), (10**9, 30),
    ]
    for max_vol, basket in ranges:
        if vol <= max_vol:
            return basket
    return 30



# -----------------------------
# Fast registry details parsers
# -----------------------------

REGISTRY_DETAILS_CACHE: Dict[str, Dict[str, str]] = {}


def _strip_html_to_lines(raw_html: str) -> List[str]:
    """Cheap visible-text extractor for registry pages without external deps."""
    text = re.sub(r"(?is)<(script|style|noscript)[^>]*>.*?</\1>", "\n", raw_html or "")
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(div|p|tr|td|th|li|h1|h2|h3|h4|section|article|label|span)>", "\n", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html.unescape(text)
    out: List[str] = []
    for line in text.splitlines():
        line = re.sub(r"[\t\r\f\v\xa0]+", " ", line)
        line = re.sub(r"\s+", " ", line).strip(" :-–—\t")
        if line:
            out.append(line)
    # Склеиваем единичные мусорные дубли умеренно: порядок важен для label->value.
    return out


def _label_value(lines: List[str], labels: Iterable[str]) -> str:
    norm_labels = [(lbl, norm_text(lbl)) for lbl in labels]
    for i, line in enumerate(lines):
        nl = norm_text(line)
        for raw_label, nlabel in norm_labels:
            if nl == nlabel or nl.startswith(nlabel + " ") or nl.startswith(nlabel + ":"):
                rest = line[len(raw_label):].strip(" :—–-\t")
                if rest:
                    return rest
                for j in range(i + 1, min(len(lines), i + 6)):
                    cand = lines[j].strip(" :—–-\t")
                    if not cand:
                        continue
                    nc = norm_text(cand)
                    # Не возвращаем следующий ярлык вместо значения.
                    if any(nc == x[1] or nc.startswith(x[1] + " ") for x in norm_labels):
                        continue
                    return cand
    return ""


def _slice_section(lines: List[str], start_markers: Iterable[str], end_markers: Iterable[str]) -> List[str]:
    start = -1
    nstarts = [norm_text(x) for x in start_markers]
    nends = [norm_text(x) for x in end_markers]
    for i, line in enumerate(lines):
        nl = norm_text(line)
        if any(m in nl for m in nstarts):
            start = i + 1
            break
    if start < 0:
        return []
    end = len(lines)
    for j in range(start, len(lines)):
        nl = norm_text(lines[j])
        if any(m in nl for m in nends):
            end = j
            break
    return lines[start:end]


def _join_short(values: Iterable[str], max_len: int = 1800) -> str:
    seen: Set[str] = set()
    out: List[str] = []
    for v in values:
        v = re.sub(r"\s+", " ", str(v or "")).strip(" ;,\t")
        if not v:
            continue
        key = norm_key(v)[:120]
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
        if len("; ".join(out)) >= max_len:
            break
    res = "; ".join(out)
    return res[:max_len]


def parse_swis_trade_kg_html(raw_html: str, url: str = "") -> Dict[str, str]:
    """Parse public swis.trade.kg /Doc/{uuid} page into flat fields."""
    lines = _strip_html_to_lines(raw_html)
    all_text = "\n".join(lines)
    lower_all = norm_text(all_text)

    if "сертификат" in lower_all:
        doc_type = "сертификат соответствия"
    elif "деклараци" in lower_all:
        doc_type = "декларация о соответствии"
    else:
        doc_type = "документ"

    applicant_sec = _slice_section(lines, ["Заявитель"], ["Изготовитель", "Сведения о продукции"])
    manufacturer_sec = _slice_section(lines, ["Изготовитель"], ["Сведения о продукции", "Товар 1"])
    product_sec = _slice_section(lines, ["Сведения о продукции"], ["Орган по сертификации", "Сведения об органе", "Документы", "Испытательная лаборатория"])

    goods_full: List[str] = []
    for i, line in enumerate(lines):
        if norm_text(line).startswith("полное наименование продукции"):
            val = line
            # Убираем сам ярлык, если значение в той же строке.
            val = re.sub(r"(?i)^полное\s+наименование\s+продукции\s+и\s+сведения[^)]*\)\s*", "", val).strip(" :—–-")
            # Если значение оказалось в следующей строке.
            if not val or norm_text(val).startswith("полное наименование продукции"):
                chunk: List[str] = []
                for j in range(i + 1, min(len(lines), i + 5)):
                    nxt = lines[j]
                    nn = norm_text(nxt)
                    if nn.startswith("код тн вэд") or re.match(r"^товар\s+\d+", nn):
                        break
                    chunk.append(nxt)
                val = " ".join(chunk)
            if val:
                goods_full.append(val)

    tnved_values: List[str] = []
    for i, line in enumerate(lines):
        nl = norm_text(line)
        if "код тн вэд" in nl:
            same = re.findall(r"\b\d{4,10}\b", line)
            tnved_values.extend(same)
            for j in range(i + 1, min(len(lines), i + 3)):
                nxt = lines[j]
                if re.match(r"(?i)^товар\s+\d+", nxt.strip()):
                    break
                tnved_values.extend(re.findall(r"\b\d{4,10}\b", nxt))

    details = {
        "registry_doc_type": doc_type,
        "registry_doc_number": _label_value(lines, ["Регистрационный номер документа ЕАЭС", "Регистрационный номер документа", "Регистрационный номер"]),
        "registry_blank_number": _label_value(lines, ["Учетный номер бланка", "Учётный номер бланка"]),
        "registry_status": _label_value(lines, ["Признак действия", "Статус", "Статус действия"]),
        "registry_status_date": _label_value(lines, ["Дата действия"]),
        "registry_status_basis": _label_value(lines, ["Основание действия"]),
        "registry_date_start": _label_value(lines, ["Дата начала действия", "Дата регистрации"]),
        "registry_date_end": _label_value(lines, ["Дата окончания действия"]),
        "registry_applicant": _label_value(applicant_sec, ["Полное наименование", "Наименование"]),
        "registry_applicant_inn": _label_value(applicant_sec, ["ИНН", "БИН", "ОГРН"]),
        "registry_manufacturer": _label_value(manufacturer_sec, ["Полное наименование организации-изготовителя продукции", "Полное наименование", "Наименование"]),
        "registry_product_group": _label_value(product_sec or lines, ["Однородное наименование продукции", "Наименование продукции"]),
        "registry_product_full": _join_short(goods_full, max_len=2400),
        "registry_tnved": " | ".join(sorted(set(tnved_values))),
        "registry_scheme": _label_value(product_sec or lines, ["Схема сертификации", "Схема декларирования"]),
        "registry_technical_regulation": _label_value(product_sec or lines, ["Обозначение ТР (НД) с указанием разделов (пунктов, подпунктов), на соответствие требованиям которых проведена сертификация", "Обозначение ТР", "Технический регламент"]),
        "registry_evidence": _label_value(product_sec or lines, ["Обозначение (наименование) документов, на основании которых выдаётся сертификат соответствия", "Документы, на основании которых", "Основание выдачи"]),
        "registry_details_source": "swis.trade.kg_html_fast",
    }
    if url:
        details["registry_details_source"] += ":" + url
    return {k: re.sub(r"\s+", " ", str(v or "")).strip() for k, v in details.items()}



def _json_leaf_items(obj: Any, path: Tuple[str, ...] = ()) -> Iterable[Tuple[Tuple[str, ...], Any]]:
    """Yield scalar JSON leaves with their key path."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from _json_leaf_items(v, path + (str(k),))
    elif isinstance(obj, list):
        for i, v in enumerate(obj[:2000]):
            yield from _json_leaf_items(v, path + (str(i),))
    else:
        yield path, obj


def _norm_path(path: Tuple[str, ...]) -> str:
    return norm_key(" ".join(str(x) for x in path))


def _clean_json_value(v: Any, max_len: int = 2000) -> str:
    if v is None or isinstance(v, bool):
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    s = re.sub(r"\s+", " ", str(v)).strip(" ;,\t\r\n")
    if s.lower() in {"null", "none", "undefined"}:
        return ""
    return s[:max_len]


def _first_json_by_path(
    leaves: List[Tuple[Tuple[str, ...], Any]],
    include_any: Iterable[str],
    include_all: Iterable[str] = (),
    exclude_any: Iterable[str] = (),
    min_len: int = 1,
    max_len: int = 2000,
    value_re: Optional[str] = None,
) -> str:
    inc_any = [norm_key(x) for x in include_any if x]
    inc_all = [norm_key(x) for x in include_all if x]
    exc_any = [norm_key(x) for x in exclude_any if x]
    val_re = re.compile(value_re, re.I | re.U) if value_re else None
    for path, value in leaves:
        p = _norm_path(path)
        if inc_any and not any(x in p for x in inc_any):
            continue
        if inc_all and not all(x in p for x in inc_all):
            continue
        if exc_any and any(x in p for x in exc_any):
            continue
        v = _clean_json_value(value, max_len=max_len)
        if len(v) < min_len:
            continue
        if val_re and not val_re.search(v):
            continue
        return v
    return ""


def _collect_json_by_path(
    leaves: List[Tuple[Tuple[str, ...], Any]],
    include_any: Iterable[str],
    include_all: Iterable[str] = (),
    exclude_any: Iterable[str] = (),
    value_re: Optional[str] = None,
    max_items: int = 50,
    max_join_len: int = 2400,
) -> str:
    inc_any = [norm_key(x) for x in include_any if x]
    inc_all = [norm_key(x) for x in include_all if x]
    exc_any = [norm_key(x) for x in exclude_any if x]
    val_re = re.compile(value_re, re.I | re.U) if value_re else None
    out: List[str] = []
    seen: Set[str] = set()
    for path, value in leaves:
        p = _norm_path(path)
        if inc_any and not any(x in p for x in inc_any):
            continue
        if inc_all and not all(x in p for x in inc_all):
            continue
        if exc_any and any(x in p for x in exc_any):
            continue
        v = _clean_json_value(value, max_len=1200)
        if not v:
            continue
        if val_re:
            vals = val_re.findall(v)
            candidates = [x if isinstance(x, str) else "".join(x) for x in vals] or []
        else:
            candidates = [v]
        for cand in candidates:
            cand = re.sub(r"\s+", " ", str(cand)).strip(" ;,")
            if not cand:
                continue
            key = norm_key(cand)[:160]
            if key in seen:
                continue
            seen.add(key)
            out.append(cand)
            if len(out) >= max_items or len("; ".join(out)) >= max_join_len:
                return "; ".join(out)[:max_join_len]
    return "; ".join(out)[:max_join_len]


def _date_from_value(s: str) -> str:
    s = str(s or "")
    m = re.search(r"\b(\d{2}\.\d{2}\.\d{4})\b", s)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{4})-(\d{2})-(\d{2})(?:[T\s]\d{2}:\d{2}:\d{2}(?:\.\d+)?)?", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"
    return ""


def _first_date_by_path(leaves: List[Tuple[Tuple[str, ...], Any]], include_any: Iterable[str], include_all: Iterable[str] = (), exclude_any: Iterable[str] = ()) -> str:
    inc_any = [norm_key(x) for x in include_any if x]
    inc_all = [norm_key(x) for x in include_all if x]
    exc_any = [norm_key(x) for x in exclude_any if x]
    for path, value in leaves:
        p = _norm_path(path)
        if inc_any and not any(x in p for x in inc_any):
            continue
        if inc_all and not all(x in p for x in inc_all):
            continue
        if exc_any and any(x in p for x in exc_any):
            continue
        d = _date_from_value(_clean_json_value(value, max_len=200))
        if d:
            return d
    return ""


def _extract_fsa_kind_id(url: str) -> Tuple[str, str]:
    """Return ('rds_declaration'|'rss_certificate', id) for pub.fsa.gov.ru urls/API urls.

    Public FSA links use singular route parts:
      /rds/declaration/view/<id>/...
      /rss/certificate/view/<id>/...

    Internal API routes have appeared both as plural and singular in different
    snippets/versions, so the extractor accepts both to avoid losing a valid URL.
    """
    try:
        u = urlparse(str(url or ""))
        host = (u.netloc or "").lower().lstrip("www.")
        if host != "pub.fsa.gov.ru":
            return "", ""
        path = u.path or ""
        m = re.search(r"/rds/declaration/view/(\d+)", path, re.I)
        if m:
            return "rds_declaration", m.group(1)
        m = re.search(r"/api/v1/rds/common/declarations?/(\d+)", path, re.I)
        if m:
            return "rds_declaration", m.group(1)
        m = re.search(r"/rss/certificate/view/(\d+)", path, re.I)
        if m:
            return "rss_certificate", m.group(1)
        m = re.search(r"/api/v1/rss/common/certificates?/(\d+)", path, re.I)
        if m:
            return "rss_certificate", m.group(1)
        return "", ""
    except Exception:
        return "", ""


def _fsa_candidates(kind: str, doc_id: str, aggressive: bool = False) -> List[Tuple[str, str, Optional[Dict[str, Any]], str]]:
    """HTTP candidates: (method, url, json_body, label).

    HAR from the real browser confirmed the exact certificate endpoint:
      GET /api/v1/rss/common/certificates/<id>
    with plural ``certificates`` even though the public page URL is singular
    ``/rss/certificate/view/<id>/baseInfo``.

    For declarations the analogous FSA frontend endpoint is normally plural:
      GET /api/v1/rds/common/declarations/<id>

    Put the confirmed/analogous endpoints first and keep older variants only as
    fallback. This avoids spending time and possible 403s on wrong singular URLs.
    """
    doc_int = safe_int(doc_id, 0) or doc_id
    if kind == "rds_declaration":
        exact = [("GET", f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{doc_id}", None, "api_rds_common_declarations_id_har")]
        extra = [
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rds/common/declaration/{doc_id}", None, "api_rds_common_declaration_id"),
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{doc_id}/common", None, "api_rds_common_declarations_id_common"),
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rds/common/declaration/{doc_id}/common", None, "api_rds_common_declaration_id_common"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rds/common/declarations/get", {"id": doc_int}, "api_rds_common_declarations_get_id_int"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rds/common/declarations/get", {"id": str(doc_id)}, "api_rds_common_declarations_get_id_str"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rds/common/declarations/get", {"declarationId": doc_int}, "api_rds_common_declarations_get_declarationId"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rds/common/declaration/get", {"id": doc_int}, "api_rds_common_declaration_get_id_int"),
        ]
        return exact + extra if aggressive else exact
    if kind == "rss_certificate":
        exact = [("GET", f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{doc_id}", None, "api_rss_common_certificates_id_har")]
        extra = [
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rss/common/certificate/{doc_id}", None, "api_rss_common_certificate_id"),
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{doc_id}/baseInfo", None, "api_rss_common_certificates_id_baseInfo"),
            ("GET", f"https://pub.fsa.gov.ru/api/v1/rss/common/certificate/{doc_id}/baseInfo", None, "api_rss_common_certificate_id_baseInfo"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rss/common/certificates/get", {"id": doc_int}, "api_rss_common_certificates_get_id_int"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rss/common/certificates/get", {"id": str(doc_id)}, "api_rss_common_certificates_get_id_str"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rss/common/certificates/get", {"certificateId": doc_int}, "api_rss_common_certificates_get_certificateId"),
            ("POST", "https://pub.fsa.gov.ru/api/v1/rss/common/certificate/get", {"id": doc_int}, "api_rss_common_certificate_get_id_int"),
        ]
        return exact + extra if aggressive else exact
    return []

def _unwrap_json_payload(obj: Any) -> Any:
    """FSA sometimes wraps payload into data/item/result/content. Keep original if uncertain."""
    cur = obj
    for _ in range(4):
        if isinstance(cur, dict):
            for k in ("data", "item", "result", "content", "payload", "declaration", "certificate"):
                v = cur.get(k)
                if isinstance(v, (dict, list)):
                    cur = v
                    break
            else:
                return cur
        elif isinstance(cur, list) and len(cur) == 1 and isinstance(cur[0], dict):
            cur = cur[0]
        else:
            return cur
    return cur


def parse_fsa_json(obj: Any, url: str, kind: str, doc_id: str, label: str = "") -> Dict[str, str]:
    """Best-effort parser for public FSA JSON document details."""
    payload = _unwrap_json_payload(obj)
    leaves = list(_json_leaf_items(payload))
    raw_text = json.dumps(payload, ensure_ascii=False, default=str)[:250000]

    is_decl = kind == "rds_declaration"
    details: Dict[str, str] = {
        "registry_doc_type": "декларация о соответствии" if is_decl else "сертификат соответствия",
        "registry_details_source": f"pub.fsa.gov.ru_api_fast:{label or kind}:{url}",
    }

    # Номер документа: сначала поля с явным register/number, затем поиск по тексту.
    number = _first_json_by_path(
        leaves,
        include_any=["declNumber", "certNumber", "certificateNumber", "declarationNumber", "number", "regNumber", "registrationNumber", "reestrNumber", "docNumber", "certificat", "declaration"],
        exclude_any=["id", "phone", "fax", "house", "building", "blank", "form", "protocol", "tnved", "code", "status", "date"],
        min_len=4,
        max_len=300,
    )
    if not number:
        m = re.search(r"(?:ЕАЭС|ТС|РОСС|RU|KZ|BY|KG|AM|EAЭС)\s*[№N]?\s*[A-ZА-Я0-9][A-ZА-Я0-9 ._/-]{8,120}", raw_text, re.I)
        if m:
            number = re.sub(r"\s+", " ", m.group(0)).strip(' ,;"')
    details["registry_doc_number"] = number

    details["registry_blank_number"] = _first_json_by_path(
        leaves,
        include_any=["blank", "formNumber", "form", "бланк"],
        exclude_any=["date", "id"],
        min_len=3,
        max_len=300,
    )
    details["registry_status"] = _first_json_by_path(
        leaves,
        include_any=["status", "state", "actual", "статус", "состояние"],
        exclude_any=["date", "id", "code"],
        min_len=2,
        max_len=300,
    )
    details["registry_status_date"] = _first_date_by_path(
        leaves,
        include_any=["statusDate", "stateDate", "suspendDate", "stopDate", "annulDate", "terminationDate", "датаСтатус", "status"],
    )
    details["registry_status_basis"] = _first_json_by_path(
        leaves,
        include_any=["statusBasis", "suspend", "termination", "annul", "reason", "basis", "основан", "причин"],
        exclude_any=["date", "id", "code"],
        min_len=6,
        max_len=900,
    )

    # Даты: названия полей в ФСА менялись, поэтому несколько групп.
    details["registry_date_start"] = (
        _first_date_by_path(leaves, include_any=["startDate", "dateStart", "validFrom", "beginDate", "regDate", "registrationDate", "declDate", "certDate", "датарег", "датанач"])
        or _first_date_by_path(leaves, include_any=["date"], include_all=["registration"])
        or _first_date_by_path(leaves, include_any=["date"], include_all=["begin"])
    )
    details["registry_date_end"] = (
        _first_date_by_path(leaves, include_any=["endDate", "dateEnd", "validTo", "validUntil", "expire", "expiration", "validity", "датаоконч", "датадо"])
        or _first_date_by_path(leaves, include_any=["date"], include_all=["end"])
    )

    # Юрлица. Берём именно имена из контекстов applicant/declarant/manufacturer.
    details["registry_applicant"] = _first_json_by_path(
        leaves,
        include_any=["applicant", "declarant", "заявител", "декларант"],
        include_all=["name"],
        exclude_any=["address", "phone", "email", "inn", "ogrn", "kpp", "id", "type"],
        min_len=3,
        max_len=1000,
    ) or _first_json_by_path(
        leaves,
        include_any=["applicant", "declarant", "заявител", "декларант"],
        exclude_any=["address", "phone", "email", "inn", "ogrn", "kpp", "id", "type", "date"],
        min_len=8,
        max_len=1000,
    )
    details["registry_applicant_inn"] = _first_json_by_path(
        leaves,
        include_any=["inn", "ИНН", "ogrn", "ОГРН"],
        include_all=["applicant"] if any("applicant" in _norm_path(p) for p, _ in leaves) else [],
        value_re=r"\b\d{9,13}\b",
        min_len=9,
        max_len=50,
    ) or _first_json_by_path(
        leaves,
        include_any=["declarant", "заявител", "декларант"],
        include_all=["inn"],
        value_re=r"\b\d{9,13}\b",
        min_len=9,
        max_len=50,
    )
    details["registry_manufacturer"] = _first_json_by_path(
        leaves,
        include_any=["manufacturer", "producer", "maker", "изготов", "производ"],
        include_all=["name"],
        exclude_any=["address", "phone", "email", "inn", "id", "type"],
        min_len=3,
        max_len=1200,
    ) or _first_json_by_path(
        leaves,
        include_any=["manufacturer", "producer", "maker", "изготов", "производ"],
        exclude_any=["address", "phone", "email", "inn", "id", "type", "date"],
        min_len=8,
        max_len=1200,
    )

    details["registry_product_group"] = _first_json_by_path(
        leaves,
        include_any=["productGroup", "productType", "productName", "homogeneous", "commonName", "groupName", "наименованиепродук", "групп"],
        exclude_any=["id", "code", "tnved", "address", "date"],
        min_len=3,
        max_len=1200,
    )
    details["registry_product_full"] = _collect_json_by_path(
        leaves,
        include_any=["product", "goods", "продукц", "объект"],
        include_all=[],
        exclude_any=["id", "uuid", "code", "tnved", "date", "status"],
        max_items=20,
        max_join_len=3000,
    )
    # Если собрали мусор из коротких значений, подстраховываемся длинными строками из сырого JSON.
    if len(details.get("registry_product_full", "")) < 20:
        long_vals: List[str] = []
        for _p, v in leaves:
            s = _clean_json_value(v, max_len=2000)
            if len(s) >= 35 and re.search(r"продукц|издел|товар|обув|одежд|сертифик|деклара", s, re.I):
                long_vals.append(s)
        details["registry_product_full"] = _join_short(long_vals, max_len=3000)

    tnved = _collect_json_by_path(
        leaves,
        include_any=["tnved", "tnVed", "tnvedCode", "tnVedCode", "кодтнвэд", "tNVED"],
        value_re=r"\b\d{4,10}\b",
        max_items=80,
        max_join_len=1200,
    )
    if not tnved:
        vals = sorted(set(re.findall(r"\b\d{10}\b", raw_text)))
        tnved = "; ".join(vals[:80])
    details["registry_tnved"] = tnved

    details["registry_scheme"] = _first_json_by_path(
        leaves,
        include_any=["scheme", "schema", "схем"],
        exclude_any=["id", "date"],
        min_len=1,
        max_len=300,
    )
    details["registry_technical_regulation"] = _collect_json_by_path(
        leaves,
        include_any=["technicalRegulation", "techReg", "trTs", "ТРТС", "регламент", "standard"],
        exclude_any=["id", "date", "code"],
        max_items=20,
        max_join_len=1800,
    )
    details["registry_evidence"] = _collect_json_by_path(
        leaves,
        include_any=["evidence", "protocol", "laboratory", "test", "basis", "основан", "протокол", "испыт", "лаборатор"],
        exclude_any=["id"],
        max_items=30,
        max_join_len=2400,
    )

    # Чистка слишком общих/технических значений.
    for k, v in list(details.items()):
        details[k] = re.sub(r"\s+", " ", str(v or "")).strip()
    return details


def parse_fsa_html(raw_html: str, url: str, kind: str, doc_id: str) -> Dict[str, str]:
    """Fallback parser for FSA page text when JSON API is not available."""
    lines = _strip_html_to_lines(raw_html)
    all_text = "\n".join(lines)
    details = {
        "registry_doc_type": "декларация о соответствии" if kind == "rds_declaration" else "сертификат соответствия",
        "registry_doc_number": _label_value(lines, ["Регистрационный номер", "Номер сертификата", "Номер декларации", "Номер документа"]),
        "registry_status": _label_value(lines, ["Статус", "Статус действия", "Состояние"]),
        "registry_date_start": _label_value(lines, ["Дата регистрации", "Дата начала действия", "Действует с"]),
        "registry_date_end": _label_value(lines, ["Дата окончания действия", "Действует до"]),
        "registry_applicant": _label_value(lines, ["Заявитель", "Полное наименование заявителя"]),
        "registry_manufacturer": _label_value(lines, ["Изготовитель", "Полное наименование изготовителя"]),
        "registry_product_full": _label_value(lines, ["Продукция", "Общее наименование продукции", "Полное наименование продукции"]),
        "registry_tnved": "; ".join(sorted(set(re.findall(r"\b\d{10}\b", all_text)))[:80]),
        "registry_details_source": f"pub.fsa.gov.ru_html_fast:{url}",
    }
    return {k: re.sub(r"\s+", " ", str(v or "")).strip() for k, v in details.items()}



def _fsa_browser_like_headers(referer: str, accept: str = "application/json, text/plain, */*", *, origin: bool = False, xhr: bool = False) -> Dict[str, str]:
    """Headers copied from the real FSA HAR as closely as possible.

    Important finding from the HAR: normal GET XHR requests to FSA API do NOT
    send ``Origin`` and do NOT send ``X-Requested-With``. They do send blank
    ``lkId``/``orgId`` headers. Earlier versions sent Origin/X-Requested-With
    for every API request; that can change the request profile and led to 403 or
    empty SPA HTML instead of document JSON.
    """
    h = {
        "Accept": accept,
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": referer,
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": DEFAULT_UA,
        "lkId": "",
        "orgId": "",
        "sec-ch-ua": '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }
    # Let requests/aiohttp negotiate supported compression. Do not advertise zstd
    # because some Python clients cannot decode it.
    if origin:
        h["Origin"] = "https://pub.fsa.gov.ru"
    if xhr:
        h["X-Requested-With"] = "XMLHttpRequest"
    return h


def _fsa_html_headers(referer: str) -> Dict[str, str]:
    h = _fsa_browser_like_headers(referer, "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8")
    h["Sec-Fetch-Dest"] = "document"
    h["Sec-Fetch-Mode"] = "navigate"
    h["Sec-Fetch-Site"] = "none"
    return h


def _fsa_warmup_urls(kind: str, doc_id: str, referer: str) -> List[Tuple[str, str, str]]:
    """Fast warm-up sequence seen in HAR, still without browser/rendering.
    v24.1: добавлен заход на главную страницу FSA первым — это устанавливает
    базовые cookies сессии, которые антибот FSA проверяет на API-запросах."""
    base = "rss" if kind == "rss_certificate" else "rds"
    return [
        ("home", "https://pub.fsa.gov.ru/", "html"),
        ("doc_page", referer, "html"),
        ("i18n_ru", "https://pub.fsa.gov.ru/assets/i18n/ru.json", "json"),
        ("lk_account", "https://pub.fsa.gov.ru/lk/api/account", "json"),
        ("common_account", f"https://pub.fsa.gov.ru/api/v1/{base}/common/account", "json"),
        ("identifiers", f"https://pub.fsa.gov.ru/api/v1/{base}/common/identifiers", "json"),
    ]


async def _fsa_polite_delay(args) -> None:
    """Serialize FSA requests and add a small delay: FSA easily returns 403/429 on bursts."""
    global FSA_LAST_REQUEST_TS
    min_delay = max(0.0, float(getattr(args, "fsa_min_delay_sec", 0.8) or 0.0))
    if min_delay <= 0:
        FSA_LAST_REQUEST_TS = time.monotonic()
        return
    now = time.monotonic()
    wait = FSA_LAST_REQUEST_TS + min_delay - now
    if wait > 0:
        await asyncio.sleep(wait + random.uniform(0.0, min(0.25, min_delay * 0.25)))
    FSA_LAST_REQUEST_TS = time.monotonic()


def _json_from_response_text(txt: str) -> Any:
    txt = (txt or "").strip()
    if not txt:
        raise ValueError("empty response")
    return json.loads(txt)


def _curl_cli_run(url: str, headers: Dict[str, str], timeout: float, cookie_file: str = "") -> Tuple[int, str, str]:
    """Run system curl.exe/curl as a fast no-browser fallback. Returns (status, body, error)."""
    curl_bin = shutil.which("curl") or shutil.which("curl.exe")
    if not curl_bin:
        return 0, "", "curl_cli:not_found"
    cmd = [curl_bin, "-sS", "-L", "--compressed", "--http1.1", "--max-time", str(int(max(3, timeout))), "-w", "\n__HTTP_CODE__:%{http_code}"]
    if cookie_file:
        cmd += ["-b", cookie_file, "-c", cookie_file]
    for k, v in headers.items():
        # libcurl treats "Header:" as an empty header; that is what FSA HAR has for lkId/orgId.
        cmd += ["-H", f"{k}: {v}" if v else f"{k}:"]
    cmd.append(url)
    try:
        cp = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=max(5, timeout + 4))
        out = cp.stdout or ""
        err = (cp.stderr or "").strip()
        m = re.search(r"\n__HTTP_CODE__:(\d{3})\s*$", out)
        if not m:
            return 0, out, f"curl_cli:no_status:{err[:120]}"
        status = int(m.group(1))
        body = out[:m.start()]
        return status, body, err[:300]
    except Exception as e:
        return 0, "", f"curl_cli:{type(e).__name__}:{str(e)[:160]}"


def _fetch_fsa_details_curl_cffi_sync(url: str, kind: str, doc_id: str, args, prior_errors: Optional[List[str]] = None) -> Dict[str, str]:
    """Fast FSA fetch with Chrome-like TLS fingerprint via curl_cffi, if installed.

    This is still not a browser/Playwright: it is a direct HTTP client, but it can
    impersonate Chrome at TLS/JA3 level. FSA may return 403 to requests/aiohttp
    while the same endpoint returns 200 in Chrome.
    """
    if _curl_requests is None:
        return {"registry_details_source": "pub.fsa.gov.ru_curl_cffi_missing:install:pip install curl_cffi"}
    referer = f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common" if kind == "rds_declaration" else f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"
    timeout = max(4.0, float(getattr(args, "registry_details_timeout_sec", 8.0)))
    ua = str(getattr(args, "user_agent", None) or DEFAULT_UA)
    errors: List[str] = list(prior_errors or [])[-4:]
    impersonates = [str(getattr(args, "fsa_curl_cffi_impersonate", "chrome") or "chrome"), "chrome120", "chrome110"]
    # Preserve order, remove duplicates.
    seen_imp = set()
    impersonates = [x for x in impersonates if not (x in seen_imp or seen_imp.add(x))]
    for imp in impersonates:
        try:
            sess = _curl_requests.Session()
            # Warm-up is cheap; use the same Chrome impersonation and headers.
            for label, warm_url, typ in _fsa_warmup_urls(kind, doc_id, referer):
                try:
                    wh = _fsa_html_headers(referer) if typ == "html" else _fsa_browser_like_headers(referer)
                    wh["User-Agent"] = ua
                    sess.get(warm_url, headers=wh, timeout=timeout, impersonate=imp)
                except TypeError:
                    sess.get(warm_url, headers=wh, timeout=timeout)
                except Exception as e:
                    errors.append(f"curl_cffi_warm_{label}:{type(e).__name__}:{str(e)[:60]}")
            get_headers = _fsa_browser_like_headers(referer)
            get_headers["User-Agent"] = ua
            post_headers = _fsa_browser_like_headers(referer, origin=True)
            post_headers["User-Agent"] = ua
            for method, api_url, body, label in _fsa_candidates(kind, doc_id, bool(getattr(args, "fsa_aggressive_candidates", False))):
                try:
                    if method == "POST":
                        try:
                            r = sess.post(api_url, json=body, headers=post_headers, timeout=timeout, impersonate=imp)
                        except TypeError:
                            r = sess.post(api_url, json=body, headers=post_headers, timeout=timeout)
                    else:
                        try:
                            r = sess.get(api_url, headers=get_headers, timeout=timeout, impersonate=imp)
                        except TypeError:
                            r = sess.get(api_url, headers=get_headers, timeout=timeout)
                    txt = getattr(r, "text", "") or ""
                    status = int(getattr(r, "status_code", 0) or 0)
                    if status != 200 or not txt:
                        errors.append(f"curl_cffi_{imp}_{label}:http_{status}")
                        continue
                    try:
                        obj = r.json()
                    except Exception:
                        try:
                            obj = json.loads(txt)
                        except Exception:
                            errors.append(f"curl_cffi_{imp}_{label}:not_json:{txt[:80]}")
                            continue
                    d = parse_fsa_json(obj, api_url, kind, doc_id, f"curl_cffi_{imp}_{label}")
                    if _details_meaningful(d):
                        return d
                    errors.append(f"curl_cffi_{imp}_{label}:parsed_empty")
                except Exception as e:
                    errors.append(f"curl_cffi_{imp}_{label}:{type(e).__name__}:{str(e)[:80]}")
        except Exception as e:
            errors.append(f"curl_cffi_session_{imp}:{type(e).__name__}:{str(e)[:100]}")
    return {"registry_details_source": f"pub.fsa.gov.ru_curl_cffi_error:fsa:{kind}:{doc_id}:" + " | ".join(errors[-8:])}


def _fetch_fsa_details_curl_cli_sync(url: str, kind: str, doc_id: str, args, prior_errors: Optional[List[str]] = None) -> Dict[str, str]:
    """Fast FSA fetch through system curl/curl.exe. No browser, but a different HTTP/TLS stack than requests."""
    referer = f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common" if kind == "rds_declaration" else f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"
    timeout = max(4.0, float(getattr(args, "registry_details_timeout_sec", 8.0)))
    ua = str(getattr(args, "user_agent", None) or DEFAULT_UA)
    errors: List[str] = list(prior_errors or [])[-4:]
    cookie_file = ""
    try:
        with tempfile.NamedTemporaryFile(prefix="fsa_cookies_", suffix=".txt", delete=False) as tf:
            cookie_file = tf.name
        for label, warm_url, typ in _fsa_warmup_urls(kind, doc_id, referer):
            wh = _fsa_html_headers(referer) if typ == "html" else _fsa_browser_like_headers(referer)
            wh["User-Agent"] = ua
            st, body, err = _curl_cli_run(warm_url, wh, timeout, cookie_file)
            if st and st not in (200, 204, 301, 302, 304, 403):
                errors.append(f"curl_cli_warm_{label}:http_{st}")
            elif err and "not_found" in err:
                return {"registry_details_source": "pub.fsa.gov.ru_curl_cli_missing:curl_not_found"}
        get_headers = _fsa_browser_like_headers(referer)
        get_headers["User-Agent"] = ua
        post_headers = _fsa_browser_like_headers(referer, origin=True)
        post_headers["User-Agent"] = ua
        for method, api_url, body, label in _fsa_candidates(kind, doc_id, bool(getattr(args, "fsa_aggressive_candidates", False))):
            if method != "GET":
                continue  # system curl fallback intentionally stays on confirmed GET endpoints
            st, txt, err = _curl_cli_run(api_url, get_headers, timeout, cookie_file)
            if err and "not_found" in err:
                return {"registry_details_source": "pub.fsa.gov.ru_curl_cli_missing:curl_not_found"}
            if st != 200 or not txt:
                errors.append(f"curl_cli_{label}:http_{st}:{err[:80]}")
                continue
            try:
                obj = json.loads(txt)
            except Exception:
                errors.append(f"curl_cli_{label}:not_json:{txt[:80]}")
                continue
            d = parse_fsa_json(obj, api_url, kind, doc_id, "curl_cli_" + label)
            if _details_meaningful(d):
                return d
            errors.append(f"curl_cli_{label}:parsed_empty")
    finally:
        if cookie_file:
            try:
                os.unlink(cookie_file)
            except Exception:
                pass
    return {"registry_details_source": f"pub.fsa.gov.ru_curl_cli_error:fsa:{kind}:{doc_id}:" + " | ".join(errors[-8:])}


def _fetch_fsa_details_requests_sync(url: str, kind: str, doc_id: str, args, prior_errors: Optional[List[str]] = None) -> Dict[str, str]:
    """Fast FSA fetch through requests.Session using the exact HAR request profile."""
    if _requests is None:
        return {"registry_details_source": "pub.fsa.gov.ru_requests_error:requests_not_available"}
    referer = f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common" if kind == "rds_declaration" else f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"
    timeout = max(4.0, float(getattr(args, "registry_details_timeout_sec", 8.0)))
    errors: List[str] = list(prior_errors or [])[-4:]
    try:
        sess = _requests.Session()
        if getattr(args, "user_agent", None):
            ua = str(getattr(args, "user_agent"))
        else:
            ua = DEFAULT_UA
        # Warm-up sequence observed in HAR. It is cheap and still not a browser.
        for label, warm_url, typ in _fsa_warmup_urls(kind, doc_id, referer):
            try:
                wh = _fsa_html_headers(referer) if typ == "html" else _fsa_browser_like_headers(referer)
                wh["User-Agent"] = ua
                r = sess.get(warm_url, headers=wh, timeout=timeout, allow_redirects=True)
                # 403 on warmup should not stop the real endpoint attempts.
                if r.status_code not in (200, 204, 301, 302, 304, 403):
                    errors.append(f"requests_warm_{label}:http_{r.status_code}")
            except Exception as e:
                errors.append(f"requests_warm_{label}:{type(e).__name__}:{str(e)[:60]}")

        get_headers = _fsa_browser_like_headers(referer)
        get_headers["User-Agent"] = ua
        post_headers = _fsa_browser_like_headers(referer, origin=True)
        post_headers["User-Agent"] = ua
        for method, api_url, body, label in _fsa_candidates(kind, doc_id, bool(getattr(args, "fsa_aggressive_candidates", False))):
            try:
                if method == "POST":
                    r = sess.post(api_url, json=body, headers=post_headers, timeout=timeout, allow_redirects=True)
                else:
                    r = sess.get(api_url, headers=get_headers, timeout=timeout, allow_redirects=True)
                txt = r.text or ""
                if r.status_code != 200 or not txt:
                    errors.append(f"requests_{label}:http_{r.status_code}")
                    continue
                try:
                    obj = r.json()
                except Exception:
                    errors.append(f"requests_{label}:not_json:{txt[:80]}")
                    continue
                d = parse_fsa_json(obj, api_url, kind, doc_id, "requests_" + label)
                if _details_meaningful(d):
                    return d
                errors.append(f"requests_{label}:parsed_empty")
            except Exception as e:
                errors.append(f"requests_{label}:{type(e).__name__}:{str(e)[:80]}")

        # Last fast fallback: parse public HTML if it is directly returned.
        try:
            hh = _fsa_html_headers(referer)
            hh["User-Agent"] = ua
            r = sess.get(url, headers=hh, timeout=timeout, allow_redirects=True)
            txt = r.text or ""
            if r.status_code == 200 and txt:
                d = parse_fsa_html(txt, url, kind, doc_id)
                if _details_meaningful(d):
                    d["registry_details_source"] = "pub.fsa.gov.ru_requests_html_fast:" + url
                    return d
                errors.append("requests_html:parsed_empty")
            else:
                errors.append(f"requests_html:http_{r.status_code}")
        except Exception as e:
            errors.append(f"requests_html:{type(e).__name__}:{str(e)[:80]}")
    except Exception as e:
        errors.append(f"requests_session:{type(e).__name__}:{str(e)[:120]}")
    return {"registry_details_source": f"pub.fsa.gov.ru_fast_error:fsa:{kind}:{doc_id}:" + " | ".join(errors[-10:])}

async def _fetch_fsa_details(url: str, args) -> Dict[str, str]:
    kind, doc_id = _extract_fsa_kind_id(url)
    if not kind or not doc_id:
        return {}
    cache_key = f"fsa:{kind}:{doc_id}"
    if cache_key in REGISTRY_DETAILS_CACHE:
        return dict(REGISTRY_DETAILS_CACHE[cache_key])

    referer = f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common" if kind == "rds_declaration" else f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"
    ua = str(getattr(args, "user_agent", None) or DEFAULT_UA)
    timeout = aiohttp.ClientTimeout(total=max(4.0, float(getattr(args, "registry_details_timeout_sec", 8.0))))
    errors: List[str] = []
    best: Dict[str, str] = {}

    # First try a Chrome-impersonating direct HTTP stack. This is still no-browser,
    # but it solves the common case where FSA blocks requests/aiohttp by TLS fingerprint.
    try:
        cd = await asyncio.to_thread(_fetch_fsa_details_curl_cffi_sync, url, kind, doc_id, args, errors)
        csource = str(cd.get("registry_details_source", "")) if cd else ""
        if cd and _details_meaningful(cd) and not csource.startswith("pub.fsa.gov.ru_curl_cffi_error"):
            REGISTRY_DETAILS_CACHE[cache_key] = dict(cd)
            return cd
        if csource:
            errors.append(csource[:300])
    except Exception as e:
        errors.append(f"curl_cffi_thread:{type(e).__name__}:{str(e)[:120]}")

    # Second no-browser fallback: system curl/curl.exe. Often available on Windows 10/11.
    try:
        clid = await asyncio.to_thread(_fetch_fsa_details_curl_cli_sync, url, kind, doc_id, args, errors)
        clisource = str(clid.get("registry_details_source", "")) if clid else ""
        if clid and _details_meaningful(clid) and not clisource.startswith("pub.fsa.gov.ru_curl_cli_error"):
            REGISTRY_DETAILS_CACHE[cache_key] = dict(clid)
            return clid
        if clisource:
            errors.append(clisource[:300])
    except Exception as e:
        errors.append(f"curl_cli_thread:{type(e).__name__}:{str(e)[:120]}")

    try:
        # Do not set session-wide headers here: GET and POST have different HAR profiles.
        async with aiohttp.ClientSession(timeout=timeout) as session:
            # Warm-up sequence copied from HAR. No Playwright/browser, just HTTP.
            for warm_label, warm_url, typ in _fsa_warmup_urls(kind, doc_id, referer):
                try:
                    wh = _fsa_html_headers(referer) if typ == "html" else _fsa_browser_like_headers(referer)
                    wh["User-Agent"] = ua
                    async with session.get(warm_url, allow_redirects=True, headers=wh) as wr:
                        await wr.text(errors="ignore")
                        if wr.status not in (200, 204, 301, 302, 304, 403):
                            errors.append(f"warm_{warm_label}:http_{wr.status}")
                except Exception as e:
                    errors.append(f"warm_{warm_label}:{type(e).__name__}:{str(e)[:60]}")

            get_headers = _fsa_browser_like_headers(referer)
            get_headers["User-Agent"] = ua
            post_headers = _fsa_browser_like_headers(referer, origin=True)
            post_headers["User-Agent"] = ua
            for method, api_url, body, label in _fsa_candidates(kind, doc_id, bool(getattr(args, "fsa_aggressive_candidates", False))):
                try:
                    if method == "POST":
                        async with session.post(api_url, json=body, allow_redirects=True, headers=post_headers) as resp:
                            txt = await resp.text(errors="ignore")
                            status = resp.status
                    else:
                        async with session.get(api_url, allow_redirects=True, headers=get_headers) as resp:
                            txt = await resp.text(errors="ignore")
                            status = resp.status
                    if status != 200 or not txt:
                        errors.append(f"{label}:http_{status}")
                        continue
                    try:
                        obj = json.loads(txt)
                    except Exception:
                        errors.append(f"{label}:not_json:{txt[:80]}")
                        continue
                    d = parse_fsa_json(obj, api_url, kind, doc_id, label)
                    if _details_meaningful(d):
                        REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
                        return d
                    best = d
                    errors.append(f"{label}:parsed_empty")
                except Exception as e:
                    errors.append(f"{label}:{type(e).__name__}:{str(e)[:80]}")

            # HTML fallback by public URL: works only if FSA embeds text in HTML.
            try:
                hh = _fsa_html_headers(referer)
                hh["User-Agent"] = ua
                async with session.get(url, allow_redirects=True, headers=hh) as resp:
                    html_body = await resp.text(errors="ignore")
                    if resp.status == 200 and html_body:
                        d = parse_fsa_html(html_body, url, kind, doc_id)
                        if _details_meaningful(d):
                            REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
                            return d
                        best = d or best
                    else:
                        errors.append(f"html:http_{resp.status}")
            except Exception as e:
                errors.append(f"html:{type(e).__name__}:{str(e)[:80]}")
    except Exception as e:
        errors.append(f"session:{type(e).__name__}:{str(e)[:120]}")

    # Second fast engine: requests.Session with the same exact HAR headers and warm-up.
    try:
        rd = await asyncio.to_thread(_fetch_fsa_details_requests_sync, url, kind, doc_id, args, errors)
        if rd and _details_meaningful(rd) and not str(rd.get("registry_details_source", "")).startswith("pub.fsa.gov.ru_fast_error"):
            REGISTRY_DETAILS_CACHE[cache_key] = dict(rd)
            return rd
        if rd and rd.get("registry_details_source"):
            errors.append(str(rd.get("registry_details_source"))[:300])
    except Exception as e:
        errors.append(f"requests_thread:{type(e).__name__}:{str(e)[:120]}")

    if best and _details_meaningful(best):
        best["registry_details_source"] = (best.get("registry_details_source") or f"pub.fsa.gov.ru_error:{cache_key}") + "; errors=" + " | ".join(errors[:8])
        REGISTRY_DETAILS_CACHE[cache_key] = dict(best)
        return best
    d = {"registry_details_source": f"pub.fsa.gov.ru_fast_error:{cache_key}:{' | '.join(errors[:10])}"}
    REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
    return d


def _details_meaningful(d: Dict[str, str]) -> bool:
    if not d:
        return False
    return any(str(d.get(k) or "").strip() for k in (
        "registry_doc_number", "registry_status", "registry_applicant", "registry_manufacturer",
        "registry_product_full", "registry_tnved", "registry_date_start", "registry_date_end"
    ))


async def _fetch_fsa_details_browser(page, url: str, args) -> Dict[str, str]:
    """Fallback for FSA details through a real Playwright browser page.

    Direct aiohttp calls to pub.fsa.gov.ru API may return 403 even when the public
    page opens in a normal browser. This fallback opens the public registry URL in
    the already running browser, captures JSON API responses if the SPA loads them,
    and otherwise parses visible page text. It is cached by unique FSA document id.
    """
    kind, doc_id = _extract_fsa_kind_id(url)
    if not kind or not doc_id or page is None:
        return {}
    cache_key = f"fsa:{kind}:{doc_id}"
    cached = REGISTRY_DETAILS_CACHE.get(cache_key)
    if cached and _details_meaningful(cached):
        return dict(cached)

    timeout_ms = int(getattr(args, "fsa_browser_timeout_ms", 25000) or 25000)
    wait_ms = int(getattr(args, "fsa_browser_wait_ms", 6500) or 6500)
    captured: List[Tuple[str, Any]] = []
    errors: List[str] = []
    tasks: List[asyncio.Task] = []

    async def _handle_response(resp):
        try:
            ru = str(resp.url or "")
            if "pub.fsa.gov.ru" not in ru or "/api/v1/" not in ru:
                return
            if doc_id not in ru and not any(x in ru for x in ("/declarations/get", "/certificates/get")):
                return
            status = int(getattr(resp, "status", 0) or 0)
            if status != 200:
                errors.append(f"browser_api:{status}:{ru[:120]}")
                return
            txt = await resp.text()
            if not txt or not re.search(r"[\[{]", txt[:50]):
                return
            try:
                obj = json.loads(txt)
            except Exception:
                errors.append(f"browser_api:not_json:{ru[:120]}")
                return
            captured.append((ru, obj))
        except Exception as e:
            errors.append(f"browser_response:{type(e).__name__}:{str(e)[:100]}")

    def _on_response(resp):
        try:
            tasks.append(asyncio.create_task(_handle_response(resp)))
        except Exception:
            pass

    try:
        page.on("response", _on_response)
    except Exception:
        pass

    try:
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        except Exception as e:
            errors.append(f"browser_goto:{type(e).__name__}:{str(e)[:120]}")
        try:
            await page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 10000))
        except Exception:
            pass
        await asyncio.sleep(max(0.3, wait_ms / 1000.0))
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

        for api_url, obj in captured:
            d = parse_fsa_json(obj, api_url, kind, doc_id, "browser_api")
            if _details_meaningful(d):
                d["registry_details_source"] = f"pub.fsa.gov.ru_browser_api:{api_url}"
                REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
                return d

        body_text = ""
        try:
            body_text = await page.locator("body").inner_text(timeout=5000)
        except Exception as e:
            errors.append(f"browser_text:{type(e).__name__}:{str(e)[:100]}")
        if body_text:
            d = parse_fsa_html(body_text, url, kind, doc_id)
            if _details_meaningful(d):
                d["registry_details_source"] = f"pub.fsa.gov.ru_browser_text:{url}"
                REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
                return d
            sample = re.sub(r"\s+", " ", body_text).strip()[:300]
            errors.append("browser_text_empty:" + sample)
    finally:
        try:
            page.remove_listener("response", _on_response)
        except Exception:
            pass

    d = {"registry_details_source": f"pub.fsa.gov.ru_browser_error:{cache_key}:" + " | ".join(errors[:8])}
    REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
    return d


async def _fetch_swis_details(url: str, args) -> Dict[str, str]:
    cache_key = url.split("#", 1)[0]
    if cache_key in REGISTRY_DETAILS_CACHE:
        return dict(REGISTRY_DETAILS_CACHE[cache_key])
    headers = {
        "User-Agent": (getattr(args, "user_agent", None) or DEFAULT_UA),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://www.wildberries.ru/",
    }
    timeout = aiohttp.ClientTimeout(total=max(3.0, float(getattr(args, "registry_details_timeout_sec", 8.0))))
    try:
        async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
            async with session.get(cache_key, allow_redirects=True) as resp:
                body = await resp.text(errors="ignore")
                if resp.status != 200 or not body:
                    d = {"registry_details_source": f"swis.trade.kg_error:http_{resp.status}:{cache_key}"}
                else:
                    d = parse_swis_trade_kg_html(body, cache_key)
    except Exception as e:
        d = {"registry_details_source": f"swis.trade.kg_error:{type(e).__name__}:{str(e)[:160]}:{cache_key}"}
    REGISTRY_DETAILS_CACHE[cache_key] = dict(d)
    return d


async def fetch_registry_details_fast(urls: List[str], args, page=None) -> Dict[str, str]:
    """Fetch details from supported public registries without browser: swis.trade.kg + FSA pub.fsa.gov.ru."""
    if not urls or not bool(getattr(args, "registry_details", True)):
        return {}
    last_error: Dict[str, str] = {}
    for raw_url in urls:
        url = str(raw_url or "").strip()
        if not url:
            continue
        parsed = urlparse(url)
        host = (parsed.netloc or "").lower().lstrip("www.")
        path = parsed.path or ""
        if host in {"swis.trade.kg", "www.swis.trade.kg"} and re.search(r"/doc/[0-9a-f-]{36}", path, re.I):
            d = await _fetch_swis_details(url, args)
            if d and not str(d.get("registry_details_source", "")).startswith("swis.trade.kg_error"):
                return d
            last_error = d or last_error
            continue
        kind, doc_id = _extract_fsa_kind_id(url)
        if host == "pub.fsa.gov.ru" and kind and doc_id:
            # FSA is sensitive to bursts: serialize unique document requests and add a small delay.
            async with FSA_DETAILS_LOCK:
                await _fsa_polite_delay(args)
                d = await _fetch_fsa_details(url, args)
            source = str(d.get("registry_details_source", "")) if d else ""
            if d and _details_meaningful(d) and not source.startswith("pub.fsa.gov.ru_error"):
                return d
            last_error = d or last_error
            # Direct API often returns 403. If a browser page is available, use it as fallback
            # for this unique FSA document only. Cache prevents repeat loads for duplicates.
            if page is not None and bool(getattr(args, "fsa_browser_fallback", True)):
                bd = await _fetch_fsa_details_browser(page, url, args)
                bsource = str(bd.get("registry_details_source", "")) if bd else ""
                if bd and _details_meaningful(bd) and not bsource.startswith("pub.fsa.gov.ru_browser_error"):
                    return bd
                if bd:
                    last_error = bd
            continue
    return last_error

def certificate_json_urls(nm_id: int) -> List[str]:
    nm = int(nm_id)
    vol, part = wb_volume_part(nm)
    primary = wb_basket_by_volume(vol)
    order: List[int] = []
    for b in [primary, primary - 1, primary + 1, 13, 12, 14, 11, 15]:
        if 1 <= b <= 30 and b not in order:
            order.append(b)
    for b in range(1, 31):
        if b not in order:
            order.append(b)
    return [f"https://basket-{b:02d}.wbbasket.ru/vol{vol}/part{part}/{nm}/info/certificate.json" for b in order]


async def fetch_registry_urls_from_certificate_json(card: BrandCard, args) -> Tuple[List[str], str]:
    """Быстрый прямой сбор ссылки на реестр из WB certificate.json без кликов.

    Возвращает (urls, detail). detail начинается с:
    - certificate_json_ok — ссылка найдена;
    - certificate_json_no_docs — certificate.json отсутствует на всех basket-хостах;
    - certificate_json_no_registry_url — JSON есть, но ссылки на разрешённый реестр внутри нет;
    - certificate_json_error — сетевой/JSON сбой, нужен UI fallback.
    """
    headers = {
        "User-Agent": args.user_agent,
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Origin": "https://www.wildberries.ru",
        "Referer": card.product_url or product_url(card.nm_id),
    }
    timeout = aiohttp.ClientTimeout(total=max(4.0, float(getattr(args, "certificate_timeout_sec", 8.0))))
    not_found = 0
    errors: List[str] = []
    tried = 0
    max_hosts = int(getattr(args, "certificate_max_hosts", 30) or 30)
    async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
        for url in certificate_json_urls(card.nm_id)[:max_hosts]:
            tried += 1
            try:
                async with session.get(url) as r:
                    status = int(r.status)
                    if status == 404:
                        not_found += 1
                        continue
                    txt = await r.text(errors="replace")
                    if status != 200:
                        errors.append(f"{status}:{url}")
                        continue
                    raw = txt.strip()
                    parsed: Any = raw
                    try:
                        parsed = json.loads(raw) if raw else {}
                    except Exception:
                        parsed = raw
                    urls = extract_registry_urls_deep(parsed)
                    if not urls and raw:
                        urls = extract_registry_urls_deep(raw)
                    if urls:
                        return urls, f"certificate_json_ok:{url}"
                    short = re.sub(r"\s+", " ", raw)[:180]
                    if raw and raw not in {"{}", "[]", "null"}:
                        return [], f"certificate_json_no_registry_url:{url}; body={short}"
                    return [], f"certificate_json_empty:{url}"
            except Exception as e:
                errors.append(f"{type(e).__name__}:{str(e)[:80]}")
                continue
    if not_found and not errors:
        return [], f"certificate_json_no_docs:tried={tried},404={not_found}"
    return [], f"certificate_json_error:tried={tried},404={not_found},errors={';'.join(errors[:4])}"


def strip_html_to_text(raw: str) -> str:
    """Грубое преобразование HTML WB/SEO-страницы в текст для поиска бейджа «Оригинал»."""
    txt = html.unescape(str(raw or ""))
    txt = re.sub(r"<script\b[\s\S]*?</script>", " ", txt, flags=re.I)
    txt = re.sub(r"<style\b[\s\S]*?</style>", " ", txt, flags=re.I)
    txt = re.sub(r"<[^>]+>", "\n", txt)
    txt = re.sub(r"[ \t\r\f\v]+", " ", txt)
    txt = re.sub(r"\n\s*\n+", "\n", txt)
    return txt.strip()


def public_html_has_original_badge(raw: str, card: BrandCard) -> Tuple[bool, str]:
    """Проверка SEO/HTML fallback.

    Нужна для случаев, когда обычный браузер WB показывает кнопку «Оригинал»,
    а Playwright/headless-рендер не успевает/не может увидеть productHeaderBadges.
    Не считаем словоформу «оригинал» внутри названия товара достаточной — ищем
    отдельную строку/маркер бейджа рядом с блоком характеристик/артикула.
    """
    if not raw:
        return False, "empty"
    raw_dec = html.unescape(str(raw))
    raw_low = norm_text(raw_dec)
    text = strip_html_to_text(raw_dec)
    lines = [re.sub(r"\s+", " ", x).strip() for x in text.splitlines()]
    lines = [x for x in lines if x]
    low_lines = [norm_text(x) for x in lines]
    nm = str(card.nm_id or "")
    brand = norm_text(card.brand or "")
    name = norm_text(card.product_name or "")

    # Страница должна быть похожа именно на карточку этого товара.
    product_evidence = False
    if nm and nm in raw_dec:
        product_evidence = True
    if brand and len(brand) >= 2 and brand in raw_low:
        product_evidence = True
    if name and len(name) >= 5 and name[:80] in raw_low:
        product_evidence = True
    if not product_evidence:
        return False, "no_product_evidence"

    # 1) Самый надёжный SEO-паттерн WB: «Оригинал. Характеристики ... Артикул ...».
    compact = norm_text(text)
    if re.search(r"(?:^|\s)оригинал\s*[\.·•|/-]?\s*(?:характеристики|описание|артикул|цвет|состав|пол|страна\s+производства)(?:\s|$)", compact, flags=re.I):
        return True, "public_html_original_before_specs"

    # 2) Отдельная строка «Оригинал» в верхней части товарного текста с товарным контекстом.
    for i, l in enumerate(low_lines[:260]):
        if re.fullmatch(r"оригинал|original", l):
            ctx = " ".join(low_lines[max(0, i-12): min(len(low_lines), i+18)])
            if re.search(r"артикул|характеристики|описание|цвет|состав|оценк|вопрос|купить|корзин", ctx):
                return True, f"public_html_exact_line_i={i}"

    # 3) JSON/React/атрибуты в HTML, если WB отдал не текст, а state/class.
    if re.search(r"originalmark|productheaderbadges|метк[аи]?\s+оригинал", raw_low, flags=re.I):
        return True, "public_html_original_dom_marker"

    return False, "not_found"


async def fetch_original_from_public_html(card: BrandCard, args) -> Tuple[bool, str]:
    """Быстрый fallback определения «Оригинал» по публичному HTML/SEO-странице WB.

    Используется только если DOM/locator в Playwright вернули НЕТ. Для 417714674
    поисковая/SEO-выдача WB содержит отдельный маркер «Оригинал», хотя headless
    проверка может его не поймать.
    """
    if not card or not card.nm_id:
        return False, "no_nm"
    if not str_to_bool(getattr(args, "original_html_fallback", True)):
        return False, "disabled"
    timeout = aiohttp.ClientTimeout(total=max(2.5, float(getattr(args, "original_html_timeout_sec", 5.0))))
    headers = {
        "User-Agent": args.user_agent,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.7,en;q=0.5",
        "Referer": "https://www.wildberries.ru/",
    }
    nm = int(card.nm_id)
    domains_raw = str(getattr(args, "original_html_domains", "ru,by,kg,ge") or "ru,by,kg,ge")
    domains = [d.strip().lower().lstrip(".") for d in re.split(r"[,;\s]+", domains_raw) if d.strip()]
    urls = []
    for d in domains:
        host = d if d.startswith("wildberries.") or d.startswith("www.wildberries.") else f"www.wildberries.{d}"
        urls.append(f"https://{host}/catalog/{nm}/detail.aspx")
    # Убираем дубли, но сохраняем порядок.
    urls = list(dict.fromkeys(urls))
    errors: List[str] = []
    async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
        for url in urls:
            try:
                async with session.get(url, allow_redirects=True) as r:
                    status = int(r.status)
                    raw = await r.text(errors="replace")
                    if status != 200:
                        errors.append(f"{status}:{url}")
                        continue
                    ok, why = public_html_has_original_badge(raw, card)
                    if ok:
                        return True, f"{why}:{url}"
                    errors.append(f"no:{why}:{url}")
            except Exception as e:
                errors.append(f"{type(e).__name__}:{url}")
    return False, ";".join(errors[:4]) or "not_found"

def parse_price_rub(p: Dict[str, Any]) -> float:
    # WB часто отдаёт цены в копейках/сотых рубля.
    for key in ("salePriceU", "salePrice", "priceU", "price"):
        val = p.get(key)
        if isinstance(val, (int, float)) and val > 0:
            return round(float(val) / 100 if val > 10000 else float(val), 2)
    try:
        sizes = p.get("sizes") or []
        for s in sizes:
            price = s.get("price") or {}
            for key in ("total", "product", "basic"):
                val = price.get(key)
                if isinstance(val, (int, float)) and val > 0:
                    return round(float(val) / 100 if val > 10000 else float(val), 2)
    except Exception:
        pass
    return 0.0

# -----------------------------
# WB JSON collection
# -----------------------------

# Широкий набор расширителей запроса. Нужен, потому что WB по короткому запросу бренда
# часто отдаёт только верхние ~100 карточек, даже если limit больше. Разбиваем бренд на
# товарные запросы и потом строго фильтруем по бренду из карточки/JSON.
DEFAULT_BRAND_QUERY_TERMS = [
    "кроссовки", "кеды", "бутсы", "сандалии", "сланцы", "шлепанцы", "обувь", "ботинки", "сапоги",
    "футболка", "майка", "лонгслив", "топ", "поло", "рубашка", "джемпер", "свитер", "кофта",
    "худи", "толстовка", "свитшот", "олимпийка", "куртка", "ветровка", "пуховик", "жилет",
    "брюки", "штаны", "джоггеры", "леггинсы", "лосины", "тайтсы", "шорты", "юбка", "платье",
    "спортивный костюм", "тренировочный костюм", "костюм", "комплект", "форма", "термобелье", "белье",
    "носки", "гетры", "колготки", "перчатки", "шапка", "бейсболка", "кепка", "панама", "шарф",
    "рюкзак", "сумка", "сумка спортивная", "кошелек", "ремень", "очки", "часы", "мяч", "бутылка",
    "мужские", "женские", "детские", "для мужчин", "для женщин", "для мальчика", "для девочки",
    "спорт", "фитнес", "бег", "футбол", "баскетбол", "теннис", "йога", "плавание", "originals", "оригинал",
    "черный", "белый", "синий", "красный", "серый", "зеленый", "розовый", "бежевый",
]

DETAIL_API_URLS = [
    "https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm={nm_id}",
    "https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm={nm_id}",
]

def get_any(d: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if isinstance(d, dict) and k in d and d.get(k) not in (None, ""):
            return d.get(k)
    return ""


def value_text(v: Any) -> str:
    """Безопасно превращает значение из WB JSON в короткий текст. Словари не превращаем в "{'name': ...}"."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, dict):
        for k in ("name", "title", "supplierName", "sellerName", "brandName", "value"):
            if k in v and isinstance(v.get(k), (str, int, float)):
                return str(v.get(k)).strip()
    return ""

def deep_first_text(obj: Any, keys: Iterable[str], max_depth: int = 7) -> str:
    """Ищет первый текст по ключам в глубине JSON-ответа WB."""
    keyset = {str(k).lower() for k in keys}
    best = ""
    def walk(x: Any, depth: int):
        nonlocal best
        if best or depth > max_depth or x is None:
            return
        if isinstance(x, dict):
            for k, v in x.items():
                if str(k).lower() in keyset:
                    txt = value_text(v)
                    if txt:
                        best = txt
                        return
            for v in x.values():
                if isinstance(v, (dict, list)):
                    walk(v, depth + 1)
                    if best:
                        return
        elif isinstance(x, list):
            for v in x[:200]:
                walk(v, depth + 1)
                if best:
                    return
    walk(obj, 0)
    return best.strip()

def best_seller_from_product_json(p: Dict[str, Any]) -> str:
    # Не берём сырой supplierInfo как dict. Нужен именно name/supplierName/sellerName.
    for key in ("supplierName", "sellerName", "supplier", "seller"):
        txt = clean_seller_name_py(value_text(p.get(key)))
        if txt:
            return txt
    for key in ("supplierInfo", "sellerInfo", "merchant", "supplier", "seller"):
        if isinstance(p.get(key), dict):
            txt = clean_seller_name_py(deep_first_text(p[key], ("name", "supplierName", "sellerName", "title"), 4))
            if txt:
                return txt
    return clean_seller_name_py(deep_first_text(p, ("supplierName", "sellerName"), 6))

def product_to_card(p: Dict[str, Any], source_query: str, brand_query: str, args) -> Optional[BrandCard]:
    nm = safe_int(p.get("id") or p.get("nmId") or p.get("nm_id"))
    if nm <= 1000:
        return None
    b = str(get_any(p, "brand", "brandName"))
    if not brand_matches(b, brand_query, args.brand_match):
        return None
    name = str(get_any(p, "name", "productName"))
    subject = str(get_any(p, "subject", "subjectName", "subjectId"))
    seller = best_seller_from_product_json(p)
    supplier_id = str(get_any(p, "supplierId", "supplierID", "sellerId", "sellerID"))
    return BrandCard(
        nm_id=nm,
        product_name=name,
        brand=b,
        subject=subject,
        price_rub=parse_price_rub(p),
        seller_name=seller,
        supplier_id=supplier_id,
        source_query=source_query,
        product_url=product_url(nm),
    )

async def enrich_card_from_detail_api(session: aiohttp.ClientSession, card: BrandCard, args) -> BrandCard:
    # Добираем продавца/цену/название из карточного API. Если API не ответил — оставляем поисковые данные.
    for tmpl in DETAIL_API_URLS:
        try:
            data = await fetch_json(session, tmpl.format(nm_id=card.nm_id), timeout=args.detail_timeout_sec)
            if not data:
                continue
            products = recursive_find_products(data)
            if not products:
                continue
            p = products[0]
            card.product_name = card.product_name or str(get_any(p, "name", "productName"))
            card.brand = card.brand or str(get_any(p, "brand", "brandName"))
            card.subject = card.subject or str(get_any(p, "subject", "subjectName", "subjectId"))
            pr = parse_price_rub(p)
            if pr:
                card.price_rub = pr
            seller = best_seller_from_product_json(p)
            if seller:
                card.seller_name = seller
            supplier_id = str(get_any(p, "supplierId", "supplierID", "sellerId", "sellerID"))
            if supplier_id:
                card.supplier_id = supplier_id
            return card
        except Exception:
            continue
    return card

def is_bad_brand_value_py(s: Any) -> bool:
    t = norm_text(str(s or ""))
    if not t:
        return True
    bad = {
        "купить сейчас", "в корзину", "выбрать размер", "главная", "распродажа", "похожие",
        "мужчинам", "женщинам", "детям", "белье", "цвет", "таблица размеров", "размер",
    }
    if t in bad:
        return True
    if len(t) > 80:
        return True
    if re.search(r"\b(купить|корзин|размер|оценк|вопрос|артикул|состав|характеристик)\b", t):
        return True
    return False

def clean_product_name_py(s: Any, brand: str = "", nm_id: int = 0) -> str:
    t = re.sub(r"\s+", " ", str(s or "")).strip().strip(" -—–/|•·")
    if not t:
        return ""
    # Убираем типовые приставки/соседние элементы WB из строки заголовка.
    t = re.sub(r"^(?:РАСПРОДАЖА\s+|Похожие\s+)+", "", t, flags=re.I).strip()
    t = re.sub(r"\bОригинал\b", " ", t, flags=re.I).strip()
    if nm_id:
        t = re.sub(rf"\s*\b{re.escape(str(nm_id))}\b\s*", " ", t).strip()
    if brand:
        # В title WB часто название выглядит как: '<товар> <brand> <nm_id> купить за...'.
        t = re.sub(rf"\s+{re.escape(str(brand))}\s*$", "", t, flags=re.I).strip()
        # А в DOM productHeader может быть: '<brand> <товар>'.
        t = re.sub(rf"^{re.escape(str(brand))}\s+", "", t, flags=re.I).strip()
    # Отсекаем хвост рейтинга/вопросов, если попал из общего header-блока.
    t = re.sub(r"\s+\d+[,.]?\d*\s*[·•]\s*\d+\s*(?:оцен|вопрос).*$", "", t, flags=re.I).strip()
    t = re.sub(r"\s+", " ", t).strip(" -—–/|•·")
    bad = {"wildberries", "купить сейчас", "в корзину", "главная", "цвет", "выбрать размер"}
    if not t or t.lower() in bad or len(t) > 220:
        return ""
    return t

def product_name_from_title(title: str, nm_id: int = 0, brand: str = "") -> str:
    """Fallback для WB title: 'Носки ... adidas 198242850 купить за ...'. Сохраняет регистр."""
    t = re.sub(r"\s+", " ", str(title or "")).strip()
    if not t:
        return ""
    t = re.sub(r"\s+купить\s+за\s+.*$", "", t, flags=re.I).strip()
    return clean_product_name_py(t, brand=brand, nm_id=nm_id)

async def enrich_single_card_from_api(card: BrandCard, args) -> BrandCard:
    headers = {
        "User-Agent": args.user_agent,
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    timeout = aiohttp.ClientTimeout(total=max(15, args.detail_timeout_sec + 5))
    try:
        async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
            return await enrich_card_from_detail_api(session, card, args)
    except Exception:
        return card


def recursive_find_products(obj: Any) -> List[Dict[str, Any]]:
    res: List[Dict[str, Any]] = []
    if isinstance(obj, dict):
        if isinstance(obj.get("products"), list):
            for p in obj["products"]:
                if isinstance(p, dict) and ("id" in p or "nmId" in p or "nm_id" in p):
                    res.append(p)
        for v in obj.values():
            if isinstance(v, (dict, list)):
                res.extend(recursive_find_products(v))
    elif isinstance(obj, list):
        for x in obj:
            if isinstance(x, (dict, list)):
                res.extend(recursive_find_products(x))
    return res

async def fetch_json(session: aiohttp.ClientSession, url: str, timeout: float = 14.0) -> Optional[Dict[str, Any]]:
    try:
        async with session.get(url, timeout=timeout) as r:
            if r.status != 200:
                return None
            txt = await r.text()
            return json.loads(txt)
    except Exception:
        return None


def recursive_find_brand_filter_ids(obj: Any, brand: str) -> List[str]:
    """Пытается найти id бренда в resultset=filters WB. Структура у WB менялась, поэтому обход общий."""
    wanted = norm_key(brand)
    ids: List[str] = []
    def add_id(v: Any):
        if v is None:
            return
        txt = str(v).strip()
        if txt and re.fullmatch(r"\d{1,10}", txt) and txt not in ids:
            ids.append(txt)
    def is_brand_like_name(x: Any) -> bool:
        k = norm_key(str(x or ""))
        return bool(k) and (k == wanted or wanted in k or k in wanted)
    def walk(x: Any, in_brand_filter: bool = False):
        if isinstance(x, dict):
            name = x.get("name") or x.get("title") or x.get("value") or x.get("label")
            key = norm_key(str(x.get("key") or x.get("type") or x.get("id") or ""))
            header = norm_key(str(name or ""))
            brand_filter = in_brand_filter or header in {"бренд", "brand", "brands"} or key in {"fbrand", "brand", "brands"}
            if is_brand_like_name(name):
                for k in ("id", "value", "key"):
                    if k in x and str(x.get(k)).strip() != str(name).strip():
                        add_id(x.get(k))
                # иногда id бренда лежит в поле "ids"/"values"
                for k in ("ids", "values"):
                    if isinstance(x.get(k), list):
                        for v in x[k]: add_id(v)
            if brand_filter:
                for arr_key in ("items", "values", "list", "filters"):
                    arr = x.get(arr_key)
                    if isinstance(arr, list):
                        for item in arr:
                            if isinstance(item, dict) and is_brand_like_name(item.get("name") or item.get("title") or item.get("value") or item.get("label")):
                                for k in ("id", "value", "key"):
                                    add_id(item.get(k))
            for v in x.values():
                if isinstance(v, (dict, list)):
                    walk(v, brand_filter)
        elif isinstance(x, list):
            for v in x:
                walk(v, in_brand_filter)
    walk(obj)
    return ids[:20]

async def discover_brand_filter_ids(session: aiohttp.ClientSession, brand: str, args) -> List[str]:
    q = urllib.parse.quote(brand)
    urls = [
        f"https://search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30&suppressSpellcheck=false",
        f"https://u-search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30&suppressSpellcheck=false",
        f"https://search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30",
        f"https://u-search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30",
    ]
    found: List[str] = []
    for url in urls:
        data = await fetch_json(session, url, timeout=max(8, args.collect_timeout_sec))
        if not data:
            continue
        for bid in recursive_find_brand_filter_ids(data, brand):
            if bid not in found:
                found.append(bid)
    # v24: дополнительный надёжный источник — поисковые подсказки WB по бренду.
    # Эндпоинт отдаёт сущности с прямым brandId. Это ловит ID даже когда фильтры пусты.
    suggest_urls = [
        f"https://search.wb.ru/exactmatch/common/v5/search?query={q}&resultset=suggestions",
        f"https://search.wb.ru/suggests/common/v5/search?query={q}",
        f"https://suggests.wb.ru/api/v6/hint?query={q}&gender=common&locale=ru",
    ]
    for url in suggest_urls:
        data = await fetch_json(session, url, timeout=max(6, args.collect_timeout_sec))
        if not data:
            continue
        for bid in _find_brand_ids_in_suggestions(data, brand):
            if bid not in found:
                found.append(bid)
    return found[:10]


def _find_brand_ids_in_suggestions(obj: Any, brand: str) -> List[str]:
    """v24: ищет brandId/id в ответах подсказок WB, где элемент относится к нашему бренду."""
    wanted = norm_key(brand)
    ids: List[str] = []
    def walk(x: Any):
        if isinstance(x, dict):
            name = x.get("name") or x.get("text") or x.get("title") or x.get("value")
            nk = norm_key(str(name or ""))
            type_hint = norm_key(str(x.get("type") or x.get("entity") or ""))
            is_brand_entity = ("brand" in type_hint) or bool(x.get("brandId")) or bool(x.get("brand_id"))
            name_matches = nk and (nk == wanted or wanted in nk or nk in wanted)
            if is_brand_entity and (name_matches or not name):
                for k in ("brandId", "brand_id", "id"):
                    v = x.get(k)
                    if v is not None and re.fullmatch(r"\d{1,10}", str(v).strip()):
                        if str(v).strip() not in ids:
                            ids.append(str(v).strip())
            for v in x.values():
                if isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)
    walk(obj)
    return ids

def brand_matches(found_brand: str, wanted: str, mode: str = "exact") -> bool:
    fb = norm_key(found_brand)
    wb = norm_key(wanted)
    if not wb:
        return True
    if mode == "any":
        return True
    if mode == "contains":
        return wb in fb or fb in wb
    return fb == wb

async def collect_brand_page(session: aiohttp.ClientSession, brand: str, source_query: str, page_num: int, sort: str, args) -> List[BrandCard]:
    q = urllib.parse.quote(source_query)
    base_urls = [
        f"https://search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={page_num}&query={q}&resultset=catalog&sort={sort}&spp=30&suppressSpellcheck=false",
        f"https://u-search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={page_num}&query={q}&resultset=catalog&sort={sort}&spp=30&suppressSpellcheck=false",
        f"https://search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page={page_num}&query={q}&resultset=catalog&sort={sort}&spp=30",
        f"https://u-search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page={page_num}&query={q}&resultset=catalog&sort={sort}&spp=30",
    ]
    urls = []
    fbrand_ids = list(getattr(args, "_fbrand_ids", []) or [])
    # v24: ПРЯМОЙ каталог бренда по fbrand БЕЗ поискового слова — отдаёт полный
    # каталог бренда (а не топ поисковой выдачи, которая обрезается на ~сотнях).
    # Это главный источник полноты; добавляем его ПЕРВЫМ.
    for bid in fbrand_ids:
        urls.append(
            f"https://catalog.wb.ru/brands/v2/catalog?ab_testing=false&appType=1&brand={urllib.parse.quote(str(bid))}"
            f"&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={page_num}&sort={sort}&spp=30"
        )
        urls.append(
            f"https://catalog.wb.ru/brands/catalog?appType=1&brand={urllib.parse.quote(str(bid))}"
            f"&curr=rub&dest=-1257786&lang=ru&page={page_num}&sort={sort}&spp=30"
        )
    for bu in base_urls:
        for bid in fbrand_ids:
            urls.append(bu + "&fbrand=" + urllib.parse.quote(str(bid)))
        urls.append(bu)
    best: List[BrandCard] = []
    for url in urls:
        data = await fetch_json(session, url)
        if not data:
            continue
        products = recursive_find_products(data)
        if not products:
            continue
        out: List[BrandCard] = []
        for p in products:
            c = product_to_card(p, source_query, brand, args)
            if c:
                out.append(c)
        if len(out) > len(best):
            best = out
        if out:
            return out
    return best

def build_collect_jobs(brand: str, args) -> List[Tuple[str, int, str, str]]:
    sorts = [s.strip() for s in args.search_sorts.split(",") if s.strip()]
    expanded_sorts = [s.strip() for s in args.expanded_sorts.split(",") if s.strip()]
    terms = list(DEFAULT_BRAND_QUERY_TERMS)
    if args.extra_terms:
        terms.extend([x.strip() for x in args.extra_terms.split(",") if x.strip()])
    # сохраняем порядок и убираем дубликаты
    seen_terms: Set[str] = set()
    clean_terms: List[str] = []
    for t in terms:
        k = norm_key(t)
        if k and k not in seen_terms:
            seen_terms.add(k)
            clean_terms.append(t)
    jobs: List[Tuple[str, int, str, str]] = []
    # 1) исходный запрос бренда: много страниц и сортировок
    for sort in sorts:
        for page in range(1, args.max_pages + 1):
            jobs.append((brand, page, sort, "brand"))
    if args.collect_strategy in {"hybrid", "expanded"}:
        # 2) расширенные запросы: бренд + товарный термин. Это ломает ограничение выдачи в ~100 карточек.
        for term in clean_terms[: max(0, args.max_query_terms)]:
            query = f"{brand} {term}".strip()
            for sort in expanded_sorts:
                for page in range(1, args.expanded_pages + 1):
                    jobs.append((query, page, sort, "expanded"))
    if args.shuffle_collect_jobs:
        random.shuffle(jobs)
    return jobs

async def collect_brand_cards(args) -> List[BrandCard]:
    brand = args.brand or input("Введите название бренда: ").strip()
    if not brand:
        raise RuntimeError("Не указан бренд")
    jobs = build_collect_jobs(brand, args)
    print(
        f"Собираю карточки бренда: '{brand}', limit={args.limit}, brand_match={args.brand_match}, "
        f"strategy={args.collect_strategy}, jobs={len(jobs)}, base_pages={args.max_pages}, expanded_pages={args.expanded_pages}"
    )

    headers = {
        "User-Agent": args.user_agent,
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    timeout = aiohttp.ClientTimeout(total=max(20, args.collect_timeout_sec))
    sem = asyncio.Semaphore(args.collect_workers)
    cards: Dict[int, BrandCard] = {}
    query_hits: Dict[str, int] = {}
    stop = asyncio.Event()
    last_save = time.time()

    async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
        if args.use_brand_filter:
            try:
                args._fbrand_ids = await discover_brand_filter_ids(session, brand, args)
                if args._fbrand_ids:
                    print(f"[collect] найден fbrand-фильтр WB для бренда: {', '.join(args._fbrand_ids)}")
                else:
                    print("[collect] fbrand-фильтр WB не найден, работаю по поисковым запросам")
            except Exception as e:
                args._fbrand_ids = []
                if args.trace:
                    print(f"[collect] ошибка fbrand-фильтра: {type(e).__name__}: {str(e)[:120]}")
        async def worker(job):
            nonlocal last_save
            if stop.is_set():
                return
            source_query, page, sort, kind = job
            async with sem:
                if stop.is_set():
                    return
                page_cards = await collect_brand_page(session, brand, source_query, page, sort, args)
                if page_cards:
                    query_hits[source_query] = query_hits.get(source_query, 0) + len(page_cards)
                added = 0
                for c in page_cards:
                    old = cards.get(c.nm_id)
                    if old is None:
                        cards[c.nm_id] = c
                        added += 1
                    else:
                        # обновляем пустые поля более полными данными
                        for attr in ("product_name", "brand", "subject", "seller_name", "supplier_id"):
                            if not getattr(old, attr, "") and getattr(c, attr, ""):
                                setattr(old, attr, getattr(c, attr))
                        if not old.price_rub and c.price_rub:
                            old.price_rub = c.price_rub
                    if len(cards) >= args.limit:
                        stop.set()
                        break
                if added and args.cards_csv and (time.time() - last_save > args.collect_save_every_sec):
                    save_cards_csv(list(cards.values())[:args.limit], Path(args.cards_csv))
                    last_save = time.time()
        tasks = [asyncio.create_task(worker(j)) for j in jobs]
        done_count = 0
        for t in asyncio.as_completed(tasks):
            try:
                await t
            except asyncio.CancelledError:
                pass
            except Exception:
                pass
            done_count += 1
            if done_count % max(1, args.collect_log_every) == 0 or stop.is_set():
                topq = sorted(query_hits.items(), key=lambda x: x[1], reverse=True)[:5]
                topq_s = "; ".join(f"{q}:{n}" for q, n in topq)
                print(f"[collect] задач={done_count}/{len(tasks)}, карточек={len(cards)}, топ-запросы=[{topq_s}]")
                emit_progress("search", done_count, len(tasks))
            if stop.is_set():
                for x in tasks:
                    if not x.done():
                        x.cancel()
                break
        await asyncio.gather(*tasks, return_exceptions=True)

        if len(cards) < args.limit and args.browser_collect_fallback:
            more = await browser_collect_cards(brand, args, set(cards.keys()))
            for c in more:
                if c.nm_id not in cards:
                    cards[c.nm_id] = c
                    if len(cards) >= args.limit:
                        break

        result = list(cards.values())[:args.limit]
        if args.enrich_details and result:
            print(f"[collect] Добираю продавца/цену через карточный API: {len(result)} карточек")
            sem2 = asyncio.Semaphore(args.detail_workers)
            async def enrich_one(i: int, c: BrandCard):
                async with sem2:
                    await enrich_card_from_detail_api(session, c, args)
                    if (i + 1) % max(1, args.collect_log_every) == 0:
                        print(f"[collect-detail] {i+1}/{len(result)}")
            await asyncio.gather(*(enrich_one(i, c) for i, c in enumerate(result)), return_exceptions=True)
            # После детального API удаляем явные чужие бренды, но не выкидываем карточки с пустым брендом — их уточнит браузер на этапе проверки.
            filtered = [c for c in result if (not c.brand) or brand_matches(c.brand, brand, args.brand_match)]
            cards = {c.nm_id: c for c in filtered}

    result = list(cards.values())[:args.limit]
    save_cards_csv(result, Path(args.cards_csv))
    fbrand_found = bool(getattr(args, "_fbrand_ids", []))
    print(f"Собрано уникальных карточек бренда: {len(result)}. Файл: {Path(args.cards_csv).resolve()}")
    if len(result) < args.limit:
        print(f"ВНИМАНИЕ: собрано меньше limit ({len(result)}/{args.limit}).")
        if not fbrand_found:
            print("  ⚠ Не найден fbrand-ID бренда — сбор шёл только по поиску (он обрезается на ~сотнях).")
            print("    Проверьте написание бренда (латиницей как на WB: 'reebok', 'nike').")
            print("    Если у бренда реально мало товаров — это нормально.")
        else:
            print(f"  fbrand-ID найден ({', '.join(getattr(args,'_fbrand_ids',[]))}), собран каталог бренда.")
            print("  Если ожидали больше — увеличьте --max-pages, или у бренда столько товаров на WB.")
    return result


async def browser_collect_cards(brand: str, args, existing_ids: Set[int]) -> List[BrandCard]:
    if async_playwright is None:
        return []
    queries = [brand]
    terms = list(DEFAULT_BRAND_QUERY_TERMS)
    if args.extra_terms:
        terms.extend([x.strip() for x in args.extra_terms.split(",") if x.strip()])
    for term in terms[: max(0, args.browser_collect_terms)]:
        queries.append(f"{brand} {term}".strip())
    seen_q: Set[str] = set()
    queries = [q for q in queries if not (q in seen_q or seen_q.add(q))]
    out: Dict[int, BrandCard] = {}
    print(f"[browser-collect] не хватает карточек, включаю браузерный добор: запросов={len(queries)}")
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=args.headless,
            args=["--disable-blink-features=AutomationControlled", "--disable-dev-shm-usage", "--no-sandbox"],
            timeout=60000,
        )
        try:
            ctx = await browser.new_context(
                user_agent=args.user_agent,
                viewport={"width": args.viewport_width, "height": args.viewport_height},
                locale="ru-RU",
                ignore_https_errors=True,
            )
            if args.block_assets:
                await ctx.route("**/*", block_assets)
            page = await ctx.new_page()
            page.set_default_timeout(args.default_timeout_ms)
            for qi, query in enumerate(queries, 1):
                if len(existing_ids) + len(out) >= args.limit:
                    break
                url = "https://www.wildberries.ru/catalog/0/search.aspx?search=" + urllib.parse.quote(query)
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=args.goto_timeout_ms)
                    await asyncio.sleep(args.browser_collect_wait_ms / 1000)
                    stable_rounds = 0
                    last_count = 0
                    for scroll_i in range(args.browser_collect_scrolls):
                        links = await page.evaluate(r"""
() => Array.from(document.querySelectorAll('a[href*="/catalog/"][href*="/detail.aspx"]'))
  .map(a => a.href || a.getAttribute('href') || '')
""")
                        added = 0
                        for href in links or []:
                            m = re.search(r"/catalog/(\d+)/detail\.aspx", str(href))
                            if not m:
                                continue
                            nm = safe_int(m.group(1))
                            if nm <= 1000 or nm in existing_ids or nm in out:
                                continue
                            out[nm] = BrandCard(nm_id=nm, source_query=query, product_url=product_url(nm))
                            added += 1
                            if len(existing_ids) + len(out) >= args.limit:
                                break
                        if len(existing_ids) + len(out) >= args.limit:
                            break
                        if len(out) == last_count:
                            stable_rounds += 1
                        else:
                            stable_rounds = 0
                        last_count = len(out)
                        if stable_rounds >= args.browser_collect_stable_rounds:
                            break
                        await page.mouse.wheel(0, args.browser_collect_scroll_px)
                        await asyncio.sleep(args.browser_collect_scroll_wait_ms / 1000)
                    if qi % max(1, args.browser_collect_log_every) == 0 or added:
                        print(f"[browser-collect] {qi}/{len(queries)} query='{query}', всего добор={len(out)}")
                except Exception as e:
                    if args.trace:
                        print(f"[browser-collect] query='{query}' error: {type(e).__name__}: {str(e)[:120]}")
                    continue
            try:
                await ctx.close()
            except Exception:
                pass
        finally:
            try:
                await browser.close()
            except Exception:
                pass
    print(f"[browser-collect] добрал карточек: {len(out)}")
    return list(out.values())

def save_cards_csv(cards: List[BrandCard], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8-sig", newline="") as f:
        fields = list(asdict(BrandCard(0)).keys())
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for c in cards:
            w.writerow(asdict(c))
    os.replace(tmp, path)

def load_cards_csv(path: Path, limit: int = 0) -> List[BrandCard]:
    cards: List[BrandCard] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            nm = safe_int(row.get("nm_id"))
            if not nm:
                continue
            cards.append(BrandCard(
                nm_id=nm,
                product_name=row.get("product_name", ""),
                brand=row.get("brand", ""),
                subject=row.get("subject", ""),
                price_rub=safe_float(row.get("price_rub")),
                seller_name=row.get("seller_name", ""),
                supplier_id=row.get("supplier_id", ""),
                source_query=row.get("source_query", ""),
                product_url=row.get("product_url") or product_url(nm),
            ))
            if limit and len(cards) >= limit:
                break
    return cards

# -----------------------------
# Browser logic
# -----------------------------

CAPTURE_INIT_SCRIPT = r"""
(() => {
  if (window.__wbRegistryCaptureInstalled) return;
  window.__wbRegistryCaptureInstalled = true;
  window.__wbCapturedUrls = [];
  const pushUrl = (u, source) => {
    try {
      if (!u) return;
      const s = String(u);
      window.__wbCapturedUrls.push({url: s, source: source || 'unknown', ts: Date.now()});
    } catch(e) {}
  };
  const oldOpen = window.open;
  window.open = function(url, name, specs) {
    pushUrl(url, 'window.open');
    return oldOpen.apply(window, arguments);
  };
  const oldAssign = window.location.assign.bind(window.location);
  window.location.assign = function(url) { pushUrl(url, 'location.assign'); return oldAssign(url); };
  const oldReplace = window.location.replace.bind(window.location);
  window.location.replace = function(url) { pushUrl(url, 'location.replace'); return oldReplace(url); };
})();
"""
GET_CAPTURED_JS = "() => window.__wbCapturedUrls || []"
CLEAR_CAPTURED_JS = "() => { window.__wbCapturedUrls = []; }"

FIND_TEXT_TARGET_JS = r"""
(args) => {
  const wanted = (args.wanted || '').toLowerCase().replace(/ё/g,'е').replace(/\s+/g,' ').trim();
  const exact = !!args.exact;
  const clickableOnly = !!args.clickableOnly;
  const norm = s => (s || '').toString().toLowerCase().replace(/ё/g,'е').replace(/\s+/g,' ').trim();
  const visible = el => {
    const r = el.getBoundingClientRect();
    const st = getComputedStyle(el);
    return r.width > 2 && r.height > 2 && st.visibility !== 'hidden' && st.display !== 'none' && r.bottom > 0 && r.right > 0 && r.top < innerHeight && r.left < innerWidth;
  };
  const isClickable = el => {
    const tag = el.tagName;
    const role = (el.getAttribute('role') || '').toLowerCase();
    const cls = (el.className || '').toString().toLowerCase();
    return ['BUTTON','A'].includes(tag) || role === 'button' || el.onclick || cls.includes('btn') || cls.includes('button') || cls.includes('chip');
  };
  const els = Array.from(document.querySelectorAll('button,a,[role="button"],div,span,li'));
  const out = [];
  for (const el of els) {
    if (!visible(el)) continue;
    if (clickableOnly && !isClickable(el)) continue;
    const text = norm(el.innerText || el.textContent || el.getAttribute('aria-label') || el.getAttribute('title') || '');
    if (!text) continue;
    let match = exact ? text === wanted : (text === wanted || text.includes(wanted));
    if (!match) continue;
    const r = el.getBoundingClientRect();
    let score = 0;
    if (text === wanted) score += 30000;
    if (['BUTTON','A'].includes(el.tagName)) score += 9000;
    if (isClickable(el)) score += 5000;
    score -= Math.abs(text.length - wanted.length) * 80;
    score -= (r.width * r.height) / 80;
    out.push({tag: el.tagName, text, score, x: r.left + r.width / 2, y: r.top + r.height / 2, w: r.width, h: r.height});
  }
  out.sort((a,b) => b.score - a.score);
  return out.slice(0, 20);
}
"""

PANEL_TEXT_JS = r"""
() => {
  const norm = s => (s || '').toString().toLowerCase().replace(/ё/g,'е').replace(/\s+/g,' ').trim();
  const txt = norm(document.body ? document.body.innerText : '');
  return {
    hasSpecs: txt.includes('характеристики и описание') || txt.includes('артикул') || txt.includes('состав'),
    hasDocs: txt.includes('документы проверены'),
    hasLook: txt.includes('смотреть на сайте'),
    hasOriginal: txt.split('\n').map(x => norm(x)).some(x => x === 'оригинал') || /(^|\s)оригинал($|\s)/.test(txt),
    textLen: txt.length
  };
}
"""

PAGE_DATA_JS = r"""
(args) => {
  const nmId = String((args && args.nmId) || '');
  const norm = s => (s || '').toString().replace(/\s+/g,' ').trim();
  const low = s => norm(s).toLowerCase().replace(/ё/g,'е');
  const cleanToken = s => low(s).replace(/[✓✔✅☑●•·✦★☆]/g,'').replace(/[\u200b-\u200f\uFEFF]/g,'').replace(/\s+/g,' ').trim();
  const visible = el => {
    try {
      if (!el || !el.getBoundingClientRect) return false;
      const r = el.getBoundingClientRect();
      const st = getComputedStyle(el);
      return r.width > 1 && r.height > 1 && st.visibility !== 'hidden' && st.display !== 'none' && st.opacity !== '0' && r.bottom > 0 && r.right > 0 && r.top < innerHeight && r.left < innerWidth;
    } catch(e) { return false; }
  };
  const rectObj = el => { const r = el.getBoundingClientRect(); return {left:r.left, right:r.right, top:r.top, bottom:r.bottom, width:r.width, height:r.height}; };
  const ownText = el => {
    try {
      let txt = '';
      for (const n of Array.from(el.childNodes || [])) if (n.nodeType === Node.TEXT_NODE) txt += ' ' + n.textContent;
      txt = norm(txt);
      return txt || norm(el.getAttribute('aria-label') || el.getAttribute('title') || el.getAttribute('alt') || '');
    } catch(e) { return norm(el && (el.innerText || el.textContent || '')); }
  };
  const textOf = el => norm(el && (el.innerText || el.textContent || el.getAttribute?.('aria-label') || el.getAttribute?.('title') || ''));
  const isPageChrome = el => {
    let x = el;
    for (let i=0; x && i<8; i++, x=x.parentElement) {
      const tag = (x.tagName || '').toLowerCase();
      const cls = low((x.className || '') + ' ' + (x.id || '') + ' ' + (x.getAttribute && (x.getAttribute('role') || '')));
      // Важно: aside НЕ считаем плохой зоной — у WB продавец часто в правом блоке покупки.
      if (['header','footer','nav'].includes(tag)) return true;
      if (/(header|footer|navbar|breadcrumbs|bread-crumb|filter|filters|catalog-page|search-page|sort|popup|modal|tooltip-content|notification|cookie)/.test(cls)) return true;
    }
    return false;
  };
  const pathClass = el => {
    const arr = [];
    let x = el;
    for (let i=0; x && i<7; i++, x=x.parentElement) {
      arr.push(low((x.tagName || '') + ' ' + (x.className || '') + ' ' + (x.id || '') + ' ' + (x.getAttribute && (x.getAttribute('data-link') || x.getAttribute('data-testid') || x.getAttribute('aria-label') || x.getAttribute('title') || ''))));
    }
    return arr.join(' | ');
  };

  let h1 = '';
  const h = document.querySelector('h1');
  let hRect = null;
  if (h && visible(h)) { h1 = norm(h.innerText || h.textContent || ''); hRect = rectObj(h); }
  const text = norm(document.body ? document.body.innerText : '');
  const docTitle = norm(document.title || '');
  const bodyLines = (document.body ? (document.body.innerText || '') : '').split(/\n+/).map(norm).filter(Boolean);
  const badBrandWords = new Set(['купить сейчас','в корзину','выбрать размер','главная','распродажа','похожие','мужчинам','женщинам','детям','белье','цвет','таблица размеров','размер']);
  const badBrandValue = s => {
    const t = low(s || '');
    if (!t || badBrandWords.has(t)) return true;
    if (t.length > 80) return true;
    if (/\b(купить|корзин|размер|оценк|вопрос|артикул|состав|характеристик)\b/.test(t)) return true;
    return false;
  };
  const escRe = s => String(s || '').replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
  const titleProductName = (brandHint) => {
    let t = docTitle.replace(/\s+купить\s+за\s+.*$/i, '').trim();
    if (nmId) t = t.replace(new RegExp('\s*\b' + escRe(nmId) + '\b\s*'), ' ').trim();
    if (brandHint) t = t.replace(new RegExp('\s+' + escRe(brandHint) + '\s*$', 'i'), '').trim();
    return cleanProductName(t);
  };
  const cleanProductName = s => {
    let t = norm(s || '');
    t = t.replace(/^[^\S\r\n]*РАСПРОДАЖА\s+/i, '').replace(/^Похожие\s+/i, '').trim();
    t = t.replace(/\bОригинал\b/ig, ' ').replace(/\s+/g, ' ').trim();
    // Если строка вида "adidas Носки ...", бренд вырежем позже после определения brandName.
    t = t.replace(/\s+\d+[,.]?\d*\s*[·•]\s*\d+\s*оцен.*$/i, '').trim();
    return t;
  };

  // -------------------------
  // Бренд
  // -------------------------
  let brandName = '';
  let brandRect = null;
  try {
    const brandSelectors = [
      'a[href*="/brands/"]','a[href*="/brand/"]','[class*="brand"] a','[data-link*="brand"] a',
      '[class*="product-page__brand"]','[class*="product-page"] [class*="brand"]'
    ];
    for (const sel of brandSelectors) {
      for (const el of Array.from(document.querySelectorAll(sel))) {
        if (!visible(el) || isPageChrome(el)) continue;
        const t = norm(el.innerText || el.textContent || el.getAttribute('title') || el.getAttribute('aria-label') || '');
        const tl = low(t);
        if (t && t.length <= 80 && !tl.includes('каталог') && !tl.includes('главная') && tl !== 'бренд') {
          brandName = t; brandRect = rectObj(el); break;
        }
      }
      if (brandName) break;
    }
  } catch(e) {}
  try {
    const badBrand = new Set(['купить сейчас','в корзину','выбрать размер','главная','распродажа','похожие','мужчинам','женщинам','детям','белье']);
    if (!brandName || badBrand.has(low(brandName))) {
      const badges = Array.from(document.querySelectorAll('[class*="productHeaderBadges"], [class*="productHeader"]')).filter(el => visible(el));
      for (const el of badges) {
        const lines = (el.innerText || el.textContent || '').split(/\n+/).map(norm).filter(Boolean);
        for (const ln of lines) {
          const cand = norm(ln.replace(/\bОригинал\b/ig, '').trim());
          const l = low(cand);
          if (cand && cand.length <= 80 && !badBrand.has(l) && !/^(оригинал|original)$/i.test(l)) { brandName = cand; brandRect = rectObj(el); break; }
        }
        if (brandName && !badBrand.has(low(brandName))) break;
      }
    }
  } catch(e) {}

  // Название товара: у нового WB часто нет h1, название лежит в productHeader рядом с брендом/плашкой.
  try {
    if (!h1) {
      const headerBlocks = Array.from(document.querySelectorAll('[class*="productHeader"], [data-testid="cardtype:colors"], [class*="mainWrap"]')).filter(el => visible(el));
      for (const el of headerBlocks) {
        let t = cleanProductName(el.innerText || el.textContent || '');
        if (brandName && low(t).startsWith(low(brandName) + ' ')) t = norm(t.slice(brandName.length));
        if (t && t.length >= 4 && t.length <= 180 && !/^(купить|в корзину|цвет|таблица|размер|артикул|состав)/i.test(t)) { h1 = t; break; }
      }
    }
    if (!h1 && docTitle) {
      let t = docTitle.replace(/\s+купить\s+за\s+.*$/i, '').trim();
      if (nmId) t = t.replace(new RegExp('\\s*\\b' + nmId + '\\b\\s*'), ' ').trim();
      if (brandName) t = t.replace(new RegExp('\\s+' + brandName.replace(/[.*+?^${}()|[\]\\]/g, '\\$&') + '\\s*$', 'i'), '').trim();
      h1 = cleanProductName(t);
    }
  } catch(e) {}

  // -------------------------
  // Подстраховка метаданных по фактической структуре WB из диагностического DOM:
  // bodyLines: "adidas / Носки..." в верхнем мини-блоке и далее "adidas", "Оригинал", "Носки...".
  // -------------------------
  try {
    // 1) Самый чистый источник названия в body: строка вида "adidas / Носки T LIN ANKLE 3P".
    const slashLine = bodyLines.slice(0,140).find(l => /^[^/]{2,80}\s+\/\s+.{3,180}$/.test(l) && !/wildberries|каталог|главная/i.test(l));
    if (slashLine) {
      const parts = slashLine.split('/').map(norm).filter(Boolean);
      if ((!brandName || badBrandValue(brandName)) && parts[0] && parts[0].length <= 80) { brandName = parts[0]; }
      if ((!h1 || /^(купить|в корзину|цвет|таблица|размер|артикул|состав)/i.test(h1)) && parts[1]) { h1 = cleanProductName(parts.slice(1).join(' / ')); }
    }

    // 2) productHeaderBadges в реальном DOM: "adidas Оригинал".
    if (!brandName || badBrandValue(brandName)) {
      for (const el of Array.from(document.querySelectorAll('[class*="productHeaderBadges"]')).filter(visible)) {
        let t = norm(el.innerText || el.textContent || '').replace(/\bОригинал\b/ig, ' ').replace(/\boriginal\b/ig, ' ').trim();
        if (t && t.length <= 80 && !badBrandValue(t)) { brandName = t; brandRect = rectObj(el); break; }
      }
    }

    // 3) productHeader: "adidas Оригинал Носки T LIN ANKLE 3P ...".
    const headerEls = Array.from(document.querySelectorAll('[class*="productHeader"]')).filter(visible);
    for (const el of headerEls) {
      const ht = norm(el.innerText || el.textContent || '');
      if (!ht || !/Оригинал|original|оценк|вопрос/i.test(ht)) continue;
      if ((!brandName || badBrandValue(brandName))) {
        const mBrand = ht.match(/^(.{2,80}?)\s+(?:Оригинал|original)\b/i);
        if (mBrand) {
          const b = norm(mBrand[1]);
          if (b && !badBrandValue(b)) brandName = b;
        }
      }
      if (!h1 || /^(купить|в корзину|цвет|таблица|размер|артикул|состав)/i.test(h1)) {
        let t = ht;
        if (brandName) t = t.replace(new RegExp('^' + escRe(brandName) + '\s+', 'i'), '');
        t = t.replace(/^(?:Оригинал|original)\s+/i, '').replace(/\b(?:Оригинал|original)\b/i, ' ');
        t = t.replace(/\s+\d+[,.]?\d*\s*[·•]\s*\d+\s*(?:оцен|вопрос).*$/i, '');
        t = cleanProductName(t);
        if (t && t.length >= 4) { h1 = t; break; }
      }
    }

    if ((!h1 || h1.length < 4) && docTitle) h1 = titleProductName(brandName);
  } catch(e) {}

  // -------------------------
  // Продавец. Не возвращаем служебную метку «Продавец».
  // -------------------------
  const badSellerWords = new Set(['продавец','о продавце','перейти к продавцу','задать вопрос','seller','wildberries продавец','товары продавца','магазин продавца','рейтинг продавца']);
  const cleanSeller = raw => {
    let t = norm(raw || '').replace(/^["'«»]+|["'«»]+$/g,'').trim();
    t = t.replace(/^продавец\s*[:\-–—]?\s*/i, '').trim();
    t = t.replace(/\s+о продавце.*$/i, '').replace(/\s+перейти к продавцу.*$/i, '').replace(/\s+задать вопрос.*$/i, '').trim();
    t = t.replace(/^магазин\s+/i,'').trim();
    const tl = low(t);
    if (!t || t.length < 2 || t.length > 120) return '';
    if (badSellerWords.has(tl)) return '';
    if (/^(продавец|о продавце|перейти к продавцу|задать вопрос|рейтинг продавца|показать продавца)$/i.test(t)) return '';
    if (/^(доставка|возврат|оплата|покупают|реклама|спонсорский|склад wb)$/i.test(t)) return '';
    if (/^(поставил|доставит|пользовател|оценк|вопрос|отзыв|отзывы|послезавтра|завтра|сегодня)/i.test(t)) return '';
    if (/^[0-9\s.,₽%+\-–—]+$/.test(t)) return '';
    return t;
  };
  const linesFrom = el => textOf(el).split(/\n|\r| {2,}/).map(norm).filter(Boolean);
  const bestSellerFromLines = lines => {
    const arr = (lines || []).map(norm).filter(Boolean);
    for (let i=0; i<arr.length; i++) {
      const l = low(arr[i]);
      let m = arr[i].match(/^продавец\s*[:\-–—]?\s*(.+)$/i);
      if (m) { const cand = cleanSeller(m[1]); if (cand) return cand; }
      if (l === 'продавец' || l === 'продавец:') {
        for (let j=i+1; j<Math.min(arr.length, i+8); j++) {
          const cand = cleanSeller(arr[j]);
          if (cand) return cand;
        }
      }
    }
    // Если в блоке продавца нет явной метки, берём первую осмысленную строку, но не служебную.
    for (const l of arr.slice(0,12)) { const cand = cleanSeller(l); if (cand) return cand; }
    return '';
  };
  const labelSellerFromLines = lines => {
    const arr = (lines || []).map(norm).filter(Boolean);
    for (let i=0; i<arr.length; i++) {
      const l = low(arr[i]);
      let m = arr[i].match(/^продавец\s*[:\-–—]?\s*(.+)$/i);
      if (m) { const cand = cleanSeller(m[1]); if (cand) return cand; }
      if (l === 'продавец' || l === 'продавец:') {
        for (let j=i+1; j<Math.min(arr.length, i+8); j++) { const cand = cleanSeller(arr[j]); if (cand) return cand; }
      }
    }
    return '';
  };

  let seller = '';
  let supplierId = '';
  try {
    const sellerLinks = Array.from(document.querySelectorAll('a[href*="/seller/"],a[href*="/sellers/"],a[href*="/supplier/"],a[href*="/suppliers/"]'))
      .filter(el => visible(el) && !isPageChrome(el));
    for (const el of sellerLinks) {
      const href = el.getAttribute('href') || '';
      const idm = href.match(/\/(?:seller|sellers|supplier|suppliers)\/(\d+)/i);
      if (!supplierId && idm) supplierId = idm[1];
      let cand = cleanSeller(el.innerText || el.textContent || el.getAttribute('title') || el.getAttribute('aria-label') || '');
      if (!cand) {
        // Очень частый случай WB: ссылка/иконка имеет текст «Продавец», а имя лежит рядом в родителе.
        let box = el;
        for (let depth=0; box && depth<7 && !cand; depth++, box=box.parentElement) cand = bestSellerFromLines(linesFrom(box));
      }
      if (cand) { seller = cand; break; }
    }
  } catch(e) {}
  if (!seller) {
    // Метка «Продавец» + соседний текст в правом блоке карточки.
    try {
      const labels = Array.from(document.querySelectorAll('body *')).filter(el => visible(el) && !isPageChrome(el) && /^продавец:?$/i.test(low(ownText(el))));
      for (const lab of labels) {
        let box = lab.parentElement;
        for (let depth=0; box && depth<7 && !seller; depth++, box=box.parentElement) seller = bestSellerFromLines(linesFrom(box));
        if (seller) break;
      }
    } catch(e) {}
  }
  if (!seller) seller = labelSellerFromLines(text.split(/\n|\r/).map(norm).filter(Boolean));

  // -------------------------
  // Оригинал.
  // Разобрано по реальному DOM карточки WB 198242850 из обычного браузера:
  //   button.mo-button...originalMark--ZeMYb
  //   aria-label="что означает метка Оригинал?"
  //   parent div.productHeaderBadges--... с текстом "adidas Оригинал"
  //   предок div.background--... header--...
  // Старый код ошибочно отбрасывал эту область как page chrome из-за класса header--...
  // Поэтому для оригинальности используем отдельную проверку зоны товара, а НЕ общий isPageChrome().
  // -------------------------
  const originalCandidates = [];

  const isBadOriginalArea = el => {
    // Отсекаем только реально внешнюю шапку/футер/модалки/попапы, но НЕ product header внутри карточки.
    let x = el;
    for (let i=0; x && i<8; i++, x=x.parentElement) {
      const tag = low(x.tagName || '');
      const cls = low((x.className || '') + ' ' + (x.id || '') + ' ' + (x.getAttribute && (x.getAttribute('role') || '')));
      const inProduct = /(product-page|productpage|productheader|productheaderbadges|mainwrap|cardtype|appreactroot|reactcontainers|body-layout|main__container)/.test(cls);
      if (inProduct) return false;
      if (['footer','nav'].includes(tag)) return true;
      if (tag === 'header') return true;
      if (/(navbar|breadcrumbs|bread-crumb|catalog-page|search-page|sort|filter|filters|popup|modal|tooltip|notification|cookie)/.test(cls)) return true;
    }
    return false;
  };

  const productAncestor = el => {
    let x = el;
    for (let i=0; x && i<10; i++, x=x.parentElement) {
      const cls = low((x.className || '') + ' ' + (x.id || '') + ' ' + (x.getAttribute && ((x.getAttribute('data-testid') || '') + ' ' + (x.getAttribute('role') || ''))));
      if (/(product-page|productpage|productheader|productheaderbadges|mainwrap|cardtype|appreactroot|reactcontainers|body-layout|main__container)/.test(cls)) return x;
    }
    return null;
  };

  const isOriginalWord = s => {
    const t = cleanToken(s || '').replace(/^['"«]+|['"»]+$/g,'').trim();
    if (t === 'оригинал' || t === 'original') return true;
    // aria-label WB: "что означает метка Оригинал?"
    if (/метк[аи]?\s+оригинал/i.test(t)) return true;
    return false;
  };

  const addOriginalCandidate = (el, source, textHint) => {
    if (!el || !visible(el) || isBadOriginalArea(el)) return;
    const r = rectObj(el);
    if (r.width < 6 || r.height < 6) return;
    if (r.top < 60 || r.top > Math.min(innerHeight - 20, 420)) return;

    const own = ownText(el);
    const all = norm(el.innerText || el.textContent || '');
    const aria = norm(el.getAttribute && (el.getAttribute('aria-label') || ''));
    const title = norm(el.getAttribute && (el.getAttribute('title') || ''));
    const cls = pathClass(el);
    const textForCheck = norm(textHint || own || all || aria || title || '');

    const classHit = /(originalmark|original-mark|оригинал)/i.test(cls);
    const exactTextHit = isOriginalWord(own) || isOriginalWord(all) || isOriginalWord(textForCheck);
    const ariaHit = isOriginalWord(aria) || /оригинал/i.test(aria);
    const headerBadgesHit = /productheaderbadges|productheader|originalmark/i.test(cls) && /(^|\s)оригинал($|\s)/i.test(low(all || textForCheck));

    if (!(classHit || exactTextHit || ariaHit || headerBadgesHit)) return;

    const prod = productAncestor(el);
    const inLikelyProductX = (r.left > innerWidth * 0.22 && r.left < innerWidth * 0.88);
    const inLikelyProductY = (r.top >= 100 && r.top <= 360);
    if (!prod && !(inLikelyProductX && inLikelyProductY)) return;

    // Не берём огромные контейнеры. Исключение: productHeaderBadges, где бывает "adidas Оригинал".
    if (r.width > 520 || r.height > 110) return;
    if (all.length > 120 && !/productheaderbadges|productheader/i.test(cls)) return;

    let score = 1000;
    if (classHit) score += 1800;
    if (exactTextHit) score += 1500;
    if (ariaHit) score += 1200;
    if (headerBadgesHit) score += 1000;
    if (prod) score += 500;
    if (r.width <= 140 && r.height <= 40) score += 400;
    if (/button/i.test(el.tagName || '')) score += 500;
    if (/productheaderbadges|productheader/i.test(cls)) score += 500;

    originalCandidates.push({
      source,
      text: norm(textForCheck || all || aria).slice(0,120),
      ownText: norm(own).slice(0,120),
      allText: norm(all).slice(0,160),
      aria: aria.slice(0,160),
      x:Math.round(r.left), y:Math.round(r.top), w:Math.round(r.width), h:Math.round(r.height),
      tag:el.tagName,
      cls:cls.slice(0,260),
      score
    });
  };

  // 1) Самый точный селектор по фактическому DOM WB: class originalMark--... / aria-label с меткой.
  try {
    const directSelectors = [
      'button[class*="originalMark"]',
      '[class*="originalMark"]',
      'button[class*="original" i]',
      '[class*="original" i]',
      'button[aria-label*="Оригинал"]',
      'button[aria-label*="оригинал"]',
      '[aria-label*="Оригинал"]',
      '[aria-label*="оригинал"]',
      '[title*="Оригинал"]',
      '[title*="оригинал"]',
      '[class*="productHeaderBadges"] button',
      '[class*="productHeaderBadges"] [role="button"]',
      '[class*="productHeader"] button'
    ];
    for (const sel of directSelectors) {
      for (const el of Array.from(document.querySelectorAll(sel))) {
        addOriginalCandidate(el, 'direct-selector:' + sel, el.innerText || el.textContent || el.getAttribute('aria-label') || '');
      }
    }
  } catch(e) {}

  // 2) Текстовые элементы с точным текстом/aria "Оригинал".
  try {
    for (const el of Array.from(document.querySelectorAll('button,span,div,a,[role="button"],[aria-label],[title]'))) {
      const rawOwn = cleanToken(ownText(el));
      const rawAll = cleanToken(el.innerText || el.textContent || '');
      const rawAria = cleanToken(el.getAttribute && (el.getAttribute('aria-label') || ''));
      const rawTitle = cleanToken(el.getAttribute && (el.getAttribute('title') || ''));
      if (isOriginalWord(rawOwn)) addOriginalCandidate(el, 'own-text', rawOwn);
      if (isOriginalWord(rawAll)) addOriginalCandidate(el, 'all-text', rawAll);
      if (isOriginalWord(rawAria) || /оригинал/i.test(rawAria)) addOriginalCandidate(el, 'aria', rawAria);
      if (isOriginalWord(rawTitle) || /оригинал/i.test(rawTitle)) addOriginalCandidate(el, 'title', rawTitle);
    }
  } catch(e) {}

  // 3) Отдельный текстовый узел "Оригинал". Поднимаемся к parent/button.
  try {
    const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT, {
      acceptNode(node) {
        const t = cleanToken(node.nodeValue || '');
        return isOriginalWord(t) ? NodeFilter.FILTER_ACCEPT : NodeFilter.FILTER_REJECT;
      }
    });
    let node; let guard = 0;
    while ((node = walker.nextNode()) && guard++ < 200) {
      let el = node.parentElement;
      for (let depth=0; el && depth<6; depth++, el=el.parentElement) {
        addOriginalCandidate(el, 'text-node-parent-' + depth, cleanToken(node.nodeValue || ''));
      }
    }
  } catch(e) {}

  // 4) Псевдоэлементы на случай A/B-тестов.
  try {
    for (const el of Array.from(document.querySelectorAll('[class],button,span,div'))) {
      const before = getComputedStyle(el, '::before').content || '';
      const after = getComputedStyle(el, '::after').content || '';
      if (/оригинал|original/i.test(before)) addOriginalCandidate(el, 'pseudo-before', before.replace(/^['"]|['"]$/g,''));
      if (/оригинал|original/i.test(after)) addOriginalCandidate(el, 'pseudo-after', after.replace(/^['"]|['"]$/g,''));
    }
  } catch(e) {}

  // 5) Резерв: структура строк body.innerText. В реальном DOM строка "Оригинал" была отдельной строкой №83,
  // рядом до/после идут бренд и название товара. Это менее точный, но полезный fallback.
  let bodyLineOriginal = false;
  let bodyOriginalContext = [];
  try {
    const lines = (document.body ? (document.body.innerText || '') : '').split(/\n+/).map(norm).filter(Boolean);
    for (let i=0; i<lines.length; i++) {
      if (/^(оригинал|original)$/i.test(low(lines[i]))) {
        const ctx = lines.slice(Math.max(0, i-4), Math.min(lines.length, i+5));
        const ctxText = low(ctx.join(' '));
        const productCtx = /adidas|nike|puma|reebok|fila|asics|skechers|носки|кеды|кроссов|футболк|брюк|куртк|товар|артикул|оценк/.test(ctxText);
        if (i < 180 && productCtx) {
          bodyLineOriginal = true;
          bodyOriginalContext = ctx;
          break;
        }
      }
    }
  } catch(e) {}

  // 6) Жёсткий confirmed-DOM fallback без геометрических отсечек.
  // Нужен потому, что в Playwright/WB координаты и layout могут отличаться от обычного браузера.
  // Приоритет: конкретный originalMark / productHeaderBadges / aria-label "метка Оригинал".
  let forcedOriginal = false;
  const forcedOriginalHits = [];
  const addForcedOriginal = (el, source) => {
    try {
      if (!el || !visible(el)) return;
      const r = rectObj(el);
      const cls = pathClass(el);
      const own = norm(ownText(el));
      const all = norm(el.innerText || el.textContent || '');
      const aria = norm(el.getAttribute && (el.getAttribute('aria-label') || ''));
      const title = norm(el.getAttribute && (el.getAttribute('title') || ''));
      const combined = low([own, all, aria, title, cls].join(' '));
      const exactText = /(^|\s)(оригинал|original)($|\s)/i.test(low([own, all].join(' ')));
      const directWbBadge = /(originalmark|productheaderbadges)/i.test(cls) && /оригинал|original/i.test(combined);
      const ariaBadge = /метк[аи]?\s+оригинал|что\s+означает\s+метк[аи]?\s+оригинал/i.test(combined);
      const inProductCard = /(product-page|productpage|productheader|productheaderbadges|mainwrap|cardtype|appreactroot|reactcontainers|body-layout|main__container)/i.test(cls);
      const saneSize = r.width >= 5 && r.height >= 5 && r.width <= 900 && r.height <= 220;
      const sanePosition = r.bottom > 0 && r.right > 0 && r.top < Math.max(innerHeight, 900);
      if ((directWbBadge || ariaBadge || (exactText && inProductCard)) && saneSize && sanePosition) {
        forcedOriginal = true;
        forcedOriginalHits.push({source, text:(all || own || aria || title).slice(0,140), aria:aria.slice(0,140), x:Math.round(r.left), y:Math.round(r.top), w:Math.round(r.width), h:Math.round(r.height), tag:el.tagName, cls:cls.slice(0,240)});
      }
    } catch(e) {}
  };
  try {
    const forcedSelectors = [
      'button[class*="originalMark"]',
      '[class*="originalMark"]',
      'button[class*="original" i]',
      '[class*="original" i]',
      'button[aria-label*="Оригинал"]',
      'button[aria-label*="оригинал"]',
      '[aria-label*="Оригинал"]',
      '[aria-label*="оригинал"]',
      '[aria-label*="метка Оригинал"]',
      '[aria-label*="метка оригинал"]',
      '[title*="Оригинал"]',
      '[title*="оригинал"]',
      '[class*="productHeaderBadges"]',
      '[class*="productHeaderBadges"] button',
      '[class*="productHeader"] button'
    ];
    for (const sel of forcedSelectors) {
      for (const el of Array.from(document.querySelectorAll(sel)).slice(0,30)) addForcedOriginal(el, 'forced-selector:' + sel);
    }
  } catch(e) {}
  try {
    // Точный текст "Оригинал" внутри товарной карточки. Не используем весь wrapperRoot, чтобы не ловить шапку/футер сайта.
    const roots = [];
    if (nmId) roots.push(...Array.from(document.querySelectorAll(`[data-testid="${nmId}"]`)));
    roots.push(...Array.from(document.querySelectorAll('[class*="product-page"], [class*="productPage"], main#body-layout, main')));
    for (const root of roots.slice(0,8)) {
      if (!root || !visible(root)) continue;
      const rt = low(root.innerText || root.textContent || '');
      const lines = (root.innerText || '').split(/\n+/).map(norm).filter(Boolean);
      const hasExactLine = lines.slice(0,400).some(l => /^(оригинал|original)$/i.test(low(l)));
      const hasCompactHeader = /(?:^|\s)(adidas|nike|puma|reebok|fila|asics|skechers)\s+(оригинал|original)\s+/i.test(rt);
      const hasSeoSpecs = /(?:^|\s)(оригинал|original)\s*[\.·•|/-]?\s*(характеристики|описание|артикул|цвет|состав)(?:\s|$)/i.test(rt);
      const productLike = /(купить|в\s+корзину|цвет|размер|артикул|оценк|вопрос|характеристики|описание)/i.test(rt);
      if ((hasExactLine || hasCompactHeader || hasSeoSpecs) && productLike) {
        forcedOriginal = true;
        const rr = rectObj(root);
        forcedOriginalHits.push({source:'forced-product-root-text', text:lines.slice(0,80).join(' | ').slice(0,300), aria:'', x:Math.round(rr.left), y:Math.round(rr.top), w:Math.round(rr.width), h:Math.round(rr.height), tag:root.tagName, cls:pathClass(root).slice(0,240)});
        break;
      }
    }
  } catch(e) {}

  const uniqOriginal = [];
  const seenOriginal = new Set();
  for (const c of originalCandidates) {
    const key = `${c.tag}|${c.text}|${c.aria}|${c.x}|${c.y}|${c.w}|${c.h}`;
    if (seenOriginal.has(key)) continue;
    seenOriginal.add(key);
    uniqOriginal.push(c);
  }
  uniqOriginal.sort((a,b)=>b.score-a.score);
  // 7) Максимально практичный fallback: если body/main товара содержит отдельную строку «Оригинал»
  // или SEO-последовательность «Оригинал. Характеристики...», считаем бейдж найденным.
  // Это нужно для карточек вроде 417714674, где кнопка видна пользователю, но CSS-модуль
  // или timing Playwright может отличаться от диагностированной карточки 198242850.
  let broadTextOriginal = false;
  let broadTextOriginalSample = '';
  try {
    const roots = [document.querySelector('main#body-layout'), document.querySelector('main'), document.body].filter(Boolean);
    for (const root of roots) {
      const raw = root.innerText || root.textContent || '';
      const ls = raw.split(/\n+/).map(norm).filter(Boolean);
      const compact = low(ls.slice(0,500).join(' '));
      const exact = ls.slice(0,500).some(l => /^(оригинал|original)$/i.test(low(l)));
      const seo = /(?:^|\s)(оригинал|original)\s*[\.·•|/-]?\s*(характеристики|описание|артикул|цвет|состав)(?:\s|$)/i.test(compact);
      const prod = /(купить|в\s+корзину|цвет|размер|артикул|оценк|вопрос|характеристики|описание)/i.test(compact);
      if ((exact || seo) && prod) {
        broadTextOriginal = true;
        broadTextOriginalSample = ls.slice(0,120).join(' | ').slice(0,500);
        break;
      }
    }
  } catch(e) {}

  const exactOriginal = uniqOriginal.length > 0 || bodyLineOriginal || forcedOriginal || broadTextOriginal;

  // -------------------------
  // Данные из JS-state: цена/бренд/продавец, но НЕ original.
  // -------------------------
  let stateSeller = '';
  let stateBrand = '';
  let statePrice = '';
  let stateSupplierId = '';
  function maybeProductObject(o) {
    if (!o || typeof o !== 'object') return false;
    if (nmId && (String(o.id || '') === nmId || String(o.nmId || '') === nmId || String(o.nm_id || '') === nmId)) return true;
    return false;
  }
  function readProductObj(o) {
    for (const [k,v] of Object.entries(o)) {
      const lk = String(k).toLowerCase();
      if (!stateSeller && ['supplier','suppliername','seller','sellername'].includes(lk) && (typeof v === 'string' || typeof v === 'number')) { const ss = cleanSeller(String(v)); if (ss) stateSeller = ss; }
      if (!stateBrand && ['brand','brandname'].includes(lk) && typeof v === 'string' && v.length <= 80) stateBrand = v;
      if (!statePrice && ['salepriceu','saleprice','priceu','price'].includes(lk) && (typeof v === 'number' || typeof v === 'string')) statePrice = String(v);
      if (!stateSupplierId && ['supplierid','supplierId','sellerid','sellerId'].map(x=>x.toLowerCase()).includes(lk) && (typeof v === 'number' || typeof v === 'string')) stateSupplierId = String(v);
      if (!stateSeller && v && typeof v === 'object' && /(supplier|seller|merchant)/.test(lk)) {
        for (const kk of ['name','supplierName','sellerName','title']) if (v[kk]) { const ss = cleanSeller(String(v[kk])); if (ss) { stateSeller = ss; break; } }
      }
    }
  }
  function walk(obj, depth) {
    if (!obj || depth > 7) return;
    if (Array.isArray(obj)) { for (const x of obj.slice(0,300)) walk(x, depth+1); return; }
    if (typeof obj !== 'object') return;
    if (maybeProductObject(obj)) readProductObj(obj);
    for (const v of Object.values(obj)) if (v && typeof v === 'object') walk(v, depth+1);
  }
  try {
    for (const key of Object.keys(window).filter(k => /state|store|redux|wb|vue|nuxt|initial/i.test(k)).slice(0,100)) {
      try { walk(window[key], 0); } catch(e) {}
    }
  } catch(e) {}

  let domPriceCandidate = '';
  let domSeller = '';
  try {
    // Цена из title надёжнее DOM: "купить за 494 ₽". В DOM рядом могут быть цена с WB-кошельком и старая цена.
    const mt = docTitle.match(/купить\s+за\s+([0-9\s.,]+)\s*₽/i);
    if (mt) domPriceCandidate = mt[1];
    if (!domPriceCandidate) {
      const idx = bodyLines.findIndex(l => /^выбрать размер$/i.test(l));
      const zone = idx >= 0 ? bodyLines.slice(idx + 1, idx + 8) : bodyLines.slice(0, 120);
      const prices = zone.map(l => (l.match(/^([0-9\s.,]+)\s*₽$/) || [])[1]).filter(Boolean);
      if (prices.length >= 2) domPriceCandidate = prices[1];
      else if (prices.length === 1) domPriceCandidate = prices[0];
    }

    // Seller fallback по видимой структуре правого блока. Для диагностической карточки:
    // "Послезавтра," -> "склад WB" -> "WILDBERRIES" -> "Оценки423".
    for (let i=0; i<bodyLines.length; i++) {
      const l = low(bodyLines[i]);
      if (l === 'продавец' || l === 'продавец:') {
        for (let j=i+1; j<Math.min(bodyLines.length, i+8); j++) { const cand = cleanSeller(bodyLines[j]); if (cand) { domSeller = cand; break; } }
      }
      if (domSeller) break;
      if (/^склад\s+wb$/i.test(l)) {
        for (let j=i+1; j<Math.min(bodyLines.length, i+5); j++) { const cand = cleanSeller(bodyLines[j]); if (cand) { domSeller = cand; break; } }
      }
      if (domSeller) break;
      if (/^(доставит|привезет|привезёт)$/i.test(l)) {
        for (let j=i+1; j<Math.min(bodyLines.length, i+6); j++) { const cand = cleanSeller(bodyLines[j]); if (cand) { domSeller = cand; break; } }
      }
      if (domSeller) break;
    }
  } catch(e) {}

  return {
    h1,
    pageTitle: docTitle,
    exactOriginal,
    originalCandidates: uniqOriginal.slice(0,12),
    forcedOriginal,
    forcedOriginalHits: forcedOriginalHits.slice(0,12),
    bodyLineOriginal,
    bodyOriginalContext,
    broadTextOriginal,
    broadTextOriginalSample,
    brandName: brandName || stateBrand,
    seller: seller || stateSeller || domSeller,
    supplierId: supplierId || stateSupplierId,
    priceCandidate: statePrice || domPriceCandidate,
    textSample: text.slice(0, 30000)
  };
}
"""
async def block_assets(route):
    try:
        req = route.request
        if req.resource_type in {"image", "media", "font"}:
            return await route.abort()
        url = req.url.lower()
        if any(x in url for x in ("google-analytics", "yandex/metrika", "doubleclick", "adsystem", "adservice")):
            return await route.abort()
        return await route.continue_()
    except Exception:
        try:
            await route.continue_()
        except Exception:
            pass

class PageCapture:
    def __init__(self, page):
        self.page = page
        self.urls: List[str] = []
        self._request_handler = None
        self._response_handler = None

    async def start(self):
        def on_request(req):
            try:
                self.urls.append(req.url)
            except Exception:
                pass
        def on_response(resp):
            try:
                self.urls.append(resp.url)
            except Exception:
                pass
        self._request_handler = on_request
        self._response_handler = on_response
        self.page.on("request", on_request)
        self.page.on("response", on_response)

    async def clear(self):
        self.urls.clear()
        try:
            await self.page.evaluate(CLEAR_CAPTURED_JS)
        except Exception:
            pass

    async def snapshot_urls(self) -> List[str]:
        out = list(self.urls)
        try:
            captured = await self.page.evaluate(GET_CAPTURED_JS)
            for item in captured or []:
                if isinstance(item, dict) and item.get("url"):
                    out.append(str(item["url"]))
        except Exception:
            pass
        try:
            out.append(self.page.url)
        except Exception:
            pass
        res: List[str] = []
        seen: Set[str] = set()
        for u in out:
            u = clean_url(u)
            if u and u not in seen:
                seen.add(u)
                res.append(u)
        return res

    async def stop(self):
        try:
            if self._request_handler:
                self.page.remove_listener("request", self._request_handler)
            if self._response_handler:
                self.page.remove_listener("response", self._response_handler)
        except Exception:
            pass

async def click_text_by_js(page, wanted: str, exact: bool = False, clickable_only: bool = False, trace: bool = False) -> bool:
    candidates = await page.evaluate(FIND_TEXT_TARGET_JS, {"wanted": wanted, "exact": exact, "clickableOnly": clickable_only})
    if trace:
        print(f"      candidates for '{wanted}': {candidates[:5]}")
    if not candidates:
        return False
    c = candidates[0]
    try:
        await page.mouse.move(c["x"], c["y"])
        await page.mouse.down()
        await page.mouse.up()
        if trace:
            print(f"    ✓ клик: {wanted}: {c.get('text')} / tag={c.get('tag')} / x={int(c['x'])},y={int(c['y'])}")
        return True
    except Exception:
        return False

async def wait_until_body_has(page, predicate, timeout_ms: int, poll_ms: int = 150) -> bool:
    start = time.monotonic()
    while (time.monotonic() - start) * 1000 < timeout_ms:
        try:
            state = await page.evaluate(PANEL_TEXT_JS)
            if predicate(state):
                return True
        except Exception:
            pass
        await asyncio.sleep(poll_ms / 1000)
    return False

async def detect_original_with_playwright_locators(page, nm_id: int = 0, trace: bool = False) -> Tuple[bool, Dict[str, Any]]:
    """Независимая проверка плашки Оригинал средствами Playwright.
    Это запасной слой на случай, если PAGE_DATA_JS не сработал из-за геометрии/DOM-фильтров.
    """
    debug: Dict[str, Any] = {"locator_hits": [], "root_text_hits": []}

    async def safe_count(locator) -> int:
        try:
            return await locator.count()
        except Exception:
            return 0

    async def inspect_locator(locator, source: str, max_items: int = 8) -> bool:
        cnt = await safe_count(locator)
        debug.setdefault("counts", {})[source] = cnt
        for i in range(min(cnt, max_items)):
            item = locator.nth(i)
            try:
                visible = await item.is_visible(timeout=250)
            except Exception:
                visible = False
            text = ""
            aria = ""
            cls = ""
            title = ""
            box = None
            try:
                text = (await item.inner_text(timeout=250) or "").strip()
            except Exception:
                pass
            try:
                aria = (await item.get_attribute("aria-label", timeout=250) or "").strip()
            except Exception:
                pass
            try:
                cls = (await item.get_attribute("class", timeout=250) or "").strip()
            except Exception:
                pass
            try:
                title = (await item.get_attribute("title", timeout=250) or "").strip()
            except Exception:
                pass
            try:
                box = await item.bounding_box(timeout=250)
            except Exception:
                box = None
            combined = " ".join([text, aria, cls, title]).lower().replace("ё", "е")
            hit = visible and (
                "originalmark" in combined
                or "productheaderbadges" in combined and "оригинал" in combined
                or "метка оригинал" in combined
                or re.search(r"(^|\s)(оригинал|original)($|\s)", combined) is not None
            )
            rec = {
                "source": source,
                "i": i,
                "visible": visible,
                "text": text[:140],
                "aria": aria[:140],
                "class": cls[:180],
                "title": title[:140],
                "box": box,
                "hit": bool(hit),
            }
            debug["locator_hits"].append(rec)
            if hit:
                return True
        return False

    selectors = [
        'button[class*="originalMark"]',
        '[class*="originalMark"]',
        'button[class*="original" i]',
        '[class*="original" i]',
        'button:has-text("Оригинал")',
        '[aria-label*="Оригинал"]',
        '[aria-label*="оригинал"]',
        '[aria-label*="метка Оригинал"]',
        '[aria-label*="метка оригинал"]',
        '[title*="Оригинал"]',
        '[title*="оригинал"]',
        '[class*="productHeaderBadges"]',
        '[class*="productHeaderBadges"] button',
        '[class*="productHeader"] button',
    ]
    for sel in selectors:
        try:
            if await inspect_locator(page.locator(sel), sel):
                return True, debug
        except Exception as e:
            debug.setdefault("errors", []).append(f"{sel}: {type(e).__name__}: {str(e)[:120]}")

    # Текстовый fallback: точная строка "Оригинал" внутри корня карточки товара.
    root_selectors = []
    if nm_id:
        root_selectors.append(f'[data-testid="{int(nm_id)}"]')
    root_selectors += ['[class*="product-page"]', '[class*="productPage"]', 'main#body-layout', 'main']
    for sel in root_selectors:
        try:
            loc = page.locator(sel).first
            if not await loc.is_visible(timeout=400):
                continue
            txt = await loc.inner_text(timeout=700)
            lines = [re.sub(r"\s+", " ", x).strip() for x in (txt or "").splitlines()]
            lines = [x for x in lines if x]
            low_txt = " ".join(lines[:260]).lower().replace("ё", "е")
            has_exact = any(re.fullmatch(r"оригинал|original", x.lower().replace("ё", "е")) for x in lines[:500])
            has_compact = re.search(r"(?:^|\s)(adidas|nike|puma|reebok|fila|asics|skechers)\s+(оригинал|original)\s+", low_txt) is not None
            has_seo_specs = re.search(r"(?:^|\s)(оригинал|original)\s*[\.·•|/-]?\s*(характеристики|описание|артикул|цвет|состав)(?:\s|$)", low_txt) is not None
            product_like = re.search(r"купить|в\s+корзину|цвет|размер|артикул|оценк|вопрос|характеристики|описание", low_txt) is not None
            rec = {"source": sel, "has_exact": bool(has_exact), "has_compact": bool(has_compact), "has_seo_specs": bool(has_seo_specs), "product_like": bool(product_like), "sample": " | ".join(lines[:120])[:650]}
            debug["root_text_hits"].append(rec)
            if (has_exact or has_compact or has_seo_specs) and product_like:
                return True, debug
        except Exception as e:
            debug.setdefault("errors", []).append(f"root {sel}: {type(e).__name__}: {str(e)[:120]}")

    # Последний fallback: видимый текст Оригинал в верхней части страницы.
    try:
        loc = page.get_by_text(re.compile(r"^\s*Оригинал\s*$", re.I))
        cnt = await safe_count(loc)
        debug.setdefault("counts", {})["get_by_text_exact_original"] = cnt
        for i in range(min(cnt, 10)):
            item = loc.nth(i)
            try:
                visible = await item.is_visible(timeout=250)
                box = await item.bounding_box(timeout=250)
            except Exception:
                visible, box = False, None
            rec = {"source": "get_by_text_exact_original", "i": i, "visible": visible, "box": box}
            debug["locator_hits"].append(rec)
            if visible and box and box.get("y", 9999) < 650 and box.get("width", 0) <= 260 and box.get("height", 0) <= 80:
                return True, debug
    except Exception as e:
        debug.setdefault("errors", []).append(f"get_by_text: {type(e).__name__}: {str(e)[:120]}")

    return False, debug


async def get_card_meta_via_browser_fetch(page, nm_id: int, trace: bool = False) -> Dict[str, Any]:
    """Пробует получить карточный JSON WB из контекста уже открытой страницы.
    Это помогает, когда aiohttp снаружи блокируется, но браузерная сессия карточку видит.
    """
    if not nm_id:
        return {}
    try:
        raw = await page.evaluate(
            r"""
            async ({nm}) => {
              const urls = [
                `https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm=${nm}`,
                `https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm=${nm}`,
                `https://card.wb.ru/cards/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm=${nm}`
              ];
              function findProducts(obj, out = [], depth = 0) {
                if (!obj || depth > 7) return out;
                if (Array.isArray(obj)) {
                  for (const x of obj.slice(0, 300)) findProducts(x, out, depth + 1);
                  return out;
                }
                if (typeof obj !== 'object') return out;
                if (Array.isArray(obj.products)) {
                  for (const p of obj.products) if (p && typeof p === 'object' && (p.id || p.nmId || p.nm_id)) out.push(p);
                }
                for (const v of Object.values(obj)) if (v && typeof v === 'object') findProducts(v, out, depth + 1);
                return out;
              }
              const errors = [];
              for (const url of urls) {
                try {
                  const r = await fetch(url, {credentials: 'include', headers: {'Accept': 'application/json,text/plain,*/*'}});
                  const text = await r.text();
                  if (!r.ok) { errors.push(`${r.status} ${url}`); continue; }
                  const j = JSON.parse(text);
                  const products = findProducts(j);
                  const p = products.find(x => String(x.id || x.nmId || x.nm_id || '') === String(nm)) || products[0];
                  if (p) return {ok: true, url, product: p};
                } catch(e) { errors.push(String(e).slice(0,160)); }
              }
              return {ok: false, errors};
            }
            """,
            {"nm": int(nm_id)},
        )
        if not isinstance(raw, dict) or not raw.get("ok") or not isinstance(raw.get("product"), dict):
            if trace:
                print(f"    browser_fetch_meta nm={nm_id}: no product, raw={str(raw)[:300]}")
            return {}
        prod = raw.get("product") or {}
        meta = {
            "name": str(get_any(prod, "name", "productName") or "").strip(),
            "brand": str(get_any(prod, "brand", "brandName") or "").strip(),
            "subject": str(get_any(prod, "subject", "subjectName", "subjectId") or "").strip(),
            "seller": best_seller_from_product_json(prod),
            "supplier_id": str(get_any(prod, "supplierId", "supplierID", "sellerId", "sellerID") or "").strip(),
            "price": parse_price_rub(prod),
            "url": raw.get("url", ""),
        }
        if trace:
            print(f"    browser_fetch_meta nm={nm_id}: {meta}")
        return meta
    except Exception as e:
        if trace:
            print(f"    browser_fetch_meta nm={nm_id}: {type(e).__name__}: {str(e)[:180]}")
        return {}

async def get_page_product_data(page, wait_ms: int = 0, nm_id: int = 0, trace: bool = False) -> Tuple[str, str, str, str, str]:
    last: Dict[str, Any] = {}
    locator_debug: Dict[str, Any] = {}
    locator_original = False
    locator_checked_at: Set[int] = set()
    started = time.monotonic()

    async def maybe_locator_check(stage_ms: int) -> bool:
        nonlocal locator_original, locator_debug, last
        # Проверяем не на каждом poll, а на нескольких контрольных точках, чтобы не убить скорость.
        bucket = int(stage_ms // 1000)
        if bucket in locator_checked_at:
            return locator_original
        locator_checked_at.add(bucket)
        try:
            locator_original, locator_debug = await detect_original_with_playwright_locators(page, nm_id, trace)
            if locator_original:
                last["exactOriginal"] = True
                last["locatorOriginal"] = True
                last["locatorDebug"] = locator_debug
                return True
        except Exception as e:
            locator_debug = {"error": f"{type(e).__name__}: {str(e)[:160]}"}
        return False

    while True:
        elapsed_ms = int((time.monotonic() - started) * 1000)
        try:
            data = await page.evaluate(PAGE_DATA_JS, {"nmId": str(nm_id or "")})
            last = data or {}
            # Если нашли оригинал JS-слоем — выходим сразу.
            if last.get("exactOriginal") or (wait_ms <= 0):
                break
        except Exception as e:
            last.setdefault("pageDataError", f"{type(e).__name__}: {str(e)[:160]}")

        # Независимый Playwright locator слой: после первой секунды, затем примерно раз в 2 секунды и в конце.
        if elapsed_ms >= 900 and (elapsed_ms < 1400 or elapsed_ms % 2000 < 350):
            if await maybe_locator_check(elapsed_ms):
                break

        if elapsed_ms >= wait_ms:
            break
        await asyncio.sleep(0.25)

    if not (last.get("exactOriginal") or locator_original):
        # Финальная проверка перед возвратом — бывает, что плашка дорисовывается в конце ожидания.
        try:
            locator_original, locator_debug = await detect_original_with_playwright_locators(page, nm_id, trace)
            if locator_original:
                last["exactOriginal"] = True
                last["locatorOriginal"] = True
                last["locatorDebug"] = locator_debug
        except Exception as e:
            locator_debug = {"error": f"{type(e).__name__}: {str(e)[:160]}"}

    brand_name = last.get("brandName") or ""
    # Важно: берём title до клика по "Смотреть на сайте". После клика страница может уйти на реестр,
    # поэтому process_card уже не сможет восстановить название из WB-title.
    name = clean_product_name_py(last.get("h1") or "", brand=brand_name, nm_id=nm_id) or product_name_from_title(last.get("pageTitle") or "", nm_id, brand_name)
    original = "ДА" if (last.get("exactOriginal") or locator_original) else "НЕТ"
    seller = clean_seller_name_py(last.get("seller") or "")
    price_candidate = last.get("priceCandidate") or ""
    if trace:
        try:
            print(
                f"    page_data nm={nm_id}: original={original}, seller={seller!r}, brand={brand_name!r}, "
                f"forcedOriginal={last.get('forcedOriginal')}, forcedHits={last.get('forcedOriginalHits') or []}, "
                f"bodyLineOriginal={last.get('bodyLineOriginal')}, bodyOriginalContext={last.get('bodyOriginalContext')}, "
                f"candidates={last.get('originalCandidates') or []}, locatorOriginal={locator_original}, locatorDebug={locator_debug}"
            )
        except Exception:
            pass
    return name, original, brand_name, seller, price_candidate


async def verify_docs_button_visible_on_card(page, args, trace: bool = False) -> bool:
    """Проверяет именно UI-признак WB: есть ли на карточке блок/кнопка «Документы проверены».

    Это нужно, потому что static certificate.json может существовать для nm_id, но WB при этом
    не показывает покупателю блок документов на карточке. Для строгого отчёта такие случаи нельзя
    считать полноценным статусом «ССЫЛКА НА РЕЕСТР СОБРАНА».
    """
    try:
        state = await page.evaluate(PANEL_TEXT_JS)
        if state and state.get("hasDocs"):
            return True
    except Exception:
        pass

    # Иногда блок документов появляется только после раскрытия секции характеристик/описания.
    try:
        await click_text_by_js(page, "Характеристики и описание", exact=False, clickable_only=False, trace=False)
        await asyncio.sleep(max(0.15, min(1.0, float(getattr(args, "after_specs_click_ms", 700)) / 1000.0)))
    except Exception:
        pass

    timeout_ms = int(getattr(args, "api_docs_verify_ms", 2500) or 2500)
    ok = await wait_until_body_has(page, lambda st: st.get("hasDocs"), timeout_ms, poll_ms=150)
    if trace:
        try:
            state = await page.evaluate(PANEL_TEXT_JS)
            print(f"    docs_button_verify: ok={ok}, state={state}")
        except Exception:
            print(f"    docs_button_verify: ok={ok}")
    return bool(ok)

async def strict_get_registry_urls(page, context, card: BrandCard, args) -> Tuple[List[str], str, str, str, str, str, str]:
    """Returns (urls, detail, page_name, is_original, page_brand, seller_name, price_candidate)."""
    cap = PageCapture(page)
    await cap.start()
    try:
        await page.add_init_script(CAPTURE_INIT_SCRIPT)
    except Exception:
        pass
    try:
        await page.goto(card.product_url, wait_until="domcontentloaded", timeout=args.goto_timeout_ms)
    except Exception as e:
        await cap.stop()
        return [], f"goto_failed: {type(e).__name__}: {str(e)[:160]}", "", "НЕ ОПРЕДЕЛЕНО", "", "", ""

    await asyncio.sleep(args.after_goto_ms / 1000)
    page_name, is_original, page_brand, seller_name, price_candidate = await get_page_product_data(page, args.original_wait_ms, card.nm_id, args.trace)

    # Дополнительный источник метаданных: карточный JSON через fetch внутри уже открытой страницы.
    # Выполняем ДО кликов по документам, пока мы точно на карточке WB.
    need_meta = (
        not clean_product_name_py(page_name, brand=page_brand or card.brand or args.brand, nm_id=card.nm_id)
        or is_bad_brand_value_py(page_brand)
        or not clean_seller_name_py(seller_name)
        or not parse_price_from_raw(price_candidate)
    )
    if need_meta:
        api_meta = await get_card_meta_via_browser_fetch(page, card.nm_id, args.trace)
        if api_meta:
            page_name = clean_product_name_py(page_name, brand=page_brand or api_meta.get("brand") or card.brand or args.brand, nm_id=card.nm_id) or api_meta.get("name") or page_name
            if is_bad_brand_value_py(page_brand) and api_meta.get("brand"):
                page_brand = api_meta.get("brand") or page_brand
            if not clean_seller_name_py(seller_name) and api_meta.get("seller"):
                seller_name = api_meta.get("seller") or seller_name
            if not parse_price_from_raw(price_candidate) and api_meta.get("price"):
                price_candidate = str(api_meta.get("price"))
            if not card.product_name and api_meta.get("name"):
                card.product_name = api_meta.get("name")
            if is_bad_brand_value_py(card.brand) and api_meta.get("brand"):
                card.brand = api_meta.get("brand")
            if not clean_seller_name_py(card.seller_name) and api_meta.get("seller"):
                card.seller_name = api_meta.get("seller")
            if not card.price_rub and api_meta.get("price"):
                card.price_rub = float(api_meta.get("price") or 0)
            if not card.subject and api_meta.get("subject"):
                card.subject = api_meta.get("subject")
            if not card.supplier_id and api_meta.get("supplier_id"):
                card.supplier_id = api_meta.get("supplier_id")

    # Быстрый путь из HAR: WB сам запрашивает certificate.json вида
    # https://basket-13.wbbasket.ru/vol1982/part198242/198242850/info/certificate.json
    # В большинстве случаев этого достаточно, клики по модалке документов не нужны.
    if str(getattr(args, "registry_mode", "api_first")).lower() in {"api_first", "direct_first", "api", "direct", "direct_only"}:
        try:
            fast_urls, fast_detail = await fetch_registry_urls_from_certificate_json(card, args)
            if args.trace:
                print(f"    certificate_json nm={card.nm_id}: urls={fast_urls}, detail={fast_detail}")
            mode = str(getattr(args, "registry_mode", "api_first")).lower()
            if fast_urls:
                # Быстрый режим: certificate.json считается самостоятельным источником.
                # UI-кнопка «Документы проверены» может отсутствовать, хотя JSON уже есть.
                # Проверяем кнопку только если пользователь явно включил --require-docs-button true.
                require_docs_button = bool(getattr(args, "require_docs_button", False)) and mode not in {"api", "direct", "direct_only"}
                if require_docs_button:
                    docs_button_ok = await verify_docs_button_visible_on_card(page, args, trace=args.trace)
                    if not docs_button_ok:
                        await cap.stop()
                        return fast_urls, "CERT_JSON_BUT_NO_DOCS_BUTTON:" + fast_detail, page_name, is_original, page_brand, seller_name, price_candidate
                await cap.stop()
                return fast_urls, fast_detail, page_name, is_original, page_brand, seller_name, price_candidate
            # В direct_only отсутствие certificate.json = нет прямой ссылки.
            # В api_first это НЕ финальный вывод: WB может показывать документы через UI без static JSON,
            # поэтому продолжаем старый UI-flow.
            if mode in {"api", "direct", "direct_only"}:
                if fast_detail.startswith("certificate_json_no_docs") or fast_detail.startswith("certificate_json_empty"):
                    await cap.stop()
                    return [], "NO_DOCS_FAST:" + fast_detail, page_name, is_original, page_brand, seller_name, price_candidate
                await cap.stop()
                return [], fast_detail, page_name, is_original, page_brand, seller_name, price_candidate
            # Иначе идём в UI fallback.
        except Exception as e:
            if args.trace:
                print(f"    certificate_json nm={card.nm_id}: ERROR {type(e).__name__}: {str(e)[:160]}")
            if str(getattr(args, "registry_mode", "api_first")).lower() in {"api", "direct", "direct_only"}:
                await cap.stop()
                return [], f"certificate_json_exception:{type(e).__name__}: {str(e)[:160]}", page_name, is_original, page_brand, seller_name, price_candidate

    specs_clicked = await click_text_by_js(page, "Характеристики и описание", exact=False, clickable_only=False, trace=args.trace)
    if specs_clicked:
        await asyncio.sleep(args.after_specs_click_ms / 1000)
    await wait_until_body_has(page, lambda s: s.get("hasSpecs") or s.get("hasDocs"), args.card_ready_timeout_ms)

    docs_found = await wait_until_body_has(page, lambda s: s.get("hasDocs"), args.docs_timeout_ms)
    if not docs_found and args.no_docs_fallback_ms > 0:
        await click_text_by_js(page, "Характеристики и описание", exact=False, clickable_only=False, trace=args.trace)
        await asyncio.sleep(args.after_specs_click_ms / 1000)
        docs_found = await wait_until_body_has(page, lambda s: s.get("hasDocs"), args.no_docs_fallback_ms)
    if not docs_found:
        await cap.stop()
        return [], "NO_DOCS", page_name, is_original, page_brand, seller_name, price_candidate

    await cap.clear()
    docs_clicked = await click_text_by_js(page, "Документы проверены", exact=False, clickable_only=False, trace=args.trace)
    if not docs_clicked:
        await cap.stop()
        return [], "docs_badge_found_but_not_clicked", page_name, is_original, page_brand, seller_name, price_candidate

    await asyncio.sleep(args.after_docs_wait_ms / 1000)
    look_found = await wait_until_body_has(page, lambda s: s.get("hasLook"), args.look_button_timeout_ms)
    if not look_found:
        await cap.stop()
        return [], "docs_modal_opened_but_no_look_button", page_name, is_original, page_brand, seller_name, price_candidate

    before_pages = set(context.pages)
    await cap.clear()
    look_clicked = await click_text_by_js(page, "Смотреть на сайте", exact=False, clickable_only=True, trace=args.trace)
    if not look_clicked:
        await cap.stop()
        return [], "look_button_found_but_not_clicked", page_name, is_original, page_brand, seller_name, price_candidate

    found: List[str] = []
    started = time.monotonic()
    while (time.monotonic() - started) * 1000 < args.after_look_wait_ms:
        urls = await cap.snapshot_urls()
        found = allowed_urls(urls)
        if found:
            break
        # check new popup pages
        try:
            for p in list(context.pages):
                if p is not page and p not in before_pages:
                    found = allowed_urls([p.url])
                    if found:
                        break
        except Exception:
            pass
        if found:
            break
        await asyncio.sleep(0.12)

    try:
        for p in list(context.pages):
            if p is not page and p not in before_pages:
                await p.close()
    except Exception:
        pass
    await cap.stop()
    if found:
        return found, "strict_flow_ok", page_name, is_original, page_brand, seller_name, price_candidate
    return [], "look_clicked_but_registry_url_not_captured", page_name, is_original, page_brand, seller_name, price_candidate



async def safe_page_title(page) -> str:
    try:
        return await page.title()
    except Exception:
        return ""

def parse_price_from_raw(raw: Any) -> float:
    try:
        s = str(raw).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        m = re.search(r"\d+(?:\.\d+)?", s)
        if not m:
            return 0.0
        v = float(m.group(0))
        return round(v / 100 if v > 10000 else v, 2)
    except Exception:
        return 0.0

async def process_card(page, context, card: BrandCard, args, worker_name: str) -> BrandResult:
    async def _inner():
        urls, detail, page_name, is_original, page_brand, seller_name, price_candidate = await strict_get_registry_urls(page, context, card, args)
        original_extra_detail = ""
        if is_original != "ДА":
            try:
                html_original, html_original_detail = await fetch_original_from_public_html(card, args)
                if args.trace:
                    print(f"    original_public_html nm={card.nm_id}: hit={html_original}, detail={html_original_detail}")
                if html_original:
                    is_original = "ДА"
                    original_extra_detail = "; original_public_html=" + str(html_original_detail)[:220]
            except Exception as e:
                if args.trace:
                    print(f"    original_public_html nm={card.nm_id}: ERROR {type(e).__name__}: {str(e)[:160]}")
        # Метаданные берём в первую очередь из API/собранной карточки, затем из DOM до клика по реестру.
        # Нельзя брать title после клика: страница уже может быть не WB.
        brand_final = ""
        for b in (card.brand, page_brand, args.brand):
            if b and not is_bad_brand_value_py(b):
                brand_final = str(b).strip()
                break
        product_name = (
            clean_product_name_py(card.product_name, brand=brand_final or args.brand, nm_id=card.nm_id)
            or clean_product_name_py(page_name, brand=brand_final or args.brand, nm_id=card.nm_id)
        )
        seller_final = clean_seller_name_py(card.seller_name) or clean_seller_name_py(seller_name)
        price_final = card.price_rub
        if not price_final and price_candidate:
            price_final = parse_price_from_raw(price_candidate)
        if str(detail).startswith("CERT_JSON_BUT_NO_DOCS_BUTTON"):
            status = STATUS_JSON_NO_DOCS_BUTTON
        elif detail == "NO_DOCS" or str(detail).startswith("NO_DOCS_FAST"):
            status = STATUS_NO_DOCS
        elif urls:
            status = STATUS_LINK_COLLECTED
        else:
            status = STATUS_NO_REGISTRY_LINK
        hosts = " | ".join(hostname(u) for u in urls)
        ids = " | ".join(extract_record_id(u) for u in urls)
        registry_details: Dict[str, str] = {}
        if urls and bool(getattr(args, "registry_details", True)):
            try:
                registry_details = await fetch_registry_details_fast(urls, args, page=page)
                if args.trace and registry_details:
                    print(f"    registry_details nm={card.nm_id}: source={registry_details.get('registry_details_source','')}, doc={registry_details.get('registry_doc_number','')}, status={registry_details.get('registry_status','')}")
            except Exception as e:
                registry_details = {"registry_details_source": f"registry_details_exception:{type(e).__name__}:{str(e)[:160]}"}
        return BrandResult(
            brand_query=args.brand,
            nm_id=card.nm_id,
            product_name=product_name,
            brand=brand_final,
            subject=card.subject,
            price_rub=price_final,
            seller_name=seller_final,
            supplier_id=card.supplier_id,
            is_original=is_original,
            status=status,
            registry_urls=" | ".join(urls),
            registry_hosts=hosts,
            registry_record_ids=ids,
            registry_doc_type=registry_details.get("registry_doc_type", ""),
            registry_doc_number=registry_details.get("registry_doc_number", ""),
            registry_blank_number=registry_details.get("registry_blank_number", ""),
            registry_status=registry_details.get("registry_status", ""),
            registry_status_date=registry_details.get("registry_status_date", ""),
            registry_status_basis=registry_details.get("registry_status_basis", ""),
            registry_date_start=registry_details.get("registry_date_start", ""),
            registry_date_end=registry_details.get("registry_date_end", ""),
            registry_applicant=registry_details.get("registry_applicant", ""),
            registry_applicant_inn=registry_details.get("registry_applicant_inn", ""),
            registry_manufacturer=registry_details.get("registry_manufacturer", ""),
            registry_product_group=registry_details.get("registry_product_group", ""),
            registry_product_full=registry_details.get("registry_product_full", ""),
            registry_tnved=registry_details.get("registry_tnved", ""),
            registry_scheme=registry_details.get("registry_scheme", ""),
            registry_technical_regulation=registry_details.get("registry_technical_regulation", ""),
            registry_evidence=registry_details.get("registry_evidence", ""),
            registry_details_source=registry_details.get("registry_details_source", ""),
            product_url=card.product_url or product_url(card.nm_id),
            details=(detail + original_extra_detail),
            worker=worker_name,
            checked_at=now_iso(),
        )
    try:
        return await asyncio.wait_for(_inner(), timeout=args.card_hard_timeout_ms / 1000)
    except asyncio.TimeoutError:
        return BrandResult(args.brand, card.nm_id, card.product_name, card.brand, card.subject, card.price_rub, card.seller_name, card.supplier_id, "НЕ ОПРЕДЕЛЕНО", STATUS_TIMEOUT, product_url=card.product_url or product_url(card.nm_id), details=f"hard_timeout_{args.card_hard_timeout_ms}ms", worker=worker_name, checked_at=now_iso())
    except Exception as e:
        return BrandResult(args.brand, card.nm_id, card.product_name, card.brand, card.subject, card.price_rub, card.seller_name, card.supplier_id, "НЕ ОПРЕДЕЛЕНО", STATUS_ERROR, product_url=card.product_url or product_url(card.nm_id), details=f"{type(e).__name__}: {str(e)[:250]}", worker=worker_name, checked_at=now_iso())


# v24: полностью HTTP-обработка одной карточки (без браузера / Playwright).
# Используется когда certificate.json + curl_cffi покрывают всё нужное.
# Скорость в 10-30 раз выше браузерного process_card.
async def process_card_no_browser(card: BrandCard, args, worker_name: str = "http") -> BrandResult:
    async def _inner():
        # 1) Ссылка на реестр — через certificate.json (HTTP, очень быстро)
        try:
            urls, cert_detail = await fetch_registry_urls_from_certificate_json(card, args)
        except Exception as e:
            urls, cert_detail = [], f"cert_json_exception:{type(e).__name__}:{str(e)[:120]}"

        # 2) Проверка «Оригинал» — через публичный HTML карточки (HTTP)
        is_original = "НЕТ ОПРЕДЕЛЕНО"
        original_detail = ""
        try:
            ok, why = await fetch_original_from_public_html(card, args)
            is_original = "ДА" if ok else "НЕТ"
            original_detail = str(why)[:220]
        except Exception as e:
            original_detail = f"original_exception:{type(e).__name__}:{str(e)[:120]}"

        # 3) Статус
        if str(cert_detail).startswith("certificate_json_no_docs") or str(cert_detail).startswith("certificate_json_empty"):
            status = STATUS_NO_DOCS
        elif urls:
            status = STATUS_LINK_COLLECTED
        else:
            status = STATUS_NO_REGISTRY_LINK

        # 4) Детали реестра (FSA — через curl_cffi если установлен; SWIS — через aiohttp)
        registry_details: Dict[str, str] = {}
        if urls and bool(getattr(args, "registry_details", True)):
            try:
                registry_details = await fetch_registry_details_fast(urls, args, page=None)
            except Exception as e:
                registry_details = {"registry_details_source": f"registry_details_exception:{type(e).__name__}:{str(e)[:160]}"}

        brand_final = card.brand or args.brand or ""
        product_name = clean_product_name_py(card.product_name, brand=brand_final, nm_id=card.nm_id)
        seller_final = clean_seller_name_py(card.seller_name)
        hosts = " | ".join(hostname(u) for u in urls)
        ids = " | ".join(extract_record_id(u) for u in urls)
        detail_str = f"no_browser; cert={cert_detail}; original={original_detail}"

        return BrandResult(
            brand_query=args.brand,
            nm_id=card.nm_id,
            product_name=product_name or card.product_name,
            brand=brand_final,
            subject=card.subject,
            price_rub=card.price_rub,
            seller_name=seller_final,
            supplier_id=card.supplier_id,
            is_original=is_original,
            status=status,
            registry_urls=" | ".join(urls),
            registry_hosts=hosts,
            registry_record_ids=ids,
            registry_doc_type=registry_details.get("registry_doc_type", ""),
            registry_doc_number=registry_details.get("registry_doc_number", ""),
            registry_blank_number=registry_details.get("registry_blank_number", ""),
            registry_status=registry_details.get("registry_status", ""),
            registry_status_date=registry_details.get("registry_status_date", ""),
            registry_status_basis=registry_details.get("registry_status_basis", ""),
            registry_date_start=registry_details.get("registry_date_start", ""),
            registry_date_end=registry_details.get("registry_date_end", ""),
            registry_applicant=registry_details.get("registry_applicant", ""),
            registry_applicant_inn=registry_details.get("registry_applicant_inn", ""),
            registry_manufacturer=registry_details.get("registry_manufacturer", ""),
            registry_product_group=registry_details.get("registry_product_group", ""),
            registry_product_full=registry_details.get("registry_product_full", ""),
            registry_tnved=registry_details.get("registry_tnved", ""),
            registry_scheme=registry_details.get("registry_scheme", ""),
            registry_technical_regulation=registry_details.get("registry_technical_regulation", ""),
            registry_evidence=registry_details.get("registry_evidence", ""),
            registry_details_source=registry_details.get("registry_details_source", ""),
            product_url=card.product_url or product_url(card.nm_id),
            details=detail_str,
            worker=worker_name,
            checked_at=now_iso(),
        )
    try:
        return await asyncio.wait_for(_inner(), timeout=max(20.0, args.card_hard_timeout_ms / 1000))
    except asyncio.TimeoutError:
        return BrandResult(
            args.brand, card.nm_id, card.product_name, card.brand, card.subject,
            card.price_rub, card.seller_name, card.supplier_id, "НЕ ОПРЕДЕЛЕНО", STATUS_TIMEOUT,
            product_url=card.product_url or product_url(card.nm_id),
            details=f"no_browser_hard_timeout", worker=worker_name, checked_at=now_iso(),
        )
    except Exception as e:
        return BrandResult(
            args.brand, card.nm_id, card.product_name, card.brand, card.subject,
            card.price_rub, card.seller_name, card.supplier_id, "НЕ ОПРЕДЕЛЕНО", STATUS_ERROR,
            product_url=card.product_url or product_url(card.nm_id),
            details=f"no_browser_exception:{type(e).__name__}:{str(e)[:200]}", worker=worker_name, checked_at=now_iso(),
        )

# -----------------------------
# Store / Excel
# -----------------------------

class BrandResultStore:
    def __init__(self, csv_path: Path, xlsx_path: Path,
                 expiry_warning_days: int = 30, make_report_xlsx: bool = True):
        self.csv_path = csv_path
        self.xlsx_path = xlsx_path
        self.expiry_warning_days = max(0, int(expiry_warning_days or 0))
        self.make_report_xlsx = bool(make_report_xlsx)
        self.rows: List[BrandResult] = []
        self.lock = asyncio.Lock()

    def load_resume(self, valid_ids: Optional[Set[int]] = None) -> Set[int]:
        processed: Set[int] = set()
        if not self.csv_path.exists():
            return processed
        with self.csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            rdr = csv.DictReader(f)
            for row in rdr:
                nm = safe_int(row.get("nm_id"))
                if not nm:
                    continue
                if valid_ids is not None and nm not in valid_ids:
                    continue
                processed.add(nm)
                self.rows.append(BrandResult(
                    brand_query=row.get("brand_query", ""),
                    nm_id=nm,
                    product_name=row.get("product_name", ""),
                    brand=row.get("brand", ""),
                    subject=row.get("subject", ""),
                    price_rub=safe_float(row.get("price_rub")),
                    seller_name=row.get("seller_name", ""),
                    supplier_id=row.get("supplier_id", ""),
                    is_original=row.get("is_original", ""),
                    status=row.get("status", ""),
                    registry_urls=row.get("registry_urls", ""),
                    registry_hosts=row.get("registry_hosts", ""),
                    registry_record_ids=row.get("registry_record_ids", ""),
                    registry_doc_type=row.get("registry_doc_type", ""),
                    registry_doc_number=row.get("registry_doc_number", ""),
                    registry_blank_number=row.get("registry_blank_number", ""),
                    registry_status=row.get("registry_status", ""),
                    registry_status_date=row.get("registry_status_date", ""),
                    registry_status_basis=row.get("registry_status_basis", ""),
                    registry_date_start=row.get("registry_date_start", ""),
                    registry_date_end=row.get("registry_date_end", ""),
                    registry_applicant=row.get("registry_applicant", ""),
                    registry_applicant_inn=row.get("registry_applicant_inn", ""),
                    registry_manufacturer=row.get("registry_manufacturer", ""),
                    registry_product_group=row.get("registry_product_group", ""),
                    registry_product_full=row.get("registry_product_full", ""),
                    registry_tnved=row.get("registry_tnved", ""),
                    registry_scheme=row.get("registry_scheme", ""),
                    registry_technical_regulation=row.get("registry_technical_regulation", ""),
                    registry_evidence=row.get("registry_evidence", ""),
                    registry_details_source=row.get("registry_details_source", ""),
                    product_url=row.get("product_url", ""),
                    details=row.get("details", ""),
                    worker=row.get("worker", ""),
                    checked_at=row.get("checked_at", ""),
                ))
        return processed

    async def add(self, row: BrandResult) -> None:
        async with self.lock:
            self.rows.append(row)

    async def save(self) -> None:
        async with self.lock:
            rows = list(self.rows)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.csv_path.with_suffix(self.csv_path.suffix + ".tmp")
        fields = list(asdict(BrandResult("", 0, "", "", "", 0.0, "", "", "", "")).keys())
        with tmp.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            for r in rows:
                w.writerow(asdict(r))
        os.replace(tmp, self.csv_path)
        save_results_xlsx(
            rows,
            self.xlsx_path,
            expiry_warning_days=self.expiry_warning_days,
            make_report_xlsx=self.make_report_xlsx,
        )


# -----------------------------
# Отчётный слой (v25-reporting)
# -----------------------------

# Русские заголовки для листа 'Подробности' — технические имена колонок на листе 'results' не трогаются.
DETAILS_HEADERS_RU: Dict[str, str] = {
    "brand_query": "Запрос по бренду",
    "nm_id": "Артикул WB",
    "product_name": "Название товара",
    "brand": "Бренд",
    "subject": "Категория WB",
    "price_rub": "Цена, ₽",
    "seller_name": "Продавец",
    "supplier_id": "ID продавца",
    "is_original": "Плашка 'Оригинал'",
    "status": "Технический статус",
    "registry_urls": "Ссылки на реестр",
    "registry_hosts": "Реестры (хосты)",
    "registry_record_ids": "ID записей реестра",
    "registry_doc_type": "Тип документа",
    "registry_doc_number": "Номер документа",
    "registry_blank_number": "Номер бланка",
    "registry_status": "Статус документа",
    "registry_status_date": "Дата статуса",
    "registry_status_basis": "Основание статуса",
    "registry_date_start": "Действует с",
    "registry_date_end": "Действует до",
    "registry_applicant": "Заявитель",
    "registry_applicant_inn": "ИНН заявителя",
    "registry_manufacturer": "Изготовитель",
    "registry_product_group": "Группа товара (реестр)",
    "registry_product_full": "Название в реестре",
    "registry_tnved": "ТН ВЭД",
    "registry_scheme": "Схема оценки",
    "registry_technical_regulation": "Техрегламент",
    "registry_evidence": "Доказательные материалы",
    "registry_details_source": "Источник деталей реестра",
    "product_url": "Ссылка на товар",
    "details": "Примечания",
    "worker": "Worker",
    "checked_at": "Проверено",
    # вычисляемые в отчёте
    "expiry_days_left": "Дней до окончания",
    "expiry_risk": "Риск по сроку",
}

EXPIRY_RISK_OK = "Действует"
EXPIRY_RISK_SOON = "Скоро истекает"
EXPIRY_RISK_EXPIRED = "Истёк"
EXPIRY_RISK_UNKNOWN = "Срок не известен"

_DATE_FORMATS_PARSE = (
    "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y",
    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
)


def _parse_date_loose(s: Any) -> Optional[dt.date]:
    if s is None:
        return None
    text = str(s).strip()
    if not text:
        return None
    # бывает приходит в виде ISO с 'T' и частичными хвостами — обрежем
    for fmt in _DATE_FORMATS_PARSE:
        try:
            return dt.datetime.strptime(text[:len(fmt) + 10], fmt).date()
        except Exception:
            pass
    # ISO с таймзоной
    try:
        return dt.datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except Exception:
        return None


def _compute_expiry(row: BrandResult, warning_days: int,
                    today: Optional[dt.date] = None) -> Tuple[Optional[int], str]:
    """Возвращает (дней_до_окончания, риск). Дни = None, если дата не распознана."""
    today = today or dt.date.today()
    end_date = _parse_date_loose(getattr(row, "registry_date_end", ""))
    if end_date is None:
        return None, EXPIRY_RISK_UNKNOWN
    days_left = (end_date - today).days
    if days_left < 0:
        return days_left, EXPIRY_RISK_EXPIRED
    if warning_days > 0 and days_left <= warning_days:
        return days_left, EXPIRY_RISK_SOON
    return days_left, EXPIRY_RISK_OK


_XLSX_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xlsx_safe(value: Any, max_len: int = 32000):
    """Безопасное значение для ячейки xlsx (см. main_v39._xlsx_safe).

    openpyxl падает с IllegalCharacterError на управляющих символах из
    спарсенного текста и роняет весь отчёт. Числа оставляем числами, остальное
    чистим от control-символов, схлопываем переводы строк/табы и режем длину.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else ""
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    if not s:
        return ""
    s = _XLSX_ILLEGAL_RE.sub("", s)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    if "  " in s:
        s = re.sub(r"\s{2,}", " ", s)
    s = s.strip()
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s


def _build_summary_sheet(wb_obj, rows: List["BrandResult"], warning_days: int) -> None:
    """Лист 'Сводка' с шапкой и агрегатами."""
    ws = wb_obj.create_sheet("Сводка", 0)
    title_font = Font(bold=True, size=12)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    ws.append(["WB Brand Checker — расширенный отчёт"])
    ws.append(["Дата формирования", now_iso()])
    ws.append(["Версия движка", APP_VERSION])
    ws.append(["Порог 'Скоро истекает', дней", warning_days])
    ws.append(["Всего товаров", len(rows)])
    ws.append([])
    ws.append(["Распределение по техническому статусу"])
    ws.append(["Статус", "Количество", "Доля, %"])
    status_counts: Dict[str, int] = {}
    for r in rows:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
    total = max(1, len(rows))
    for status_name, count in sorted(status_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([status_name, count, round(count * 100.0 / total, 1)])
    ws.append([])
    ws.append(["Плашка 'Оригинал'"])
    ws.append(["Значение", "Количество", "Доля, %"])
    original_counts: Dict[str, int] = {}
    for r in rows:
        key = r.is_original or "—"
        original_counts[key] = original_counts.get(key, 0) + 1
    for key, count in sorted(original_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([key, count, round(count * 100.0 / total, 1)])
    ws.append([])
    ws.append(["Риски по сроку действия документа"])
    ws.append(["Категория", "Количество", "Доля, %"])
    risk_counts: Dict[str, int] = {EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
                                   EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0}
    for r in rows:
        _, risk = _compute_expiry(r, warning_days)
        risk_counts[risk] = risk_counts.get(risk, 0) + 1
    for key in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
        ws.append([key, risk_counts[key], round(risk_counts[key] * 100.0 / total, 1)])
    # форматирование
    ws["A1"].font = Font(bold=True, size=14)
    for row_idx in (2, 3, 4, 5):
        ws[f"A{row_idx}"].font = Font(bold=True)
    for row_idx in (7, 12, 17):
        try:
            ws[f"A{row_idx}"].font = title_font
        except Exception:
            pass
    for cell_addr in ("A8", "B8", "C8", "A13", "B13", "C13", "A18", "B18", "C18"):
        try:
            ws[cell_addr].fill = hdr_fill
            ws[cell_addr].font = hdr_font
            ws[cell_addr].alignment = Alignment(horizontal="center")
        except Exception:
            pass
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A7"


# v25-reporting: единый порядок ядра 'Подробностей'. Специфичные поля бренд-движка пойдут после.
CORE_DETAILS_ORDER: Tuple[str, ...] = (
    "brand_query", "nm_id", "product_name", "brand", "subject", "product_url",
    "status", "price_rub", "seller_name", "supplier_id", "is_original",
    "registry_urls", "registry_doc_type", "registry_status",
    "registry_doc_number", "registry_date_start", "registry_date_end",
)


def _details_field_order(base_fields: List[str]) -> List[str]:
    seen: Set[str] = set()
    ordered: List[str] = []
    for f in CORE_DETAILS_ORDER:
        if f in base_fields and f not in seen:
            ordered.append(f); seen.add(f)
    for f in base_fields:
        if f not in seen:
            ordered.append(f); seen.add(f)
    return ordered


def _build_details_sheet(wb_obj, rows: List["BrandResult"], warning_days: int) -> None:
    """Лист 'Подробности' — все товары с русскими заголовками и флагами по сроку."""
    ws = wb_obj.create_sheet("Подробности")
    base_fields = list(asdict(BrandResult("", 0, "", "", "", 0.0, "", "", "", "")).keys())
    ordered = _details_field_order(base_fields)
    fields = ordered + ["expiry_days_left", "expiry_risk"]
    ws.append([DETAILS_HEADERS_RU.get(f, f) for f in fields])
    for r in rows:
        d = asdict(r)
        days_left, risk = _compute_expiry(r, warning_days)
        d["expiry_days_left"] = "" if days_left is None else days_left
        d["expiry_risk"] = risk
        ws.append([_xlsx_safe(d.get(f, "")) for f in fields])
    # шапка
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions
    # разумные ширины колонок
    for idx, f in enumerate(fields, start=1):
        if f in {"product_name", "registry_urls", "registry_product_full", "registry_evidence",
                 "registry_technical_regulation", "registry_details_source", "details", "product_url"}:
            width = 50
        elif f in {"nm_id", "price_rub", "expiry_days_left"}:
            width = 14
        elif f.startswith("registry_"):
            width = 26
        else:
            width = 20
        ws.column_dimensions[get_column_letter(idx)].width = width
    # цвета по expiry_risk
    risk_col_idx = fields.index("expiry_risk") + 1
    fills = {
        EXPIRY_RISK_OK: PatternFill("solid", fgColor="D9EAD3"),
        EXPIRY_RISK_SOON: PatternFill("solid", fgColor="FFF2CC"),
        EXPIRY_RISK_EXPIRED: PatternFill("solid", fgColor="F4CCCC"),
        EXPIRY_RISK_UNKNOWN: PatternFill("solid", fgColor="EFEFEF"),
    }
    for row in ws.iter_rows(min_row=2):
        risk_val = str(row[risk_col_idx - 1].value or "")
        fill = fills.get(risk_val)
        if fill is not None:
            row[risk_col_idx - 1].fill = fill
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)


def _write_run_log(rows: List["BrandResult"], xlsx_path: Path,
                   warning_days: int, started_at: float,
                   log_path: Optional[Path] = None) -> Optional[Path]:
    """v25-reporting: текстовый лог прогона. Разборы 'wtf в проде'."""
    try:
        if log_path is None:
            log_path = xlsx_path.with_suffix("") .with_name(xlsx_path.stem + "_run.log")
        finished = time.time()
        status_counts: Dict[str, int] = {}
        risk_counts: Dict[str, int] = {EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
                                       EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0}
        for r in rows:
            status_counts[r.status] = status_counts.get(r.status, 0) + 1
            _, risk = _compute_expiry(r, warning_days)
            risk_counts[risk] = risk_counts.get(risk, 0) + 1
        lines: List[str] = []
        lines.append("WB Brand Checker — лог прогона")
        lines.append(f"Версия движка: {APP_VERSION}")
        lines.append(f"Старт:     {dt.datetime.fromtimestamp(started_at).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Финиш:    {dt.datetime.fromtimestamp(finished).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Длительность: {int(finished - started_at)} сек")
        lines.append(f"Итоговый файл: {xlsx_path}")
        lines.append(f"Порог 'Скоро истекает', дней: {warning_days}")
        lines.append(f"Всего товаров: {len(rows)}")
        lines.append("")
        lines.append("Распределение по техническому статусу:")
        for k, v in sorted(status_counts.items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {k:<40} {v}")
        lines.append("")
        lines.append("Распределение по риску по сроку:")
        for k in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
            lines.append(f"  {k:<20} {risk_counts.get(k, 0)}")
        log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return log_path
    except Exception as e:
        print(f"[run-log] Не удалось записать лог прогона: {e}")
        return None


def save_results_xlsx(rows: List[BrandResult], path: Path,
                      expiry_warning_days: int = 30,
                      make_report_xlsx: bool = True) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    headers = list(asdict(BrandResult("", 0, "", "", "", 0.0, "", "", "", "")).keys())
    ws.append(headers)
    for r in rows:
        d = asdict(r)
        ws.append([_xlsx_safe(d.get(h, "")) for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    widths = {
        "A": 18, "B": 14, "C": 45, "D": 20, "E": 20, "F": 12, "G": 30, "H": 14,
        "I": 12, "J": 24, "K": 55, "L": 25, "M": 25, "N": 45, "O": 35, "P": 12, "Q": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        if col not in widths:
            if header in {"registry_product_full", "registry_evidence", "registry_technical_regulation", "registry_details_source"}:
                ws.column_dimensions[col].width = 55
            elif header.startswith("registry_"):
                ws.column_dimensions[col].width = 28
            else:
                ws.column_dimensions[col].width = 18
    for row in ws.iter_rows(min_row=2):
        status = str(row[9].value or "")
        fill = None
        if status == STATUS_LINK_COLLECTED:
            fill = PatternFill("solid", fgColor="D9EAD3")
        elif status == STATUS_NO_DOCS:
            fill = PatternFill("solid", fgColor="FCE5CD")
        elif status == STATUS_JSON_NO_DOCS_BUTTON:
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif status in {STATUS_TIMEOUT, STATUS_ERROR}:
            fill = PatternFill("solid", fgColor="F4CCCC")
        if fill:
            for c in row:
                c.fill = fill
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    # summary
    ws2 = wb.create_sheet("summary")
    counts: Dict[str, int] = {}
    originals: Dict[str, int] = {}
    for r in rows:
        counts[r.status] = counts.get(r.status, 0) + 1
        originals[r.is_original] = originals.get(r.is_original, 0) + 1
    ws2.append(["metric", "value"])
    ws2.append(["total", len(rows)])
    ws2.append(["with_registry_links", counts.get(STATUS_LINK_COLLECTED, 0)])
    ws2.append(["json_link_no_docs_button", counts.get(STATUS_JSON_NO_DOCS_BUTTON, 0)])
    ws2.append(["no_docs", counts.get(STATUS_NO_DOCS, 0)])
    ws2.append(["no_registry_link", counts.get(STATUS_NO_REGISTRY_LINK, 0)])
    ws2.append(["timeouts_errors", counts.get(STATUS_TIMEOUT, 0) + counts.get(STATUS_ERROR, 0)])
    for k, v in originals.items():
        ws2.append([f"original_{k}", v])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16
    # Отчётный слой (v25-reporting): русские листы 'Сводка' и 'Подробности' в том же файле.
    if make_report_xlsx:
        try:
            _build_summary_sheet(wb, rows, expiry_warning_days)
            _build_details_sheet(wb, rows, expiry_warning_days)
        except Exception as e:
            # Отчёт — бонусный слой, из-за его ошибки основной results-лист терять нельзя.
            print(f"[report] Не удалось сформировать листы 'Сводка'/'Подробности': {e}")
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    os.replace(tmp, path)

# -----------------------------
# Browser pool / workers
# -----------------------------

async def launch_browser(p, args):
    return await p.chromium.launch(
        headless=args.headless,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-background-timer-throttling",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-extensions",
        ],
        timeout=60000,
    )

class BrowserPool:
    def __init__(self, p, args):
        self.p = p
        self.args = args
        self.count = max(1, int(args.browser_count or 1))
        self.browsers = [None] * self.count
        self.locks = [asyncio.Lock() for _ in range(self.count)]

    async def get_browser(self, idx: int):
        idx = idx % self.count
        async with self.locks[idx]:
            br = self.browsers[idx]
            try:
                if br is not None and br.is_connected():
                    return br
            except Exception:
                pass
            try:
                if br is not None:
                    await br.close()
            except Exception:
                pass
            br = await launch_browser(self.p, self.args)
            self.browsers[idx] = br
            return br

    async def new_context(self, idx: int):
        br = await self.get_browser(idx)
        ctx = await br.new_context(
            user_agent=self.args.user_agent,
            viewport={"width": self.args.viewport_width, "height": self.args.viewport_height},
            locale="ru-RU",
            ignore_https_errors=True,
        )
        if self.args.block_assets:
            await ctx.route("**/*", block_assets)
        try:
            await ctx.add_init_script(CAPTURE_INIT_SCRIPT)
        except Exception:
            pass
        return ctx

    async def close_all(self):
        for br in list(self.browsers):
            try:
                if br:
                    await br.close()
            except Exception:
                pass

async def make_page(pool: BrowserPool, pool_idx: int, args):
    ctx = await pool.new_context(pool_idx)
    page = await ctx.new_page()
    async def _safe_dialog_close(dialog):
        try:
            await dialog.dismiss()
        except Exception:
            pass
    try:
        page.on("dialog", lambda d: asyncio.create_task(_safe_dialog_close(d)))
    except Exception:
        pass
    page.set_default_timeout(args.default_timeout_ms)
    try:
        await page.add_init_script(CAPTURE_INIT_SCRIPT)
    except Exception:
        pass
    return ctx, page

async def close_context_quiet(context):
    try:
        if context:
            await context.close()
    except Exception:
        pass

# v24: воркер без браузера. Используется когда --no-browser=true.
# Не открывает Playwright, всё делает через HTTP. В 10-30 раз быстрее обычного.
async def brand_worker_no_browser(worker_id: int, q: asyncio.Queue, store: BrandResultStore, args, progress: Dict[str, Any]):
    worker_name = f"h{worker_id}"
    while True:
        try:
            card = await q.get()
        except asyncio.CancelledError:
            break
        progress["active"][worker_name] = {"nm_id": card.nm_id, "started": time.time(), "attempt": 1}
        try:
            row = await process_card_no_browser(card, args, worker_name)
            await store.add(row)
            progress["done"] += 1
            if row.is_original == "ДА":
                progress["original_yes"] += 1
            if row.status == STATUS_LINK_COLLECTED:
                progress["links"] += 1
            elif row.status == STATUS_NO_DOCS:
                progress["no_docs"] += 1
            elif row.status == STATUS_NO_REGISTRY_LINK:
                progress["no_link"] += 1
            elif row.status in {STATUS_TIMEOUT, STATUS_ERROR}:
                progress["tech"] += 1
            if args.verbose_each or progress["done"] <= 10 or (row.status == STATUS_LINK_COLLECTED and args.print_links):
                print(f"[{worker_name}] #{progress['done']} nm_id={card.nm_id}: {row.status}, original={row.is_original}, price={row.price_rub}, doc_number={row.registry_doc_number[:30]}")
            if progress["done"] % max(1, args.autosave_every) == 0:
                await store.save()
        finally:
            progress["active"].pop(worker_name, None)
            q.task_done()


async def brand_worker(worker_id: int, pool: BrowserPool, q: asyncio.Queue, store: BrandResultStore, args, progress: Dict[str, Any]):
    pool_idx = (worker_id - 1) % max(1, int(args.browser_count or 1))
    worker_name = f"w{worker_id}"
    context = None
    page = None
    try:
        context, page = await make_page(pool, pool_idx, args)
        while True:
            try:
                card = await q.get()
            except asyncio.CancelledError:
                break
            progress["active"][worker_name] = {"nm_id": card.nm_id, "started": time.time(), "attempt": 1}
            try:
                final: Optional[BrandResult] = None
                for attempt in range(1, args.max_card_retries + 2):
                    progress["active"][worker_name] = {"nm_id": card.nm_id, "started": time.time(), "attempt": attempt}
                    row = await process_card(page, context, card, args, worker_name)
                    if row.status in {STATUS_TIMEOUT, STATUS_ERROR} and attempt <= args.max_card_retries:
                        await close_context_quiet(context)
                        context, page = await make_page(pool, pool_idx, args)
                        await asyncio.sleep(0.2 * attempt)
                        continue
                    final = row
                    break
                row = final or row
                await store.add(row)
                progress["done"] += 1
                if row.is_original == "ДА":
                    progress["original_yes"] += 1
                if row.status == STATUS_LINK_COLLECTED:
                    progress["links"] += 1
                elif row.status == STATUS_NO_DOCS:
                    progress["no_docs"] += 1
                elif row.status == STATUS_JSON_NO_DOCS_BUTTON:
                    progress["json_no_button"] += 1
                elif row.status == STATUS_NO_REGISTRY_LINK:
                    progress["no_link"] += 1
                elif row.status in {STATUS_TIMEOUT, STATUS_ERROR}:
                    progress["tech"] += 1
                if args.verbose_each or progress["done"] <= 10 or (row.status == STATUS_LINK_COLLECTED and args.print_links):
                    print(f"[{worker_name}] #{progress['done']} nm_id={card.nm_id}: {row.status}, original={row.is_original}, price={row.price_rub} — {row.details[:120]}")
                if progress["done"] % max(1, args.context_refresh_every) == 0:
                    await close_context_quiet(context)
                    context, page = await make_page(pool, pool_idx, args)
                if progress["done"] % max(1, args.autosave_every) == 0:
                    await store.save()
            finally:
                progress["active"].pop(worker_name, None)
                q.task_done()
    finally:
        await close_context_quiet(context)

async def progress_loop(q: asyncio.Queue, store: BrandResultStore, args, progress: Dict[str, Any], total: int):
    try:
        while True:
            await asyncio.sleep(args.progress_interval_sec)
            elapsed = max(1, time.time() - progress["start"])
            speed = progress["done"] / elapsed * 60
            link_speed = progress["links"] / elapsed * 60
            active = []
            now = time.time()
            for w, info in progress["active"].items():
                active.append(f"{w}:{info.get('nm_id')}:{int(now-info.get('started', now))}s/a{info.get('attempt',1)}")
            print(
                f"Прогресс: обработано={progress['done']}/{total}, скорость≈{speed:.1f}/мин, ссылки≈{link_speed:.1f}/мин, "
                f"найдено ссылок={progress['links']}, оригинал={progress['original_yes']}, нет документов={progress['no_docs']}, "
                f"нет ссылки={progress['no_link']}, очередь={q.qsize()}, тех={progress['tech']}, активные=[{'; '.join(active[:8])}]"
            )
            emit_progress("links", progress['done'], total)
            await store.save()
    except asyncio.CancelledError:
        return

async def run_check(args):
    # v25-reporting: время старта для финального лога прогона.
    _run_started_at = time.time()
    # v24: режим --no-browser. Если включён, не запускаем Playwright вообще —
    # всё работает через HTTP (certificate.json + curl_cffi для FSA + aiohttp для SWIS).
    # Это в 10-30 раз быстрее браузерного режима.
    no_browser = bool(getattr(args, "no_browser", False))

    if not no_browser and async_playwright is None:
        raise RuntimeError(
            "Playwright не установлен. Либо: python -m pip install playwright && python -m playwright install chromium\n"
            "Либо запусти с --no-browser true (если установлен curl_cffi — для FSA это даже быстрее)."
        )
    # Точечная проверка одной карточки без CSV — удобно для спорных ссылок.
    if args.single_url:
        nm = nm_id_from_url_or_text(args.single_url)
        if not nm:
            raise RuntimeError(f"Не смог извлечь nm_id из --single-url: {args.single_url}")
        card = BrandCard(nm_id=nm, source_query=args.brand or "single", brand=args.brand or "", product_url=product_url(nm))
        # Для одиночной проверки раньше карточка шла пустой: без product_name / seller / price.
        # Добираем эти поля через тот же карточный API WB, который используется в массовом сборе.
        card = await enrich_single_card_from_api(card, args)
        cards = [card]
        args.resume = False
        print(f"Точечная проверка одной карточки: nm_id={nm}, name={card.product_name!r}, brand={card.brand!r}, seller={card.seller_name!r}, price={card.price_rub}")
    # Важно: если НЕ указан --check-only, программа собирает карточки заново.
    # Иначе при существующем старом adidas_cards.csv на 100 строк она снова обработает только 100.
    elif args.check_only:
        cards = load_cards_csv(Path(args.cards_csv), args.limit)
        print(f"Загружено карточек из {args.cards_csv}: {len(cards)}")
    elif args.reuse_cards and args.cards_csv and Path(args.cards_csv).exists():
        cards = load_cards_csv(Path(args.cards_csv), args.limit)
        if len(cards) >= args.limit:
            print(f"Использую существующий файл карточек {args.cards_csv}: {len(cards)}")
        else:
            print(f"Существующий файл карточек содержит только {len(cards)}/{args.limit}; собираю заново")
            cards = await collect_brand_cards(args)
    else:
        cards = await collect_brand_cards(args)
    if not cards:
        print("Карточки не найдены. Проверьте бренд или --brand-match contains/any.")
        return
    if args.reset_output:
        for p in (Path(args.result_csv), Path(args.output)):
            try:
                if p.exists():
                    p.unlink()
            except Exception:
                pass
    store = BrandResultStore(
        Path(args.result_csv),
        Path(args.output),
        expiry_warning_days=getattr(args, "expiry_warning_days", 30),
        make_report_xlsx=getattr(args, "make_report_xlsx", True),
    )
    valid_ids = {c.nm_id for c in cards}
    processed = store.load_resume(valid_ids=valid_ids) if args.resume else set()
    remaining = [c for c in cards if c.nm_id not in processed]
    print(f"К обработке: всего={len(cards)}, уже есть={len(processed)}, осталось={len(remaining)}")
    if not remaining:
        await store.save()
        print(f"Готово. Все карточки уже обработаны: {Path(args.output).resolve()}")
        return

    q: asyncio.Queue = asyncio.Queue()
    for c in remaining:
        await q.put(c)
    progress = {
        "done": len(processed),
        "links": sum(1 for r in store.rows if r.status == STATUS_LINK_COLLECTED),
        "no_docs": sum(1 for r in store.rows if r.status == STATUS_NO_DOCS),
        "json_no_button": sum(1 for r in store.rows if r.status == STATUS_JSON_NO_DOCS_BUTTON),
        "no_link": sum(1 for r in store.rows if r.status == STATUS_NO_REGISTRY_LINK),
        "tech": sum(1 for r in store.rows if r.status in {STATUS_TIMEOUT, STATUS_ERROR}),
        "original_yes": sum(1 for r in store.rows if r.is_original == "ДА"),
        "active": {},
        "start": time.time(),
    }

    if no_browser:
        # v24: HTTP-only режим. Параллельность можно ставить выше — это HTTP, не браузер.
        n_workers = max(1, int(getattr(args, "http_workers", 0) or args.workers * 4))
        print(f"🌐 Режим --no-browser: запускаю {n_workers} HTTP-воркеров (без Playwright)")
        tasks = [asyncio.create_task(brand_worker_no_browser(i + 1, q, store, args, progress)) for i in range(n_workers)]
        prog = asyncio.create_task(progress_loop(q, store, args, progress, len(cards)))
        try:
            await q.join()
        finally:
            for t in tasks:
                t.cancel()
            await asyncio.gather(*tasks, return_exceptions=True)
            prog.cancel()
            await asyncio.gather(prog, return_exceptions=True)
    else:
        async with async_playwright() as p:
            pool = BrowserPool(p, args)
            tasks = [asyncio.create_task(brand_worker(i + 1, pool, q, store, args, progress)) for i in range(args.workers)]
            prog = asyncio.create_task(progress_loop(q, store, args, progress, len(cards)))
            try:
                await q.join()
            finally:
                for t in tasks:
                    t.cancel()
                await asyncio.gather(*tasks, return_exceptions=True)
                prog.cancel()
                await asyncio.gather(prog, return_exceptions=True)
                await pool.close_all()
    await store.save()
    elapsed = time.time() - progress["start"]
    print(f"Финальный прогресс: обработано={progress['done']}/{len(cards)}, ссылок={progress['links']}, нет документов={progress['no_docs']}, нет ссылки={progress['no_link']}, тех={progress['tech']}")
    if elapsed > 0:
        print(f"Время: {elapsed:.1f}с, скорость: {progress['done'] / elapsed * 60:.1f} карточек/мин")
    # v25-reporting: финальный текстовый лог прогона
    try:
        _log_arg = getattr(args, "run_log", "") or ""
        _log_path = Path(_log_arg) if _log_arg else None
        _wd = int(getattr(args, "expiry_warning_days", 30) or 0)
        _written = _write_run_log(list(store.rows), Path(args.output), _wd,
                                  _run_started_at, log_path=_log_path)
        if _written:
            print(f"Лог прогона записан: {_written}")
    except Exception as _e:
        print(f"[run-log] ошибка: {_e}")

    # v24.1: диагностика FSA — сколько распарсилось, сколько 403.
    try:
        fsa_rows = [r for r in store.rows if 'fsa.gov.ru' in (r.registry_hosts or '')]
        fsa_ok = [r for r in fsa_rows if (r.registry_doc_number or '').strip()]
        fsa_403 = [r for r in fsa_rows if 'http_403' in (r.registry_details_source or '')]
        swis_rows = [r for r in store.rows if 'swis' in (r.registry_hosts or '') or 'trade.kg' in (r.registry_hosts or '')]
        swis_ok = [r for r in swis_rows if (r.registry_doc_number or '').strip()]
        if fsa_rows or swis_rows:
            print("-" * 70)
            print(f"Реестры: FSA {len(fsa_ok)}/{len(fsa_rows)} распарсено, SWIS {len(swis_ok)}/{len(swis_rows)} распарсено")
            if fsa_403 and len(fsa_ok) == 0 and len(fsa_rows) > 0:
                print("⚠️  ВСЕ FSA-документы вернули HTTP 403 — антибот FSA блокирует HTTP-запросы в твоей сети.")
                print("    Это значит curl_cffi не пробивает защиту здесь. Решения:")
                print("    1) Запусти БЕЗ --no-browser (браузерный режим парсит FSA надёжно):")
                print(f"       python main_brand.py --brand {args.brand} --limit {args.limit} --workers 4")
                print("    2) Или проверь что pub.fsa.gov.ru доступен (возможно нужен hosts/VPN).")
    except Exception:
        pass

    print(f"CSV сохранён: {Path(args.result_csv).resolve()}")
    print(f"Excel сохранён: {Path(args.output).resolve()}")

# -----------------------------
# CLI
# -----------------------------

def build_parser():
    ap = argparse.ArgumentParser(description="WB Brand Checker: бренд, цена, Оригинал, ссылки на реестры документов")
    ap.add_argument("--brand", default="", help="Название бренда, например adidas")
    ap.add_argument("--single-url", "--url", dest="single_url", default="", help="Проверить одну карточку WB по ссылке или nm_id без cards-csv")
    ap.add_argument("--limit", type=int, default=1000)
    ap.add_argument("--cards-csv", default="brand_cards.csv")
    ap.add_argument("--result-csv", default="brand_result.csv")
    ap.add_argument("--output", default="brand_result.xlsx")
    ap.add_argument("--collect-only", action="store_true")
    ap.add_argument("--check-only", action="store_true")
    ap.add_argument("--reuse-cards", type=str_to_bool, default=False, help="Использовать существующий cards-csv вместо нового сбора")
    ap.add_argument("--reset-output", action="store_true")
    ap.add_argument("--resume", type=str_to_bool, default=True)

    ap.add_argument("--brand-match", choices=["exact", "contains", "any"], default="contains", help="Как фильтровать бренд из выдачи. contains лучше для adidas/adidas Originals")
    ap.add_argument("--use-brand-filter", type=str_to_bool, default=True, help="Пытаться найти внутренний fbrand-фильтр WB и собирать выдачу уже с фильтром бренда")
    ap.add_argument("--collect-strategy", choices=["base", "expanded", "hybrid"], default="hybrid")
    ap.add_argument("--max-pages", type=int, default=80, help="Сколько страниц проверять по исходному запросу бренда")
    ap.add_argument("--expanded-pages", type=int, default=12, help="Сколько страниц проверять по запросам бренд+товар")
    ap.add_argument("--max-query-terms", type=int, default=80, help="Сколько расширителей запроса использовать")
    ap.add_argument("--search-sorts", default="popular,rate,priceup,pricedown,newly,benefit")
    ap.add_argument("--expanded-sorts", default="popular,rate,priceup,pricedown")
    ap.add_argument("--extra-terms", default="", help="Дополнительные запросы через запятую")
    ap.add_argument("--shuffle-collect-jobs", type=str_to_bool, default=False)
    ap.add_argument("--collect-workers", type=int, default=24)
    ap.add_argument("--collect-timeout-sec", type=int, default=30)
    ap.add_argument("--collect-save-every-sec", type=int, default=15)
    ap.add_argument("--collect-log-every", type=int, default=25)
    ap.add_argument("--enrich-details", type=str_to_bool, default=True)
    ap.add_argument("--detail-workers", type=int, default=32)
    ap.add_argument("--detail-timeout-sec", type=float, default=12.0)
    ap.add_argument("--browser-collect-fallback", type=str_to_bool, default=True)
    ap.add_argument("--browser-collect-terms", type=int, default=35)
    ap.add_argument("--browser-collect-scrolls", type=int, default=25)
    ap.add_argument("--browser-collect-wait-ms", type=int, default=2500)
    ap.add_argument("--browser-collect-scroll-px", type=int, default=1700)
    ap.add_argument("--browser-collect-scroll-wait-ms", type=int, default=650)
    ap.add_argument("--browser-collect-stable-rounds", type=int, default=4)
    ap.add_argument("--browser-collect-log-every", type=int, default=3)

    ap.add_argument("--registry-mode", choices=["api_first", "ui_only", "direct_only"], default="api_first", help="Как собирать ссылку на реестр: api_first=сначала WB certificate.json, потом UI fallback; ui_only=старые клики; direct_only=только certificate.json")
    ap.add_argument("--registry-details", type=str_to_bool, default=True, help="Быстро собирать сведения о документе из поддерживаемых реестров по найденной ссылке. Поддержаны swis.trade.kg и pub.fsa.gov.ru/rds/rss. Для ФСА пробуются singular и plural API-варианты. По умолчанию true.")
    ap.add_argument("--fsa-browser-fallback", type=str_to_bool, default=False, help="Если прямой сбор ФСА дал 403/пустой ответ, открыть уникальный документ ФСА в Playwright. По умолчанию false: ФСА собирается только быстрыми HTTP-запросами без браузера.")
    ap.add_argument("--fsa-min-delay-sec", type=float, default=0.8, help="Минимальная пауза между прямыми запросами к ФСА. По умолчанию 0.8 сек, чтобы не ловить 403/429 на пачке документов.")
    ap.add_argument("--fsa-aggressive-candidates", type=str_to_bool, default=False, help="Пробовать дополнительные неподтверждённые API-варианты ФСА после точного HAR endpoint. По умолчанию false, чтобы не провоцировать 403/429.")
    ap.add_argument("--fsa-curl-cffi-impersonate", default="chrome", help="Профиль TLS-имперсонации curl_cffi для ФСА: chrome, chrome120, chrome110 и т.п. Используется только если установлен пакет curl_cffi.")
    ap.add_argument("--fsa-browser-timeout-ms", type=int, default=25000, help="Таймаут открытия страницы ФСА в браузерном fallback.")
    ap.add_argument("--fsa-browser-wait-ms", type=int, default=6500, help="Сколько ждать рендеринга/загрузки API на странице ФСА в браузерном fallback.")
    ap.add_argument("--registry-details-timeout-sec", type=float, default=8.0, help="Таймаут быстрого запроса к реестру для сведений о документе")
    ap.add_argument("--require-docs-button", type=str_to_bool, default=False, help="Не рекомендуется для быстрого режима. Если true — дополнительно проверять UI-кнопку Документы проверены. По умолчанию false: certificate.json считается самостоятельным источником истины.")
    ap.add_argument("--api-docs-verify-ms", type=int, default=2500, help="Сколько ждать UI-признак 'Документы проверены' при найденном certificate.json")
    ap.add_argument("--certificate-timeout-sec", type=float, default=8.0, help="Таймаут прямого запроса WB certificate.json")
    ap.add_argument("--certificate-max-hosts", type=int, default=30, help="Сколько basket-хостов пробовать для certificate.json, если основной шард не подошёл")

    ap.add_argument("--browser-count", type=int, default=2)
    ap.add_argument("--workers", type=int, default=4)
    # v24: режим без браузера
    ap.add_argument("--no-browser", type=str_to_bool, default=False,
                    help="v24: НЕ запускать Playwright. Всё через HTTP: certificate.json для ссылок на реестр, curl_cffi для FSA-парсинга, aiohttp для SWIS. В 10-30 раз быстрее. ТРЕБУЕТ pip install curl_cffi для FSA.")
    ap.add_argument("--http-workers", type=int, default=0,
                    help="v24: параллельность для --no-browser режима. По умолчанию = workers × 4. Можно ставить 30-100 — HTTP-запросы лёгкие.")
    ap.add_argument("--headless", type=str_to_bool, default=True)
    ap.add_argument("--block-assets", type=str_to_bool, default=False)
    ap.add_argument("--autosave-every", type=int, default=25)
    ap.add_argument("--context-refresh-every", type=int, default=150)
    ap.add_argument("--max-card-retries", type=int, default=1)
    ap.add_argument("--card-hard-timeout-ms", type=int, default=45000)
    ap.add_argument("--goto-timeout-ms", type=int, default=18000)
    ap.add_argument("--default-timeout-ms", type=int, default=5000)
    ap.add_argument("--after-goto-ms", type=int, default=1200)
    ap.add_argument("--original-wait-ms", type=int, default=7000, help="Сколько ждать появления плашки Оригинал после открытия карточки")
    ap.add_argument("--original-html-fallback", type=str_to_bool, default=True, help="Если DOM/Playwright не увидел Оригинал, проверить публичный HTML/SEO WB")
    ap.add_argument("--original-html-timeout-sec", type=float, default=5.0, help="Таймаут HTML fallback для проверки Оригинал")
    ap.add_argument("--original-html-domains", default="ru,by,kg,ge", help="Домены WB для HTML fallback проверки Оригинал: ru,by,kg,ge")
    ap.add_argument("--after-specs-click-ms", type=int, default=700)
    ap.add_argument("--after-docs-wait-ms", type=int, default=700)
    ap.add_argument("--after-look-wait-ms", type=int, default=2500)
    ap.add_argument("--card-ready-timeout-ms", type=int, default=3500)
    ap.add_argument("--docs-timeout-ms", type=int, default=5000)
    ap.add_argument("--no-docs-fallback-ms", type=int, default=2500)
    ap.add_argument("--look-button-timeout-ms", type=int, default=4000)
    ap.add_argument("--viewport-width", type=int, default=1365)
    ap.add_argument("--viewport-height", type=int, default=900)
    ap.add_argument("--user-agent", default="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    ap.add_argument("--trace", nargs="?", const=True, default=False, type=str_to_bool)
    ap.add_argument("--verbose-each", action="store_true")
    # --- Отчётный слой ---
    ap.add_argument("--expiry-warning-days", type=int, default=30,
                    help="Сколько дней до окончания действия документа считать риском 'Скоро истекает'. По умолчанию 30.")
    ap.add_argument("--make-report-xlsx", type=str_to_bool, default=True,
                    help="Добавлять в итоговый XLSX листы 'Сводка' и 'Подробности' с русскими заголовками и подсветкой по сроку. По умолчанию true.")
    ap.add_argument("--run-log", default="",
                    help="Путь к текстовому логу прогона (run-summary). Если пусто — положится рядом с итоговым XLSX как *_run.log.")
    ap.add_argument("--print-links", action="store_true")
    ap.add_argument("--progress-interval-sec", type=int, default=15)
    return ap

async def main_async(args):
    # Версия уже напечатана до разбора аргументов
    # v24: критическое предупреждение про curl_cffi.
    # БЕЗ него FSA вернёт 403 на все HTTP-запросы (антибот по TLS-fingerprint).
    if _curl_requests is None:
        print("=" * 80)
        print("⚠️  ВНИМАНИЕ: библиотека curl_cffi не установлена.")
        print("    Без неё FSA (pub.fsa.gov.ru) ВСЕГДА вернёт 403 Forbidden — это")
        print("    антибот по TLS-fingerprint. Результат: для FSA-документов поля")
        print("    registry_doc_number, registry_status и др. БУДУТ ПУСТЫМИ.")
        print("")
        print("    УСТАНОВИ ОДНОЙ КОМАНДОЙ:")
        print("        pip install curl_cffi")
        print("")
        print("    SWIS (Киргизстан) работает и без неё через обычный HTTP.")
        print("=" * 80)
        # Не падаем — даём пользователю продолжить, но он знает что FSA не будет работать.
    else:
        print(f"✓ curl_cffi доступен — FSA-документы будут парситься через HTTP (TLS impersonate Chrome).")

    if not args.brand and not args.check_only:
        args.brand = input("Введите название бренда: ").strip()
    if args.collect_only:
        await collect_brand_cards(args)
        return
    await run_check(args)

if __name__ == "__main__":
    # Печатаем версию ДО argparse, чтобы сразу было видно, какой файл реально запущен
    print(f"WB Brand Checker {APP_VERSION} | file={Path(__file__).resolve()}")
    parser = build_parser()
    ns = parser.parse_args()
    try:
        asyncio.run(main_async(ns))
    except KeyboardInterrupt:
        print("Остановлено пользователем")
