# -*- coding: utf-8 -*-
"""
WB+Ozon Checker — единая программа проверки карточек Wildberries и Ozon
против реестров (ФСА, SWIS, Belgiss, EAEU).

Версия: 2026-06-05-v27-unified-ozon

Архитектура:
  • Один Python-файл с встроенным современным интерфейсом на pywebview
    (нативное окно с HTML/CSS/JS внутри).
  • Backend = существующие движки main_v39.py, main_brand.py, ozon_parser.py,
    а также wb_enhanced.py и fsa_enhanced.py (улучшения).
  • 6 режимов: query_full, query_stage1, query_stage2, brand, ozon, unified.
  • 5 экранов в UI: 🚀 Запуск, 📋 Очередь, 📊 Результаты, ⚙️ Настройки, 📜 Логи.
  • Bridge API между JS и Python через webview.api.*.
  • Поддержка слияния результатов WB + Ozon в unified-режиме.

Запуск:
    python wb_checker.py

Зависимости:
    pip install pywebview openpyxl aiohttp curl_cffi playwright beautifulsoup4 lxml
    python -m playwright install chromium

main_v39.py, main_brand.py и ozon_parser.py должны лежать рядом с wb_checker.py.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import platform
import queue
import re
import subprocess
import sys
import threading
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# pywebview import — обязательная зависимость
# ---------------------------------------------------------------------------
try:
    import webview  # type: ignore
except ImportError:
    print(
        "Не установлен pywebview. Установите: pip install pywebview",
        file=sys.stderr,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------
APP_VERSION = "2026-06-20-v54.4"
APP_DIR = Path(__file__).resolve().parent
# pywebview на Windows часто запускается через pythonw.exe (без консоли) — это ломает stdout pipe в дочерних
# процессах. Сила принуждаем использовать python.exe (с консолью) для subprocess.
def _resolve_python() -> str:
    exe = sys.executable or "python3"
    if os.name == "nt" and exe.lower().endswith("pythonw.exe"):
        alt = exe[:-len("pythonw.exe")] + "python.exe"
        if Path(alt).exists():
            return alt
    return exe

PYTHON = _resolve_python()
SETTINGS_PATH = APP_DIR / "wb_checker_settings.json"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stderr)],
)
log = logging.getLogger("wb_checker")


def _pick_engine_v39() -> Path:
    """Возвращает путь к движку WB query mode (берёт первый найденный файл)."""
    for name in ("main_v39.py", "main_v39-2.py", "main_v39_reporting.py"):
        p = APP_DIR / name
        if p.exists():
            return p
    return APP_DIR / "main_v39.py"


ENGINE_V39 = _pick_engine_v39()
ENGINE_BRAND = APP_DIR / "main_brand.py"
ENGINE_OZON = APP_DIR / "ozon_parser.py"
ENGINE_WB_ENHANCED = APP_DIR / "wb_enhanced.py"
ENGINE_FSA_ENHANCED = APP_DIR / "fsa_enhanced.py"


# ---------------------------------------------------------------------------
# Настройки
# ---------------------------------------------------------------------------

def load_settings() -> dict:
    """Загружает настройки из JSON-файла. Возвращает пустой словарь при ошибке."""
    if SETTINGS_PATH.exists():
        try:
            return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        except Exception as exc:
            log.warning("Ошибка загрузки настроек: %s", exc)
    return {}


def save_settings(data: dict) -> None:
    """Сохраняет настройки в JSON-файл."""
    try:
        SETTINGS_PATH.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as exc:
        log.error("Ошибка сохранения настроек: %s", exc)


# ---------------------------------------------------------------------------
# Модели данных
# ---------------------------------------------------------------------------

@dataclass
class RunSpec:
    """Единая модель параметров запуска для всех режимов."""

    mode: str = "query_full"
    """Режим: query_full | query_stage1 | query_stage2 | brand | ozon | unified"""

    query: str = ""
    """Поисковый запрос (для query_* и ozon/unified)."""

    brand: str = ""
    """Название бренда (для brand-режима)."""

    brand_match: str = "exact"
    """Тип совпадения бренда: exact | contains | any."""

    brand_category: str = ""
    """v27.9.x: товарная категория для поиска по бренду (RU-метка). Пусто/«любая»
    — без сужения. Маппится на доменный профиль движка (--query-profile)."""

    query_profile: str = "auto"
    """v27.9.x: доменный профиль движка для stage1 (clothing/shoes/appliances/...)."""

    catalog_sweep: bool = False
    """v48: «поиск без запроса» — программа сама сметает каталог WB по типам товаров."""

    catalog_categories: str = ""
    """v48: RU-категории (через запятую) для сметания каталога. Пусто = ВСЕ категории."""

    registry_fsa_retry: bool = False
    """v27.9.x: второй проход по упавшим FSA (по кнопке)."""

    limit: int = 5000
    """Лимит карточек."""

    workers: int = 4
    """Количество параллельных воркеров/браузеров."""

    expiry_warning_days: int = 30
    """Порог (в днях) для метки «Скоро истекает»."""

    fsa_slow_mode: bool = False
    """v46: медленный режим ФСА (без блокировок) — для больших прогонов."""

    fsa_slow_delay_sec: float = 0
    """v46: пауза между ФСА-документами в медл. режиме, сек. 0 = по умолчанию (~2.5–5с)."""

    make_report_xlsx: bool = True
    """Создавать листы Сводка/Подробности в XLSX."""

    headless: bool = True
    """Запускать браузер в скрытом режиме."""

    strict_brand: str = ""
    """Строгий фильтр по бренду (для Stage 2)."""

    strict_brand_match: str = "any"
    """Тип строгого совпадения: any | exact | contains."""

    output: str = ""
    """Путь к итоговому XLSX (пусто = авто)."""

    output_links_csv: str = ""
    """CSV со ссылками (для Stage 1 → Stage 2)."""

    input_links_csv: str = ""
    """Входной CSV ссылок (для Stage 2)."""

    use_wb_enhanced: bool = False
    """Использовать wb_enhanced.py для улучшенного seller_name и плашки «Оригинал»."""

    use_fsa_enhanced: bool = False
    """Использовать fsa_enhanced.py для расширенных полей ФСА."""

    # Ozon-специфичные поля
    ozon_delay_min_ms: int = 200
    """Минимальная задержка между запросами Ozon (мс)."""

    ozon_delay_max_ms: int = 500
    """Максимальная задержка между запросами Ozon (мс)."""

    unified_report: bool = True
    """В unified-режиме: объединять WB и Ozon в один XLSX."""

    def wb_args(self) -> List[str]:
        """Формирует CLI-аргументы для WB-движка (query или brand)."""
        if self.mode == "brand":
            out = self.output or "brand_result.xlsx"
            return [
                PYTHON, str(ENGINE_BRAND),
                "--brand", self.brand,
                "--limit", str(self.limit),
                "--brand-match", self.brand_match,
                "--workers", str(self.workers),
                "--expiry-warning-days", str(self.expiry_warning_days),
                "--make-report-xlsx", "true" if self.make_report_xlsx else "false",
                "--output", out,
            ]
        engine = str(ENGINE_V39)
        if self.mode == "query_stage1":
            out = self.output or "links.xlsx"
            a = [
                PYTHON, engine,
                "--query", self.query,
                "--limit", str(self.limit),
                "--link-only", "true",
            ]
            # v48: «поиск без запроса» — сметание каталога вместо одного запроса.
            if self.catalog_sweep:
                a += ["--catalog-sweep", "true"]
                if self.catalog_categories:
                    a += ["--catalog-categories", self.catalog_categories]
            a += [
                "--link-mode", "http_only",
                # v45.11: было workers*10 (=30) — слишком много одновременных
                # запросов к wbbasket.ru, WB начинал троттлить IP (растущие «сетевые
                # ошибки»). Шарды теперь пробуются по одному (1 запрос/карточку),
                # поэтому 12-16 воркеров достаточно для скорости и НЕ ловят троттлинг.
                "--http-link-workers", str(max(8, min(16, self.workers * 4))),
                "--output", out,
                "--output-links-csv", self.output_links_csv or "registry_links.csv",
            ]
            # v27.9.x: поиск по бренду через тот же движок — строгий бренд-фильтр.
            if self.strict_brand and self.strict_brand_match != "any":
                a += ["--brand", self.strict_brand, "--brand-match", self.strict_brand_match]
            # v27.9.x: товарная категория бренда -> доменный профиль (сужение выдачи).
            if self.query_profile and self.query_profile != "auto":
                a += ["--query-profile", self.query_profile]
            return a
        if self.mode == "query_stage2":
            out = self.output or "result.xlsx"
            args = [
                PYTHON, engine,
                "--input-links-csv", self.input_links_csv or "registry_links.csv",
                "--limit", str(self.limit),
                "--registry-headless", "true" if self.headless else "false",
                "--registry-browser-workers", str(self.workers),
                "--output", out,
                "--expiry-warning-days", str(self.expiry_warning_days),
                "--make-report-xlsx", "true" if self.make_report_xlsx else "false",
                "--registry-fsa-retry", "true" if self.registry_fsa_retry else "false",
                "--fsa-slow-mode", "true" if self.fsa_slow_mode else "false",
            ]
            if self.fsa_slow_mode and float(self.fsa_slow_delay_sec or 0) > 0:
                _v = float(self.fsa_slow_delay_sec)
                args += ["--fsa-slow-delay-ms", f"{int(_v*1000)},{int(_v*2000)}"]
            if self.strict_brand and self.strict_brand_match != "any":
                args += ["--brand", self.strict_brand, "--brand-match", self.strict_brand_match]
            return args
        # query_full / unified (WB-часть)
        out = self.output or ("wb_result.xlsx" if self.mode == "unified" else "result.xlsx")
        return [
            PYTHON, engine,
            "--query", self.query,
            "--limit", str(self.limit),
            "--registry-headless", "true" if self.headless else "false",
            "--registry-browser-workers", str(self.workers),
            "--output", out,
            "--expiry-warning-days", str(self.expiry_warning_days),
            "--make-report-xlsx", "true" if self.make_report_xlsx else "false",
        ]

    def ozon_args(self) -> List[str]:
        """Формирует CLI-аргументы для движка Ozon."""
        out = self.output or ("ozon_result.xlsx" if self.mode != "unified" else "ozon_result.xlsx")
        args = [
            PYTHON, str(ENGINE_OZON),
            "--query", self.query,
            "--limit", str(self.limit),
            "--workers", str(self.workers),
            "--headless", "true" if self.headless else "false",
            "--output", out,
            "--expiry-warning-days", str(self.expiry_warning_days),
            "--make-report-xlsx", "true" if self.make_report_xlsx else "false",
            "--delay-min-ms", str(self.ozon_delay_min_ms),
            "--delay-max-ms", str(self.ozon_delay_max_ms),
        ]
        return args


@dataclass
class AppState:
    """Live-состояние прогона. Читается фронтом через polling."""

    running: bool = False
    mode: str = ""
    started_at: float = 0.0
    finished_at: float = 0.0
    progress_done: int = 0
    progress_total: int = 0
    progress_pct: float = 0.0
    progress_stage: str = ""
    progress_speed: int = 0   # карточек/мин (из движка)
    progress_eta: int = 0     # сек до конца этапа (из движка)
    log_lines: List[str] = field(default_factory=list)
    output_path: str = ""
    ozon_output_path: str = ""
    log_path: str = ""
    metrics: Dict[str, Any] = field(default_factory=dict)
    last_cmd: str = ""
    error: str = ""
    stage_label: str = ""

    # Временные ряды для графика активности (последние 60 точек, каждая = 1 сек)
    activity_series: List[int] = field(default_factory=list)
    _last_activity_ts: float = field(default=0.0, repr=False)
    _activity_counter: int = field(default=0, repr=False)
    # v45.7: выборки (время, обработано) для ЧЕСТНОЙ скорости строк/мин — считаем по
    # реальному приросту progress_done, а не по числу лог-событий.
    _progress_samples: List = field(default_factory=list, repr=False)

    def record_progress_sample(self) -> None:
        """Запоминает точку (время, progress_done) для расчёта реальной скорости.
        При смене этапа (done пошёл назад) — сбрасываем историю, чтобы скорость не
        прыгала в минус."""
        now = time.time()
        if self._progress_samples and self.progress_done < self._progress_samples[-1][1]:
            self._progress_samples = []
        self._progress_samples.append((now, self.progress_done))
        # держим окно ~40 секунд (но не меньше последних 2 точек)
        cutoff = now - 40.0
        kept = [(t, d) for (t, d) in self._progress_samples if t >= cutoff]
        self._progress_samples = kept if len(kept) >= 2 else self._progress_samples[-2:]

    def tick_activity(self) -> None:
        """Увеличивает счётчик активности — вызывается на каждую обработанную строку."""
        now = time.time()
        self._activity_counter += 1
        if now - self._last_activity_ts >= 1.0:
            self.activity_series.append(self._activity_counter)
            if len(self.activity_series) > 60:
                self.activity_series = self.activity_series[-60:]
            self._activity_counter = 0
            self._last_activity_ts = now

    def to_dict(self) -> dict:
        """Сериализует состояние для передачи во фронтенд."""
        elapsed = 0.0
        if self.started_at:
            end = self.finished_at if (not self.running and self.finished_at) else time.time()
            elapsed = max(0.0, end - self.started_at)
        # v45.7: ЧЕСТНАЯ скорость строк/мин — по реальному приросту обработанных
        # строк (progress_done) за последние ~30 сек, как и пишет движок в логе
        # («460/мин»). Раньше считалось число лог-событий, поэтому показывало
        # заниженную и непонятную цифру (≈60), не совпадавшую с логом.
        speed = 0.0
        samples = [s for s in self._progress_samples if s[0] >= time.time() - 30.0]
        if len(samples) >= 2:
            (t0, d0), (t1, d1) = samples[0], samples[-1]
            dt, dd = (t1 - t0), (d1 - d0)
            if dt > 0 and dd >= 0:
                speed = round(dd / dt * 60.0, 1)
        return {
            "running": self.running,
            "mode": self.mode,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "elapsed_sec": elapsed,
            "progress_done": self.progress_done,
            "progress_total": self.progress_total,
            "progress_pct": self.progress_pct,
            "progress_stage": self.progress_stage,
            "progress_speed": self.progress_speed,
            "progress_eta": self.progress_eta,
            "output_path": self.output_path,
            "ozon_output_path": self.ozon_output_path,
            "log_path": self.log_path,
            "metrics": self.metrics,
            "last_cmd": self.last_cmd,
            "error": self.error,
            "stage_label": self.stage_label,
            "log_total": len(self.log_lines),
            "activity_series": list(self.activity_series),
            "speed_per_min": speed,
        }


# ---------------------------------------------------------------------------
# Regex для парсинга stdout движков
# ---------------------------------------------------------------------------
# v27.9.x: однозначный маркер прогресса от движков (emit_progress). Парсится
# в первую очередь — это убирает скачки полосы из-за «повтор 2/5» и т.п.
PROGRESS_SENTINEL_RX = re.compile(
    r"@@PROGRESS@@\s+stage=(\S+)\s+done=(\d+)\s+total=(\d+)"
    r"(?:\s+speed=(\d+)\s+eta=(\d+))?"
)
PROGRESS_RX = re.compile(r"(\d+)\s*/\s*(\d+)")
PROGRESS_PCT_RX = re.compile(r"(\d+(?:\.\d+)?)\s*%")
# Строки, где «X/Y» — это НЕ прогресс (счётчик повторов, попытки и т.п.).
PROGRESS_FALSE_RX = re.compile(r"повтор|retry|попытк|/мин|/час", re.IGNORECASE)
STATUS_RX = re.compile(
    r"\[(OK|ОШИБКА|ТАЙМАУТ|НЕТ ДОКУМЕНТОВ|НЕТ ССЫЛКИ НА РЕЕСТР|НЕСООТВЕТСТВИЕ|ССЫЛКА НА РЕЕСТР СОБРАНА)\]"
)
OUTPUT_RX = re.compile(r"(?:output|output_path|сохранено|saved)[:\s]+([^\s]+\.xlsx)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# EngineRunner — запуск движков и парсинг stdout
# ---------------------------------------------------------------------------

_CATEGORY_TO_PROFILE = {
    "одежда": "clothing", "обувь": "shoes", "бытовая техника": "appliances",
    "электроника": "electronics", "игрушки": "toys", "косметика": "cosmetics",
    "детские аксессуары": "kids_accessories", "детский транспорт": "baby_gear",
    "дом и текстиль": "home", "посуда": "kitchenware", "продукты": "food",
}


def _categories_to_profile(category_value: str) -> str:
    """v27.9.x: RU-категории (через запятую) -> доменный профиль(и) движка для
    --query-profile. Пусто -> 'auto'. Используется и брендом, и запросом."""
    cats = [c.strip().lower() for c in (category_value or "").split(",") if c.strip()]
    profiles = [_CATEGORY_TO_PROFILE[c] for c in cats if c in _CATEGORY_TO_PROFILE]
    return ",".join(dict.fromkeys(profiles)) if profiles else "auto"


# v45.5: алиасы «сырых» английских заголовков листа results -> русские
# отображаемые имена (как в листе «Подробности»). Нужны, чтобы ЛЮБОЙ result.xlsx
# (старый с английскими колонками или новый) корректно показывался в таблице:
# колонка «Название в реестре» и кликабельные ссылки на товар/реестр опираются на
# русские имена. Ключи — и оригинальные, и в нижнем регистре (для надёжности).
_RESULT_HEADER_ALIASES = {
    "query": "Запрос",
    "nm_id": "Артикул WB",
    "product_name": "Название товара",
    "brand": "Бренд",
    "subject": "Категория WB",
    "product_url": "Ссылка на товар",
    "status": "Технический статус",
    "price_rub": "Цена, ₽",
    "sale_price_rub": "Цена со скидкой, ₽",
    "seller_name": "Продавец",
    "supplier_id": "ID продавца",
    "rating": "Рейтинг",
    "feedbacks": "Отзывы",
    "is_original": "Плашка 'Оригинал'",
    "docs_verified": "Документ проверен WB",
    "colors": "Цвет",
    "wb_root": "Корневой ID (WB)",
    "registry_url": "Ссылка на реестр",
    "registry_host": "Реестр (хост)",
    "registry_record_id": "ID записи реестра",
    "certificate_number": "Номер документа",
    "document_type": "Тип документа",
    "document_status": "Статус документа",
    "rf_status": "Статус на территории РФ",
    "certificate_product_name": "Название в реестре",
    "document_date_start": "Действует с",
    "document_date_end": "Действует до",
    "applicant_name": "Заявитель",
    "applicant_inn": "ИНН заявителя",
    "manufacturer_name": "Изготовитель",
    "tnved": "ТН ВЭД",
    "scheme": "Схема оценки",
    "technical_regulation": "Техрегламент",
    "score": "Оценка совпадения",
    "details": "Примечания",
    "worker": "Worker",
    "checked_at": "Проверено",
}


def freshest_xlsx(path: Path) -> Path:
    """v27.9.x: если основной XLSX был занят (открыт в Excel), движок пишет в
    `<stem>_live.xlsx`. Тогда таблица/графики/сводка ДОЛЖНЫ читать именно его,
    иначе показывают устаревшие данные прошлого прогона. Возвращает самый свежий
    из (основной, _live). Используется и EngineRunner, и Bridge."""
    try:
        path = Path(path)
        live = path.with_name(path.stem + "_live" + path.suffix)
        if live.exists():
            if not path.exists():
                return live
            if live.stat().st_mtime > path.stat().st_mtime + 0.5:
                return live
    except Exception:
        pass
    return path


class EngineRunner:
    """
    Запускает subprocess-движки, читает stdout в фоновом потоке,
    парсит прогресс и метрики, обновляет AppState.
    """

    def __init__(self, state: AppState) -> None:
        self.state = state
        self.proc: Optional[subprocess.Popen] = None
        self.proc_ozon: Optional[subprocess.Popen] = None
        self._thread: Optional[threading.Thread] = None
        self._stop_flag = threading.Event()

    def start(self, spec: RunSpec) -> None:
        """Запускает прогон по переданному RunSpec."""
        if self.state.running:
            raise RuntimeError("Прогон уже выполняется")
        self._reset_state(spec)
        self._stop_flag.clear()
        self._thread = threading.Thread(
            target=self._run, args=(spec,), daemon=True, name="runner"
        )
        self._thread.start()

    def stop(self) -> None:
        """Останавливает текущий прогон."""
        self._stop_flag.set()
        for proc in (self.proc, self.proc_ozon):
            if proc and proc.poll() is None:
                try:
                    proc.terminate()
                except Exception:
                    pass

    def _reset_state(self, spec: RunSpec) -> None:
        s = self.state
        s.running = True
        s.mode = spec.mode
        s.started_at = time.time()
        s.finished_at = 0.0
        s.progress_done = 0
        s.progress_total = 0
        s.progress_pct = 0.0
        s.progress_stage = ""
        s.log_lines = []
        s.error = ""
        s.output_path = ""
        s.ozon_output_path = ""
        s.log_path = ""
        s.metrics = {"status": {}, "risk": {}, "registry": {}, "marketplace": {}}
        s.last_cmd = ""
        s.stage_label = ""
        s.activity_series = []
        s._last_activity_ts = 0.0
        s._activity_counter = 0
        s._progress_samples = []

    def _run(self, spec: RunSpec) -> None:
        try:
            if spec.mode == "unified":
                self._run_unified(spec)
            elif spec.mode in ("query_full", "query_auto"):
                # Цепочка: Stage1 → Stage2. v27.9.x: товарная категория (если
                # выбрана) -> --query-profile, чтобы запрос к WB был точнее.
                # v48: query_auto — «поиск без запроса» (сметание каталога WB).
                _is_auto = spec.mode == "query_auto"
                s1 = RunSpec(**{**asdict(spec), "mode": "query_stage1",
                                "query_profile": _categories_to_profile(spec.brand_category),
                                "catalog_sweep": _is_auto,
                                "catalog_categories": (spec.catalog_categories if _is_auto else ""),
                                "output_links_csv": "registry_links.csv",
                                "output": "links.xlsx"})
                _lbl1 = "🧭 Этап 1 — сбор каталога WB (без запроса)" if _is_auto else "🔍 Этап 1 — сбор ссылок WB"
                rc = self._run_one(s1.wb_args(), _lbl1)
                if rc != 0 or self._stop_flag.is_set():
                    return
                s2 = RunSpec(**{**asdict(spec), "mode": "query_stage2",
                                "input_links_csv": "registry_links.csv"})
                self._run_one(s2.wb_args(), "📋 Этап 2 — парсинг реестров WB")
            elif spec.mode == "ozon":
                self._run_one(spec.ozon_args(), "🛒 Ozon — поиск и проверка")
            elif spec.mode == "brand":
                # v27.9.x: «по бренду» теперь работает ЧЕРЕЗ ТОТ ЖЕ движок main_v39,
                # что и «по запросу». Stage1 ищет карточки бренда (поиск по названию
                # бренда + строгий бренд-фильтр), Stage2 парсит реестры тем же
                # надёжным способом. Итог: идентичные столбцы result.xlsx, «Оригинал»
                # по card.json и корректное извлечение реестров (FSA/SWIS/BelGISS).
                _bm = spec.brand_match if spec.brand_match and spec.brand_match != "any" else "contains"
                _profile = _categories_to_profile(spec.brand_category)
                s1 = RunSpec(**{**asdict(spec), "mode": "query_stage1",
                                "query": spec.brand or spec.query,
                                "strict_brand": spec.brand,
                                "strict_brand_match": _bm,
                                "query_profile": _profile,
                                "output_links_csv": "registry_links.csv",
                                "output": "links.xlsx"})
                rc = self._run_one(s1.wb_args(), "🔍 Этап 1 — сбор карточек бренда WB")
                if rc != 0 or self._stop_flag.is_set():
                    return
                s2 = RunSpec(**{**asdict(spec), "mode": "query_stage2",
                                "input_links_csv": "registry_links.csv"})
                self._run_one(s2.wb_args(), "📋 Этап 2 — парсинг реестров WB")
            else:
                label_map = {
                    "query_stage1": "🔍 Этап 1 — сбор ссылок",
                    "query_stage2": "📋 Этап 2 — парсинг реестров",
                    "brand": "🏷️ Поиск по бренду WB",
                }
                self._run_one(spec.wb_args(), label_map.get(spec.mode, spec.mode))
        except Exception as exc:
            self.state.error = f"{type(exc).__name__}: {exc}"
            self.state.log_lines.append(f"[ОШИБКА] {self.state.error}")
            log.exception("Runner error")
        finally:
            self.state.running = False
            self.state.finished_at = time.time()
            # v27.9.x: явно «закрываем» прогресс, иначе после завершения полоса
            # остаётся на старом значении/этапе (особенно если движок нашёл 0
            # результатов и прогресс не двигался) — и кажется, что работа идёт.
            self.state.progress_stage = ""
            if not self.state.error:
                self.state.progress_pct = 100.0
                if self.state.progress_total:
                    self.state.progress_done = self.state.progress_total
            self._finalize_paths(spec)

    def _run_unified(self, spec: RunSpec) -> None:
        """Запускает WB и Ozon параллельно, затем мержит результаты."""
        self.state.stage_label = "🔄 Unified: WB + Ozon параллельно"
        wb_spec = RunSpec(**{**asdict(spec),
                             "mode": "query_full",
                             "output": str(APP_DIR / "wb_result.xlsx")})
        ozon_spec = RunSpec(**{**asdict(spec),
                               "mode": "ozon",
                               "output": str(APP_DIR / "ozon_result.xlsx")})

        wb_lines: List[str] = []
        ozon_lines: List[str] = []
        results: Dict[str, int] = {}

        def run_wb() -> None:
            rc = self._run_one_collect(wb_spec.wb_args(), "📦 WB — поиск + реестры", wb_lines)
            results["wb_rc"] = rc

        def run_ozon() -> None:
            rc = self._run_one_collect(ozon_spec.ozon_args(), "🛒 Ozon — поиск + реестры", ozon_lines)
            results["ozon_rc"] = rc

        t_wb = threading.Thread(target=run_wb, daemon=True, name="runner_wb")
        t_ozon = threading.Thread(target=run_ozon, daemon=True, name="runner_ozon")
        t_wb.start()
        t_ozon.start()
        t_wb.join()
        t_ozon.join()

        if self._stop_flag.is_set():
            return

        # Мерж результатов
        wb_xlsx = APP_DIR / "wb_result.xlsx"
        ozon_xlsx = APP_DIR / "ozon_result.xlsx"
        unified_xlsx = APP_DIR / (spec.output or "unified_result.xlsx")

        if wb_xlsx.exists() and ozon_xlsx.exists() and spec.unified_report:
            self.state.log_lines.append("\n━━━ 🔀 Объединение отчётов WB + Ozon ━━━")
            try:
                merge_xlsx(wb_xlsx, ozon_xlsx, unified_xlsx)
                self.state.output_path = str(unified_xlsx)
                self.state.log_lines.append(f"[OK] Unified отчёт: {unified_xlsx}")
            except Exception as exc:
                self.state.log_lines.append(f"[ОШИБКА мержа] {exc}")
                self.state.output_path = str(wb_xlsx) if wb_xlsx.exists() else str(ozon_xlsx)
        else:
            self.state.output_path = str(wb_xlsx) if wb_xlsx.exists() else ""
        self.state.ozon_output_path = str(ozon_xlsx) if ozon_xlsx.exists() else ""

    def _run_one(self, args: List[str], label: str) -> int:
        """Запускает один subprocess, читает stdout, возвращает код возврата."""
        self.state.stage_label = label
        self.state.log_lines.append(f"\n━━━ {label} ━━━")
        cmd_str = " ".join(f'"{a}"' if " " in str(a) else str(a) for a in args)
        self.state.last_cmd = cmd_str
        self.state.log_lines.append(f"[команда] {cmd_str}")
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUNBUFFERED"] = "1"
            env["PYTHONUTF8"] = "1"
            proc = subprocess.Popen(
                args,
                cwd=str(APP_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env,
                creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
            )
            self.proc = proc
        except FileNotFoundError as exc:
            msg = f"[ОШИБКА ЗАПУСКА] Файл не найден: {exc}\n  args[0]={args[0] if args else '?'}\n  PYTHON={PYTHON}"
            self.state.log_lines.append(msg)
            self.state.error = msg
            return -1
        except Exception as exc:
            msg = f"[ОШИБКА ЗАПУСКА] {type(exc).__name__}: {exc}"
            self.state.log_lines.append(msg)
            self.state.error = msg
            return -1

        assert proc.stdout is not None
        for raw_line in proc.stdout:
            if self._stop_flag.is_set():
                break
            line = raw_line.rstrip("\n")
            self.state.log_lines.append(line)
            if len(self.state.log_lines) > 6000:
                self.state.log_lines = self.state.log_lines[-5000:]
            self._parse_line(line)

        proc.wait()
        rc = proc.returncode or 0
        self.state.log_lines.append(f"[завершено, код {rc}]")
        if rc != 0 and not self.state.error:
            self.state.error = f"Subprocess завершился с кодом {rc} (см. Логи)"
        return rc

    def _run_one_collect(self, args: List[str], label: str, lines_out: List[str]) -> int:
        """Версия _run_one для параллельного запуска — добавляет строки в lines_out."""
        self.state.log_lines.append(f"\n━━━ {label} ━━━")
        cmd_str = " ".join(f'"{a}"' if " " in str(a) else str(a) for a in args)
        self.state.log_lines.append(f"[команда] {cmd_str}")
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            env["PYTHONUNBUFFERED"] = "1"
            env["PYTHONUTF8"] = "1"
            proc = subprocess.Popen(
                args,
                cwd=str(APP_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env,
                creationflags=(subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0),
            )
        except Exception as exc:
            msg = f"[ОШИБКА ЗАПУСКА {label}] {type(exc).__name__}: {exc}"
            self.state.log_lines.append(msg)
            self.state.error = msg
            return -1

        assert proc.stdout is not None
        for raw_line in proc.stdout:
            if self._stop_flag.is_set():
                break
            line = raw_line.rstrip("\n")
            lines_out.append(line)
            self.state.log_lines.append(f"  [{label[:2]}] {line}")
            if len(self.state.log_lines) > 6000:
                self.state.log_lines = self.state.log_lines[-5000:]
            self._parse_line(line)

        proc.wait()
        return proc.returncode or 0

    def _parse_line(self, line: str) -> None:
        """Парсит строку stdout: прогресс, метрики, пути к файлам."""
        # 1) Однозначный машиночитаемый маркер прогресса (приоритет).
        m = PROGRESS_SENTINEL_RX.search(line)
        if m:
            try:
                stage, d, t = m.group(1), int(m.group(2)), int(m.group(3))
                if 0 < t < 10_000_000:
                    self.state.progress_stage = stage
                    self.state.progress_done = d
                    self.state.progress_total = t
                    self.state.progress_pct = min(100.0, 100.0 * d / t)
                    # v54.4 (улучшение №7): скорость (карточек/мин) и ETA (сек) из движка
                    if m.group(4) is not None:
                        self.state.progress_speed = int(m.group(4))
                        self.state.progress_eta = int(m.group(5))
                    self.state.tick_activity()
                    self.state.record_progress_sample()
                    return
            except ValueError:
                pass
        # 2) Запасной разбор «X/Y» из обычного лога — но не из строк, где
        #    «X/Y» означает счётчик повторов/скорость (иначе полоса скачет).
        if not PROGRESS_FALSE_RX.search(line):
            m = PROGRESS_RX.search(line)
            if m:
                try:
                    d, t = int(m.group(1)), int(m.group(2))
                    if 0 < t < 10_000_000 and d <= t:
                        self.state.progress_done = d
                        self.state.progress_total = t
                        self.state.progress_pct = min(100.0, 100.0 * d / t)
                        self.state.tick_activity()
                        self.state.record_progress_sample()
                        return
                except ValueError:
                    pass
        # Прогресс X%
        m = PROGRESS_PCT_RX.search(line)
        if m:
            try:
                self.state.progress_pct = min(100.0, float(m.group(1)))
            except ValueError:
                pass
        # Статус OK/ОШИБКА и т.д.
        m = STATUS_RX.search(line)
        if m:
            k = m.group(1)
            st = self.state.metrics.setdefault("status", {})
            st[k] = st.get(k, 0) + 1
            self.state.tick_activity()
        # v27.9.x: распределение по реестрам из строк прогресса (FSA=.., SWIS=..).
        # Это наполняет график «Реестры», который раньше оставался «нет данных».
        reg_hits = re.findall(r"\b(FSA|SWIS|BELGISS|BelGISS)\s*=\s*(\d+)", line)
        if reg_hits:
            reg = self.state.metrics.setdefault("registry", {})
            for _name, _val in reg_hits:
                _up = _name.upper()
                _key = "ФСА" if _up == "FSA" else ("SWIS" if _up == "SWIS" else "BelGISS")
                try:
                    reg[_key] = int(_val)  # в логе кумулятивные тоталы — присваиваем
                except ValueError:
                    pass
            self.state.tick_activity()
        # Путь к выходному файлу
        m = OUTPUT_RX.search(line)
        if m:
            p = Path(m.group(1).strip())
            if not p.is_absolute():
                p = APP_DIR / p
            if p.suffix == ".xlsx":
                name_lower = p.name.lower()
                if "ozon" in name_lower:
                    self.state.ozon_output_path = str(p)
                elif not self.state.output_path:
                    self.state.output_path = str(p)

    def _finalize_paths(self, spec: RunSpec) -> None:
        """Определяет итоговые пути к XLSX и лог-файлу после завершения прогона."""
        if not self.state.output_path:
            out_name = spec.output or {
                "query_stage1": "links.xlsx",
                "query_stage2": "result.xlsx",
                "query_full": "result.xlsx",
                "brand": "brand_result.xlsx",
                "ozon": "ozon_result.xlsx",
                "unified": "unified_result.xlsx",
            }.get(spec.mode, "result.xlsx")
            p = Path(out_name)
            if not p.is_absolute():
                p = APP_DIR / p
            if p.exists():
                self.state.output_path = str(p)

        # Для ozon-only режима: явно выставляем ozon_output_path к общему файлу,
        # чтобы кнопка «Ozon XLSX» заработала. Парсер всегда пишет файл (даже диагностический).
        if spec.mode == "ozon" and self.state.output_path and not self.state.ozon_output_path:
            self.state.ozon_output_path = self.state.output_path

        # Лог-файл рядом с xlsx
        if self.state.output_path:
            p = Path(self.state.output_path)
            lp = p.with_name(p.stem + "_run.log")
            if lp.exists():
                self.state.log_path = str(lp)
            # Парсим Сводку из xlsx
            self._parse_summary(p)

    def _parse_summary(self, xlsx_path: Path) -> None:
        """Читает лист 'Сводка' из XLSX и заполняет metrics.status/risk/registry."""
        try:
            from openpyxl import load_workbook  # type: ignore
            xlsx_path = freshest_xlsx(Path(xlsx_path))
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            if "Сводка" not in wb.sheetnames:
                return
            ws = wb["Сводка"]
            status: Dict[str, int] = {}
            risk: Dict[str, int] = {}
            registry: Dict[str, int] = {}
            section = None
            for row in ws.iter_rows(values_only=True):
                cells = [c for c in row if c is not None]
                if not cells:
                    section = None
                    continue
                head = str(cells[0])
                if "Распределение по техническому статусу" in head:
                    section = "status"
                    continue
                if "Риски по сроку" in head:
                    section = "risk"
                    continue
                if "Реестр" in head or "По реестрам" in head:
                    section = "registry"
                    continue
                if section in ("status", "risk", "registry") and len(cells) >= 2:
                    if head in ("Статус", "Категория", "Реестр"):
                        continue
                    try:
                        cnt = int(cells[1])
                        if section == "status":
                            status[head] = cnt
                        elif section == "risk":
                            risk[head] = cnt
                        elif section == "registry":
                            registry[head] = cnt
                    except Exception:
                        pass
            if status:
                self.state.metrics["status"] = status
            if risk:
                self.state.metrics["risk"] = risk
            if registry:
                self.state.metrics["registry"] = registry
            # Обнаружение диагностического xlsx (Ozon вернул 0 товаров)
            if "Диагностика" in wb.sheetnames:
                ws_d = wb["Диагностика"]
                diag_msg = ""
                for row in ws_d.iter_rows(values_only=True):
                    if row and row[0] and str(row[0]).strip() == "diagnosis" and len(row) > 1:
                        diag_msg = str(row[1] or "")
                        break
                if diag_msg:
                    self.state.log_lines.append(f"[⚠️ Ozon] {diag_msg}")
                    self.state.metrics["ozon_diagnosis"] = diag_msg
        except Exception as exc:
            self.state.log_lines.append(f"[summary parse fail] {exc}")


# ---------------------------------------------------------------------------
# merge_xlsx — объединение WB + Ozon отчётов
# ---------------------------------------------------------------------------

def merge_xlsx(wb_xlsx: Path, ozon_xlsx: Path, output_xlsx: Path) -> None:
    """
    Сливает два отчёта WB и Ozon в единый XLSX.

    Структура итогового файла:
    - Подробности (общий лист с колонкой «Маркетплейс»)
    - Подробности WB (исходный лист)
    - Подробности Ozon (исходный лист)
    - Сводка (объединённая статистика)
    """
    try:
        from openpyxl import load_workbook, Workbook  # type: ignore
        from openpyxl.styles import PatternFill, Font, Alignment  # type: ignore
    except ImportError as exc:
        raise RuntimeError(f"openpyxl не установлен: {exc}") from exc

    def read_sheet(path: Path, sheet_hint: str = "Подробности") -> Tuple[List[str], List[List]]:
        """Читает указанный лист или первый доступный."""
        if not path.exists():
            return [], []
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = None
        for name in (sheet_hint, "results", wb.sheetnames[0] if wb.sheetnames else None):
            if name and name in wb.sheetnames:
                sheet = wb[name]
                break
        if sheet is None:
            return [], []
        headers: List[str] = []
        rows: List[List] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
            else:
                rows.append([c if c is not None else "" for c in row])
        return headers, rows

    wb_headers, wb_rows = read_sheet(wb_xlsx)
    oz_headers, oz_rows = read_sheet(ozon_xlsx)

    out = Workbook()
    out.remove(out.active)  # удаляем дефолтный лист

    header_fill = PatternFill("solid", fgColor="1E1F29")
    header_font = Font(bold=True, color="FFFFFF")

    def write_sheet(ws, headers: List[str], rows: List[List]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)

    # Лист «Подробности» — объединённый с колонкой «Маркетплейс»
    ws_all = out.create_sheet("Подробности")
    mp_col = "Маркетплейс"
    all_headers = [mp_col] + wb_headers if wb_headers else [mp_col] + oz_headers
    wb_header_set = set(wb_headers)
    oz_header_set = set(oz_headers)
    # Добавляем уникальные Ozon-колонки
    for h in oz_headers:
        if h not in wb_header_set and h not in all_headers:
            all_headers.append(h)

    write_sheet(ws_all, all_headers, [])
    # Убираем заголовки (уже записаны), добавляем данные
    # WB строки
    for row in wb_rows:
        padded = ["WB"] + list(row) + [""] * (len(all_headers) - len(row) - 1)
        ws_all.append(padded[:len(all_headers)])
    # Ozon строки — маппим по заголовкам
    for row in oz_rows:
        oz_dict = dict(zip(oz_headers, row))
        new_row = ["Ozon"] + [oz_dict.get(h, "") for h in all_headers[1:]]
        ws_all.append(new_row[:len(all_headers)])

    # Отдельные листы
    if wb_rows:
        ws_wb = out.create_sheet("Подробности WB")
        write_sheet(ws_wb, wb_headers, wb_rows)

    if oz_rows:
        ws_oz = out.create_sheet("Подробности Ozon")
        write_sheet(ws_oz, oz_headers, oz_rows)

    # Сводный лист
    ws_sum = out.create_sheet("Сводка")
    ws_sum.append(["Параметр", "Значение"])
    ws_sum.append(["WB карточек", len(wb_rows)])
    ws_sum.append(["Ozon карточек", len(oz_rows)])
    ws_sum.append(["Всего карточек", len(wb_rows) + len(oz_rows)])
    ws_sum.append(["Дата объединения", time.strftime("%Y-%m-%d %H:%M:%S")])

    out.save(str(output_xlsx))
    log.info("Unified XLSX сохранён: %s", output_xlsx)


# ---------------------------------------------------------------------------
# Bridge API (JS ↔ Python)
# ---------------------------------------------------------------------------

class Bridge:
    """Методы Bridge, доступные из JS как window.pywebview.api.*"""

    def __init__(self) -> None:
        self.state = AppState()
        self.runner = EngineRunner(self.state)
        s = load_settings()
        self._last_spec: dict = s.get("last_spec", {})
        self._missing_deps: List[str] = []
        self._window = None  # ссылка на окно webview (для файловых диалогов)
        self._loaded_result_path: str = ""  # внешний result.xlsx, загруженный для анализа

    def diagnose(self) -> dict:
        """Возвращает диагностику: python, движки, зависимости, последняя команда."""
        return {
            "ok": True,
            "app_version": APP_VERSION,
            "python": sys.executable,
            "python_version": sys.version.split()[0],
            "platform": platform.platform(),
            "app_dir": str(APP_DIR),
            "engines": {
                "main_v39.py": ENGINE_V39.exists(),
                "main_brand.py": ENGINE_BRAND.exists(),
                "ozon_parser.py": ENGINE_OZON.exists(),
                "wb_enhanced.py": ENGINE_WB_ENHANCED.exists(),
                "fsa_enhanced.py": ENGINE_FSA_ENHANCED.exists(),
            },
            "missing_deps": self._missing_deps,
            "last_cmd": self.state.last_cmd,
            "last_error": self.state.error,
        }

    # ---- статус ----

    def get_state(self) -> dict:
        """Возвращает полное состояние прогона для фронтенда."""
        return self.state.to_dict()

    def get_log_lines(self, offset: int = 0) -> dict:
        """Возвращает порцию лога начиная с offset (для инкрементального polling)."""
        offset = max(0, int(offset))
        chunk = self.state.log_lines[offset: offset + 500]
        return {
            "from": offset,
            "total": len(self.state.log_lines),
            "lines": chunk,
        }

    def get_version(self) -> dict:
        """Возвращает версию приложения и статус наличия движков."""
        return {
            "app_version": APP_VERSION,
            "engine_v39": str(ENGINE_V39),
            "engine_v39_exists": ENGINE_V39.exists(),
            "engine_brand": str(ENGINE_BRAND),
            "engine_brand_exists": ENGINE_BRAND.exists(),
            "engine_ozon": str(ENGINE_OZON),
            "engine_ozon_exists": ENGINE_OZON.exists(),
            "engine_wb_enhanced_exists": ENGINE_WB_ENHANCED.exists(),
            "engine_fsa_enhanced_exists": ENGINE_FSA_ENHANCED.exists(),
            "python": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
            "platform": platform.system(),
            "app_dir": str(APP_DIR),
        }

    # ---- настройки ----

    def get_settings(self) -> dict:
        """Загружает и возвращает текущие настройки."""
        return load_settings()

    def save_settings(self, data: dict) -> dict:
        """Сохраняет настройки полностью."""
        s = load_settings()
        s.update(data or {})
        save_settings(s)
        return {"ok": True}

    def get_last_spec(self) -> dict:
        """Возвращает последний использованный RunSpec."""
        return dict(self._last_spec or {})

    # ---- запуск ----

    def start_run(self, spec: dict) -> dict:
        """Запускает прогон по переданным параметрам."""
        if self.state.running:
            return {"ok": False, "error": "Прогон уже выполняется"}
        spec = spec or {}
        # Фильтруем только известные поля
        valid_keys = set(RunSpec.__dataclass_fields__.keys())
        filtered = {k: v for k, v in spec.items() if k in valid_keys}
        try:
            run_spec = RunSpec(**filtered)
        except Exception as exc:
            return {"ok": False, "error": f"Неверные параметры: {exc}"}

        # Валидация
        mode = run_spec.mode
        if mode in ("query_full", "query_stage1", "ozon", "unified") and not run_spec.query.strip():
            return {"ok": False, "error": "Укажите поисковый запрос"}
        if mode == "brand" and not run_spec.brand.strip():
            return {"ok": False, "error": "Укажите название бренда"}
        if mode in ("query_full", "query_auto", "query_stage1", "query_stage2") and not ENGINE_V39.exists():
            return {"ok": False, "error": f"Движок WB Query не найден: {ENGINE_V39.name}"}
        if mode == "brand" and not ENGINE_V39.exists():
            # v27.9.x: бренд-режим теперь использует движок main_v39 (как «по запросу»).
            return {"ok": False, "error": f"Движок WB Query не найден: {ENGINE_V39.name}"}
        if mode in ("ozon", "unified") and not ENGINE_OZON.exists():
            return {"ok": False, "error": f"Движок Ozon не найден: {ENGINE_OZON.name}"}
        if mode == "unified" and not ENGINE_V39.exists():
            return {"ok": False, "error": f"Движок WB Query не найден для unified-режима: {ENGINE_V39.name}"}

        # Сохраняем последний spec
        self._last_spec = spec
        s = load_settings()
        s["last_spec"] = spec
        save_settings(s)

        # Новый прогон — вкладка «Результаты» должна показывать ЕГО результат, а не
        # ранее загруженный внешний файл. Сбрасываем загруженный путь.
        self._loaded_result_path = ""

        try:
            self.runner.start(run_spec)
        except Exception as exc:
            return {"ok": False, "error": str(exc)}
        return {"ok": True}

    def stop_run(self) -> dict:
        """Останавливает текущий прогон."""
        self.runner.stop()
        return {"ok": True}

    # ---- результаты ----

    def browse_result_file(self) -> dict:
        """Открывает файловый диалог и загружает выбранный result.xlsx (с другого
        прогона) для анализа во вкладке «Результаты». Возвращает выбранный путь —
        дальше JS вызывает get_results(path) с ним. Путь запоминается, чтобы
        «Обновить»/CSV/графики работали с этим же файлом."""
        try:
            import webview  # type: ignore
        except Exception:
            return {"ok": False, "error": "webview недоступен"}
        if self._window is None:
            return {"ok": False, "error": "Окно не готово"}
        try:
            file_types = ("Excel (*.xlsx)", "Все файлы (*.*)")
            res = self._window.create_file_dialog(
                webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types)
        except Exception as exc:
            return {"ok": False, "error": f"Не удалось открыть диалог: {exc}"}
        if not res:
            return {"ok": False, "cancelled": True}
        path = res[0] if isinstance(res, (list, tuple)) else str(res)
        p = Path(path)
        if not p.exists():
            return {"ok": False, "error": "Файл не найден"}
        if p.suffix.lower() != ".xlsx":
            return {"ok": False, "error": "Нужен файл .xlsx"}
        self._loaded_result_path = str(p)
        return {"ok": True, "path": str(p), "name": p.name}

    def clear_loaded_result(self) -> dict:
        """Сбрасывает загруженный внешний файл — вкладка снова показывает результат
        текущего прогона."""
        self._loaded_result_path = ""
        return {"ok": True}

    def _kg_status_path(self) -> Path:
        return APP_DIR / "kg_rf_status.xlsx"

    @staticmethod
    def _registry_country(host_or_url: str) -> str:
        """РФ (ФСА) / КГ (киргизский SWIS) / BY (БелГИСС) / ЕАЭС — по хосту реестра."""
        s = str(host_or_url or "").lower()
        if "fsa.gov.ru" in s:
            return "РФ"
        if "trade.kg" in s or "swis" in s:
            return "КГ"
        if "belgiss" in s:
            return "BY"
        if "eaeunion" in s:
            return "ЕАЭС"
        return ""

    def _add_registry_country(self, headers: List[str], rows: List[List]) -> None:
        """v46: добавляет колонку «Реестр (страна)» (РФ/КГ/BY) по хосту реестра —
        и для свежего прогона, и для загруженного файла."""
        low = [h.strip().lower() for h in headers]
        ci_host = next((i for i, h in enumerate(low) if h in ("реестр (хост)", "registry_host")), -1)
        ci_url = next((i for i, h in enumerate(low) if h in ("ссылка на реестр", "registry_url")), -1)
        if ci_host < 0 and ci_url < 0:
            return
        ci_reg = next((i for i, h in enumerate(low) if h in ("реестр (страна)", "registry_country")), -1)
        if ci_reg < 0:
            headers.append("Реестр (страна)")
            ci_reg = len(headers) - 1
            for r in rows:
                r.append("")
        for r in rows:
            while len(r) < len(headers):
                r.append("")
            src = ""
            if ci_host >= 0 and ci_host < len(r) and str(r[ci_host]).strip():
                src = str(r[ci_host])
            elif ci_url >= 0 and ci_url < len(r):
                src = str(r[ci_url])
            r[ci_reg] = self._registry_country(src)

    def kg_status_info(self) -> dict:
        """Сколько записей в загруженной таблице статусов КГ-документов в РФ."""
        p = next((q for q in (APP_DIR / "kg_rf_status.xlsx", APP_DIR / "kg_rf_status.csv")
                  if q.exists()), None)
        if p is None:
            return {"ok": True, "loaded": False, "count": 0}
        try:
            import main_v39 as _mv
            n = _mv.load_kg_rf_status(str(p))
            return {"ok": True, "loaded": n > 0, "count": int(n)}
        except Exception:
            return {"ok": True, "loaded": True, "count": 0}

    def browse_kg_status_file(self) -> dict:
        """Открывает файловый диалог, копирует выбранную таблицу (xlsx/csv) статусов
        КГ-документов в РФ в kg_rf_status.xlsx рядом с программой. Дальше движок
        автоматически её использует: совпавшие по номеру киргизские документы
        получают «Статус на территории РФ» и вердикт «НЕДЕЙСТВУЕТ В РФ»."""
        try:
            import webview  # type: ignore
        except Exception:
            return {"ok": False, "error": "webview недоступен"}
        if self._window is None:
            return {"ok": False, "error": "Окно не готово"}
        try:
            res = self._window.create_file_dialog(
                webview.OPEN_DIALOG, allow_multiple=False,
                file_types=("Таблицы (*.xlsx;*.csv)", "Все файлы (*.*)"))
        except Exception as exc:
            return {"ok": False, "error": f"Не удалось открыть диалог: {exc}"}
        if not res:
            return {"ok": False, "cancelled": True}
        src = Path(res[0] if isinstance(res, (list, tuple)) else str(res))
        if not src.exists():
            return {"ok": False, "error": "Файл не найден"}
        try:
            import shutil
            dst = self._kg_status_path()
            # .csv тоже принимаем — кладём как kg_rf_status.csv, движок ищет оба
            if src.suffix.lower() == ".csv":
                dst = APP_DIR / "kg_rf_status.csv"
                try:
                    (APP_DIR / "kg_rf_status.xlsx").unlink()
                except Exception:
                    pass
            shutil.copyfile(src, dst)
        except Exception as exc:
            return {"ok": False, "error": f"Не удалось скопировать файл: {exc}"}
        info = self.kg_status_info()
        return {"ok": True, "name": src.name, "count": info.get("count", 0)}

    def _apply_kg_rf_status(self, headers: List[str], rows: List[List]) -> None:
        """v46: для уже прочитанных строк (загруженный файл) проставляет «Статус на
        территории РФ» и вердикт «НЕДЕЙСТВУЕТ В РФ» по таблице КГ-документов.
        Колонку добавляет, если её ещё нет."""
        kg_path = next((q for q in (APP_DIR / "kg_rf_status.xlsx", APP_DIR / "kg_rf_status.csv")
                        if q.exists()), None)
        if kg_path is None:
            return
        try:
            import main_v39 as _mv
            if not getattr(_mv, "_KG_RF_STATUS_MAP", None):
                _mv.load_kg_rf_status(str(kg_path))
            if not _mv._KG_RF_STATUS_MAP:
                return
        except Exception:
            return

        def _col(*names):
            low = [h.strip().lower() for h in headers]
            for n in names:
                if n in low:
                    return low.index(n)
            return -1

        ci_num = _col("номер документа", "certificate_number")
        ci_status = _col("технический статус", "status")
        if ci_num < 0:
            return
        ci_rf = _col("статус на территории рф", "rf_status")
        if ci_rf < 0:
            headers.append("Статус на территории РФ")
            ci_rf = len(headers) - 1
            for r in rows:
                r.append("")
        for r in rows:
            # выравниваем длину строки под заголовки
            while len(r) < len(headers):
                r.append("")
            num = r[ci_num] if ci_num < len(r) else ""
            rf = _mv.kg_rf_status_text(num)
            if rf:
                r[ci_rf] = rf
                if 0 <= ci_status < len(r):
                    r[ci_status] = _mv.STATUS_INVALID_IN_RF

    def get_results(self, xlsx_path: Optional[str] = None, limit: int = 100000) -> dict:
        """
        Читает лист «Подробности» из XLSX и возвращает данные для таблицы.
        Также считает статистику по статусу, реестру и маркетплейсу.
        """
        # Приоритет: явный путь из JS → загруженный внешний файл → текущий прогон.
        chosen = xlsx_path or self._loaded_result_path or self.state.output_path or ""
        # freshest_xlsx ищет более свежий результат рядом — но для ЯВНО загруженного
        # пользователем файла этого делать НЕ нужно (показываем именно его).
        if xlsx_path or self._loaded_result_path:
            path = Path(chosen)
        else:
            path = freshest_xlsx(Path(chosen))
        if not path.exists():
            return {"ok": False, "error": "Файл результата не найден. Запустите прогон или загрузите файл."}
        # v47: КЭШ для больших файлов. На 20k+ строк чтение xlsx + статистика
        # занимает секунды — без кэша КАЖДОЕ переключение на экран «Результаты»
        # замораживало окно. Перечитываем только если файл реально изменился.
        try:
            st = path.stat()
            cache_key = (str(path), st.st_mtime_ns, st.st_size)
        except Exception:
            cache_key = None
        cached = getattr(self, "_results_cache", None)
        if cache_key and cached and cached.get("key") == cache_key:
            return cached["payload"]
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_name = "Подробности"
            if sheet_name not in wb.sheetnames:
                for candidate in ("results", wb.sheetnames[0] if wb.sheetnames else None):
                    if candidate and candidate in wb.sheetnames:
                        sheet_name = candidate
                        break
            ws = wb[sheet_name]
            headers: List[str] = []
            rows: List[List] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else "" for c in row]
                elif i <= limit:
                    rows.append([(c if c is not None else "") for c in row])
                else:
                    break
            # Нормализуем «сырые» английские заголовки листа results в русские
            # отображаемые имена (как в листе «Подробности»). Без этого старые файлы
            # (или лист results) показывались без «Названия в реестре» и без
            # кликабельных ссылок на реестр — таблица ориентируется на русские имена.
            headers = [_RESULT_HEADER_ALIASES.get(h, _RESULT_HEADER_ALIASES.get(h.strip().lower(), h))
                       for h in headers]
            total_rows = (ws.max_row or 1) - 1

            # v46: применяем таблицу статусов КГ-документов в РФ и к ЗАГРУЖЕННОМУ
            # файлу (не только к свежему прогону). Совпавшие по номеру киргизские
            # документы получают «Статус на территории РФ» и вердикт «НЕДЕЙСТВУЕТ В РФ».
            try:
                self._apply_kg_rf_status(headers, rows)
            except Exception:
                pass
            # v46: колонка «Реестр (страна)» — РФ/КГ/BY по хосту реестра.
            try:
                self._add_registry_country(headers, rows)
            except Exception:
                pass

            # Статистика
            stats: Dict[str, Any] = {
                "by_status": {}, "by_registry": {}, "by_marketplace": {},
                "by_original": {}, "by_doc_status": {}, "by_brand": {},
                "by_risk": {},
            }
            h_lower = [h.lower() for h in headers]

            def find_col(*needles: str) -> Optional[int]:
                # Ищем точное совпадение сначала, потом фрагмент
                for needle in needles:
                    if needle in h_lower:
                        return h_lower.index(needle)
                for needle in needles:
                    for i, h in enumerate(h_lower):
                        if needle in h:
                            return i
                return None

            def _host_of(url: str) -> str:
                try:
                    from urllib.parse import urlparse
                    return (urlparse(url).hostname or "").replace("www.", "")
                except Exception:
                    return url

            try:
                idx_status      = find_col("технический статус", "status", "статус")
                idx_registry    = find_col("registry_host", "реестр (хост)", "реестр (host)", "реестр (хост", "registry_url")
                idx_marketplace = find_col("marketplace", "маркетплейс")
                idx_original    = find_col("is_original", "оригинал")
                idx_docstatus   = find_col("document_status", "статус документа")
                idx_brand       = find_col("brand", "бренд")
                idx_risk        = find_col("риск по сроку", "риск", "risk")
                idx_details     = find_col("примечания", "details", "детали")
                # Если маркетплейс не в файле — определяем по product_url
                # (в листе «Подробности» колонка называется «Ссылка на товар»).
                idx_purl = find_col("product_url", "ссылка на товар", "ссылка на товар (wb)")
                # v54.4 (улучшение №8): группировка «почему ПРОВЕРИТЬ ВРУЧНУЮ» по причине
                stats["by_review_reason"] = {}

                def _review_reason(det: str) -> str:
                    d = (det or "").lower()
                    if "категори" in d and ("не совпада" in d or "другой групп" in d):
                        return "конфликт категории"
                    if "слой" in d and "не совпада" in d:
                        return "не совпал слой/вид одежды"
                    if "низкое совпадение" in d:
                        return "нет категории / низкое совпадение"
                    if "частичное совпадение" in d:
                        return "частичное совпадение"
                    if "коды тн вэд" in d or "тн вэд, а не название" in d:
                        return "в реестре только коды ТН ВЭД"
                    if "не извлечено" in d:
                        return "не извлечено наименование"
                    return "прочее"
                for row in rows:
                    if idx_status is not None and idx_status < len(row):
                        v = str(row[idx_status] or "").strip()
                        if v: stats["by_status"][v] = stats["by_status"].get(v, 0) + 1
                    if idx_registry is not None and idx_registry < len(row):
                        v = str(row[idx_registry] or "").strip()
                        if v.startswith("http"):
                            v = _host_of(v)
                        if v: stats["by_registry"][v] = stats["by_registry"].get(v, 0) + 1
                    if idx_marketplace is not None and idx_marketplace < len(row):
                        v = str(row[idx_marketplace] or "").strip()
                        if v: stats["by_marketplace"][v] = stats["by_marketplace"].get(v, 0) + 1
                    elif idx_purl is not None and idx_purl < len(row):
                        u = str(row[idx_purl] or "").lower()
                        if "wildberries" in u: mk = "Wildberries"
                        elif "ozon" in u: mk = "Ozon"
                        else: mk = "Не определен"
                        stats["by_marketplace"][mk] = stats["by_marketplace"].get(mk, 0) + 1
                    if idx_original is not None and idx_original < len(row):
                        v = str(row[idx_original] or "").strip()
                        # Нормализуем в два стабильных значения
                        vl = v.lower()
                        if vl in ("true", "да", "оригинал"):
                            key = "Оригинал"
                        elif vl in ("false", "нет"):
                            key = "Не оригинал"
                        elif not v or vl in ("не указано", "none"):
                            key = "Не указано"
                        else:
                            key = v
                        stats["by_original"][key] = stats["by_original"].get(key, 0) + 1
                    if idx_docstatus is not None and idx_docstatus < len(row):
                        v = str(row[idx_docstatus] or "").strip()
                        if v: stats["by_doc_status"][v] = stats["by_doc_status"].get(v, 0) + 1
                    if idx_brand is not None and idx_brand < len(row):
                        v = str(row[idx_brand] or "").strip()
                        if v: stats["by_brand"][v] = stats["by_brand"].get(v, 0) + 1
                    if idx_risk is not None and idx_risk < len(row):
                        v = str(row[idx_risk] or "").strip()
                        if v: stats["by_risk"][v] = stats["by_risk"].get(v, 0) + 1
                    # причина ручной проверки (только для ПРОВЕРИТЬ ВРУЧНУЮ)
                    if (idx_status is not None and idx_status < len(row)
                            and str(row[idx_status] or "").strip() == "ПРОВЕРИТЬ ВРУЧНУЮ"):
                        det = str(row[idx_details] or "") if idx_details is not None and idx_details < len(row) else ""
                        rk = _review_reason(det)
                        stats["by_review_reason"][rk] = stats["by_review_reason"].get(rk, 0) + 1
                # Обрезаем by_brand до топ-12, остальные в «Прочее»
                if len(stats["by_brand"]) > 12:
                    sb = sorted(stats["by_brand"].items(), key=lambda kv: -kv[1])
                    top = dict(sb[:11])
                    rest = sum(v for _, v in sb[11:])
                    if rest > 0:
                        top["Прочее"] = rest
                    stats["by_brand"] = top
            except Exception as exc:
                log.warning("Сборка статистики не удалась: %s", exc)

            payload = {
                "ok": True,
                "sheet": sheet_name,
                "columns": headers,
                "rows": rows,
                "total": total_rows,
                "stats": stats,
                "xlsx_path": str(path),
            }
            if cache_key:
                self._results_cache = {"key": cache_key, "payload": payload}
            return payload
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def open_path(self, path: str) -> dict:
        """Открывает файл или папку в системном проводнике/приложении."""
        try:
            target = Path(path)
            if not target.exists():
                return {"ok": False, "error": f"Путь не найден: {path}"}
            if sys.platform == "win32":
                os.startfile(str(target))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(target)])
            else:
                subprocess.Popen(["xdg-open", str(target)])
            return {"ok": True}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def open_workspace(self) -> dict:
        """Открывает рабочую папку программы."""
        return self.open_path(str(APP_DIR))

    def retry_failed_fsa(self) -> dict:
        """v27.9.x: ПОВТОР по упавшим FSA-ссылкам (по кнопке). Запускает этап 2
        на том же registry_links.csv с включённым режимом повтора FSA. Движок
        переносит успешные документы из предыдущего result.xlsx как есть и
        пере-проверяет ТОЛЬКО упавшие FSA-ссылки (нет номера/названия/статуса).
        Нажимать, когда pub.fsa.gov.ru снова доступен."""
        if self.state.running:
            return {"ok": False, "error": "Дождитесь завершения текущего прогона"}
        last = dict(self._last_spec or {})
        out = self.state.output_path or last.get("output") or "result.xlsx"
        out_name = Path(out).name
        links = last.get("output_links_csv") or last.get("input_links_csv") or "registry_links.csv"
        if not (APP_DIR / links).exists() and not Path(links).exists():
            return {"ok": False, "error": f"Не найден файл ссылок {links}. Сначала выполните этап 1."}
        spec = {
            "mode": "query_stage2",
            "input_links_csv": links,
            "output": out_name,
            "limit": int(last.get("limit", 10000) or 10000),
            "workers": int(last.get("workers", 5) or 5),
            "headless": bool(last.get("headless", True)),
            "expiry_warning_days": int(last.get("expiry_warning_days", 30) or 30),
            "make_report_xlsx": bool(last.get("make_report_xlsx", True)),
            "registry_fsa_retry": True,
            "strict_brand": last.get("strict_brand", "") or "",
            "strict_brand_match": last.get("strict_brand_match", "any") or "any",
        }
        run_spec = RunSpec(**{k: v for k, v in spec.items() if k in RunSpec.__dataclass_fields__})
        try:
            self.runner.start(run_spec)
        except Exception as exc:
            return {"ok": False, "error": str(exc)}
        return {"ok": True}

    def open_url(self, url: str) -> dict:
        """v27.9.x: открывает ссылку (товар WB / реестр документа) в системном
        браузере. Используется кликами в таблице результатов."""
        try:
            u = (url or "").strip()
            if not (u.startswith("http://") or u.startswith("https://")):
                return {"ok": False, "error": "Некорректная ссылка"}
            import webbrowser
            webbrowser.open(u)
            return {"ok": True}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def export_csv(self, xlsx_path: Optional[str] = None, filter_text: str = "") -> dict:
        """
        Экспортирует отфильтрованные данные из XLSX в CSV.
        Возвращает путь к сохранённому CSV-файлу.
        """
        res = self.get_results(xlsx_path)
        if not res.get("ok"):
            return res
        rows = res["rows"]
        headers = res["columns"]
        q = filter_text.lower().strip() if filter_text else ""
        if q:
            rows = [r for r in rows if any(q in str(c).lower() for c in r)]
        ts = time.strftime("%Y%m%d_%H%M%S")
        out_path = APP_DIR / f"export_{ts}.csv"
        try:
            with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            return {"ok": True, "path": str(out_path), "rows": len(rows)}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    # v53 (улучшение №10): выгрузка ТОЛЬКО спорных строк («ПРОВЕРИТЬ ВРУЧНУЮ» +
    # «НЕСООТВЕТСТВИЕ») отдельным Excel с КЛИКАБЕЛЬНЫМИ ссылками на карточку WB и
    # на реестр + причиной вердикта. Ревью 300-800 строк по ссылкам вместо листания
    # всех 50 000.
    DISPUTED_STATUSES = ("ПРОВЕРИТЬ ВРУЧНУЮ", "НЕСООТВЕТСТВИЕ")

    def export_disputed(self, xlsx_path: Optional[str] = None) -> dict:
        res = self.get_results(xlsx_path)
        if not res.get("ok"):
            return res
        rows = res["rows"]
        headers = res["columns"]
        h_lower = [str(h).lower() for h in headers]

        def col(*needles: str) -> Optional[int]:
            for needle in needles:
                if needle in h_lower:
                    return h_lower.index(needle)
            for needle in needles:
                for i, h in enumerate(h_lower):
                    if needle in h:
                        return i
            return None

        i_status = col("технический статус", "status", "статус")
        i_nm     = col("артикул", "nm_id", "артикул wb")
        i_name   = col("название товара", "product_name", "наименование")
        i_brand  = col("бренд", "brand")
        i_cert   = col("название в реестре", "certificate_product_name")
        i_reason = col("примечания", "details", "детали", "причина")
        i_purl   = col("ссылка на товар (wb)", "ссылка на товар", "product_url")
        i_rurl   = col("ссылка на реестр", "registry_url", "реестр (url)")
        if i_status is None:
            return {"ok": False, "error": "В файле нет колонки статуса — нечего фильтровать."}

        def cell(r, idx):
            return "" if idx is None or idx >= len(r) else (r[idx] if r[idx] is not None else "")

        disputed = [r for r in rows
                    if str(cell(r, i_status)).strip() in self.DISPUTED_STATUSES]
        if not disputed:
            return {"ok": False, "error": "Спорных строк (ПРОВЕРИТЬ ВРУЧНУЮ / НЕСООТВЕТСТВИЕ) не найдено."}

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "На проверку"
            out_headers = ["Артикул", "Название товара", "Бренд", "Вердикт",
                           "Название в реестре", "Причина", "Карточка WB", "Реестр"]
            ws.append(out_headers)
            hfont = Font(bold=True, color="FFFFFF")
            hfill = PatternFill("solid", fgColor="3A5FBF")
            for c in ws[1]:
                c.font = hfont
                c.fill = hfill
                c.alignment = Alignment(vertical="center")
            link_font = Font(color="0563C1", underline="single")
            warn_fill = PatternFill("solid", fgColor="FFF3CD")   # ПРОВЕРИТЬ ВРУЧНУЮ
            bad_fill = PatternFill("solid", fgColor="F8D7DA")     # НЕСООТВЕТСТВИЕ
            for r in disputed:
                verdict = str(cell(r, i_status)).strip()
                purl = str(cell(r, i_purl)).strip()
                rurl = str(cell(r, i_rurl)).strip()
                ws.append([
                    cell(r, i_nm), cell(r, i_name), cell(r, i_brand), verdict,
                    cell(r, i_cert), cell(r, i_reason),
                    "Открыть" if purl else "", "Открыть" if rurl else "",
                ])
                row_i = ws.max_row
                vcell = ws.cell(row=row_i, column=4)
                vcell.fill = bad_fill if verdict == "НЕСООТВЕТСТВИЕ" else warn_fill
                if purl:
                    lc = ws.cell(row=row_i, column=7)
                    lc.hyperlink = purl
                    lc.font = link_font
                if rurl:
                    lc = ws.cell(row=row_i, column=8)
                    lc.hyperlink = rurl
                    lc.font = link_font
            widths = [14, 42, 18, 20, 46, 50, 12, 12]
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
            ws.freeze_panes = "A2"
            ts = time.strftime("%Y%m%d_%H%M%S")
            out_path = APP_DIR / f"на_проверку_{ts}.xlsx"
            wb.save(str(out_path))
            return {"ok": True, "path": str(out_path), "rows": len(disputed)}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def _results_columns_rows(self, xlsx_path):
        res = self.get_results(xlsx_path)
        if not res.get("ok"):
            return None, None, res
        return res["columns"], res["rows"], None

    @staticmethod
    def _col_finder(headers):
        h_lower = [str(h).lower() for h in headers]

        def col(*needles):
            for needle in needles:
                if needle in h_lower:
                    return h_lower.index(needle)
            for needle in needles:
                for i, h in enumerate(h_lower):
                    if needle in h:
                        return i
            return None
        return col

    def export_supplier_stats(self, xlsx_path: Optional[str] = None) -> dict:
        """v54.4 (улучшение №12): сводка по продавцам/брендам — у кого больше всего
        несоответствий/спорных. Помогает приоритизировать ручные проверки."""
        headers, rows, err = self._results_columns_rows(xlsx_path)
        if err:
            return err
        col = self._col_finder(headers)
        i_status = col("технический статус", "status", "статус")
        i_seller = col("продавец", "seller_name", "поставщик")
        i_brand = col("бренд", "brand")
        if i_status is None or (i_seller is None and i_brand is None):
            return {"ok": False, "error": "В файле нет нужных колонок (статус/продавец/бренд)."}

        def cell(r, idx):
            return "" if idx is None or idx >= len(r) else str(r[idx] or "")

        BAD = {"НЕСООТВЕТСТВИЕ", "ПРОВЕРИТЬ ВРУЧНУЮ", "НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ",
               "НЕДЕЙСТВУЕТ В РФ", "ДОКУМЕНТ НЕ ПРОВЕРЕН"}
        from collections import defaultdict
        agg = defaultdict(lambda: defaultdict(int))
        keyname = "продавец" if i_seller is not None else "бренд"
        ikey = i_seller if i_seller is not None else i_brand
        for r in rows:
            key = cell(r, ikey).strip() or "(не указан)"
            st = cell(r, i_status).strip()
            agg[key]["всего"] += 1
            if st == "OK":
                agg[key]["OK"] += 1
            if st in BAD:
                agg[key]["проблемных"] += 1
            agg[key][st] += 1
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "По продавцам"
            ws.append([keyname.capitalize(), "Всего", "OK", "Проблемных",
                       "% проблемных", "НЕСООТВЕТСТВИЕ", "ПРОВЕРИТЬ ВРУЧНУЮ",
                       "НЕДЕЙСТВ. ДОК", "НЕДЕЙСТВ. В РФ", "ДОК НЕ ПРОВЕРЕН"])
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="3A5FBF")
            ordered = sorted(agg.items(), key=lambda kv: -kv[1]["проблемных"])
            for key, d in ordered:
                tot = d["всего"]
                pct = round(100.0 * d["проблемных"] / tot, 1) if tot else 0.0
                ws.append([key, tot, d["OK"], d["проблемных"], pct,
                           d["НЕСООТВЕТСТВИЕ"], d["ПРОВЕРИТЬ ВРУЧНУЮ"],
                           d["НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ"], d["НЕДЕЙСТВУЕТ В РФ"],
                           d["ДОКУМЕНТ НЕ ПРОВЕРЕН"]])
            for ci, w in enumerate([40, 9, 8, 12, 13, 16, 18, 14, 14, 16], 1):
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
            ws.freeze_panes = "A2"
            ts = time.strftime("%Y%m%d_%H%M%S")
            out_path = APP_DIR / f"по_продавцам_{ts}.xlsx"
            wb.save(str(out_path))
            return {"ok": True, "path": str(out_path), "rows": len(ordered)}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}

    def suggest_dictionary(self, xlsx_path: Optional[str] = None) -> dict:
        """v54.4 (улучшение №9): «обучить словарь». Сканирует товары, у которых
        НЕ определилась категория, собирает частые незнакомые слова и выгружает их
        кандидатами в словарь (слово, частота, пример). Пользователь проставляет
        категорию и кладёт файл как dictionary.csv рядом с программой."""
        headers, rows, err = self._results_columns_rows(xlsx_path)
        if err:
            return err
        col = self._col_finder(headers)
        i_name = col("название товара", "product_name", "наименование")
        i_status = col("технический статус", "status", "статус")
        if i_name is None:
            return {"ok": False, "error": "В файле нет колонки с названием товара."}

        def cell(r, idx):
            return "" if idx is None or idx >= len(r) else str(r[idx] or "")
        try:
            import main_v39 as _mv
            import re as _re
            from collections import Counter
        except Exception as exc:
            return {"ok": False, "error": f"Движок недоступен: {exc}"}

        STOP = {"для", "под", "без", "при", "это", "или", "как", "что", "над",
                "из", "на", "по", "до", "от", "со", "шт", "см", "мл", "гр", "кг"}
        freq = Counter()
        example = {}
        focus = {"ПРОВЕРИТЬ ВРУЧНУЮ", "НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА"}
        for r in rows:
            name = cell(r, i_name)
            if not name:
                continue
            st = cell(r, i_status).strip()
            # карточки без определённой категории — кандидаты на пополнение словаря
            try:
                cats = _mv._detect_categories(name)
            except Exception:
                cats = set()
            if cats:
                continue
            # на спорных/неизвлечённых акцент сильнее (но берём и прочие без категории)
            weight = 3 if st in focus else 1
            for w in _re.findall(r"[а-яёa-z]{4,}", name.lower()):
                if w in STOP:
                    continue
                freq[w] += weight
                example.setdefault(w, name)
        cands = [(w, c) for w, c in freq.most_common(400) if c >= 3]
        if not cands:
            return {"ok": False, "error": "Незнакомых частых слов не найдено — словарь уже хорошо покрывает товары."}
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "Кандидаты словаря"
            ws.append(["слово/стем", "категория (впишите)", "частота", "пример названия"])
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="3A5FBF")
            for w, c in cands:
                ws.append([w, "", c, example.get(w, "")[:80]])
            for ci, wdt in enumerate([22, 22, 10, 60], 1):
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = wdt
            ws.freeze_panes = "A2"
            ts = time.strftime("%Y%m%d_%H%M%S")
            out_path = APP_DIR / f"словарь_кандидаты_{ts}.xlsx"
            wb.save(str(out_path))
            return {"ok": True, "path": str(out_path), "rows": len(cands),
                    "hint": "Впишите категорию рядом с каждым словом, сохраните как dictionary.csv (слово,категория) рядом с программой."}
        except Exception as exc:
            return {"ok": False, "error": str(exc)}


# ---------------------------------------------------------------------------
# Frontend HTML/CSS/JS
# ---------------------------------------------------------------------------

FRONTEND_HTML = r"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WB+Ozon Checker v27</title>
<style>
/* ===================== CSS Variables ===================== */
:root {
  --bg:       #0a0a0f;
  --surface:  #15161d;
  --card:     #1e1f29;
  --card2:    #252636;
  --border:   rgba(255,255,255,0.08);
  --border2:  rgba(255,255,255,0.12);
  --fg:       #e8eaf2;
  --fg2:      #a0a3b1;
  --muted:    #6b6e82;
  --accent:   #5b8cff;
  --accent2:  #4070e0;
  --accent3:  #3a5fbf;
  --success:  #10b981;
  --warning:  #f59e0b;
  --error:    #ef4444;
  --info:     #38bdf8;
  --ozon:     #005bff;
  --wb:       #cb11ab;
}
[data-theme="light"] {
  --bg:       #f5f6fa;
  --surface:  #ffffff;
  --card:     #ffffff;
  --card2:    #f0f1f8;
  --border:   rgba(0,0,0,0.1);
  --border2:  rgba(0,0,0,0.15);
  --fg:       #111827;
  --fg2:      #374151;
  --muted:    #6b7280;
  --accent:   #3b6ff5;
  --accent2:  #2d5de0;
}

/* ===================== Reset ===================== */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body {
  height: 100%; overflow: hidden;
  background: var(--bg); color: var(--fg);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, "Helvetica Neue", sans-serif;
  font-size: 14px; line-height: 1.5;
}

/* ===================== Layout ===================== */
.app { display: grid; grid-template-columns: 240px 1fr; height: 100vh; }

/* ===================== Sidebar ===================== */
.sidebar {
  background: var(--surface);
  border-right: 1px solid var(--border);
  display: flex; flex-direction: column;
  padding: 0;
  overflow: hidden;
}
.sidebar-logo {
  padding: 20px 20px 16px;
  border-bottom: 1px solid var(--border);
}
.sidebar-logo .brand-name {
  font-size: 16px; font-weight: 700; letter-spacing: -0.3px;
  background: linear-gradient(135deg, var(--accent), #a78bfa);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
.sidebar-logo .brand-sub { font-size: 11px; color: var(--muted); margin-top: 3px; }
.nav { padding: 10px 8px; flex: 1; display: flex; flex-direction: column; gap: 2px; }
.nav-btn {
  background: transparent; border: none; color: var(--fg2);
  padding: 10px 12px; text-align: left; cursor: pointer;
  border-radius: 8px; font-size: 13px; font-family: inherit;
  display: flex; align-items: center; gap: 10px;
  transition: background 0.15s ease, color 0.15s ease;
  width: 100%;
}
.nav-btn:hover { background: var(--card); color: var(--fg); }
.nav-btn.active {
  background: rgba(91,140,255,0.15);
  color: var(--accent);
}
.nav-btn.active .nav-ico { opacity: 1; }
.nav-ico { font-size: 16px; width: 20px; text-align: center; flex-shrink: 0; }
.nav-label { flex: 1; }
.nav-badge {
  font-size: 10px; font-weight: 600;
  background: var(--accent); color: white;
  border-radius: 10px; padding: 1px 6px;
  display: none;
}
.nav-badge.visible { display: inline-block; }
.sidebar-footer {
  padding: 12px 16px 14px;
  border-top: 1px solid var(--border);
  font-size: 11px; color: var(--muted);
  line-height: 1.6;
}
.sidebar-footer .engine-dot {
  display: inline-block; width: 7px; height: 7px;
  border-radius: 50%; background: var(--muted);
  margin-right: 4px;
}
.engine-dot.ok { background: var(--success); }
.engine-dot.miss { background: var(--error); }

/* ===================== Main area ===================== */
.main { display: flex; flex-direction: column; overflow: hidden; background: var(--bg); }
.topbar {
  background: var(--surface);
  border-bottom: 1px solid var(--border);
  padding: 0 22px;
  height: 56px;
  display: flex; align-items: center; justify-content: space-between;
  flex-shrink: 0;
}
.topbar-left { display: flex; align-items: center; gap: 12px; }
.topbar-title { font-size: 16px; font-weight: 600; }
.topbar-right { display: flex; align-items: center; gap: 16px; }
.status-indicator {
  display: flex; align-items: center; gap: 7px;
  font-size: 12.5px; color: var(--muted);
}
.status-dot {
  width: 8px; height: 8px; border-radius: 50%;
  background: var(--muted); flex-shrink: 0;
}
.status-dot.running {
  background: var(--accent);
  animation: pulse-dot 1.4s ease-in-out infinite;
}
.status-dot.ok { background: var(--success); }
.status-dot.error { background: var(--error); }
@keyframes pulse-dot {
  0%, 100% { box-shadow: 0 0 0 0 rgba(91,140,255,0.5); }
  50% { box-shadow: 0 0 0 5px rgba(91,140,255,0); }
}
.content {
  flex: 1; overflow-y: auto; overflow-x: hidden;
  padding: 22px; scroll-behavior: smooth;
}
/* Custom scrollbar */
.content::-webkit-scrollbar,
.log-view::-webkit-scrollbar,
.tbl-wrap::-webkit-scrollbar { width: 8px; height: 8px; }
.content::-webkit-scrollbar-track,
.log-view::-webkit-scrollbar-track,
.tbl-wrap::-webkit-scrollbar-track { background: transparent; }
.content::-webkit-scrollbar-thumb,
.log-view::-webkit-scrollbar-thumb,
.tbl-wrap::-webkit-scrollbar-thumb {
  background: var(--card2); border-radius: 4px;
}
.content::-webkit-scrollbar-thumb:hover,
.log-view::-webkit-scrollbar-thumb:hover,
.tbl-wrap::-webkit-scrollbar-thumb:hover { background: var(--border2); }

/* ===================== Screens ===================== */
.screen { display: none; }
.screen.active {
  display: block;
  animation: fadeUp 0.18s ease-out;
}
@keyframes fadeUp {
  from { opacity: 0; transform: translateY(6px); }
  to   { opacity: 1; transform: translateY(0); }
}

/* ===================== Cards ===================== */
.card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 20px;
  margin-bottom: 16px;
}
.card-title {
  font-size: 13.5px; font-weight: 600; color: var(--fg);
  margin-bottom: 14px;
  display: flex; align-items: center; gap: 8px;
}
.card-title .ctag {
  font-size: 10px; font-weight: 500; padding: 2px 7px;
  border-radius: 10px; background: rgba(91,140,255,0.15); color: var(--accent);
}

/* ===================== KPI Grid ===================== */
.kpi-grid {
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 12px;
  margin-bottom: 16px;
}
.kpi-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 16px 18px;
  position: relative;
  overflow: hidden;
}
.kpi-card::before {
  content: '';
  position: absolute; top: 0; left: 0; right: 0; height: 2px;
  background: var(--kpi-color, var(--accent));
}
.kpi-label {
  font-size: 11px; font-weight: 500;
  color: var(--muted); text-transform: uppercase;
  letter-spacing: 0.6px; margin-bottom: 8px;
}
.kpi-value {
  font-size: 26px; font-weight: 700;
  color: var(--kpi-color, var(--fg));
  line-height: 1;
}
.kpi-sub { font-size: 11.5px; color: var(--muted); margin-top: 5px; }

/* ===================== Segmented Control / Tabs ===================== */
.seg-ctrl {
  display: inline-flex;
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 10px; padding: 4px; gap: 2px;
  margin-bottom: 16px;
}
.seg-btn {
  background: transparent; border: none;
  color: var(--muted); cursor: pointer;
  padding: 8px 16px; border-radius: 7px;
  font-size: 13px; font-family: inherit;
  transition: background 0.15s, color 0.15s;
  white-space: nowrap;
}
.seg-btn:hover { color: var(--fg); }
.seg-btn.active { background: var(--accent); color: white; }
.seg-btn .mkt-tag {
  font-size: 10px; opacity: 0.75; margin-left: 5px;
}

/* ===================== Form fields ===================== */
.form-grid {
  display: grid; gap: 14px 18px;
}
.form-grid.cols-2 { grid-template-columns: 1fr 1fr; }
.form-grid.cols-3 { grid-template-columns: 1fr 1fr 1fr; }

.field {
  display: flex; flex-direction: column; gap: 5px;
}
.field-label {
  font-size: 12px; font-weight: 500; color: var(--fg2);
}
.field-hint { font-size: 11px; color: var(--muted); margin-top: -2px; }
.field input[type=text],
.field input[type=number],
.field select {
  background: var(--surface);
  border: 1px solid var(--border2);
  color: var(--fg);
  padding: 9px 12px; border-radius: 8px;
  font-size: 13.5px; font-family: inherit;
  outline: none;
  transition: border-color 0.15s, box-shadow 0.15s;
  width: 100%;
}
.field input:focus,
.field select:focus {
  border-color: var(--accent);
  box-shadow: 0 0 0 3px rgba(91,140,255,0.12);
}
.field select option { background: var(--surface); }

/* Toggle Switch */
.toggle-wrap {
  display: flex; align-items: center; gap: 10px; cursor: pointer;
}
.toggle-wrap input[type=checkbox] { display: none; }
.toggle-track {
  width: 38px; height: 22px; border-radius: 12px;
  background: var(--card2); border: 1px solid var(--border2);
  position: relative; transition: background 0.2s, border-color 0.2s;
  flex-shrink: 0;
}
.toggle-track::after {
  content: '';
  position: absolute; top: 2px; left: 2px;
  width: 16px; height: 16px; border-radius: 50%;
  background: var(--muted); transition: left 0.2s, background 0.2s;
}
.toggle-wrap input:checked + .toggle-track {
  background: var(--accent); border-color: var(--accent);
}
.toggle-wrap input:checked + .toggle-track::after {
  left: 18px; background: white;
}
.toggle-label { font-size: 13px; color: var(--fg2); user-select: none; }
.toggle-sub { font-size: 11px; color: var(--muted); }

/* ===================== Buttons ===================== */
.btn {
  display: inline-flex; align-items: center; gap: 7px;
  border: none; cursor: pointer;
  padding: 10px 20px; border-radius: 8px;
  font-size: 13.5px; font-weight: 500; font-family: inherit;
  transition: background 0.15s, transform 0.08s, box-shadow 0.15s;
  white-space: nowrap;
}
.btn:active { transform: scale(0.97); }
.btn:disabled { opacity: 0.4; cursor: not-allowed; transform: none !important; }
.btn-primary {
  background: var(--accent); color: white;
  box-shadow: 0 2px 8px rgba(91,140,255,0.3);
}
.btn-primary:hover:not(:disabled) {
  background: var(--accent2);
  box-shadow: 0 4px 12px rgba(91,140,255,0.4);
}
.btn-danger { background: var(--error); color: white; }
.btn-danger:hover:not(:disabled) { background: #dc2626; }
.btn-ghost {
  background: transparent; color: var(--fg2);
  border: 1px solid var(--border2);
}
.btn-ghost:hover:not(:disabled) { background: var(--card); color: var(--fg); }
.btn-lg { padding: 13px 28px; font-size: 14.5px; }
.btn-sm { padding: 7px 14px; font-size: 12.5px; }

.actions-row {
  display: flex; flex-wrap: wrap; gap: 10px;
  align-items: center; margin-top: 16px;
}

/* ===================== Progress ===================== */
.progress-wrap { margin: 12px 0 16px; }
.progress-bar-bg {
  height: 8px; background: var(--card2); border-radius: 4px; overflow: hidden;
}
.progress-bar-fill {
  height: 100%;
  background: linear-gradient(90deg, var(--accent) 0%, #a78bfa 100%);
  width: 0%; border-radius: 4px;
  transition: width 0.3s ease;
}
.progress-info {
  display: flex; justify-content: space-between;
  font-size: 12px; color: var(--muted); margin-top: 7px;
}

/* ===================== Chart containers ===================== */
.charts-row {
  display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 14px;
  margin-bottom: 16px;
}
.chart-card {
  background: var(--card);
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 16px;
}
.chart-card .chart-title {
  font-size: 12px; font-weight: 600; color: var(--fg2);
  margin-bottom: 12px;
}
.chart-canvas-wrap { position: relative; height: 180px; }
.chart-fallback {
  height: 180px; display: flex; align-items: center; justify-content: center;
  color: var(--muted); font-size: 12px; text-align: center;
}

/* Activity sparkline */
.activity-chart-wrap {
  position: relative; height: 80px;
  background: var(--card2); border-radius: 8px;
  overflow: hidden; margin-top: 10px;
}

/* ===================== Badges ===================== */
.badge {
  display: inline-flex; align-items: center;
  padding: 3px 9px; border-radius: 5px;
  font-size: 11.5px; font-weight: 500; white-space: nowrap;
}
.badge-ok      { background: rgba(16,185,129,0.12); color: var(--success); }
.badge-warn    { background: rgba(245,158,11,0.12); color: var(--warning); }
.badge-error   { background: rgba(239,68,68,0.12);  color: var(--error); }
.badge-info    { background: rgba(56,189,248,0.12); color: var(--info); }
.badge-muted   { background: var(--card2); color: var(--muted); }
.badge-accent  { background: rgba(91,140,255,0.12); color: var(--accent); }
.badge-wb      { background: rgba(203,17,171,0.12); color: var(--wb); }
.badge-ozon    { background: rgba(0,91,255,0.12);   color: var(--ozon); }

.badge-group { display: flex; flex-wrap: wrap; gap: 6px; }

/* ===================== Table ===================== */
.tbl-wrap {
  overflow: auto;
  max-height: calc(100vh - 280px);
  border: 1px solid var(--border);
  border-radius: 10px;
}
table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
thead th {
  position: sticky; top: 0; z-index: 1;
  background: var(--surface); padding: 10px 12px;
  text-align: left; font-weight: 600; font-size: 11.5px;
  color: var(--muted); text-transform: uppercase; letter-spacing: 0.4px;
  border-bottom: 1px solid var(--border2); white-space: nowrap;
}
tbody td {
  padding: 8px 12px;
  border-bottom: 1px solid var(--border);
  color: var(--fg2); max-width: 280px;
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
tbody tr:hover td { background: var(--card); color: var(--fg); }
tbody tr:last-child td { border-bottom: none; }

/* ===================== Log ===================== */
.log-toolbar {
  display: flex; align-items: center; gap: 10px; margin-bottom: 10px;
  flex-wrap: wrap;
}
.log-filter-btns { display: flex; gap: 4px; }
.log-filter-btn {
  background: var(--card); border: 1px solid var(--border);
  color: var(--muted); padding: 4px 10px; border-radius: 6px;
  font-size: 11.5px; cursor: pointer; font-family: inherit;
  transition: background 0.12s, color 0.12s;
}
.log-filter-btn.active { background: var(--accent); color: white; border-color: var(--accent); }
.log-view {
  background: #09090e;
  border: 1px solid var(--border);
  border-radius: 10px;
  font-family: "JetBrains Mono", "Fira Code", Consolas, monospace;
  font-size: 12px; line-height: 1.55;
  padding: 14px; overflow: auto;
  height: calc(100vh - 220px);
  white-space: pre-wrap; word-break: break-all;
}
.log-line-err  { color: #f87171; }
.log-line-warn { color: #fbbf24; }
.log-line-ok   { color: #34d399; }
.log-line-info { color: #94a3b8; }
.log-line-head { color: var(--accent); font-weight: 700; }
.log-line-cmd  { color: #64748b; font-style: italic; }

/* ===================== Filter bar ===================== */
.filter-bar {
  display: flex; gap: 10px; align-items: center; margin-bottom: 14px;
}
.filter-bar input {
  flex: 1; background: var(--card); border: 1px solid var(--border2);
  color: var(--fg); padding: 9px 14px; border-radius: 8px;
  font-family: inherit; font-size: 13px; outline: none;
}
.filter-bar input:focus { border-color: var(--accent); }

/* ===================== Column picker (v48) ===================== */
.col-picker {
  position: fixed; z-index: 1000; width: 300px; max-height: 60vh; overflow: auto;
  background: var(--card); border: 1px solid var(--border2); border-radius: 10px;
  box-shadow: 0 10px 30px rgba(0,0,0,0.35); padding: 10px 12px;
}
.col-picker .cp-head {
  display: flex; justify-content: space-between; align-items: center;
  gap: 8px; margin-bottom: 8px; font-size: 13px;
}
.col-picker .cp-actions { font-size: 12px; }
.col-picker .cp-actions a { color: var(--accent); cursor: pointer; }
.col-picker .cp-list { display: flex; flex-direction: column; gap: 4px; }
.col-picker label {
  display: flex; align-items: center; gap: 8px; font-size: 13px;
  padding: 3px 4px; border-radius: 6px; cursor: pointer;
}
.col-picker label:hover { background: var(--border); }

/* ===================== Empty state ===================== */
.empty-state {
  padding: 60px 20px; text-align: center; color: var(--muted);
}
.empty-state .empty-ico { font-size: 44px; margin-bottom: 14px; opacity: 0.35; }
.empty-state p { font-size: 13.5px; }

/* ===================== Settings ===================== */
.settings-row {
  display: flex; align-items: center; justify-content: space-between;
  padding: 12px 0; border-bottom: 1px solid var(--border);
}
.settings-row:last-child { border-bottom: none; }
.settings-key { font-size: 13px; color: var(--fg); }
.settings-desc { font-size: 11.5px; color: var(--muted); margin-top: 2px; }
.settings-val {
  font-size: 12.5px; color: var(--accent);
  font-family: monospace; text-align: right;
}
.engine-status {
  display: flex; align-items: center; gap: 8px; padding: 8px 0;
  border-bottom: 1px solid var(--border); font-size: 13px;
}
.engine-status:last-child { border-bottom: none; }

/* ===================== Toast notifications ===================== */
#toast-container {
  position: fixed; top: 16px; right: 16px;
  display: flex; flex-direction: column; gap: 8px;
  z-index: 9999; pointer-events: none;
}
.toast {
  background: var(--card2); border: 1px solid var(--border2);
  border-radius: 10px; padding: 12px 16px;
  font-size: 13px; color: var(--fg);
  display: flex; align-items: center; gap: 10px;
  box-shadow: 0 8px 24px rgba(0,0,0,0.4);
  animation: toastIn 0.25s ease;
  min-width: 240px; max-width: 360px;
  pointer-events: all;
}
.toast.removing { animation: toastOut 0.2s ease forwards; }
.toast-ico { font-size: 16px; flex-shrink: 0; }
.toast.toast-ok   { border-left: 3px solid var(--success); }
.toast.toast-err  { border-left: 3px solid var(--error); }
.toast.toast-info { border-left: 3px solid var(--accent); }
@keyframes toastIn { from { opacity:0; transform: translateX(20px); } to { opacity:1; transform:none; } }
@keyframes toastOut { to { opacity:0; transform: translateX(20px); } }

/* ===================== Misc ===================== */
.divider { height: 1px; background: var(--border); margin: 16px 0; }
.text-muted { color: var(--muted); }
.text-sm { font-size: 12px; }
.ml-auto { margin-left: auto; }
hr.section-sep { border: none; border-top: 1px solid var(--border); margin: 0; }

/* ===================== v27.9.x: Polish (additive) ===================== */
/* Дополняющие правила в конце каскада: глубина, мягкие микро-взаимодействия,
   доступный фокус. Не переопределяют структуру — только визуальный лоск. */
:root { --shadow-1: 0 1px 2px rgba(0,0,0,0.18), 0 4px 16px rgba(0,0,0,0.22);
        --shadow-2: 0 6px 24px rgba(0,0,0,0.30);
        --ring: 0 0 0 3px rgba(91,140,255,0.35); }
[data-theme="light"] { --shadow-1: 0 1px 2px rgba(16,24,40,0.06), 0 4px 14px rgba(16,24,40,0.08);
        --shadow-2: 0 10px 28px rgba(16,24,40,0.12); }

.card, .kpi-card, .chart-card {
  box-shadow: var(--shadow-1);
  transition: box-shadow 0.2s ease, transform 0.2s ease, border-color 0.2s ease;
  will-change: transform;
}
.kpi-card:hover, .chart-card:hover {
  transform: translateY(-2px);
  box-shadow: var(--shadow-2);
  border-color: var(--border2);
}

/* Главная кнопка — мягкий градиент и выразительный hover */
.btn-primary {
  background-image: linear-gradient(135deg, var(--accent), var(--accent2));
}
.btn-primary:hover:not(:disabled) {
  background-image: linear-gradient(135deg, var(--accent), var(--accent3, var(--accent2)));
  transform: translateY(-1px);
}
.btn-lg { box-shadow: 0 4px 14px rgba(91,140,255,0.35); }

/* Доступный фокус по клавиатуре (не мешает мыши) */
.btn:focus-visible, .nav-btn:focus-visible, .seg-btn:focus-visible,
.field input:focus-visible, .field select:focus-visible {
  outline: none; box-shadow: var(--ring);
}

/* Активный пункт меню — аккуратная акцентная полоса слева */
.nav-btn { position: relative; }
.nav-btn.active::before {
  content: ''; position: absolute; left: 0; top: 8px; bottom: 8px;
  width: 3px; border-radius: 0 3px 3px 0; background: var(--accent);
}

/* Числа KPI — моноширинные цифры, чтобы не «прыгали» при обновлении */
.kpi-value { font-variant-numeric: tabular-nums; }

/* Чуть живее переходы экранов */
.screen.active { animation: fadeUp 0.22s cubic-bezier(0.16,1,0.3,1); }

/* v27.9.x: мультиселект категорий бренда */
.ms-group { display:flex; flex-wrap:wrap; gap:6px 14px; padding:6px 0; }
.ms-item { display:flex; align-items:center; gap:6px; font-size:13px; color:var(--text); cursor:pointer; white-space:nowrap; }
.ms-item input { width:15px; height:15px; cursor:pointer; }

/* v27.9.x: сортируемые заголовки + кликабельные длинные ячейки */
th.sortable { cursor:pointer; user-select:none; }
th.sortable:hover { color:#fff; background:rgba(91,140,255,0.12); }
td.cell-expand { cursor:pointer; text-decoration:underline dotted rgba(255,255,255,0.25); }
td.cell-expand:hover { color:#fff; background:rgba(91,140,255,0.10); }
td.cell-link { cursor:pointer; color:#5b8cff; white-space:nowrap; }
td.cell-link:hover { color:#88aaff; background:rgba(91,140,255,0.12); text-decoration:underline; }

/* v27.9.x: модалка полного текста названия */
#full-text-modal { display:none; position:fixed; inset:0; z-index:9999; }
#full-text-modal .ftm-backdrop { position:absolute; inset:0; background:rgba(0,0,0,0.55); }
#full-text-modal .ftm-box { position:relative; max-width:680px; margin:12vh auto 0; background:var(--card,#1b2030); border:1px solid rgba(255,255,255,0.12); border-radius:12px; padding:20px; box-shadow:0 20px 60px rgba(0,0,0,0.5); }
#full-text-modal .ftm-text { color:var(--text,#e8ecf3); font-size:14px; line-height:1.5; white-space:pre-wrap; max-height:60vh; overflow:auto; }
#full-text-modal .ftm-close { margin-top:16px; padding:8px 18px; border-radius:8px; border:none; background:#5b8cff; color:#fff; cursor:pointer; font-size:13px; }
</style>
<!--__CHARTJS_INLINE__-->
</head>
<body>
<div id="toast-container"></div>
<div class="app">

  <!-- ========== SIDEBAR ========== -->
  <aside class="sidebar">
    <div class="sidebar-logo">
      <div class="brand-name">WB+Ozon Checker</div>
      <div class="brand-sub" id="ver-sub">v27.6 · ozon-playwright</div>
    </div>
    <nav class="nav">
      <button class="nav-btn active" data-screen="run">
        <span class="nav-ico">🚀</span>
        <span class="nav-label">Запуск</span>
      </button>
      <button class="nav-btn" data-screen="queue">
        <span class="nav-ico">📋</span>
        <span class="nav-label">Очередь</span>
        <span class="nav-badge" id="queue-badge">●</span>
      </button>
      <button class="nav-btn" data-screen="results">
        <span class="nav-ico">📊</span>
        <span class="nav-label">Результаты</span>
      </button>
      <button class="nav-btn" data-screen="settings">
        <span class="nav-ico">⚙️</span>
        <span class="nav-label">Настройки</span>
      </button>
      <button class="nav-btn" data-screen="logs">
        <span class="nav-ico">📜</span>
        <span class="nav-label">Логи</span>
      </button>
    </nav>
    <div class="sidebar-footer">
      <div id="foot-py">Python …</div>
      <div id="foot-engines" style="margin-top:5px; display:flex; flex-direction:column; gap:3px;"></div>
      <div id="foot-dir" style="margin-top:5px; opacity:0.6; word-break:break-all;"></div>
    </div>
  </aside>

  <!-- ========== MAIN ========== -->
  <main class="main">
    <!-- Topbar -->
    <div class="topbar">
      <div class="topbar-left">
        <span class="topbar-title" id="screen-title">Запуск</span>
      </div>
      <div class="topbar-right">
        <button class="btn btn-sm" id="btn-diagnose" style="margin-right:8px;" title="Диагностика среды">Диагностика</button>
        <div class="status-indicator">
          <div class="status-dot" id="status-dot"></div>
          <span id="status-text">готов</span>
        </div>
      </div>
    </div>
    <div id="deps-banner" style="display:none; background:#3b1d1d; color:#ffb3b3; padding:10px 16px; font-size:13px; border-bottom:1px solid #5a2a2a;">
      <strong>Внимание:</strong> <span id="deps-banner-text"></span>
    </div>

    <div class="content" id="main-content">

      <!-- ========================= RUN SCREEN ========================= -->
      <section class="screen active" id="screen-run">

        <!-- Marketplace selector -->
        <div class="seg-ctrl" id="mkt-sel">
          <button class="seg-btn active" data-mkt="wb">
            <span style="color:var(--wb)">●</span> Wildberries
          </button>
          <button class="seg-btn" data-mkt="ozon">
            <span style="color:var(--ozon)">●</span> Ozon
          </button>
          <button class="seg-btn" data-mkt="both">
            ⚡ Оба
            <span class="mkt-tag">unified</span>
          </button>
        </div>

        <!-- WB mode tabs — скрывается при ozon/both -->
        <div id="wb-mode-row" class="seg-ctrl" style="margin-left:12px; margin-bottom:16px;">
          <button class="seg-btn active" data-mode="query_full">По запросу (полный)</button>
          <button class="seg-btn" data-mode="query_auto">Без запроса (каталог)</button>
          <button class="seg-btn" data-mode="query_stage1">Только ссылки</button>
          <button class="seg-btn" data-mode="query_stage2">Только реестры (CSV)</button>
          <button class="seg-btn" data-mode="brand">По бренду</button>
        </div>

        <!-- Form card -->
        <div class="card" id="launch-card">
          <div class="card-title" id="form-card-title">
            🔍 По запросу — полный прогон
            <span class="ctag" id="form-mode-tag">WB</span>
          </div>
          <div class="form-grid cols-2" id="form-grid">
            <!-- поля рендерит JS -->
          </div>

          <!-- Enhancements row -->
          <div class="divider"></div>
          <div class="form-grid cols-2" style="margin-top: 0;">
            <div>
              <label class="toggle-wrap">
                <input type="checkbox" id="use-wb-enhanced">
                <div class="toggle-track"></div>
                <div>
                  <div class="toggle-label">wb_enhanced: улучшенный seller_name + «Оригинал»</div>
                  <div class="toggle-sub">Требует wb_enhanced.py</div>
                </div>
              </label>
            </div>
            <div>
              <label class="toggle-wrap">
                <input type="checkbox" id="use-fsa-enhanced">
                <div class="toggle-track"></div>
                <div>
                  <div class="toggle-label">fsa_enhanced: расширенные поля ФСА (54 поля)</div>
                  <div class="toggle-sub">Требует fsa_enhanced.py</div>
                </div>
              </label>
            </div>
          </div>

          <div class="actions-row">
            <button class="btn btn-primary btn-lg" id="btn-run">
              🚀 Запустить
            </button>
            <button class="btn btn-danger" id="btn-stop" disabled>
              ⏹ Остановить
            </button>
            <button class="btn btn-ghost" id="btn-open-folder">
              📁 Папка
            </button>
            <span class="text-muted text-sm ml-auto">Параметры сохраняются автоматически</span>
          </div>
        </div>

        <!-- Last run card -->
        <div class="card" id="last-run-card" style="display:none;">
          <div class="card-title">📂 Последний прогон</div>
          <div id="last-run-info" style="font-size:13px; color:var(--fg2);"></div>
          <div class="actions-row" style="margin-top:12px;">
            <button class="btn btn-ghost btn-sm" id="btn-open-last-xlsx">📄 Открыть XLSX</button>
            <button class="btn btn-ghost btn-sm" id="btn-view-last-results">📊 Таблица</button>
          </div>
        </div>
      </section>


      <!-- ========================= QUEUE SCREEN ========================= -->
      <section class="screen" id="screen-queue">
        <!-- KPI row -->
        <div class="kpi-grid">
          <div class="kpi-card" style="--kpi-color:var(--accent)">
            <div class="kpi-label">Прогресс</div>
            <div class="kpi-value" id="kpi-pct">0%</div>
            <div class="kpi-sub" id="kpi-pct-sub">— из —</div>
          </div>
          <div class="kpi-card" style="--kpi-color:var(--fg)">
            <div class="kpi-label">Прошло</div>
            <div class="kpi-value" id="kpi-elapsed">—</div>
            <div class="kpi-sub" id="kpi-mode-lbl">—</div>
          </div>
          <div class="kpi-card" style="--kpi-color:var(--success)">
            <div class="kpi-label">Скорость</div>
            <div class="kpi-value" id="kpi-speed">—</div>
            <div class="kpi-sub">строк/мин</div>
          </div>
          <div class="kpi-card" style="--kpi-color:var(--info)">
            <div class="kpi-label">Осталось (ETA)</div>
            <div class="kpi-value" id="kpi-eta">—</div>
            <div class="kpi-sub">оценка времени</div>
          </div>
          <div class="kpi-card" style="--kpi-color:var(--error)">
            <div class="kpi-label">Ошибок</div>
            <div class="kpi-value" id="kpi-errors">0</div>
            <div class="kpi-sub" id="kpi-ok-sub">OK: 0</div>
          </div>
        </div>

        <!-- Progress card -->
        <div class="card">
          <div class="card-title">📈 Прогресс выполнения</div>
          <div class="progress-wrap">
            <div class="progress-bar-bg">
              <div class="progress-bar-fill" id="prog-bar"></div>
            </div>
            <div class="progress-info">
              <span id="prog-label">ожидание</span>
              <span id="prog-stage" style="color:var(--accent);"></span>
            </div>
          </div>
          <!-- Стек-бар по статусам (визуальный расклад результата) -->
          <div id="status-stack" style="display:flex; height:14px; width:100%; border-radius:4px; overflow:hidden; margin-top:14px; background:rgba(255,255,255,0.04);">
          </div>
          <div id="status-stack-legend" style="display:flex; flex-wrap:wrap; gap:14px; margin-top:10px; font-size:12px; color:var(--fg2);"></div>
          <div class="actions-row" style="margin-top:14px;">
            <button class="btn btn-ghost btn-sm" id="btn-open-result" disabled>📄 Открыть XLSX</button>
            <button class="btn btn-ghost btn-sm" id="btn-open-ozon-result" disabled>🛒 Ozon XLSX</button>
            <button class="btn btn-ghost btn-sm" id="btn-open-log" disabled>📝 Лог файл</button>
            <button class="btn btn-ghost btn-sm" id="btn-goto-results">📊 Таблица</button>
            <button class="btn btn-ghost btn-sm" id="btn-retry-fsa" title="Перезапустить этап 2 по упавшим FSA-ссылкам (когда FSA снова доступен)">🔁 Повторить упавшие FSA</button>
          </div>
        </div>

        <!-- v27.6: Live-лог (последние ~30 строк stdout) -->
        <div class="card">
          <div class="card-title" style="display:flex; justify-content:space-between; align-items:center;">
            <span>📝 Живой лог</span>
            <span style="font-size:11px; color:var(--fg2); font-weight:normal;" id="livelog-total">0 строк</span>
          </div>
          <div id="livelog-box" style="max-height:240px; overflow-y:auto; background:rgba(0,0,0,0.25); border-radius:6px; padding:10px 12px; font-family:'JetBrains Mono', Consolas, monospace; font-size:11.5px; line-height:1.45; color:#cfd6e0; white-space:pre-wrap; word-break:break-word;">Ожидание запуска…</div>
        </div>

        <!-- Charts row -->
        <div class="charts-row">
          <div class="chart-card">
            <div class="chart-title">Технический статус</div>
            <div class="chart-canvas-wrap">
              <canvas id="chart-status"></canvas>
              <div class="chart-fallback" id="chart-status-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Реестры</div>
            <div class="chart-canvas-wrap">
              <canvas id="chart-registry"></canvas>
              <div class="chart-fallback" id="chart-registry-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Маркетплейсы</div>
            <div class="chart-canvas-wrap">
              <canvas id="chart-marketplace"></canvas>
              <div class="chart-fallback" id="chart-marketplace-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Оригинальность</div>
            <div class="chart-canvas-wrap">
              <canvas id="chart-original"></canvas>
              <div class="chart-fallback" id="chart-original-fb" style="display:none;">нет данных</div>
            </div>
          </div>
        </div>
      </section>


      <!-- ========================= RESULTS SCREEN ========================= -->
      <section class="screen" id="screen-results">
        <!-- Charts -->
        <div class="charts-row" id="results-charts-row">
          <div class="chart-card">
            <div class="chart-title">Технический статус</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-status"></canvas>
              <div class="chart-fallback" id="res-chart-status-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Реестры</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-registry"></canvas>
              <div class="chart-fallback" id="res-chart-registry-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Маркетплейсы</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-marketplace"></canvas>
              <div class="chart-fallback" id="res-chart-marketplace-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Оригинальность</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-original"></canvas>
              <div class="chart-fallback" id="res-chart-original-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Риск по сроку</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-risk"></canvas>
              <div class="chart-fallback" id="res-chart-risk-fb" style="display:none;">нет данных</div>
            </div>
          </div>
          <div class="chart-card">
            <div class="chart-title">Топ брендов</div>
            <div class="chart-canvas-wrap">
              <canvas id="res-chart-brand"></canvas>
              <div class="chart-fallback" id="res-chart-brand-fb" style="display:none;">нет данных</div>
            </div>
          </div>
        </div>

        <!-- Filter + table -->
        <div class="filter-bar">
          <input type="text" id="filter-input"
                 placeholder="🔍  Поиск по названию, бренду, артикулу, статусу…">
          <select id="filter-status" class="btn btn-ghost btn-sm" style="padding:8px 12px;">
            <option value="">Все статусы</option>
          </select>
          <button class="btn btn-ghost btn-sm" id="btn-reload-results">⟳ Обновить</button>
          <button class="btn btn-ghost btn-sm" id="btn-load-result">📂 Загрузить файл</button>
          <button class="btn btn-ghost btn-sm" id="btn-clear-loaded" style="display:none">✖ Текущий прогон</button>
          <button class="btn btn-ghost btn-sm" id="btn-export-csv">⬇ CSV</button>
          <button class="btn btn-ghost btn-sm" id="btn-export-disputed" title="Выгрузить только спорные строки (ПРОВЕРИТЬ ВРУЧНУЮ + НЕСООТВЕТСТВИЕ) со ссылками на карточку и реестр">⬇ На проверку</button>
          <button class="btn btn-ghost btn-sm" id="btn-supplier-stats" title="Сводка по продавцам/брендам: у кого больше всего несоответствий и спорных">⬇ По продавцам</button>
          <button class="btn btn-ghost btn-sm" id="btn-suggest-dict" title="Собрать частые незнакомые слова из товаров без категории — кандидаты для пополнения словаря">🧠 Обучить словарь</button>
          <button class="btn btn-ghost btn-sm" id="btn-columns">⚙ Колонки</button>
          <span class="text-muted text-sm" id="results-count"></span>
          <span class="text-muted text-sm" id="loaded-file-badge" style="display:none"></span>
        </div>

        <div id="review-reasons" style="display:none; margin:0 0 10px; padding:8px 12px; background:rgba(245,158,11,0.08); border:1px solid rgba(245,158,11,0.25); border-radius:8px; font-size:12.5px; color:var(--fg2);"></div>

        <div class="tbl-wrap" id="tbl-wrap">
          <div class="empty-state">
            <div class="empty-ico">📋</div>
            <p>Результатов пока нет.<br>Запустите прогон и нажмите «Результаты».</p>
          </div>
        </div>
      </section>


      <!-- ========================= SETTINGS SCREEN ========================= -->
      <section class="screen" id="screen-settings">
        <div class="card">
          <div class="card-title">🔖 О программе</div>
          <div id="settings-version-info" style="font-size:13px; color:var(--fg2); line-height:1.7;"></div>
        </div>

        <div class="card">
          <div class="card-title">⚙️ Движки</div>
          <div id="settings-engines"></div>
        </div>

        <div class="card">
          <div class="card-title">🎛 Параметры по умолчанию</div>
          <div class="form-grid cols-2">
            <div class="field">
              <span class="field-label">expiry_warning_days</span>
              <input type="number" id="s-expiry" value="30" min="1" max="365">
              <span class="field-hint">Порог «Скоро истекает» (дней)</span>
            </div>
            <div class="field">
              <span class="field-label">Воркеры по умолчанию</span>
              <input type="number" id="s-workers" value="4" min="1" max="20">
              <span class="field-hint">Параллельных браузеров</span>
            </div>
          </div>
          <div class="actions-row">
            <button class="btn btn-primary btn-sm" id="btn-save-defaults">💾 Сохранить</button>
          </div>
        </div>

        <div class="card">
          <div class="card-title">🇰🇬 Статусы киргизских документов в РФ</div>
          <div style="font-size:13px; color:var(--fg2); line-height:1.6; margin-bottom:10px;">
            Загрузите таблицу (xlsx/csv) с колонками <b>number</b> и <b>id_status_in_rf</b>
            (14 = прекращён, 15 = приостановлен). Совпавшие по номеру киргизские документы
            получат колонку «Статус на территории РФ» и итоговый вердикт
            «<b style="color:#e06666;">НЕДЕЙСТВУЕТ В РФ</b>».
          </div>
          <div class="actions-row">
            <button class="btn btn-ghost btn-sm" id="btn-load-kg">📋 Загрузить таблицу КГ-статусов</button>
            <span class="text-muted text-sm" id="kg-status-info"></span>
          </div>
        </div>

        <div class="card">
          <div class="card-title">🗂 Рабочая папка</div>
          <div id="settings-dir" style="font-family:monospace; font-size:12.5px; color:var(--accent); word-break:break-all;"></div>
          <div class="actions-row" style="margin-top:12px;">
            <button class="btn btn-ghost btn-sm" id="btn-open-folder3">📁 Открыть</button>
          </div>
        </div>

        <div class="card">
          <div class="card-title">🌓 Тема интерфейса</div>
          <div style="display:flex; gap:10px; flex-wrap:wrap;">
            <button class="btn btn-ghost btn-sm theme-btn active" data-theme="dark">🌙 Тёмная</button>
            <button class="btn btn-ghost btn-sm theme-btn" data-theme="light">☀️ Светлая</button>
          </div>
        </div>
      </section>


      <!-- ========================= LOGS SCREEN ========================= -->
      <section class="screen" id="screen-logs">
        <div class="log-toolbar">
          <div class="log-filter-btns">
            <button class="log-filter-btn active" data-lvl="all">Все</button>
            <button class="log-filter-btn" data-lvl="error">Ошибки</button>
            <button class="log-filter-btn" data-lvl="warn">Предупреждения</button>
            <button class="log-filter-btn" data-lvl="ok">OK</button>
          </div>
          <span class="text-muted text-sm ml-auto" id="log-info"></span>
          <button class="btn btn-ghost btn-sm" id="btn-clear-log">🧹 Очистить</button>
          <button class="btn btn-ghost btn-sm" id="btn-copy-log">⧉ Копировать</button>
          <button class="btn btn-ghost btn-sm" id="btn-save-log">💾 Сохранить</button>
        </div>
        <div class="log-view" id="log-view">
          <span style="color:var(--muted);">— журнал пуст —</span>
        </div>
      </section>

    </div><!-- /content -->
  </main>
</div><!-- /app -->

<script>
// ============================================================
// Глобальные переменные
// ============================================================
let _currentMkt    = 'wb';       // wb | ozon | both
let _currentMode   = 'query_full'; // query_full | query_stage1 | query_stage2 | brand | ozon | unified
let _logSeen       = 0;
let _pollInterval  = null;
let _cachedSettings = {};
let _allRows       = [];
let _allHeaders    = [];
let _logFilterLvl  = 'all';
let _chartjs       = null;       // Chart.js объект после загрузки
let _charts        = {};         // храним Chart-инстансы
let _activityCtx   = null;

// ============================================================
// Утилиты
// ============================================================
const $ = s => document.querySelector(s);
const $$ = s => document.querySelectorAll(s);

function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'})[c]);
}

function fmtElapsed(secs) {
  secs = Math.floor(Math.max(0, secs));
  const h = Math.floor(secs / 3600);
  const m = Math.floor((secs % 3600) / 60);
  const s = secs % 60;
  if (h) return `${h}ч ${String(m).padStart(2,'0')}м`;
  if (m) return `${m}м ${String(s).padStart(2,'0')}с`;
  return `${s}с`;
}

function toast(msg, type = 'info') {
  const c = $('#toast-container');
  const el = document.createElement('div');
  const icons = { ok: '✅', err: '❌', info: 'ℹ️' };
  el.className = `toast toast-${type}`;
  el.innerHTML = `<span class="toast-ico">${icons[type] || '💬'}</span><span>${escapeHtml(msg)}</span>`;
  c.appendChild(el);
  setTimeout(() => {
    el.classList.add('removing');
    setTimeout(() => el.remove(), 250);
  }, 3000);
}

// ============================================================
// pywebview ready
// ============================================================
function waitForApi() {
  return new Promise(resolve => {
    if (window.pywebview && window.pywebview.api) return resolve();
    window.addEventListener('pywebviewready', () => resolve(), { once: true });
    const t = setInterval(() => {
      if (window.pywebview && window.pywebview.api) { clearInterval(t); resolve(); }
    }, 40);
  });
}

// ============================================================
// Навигация
// ============================================================
const SCREEN_TITLES = {
  run: '🚀 Запуск', queue: '📋 Очередь',
  results: '📊 Результаты', settings: '⚙️ Настройки', logs: '📜 Логи'
};

function go(name) {
  $$('.screen').forEach(s => s.classList.remove('active'));
  $$('.nav-btn').forEach(b => b.classList.remove('active'));
  const sc = $('#screen-' + name);
  if (sc) sc.classList.add('active');
  const nb = $(`.nav-btn[data-screen="${name}"]`);
  if (nb) nb.classList.add('active');
  $('#screen-title').textContent = SCREEN_TITLES[name] || name;
  if (name === 'results') loadResults();
  if (name === 'logs') renderLogFull();
  if (name === 'settings') fillSettings();
  if (name === 'queue') tryInitCharts();
}

$$('.nav-btn').forEach(b => b.addEventListener('click', () => go(b.dataset.screen)));

// ============================================================
// Маркетплейс и режим
// ============================================================
function setMkt(mkt) {
  _currentMkt = mkt;
  $$('#mkt-sel .seg-btn').forEach(b => b.classList.toggle('active', b.dataset.mkt === mkt));
  const wbModes = $('#wb-mode-row');
  if (mkt === 'wb') {
    wbModes.style.display = '';
    _currentMode = _currentMode === 'ozon' || _currentMode === 'unified' ? 'query_full' : _currentMode;
  } else if (mkt === 'ozon') {
    wbModes.style.display = 'none';
    _currentMode = 'ozon';
  } else { // both
    wbModes.style.display = 'none';
    _currentMode = 'unified';
  }
  setWbMode(_currentMode);
  renderForm();
}

function setWbMode(mode) {
  _currentMode = mode;
  $$('#wb-mode-row .seg-btn').forEach(b => b.classList.toggle('active', b.dataset.mode === mode));
  renderForm();
}

$$('#mkt-sel .seg-btn').forEach(b => b.addEventListener('click', () => setMkt(b.dataset.mkt)));
$$('#wb-mode-row .seg-btn').forEach(b => b.addEventListener('click', () => setWbMode(b.dataset.mode)));

// ============================================================
// Определения полей формы
// ============================================================
const FORM_FIELDS = {
  query_full: [
    {key:'query',   lbl:'Поисковый запрос',    type:'text',   def:'детская обувь', hint:'Например: «детская обувь»'},
    {key:'brand_category', lbl:'Категория товаров (уточнить запрос)', type:'multiselect', def:'',
      options:['одежда','обувь','бытовая техника','электроника','игрушки','косметика','детские аксессуары','детский транспорт','дом и текстиль','посуда','продукты'],
      hint:'Необязательно. Сужает поиск до выбранных категорий — запрос к WB точнее. Ничего не выбрано — авто-определение'},
    {key:'limit',   lbl:'Лимит карточек',       type:'number', def:5000, min:1, max:200000},
    {key:'workers', lbl:'Браузер-воркеры',       type:'number', def:5, min:1, max:12, hint:'Параллельных браузеров для парсинга реестров (4–6 оптимально; больше = быстрее FSA, но больше памяти)'},
    {key:'expiry_warning_days', lbl:'Скоро истекает (дней)', type:'number', def:30, min:1, max:365},
    {key:'headless',        lbl:'Скрытый браузер',    type:'switch', def:true},
    {key:'make_report_xlsx',lbl:'Расширенный отчёт',   type:'switch', def:true, hint:'Листы «Сводка» + «Подробности»'},
    {key:'fsa_slow_mode',   lbl:'Медленный режим ФСА (без блокировок)', type:'switch', def:false,
      hint:'Для больших прогонов: ФСА по одному документу с паузой — IP не банится, но медленно (часы для тысяч ссылок). SWIS/прочие реестры идут параллельно.'},
    {key:'fsa_slow_delay_sec', lbl:'Пауза ФСА, сек (медл. режим)', type:'number', def:0, min:0, max:30,
      hint:'0 = авто (~2.5–5с). Меньше — быстрее, но выше риск бана; при блокировках пауза сама растёт.'},
  ],
  query_auto: [
    {key:'limit',   lbl:'Сколько карточек собрать',  type:'number', def:50000, min:1, max:500000,
      hint:'Программа сама подберёт запросы и сметёт каталог WB до этого числа карточек'},
    {key:'catalog_categories', lbl:'Категории (необязательно)', type:'multiselect', def:'',
      options:['одежда','обувь','бытовая техника','электроника','игрушки','косметика','детские аксессуары','детский транспорт','дом и текстиль','посуда','продукты'],
      hint:'Ничего не выбрано — сметаются ВСЕ категории. Выбор сужает сбор до них.'},
    {key:'workers', lbl:'Браузер-воркеры',       type:'number', def:5, min:1, max:12, hint:'Параллельных браузеров для парсинга реестров (4–6 оптимально)'},
    {key:'expiry_warning_days', lbl:'Скоро истекает (дней)', type:'number', def:30, min:1, max:365},
    {key:'headless',        lbl:'Скрытый браузер',    type:'switch', def:true},
    {key:'make_report_xlsx',lbl:'Расширенный отчёт',   type:'switch', def:true, hint:'Листы «Сводка» + «Подробности»'},
    {key:'fsa_slow_mode',   lbl:'Медленный режим ФСА (без блокировок)', type:'switch', def:true,
      hint:'Для больших прогонов 50k+ рекомендуется ВКЛ: ФСА по одному документу с паузой — IP не банится. SWIS/прочие идут параллельно.'},
    {key:'fsa_slow_delay_sec', lbl:'Пауза ФСА, сек (медл. режим)', type:'number', def:0, min:0, max:30,
      hint:'0 = авто (~2.5–5с). При блокировках пауза сама растёт.'},
  ],
  query_stage1: [
    {key:'query',            lbl:'Поисковый запрос',  type:'text',   def:'детская обувь'},
    {key:'limit',            lbl:'Лимит карточек',    type:'number', def:10000, min:1, max:200000},
    {key:'workers',          lbl:'HTTP-воркеры',      type:'number', def:3, min:1, max:20, hint:'×10 в команде'},
    {key:'output_links_csv', lbl:'Ссылки CSV',        type:'text',   def:'registry_links.csv'},
    {key:'output',           lbl:'Карточки XLSX',     type:'text',   def:'links.xlsx'},
  ],
  query_stage2: [
    {key:'input_links_csv',  lbl:'Входной CSV ссылок', type:'text',  def:'registry_links.csv'},
    {key:'output',           lbl:'Результат XLSX',     type:'text',  def:'result.xlsx'},
    {key:'limit',            lbl:'Лимит',              type:'number',def:10000, min:1, max:200000},
    {key:'workers',          lbl:'Браузер-воркеры',    type:'number',def:5, min:1, max:10, hint:'Параллельных браузеров (4–6 оптимально)'},
    {key:'expiry_warning_days',lbl:'Скоро истекает (дней)',type:'number',def:30,min:1,max:365},
    {key:'headless',         lbl:'Скрытый браузер',   type:'switch', def:true},
    {key:'make_report_xlsx', lbl:'Расширенный отчёт', type:'switch', def:true},
    {key:'fsa_slow_mode',    lbl:'Медленный режим ФСА (без блокировок)', type:'switch', def:false,
      hint:'ФСА по одному документу с паузой — IP не банится на больших прогонах, но медленно. SWIS/прочие параллельно.'},
    {key:'fsa_slow_delay_sec', lbl:'Пауза ФСА, сек (медл. режим)', type:'number', def:0, min:0, max:30,
      hint:'0 = авто (~2.5–5с). Меньше — быстрее, но выше риск бана.'},
    {key:'strict_brand',     lbl:'Строгий бренд',     type:'text',  def:'', hint:'Опционально'},
    {key:'strict_brand_match',lbl:'Тип совпадения',   type:'select',def:'any', options:['any','exact','contains']},
  ],
  brand: [
    {key:'brand',       lbl:'Бренд',                   type:'text',  def:'adidas', hint:'Латиницей, как на WB'},
    {key:'brand_category', lbl:'Категории товаров (можно несколько)', type:'multiselect', def:'',
      options:['одежда','обувь','бытовая техника','электроника','игрушки','косметика','детские аксессуары','детский транспорт','дом и текстиль','посуда','продукты'],
      hint:'Сузить поиск до выбранных категорий (reebok→одежда+обувь, indesit→бытовая техника). Ничего не выбрано — все товары бренда'},
    {key:'brand_match', lbl:'Тип совпадения',           type:'select',def:'exact',options:['exact','contains','any']},
    {key:'limit',       lbl:'Лимит карточек',           type:'number',def:5000, min:1, max:200000},
    {key:'workers',     lbl:'Браузер-воркеры',          type:'number',def:5, min:1, max:12, hint:'Параллельных браузеров для реестров (4–6 оптимально)'},
    {key:'expiry_warning_days',lbl:'Скоро истекает (дней)',type:'number',def:30,min:1,max:365},
    {key:'output',      lbl:'Результат XLSX',           type:'text',  def:'brand_result.xlsx'},
    {key:'make_report_xlsx',lbl:'Расширенный отчёт',    type:'switch',def:true},
    {key:'fsa_slow_mode',   lbl:'Медленный режим ФСА (без блокировок)', type:'switch', def:false,
      hint:'ФСА по одному документу с паузой — IP не банится на больших прогонах, но медленно. SWIS/прочие параллельно.'},
    {key:'fsa_slow_delay_sec', lbl:'Пауза ФСА, сек (медл. режим)', type:'number', def:0, min:0, max:30,
      hint:'0 = авто (~2.5–5с). Меньше — быстрее, но выше риск бана.'},
  ],
  ozon: [
    {key:'query',   lbl:'Поисковый запрос Ozon', type:'text',  def:'детская обувь', hint:'Как в поиске Ozon.ru'},
    {key:'limit',   lbl:'Лимит товаров',         type:'number',def:1000, min:1, max:50000},
    {key:'workers', lbl:'Воркеры',               type:'number',def:10, min:1, max:30, hint:'Параллельных потоков (5–15)'},
    {key:'expiry_warning_days',lbl:'Скоро истекает (дней)',type:'number',def:30,min:1,max:365},
    {key:'headless',lbl:'Скрытый браузер',       type:'switch',def:true},
    {key:'make_report_xlsx',lbl:'Расширенный отчёт',type:'switch',def:true},
    {key:'output',  lbl:'Результат XLSX',        type:'text',  def:'ozon_result.xlsx'},
    {key:'ozon_delay_min_ms',lbl:'Мин. задержка (мс)',type:'number',def:200,min:50,max:2000},
    {key:'ozon_delay_max_ms',lbl:'Макс. задержка (мс)',type:'number',def:500,min:100,max:5000},
  ],
  unified: [
    {key:'query',   lbl:'Поисковый запрос',   type:'text',  def:'детская обувь', hint:'Одновременно WB + Ozon'},
    {key:'limit',   lbl:'Лимит карточек',     type:'number',def:3000, min:1, max:100000},
    {key:'workers', lbl:'Воркеры',            type:'number',def:4, min:1, max:12},
    {key:'expiry_warning_days',lbl:'Скоро истекает (дней)',type:'number',def:30,min:1,max:365},
    {key:'headless',lbl:'Скрытый браузер',   type:'switch',def:true},
    {key:'make_report_xlsx',lbl:'Расширенный отчёт',type:'switch',def:true},
    {key:'unified_report',lbl:'Объединить отчёты WB+Ozon',type:'switch',def:true,hint:'Создать общий XLSX'},
    {key:'output',  lbl:'Результат XLSX',    type:'text',  def:'unified_result.xlsx'},
  ],
};

const MODE_LABELS = {
  query_full:   '🔍 По запросу — полный прогон',
  query_auto:   '🧭 Без запроса — сметание каталога WB',
  query_stage1: '🔗 Только сбор ссылок (Этап 1)',
  query_stage2: '📋 Только реестры из CSV (Этап 2)',
  brand:        '🏷️ По бренду WB',
  ozon:         '🛒 Ozon — поиск + реестры',
  unified:      '⚡ Unified — WB + Ozon параллельно',
};

const MKT_TAGS = {
  query_full: 'WB', query_auto: 'WB', query_stage1: 'WB', query_stage2: 'WB',
  brand: 'WB', ozon: 'Ozon', unified: 'WB+Ozon'
};

// ============================================================
// Рендер формы
// ============================================================
function renderForm() {
  const fields = FORM_FIELDS[_currentMode] || [];
  const saved  = _cachedSettings.last_spec || {};
  const grid   = $('#form-grid');
  grid.innerHTML = '';
  $('#form-card-title').innerHTML = `${MODE_LABELS[_currentMode] || _currentMode}
    <span class="ctag">${MKT_TAGS[_currentMode] || ''}</span>`;

  fields.forEach(f => {
    const val = (saved[f.key] !== undefined) ? saved[f.key] : f.def;
    const wrap = document.createElement('div');
    if (f.type === 'switch') {
      wrap.innerHTML = `
        <div style="padding:8px 0;">
          <label class="toggle-wrap">
            <input type="checkbox" data-key="${f.key}" ${val ? 'checked' : ''}>
            <div class="toggle-track"></div>
            <div>
              <div class="toggle-label">${escapeHtml(f.lbl)}</div>
              ${f.hint ? `<div class="toggle-sub">${escapeHtml(f.hint)}</div>` : ''}
            </div>
          </label>
        </div>`;
    } else if (f.type === 'select') {
      const opts = (f.options||[]).map(o =>
        `<option value="${o}" ${val===o?'selected':''}>${o}</option>`).join('');
      wrap.innerHTML = `
        <div class="field">
          <span class="field-label">${escapeHtml(f.lbl)}</span>
          <select data-key="${f.key}">${opts}</select>
          ${f.hint ? `<span class="field-hint">${escapeHtml(f.hint)}</span>` : ''}
        </div>`;
    } else if (f.type === 'textarea') {
      const safeVal = String(val).replace(/"/g, '&quot;');
      wrap.innerHTML = `
        <div class="field">
          <span class="field-label">${escapeHtml(f.lbl)}</span>
          <textarea data-key="${f.key}" rows="${f.rows||4}" placeholder="${escapeHtml(f.ph||'')}" style="width:100%;resize:vertical;font-family:monospace;font-size:12px;">${escapeHtml(String(val))}</textarea>
          ${f.hint ? `<span class="field-hint">${escapeHtml(f.hint)}</span>` : ''}
        </div>`;
    } else if (f.type === 'multiselect') {
      // v27.9.x: чекбоксы — можно выбрать НЕСКОЛЬКО значений. Хранится строкой
      // через запятую. Используется для выбора нескольких категорий бренда.
      const cur = String(val || '').split(',').map(s => s.trim()).filter(Boolean);
      const boxes = (f.options||[]).map(o =>
        `<label class="ms-item"><input type="checkbox" data-mskey="${f.key}" value="${escapeHtml(o)}" ${cur.includes(o)?'checked':''}> ${escapeHtml(o)}</label>`
      ).join('');
      wrap.innerHTML = `
        <div class="field">
          <span class="field-label">${escapeHtml(f.lbl)}</span>
          <div class="ms-group" data-mswrap="${f.key}">${boxes}</div>
          ${f.hint ? `<span class="field-hint">${escapeHtml(f.hint)}</span>` : ''}
        </div>`;
    } else {
      const t = f.type === 'number' ? 'number' : 'text';
      const minattr = f.min !== undefined ? `min="${f.min}"` : '';
      const maxattr = f.max !== undefined ? `max="${f.max}"` : '';
      const safeVal = String(val).replace(/"/g, '&quot;');
      wrap.innerHTML = `
        <div class="field">
          <span class="field-label">${escapeHtml(f.lbl)}</span>
          <input type="${t}" data-key="${f.key}" ${minattr} ${maxattr} value="${safeVal}">
          ${f.hint ? `<span class="field-hint">${escapeHtml(f.hint)}</span>` : ''}
        </div>`;
    }
    grid.appendChild(wrap);
  });
}

function collectSpec() {
  const spec = { mode: _currentMode };
  $$('#form-grid [data-key]').forEach(el => {
    const k = el.dataset.key;
    if (el.type === 'checkbox') spec[k] = el.checked;
    else if (el.type === 'number') spec[k] = Number(el.value) || 0;
    else spec[k] = el.value;
  });
  // v27.9.x: мультиселект (категории бренда) — собираем отмеченные в строку.
  $$('#form-grid [data-mswrap]').forEach(group => {
    const k = group.dataset.mswrap;
    const vals = Array.from(group.querySelectorAll('input[type=checkbox]:checked')).map(c => c.value);
    spec[k] = vals.join(',');
  });
  spec.use_wb_enhanced = $('#use-wb-enhanced').checked;
  spec.use_fsa_enhanced = $('#use-fsa-enhanced').checked;
  return spec;
}

// ============================================================
// Run / Stop
// ============================================================
async function startRun() {
  const spec = collectSpec();
  await window.pywebview.api.save_settings({ last_spec: spec });
  resetLiveLog();  // v27.6: сброс live-лога при новом прогоне
  const res = await window.pywebview.api.start_run(spec);
  if (!res.ok) {
    toast(res.error || 'Не удалось запустить', 'err');
    return;
  }
  toast('Прогон запущен!', 'ok');
  // Новый прогон — снимаем метку ранее загруженного внешнего файла.
  const _lb = $('#loaded-file-badge'); if (_lb) _lb.style.display = 'none';
  const _cb = $('#btn-clear-loaded'); if (_cb) _cb.style.display = 'none';
  go('queue');
}

async function stopRun() {
  await window.pywebview.api.stop_run();
  toast('Команда «Стоп» отправлена', 'info');
}

$('#btn-run').addEventListener('click', startRun);
$('#btn-stop').addEventListener('click', stopRun);
$('#btn-diagnose').addEventListener('click', async () => {
  const d = await window.pywebview.api.diagnose();
  const eng = Object.entries(d.engines).map(([k, v]) => `  ${v ? '✓' : '✗'} ${k}`).join('\n');
  const deps = d.missing_deps.length ? 'ОТСУТСТВУЮТ: ' + d.missing_deps.join(', ') : 'все на месте';
  const msg = [
    `Версия: ${d.app_version}`,
    `Python: ${d.python_version}`,
    `Платформа: ${d.platform}`,
    `Папка: ${d.app_dir}`,
    ``,
    `Движки:`,
    eng,
    ``,
    `Зависимости: ${deps}`,
    ``,
    `Последняя команда:`,
    d.last_cmd || 'ещё не запускалась',
    ``,
    `Последняя ошибка:`,
    d.last_error || 'нет',
  ].join('\n');
  alert(msg);
});
// Баннер о недостающих зависимостях
(async () => {
  try {
    const d = await window.pywebview.api.diagnose();
    if (d.missing_deps && d.missing_deps.length) {
      const b = $('#deps-banner');
      $('#deps-banner-text').textContent =
        'не установлены пакеты: ' + d.missing_deps.join(', ') +
        ' — запустите install_windows.bat или pip install -U ' + d.missing_deps.join(' ');
      b.style.display = 'block';
    }
  } catch (e) {}
})();
$('#btn-open-folder').addEventListener('click', () => window.pywebview.api.open_workspace());
$('#btn-open-folder3').addEventListener('click', () => window.pywebview.api.open_workspace());
$('#btn-goto-results').addEventListener('click', () => go('results'));
$('#btn-view-last-results').addEventListener('click', () => go('results'));

$('#btn-open-result').addEventListener('click', async () => {
  const s = await window.pywebview.api.get_state();
  if (s.output_path) window.pywebview.api.open_path(s.output_path);
});
$('#btn-open-ozon-result').addEventListener('click', async () => {
  const s = await window.pywebview.api.get_state();
  if (s.ozon_output_path) window.pywebview.api.open_path(s.ozon_output_path);
});
$('#btn-open-log').addEventListener('click', async () => {
  const s = await window.pywebview.api.get_state();
  if (s.log_path) window.pywebview.api.open_path(s.log_path);
});
$('#btn-open-last-xlsx').addEventListener('click', async () => {
  const s = await window.pywebview.api.get_state();
  if (s.output_path) window.pywebview.api.open_path(s.output_path);
});

// ============================================================
// Polling
// ============================================================
let _errorShown = false;
async function pollOnce() {
  try {
    const s = await window.pywebview.api.get_state();
    updateStatusBar(s);
    updateQueueScreen(s);
    streamLog(s);
    // Автопереход на «Логи» при ошибке
    if (s.error && !_errorShown) {
      _errorShown = true;
      toast('Ошибка прогона — открываю логи', 'err');
      setTimeout(() => go('logs'), 600);
    }
    if (!s.error && !s.running) _errorShown = false;
  } catch (e) { /* pywebview ещё не готов */ }
}

function updateStatusBar(s) {
  const dot = $('#status-dot');
  const txt = $('#status-text');
  if (s.running) {
    dot.className = 'status-dot running';
    txt.textContent = s.stage_label || 'выполняется…';
  } else if (s.error) {
    dot.className = 'status-dot error';
    txt.textContent = 'ошибка';
  } else if (s.output_path) {
    dot.className = 'status-dot ok';
    txt.textContent = 'готово';
  } else {
    dot.className = 'status-dot';
    txt.textContent = 'готов';
  }
  $('#btn-run').disabled  = !!s.running;
  $('#btn-stop').disabled = !s.running;
  $('#queue-badge').classList.toggle('visible', !!s.running);

  // Last run card
  if (s.output_path && !s.running) {
    $('#last-run-card').style.display = '';
    const fname = s.output_path.split(/[\/\\]/).pop();
    $('#last-run-info').innerHTML = `
      <span class="badge badge-ok">✓ готово</span>
      <span style="margin-left:8px; color:var(--accent); font-family:monospace;">${escapeHtml(fname)}</span>
      <span class="text-muted" style="margin-left:8px; font-size:12px;">
        ${fmtElapsed(s.elapsed_sec)}
      </span>`;
  }
}

// v27.6: Пуллинг live-лога. Держим последние ~30 строк stdout.
let _liveLogOffset = 0;
let _liveLogLines = [];
let _liveLogBusy = false;
async function refreshLiveLog(state) {
  try {
    if (!state || !state.running) {
      // Прогон завершён — всё равно дотягиваем остаток один раз
      if (_liveLogOffset >= (state.log_total || 0)) return;
    }
    if (_liveLogBusy) return;
    _liveLogBusy = true;
    const data = await window.pywebview.api.get_log_lines(_liveLogOffset);
    if (data && data.lines && data.lines.length) {
      _liveLogOffset += data.lines.length;
      _liveLogLines = _liveLogLines.concat(data.lines).slice(-30);
      const box = $('#livelog-box');
      if (box) {
        const html = _liveLogLines.map(l => colorLogLine(l)).join('\n');
        box.innerHTML = html;
        box.scrollTop = box.scrollHeight;
      }
      const lblTotal = $('#livelog-total');
      if (lblTotal) lblTotal.textContent = data.total + ' строк';
    }
  } catch(e) {
    // тихо
  } finally {
    _liveLogBusy = false;
  }
}
function colorLogLine(line) {
  const s = escapeHtml(line);
  if (/\[ОШИБКА\]|\[ERROR\]|\bERROR\b|exception|traceback/i.test(line))
    return '<span style="color:#ff6b6b">' + s + '</span>';
  if (/\[ТАЙМАУТ\]|\[WARN\]|warning/i.test(line))
    return '<span style="color:#ffd166">' + s + '</span>';
  if (/\[OK\]|\u2713|готово|success/i.test(line))
    return '<span style="color:#7ce087">' + s + '</span>';
  if (/\[Ozon\]|\[Ozon-Playwright\]/i.test(line))
    return '<span style="color:#4cc9f0">' + s + '</span>';
  if (/^━━━|━━━ /.test(line))
    return '<span style="color:#a78bfa; font-weight:600">' + s + '</span>';
  return s;
}
function resetLiveLog() {
  _liveLogOffset = 0;
  _liveLogLines = [];
  const box = $('#livelog-box');
  if (box) box.innerHTML = 'Ожидание запуска…';
}

function updateQueueScreen(s) {
  const pct = s.progress_pct || 0;
  $('#prog-bar').style.width = pct.toFixed(1) + '%';
  $('#kpi-pct').textContent  = pct.toFixed(0) + '%';
  $('#kpi-pct-sub').textContent = s.progress_total
    ? `${s.progress_done} из ${s.progress_total}` : '— из —';
  $('#kpi-elapsed').textContent = fmtElapsed(s.elapsed_sec || 0);
  $('#kpi-mode-lbl').textContent = s.mode || '—';
  $('#kpi-speed').textContent = s.speed_per_min > 0 ? s.speed_per_min.toFixed(1)
    : (s.progress_speed > 0 ? s.progress_speed : '—');
  // v54.4 (улучшение №7): ETA из движка
  const etaEl = $('#kpi-eta');
  if (etaEl) {
    const eta = s.progress_eta || 0;
    etaEl.textContent = (s.running && eta > 0) ? fmtElapsed(eta) : '—';
  }
  $('#prog-label').textContent = s.running ? 'выполняется' : (s.output_path ? 'завершено' : 'ожидание');
  const STAGE_LABELS = {
    links: 'этап: сбор ссылок на документы',
    registry: 'этап: проверка реестров',
    search: 'этап: поиск карточек',
    cards: 'этап: загрузка карточек',
  };
  $('#prog-stage').textContent = s.stage_label || STAGE_LABELS[s.progress_stage] || '';

  const st  = (s.metrics && s.metrics.status)      || {};
  const reg = (s.metrics && s.metrics.registry)     || {};
  const mkt = (s.metrics && s.metrics.marketplace)  || {};
  _lastRunMode = s.mode || _lastRunMode;

  let okCnt = 0, errCnt = 0;
  Object.entries(st).forEach(([k,v]) => {
    if (k==='OK' || k.includes('СОБРАНА')) okCnt += v;
    else if (k==='ОШИБКА' || k==='ТАЙМАУТ' || k==='НЕСООТВЕТСТВИЕ') errCnt += v;
  });
  $('#kpi-errors').textContent = errCnt;
  $('#kpi-ok-sub').textContent = `OK: ${okCnt}`;

  $('#btn-open-result').disabled      = !s.output_path;
  $('#btn-open-ozon-result').disabled = !s.ozon_output_path;
  $('#btn-open-log').disabled         = !s.log_path;

  // v27.6: Live-лог
  refreshLiveLog(s);

  // Стек-бар статусов
  renderStatusStack(st);

  // Сброс кэша при новом запуске
  if (s.running && _queueChartsLoaded) _queueChartsLoaded = false;

  // Пока идёт прогон — показываем лайв-статистику из metrics (stdout-парсер движка)
  // После завершения — подгружаем из xlsx полную статистику.
  if (_chartjs) {
    if (s.running) {
      // Маркетплейс: движок WB/Ozon не присылает распределение — выводим по
      // режиму, иначе график всегда «нет данных».
      let mktData = mkt;
      if (!mktData || !Object.keys(mktData).length) {
        const totalSt = Object.values(st || {}).reduce((a, b) => a + b, 0);
        if (totalSt > 0) {
          const md = (s.mode || '').toLowerCase();
          mktData = (md.indexOf('ozon') >= 0) ? { 'Ozon': totalSt } : { 'Wildberries': totalSt };
        }
      }
      updateDonutChart('chart-status',      'chart-status-fb',      st,      STATUS_COLORS);
      updateDonutChart('chart-registry',    'chart-registry-fb',    reg,     REGISTRY_COLORS);
      updateDonutChart('chart-marketplace', 'chart-marketplace-fb', mktData, MKT_COLORS);
      // Оригинальность лайв не можем — она в xlsx
      updateDonutChart('chart-original', 'chart-original-fb', {}, ORIGINAL_COLORS);
    } else if (s.output_path && !_queueChartsLoaded) {
      _queueChartsLoaded = true;
      refreshQueueCharts();
    }
  }
}
let _queueChartsLoaded = false;

function renderStatusStack(stat) {
  const stack = $('#status-stack');
  const leg = $('#status-stack-legend');
  if (!stack || !leg) return;
  const entries = Object.entries(stat).filter(([k, v]) => v > 0).sort((a, b) => b[1] - a[1]);
  const total = entries.reduce((s, [, v]) => s + v, 0);
  if (!total) { stack.innerHTML = ''; leg.innerHTML = ''; return; }
  stack.innerHTML = entries.map(([k, v], i) => {
    const w = (v / total * 100).toFixed(2);
    const c = STATUS_COLORS[k] || colorFor(i, entries.length);
    return `<div title="${escapeHtml(k)}: ${v}" style="width:${w}%; background:${c};"></div>`;
  }).join('');
  leg.innerHTML = entries.map(([k, v], i) => {
    const c = STATUS_COLORS[k] || colorFor(i, entries.length);
    return `<span style="display:inline-flex; align-items:center; gap:6px;">
      <span style="width:10px; height:10px; border-radius:2px; background:${c};"></span>
      ${escapeHtml(k)} — <strong style="color:var(--fg);">${v}</strong>
    </span>`;
  }).join('');
}

// ============================================================
// Log streaming
// ============================================================
async function streamLog(s) {
  if (s.log_total > _logSeen) {
    const chunk = await window.pywebview.api.get_log_lines(_logSeen);
    _logSeen = chunk.from + chunk.lines.length;
    appendLogLines(chunk.lines);
  }
}

function classifyLine(ln) {
  if (ln.startsWith('━━━'))            return 'log-line-head';
  if (ln.startsWith('[команда]'))       return 'log-line-cmd';
  if (ln.startsWith('[ОШИБКА') || ln.includes('[ERROR]') || ln.includes('ERROR')) return 'log-line-err';
  if (ln.includes('[OK]') || ln.includes('[OK ') || ln.startsWith('[OK]')) return 'log-line-ok';
  if (ln.includes('[ТАЙМАУТ]') || ln.includes('[ОШИБКА]') || ln.includes('ОШИБК')) return 'log-line-err';
  if (ln.includes('WARN') || ln.includes('ПРЕДУПРЕЖ')) return 'log-line-warn';
  if (ln.startsWith('[завершено') || ln.startsWith('[команда')) return 'log-line-cmd';
  return 'log-line-info';
}

function passesFilter(cls) {
  if (_logFilterLvl === 'all') return true;
  if (_logFilterLvl === 'error') return cls === 'log-line-err';
  if (_logFilterLvl === 'warn')  return cls === 'log-line-warn';
  if (_logFilterLvl === 'ok')    return cls === 'log-line-ok';
  return true;
}

function appendLogLines(lines) {
  const el = $('#log-view');
  if (el.children.length === 1 && el.querySelector('span')) el.innerHTML = '';
  const atBottom = el.scrollTop + el.clientHeight >= el.scrollHeight - 40;
  const frag = document.createDocumentFragment();
  for (const ln of lines) {
    const cls = classifyLine(ln);
    if (!passesFilter(cls)) continue;
    const div = document.createElement('div');
    div.className = cls;
    div.textContent = ln;
    frag.appendChild(div);
  }
  el.appendChild(frag);
  if (atBottom) el.scrollTop = el.scrollHeight;
  $('#log-info').textContent = `${_logSeen} строк`;
}

function renderLogFull() {
  $('#log-info').textContent = `${_logSeen} строк`;
}

$$('.log-filter-btn').forEach(b => {
  b.addEventListener('click', () => {
    $$('.log-filter-btn').forEach(x => x.classList.remove('active'));
    b.classList.add('active');
    _logFilterLvl = b.dataset.lvl;
    rebuildLog();
  });
});

function rebuildLog() {
  // Перефильтровываем уже загруженные строки
  const el = $('#log-view');
  const existing = Array.from(el.children);
  existing.forEach(div => {
    if (!div.className) return;
    div.style.display = passesFilter(div.className) ? '' : 'none';
  });
}

$('#btn-clear-log').addEventListener('click', () => {
  $('#log-view').innerHTML = '<span style="color:var(--muted);">— журнал очищен —</span>';
  _logSeen = 0;
});
function copyTextRobust(text) {
  // v27.9.x: navigator.clipboard недоступен в pywebview (не secure-context) —
  // делаем надёжный fallback через скрытую textarea + execCommand('copy').
  return new Promise((resolve, reject) => {
    try {
      if (navigator.clipboard && window.isSecureContext) {
        navigator.clipboard.writeText(text).then(resolve).catch(() => fallback());
      } else { fallback(); }
    } catch (e) { fallback(); }
    function fallback() {
      try {
        const ta = document.createElement('textarea');
        ta.value = text;
        ta.style.position = 'fixed';
        ta.style.left = '-9999px';
        ta.style.top = '0';
        document.body.appendChild(ta);
        ta.focus(); ta.select();
        const ok = document.execCommand('copy');
        document.body.removeChild(ta);
        ok ? resolve() : reject(new Error('execCommand failed'));
      } catch (e2) { reject(e2); }
    }
  });
}

$('#btn-copy-log').addEventListener('click', async () => {
  const el = $('#log-view');
  const text = el ? (el.innerText || el.textContent || '') : '';
  try { await copyTextRobust(text); toast('Лог скопирован', 'ok'); }
  catch (e) { toast('Не удалось скопировать', 'err'); }
});
$('#btn-save-log').addEventListener('click', async () => {
  const res = await window.pywebview.api.export_csv(null, '');
  // Лог отдельно не экспортируется — открываем log_path если есть
  const s = await window.pywebview.api.get_state();
  if (s.log_path) { window.pywebview.api.open_path(s.log_path); }
  else toast('Лог-файл не найден', 'info');
});

// ============================================================
// Results table
// ============================================================
async function loadResults() {
  const wrap = $('#tbl-wrap');
  wrap.innerHTML = '<div class="empty-state"><div class="empty-ico">⏳</div><p>Загрузка…</p></div>';
  // v48: грузим ВСЕ строки (раньше таблица ограничивалась — пользователь видел не всё)
  const res = await window.pywebview.api.get_results(null, 1000000);
  if (!res.ok) {
    wrap.innerHTML = `<div class="empty-state"><div class="empty-ico">📋</div><p>${escapeHtml(res.error)}</p></div>`;
    return;
  }
  _allHeaders = res.columns;
  _allRows    = res.rows;
  _visibleCols = null;  // v48: новый файл — колонки по умолчанию
  const _moreNote = (res.total > _allRows.length)
    ? ` (загружено ${_allRows.length})` : '';
  $('#results-count').textContent = `${res.total} строк${_moreNote} · «${res.sheet}»`;

  // Fill status filter
  const statSel = $('#filter-status');
  const statuses = Object.keys(res.stats.by_status || {});
  statSel.innerHTML = '<option value="">Все статусы</option>' +
    statuses.map(s => `<option value="${escapeHtml(s)}">${escapeHtml(s)}</option>`).join('');

  // v46: новый файл/обновление — сбрасываем все активные фильтры.
  clearAllFilters();
  const _fi = $('#filter-input'); if (_fi) _fi.value = '';
  const _fs = $('#filter-status'); if (_fs) _fs.value = '';
  const _chips = $('#filter-chips'); if (_chips) _chips.innerHTML = '';
  renderTable(_allRows);

  // Charts — из статистики Python по всем строкам (authoritative).
  // v49: инициализируем Chart.js здесь же и строим БЕЗУСЛОВНО. Раньше стоял
  // guard `if (_chartjs)`, а _chartjs выставляется лениво — при открытии экрана
  // «Результаты» (особенно через «Загрузить файл») до первого опроса очереди он
  // мог быть ещё null → графики не рисовались. tryInitCharts() подхватывает уже
  // встроенный window.Chart; buildResultsCharts сам выходит, если его нет.
  tryInitCharts();
  buildResultsCharts(res.stats);

  // v54.4 (улучшение №8): почему «ПРОВЕРИТЬ ВРУЧНУЮ» — разбивка по причинам.
  const rr = res.stats.by_review_reason || {};
  const rrPanel = $('#review-reasons');
  if (rrPanel) {
    const items = Object.entries(rr).sort((a, b) => b[1] - a[1]);
    if (items.length) {
      rrPanel.style.display = '';
      rrPanel.innerHTML = '<b>Почему «Проверить вручную»:</b> ' +
        items.map(([k, v]) => `${escapeHtml(k)} — <b>${v}</b>`).join(' · ') +
        ' <span style="opacity:.7">(подсказывает, где пополнить словарь — кнопка «Обучить словарь»)</span>';
    } else {
      rrPanel.style.display = 'none';
    }
  }
}

// v27.9.x: КУРИРУЕМЫЙ набор колонок таблицы — включает «Название в реестре»
// (наименование товара из реестра) и ключевые поля документа, которые раньше
// были за пределами первых 14 столбцов и не показывались.
const PREFERRED_COLS = [
  'Запрос', 'Артикул WB', 'Название товара', 'Бренд', 'Категория WB',
  'Технический статус', 'Цена со скидкой, ₽', 'Продавец', "Плашка 'Оригинал'",
  'Документ проверен WB', 'Рейтинг', 'Отзывы',
  'Реестр (страна)', 'Название в реестре', 'Статус документа', 'Статус на территории РФ',
  'Номер документа', 'Тип документа',
  'ТН ВЭД', 'Изготовитель', 'Действует до', 'Риск по сроку',
];
// URL-колонки в таблице НЕ показываем (по просьбе), но используем для кликов:
// «Артикул WB» -> страница товара, «Номер документа» -> страница реестра.

// v48: видимые колонки — пользователь может скрывать/показывать любые столбцы.
// null = набор по умолчанию (PREFERRED_COLS, что есть в данных).
let _visibleCols = null;

function _allDisplayableCols() {
  // PREFERRED (что есть в данных) первыми, затем остальные колонки файла —
  // URL-колонки исключаем (они используются для кликов, не для показа).
  const pref = PREFERRED_COLS.filter(h => _allHeaders.includes(h));
  const rest = _allHeaders.filter(h =>
    h && !pref.includes(h) && !/^(ссылка на товар|ссылка на реестр|product_url|registry_url)$/i.test(h));
  return pref.concat(rest);
}

function _defaultVisibleCols() {
  return PREFERRED_COLS.filter(h => _allHeaders.includes(h));
}

function _tableColIndices() {
  // Сопоставляем выбранные (или дефолтные) заголовки с реальными; чего нет — пропускаем.
  const cols = (_visibleCols && _visibleCols.length) ? _visibleCols : _defaultVisibleCols();
  let idx = [];
  for (const name of cols) {
    const i = _allHeaders.indexOf(name);
    if (i >= 0 && !idx.includes(i)) idx.push(i);
  }
  if (!idx.length) idx = _allHeaders.map((_, i) => i).slice(0, 16);
  return idx;
}

// v48: панель выбора колонок (скрыть/показать любые столбцы результата).
function toggleColPicker() {
  let p = document.getElementById('col-picker');
  if (p) { p.remove(); return; }
  if (!_allHeaders.length) return;
  const btn = document.getElementById('btn-columns');
  const opts = _allDisplayableCols();
  const visible = new Set((_visibleCols && _visibleCols.length) ? _visibleCols : _defaultVisibleCols());
  p = document.createElement('div');
  p.id = 'col-picker';
  p.className = 'col-picker';
  p.innerHTML =
    '<div class="cp-head"><b>Колонки</b>' +
    '<span class="cp-actions"><a data-cp="pref">по умолчанию</a> · <a data-cp="all">все</a> · <a data-cp="none">снять все</a></span></div>' +
    '<div class="cp-list">' +
    opts.map(h => `<label><input type="checkbox" value="${escapeHtml(h)}" ${visible.has(h) ? 'checked' : ''}> <span>${escapeHtml(h)}</span></label>`).join('') +
    '</div>';
  document.body.appendChild(p);
  // позиционирование под кнопкой
  if (btn) {
    const r = btn.getBoundingClientRect();
    p.style.top = (r.bottom + 6) + 'px';
    p.style.left = Math.max(8, Math.min(r.left, window.innerWidth - 320)) + 'px';
  }
  function _applyFromChecks() {
    _visibleCols = Array.from(p.querySelectorAll('input[type=checkbox]:checked')).map(c => c.value);
    renderTable(_lastRenderRows);
  }
  p.querySelectorAll('input[type=checkbox]').forEach(cb => cb.addEventListener('change', _applyFromChecks));
  p.querySelectorAll('[data-cp]').forEach(a => a.addEventListener('click', () => {
    const mode = a.dataset.cp;
    if (mode === 'all') _visibleCols = opts.slice();
    else if (mode === 'none') _visibleCols = [];
    else _visibleCols = _defaultVisibleCols();
    const vis = new Set(_visibleCols);
    p.querySelectorAll('input[type=checkbox]').forEach(cb => { cb.checked = vis.has(cb.value); });
    renderTable(_lastRenderRows);
  }));
  // закрытие по клику вне панели
  setTimeout(() => {
    document.addEventListener('click', function _close(ev) {
      if (!p.contains(ev.target) && ev.target.id !== 'btn-columns') {
        p.remove(); document.removeEventListener('click', _close);
      }
    });
  }, 0);
}

let _sortCol = -1;      // индекс колонки сортировки (в _allHeaders)
let _sortDir = 1;       // 1 = по возрастанию, -1 = по убыванию
let _lastRenderRows = [];

function _sortRows(rows, ci, dir) {
  const num = rows.every(r => r[ci] === '' || r[ci] === null || r[ci] === undefined || !isNaN(Number(r[ci])));
  return rows.slice().sort((a, b) => {
    let va = a[ci] ?? '', vb = b[ci] ?? '';
    if (num) { va = Number(va) || 0; vb = Number(vb) || 0; return (va - vb) * dir; }
    return String(va).localeCompare(String(vb), 'ru') * dir;
  });
}

// v47: ПОРЦИОННЫЙ рендер таблицы. На 20k+ строк построение DOM одним куском
// (раньше 5000 строк × ~20 колонок) замораживало окно на секунды. Теперь
// рендерим порциями по TABLE_CHUNK с кнопкой «Показать ещё»; фильтры,
// сортировка и диаграммы по-прежнему работают по ПОЛНОМУ набору строк.
const TABLE_CHUNK = 1000;
let _tblView = [];   // строки текущего вида (после сортировки)
let _tblShown = 0;   // сколько строк уже в DOM
let _tblCtx = null;  // {colIdx, hl, purlIdx, rurlIdx}

function _rowHtml(row) {
  const { colIdx, hl, purlIdx, rurlIdx } = _tblCtx;
  const purl = purlIdx >= 0 ? String(row[purlIdx] ?? '') : '';
  const rurl = rurlIdx >= 0 ? String(row[rurlIdx] ?? '') : '';
  let html = '<tr>';
  colIdx.forEach((ci) => {
    const v = String(row[ci] ?? '');
    const head = hl[ci] || '';
    let badge = '';
    if (head.includes('риск')) {
      if (v === 'Действует')         badge = 'badge-ok';
      else if (v === 'Скоро истекает') badge = 'badge-warn';
      else if (v === 'Истёк')          badge = 'badge-error';
      else if (v)                      badge = 'badge-muted';
    } else if (head.includes('статус') || head.includes('status')) {
      // Технический статус + Статус документа + Статус на территории РФ.
      if (v.includes('OK') || v.includes('СОБРАНА') || v === 'Действует') badge = 'badge-ok';
      else if (v.includes('НЕДЕЙСТВ') || v.includes('НЕСООТВЕТ') || v.includes('ОШИБКА')
               || v.includes('ТАЙМАУТ') || v.includes('Прекращ') || v.includes('Аннулир')) badge = 'badge-error';
      else if (v.includes('ПРОВЕРИТЬ') || v.includes('НЕ УДАЛОСЬ') || v.includes('НЕ ПРОВЕРЕН')
               || v.includes('Приостановл') || v.includes('Архивн')) badge = 'badge-warn';
      else if (v) badge = 'badge-muted';
    } else if (head.includes('маркетплейс') || head.includes('marketplace')) {
      if (v === 'WB' || v.toLowerCase().includes('wildberries')) badge = 'badge-wb';
      else if (v.toLowerCase().includes('ozon')) badge = 'badge-ozon';
    }
    // Кликабельные ссылки: «Артикул WB» -> товар, «Номер документа» -> реестр.
    if ((head.includes('артикул') || head === 'nm_id') && purl) {
      html += `<td class="cell-link" data-url="${escapeHtml(purl)}" title="Открыть товар на WB">${escapeHtml(v)} ↗</td>`;
      return;
    }
    if ((head.includes('номер документа') || head.includes('certificate_number')) && rurl) {
      html += `<td class="cell-link" data-url="${escapeHtml(rurl)}" title="Открыть реестр документа">${escapeHtml(v || 'реестр')} ↗</td>`;
      return;
    }
    // длинные текстовые поля (название товара / название в реестре) — кликабельны (полный текст)
    const isLong = (head.includes('название') || head.includes('изготовитель') || head.includes('примечан')) && v.length > 28;
    const cellCls = isLong ? ' class="cell-expand"' : '';
    html += badge
      ? `<td><span class="badge ${badge}">${escapeHtml(v)}</span></td>`
      : `<td${cellCls} title="${escapeHtml(v)}" data-full="${escapeHtml(v)}">${escapeHtml(v)}</td>`;
  });
  return html + '</tr>';
}

function _nextChunkHtml() {
  const slice = _tblView.slice(_tblShown, _tblShown + TABLE_CHUNK);
  _tblShown += slice.length;
  return slice.map(_rowHtml).join('');
}

function _updateMoreBtn(wrap) {
  const more = wrap.querySelector('#tbl-more');
  if (!more) return;
  const left = _tblView.length - _tblShown;
  if (left <= 0) {
    more.innerHTML = `<span class="text-muted text-sm">Показаны все ${_tblView.length} строк</span>`;
    return;
  }
  more.innerHTML =
    `<button class="btn btn-ghost btn-sm" data-more="chunk">Показать ещё ${Math.min(TABLE_CHUNK, left)}</button> ` +
    `<button class="btn btn-ghost btn-sm" data-more="all">Показать все (${left})</button> ` +
    `<span class="text-muted text-sm">показано ${_tblShown} из ${_tblView.length}</span>`;
}

function renderTable(rows) {
  const wrap = $('#tbl-wrap');
  if (!rows.length) {
    wrap.innerHTML = '<div class="empty-state"><div class="empty-ico">🤷</div><p>Ничего не найдено</p></div>';
    return;
  }
  _lastRenderRows = rows;
  const colIdx = _tableColIndices();
  // Сортировка (если выбрана колонка)
  let viewRows = rows;
  if (_sortCol >= 0) viewRows = _sortRows(rows, _sortCol, _sortDir);
  _tblView = viewRows;
  _tblShown = 0;
  _tblCtx = {
    colIdx,
    hl: _allHeaders.map(x => x.toLowerCase()),
    purlIdx: _allHeaders.findIndex(h => /ссылка на товар|product_url/i.test(h)),
    rurlIdx: _allHeaders.findIndex(h => /ссылка на реестр|registry_url/i.test(h)),
  };

  let html = '<table><thead><tr>' +
    colIdx.map(i => {
      const h = _allHeaders[i] || '';
      const arrow = (_sortCol === i) ? (_sortDir === 1 ? ' ▲' : ' ▼') : '';
      const lbl = escapeHtml(h.length > 22 ? h.slice(0,20)+'…' : h) + arrow;
      return `<th class="sortable" data-col="${i}" title="${escapeHtml(h)} (клик — сортировать)">${lbl}</th>`;
    }).join('') +
    '</tr></thead><tbody>' + _nextChunkHtml() + '</tbody></table>' +
    '<div id="tbl-more" style="text-align:center;padding:10px"></div>';
  wrap.innerHTML = html;
  _updateMoreBtn(wrap);

  // v47: ДЕЛЕГИРОВАНИЕ кликов (один обработчик вместо тысяч на ячейках) —
  // дешевле на больших таблицах и автоматически работает для дорендеренных порций.
  wrap.onclick = (e) => {
    const moreBtn = e.target.closest('#tbl-more button');
    if (moreBtn) {
      const tbody = wrap.querySelector('tbody');
      if (moreBtn.dataset.more === 'all') {
        // дорисовываем ВСЕ оставшиеся строки порциями (чтобы не блокировать надолго)
        let html = '';
        while (_tblShown < _tblView.length) html += _nextChunkHtml();
        if (tbody) tbody.insertAdjacentHTML('beforeend', html);
      } else if (tbody) {
        tbody.insertAdjacentHTML('beforeend', _nextChunkHtml());
      }
      _updateMoreBtn(wrap);
      return;
    }
    const link = e.target.closest('td.cell-link');
    if (link) {
      if (link.dataset.url) window.pywebview.api.open_url(link.dataset.url);
      return;
    }
    const th = e.target.closest('th.sortable');
    if (th) {
      const ci = Number(th.dataset.col);
      if (_sortCol === ci) _sortDir = -_sortDir; else { _sortCol = ci; _sortDir = 1; }
      renderTable(_lastRenderRows);
      return;
    }
    const exp = e.target.closest('td.cell-expand');
    if (exp) showFullTextModal(exp.dataset.full || exp.textContent);
  };
}

function showFullTextModal(text) {
  let m = document.getElementById('full-text-modal');
  if (!m) {
    m = document.createElement('div');
    m.id = 'full-text-modal';
    m.innerHTML = '<div class="ftm-backdrop"></div><div class="ftm-box"><div class="ftm-text"></div><button class="ftm-close">Закрыть</button></div>';
    document.body.appendChild(m);
    m.querySelector('.ftm-backdrop').addEventListener('click', () => m.style.display = 'none');
    m.querySelector('.ftm-close').addEventListener('click', () => m.style.display = 'none');
  }
  m.querySelector('.ftm-text').textContent = text;
  m.style.display = 'block';
}

// v46: текстовый поиск и выпадающий статус — тоже часть ЕДИНОЙ системы фильтров
// (комбинируются с фильтрами диаграмм и пересчитывают все диаграммы).
// v48: ДЕБАУНС поиска — на 50k+ строк пересчёт фильтра+диаграмм на каждое нажатие
// клавиши заметно лагал; ждём 220мс тишины и считаем один раз.
let _filterDebounce = null;
$('#filter-input').addEventListener('input', e => {
  const q = e.target.value.trim();
  if (_filterDebounce) clearTimeout(_filterDebounce);
  _filterDebounce = setTimeout(() => {
    const ql = q.toLowerCase();
    if (q) setFilter('text', 'Поиск: ' + q,
                     r => r.some(c => String(c).toLowerCase().includes(ql)));
    else removeFilter('text');
  }, 220);
});
$('#filter-status').addEventListener('change', e => {
  const st = e.target.value;
  if (st) {
    const si = _colIdxByName('технический статус', 'status', 'статус');
    setFilter('status-dd', 'Статус: ' + st,
              r => si >= 0 && String(r[si] ?? '').toLowerCase().includes(st.toLowerCase()));
  } else removeFilter('status-dd');
});

$('#btn-reload-results').addEventListener('click', loadResults);
{ const _bc = document.getElementById('btn-columns'); if (_bc) _bc.addEventListener('click', (e) => { e.stopPropagation(); toggleColPicker(); }); }

// Загрузка внешнего result.xlsx (с другого прогона) для анализа во вкладке.
$('#btn-load-result').addEventListener('click', async () => {
  let res;
  try { res = await window.pywebview.api.browse_result_file(); }
  catch (e) { toast('Ошибка диалога: ' + e, 'err'); return; }
  if (!res || res.cancelled) return;
  if (!res.ok) { toast(res.error || 'Не удалось загрузить файл', 'err'); return; }
  const badge = $('#loaded-file-badge');
  badge.textContent = '📂 ' + (res.name || 'файл');
  badge.style.display = '';
  $('#btn-clear-loaded').style.display = '';
  toast('Загружен файл: ' + (res.name || ''), 'ok');
  await loadResults();
});

// Вернуться к результату текущего прогона.
$('#btn-clear-loaded').addEventListener('click', async () => {
  try { await window.pywebview.api.clear_loaded_result(); } catch (e) {}
  $('#loaded-file-badge').style.display = 'none';
  $('#btn-clear-loaded').style.display = 'none';
  toast('Показан результат текущего прогона', 'ok');
  await loadResults();
});

const _btnRetryFsa = $('#btn-retry-fsa');
if (_btnRetryFsa) _btnRetryFsa.addEventListener('click', async () => {
  const res = await window.pywebview.api.retry_failed_fsa();
  if (res && res.ok) {
    toast('Повтор FSA запущен — пере-проверяются только упавшие ссылки', 'ok');
    go('queue');
  } else {
    toast((res && res.error) || 'Не удалось запустить повтор', 'err');
  }
});
$('#btn-export-csv').addEventListener('click', async () => {
  const q = $('#filter-input').value.trim();
  const res = await window.pywebview.api.export_csv(null, q);
  if (res.ok) toast(`CSV сохранён: ${res.rows} строк`, 'ok');
  else toast(res.error || 'Ошибка экспорта', 'err');
});

// v53 (улучшение №10): выгрузка только спорных строк со ссылками и причиной.
$('#btn-export-disputed').addEventListener('click', async () => {
  const btn = $('#btn-export-disputed');
  btn.disabled = true;
  try {
    const res = await window.pywebview.api.export_disputed(null);
    if (res.ok) toast(`Файл «на проверку» сохранён: ${res.rows} спорных строк`, 'ok');
    else toast(res.error || 'Ошибка экспорта', 'err');
  } finally {
    btn.disabled = false;
  }
});

// v54.4 (улучшение №12): сводка по продавцам/брендам.
$('#btn-supplier-stats').addEventListener('click', async () => {
  const btn = $('#btn-supplier-stats');
  btn.disabled = true;
  try {
    const res = await window.pywebview.api.export_supplier_stats(null);
    if (res.ok) toast(`Сводка по продавцам сохранена: ${res.rows} строк`, 'ok');
    else toast(res.error || 'Ошибка экспорта', 'err');
  } finally { btn.disabled = false; }
});

// v54.4 (улучшение №9): «обучить словарь» — кандидаты из товаров без категории.
$('#btn-suggest-dict').addEventListener('click', async () => {
  const btn = $('#btn-suggest-dict');
  btn.disabled = true;
  try {
    const res = await window.pywebview.api.suggest_dictionary(null);
    if (res.ok) toast(`Кандидаты словаря: ${res.rows} слов. ${res.hint || ''}`, 'ok', 9000);
    else toast(res.error || 'Ошибка', 'err');
  } finally { btn.disabled = false; }
});

// ============================================================
// Settings screen
// ============================================================
async function fillSettings() {
  const v = await window.pywebview.api.get_version();

  $('#settings-version-info').innerHTML = `
    <div>Версия: <b>${escapeHtml(v.app_version)}</b></div>
    <div>Python ${escapeHtml(v.python)} · ${escapeHtml(v.platform)}</div>
    <div style="margin-top:4px; color:var(--muted); font-family:monospace; font-size:12px;">${escapeHtml(v.app_dir)}</div>`;

  $('#settings-dir').textContent = v.app_dir;

  // Engine statuses
  const engines = [
    {label:'main_v39.py (WB Query)',  exists: v.engine_v39_exists,  path: v.engine_v39},
    {label:'main_brand.py (WB Brand)',exists: v.engine_brand_exists, path: v.engine_brand},
    {label:'ozon_parser.py (Ozon)',   exists: v.engine_ozon_exists,  path: v.engine_ozon},
    {label:'wb_enhanced.py',          exists: v.engine_wb_enhanced_exists},
    {label:'fsa_enhanced.py',         exists: v.engine_fsa_enhanced_exists},
  ];
  $('#settings-engines').innerHTML = engines.map(e => `
    <div class="engine-status">
      <span class="engine-dot ${e.exists ? 'ok' : 'miss'}"></span>
      <span style="font-family:monospace; font-size:12.5px; flex:1;">${escapeHtml(e.label)}</span>
      <span class="badge ${e.exists ? 'badge-ok' : 'badge-error'}">${e.exists ? 'найден' : 'отсутствует'}</span>
    </div>`).join('');

  // Sidebar footer
  $('#foot-py').textContent = `Python ${v.python} · ${v.platform}`;
  $('#foot-dir').textContent = v.app_dir;
  const footEl = $('#foot-engines');
  footEl.innerHTML = engines.slice(0,3).map(e => `
    <div>
      <span class="engine-dot ${e.exists?'ok':'miss'}"></span>
      <span style="font-size:10.5px;">${escapeHtml(e.label.split(' ')[0])}</span>
    </div>`).join('');

  // Load saved defaults
  const s = await window.pywebview.api.get_settings();
  if (s.defaults) {
    if (s.defaults.expiry_warning_days) $('#s-expiry').value = s.defaults.expiry_warning_days;
    if (s.defaults.workers) $('#s-workers').value = s.defaults.workers;
  }
  refreshKgInfo();
}

$('#btn-save-defaults').addEventListener('click', async () => {
  const data = {
    defaults: {
      expiry_warning_days: parseInt($('#s-expiry').value) || 30,
      workers: parseInt($('#s-workers').value) || 4,
    }
  };
  const res = await window.pywebview.api.save_settings(data);
  toast(res.ok ? 'Настройки сохранены' : 'Ошибка сохранения', res.ok ? 'ok' : 'err');
});

// Таблица статусов КГ-документов в РФ
async function refreshKgInfo() {
  try {
    const r = await window.pywebview.api.kg_status_info();
    const el = $('#kg-status-info');
    if (el) el.textContent = (r && r.loaded)
      ? `✓ загружено: ${r.count} записей` : 'таблица не загружена';
  } catch (e) {}
}
const _btnKg = $('#btn-load-kg');
if (_btnKg) _btnKg.addEventListener('click', async () => {
  let r;
  try { r = await window.pywebview.api.browse_kg_status_file(); }
  catch (e) { toast('Ошибка диалога: ' + e, 'err'); return; }
  if (!r || r.cancelled) return;
  if (!r.ok) { toast(r.error || 'Не удалось загрузить', 'err'); return; }
  toast(`Таблица КГ загружена: ${r.count} записей`, 'ok');
  refreshKgInfo();
});

// Theme toggle
$$('.theme-btn').forEach(b => {
  b.addEventListener('click', () => {
    const th = b.dataset.theme;
    document.documentElement.setAttribute('data-theme', th === 'light' ? 'light' : '');
    $$('.theme-btn').forEach(x => x.classList.toggle('active', x.dataset.theme === th));
    window.pywebview.api.save_settings({ theme: th });
  });
});

// ============================================================
// Chart.js integration
// ============================================================
// v27.9.x: РАЗЛИЧИМЫЕ оттенки — раньше OK и «ССЫЛКА СОБРАНА» были оба зелёными,
// а НЕСООТВЕТСТВИЕ/ОШИБКА/ТАЙМАУТ — почти одинаково красно-оранжевыми.
// v49: перцептивно-РАЗЛИЧИМАЯ палитра (Tableau-стиль) — цвета максимально
// разнесены по тону/светлоте, чтобы статусы на графике не сливались.
const STATUS_COLORS = {
  'OK':                                    '#4e9a51', // зелёный
  'ДОКУМЕНТ НЕ ПРОВЕРЕН':                  '#17becf', // бирюзово-циан
  'ССЫЛКА НА РЕЕСТР СОБРАНА':           '#1f77b4', // синий
  'НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ':                 '#ff7f0e', // оранжевый
  'ПРОВЕРИТЬ ВРУЧНУЮ':                     '#e7c000', // золотисто-жёлтый
  'НЕСООТВЕТСТВИЕ':                        '#d62728', // красный
  'НЕДЕЙСТВУЕТ В РФ':                       '#e377c2', // розово-пурпурный
  'НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА': '#9467bd', // фиолетовый
  'ОШИБКА':                                '#8c564b', // коричнево-бордовый
  'ТАЙМАУТ':                               '#bcbd22', // оливковый
  'НЕТ ДОКУМЕНТОВ':                        '#7f7f7f', // серый
  'НЕТ ССЫЛКИ НА РЕЕСТР':                 '#a2b3c2', // сине-серый
};
const MKT_COLORS = {
  'WB': '#cb11ab', 'Wildberries': '#cb11ab', 'wildberries': '#cb11ab',
  'Ozon': '#005bff', 'OZON': '#005bff', 'ozon': '#005bff',
  'Не определен': '#6b7280',
};
const REGISTRY_COLORS = {
  'pub.fsa.gov.ru':      '#5b8cff',
  'swis.trade.kg':       '#10b981',
  'belgiss.by':          '#a78bfa',
  'tsouz.belgiss.by':    '#8b5cf6',
  'portal.eaeunion.org': '#fb923c',
  // v27.9.x: ключи живого графика (из лога прогресса) — те же цвета, чтобы
  // BelGISS отображался наравне с ФСА/SWIS, если на него собраны ссылки.
  'ФСА':                 '#5b8cff',
  'SWIS':                '#10b981',
  'BelGISS':             '#8b5cf6',
};
const ORIGINAL_COLORS = {
  'Оригинал':     '#10b981',
  'Не оригинал': '#ef4444',
  'Не указано':  '#6b7280',
};
const RISK_COLORS = {
  'Действует':          '#10b981',
  'Скоро истекает':     '#fbbf24',
  'Истёк':              '#ef4444',
  'Срок не известен': '#6b7280',
};

// 24-цветная палитра без повторов, подобранная по оттенкам
const PALETTE = [
  '#5b8cff','#10b981','#a78bfa','#fb923c','#f87171','#38bdf8',
  '#fbbf24','#34d399','#ec4899','#8b5cf6','#06b6d4','#84cc16',
  '#f59e0b','#14b8a6','#d946ef','#0ea5e9','#22c55e','#eab308',
  '#f43f5e','#6366f1','#7c3aed','#f97316','#65a30d','#3b82f6',
];

function colorFor(i, total) {
  // Равномерно распределяем цвета по палитре, чтобы соседние не сливались
  if (total <= 0) return PALETTE[0];
  const step = Math.max(1, Math.floor(PALETTE.length / Math.min(total, PALETTE.length)));
  return PALETTE[(i * step) % PALETTE.length];
}

function randomColor(i) { return colorFor(i, 8); }

function tryInitCharts() {
  if (_chartjs || !window.Chart) return;
  _chartjs = window.Chart;
  // Глобальные настройки Chart.js
  _chartjs.defaults.color = '#6b6e82';
  _chartjs.defaults.borderColor = 'rgba(255,255,255,0.06)';
  _chartjs.defaults.font.family = "-apple-system, 'Segoe UI', Roboto, sans-serif";
  _chartjs.defaults.plugins.legend.labels.boxWidth = 12;
  _chartjs.defaults.plugins.legend.labels.padding  = 14;
}

function destroyChart(id) {
  if (_charts[id]) { _charts[id].destroy(); delete _charts[id]; }
}

function _truncateLabel(s, n) {
  s = String(s || '');
  return s.length > n ? s.slice(0, n - 1) + '…' : s;
}

function updateDonutChart(canvasId, fallbackId, data, colorMap, opts) {
  tryInitCharts();
  opts = opts || {};
  const canvas = $('#' + canvasId);
  const fb = fallbackId ? $('#' + fallbackId) : null;
  if (!canvas) return;
  // Сортируем по убыванию
  let entries = Object.entries(data).filter(([k, v]) => v > 0).sort((a, b) => b[1] - a[1]);
  if (!entries.length) {
    if (fb) { fb.style.display = ''; canvas.style.display = 'none'; }
    destroyChart(canvasId);
    return;
  }
  if (fb) { fb.style.display = 'none'; canvas.style.display = ''; }
  const fullLabels = entries.map(([k]) => k);
  const labels = fullLabels.map(k => _truncateLabel(k, opts.maxLabel || 28));
  const values = entries.map(([, v]) => v);
  // v46: цвета в пределах ОДНОГО графика не повторяются. Сначала берём
  // фиксированный цвет из colorMap; для остальных — следующий НЕзанятый из палитры.
  const _used = new Set();
  const colors = fullLabels.map(k => {
    const c = colorMap && colorMap[k];
    if (c) { _used.add(c); return c; }
    return null;
  });
  let _pi = 0;
  for (let i = 0; i < colors.length; i++) {
    if (colors[i]) continue;
    let tries = 0;
    while (_used.has(PALETTE[_pi % PALETTE.length]) && tries < PALETTE.length) { _pi++; tries++; }
    const c = PALETTE[_pi % PALETTE.length];
    _used.add(c); colors[i] = c; _pi++;
  }

  const chartType = opts.type || 'doughnut';

  // v27.9.x: ОБНОВЛЯЕМ существующий график на месте (а не destroy+create на
  // каждом опросе) — иначе график мерцал каждые 600мс. Полные подписи храним на
  // самом графике ($fullLabels), чтобы tooltip работал и после обновления.
  const existing = _charts[canvasId];
  if (existing && existing.$chartType === chartType) {
    existing.$fullLabels = fullLabels;
    existing.data.labels = labels;
    existing.data.datasets[0].data = values;
    existing.data.datasets[0].backgroundColor = colors;
    existing.update('none');  // без анимации — плавно и без мерцания
    return;
  }

  const cfg = {
    type: chartType,
    data: {
      labels,
      datasets: [{
        data: values,
        backgroundColor: colors,
        borderWidth: opts.type === 'bar' ? 0 : 2,
        borderColor: '#1e1f29',
      }],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      animation: { duration: 200 },
      // v46: клик по сегменту -> фильтр таблицы по этому значению.
      onClick: (evt, elements, chart) => {
        if (!opts.onSegmentClick || !elements || !elements.length) return;
        const full = (chart.$fullLabels || [])[elements[0].index];
        if (full != null) opts.onSegmentClick(full);
      },
      onHover: (evt, elements) => {
        try { evt.native.target.style.cursor =
          (opts.onSegmentClick && elements && elements.length) ? 'pointer' : 'default'; } catch(e){}
      },
      cutout: opts.type === 'bar' ? undefined : '62%',
      indexAxis: opts.type === 'bar' ? 'y' : undefined,
      plugins: {
        legend: {
          display: opts.type !== 'bar',
          position: 'right',
          labels: { font: { size: 11 }, boxWidth: 12, padding: 10 },
        },
        tooltip: {
          callbacks: {
            // полные подписи берём с самого графика (переживает обновления)
            title: (items) => ((items[0].chart.$fullLabels || [])[items[0].dataIndex]) || '',
            label: ctx => {
              const arr = ctx.chart.data.datasets[0].data || [];
              const total = arr.reduce((a, b) => a + b, 0) || 1;
              const v = ctx.parsed.x ?? ctx.parsed.y ?? ctx.parsed;
              return ` ${v} (${(v / total * 100).toFixed(1)}%)`;
            },
          },
        },
      },
      scales: opts.type === 'bar' ? {
        x: { grid: { color: 'rgba(255,255,255,0.04)' }, ticks: { color: '#9aa0b4' } },
        y: { grid: { display: false }, ticks: { color: '#d4d6e0', font: { size: 11 } } },
      } : undefined,
    },
  };
  if (_charts[canvasId]) { destroyChart(canvasId); }
  if (!_chartjs) return;
  _charts[canvasId] = new _chartjs(canvas, cfg);
  _charts[canvasId].$fullLabels = fullLabels;
  _charts[canvasId].$chartType = chartType;
}

// v46: фильтр таблицы по клику на сегмент диаграммы.
function _colIdxByName() {
  const low = _allHeaders.map(h => String(h).toLowerCase());
  for (const n of arguments) { const i = low.indexOf(n); if (i >= 0) return i; }
  return -1;
}
// v46: ЕДИНАЯ СИСТЕМА ФИЛЬТРОВ. Любой фильтр (клик по диаграмме, текст, статус)
// пересчитывает И таблицу, И ВСЕ остальные диаграммы по отфильтрованному набору.
// Фильтры КОМБИНИРУЮТСЯ (AND): клик по нескольким сегментам = пересечение.
let _activeFilters = {};   // ключ -> {label, matcher}

function _filteredRows() {
  const fs = Object.values(_activeFilters);
  if (!fs.length) return _allRows || [];
  return (_allRows || []).filter(r => fs.every(f => f.matcher(r)));
}
function recomputeView() {
  const rows = _filteredRows();
  renderTable(rows);
  const total = (_allRows || []).length;
  $('#results-count').textContent = Object.keys(_activeFilters).length
    ? `${rows.length} из ${total} строк (фильтр)` : `${total} строк`;
  if (_chartjs || window.Chart) { tryInitCharts(); buildResultsCharts(computeStatsFromRows(rows)); }
  renderFilterChips();
}
function setFilter(key, label, matcher) {
  _activeFilters[key] = { label, matcher };
  recomputeView();
}
function removeFilter(key) {
  delete _activeFilters[key];
  recomputeView();
}
function clearAllFilters() {
  _activeFilters = {};
}
function renderFilterChips() {
  let box = $('#filter-chips');
  if (!box) {
    box = document.createElement('span');
    box.id = 'filter-chips';
    box.style.cssText = 'display:inline-flex;flex-wrap:wrap;gap:6px;margin-left:8px;vertical-align:middle;';
    const cnt = $('#results-count');
    if (cnt && cnt.parentNode) cnt.parentNode.insertBefore(box, cnt.nextSibling);
  }
  const keys = Object.keys(_activeFilters);
  box.innerHTML = keys.map(k =>
    `<span class="flt-chip" data-key="${escapeHtml(k)}" title="Убрать фильтр"
       style="display:inline-flex;align-items:center;gap:6px;padding:3px 10px;border-radius:12px;
       background:var(--accent,#5b8cff);color:#fff;font-size:12px;cursor:pointer;">
       🔎 ${escapeHtml(_activeFilters[k].label)} ✕</span>`).join('') +
    (keys.length > 1 ? `<span class="flt-chip" data-key="__all__"
       style="display:inline-flex;align-items:center;padding:3px 10px;border-radius:12px;
       background:#6b7280;color:#fff;font-size:12px;cursor:pointer;">Сбросить всё ✕</span>` : '');
  box.querySelectorAll('.flt-chip').forEach(el => el.addEventListener('click', () => {
    const k = el.getAttribute('data-key');
    if (k === '__all__') { clearAllFilters(); recomputeView(); }
    else removeFilter(k);
  }));
}
// Пересчёт статистики для диаграмм из набора строк (как в Python get_results).
function computeStatsFromRows(rows) {
  const st = { by_status:{}, by_registry:{}, by_marketplace:{}, by_original:{},
               by_doc_status:{}, by_brand:{}, by_risk:{} };
  const ci = (...n) => _colIdxByName.apply(null, n);
  const iStatus = ci('технический статус','status','статус');
  const iHost   = ci('реестр (хост)','registry_host');
  const iRurl   = ci('ссылка на реестр','registry_url');
  const iMkt    = ci('marketplace','маркетплейс');
  const iPurl   = ci('ссылка на товар','product_url');
  const iOrig   = ci("плашка 'оригинал'",'is_original','оригинал');
  const iDoc    = ci('статус документа','document_status');
  const iBrand  = ci('бренд','brand');
  const iRisk   = ci('риск по сроку','риск','risk');
  const inc = (o,k) => { if (k) o[k] = (o[k]||0)+1; };
  const hostOf = u => { try { return new URL(u).hostname.replace('www.',''); } catch(e){ return u; } };
  for (const r of rows) {
    if (iStatus>=0) inc(st.by_status, String(r[iStatus]??'').trim());
    if (iHost>=0 && String(r[iHost]??'').trim()) { let v=String(r[iHost]).trim(); if(v.startsWith('http'))v=hostOf(v); inc(st.by_registry,v); }
    else if (iRurl>=0 && String(r[iRurl]??'').trim()) inc(st.by_registry, hostOf(String(r[iRurl])));
    if (iMkt>=0 && String(r[iMkt]??'').trim()) inc(st.by_marketplace, String(r[iMkt]).trim());
    else if (iPurl>=0) { const u=String(r[iPurl]??'').toLowerCase(); inc(st.by_marketplace, u.includes('wildberries')?'Wildberries':(u.includes('ozon')?'Ozon':'Не определен')); }
    if (iOrig>=0) { const v=String(r[iOrig]??'').toLowerCase().trim(); let k; if(['true','да','оригинал'].includes(v))k='Оригинал'; else if(['false','нет'].includes(v)||v.includes('не ориг'))k='Не оригинал'; else if(!v||v.includes('не указан')||v==='none')k='Не указано'; else k=String(r[iOrig]); inc(st.by_original,k); }
    if (iDoc>=0) inc(st.by_doc_status, String(r[iDoc]??'').trim());
    if (iBrand>=0) inc(st.by_brand, String(r[iBrand]??'').trim());
    if (iRisk>=0) inc(st.by_risk, String(r[iRisk]??'').trim());
  }
  const be = Object.entries(st.by_brand).sort((a,b)=>b[1]-a[1]);
  if (be.length > 12) { const top={}; be.slice(0,11).forEach(([k,v])=>top[k]=v); const rest=be.slice(11).reduce((s,[,v])=>s+v,0); if(rest>0)top['Прочее']=rest; st.by_brand=top; }
  return st;
}
function _matchExact(names, lbl) {
  const ci = _colIdxByName.apply(null, names);
  const t = String(lbl).trim();
  return r => ci >= 0 && String(r[ci] ?? '').trim() === t;
}
function _matchOriginal(lbl) {
  const ci = _colIdxByName("плашка 'оригинал'", 'is_original', 'оригинал');
  const L = String(lbl).toLowerCase();
  return r => {
    if (ci < 0) return false;
    const v = String(r[ci] ?? '').toLowerCase().trim();
    if (L.includes('не указан') || L === 'none') return !v || v.includes('не указан') || v === 'none';
    if (L.includes('не ориг')) return v.includes('не ') ;
    if (L.includes('ориг')) return v.includes('ориг') && !v.includes('не ');
    return v === L;
  };
}
function _matchRegistry(lbl) {
  const ciHost = _colIdxByName('реестр (хост)', 'registry_host');
  const ciUrl  = _colIdxByName('ссылка на реестр', 'registry_url');
  const t = String(lbl).trim();
  return r => {
    if (ciHost >= 0 && String(r[ciHost] ?? '').trim()) return String(r[ciHost]).trim() === t;
    if (ciUrl >= 0) return String(r[ciUrl] ?? '').includes(t);
    return false;
  };
}
function _matchMarketplace(lbl) {
  const ciM = _colIdxByName('marketplace', 'маркетплейс');
  const ciU = _colIdxByName('ссылка на товар', 'product_url');
  const L = String(lbl).toLowerCase();
  const key = L.includes('ozon') ? 'ozon' : (L.includes('wild') || L === 'wb' ? 'wildberries' : L);
  return r => {
    if (ciM >= 0 && String(r[ciM] ?? '').trim()) return String(r[ciM]).toLowerCase().includes(L);
    if (ciU >= 0) return String(r[ciU] ?? '').toLowerCase().includes(key);
    return false;
  };
}

function buildResultsCharts(stats) {
  tryInitCharts();
  if (!_chartjs) {
    // Chart.js ещё не подгрузился (редкий CDN-fallback) — повторим чуть позже,
    // чтобы графики всё-таки нарисовались, а не остались пустыми.
    if (stats) {
      _pendingChartStats = stats;
      if (!_chartRetryTimer) {
        _chartRetryTimer = setInterval(() => {
          tryInitCharts();
          if (_chartjs) {
            clearInterval(_chartRetryTimer); _chartRetryTimer = null;
            const s = _pendingChartStats; _pendingChartStats = null;
            if (s) buildResultsCharts(s);
          }
        }, 400);
      }
    }
    return;
  }
  const _draw = (fn) => { try { fn(); } catch (e) { console && console.warn && console.warn('chart', e); } };
  _draw(() => updateDonutChart('res-chart-status', 'res-chart-status-fb', stats.by_status || {}, STATUS_COLORS,
    { onSegmentClick: l => setFilter('chart:status', 'Статус: ' + l, _matchExact(['технический статус', 'status'], l)) }));
  _draw(() => updateDonutChart('res-chart-registry', 'res-chart-registry-fb', stats.by_registry || {}, REGISTRY_COLORS,
    { onSegmentClick: l => setFilter('chart:registry', 'Реестр: ' + l, _matchRegistry(l)) }));
  _draw(() => updateDonutChart('res-chart-marketplace', 'res-chart-marketplace-fb', stats.by_marketplace || {}, MKT_COLORS,
    { onSegmentClick: l => setFilter('chart:marketplace', 'Маркетплейс: ' + l, _matchMarketplace(l)) }));
  _draw(() => updateDonutChart('res-chart-original', 'res-chart-original-fb', stats.by_original || {}, ORIGINAL_COLORS,
    { onSegmentClick: l => setFilter('chart:original', 'Оригинал: ' + l, _matchOriginal(l)) }));
  _draw(() => updateDonutChart('res-chart-risk', 'res-chart-risk-fb', stats.by_risk || {}, RISK_COLORS,
    { onSegmentClick: l => setFilter('chart:risk', 'Риск: ' + l, _matchExact(['риск по сроку'], l)) }));
  _draw(() => updateDonutChart('res-chart-brand', 'res-chart-brand-fb', stats.by_brand || {}, null,
    { type: 'bar', maxLabel: 24, onSegmentClick: l => setFilter('chart:brand', 'Бренд: ' + l, _matchExact(['бренд', 'brand'], l)) }));
}
let _pendingChartStats = null;
let _chartRetryTimer = null;

// Кэш последней статистики для вкладки «Очередь»
let _lastQueueStats = null;
let _lastRunMode = '';
async function refreshQueueCharts() {
  // Гружаем статистику из файла-результата (если он есть)
  try {
    const res = await window.pywebview.api.get_results(null, 5000);
    if (res && res.ok && res.stats) {
      _lastQueueStats = res.stats;
    } else {
      _lastQueueStats = null;
    }
  } catch (e) { _lastQueueStats = null; }
  const st = _lastQueueStats || { by_status: {}, by_registry: {}, by_marketplace: {}, by_original: {} };
  // Маркетплейс: если в выгрузке нет распределения — выводим по режиму прогона.
  let mkt2 = st.by_marketplace || {};
  if (!Object.keys(mkt2).length) {
    const totalSt = Object.values(st.by_status || {}).reduce((a, b) => a + b, 0);
    if (totalSt > 0) {
      mkt2 = ((_lastRunMode || '').toLowerCase().indexOf('ozon') >= 0)
        ? { 'Ozon': totalSt } : { 'Wildberries': totalSt };
    }
  }
  updateDonutChart('chart-status',      'chart-status-fb',      st.by_status      || {}, STATUS_COLORS);
  updateDonutChart('chart-registry',    'chart-registry-fb',    st.by_registry    || {}, REGISTRY_COLORS);
  updateDonutChart('chart-marketplace', 'chart-marketplace-fb', mkt2, MKT_COLORS);
  updateDonutChart('chart-original',    'chart-original-fb',    st.by_original    || {}, ORIGINAL_COLORS);
}

// Activity sparkline (mini canvas)
function drawSparkline(series) {
  const canvas = $('#activity-canvas');
  if (!canvas) return;
  const wrap = canvas.parentElement;
  canvas.width  = wrap.offsetWidth  || 400;
  canvas.height = wrap.offsetHeight || 80;
  const ctx = canvas.getContext('2d');
  ctx.clearRect(0, 0, canvas.width, canvas.height);
  if (!series.length) return;
  const max = Math.max(...series, 1);
  const w = canvas.width;
  const h = canvas.height;
  const step = w / Math.max(series.length - 1, 1);

  // Fill
  const grad = ctx.createLinearGradient(0, 0, 0, h);
  grad.addColorStop(0, 'rgba(91,140,255,0.25)');
  grad.addColorStop(1, 'rgba(91,140,255,0)');
  ctx.beginPath();
  series.forEach((v, i) => {
    const x = i * step;
    const y = h - (v / max) * (h - 4) - 2;
    i === 0 ? ctx.moveTo(x, y) : ctx.lineTo(x, y);
  });
  ctx.lineTo((series.length - 1) * step, h);
  ctx.lineTo(0, h);
  ctx.closePath();
  ctx.fillStyle = grad;
  ctx.fill();

  // Line
  ctx.beginPath();
  series.forEach((v, i) => {
    const x = i * step;
    const y = h - (v / max) * (h - 4) - 2;
    i === 0 ? ctx.moveTo(x, y) : ctx.lineTo(x, y);
  });
  ctx.strokeStyle = '#5b8cff';
  ctx.lineWidth = 2;
  ctx.stroke();
}

// Загрузка Chart.js. v27.7: библиотека встроена локально (offline) через
// _build_frontend_html на стороне Python — поэтому window.Chart обычно уже
// доступен сразу. CDN остаётся лишь аварийным fallback'ом, если локальный
// файл vendor/chart.umd.min.js по какой-то причине не встроился.
function loadChartJs() {
  return new Promise((resolve, reject) => {
    if (window.Chart) { resolve(); return; }
    const s = document.createElement('script');
    s.src = 'https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js';
    s.onload  = resolve;
    s.onerror = reject;
    document.head.appendChild(s);
  });
}

// ============================================================
// Init
// ============================================================
(async function init() {
  await waitForApi();

  _cachedSettings = await window.pywebview.api.get_settings();
  renderForm();

  // Восстановить тему
  if (_cachedSettings.theme === 'light') {
    document.documentElement.setAttribute('data-theme', 'light');
    $$('.theme-btn').forEach(b => b.classList.toggle('active', b.dataset.theme === 'light'));
  }

  await fillSettings();

  // Восстановить последний режим/маркет
  const last = _cachedSettings.last_spec || {};
  if (last.mode) {
    if (last.mode === 'ozon') { setMkt('ozon'); }
    else if (last.mode === 'unified') { setMkt('both'); }
    else { setMkt('wb'); setWbMode(last.mode); }
  }

  // Начальный опрос
  await pollOnce();

  // Загружаем Chart.js в фоне — не блокируем UI
  loadChartJs().then(() => {
    tryInitCharts();
  }).catch(() => {
    // Без интернета — charts работать не будут, показываем fallback
  });

  // Polling каждые 600 мс
  _pollInterval = setInterval(pollOnce, 600);
})();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Проверка движков при запуске
# ---------------------------------------------------------------------------

def check_engines() -> None:
    """Печатает в stderr статус наличия каждого движка."""
    engines = [
        (ENGINE_V39,         "WB Query   (main_v39.py)"),
        (ENGINE_BRAND,       "WB Brand   (main_brand.py)"),
        (ENGINE_OZON,        "Ozon       (ozon_parser.py)"),
        (ENGINE_WB_ENHANCED, "WBEnhanced (wb_enhanced.py)"),
        (ENGINE_FSA_ENHANCED,"FSAEnhanced(fsa_enhanced.py)"),
    ]
    log.info("=" * 60)
    log.info("WB+Ozon Checker v27 — проверка движков")
    log.info("  Рабочая папка: %s", APP_DIR)
    for path, label in engines:
        status = "OK" if path.exists() else "ОТСУТСТВУЕТ"
        log.info("  [%s] %s", status.ljust(10), label)
    log.info("=" * 60)


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------

def check_python_deps() -> List[str]:
    """Проверяет наличие критичных Python-зависимостей. Возвращает список отсутствующих."""
    missing: List[str] = []
    for mod, pkg in [
        ("webview", "pywebview"),
        ("openpyxl", "openpyxl"),
        ("aiohttp", "aiohttp"),
        ("bs4", "beautifulsoup4"),
        ("lxml", "lxml"),
        ("curl_cffi", "curl_cffi"),
        ("playwright", "playwright"),
    ]:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    return missing


def _build_frontend_html() -> str:
    """
    Возвращает HTML интерфейса со ВСТРОЕННОЙ библиотекой Chart.js (offline).

    v27.7: раньше Chart.js грузился с интернет-CDN (cdn.jsdelivr.net) — без
    сети графики молча не рисовались (главная причина жалоб на интерфейс).
    Теперь локальный vendor/chart.umd.min.js инлайнится прямо в HTML, и графики
    работают полностью офлайн. Если файла нет — остаётся CDN-fallback в JS.
    """
    html = FRONTEND_HTML
    try:
        chart_path = APP_DIR / "vendor" / "chart.umd.min.js"
        if chart_path.exists():
            js = chart_path.read_text(encoding="utf-8")
            # Закрывающий </script> внутри библиотеки маловероятен, но на всякий
            # случай экранируем, чтобы не разорвать наш inline-блок.
            js = js.replace("</script>", "<\\/script>")
            inline = "<script>\n" + js + "\n</script>"
            html = html.replace("<!--__CHARTJS_INLINE__-->", inline, 1)
            log.info("Chart.js встроен локально (%d КБ) — графики работают офлайн", len(js) // 1024)
        else:
            log.warning("vendor/chart.umd.min.js не найден — графики будут грузиться с CDN (нужен интернет)")
    except Exception as e:
        log.warning("Не удалось встроить Chart.js локально: %s — fallback на CDN", e)
    return html


def main() -> None:
    """Запускает приложение WB+Ozon Checker v27."""
    check_engines()
    missing = check_python_deps()
    if missing:
        log.warning("=" * 60)
        log.warning("ОТСУТСТВУЮТ ПАКЕТЫ: %s", ", ".join(missing))
        log.warning("Запустите: pip install -U %s", " ".join(missing))
        log.warning("Или дважды кликните install_windows.bat")
        log.warning("=" * 60)
    bridge = Bridge()
    bridge._missing_deps = missing
    window = webview.create_window(
        title="WB+Ozon Checker v27",
        html=_build_frontend_html(),
        js_api=bridge,
        width=1400,
        height=900,
        min_size=(1100, 720),
        background_color="#0a0a0f",
    )
    bridge._window = window
    webview.start(debug=False)


if __name__ == "__main__":
    main()
