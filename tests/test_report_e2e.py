"""
Сквозной офлайн-тест генерации Excel-отчёта WB (реальный код ResultStore).

Проверяем НЕ helper в изоляции, а весь конвейер отчёта: results / review /
summary / Сводка (с графиками) / Подробности — на «грязных» данных с
управляющими символами. Это регрессионная защита от «данные сохраняются криво»
на настоящем пути сохранения.

Запуск:  python3 tests/test_report_e2e.py
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402

import main_v39 as m  # noqa: E402

RESULTS = []


def check(name, cond, extra=""):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}  {extra}")


def make_rows():
    nasty = "Носки\x00\x07 хлопок\nстрока2\tтаб"
    return [
        m.ResultRow(
            query="носки", nm_id=198234567, product_name=nasty,
            brand="BrandX\x01", subject="Носки",
            product_url="https://www.wildberries.ru/catalog/198234567/detail.aspx",
            status="OK", price_rub=1234.5, sale_price_rub=999.0,
            seller_name="ООО\x08 Продавец", is_original="оригинал",
            registry_url="https://pub.fsa.gov.ru/rss/certificate/view/1/baseInfo",
            certificate_product_name="Изделия чулочно-носочные\x0b для взрослых",
            document_status="Действует", document_date_end="2027-01-01",
            details="ok\x00", score=82.0,
        ),
        m.ResultRow(
            query="носки", nm_id=600000123, product_name="Платье",
            brand="Y", subject="Платья",
            product_url="https://www.wildberries.ru/catalog/600000123/detail.aspx",
            status="НЕСООТВЕТСТВИЕ", price_rub=0.0,
            registry_url="", certificate_product_name="",
            document_status="", details="нет ссылки",
        ),
    ]


tmp = Path(tempfile.mkdtemp())
xlsx_path = tmp / "wb_report.xlsx"
csv_path = tmp / "wb_report.csv"

store = m.ResultStore(xlsx_path, csv_path=csv_path, expiry_warning_days=30,
                      make_report_xlsx=True)
rows = make_rows()

saved_ok = True
try:
    store._save_csv(rows, csv_path)
    store._save_xlsx(rows, xlsx_path)
except Exception as e:
    saved_ok = False
    print("save error:", repr(e))

check("отчёт сохранён без исключений (грязные данные)", saved_ok)
check("xlsx-файл создан", xlsx_path.exists())
check("csv-файл создан", csv_path.exists())

if xlsx_path.exists():
    wb = load_workbook(xlsx_path)
    names = set(wb.sheetnames)
    for sheet in ("results", "review", "summary", "Сводка", "Подробности"):
        check(f"лист '{sheet}' присутствует", sheet in names, str(sorted(names)))
    # данные читаются и control-символы вычищены
    ws = wb["results"]
    all_text = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    check("в листе results нет control-символов",
          all(ch not in all_text for ch in ("\x00", "\x07", "\x08", "\x01", "\x0b")))
    # графики на «Сводке»
    svodka = wb["Сводка"]
    check("на 'Сводке' есть графики", len(getattr(svodka, "_charts", [])) >= 1,
          f"charts={len(getattr(svodka, '_charts', []))}")

# --------------------------------------------------------------------------
# Ozon-отчёт тем же сценарием (был ключевой жалобой)
import ozon_parser as oz  # noqa: E402

oz_xlsx = tmp / "ozon_report.xlsx"
oz_rows = [
    oz.OzonResultRow(
        nm_id=123456789, product_name="Кроссовки\x00 Nike", subject="Кроссовки",
        status="OK",
        certificate_product_name="Обувь спортивная\x07 для взрослых",
        document_status="Действует", score=82.0,
        registry_url="https://pub.fsa.gov.ru/rss/certificate/view/1/baseInfo",
    ),
    oz.OzonResultRow(
        nm_id=987654321, product_name="Платье", subject="Платья",
        status="НЕСООТВЕТСТВИЕ",
    ),
]
oz_ok = True
try:
    oz.save_xlsx(oz_rows, oz_xlsx, expiry_warning_days=30, make_report_xlsx=True)
except Exception as e:
    oz_ok = False
    print("ozon save error:", repr(e))
check("Ozon-отчёт сохранён без исключений (грязные данные)", oz_ok)
check("Ozon xlsx создан", oz_xlsx.exists())
if oz_xlsx.exists():
    owb = load_workbook(oz_xlsx)
    otext = " ".join(
        str(c.value) for ws in owb.worksheets
        for row in ws.iter_rows() for c in row if c.value is not None
    )
    check("в Ozon-отчёте нет control-символов",
          all(ch not in otext for ch in ("\x00", "\x07")))


if __name__ == "__main__":
    passed = sum(RESULTS)
    total = len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {passed}/{total} прошло")
    sys.exit(0 if passed == total else 1)
