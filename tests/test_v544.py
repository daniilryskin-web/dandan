"""
Тест (v54.4): улучшения №2, №7, №8, №9, №11, №12.

  №2  — вырезание ссылок на стандарты (ГОСТ/ТР ТС) из наименования продукции;
  №7  — прогресс с ETA/скоростью (формат @@PROGRESS@@ speed= eta=);
  №8  — группировка причин «ПРОВЕРИТЬ ВРУЧНУЮ» (by_review_reason);
  №9  — загрузка пользовательского словаря + «обучить словарь» (suggest_dictionary);
  №11 — расширенная карта ТН ВЭД + загрузка из файла;
  №12 — сводка по продавцам/брендам (export_supplier_stats).

Запуск:  python3 tests/test_v544.py
"""
import sys
import tempfile
import types
import csv
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.modules.setdefault("webview", types.ModuleType("webview"))

import main_v39 as m  # noqa: E402
import wb_checker as w  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

RESULTS = []


def check(name, cond):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}")


def _make(path, rows, headers=None):
    wb = Workbook(); ws = wb.active; ws.title = "Подробности"
    ws.append(headers or ["Артикул", "Название товара", "Бренд", "Продавец",
                          "Технический статус", "Название в реестре", "Примечания",
                          "Ссылка на товар (WB)", "Ссылка на реестр"])
    for r in rows:
        ws.append(r)
    wb.save(str(path))


def main():
    # --- №2: вырезание ссылок на стандарты ---
    g = 'ГОСТ CISPR 14-2-2016 "Электромагнитная совместимость. Требования для бытовых приборов"'
    check("№2: ГОСТ-номер вырезан, описание осталось",
          "гост" not in m._strip_standard_refs(g).lower() and "бытовых приборов" in m._strip_standard_refs(g))
    check("№2: ТР ТС вырезан", "тр тс" not in m._strip_standard_refs("ТР ТС 020/2011 Электромагнитная совместимость").lower())
    check("№2: обычное название не тронуто",
          m._strip_standard_refs("Аэрогриль, модель FF-01") == "Аэрогриль, модель FF-01")
    check("№2: аэрогриль vs ГОСТ-сертификат → OK (не ручная)",
          m.compare_product_names("Аэрогриль электрический 8 л", g, "", "", "Действует")[0] == "OK")

    # --- №11: расширенная карта ТН ВЭД + загрузка из файла ---
    for code, exp in [("0401", "food"), ("9506", "sport_equipment"), ("8201", "tools"),
                      ("4011", "auto"), ("2309", "pet"), ("9608", "stationery"),
                      ("7201", "")]:
        check(f"№11: ТН ВЭД {code} → {exp or '∅'}", m._tnved_category(code + "000000") == exp)
    with tempfile.TemporaryDirectory() as td:
        tp = Path(td) / "t.csv"
        csv.writer(open(tp, "w", newline="", encoding="utf-8")).writerows([["9999", "toys"]])
        check("№11: загрузка доп. карты ТН ВЭД из CSV", m.load_tnved_map(tp) == 1)
        check("№11: загруженный префикс работает", m._tnved_category("9999000000") == "toys")

    # --- №9: загрузка пользовательского словаря ---
    with tempfile.TemporaryDirectory() as td:
        dp = Path(td) / "d.csv"
        csv.writer(open(dp, "w", newline="", encoding="utf-8")).writerows(
            [["вейп", "electronics"], ["# комментарий", "x"]])
        check("№9: загрузка словаря из CSV (без комментариев)", m.load_user_dictionary(dp) == 1)
        check("№9: новое слово распознаётся", "electronics" in m._detect_categories("Вейп одноразовый"))

    # --- №7: формат прогресса со скоростью/ETA ---
    src = (Path(__file__).resolve().parents[1] / "main_v39.py").read_text(encoding="utf-8")
    check("№7: emit_progress печатает speed= и eta=", "speed=" in src and "eta=" in src)
    gui = (Path(__file__).resolve().parents[1] / "wb_checker.py").read_text(encoding="utf-8")
    check("№7: GUI парсит speed/eta и показывает ETA", "progress_eta" in gui and "kpi-eta" in gui)

    # --- бэкенд-методы №8/№9/№12 на синтетике ---
    b = w.Bridge()
    gen = []
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "result.xlsx"
        _make(p, [
            [1, "Аэрогриль 7л", "BR1", "ООО Альфа", "OK", "прибор", "ok", "u", "r"],
            [2, "Квазиблабер зугвальд", "BR2", "ООО Альфа", "ПРОВЕРИТЬ ВРУЧНУЮ", "x",
             "Низкое совпадение без доказанного конфликта", "u", "r"],
            [3, "Платье", "BR3", "ИП Бета", "НЕСООТВЕТСТВИЕ", "игрушки",
             "conflicts=категория карточки не совпадает", "u", "r"],
            [4, "Квазиблабер прибор2", "BR2", "ООО Альфа", "ПРОВЕРИТЬ ВРУЧНУЮ", "y",
             "Низкое совпадение", "u", "r"],
        ])
        # №12
        r12 = b.export_supplier_stats(str(p))
        check("№12: сводка по продавцам сформирована", r12.get("ok") and r12["rows"] == 2)
        if r12.get("ok"):
            gen.append(r12["path"])
            ws = load_workbook(r12["path"]).active
            hdr = [c.value for c in ws[1]]
            check("№12: есть колонки Проблемных/% проблемных",
                  "Проблемных" in hdr and "% проблемных" in hdr)
        # №9 suggest
        r9 = b.suggest_dictionary(str(p))
        check("№9: suggest_dictionary нашёл незнакомые слова", r9.get("ok") and r9["rows"] >= 1)
        if r9.get("ok"):
            gen.append(r9["path"])
            ws = load_workbook(r9["path"]).active
            words = [str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True)]
            check("№9: незнакомое слово попало в кандидаты", any("квазиблабер" in x for x in words))
        # №8 review reasons
        res = b.get_results(str(p))
        rr = res["stats"].get("by_review_reason", {})
        check("№8: by_review_reason посчитан (2 низкого совпадения)",
              rr.get("нет категории / низкое совпадение") == 2)

    for f in gen:
        try:
            Path(f).unlink()
        except Exception:
            pass
    check("№8/№9/№12: кнопки в окне есть",
          "btn-supplier-stats" in gui and "btn-suggest-dict" in gui and "review-reasons" in gui)


if __name__ == "__main__":
    main()
    p, t = sum(RESULTS), len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {p}/{t} прошло")
    sys.exit(0 if p == t else 1)
