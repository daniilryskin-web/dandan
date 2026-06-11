#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WB certificate checker strict v38.5 browser registry stage

Главная идея:
1) nm_id собираются быстро через JSON-выдачу WB.
2) registry_url собирается ТОЛЬКО строгим путём:
   карточка -> "Характеристики и описание" -> "Документы проверены" -> "Смотреть на сайте".
3) Принимаются только разрешённые домены реестров.
4) Есть autosave/resume/watchdog, чтобы ночной запуск не пропадал.
5) browser_count теперь реально используется: N браузеров + M контекстов/страниц, а не один Chromium на каждый воркер.
6) v37: убраны диалоги/response-body memory leak, добавлен supervisor воркеров и восстановление потерянных карточек.
7) v38: усилен второй этап: SWIS/TULPAR field parser, ФСА/SWIS table parser, dedupe parsing registry URLs.
8) v38.2: не трогая сбор ссылок, исправлен парсер SWIS (только два продуктовых поля) и ФСА, усилено сравнение.
9) v38.3: не трогая сбор ссылок, ФСА парсится по точным полям: «Наименование (обозначение) продукции», «Регистрационный номер сертификата», «Регистрационный номер декларации о соответствии».
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import datetime as dt
import html
import json
import math
import os
import random
import re
import sys
import time
import traceback
import logging as _logging

# v27.9.x: модульный логгер. Раньше код местами вызывал log.warning(...), но
# `log` нигде не определялся -> при ЛЮБОМ исключении в построении «Сводки»
# падало с NameError и рушило сохранение всего файла. Теперь логгер есть.
log = _logging.getLogger("wb_registry")


class _SkipSecondPass(Exception):
    """v27.9.x: внутренний сигнал — пропустить второй проход FSA (он теперь по кнопке)."""
import difflib
import urllib.parse
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import urlparse, parse_qs, unquote

import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None

try:
    from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
except Exception:
    async_playwright = None
    PlaywrightTimeoutError = TimeoutError

# =============================================================================
# v39.15: ВСТРОЕННЫЙ HTTP-парсер FSA (раньше был отдельным файлом fsa_http_fetcher.py).
# Теперь main_v39.py самодостаточна — внешний файл не нужен.
# FSA блокирует обычные HTTP-клиенты по TLS-fingerprint; обойти можно только
# через curl_cffi (имитирует Chrome). Без `pip install curl_cffi` функции вернут None
# и программа откатится на браузерный путь.
# =============================================================================
try:
    from curl_cffi import requests as _curl_requests
except Exception:
    _curl_requests = None

DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"
)


# Кэш по (kind, doc_id) чтобы не запрашивать один и тот же документ многократно.
_FSA_CACHE: Dict[str, Dict[str, str]] = {}


def is_curl_cffi_available() -> bool:
    """Главная проверка. Без curl_cffi функции вернут None."""
    return _curl_requests is not None


def extract_fsa_kind_id(url: str) -> Tuple[str, str]:
    """Разбирает FSA-URL вида
       https://pub.fsa.gov.ru/rss/certificate/view/{id}/baseInfo
       https://pub.fsa.gov.ru/rds/declaration/view/{id}/common
    и возвращает ('rss_certificate' | 'rds_declaration', '{id}').
    Возвращает ('', '') если URL не FSA или формат другой.
    """
    try:
        u = urlparse(url)
        if "fsa.gov.ru" not in (u.netloc or "").lower():
            return "", ""
        path = u.path or ""
        m = re.search(r"/rss/certificate/(?:view|details|card|api)?/?(\d{3,})", path)
        if m:
            return "rss_certificate", m.group(1)
        m = re.search(r"/rds/declaration/(?:view|details|card|api)?/?(\d{3,})", path)
        if m:
            return "rds_declaration", m.group(1)
        # API URLs тоже могут прийти на вход
        m = re.search(r"/api/v\d+/rss/(?:common/)?certificates?/(\d{3,})", path)
        if m:
            return "rss_certificate", m.group(1)
        m = re.search(r"/api/v\d+/rds/(?:common/)?declarations?/(\d{3,})", path)
        if m:
            return "rds_declaration", m.group(1)
    except Exception:
        pass
    return "", ""


def _fsa_browser_headers(referer: str, *, origin: bool = False) -> Dict[str, str]:
    """Заголовки скопированные с реального Chrome из HAR. Не отправляем Origin/XHR
    для GET-XHR — иначе FSA возвращает 403."""
    h = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Referer": referer,
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": DEFAULT_UA,
        "lkId": "",
        "orgId": "",
        "sec-ch-ua": '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }
    if origin:
        h["Origin"] = "https://pub.fsa.gov.ru"
    return h


def _fsa_html_headers(referer: str) -> Dict[str, str]:
    h = _fsa_browser_headers(referer)
    h["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    h["Sec-Fetch-Dest"] = "document"
    h["Sec-Fetch-Mode"] = "navigate"
    h["Sec-Fetch-Site"] = "none"
    return h


def _fsa_warmup_urls(kind: str, doc_id: str, referer: str) -> List[Tuple[str, str, str]]:
    """Warm-up последовательность из HAR — без неё FSA может вернуть 403."""
    base = "rss" if kind == "rss_certificate" else "rds"
    return [
        ("doc_page", referer, "html"),
        ("i18n_ru", "https://pub.fsa.gov.ru/assets/i18n/ru.json", "json"),
        ("lk_account", "https://pub.fsa.gov.ru/lk/api/account", "json"),
        ("common_account", f"https://pub.fsa.gov.ru/api/v1/{base}/common/account", "json"),
    ]


def _fsa_candidates(kind: str, doc_id: str, aggressive: bool = False) -> List[Tuple[str, str]]:
    """Список (label, url) для перебора. Сначала подтверждённый из HAR."""
    if kind == "rss_certificate":
        out = [("api_rss_common_certificates_id_har",
                f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{doc_id}")]
        if aggressive:
            out += [
                ("api_rss_common_certificate_id",
                 f"https://pub.fsa.gov.ru/api/v1/rss/common/certificate/{doc_id}"),
            ]
        return out
    if kind == "rds_declaration":
        out = [("api_rds_common_declarations_id_har",
                f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{doc_id}")]
        if aggressive:
            out += [
                ("api_rds_common_declaration_id",
                 f"https://pub.fsa.gov.ru/api/v1/rds/common/declaration/{doc_id}"),
            ]
        return out
    return []


# =============================================================================
# Парсинг JSON ответа FSA
# =============================================================================

def _unwrap_payload(obj: Any, max_depth: int = 4) -> Any:
    """FSA иногда оборачивает payload в {data: ...} / {item: ...}."""
    cur = obj
    for _ in range(max_depth):
        if isinstance(cur, dict):
            for k in ("data", "item", "result", "content", "payload", "declaration", "certificate"):
                v = cur.get(k)
                if isinstance(v, (dict, list)):
                    cur = v
                    break
            else:
                return cur
        elif isinstance(cur, list) and len(cur) == 1 and isinstance(cur[0], dict):
            cur = cur[0]
        else:
            return cur
    return cur


def _walk_leaves(obj: Any, path: Tuple[str, ...] = ()) -> List[Tuple[Tuple[str, ...], Any]]:
    """Глубокий обход JSON, возвращает все листовые значения с путями."""
    out: List[Tuple[Tuple[str, ...], Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            out.extend(_walk_leaves(v, path + (str(k),)))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            out.extend(_walk_leaves(v, path + (f"[{i}]",)))
    else:
        out.append((path, obj))
    return out


def _find_value(leaves, include_any, include_all=(), exclude_any=()):
    """Найти первое значение чей путь содержит ключевые слова."""
    inc_any = [s.lower() for s in include_any]
    inc_all = [s.lower() for s in include_all]
    exc_any = [s.lower() for s in exclude_any]
    for path, v in leaves:
        if v in (None, "", []):
            continue
        pl = "/".join(path).lower()
        if any(x in pl for x in exc_any):
            continue
        if inc_all and not all(x in pl for x in inc_all):
            continue
        if inc_any and not any(x in pl for x in inc_any):
            continue
        if isinstance(v, (dict, list)):
            continue
        s = str(v).strip()
        if s and len(s) < 2000:
            return s
    return ""


def _format_date(s: Any) -> str:
    """ISO-датировки → DD.MM.YYYY."""
    if not s:
        return ""
    s = str(s).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"
    return s[:10]


def parse_fsa_json(obj: Any, url: str, kind: str, doc_id: str) -> Dict[str, str]:
    """Достаёт ключевые поля из JSON-ответа FSA. Возвращает dict с registry_*.

    v27.9.x: разбор по РЕАЛЬНОЙ структуре API ФСА (подтверждена живым ответом).
    Реальный JSON вложенный и кодирует status/scheme/техрегламент числами, а
    название/изготовителя держит в под-объектах (product.fullName,
    manufacturer.fullName и т.п.) — наивный поиск по подстроке ключа брал не те
    поля (idLegalSubject вместо fullName, idProduct вместо названия). Поэтому
    сначала пытаемся СТРОГО по структуре, и только то, что не нашли — добираем
    старым обобщённым поиском.
    """
    payload = _unwrap_payload(obj)
    is_cert = (kind == "rss_certificate")
    out: Dict[str, str] = {
        "doc_type": "Сертификат" if is_cert else "Декларация",
        "source": f"json:{url}",
    }

    if isinstance(payload, dict):
        def _org_name(d: Any) -> str:
            if isinstance(d, dict):
                return str(d.get("fullName") or d.get("shortName") or "").strip()
            return ""

        def _org_inn(d: Any) -> str:
            if isinstance(d, dict):
                return str(d.get("inn") or "").strip()
            return ""

        prod = payload.get("product") if isinstance(payload.get("product"), dict) else {}
        applicant = payload.get("applicant")
        manufacturer = payload.get("manufacturer")

        # Номер документа
        if payload.get("number"):
            out["doc_number"] = str(payload["number"]).strip()
        if payload.get("blankNumber"):
            out["blank_number"] = str(payload["blankNumber"]).strip()

        # Статус (числовой код idStatus -> текст). 6 = «Действует» (подтверждено).
        st_id = payload.get("idStatus")
        if st_id is None:
            chg = payload.get("statusChanges")
            if isinstance(chg, list) and chg and isinstance(chg[-1], dict):
                st_id = chg[-1].get("idStatus")
        if st_id is not None:
            out["status"] = _FSA_STATUS_MAP.get(int(st_id), f"Статус {st_id}") \
                if isinstance(st_id, (int, float)) or str(st_id).isdigit() else str(st_id)
            out["status_code"] = str(st_id)

        # Даты
        ds = _format_date(payload.get("certRegDate") or payload.get("declRegDate")
                          or payload.get("regDate"))
        de = _format_date(payload.get("certEndDate") or payload.get("declEndDate")
                          or payload.get("endDate"))
        if ds:
            out["date_start"] = ds
        if de:
            out["date_end"] = de

        # Заявитель / изготовитель — берём имя из под-объекта (НЕ id!)
        if _org_name(applicant):
            out["applicant"] = _org_name(applicant)
        if _org_inn(applicant):
            out["applicant_inn"] = _org_inn(applicant)
        if _org_name(manufacturer):
            out["manufacturer"] = _org_name(manufacturer)
        if _org_inn(manufacturer):
            out["manufacturer_inn"] = _org_inn(manufacturer)

        # Название продукции. ФСА держит ДВА названия:
        #   • product.fullName — «Общее наименование продукции» (часто общее, напр.
        #     «Электрические приборы бытового назначения»);
        #   • product.identifications[].name — «Наименование (обозначение) продукции»
        #     (конкретика: «электрические чайники, торговой марки …, модель …»).
        # Раньше брали только fullName → общее имя не совпадало с названием WB и
        # давало ложное «НЕСООТВЕТСТВИЕ». Собираем ОБА поля в одно через «; »
        # (с дедупликацией: если они совпадают — не дублируем).
        _name_parts: List[str] = []
        if prod.get("fullName"):
            _name_parts.append(str(prod["fullName"]).strip())
        _idents = prod.get("identifications")
        if isinstance(_idents, list):
            for _it in _idents:
                if isinstance(_it, dict) and _it.get("name"):
                    _name_parts.append(str(_it["name"]).strip())
        _seen_names: List[str] = []
        for _p in _name_parts:
            if _p and _p not in _seen_names:
                _seen_names.append(_p)
        if _seen_names:
            out["product_full"] = "; ".join(_seen_names)

        # Схема. У СЕРТИФИКАТОВ idCertScheme = реальный номер схемы (1..7 -> «1с»).
        # У ДЕКЛАРАЦИЙ idDeclScheme — это ВНУТРЕННИЙ id записи (напр. 3581), а НЕ номер
        # схемы. Поэтому код берём только если это настоящий однозначный/двузначный
        # номер схемы (1..9). Для деклараций реальную схему («1д») добираем из ТЕКСТА
        # страницы (см. _parse_fsa_with_existing_page_v386).
        sch = payload.get("idCertScheme") if is_cert else payload.get("idDeclScheme")
        if sch is None:
            sch = payload.get("idCertScheme") or payload.get("idDeclScheme")
        if sch is not None and str(sch).strip().isdigit() and 1 <= int(sch) <= 9:
            out["scheme"] = f"{int(sch)}{'с' if is_cert else 'д'}"

        # Технические регламенты (idTechnicalReglaments -> текст по словарю).
        # v27.9.x: словарь FSA НЕ последовательный (007/2011=id39), поэтому
        # неизвестные id НЕ выводим как «ТР (код N)» — это путало пользователя.
        # Известные id маппим; для неизвестных оставляем technical_regulation
        # пустым, чтобы его заполнил настоящий «ТР ТС NNN/YYYY» из текста
        # страницы/вкладок (см. вызов в _parse_fsa_with_existing_page_v386).
        tregs = payload.get("idTechnicalReglaments")
        if isinstance(tregs, list) and tregs:
            names = []
            for t in tregs:
                try:
                    nm = _FSA_TECHREG_MAP.get(int(t))
                    if nm:
                        names.append(nm)
                except Exception:
                    pass
            if names:
                out["technical_regulation"] = "; ".join(dict.fromkeys(names))

        # ТН ВЭД — в JSON только внутренние idTnveds (не сами коды), пропускаем.

    # Добор обобщённым поиском только тех полей, что не нашли строго.
    leaves = _walk_leaves(payload)
    if not out.get("doc_number"):
        out["doc_number"] = _find_value(leaves, ("number", "regnumber"), exclude_any=("blank", "applicant", "manufacturer"))
    if not out.get("applicant"):
        out["applicant"] = _find_value(leaves, ("applicantname",)) or _find_value(leaves, ("applicant", "fullname"), include_all=("applicant", "fullname"))
    if not out.get("manufacturer"):
        out["manufacturer"] = _find_value(leaves, ("manufacturername",)) or _find_value(leaves, ("manufacturer", "fullname"), include_all=("manufacturer", "fullname"))
    if not out.get("product_full"):
        out["product_full"] = _find_value(leaves, ("productfullname", "product_fullname"))
    # Схема/техрегламент: если в JSON они даны строкой (не числовым кодом) —
    # добираем обобщённо. exclude id*, чтобы не схватить idCertScheme и пр.
    if not out.get("scheme"):
        out["scheme"] = _find_value(leaves, ("scheme",), exclude_any=("id", "object"))
    if not out.get("technical_regulation"):
        out["technical_regulation"] = _find_value(
            leaves, ("techreg", "technicalregulation", "technical_regulation"), exclude_any=("id",))
    # v45.5: техрегламент НАДЁЖНО достаём из ТЕКСТА самого JSON (без словаря id).
    # В ответе ФСА сам код ТР встречается в названиях документов/стандартов и тексте
    # типа сертификата: «ТР ТС 017/2011», «Технического регламента Таможенного союза
    # 007/2011», «Евразийского экономического союза 037/2016». Сканируем все строковые
    # листья и собираем нормализованные «ТР ТС/ЕАЭС NNN/YYYY».
    if not out.get("technical_regulation"):
        _tr_found: List[str] = []
        for _p, _v in leaves:
            if not isinstance(_v, str) or "/" not in _v:
                continue
            for m in re.finditer(r"ТР\s+(ТС|ЕАЭС)\s+(\d{3}/\d{4})", _v):
                _s = f"ТР {m.group(1)} {m.group(2)}"
                if _s not in _tr_found:
                    _tr_found.append(_s)
            for m in re.finditer(r"регламента\s+Таможенного\s+союза\s+(\d{3}/\d{4})", _v, re.I):
                _s = f"ТР ТС {m.group(1)}"
                if _s not in _tr_found:
                    _tr_found.append(_s)
            for m in re.finditer(r"(?:Евразийского\s+экономического\s+союза|ЕАЭС)\s+(\d{3}/\d{4})", _v, re.I):
                _s = f"ТР ЕАЭС {m.group(1)}"
                if _s not in _tr_found:
                    _tr_found.append(_s)
        if _tr_found:
            out["technical_regulation"] = "; ".join(_tr_found)
    if not out.get("status"):
        out["status"] = _find_value(leaves, ("status",), exclude_any=("id", "change", "history", "date"))

    return {k: v for k, v in out.items() if v}


# Словарь статусов ФСА (idStatus). Коды подтверждены пользователем по карточкам
# реестра. ВАЖНО для вердикта: действующий — только «Действует» (6); остальные
# (приостановлен/прекращён/недействителен/архивный) = документ НЕ действует.
_FSA_STATUS_MAP: Dict[int, str] = {
    1: "Архивный",
    6: "Действует",
    11: "Недействителен",
    14: "Прекращён",
    15: "Приостановлен",
}
# Статусы, при которых документ считается НЕдействующим (для вердикта/окраски).
_FSA_INACTIVE_STATUS_CODES = {1, 11, 14, 15}


# =============================================================================
# v46: Статус КИРГИЗСКИХ документов на территории РФ
# =============================================================================
# Пользователь ведёт таблицу (xlsx) с киргизскими (ЕАЭС KG…) документами, которые
# приостановлены/прекращены именно на территории РФ. Колонки: number, [reg_date],
# id_status_in_rf (14 = прекращён, 15 = приостановлен). Программа сверяет номер
# документа из реестра с этой таблицей и, если документ там есть, выставляет
# колонку «Статус на территории РФ» и итоговый вердикт «НЕДЕЙСТВУЕТ В РФ».
STATUS_INVALID_IN_RF = "НЕДЕЙСТВУЕТ В РФ"
_KG_RF_STATUS_MAP: Dict[str, int] = {}     # norm_number -> код (14/15/…)
_KG_RF_CODE_TEXT = {14: "Прекращён", 15: "Приостановлен"}


def _norm_doc_number(s: Any) -> str:
    """Нормализация номера документа для сверки: убираем всё кроме букв/цифр,
    приводим к нижнему регистру. «ЕАЭС KG 417/043.RU.02.09525» и
    «ЕАЭС KG417/043.RU.02.09525.» дают одинаковый ключ."""
    return re.sub(r'[^0-9a-zA-Zа-яА-ЯёЁ]', '', str(s or '')).lower()


def load_kg_rf_status(path: Any) -> int:
    """Загружает таблицу статусов КГ-документов в РФ в _KG_RF_STATUS_MAP.
    Поддерживает .xlsx (колонки number/…/id_status_in_rf — по именам или по
    позиции 0 и последняя) и .csv. Возвращает число загруженных записей."""
    global _KG_RF_STATUS_MAP
    p = Path(str(path))
    if not p.exists():
        return 0
    rows_iter = []
    try:
        if p.suffix.lower() in ('.xlsx', '.xlsm'):
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            rows_iter = list(ws.iter_rows(values_only=True))
        else:
            import csv as _csv
            with open(p, 'r', encoding='utf-8-sig', newline='') as f:
                rows_iter = [tuple(r) for r in _csv.reader(f)]
    except Exception as e:
        print(f"⚠️  Не удалось прочитать таблицу статусов КГ-РФ ({p.name}): {type(e).__name__}: {e}")
        return 0
    if not rows_iter:
        return 0
    header = [str(c).strip().lower() if c is not None else '' for c in rows_iter[0]]
    # ищем колонки number и id_status_in_rf по именам; иначе 0 и последняя
    num_i = next((i for i, h in enumerate(header) if 'number' in h or 'номер' in h), 0)
    st_i = next((i for i, h in enumerate(header)
                 if 'status' in h or 'статус' in h), len(header) - 1)
    has_header = ('number' in header[num_i]) or ('номер' in header[num_i]) or \
                 ('status' in header[st_i]) or ('статус' in header[st_i])
    data = rows_iter[1:] if has_header else rows_iter
    m: Dict[str, int] = {}
    for r in data:
        if not r or num_i >= len(r) or st_i >= len(r):
            continue
        key = _norm_doc_number(r[num_i])
        if not key:
            continue
        try:
            code = int(str(r[st_i]).strip())
        except Exception:
            continue
        m[key] = code
    _KG_RF_STATUS_MAP = m
    return len(m)


def kg_rf_status_text(cert_number: Any) -> str:
    """Если номер документа есть в таблице КГ-РФ — возвращает русский статус
    («Прекращён»/«Приостановлен»/«Статус N»), иначе пустую строку."""
    if not _KG_RF_STATUS_MAP or not cert_number:
        return ""
    code = _KG_RF_STATUS_MAP.get(_norm_doc_number(cert_number))
    if code is None:
        return ""
    return _KG_RF_CODE_TEXT.get(code, f"Статус {code}")

# Словарь технических регламентов ФСА (idTechnicalReglaments). 39 = ТР ТС
# 007/2011 подтверждён живым ответом (в тексте документов). Остальные —
# по мере подтверждения; неизвестные отдаются как «ТР (код N)».
_FSA_TECHREG_MAP: Dict[int, str] = {
    39: "ТР ТС 007/2011",
}



# =============================================================================
# Главная функция: HTTP-fetch с curl_cffi
# =============================================================================

def fetch_fsa_via_http(
    url: str,
    *,
    timeout_sec: float = 8.0,
    impersonate: str = "chrome",
    user_agent: Optional[str] = None,
    aggressive: bool = False,
    skip_warmup: bool = False,
    cookies: Optional[Dict[str, str]] = None,
) -> Optional[Dict[str, str]]:
    """Главный публичный API.

    Возвращает dict с распарсенными полями FSA-документа, либо None если:
      - URL не FSA-документ
      - curl_cffi не установлен (FSA вернёт 403)
      - все попытки вернули 403/500/пустые ответы

    v45.2: cookies — куки сессии, СНЯТЫЕ С НАСТОЯЩЕГО БРАУЗЕРА (который прошёл
    JS-антибот ФСА). С ними curl_cffi обращается к /api/v1/.../{id} напрямую и
    получает тот же JSON, что браузер — но в разы быстрее и одним лёгким запросом
    вместо полной отрисовки SPA. Это ключ к скорости без блокировок: после первого
    документа (через браузер) все остальные тянем по HTTP с этими куками.

    Использование:
        result = fetch_fsa_via_http("https://pub.fsa.gov.ru/rss/certificate/view/123/baseInfo")
        if result:
            print(result['doc_number'], result['status'])
    """
    kind, doc_id = extract_fsa_kind_id(url)
    if not kind or not doc_id:
        return None
    cache_key = f"{kind}:{doc_id}"
    if cache_key in _FSA_CACHE:
        return dict(_FSA_CACHE[cache_key])
    if _curl_requests is None:
        return None

    referer = (
        f"https://pub.fsa.gov.ru/rds/declaration/view/{doc_id}/common"
        if kind == "rds_declaration"
        else f"https://pub.fsa.gov.ru/rss/certificate/view/{doc_id}/baseInfo"
    )
    ua = user_agent or DEFAULT_UA
    timeout = max(4.0, float(timeout_sec))
    impersonates = [impersonate, "chrome120", "chrome110"]
    seen = set()
    impersonates = [x for x in impersonates if not (x in seen or seen.add(x))]
    _ck = {str(k): str(v) for k, v in (cookies or {}).items() if k}

    for imp in impersonates:
        try:
            sess = _curl_requests.Session()
            # v45.2: подкладываем браузерные куки в сессию curl_cffi — с ними API ФСА
            # отвечает там, где «голый» HTTP получает 403 (антибот ставит куку через JS).
            if _ck:
                try:
                    for _k, _v in _ck.items():
                        sess.cookies.set(_k, _v, domain="pub.fsa.gov.ru")
                except Exception:
                    pass
            # Warm-up — это критично, без него FSA часто отдаёт 403

            if not skip_warmup:
                for label, warm_url, typ in _fsa_warmup_urls(kind, doc_id, referer):
                    try:
                        wh = _fsa_html_headers(referer) if typ == "html" else _fsa_browser_headers(referer)
                        wh["User-Agent"] = ua
                        try:
                            sess.get(warm_url, headers=wh, timeout=timeout, impersonate=imp)
                        except TypeError:
                            sess.get(warm_url, headers=wh, timeout=timeout)
                    except Exception:
                        pass  # warm-up errors ignored

            for label, api_url in _fsa_candidates(kind, doc_id, aggressive):
                try:
                    h = _fsa_browser_headers(referer)
                    h["User-Agent"] = ua
                    try:
                        r = sess.get(api_url, headers=h, timeout=timeout, impersonate=imp)
                    except TypeError:
                        r = sess.get(api_url, headers=h, timeout=timeout)
                    status = int(getattr(r, "status_code", 0) or 0)
                    if status != 200:
                        continue
                    txt = getattr(r, "text", "") or ""
                    if not txt:
                        continue
                    try:
                        obj = r.json()
                    except Exception:
                        try:
                            obj = json.loads(txt)
                        except Exception:
                            continue
                    parsed = parse_fsa_json(obj, api_url, kind, doc_id)
                    if parsed and parsed.get("doc_number"):
                        # Успех — сохраняем в кэш
                        _FSA_CACHE[cache_key] = dict(parsed)
                        return parsed
                except Exception:
                    continue
        except Exception:
            continue

    return None




# -----------------------------
# Constants
# -----------------------------

STATUS_LINK_COLLECTED = "ССЫЛКА НА РЕЕСТР СОБРАНА"
# v27.9.x: реестры, которые по требованию НЕ парсим (оставляем только ссылку).
_BELGISS_EAEU_HOSTS = {
    "belgiss.by", "www.belgiss.by", "tsouz.belgiss.by",
    "portal.eaeunion.org", "eaeunion.org",
}
STATUS_NO_DOCS = "НЕТ ДОКУМЕНТОВ"
STATUS_NO_REGISTRY_LINK = "НЕТ ССЫЛКИ НА РЕЕСТР"
STATUS_TIMEOUT = "ТАЙМАУТ"
STATUS_ERROR = "ОШИБКА"

# v27.6-playwright: версия движка для шапки расширенного отчёта.
APP_VERSION = "2026-06-06-v27.6-playwright"

ALLOWED_REGISTRY_HOSTS = {
    "pub.fsa.gov.ru",
    "fsa.gov.ru",
    "swis.trade.kg",
    "trade.kg",
    "belgiss.by",
    "www.belgiss.by",
    "tsouz.belgiss.by",
    "portal.eaeunion.org",
    "eaeunion.org",
    "www.eaeunion.org",
}

BLOCKED_HOST_PARTS = (
    "wildberries",
    "wb.ru",
    "wb-bank",
    "eapteka",
    "datatracker.ietf.org",
    "ietf.org",
    "google",
    "yandex",
    "facebook",
    "vk.com",
    "t.me",
    "telegram",
    "doubleclick",
    "adservice",
)

# Реестровые URL часто появляются внутри JS/JSON и в обычных текстах.
# v39.12: исправлено экранирование, чтобы регулярка корректно ловила URLs даже когда
# вокруг них кириллица. Раньше из-за двойного бэкслэша не находила URLs в текстах типа
# "https://... и ещё https://...".
HTTP_URL_RE = re.compile(r"https?://[^\s\"'<>)\]]+", re.IGNORECASE)
UUID_RE = re.compile(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}")


# -----------------------------
# Data classes
# -----------------------------

@dataclass
class Card:
    nm_id: int
    product_name: str = ""
    brand: str = ""
    subject: str = ""
    source_query: str = ""
    product_url: str = ""
    # v39.14: дополнительные поля из WB API
    price_rub: float = 0.0
    sale_price_rub: float = 0.0
    seller_name: str = ""
    supplier_id: str = ""
    rating: float = 0.0
    feedbacks: int = 0
    pics_count: int = 0
    in_stock: int = 0
    is_original: str = ""  # v43: признак оригинальности из карточки WB
    docs_verified: str = ""  # v46: бейдж «Документ проверен WB» (Да/Нет) из card.json
    view_flags: int = 0     # v46: сырой viewFlags из поиска WB (для вывода бита «Документы проверены»)
    colors: str = ""        # v27.9.x: цвет(а) товара из карточки WB
    wb_root: str = ""       # v27.9.x: корневой ID карточки WB (группировка вариантов)

@dataclass
class ResultRow:
    query: str
    nm_id: int
    product_name: str
    brand: str
    subject: str
    product_url: str
    status: str
    # WB-данные карточки (v39.14)
    price_rub: float = 0.0
    sale_price_rub: float = 0.0
    seller_name: str = ""
    supplier_id: str = ""
    rating: float = 0.0
    feedbacks: int = 0
    is_original: str = ""  # v43: признак оригинальности
    docs_verified: str = ""  # v46: бейдж «Документ проверен WB» (Да/Нет)
    colors: str = ""        # v27.9.x: цвет(а) товара из карточки WB
    wb_root: str = ""       # v27.9.x: корневой ID карточки WB
    # Реестровая ссылка
    registry_url: str = ""
    registry_host: str = ""
    registry_record_id: str = ""
    # Поля документа из реестра
    certificate_number: str = ""
    document_type: str = ""
    document_status: str = ""  # v39.5: "Действует" / "Прекращён" / "Приостановлен" и т.п.
    rf_status: str = ""         # v46: статус киргизского документа на территории РФ
    certificate_product_name: str = ""
    # v39.14: расширенные поля из реестра (если HTTP-парсер их нашёл)
    document_date_start: str = ""
    document_date_end: str = ""
    applicant_name: str = ""
    applicant_inn: str = ""
    manufacturer_name: str = ""
    tnved: str = ""
    scheme: str = ""
    technical_regulation: str = ""
    # Сравнение
    score: float = 0.0
    details: str = ""
    worker: str = ""
    checked_at: str = ""


# -----------------------------
# Utility
# -----------------------------

def now_iso() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# v27.9.x: машиночитаемый прогресс для GUI. Раньше окно вытаскивало прогресс
# слабым regex (\d+/\d+) прямо из логов и ловило ложные срабатывания —
# «повтор 2/5», «заполнено 120/200» и т.п. дёргали полосу прогресса. Этот
# однозначный маркер парсится в первую очередь; печатается рядом с обычным
# человекочитаемым логом, поэтому логи остаются прежними.
def emit_progress(stage: str, done: int, total: int) -> None:
    try:
        print(f"@@PROGRESS@@ stage={stage} done={int(done)} total={int(total)}",
              flush=True)
    except Exception:
        pass


# v39.8: precompiled regex и LRU cache для горячих утилит.
# Эти функции вызываются десятки тысяч раз за прогон (compare_product_names
# вызывает их через _detect_*). Inline re.compile внутри них раньше ел ~3-5%
# CPU прогона. Кэширование результата для одинаковых входов даёт ещё ускорение.
from functools import lru_cache as _lru_cache

_RE_NORM_SPACE = re.compile(r"\s+")
_RE_WORD_BOUNDARY_PREFIX = re.compile(r"(?:^|[^а-яёa-z0-9])")


@_lru_cache(maxsize=8192)
def norm_text(s: str) -> str:
    # v39.8: сначала lower(), потом replace ё→е. До этого было наоборот, и заглавная «Ё»
    # не нормализовалась — «ЁЛКА» оставалось «ёлка» вместо «елка».
    return _RE_NORM_SPACE.sub(" ", (s or "").lower().replace("ё", "е")).strip()


# v25-reporting: единый ключ для строгого сравнения брендов
# (выровнено с norm_key из main_brand.py).
def norm_key_brand(s: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", norm_text(s or ""))


def brand_matches_v39(found_brand: str, wanted: str, mode: str = "any") -> bool:
    """True, если бренд карточки совпадает с искомым по выбранному режиму.

    mode='any'      — фильтр выключен (по умолчанию для запроса).
    mode='exact'    — строгое совпадение нормализованных строк.
    mode='contains' — wanted ⊂ found или found ⊂ wanted.
    Пустой wanted всегда True.
    """
    wb = norm_key_brand(wanted)
    if not wb or mode == "any":
        return True
    fb = norm_key_brand(found_brand)
    if mode == "contains":
        return wb in fb or fb in wb
    return fb == wb


# Кэш скомпилированных regex для _contains_word_token.
# Term_low фиксированных значений (название категорий) — всего ~200 уникальных,
# но без кэша они компилируются на каждый вызов.
_WORD_BOUNDARY_CACHE: Dict[str, re.Pattern] = {}


def _word_boundary_pattern(term_low: str) -> re.Pattern:
    p = _WORD_BOUNDARY_CACHE.get(term_low)
    if p is None:
        p = re.compile(r'(?:^|[^а-яёa-z0-9])' + re.escape(term_low))
        _WORD_BOUNDARY_CACHE[term_low] = p
    return p


def str_to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in {"1", "true", "yes", "y", "да", "on"}

def safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return default

def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(str(v).replace(",", "."))
    except Exception:
        return default

def product_url(nm_id: int) -> str:
    return f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx"

def hostname(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().split(":")[0].lstrip("www.")
    except Exception:
        return ""

def extract_record_id(url: str) -> str:
    u = urlparse(url)
    path = u.path
    m = UUID_RE.search(path)
    if m:
        return m.group(0)
    # FSA: /view/<id>/...
    parts = [p for p in path.split("/") if p]
    for i, p in enumerate(parts):
        if p == "view" and i + 1 < len(parts):
            return parts[i + 1]
    # fallback last number
    nums = re.findall(r"\d{4,}", path)
    return nums[-1] if nums else ""

def is_allowed_registry_url(url: str) -> bool:
    if not url:
        return False
    try:
        decoded = urllib.parse.unquote(url.strip())
        parsed = urlparse(decoded)
        if parsed.scheme not in {"http", "https"}:
            return False
        h = parsed.netloc.lower().split(":")[0].lstrip("www.")
        full = (h + parsed.path).lower()
        if any(b in full for b in BLOCKED_HOST_PARTS):
            # важно: fsa.gov.ru содержит ".gov", не блокируем
            if "pub.fsa.gov.ru" not in h and "fsa.gov.ru" not in h:
                return False
        if h == "pub.fsa.gov.ru":
            # принимаем только страницы реестров/сертификатов/деклараций/API ФСА
            p = parsed.path.lower()
            return any(x in p for x in (
                "/rds/declaration/",
                "/rss/certificate/",
                "/api/v1/rds/",
                "/api/v1/rss/",
            ))
        if h == "fsa.gov.ru":
            return True
        if h == "swis.trade.kg":
            return parsed.path.lower().startswith("/doc/")
        if h == "trade.kg":
            return True
        if h in {"belgiss.by", "tsouz.belgiss.by", "portal.eaeunion.org", "eaeunion.org"}:
            return True
        return False
    except Exception:
        return False

def clean_url(url: str) -> str:
    url = html.unescape(urllib.parse.unquote(str(url or "").strip()))
    url = url.rstrip(").,;\"'<>")
    return url


# =============================================================================
# v39.12: HTTP fast-path через WB certificate.json
# =============================================================================
# Wildberries для каждой карточки с прикреплённым документом отдаёт JSON по адресу
# https://basket-NN.wbbasket.ru/volXXXX/partXXXXXX/{nm_id}/info/certificate.json
# где XXXX = nm_id // 100000, XXXXXX = nm_id // 1000, NN = шард basket (1..30).
# В JSON находится прямая ссылка на реестр (FSA / SWIS / Belgiss / EAEUnion).
# Это в 30-50 раз быстрее чем открывать карточку в браузере и кликать кнопки.

def wb_volume_part(nm_id: int) -> Tuple[int, int]:
    return int(nm_id) // 100000, int(nm_id) // 1000


def wb_basket_by_volume(vol: int) -> int:
    """Официальная шардировка WB media-basket по vol.
    Источник: реальные HAR-дампы. Для vol=1982 → basket-13.

    v27.9.x: таблица точна для vol≤5429. Для бОльших vol мэппинг продолжается,
    но НЕ шагом 216 — по реальным прогонам (5514→29, 6708→32, 7627→36, 8004→37)
    шаг более пологий, ≈311 vol на шард. Прежняя экстраполяция по 216 «перелетала»
    в несуществующие basket-50+ (vol≈9800 давал basket-50, а реальный — ~40),
    из-за чего запросы шли к нерезолвящимся хостам (DNS-ошибки) и карточка зря
    помечалась ОШИБКОЙ. Теперь наклон ≈311 и осторожный кап.
    """
    ranges = [
        (143, 1), (287, 2), (431, 3), (719, 4), (1007, 5),
        (1061, 6), (1115, 7), (1169, 8), (1313, 9), (1601, 10),
        (1655, 11), (1919, 12), (2045, 13), (2189, 14), (2405, 15),
        (2621, 16), (2837, 17), (3053, 18), (3269, 19), (3485, 20),
        (3701, 21), (3917, 22), (4133, 23), (4349, 24), (4565, 25),
        (4781, 26), (4997, 27), (5213, 28), (5429, 29),
    ]
    for max_vol, basket in ranges:
        if vol <= max_vol:
            return basket
    # vol > 5429 — пологое продолжение ≈311 vol/шард (подтверждено прогонами).
    extra = round((int(vol) - 5429) / 311.0)
    return min(60, 29 + extra)


def certificate_json_urls(nm_id: int, max_hosts: int = 12) -> List[str]:
    """Возвращает упорядоченный список URL'ов для перебора basket-хостов.
    Первым идёт наиболее вероятный (по vol), затем соседи (±4), затем
    исторически «населённые» шарды и полный диапазон.

    Под конкурентным перебором (см. fetch_certificate_json_for_nm) порядок не
    критичен для скорости, но primary + соседи дают самый быстрый HTTP 200 в
    типичном случае. Диапазон fill больше не обрывается на 30 — иначе новые
    высоковолюмные товары (primary>30) никогда не нашли бы свой шард.
    """
    nm = int(nm_id)
    vol, part = wb_volume_part(nm)
    primary = wb_basket_by_volume(vol)
    order: List[int] = []

    def push(b: int) -> None:
        if 1 <= b <= 199 and b not in order:
            order.append(b)

    # 1) расчётный шард и широкий круг соседей (страховка от неточности мэппинга
    #    высоких vol — там границы шардов известны хуже)
    for d in (0, -1, 1, -2, 2, -3, 3, -4, 4, -5, 5, -6, 6):
        push(primary + d)
    # 2) для высоких vol реальные шарды лежат в диапазоне ~30..40 (по прогонам
    #    встречаются вплоть до basket-37). Плотно добиваем именно его, иначе при
    #    неточном primary карточка не нашла бы свой шард среди первых кандидатов.
    if primary >= 28:
        for b in (37, 36, 38, 35, 39, 34, 40, 33, 32, 31, 30, 41, 42, 43):
            push(b)
    # 3) исторически «населённые» низкие шарды
    for b in (13, 12, 14, 15, 16, 11, 10, 17, 18, 1):
        push(b)
    # 4) добиваем остальной диапазон
    upper = max(42, primary + 6)
    for b in range(1, upper + 1):
        push(b)

    return [
        f"https://basket-{b:02d}.wbbasket.ru/vol{vol}/part{part}/{nm}/info/certificate.json"
        for b in order[:max(1, int(max_hosts))]
    ]


def extract_registry_urls_from_certificate_json(parsed: Any) -> List[str]:
    """Глубокий обход JSON-структуры certificate.json. Возвращает все найденные
    ссылки на разрешённые реестры (по is_allowed_registry_url).
    """
    found: List[str] = []
    seen: Set[str] = set()

    def add(u: str):
        u2 = clean_url(u)
        if u2 and u2 not in seen and is_allowed_registry_url(u2):
            seen.add(u2)
            found.append(u2)

    def walk(obj: Any):
        if isinstance(obj, str):
            for m in HTTP_URL_RE.findall(obj):
                add(m)
        elif isinstance(obj, dict):
            for key in ("registryUrl", "registry_url", "url", "link", "href",
                        "documentUrl", "document_url", "certificateUrl"):
                v = obj.get(key)
                if isinstance(v, str):
                    add(v)
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for v in obj:
                walk(v)

    walk(parsed)
    return found


async def fetch_certificate_json_for_nm(
    session: "aiohttp.ClientSession", nm_id: int, timeout_sec: float = 6.0,
    max_hosts: int = 30, concurrency: int = 0
) -> Tuple[List[str], str]:
    """v40: HTTP-определение наличия документа по certificate.json. БЕЗ браузера.

    Логика (по требованию: наличие документа однозначно определяется по json):
      - перебираем basket-шарды;
      - КАК ТОЛЬКО находим существующий certificate.json (HTTP 200) — это значит
        у карточки ЕСТЬ документ. Внутри ищем ссылку на реестр;
      - если json есть и ссылка извлеклась → cert_json_ok;
      - если json есть, но ссылку извлечь не смогли → cert_json_has_doc_no_url
        (документ ЕСТЬ, но реестровая ссылка не распозналась — это НЕ «нет документов»);
      - если НИ ОДИН шард не отдал json (все честные 404) → cert_json_no_docs
        (документа действительно нет);
      - если были сетевые ошибки (timeout/connection) и не было ни одного 200 →
        cert_json_neterror (надо повторить; это НЕ «нет документов»).

    ВАЖНО: 404 и сетевая ошибка — РАЗНЫЕ вещи. 404 = файла нет = документа нет.
    Сетевая ошибка = техническая проблема, карточку надо повторить, а не списывать.
    """
    headers = {
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    timeout = aiohttp.ClientTimeout(total=max(3.0, float(timeout_sec)))
    urls = certificate_json_urls(nm_id, max_hosts=max_hosts)
    tried = len(urls)

    # v45.10: ВОЗВРАТ К СТАРОМУ РАБОЧЕМУ ПОВЕДЕНИЮ. basket-шарды пробуются ПО ОДНОМУ
    # в порядке «самый вероятный первым» (шард считается детерминированно по vol), и
    # как только найден json со ссылкой — остальные пробы мгновенно отменяются. У
    # карточки с документом он почти всегда на расчётном шарде, поэтому это 1 запрос
    # на карточку — ровно как в старых версиях, где сбор шёл за секунды.
    # Параллельный «залп» всех 16 шардов (его добавили позже ради скорости no-doc
    # карточек) на деле ЛОМАЛ сбор: 30 воркеров × 16 = ~480 соединений к wbbasket.ru
    # разом → WB CDN включал троттлинг → пробы отваливались по таймауту («сетевые
    # ошибки») и скорость падала. Один запрос на карточку нагрузку убирает.
    if concurrency and concurrency > 0:
        probe_concurrency = int(concurrency)
    else:
        probe_concurrency = 1
    probe_concurrency = max(1, probe_concurrency)
    sem = asyncio.Semaphore(probe_concurrency)

    state = {
        "not_found": 0,            # честные 404
        "host_absent": 0,          # DNS не резолвится = такого basket-шарда НЕТ (как 404)
        "net_errors": [],          # timeout / connection reset — это НЕ 404 (повторить)
        "other_status": [],        # 5xx и прочее
        "json_hosts": [],          # шарды где json реально есть (200)
        "urls": [],                # извлечённые ссылки на реестр
    }
    found = asyncio.Event()

    def _is_host_absent(exc: BaseException) -> bool:
        # Несуществующий basket-хост: DNS не резолвится. Это НЕ временный сбой —
        # такого шарда просто нет, повтор не поможет, трактуем как «нет файла».
        name = type(exc).__name__
        if "DNS" in name or "gaierror" in name:
            return True
        s = str(exc).lower()
        return ("getaddrinfo" in s or "name or service not known" in s
                or "nodename nor servname" in s or "temporary failure in name resolution" in s)

    async def probe(url: str) -> None:
        if found.is_set():
            return
        async with sem:
            if found.is_set():
                return
            try:
                async with session.get(url, headers=headers, timeout=timeout) as r:
                    status = int(r.status)
                    if status == 404:
                        state["not_found"] += 1
                        return
                    if status != 200:
                        state["other_status"].append(f"{status}:{hostname(url)}")
                        return
                    # HTTP 200 — json существует, значит документ ЕСТЬ
                    raw = (await r.text(errors="replace")).strip()
                    state["json_hosts"].append(hostname(url))
                    if not raw:
                        return  # json пустой, но он есть — документ есть, ссылки нет
                    parsed: Any = raw
                    try:
                        parsed = json.loads(raw)
                    except Exception:
                        pass
                    urls_found = extract_registry_urls_from_certificate_json(parsed)
                    if not urls_found and isinstance(raw, str):
                        urls_found = extract_registry_urls_from_certificate_json(raw)
                    for u in urls_found:
                        if u not in state["urls"]:
                            state["urls"].append(u)
                    if state["urls"]:
                        found.set()  # ссылка найдена — сигнал отменить остальные пробы
            except asyncio.CancelledError:
                raise
            except asyncio.TimeoutError:
                state["net_errors"].append(f"timeout:{hostname(url)}")
            except aiohttp.ClientError as e:
                if _is_host_absent(e):
                    state["host_absent"] += 1   # такого basket-шарда нет = как 404
                else:
                    state["net_errors"].append(f"{type(e).__name__}:{hostname(url)}")
            except Exception as e:
                if _is_host_absent(e):
                    state["host_absent"] += 1
                else:
                    state["net_errors"].append(f"{type(e).__name__}:{hostname(url)}")

    tasks = [asyncio.create_task(probe(u)) for u in urls]
    try:
        pending = set(tasks)
        while pending:
            _done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
            if found.is_set():
                break
    finally:
        for t in tasks:
            if not t.done():
                t.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)

    # Классифицируем исход (контракт статусов сохранён 1:1 со старой версией).
    all_urls = state["urls"]
    json_hosts = state["json_hosts"]
    not_found = state["not_found"]
    host_absent = state["host_absent"]
    net_errors = state["net_errors"]
    other_status = state["other_status"]
    # «нет файла» = честный 404 ИЛИ несуществующий хост (DNS). И то и другое —
    # достоверный сигнал отсутствия документа на этом шарде, а не сетевой сбой.
    absent = not_found + host_absent
    if all_urls:
        return all_urls, f"cert_json_ok:{json_hosts[0] if json_hosts else '?'}"
    if json_hosts:
        # json существует (документ есть), но ссылку извлечь не удалось
        return [], f"cert_json_has_doc_no_url:hosts={','.join(json_hosts[:3])}"
    if net_errors:
        # БЫЛИ настоящие сетевые сбои (timeout/reset) без 200 → надо повторить
        return [], f"cert_json_neterror:tried={tried},404={not_found},dns={host_absent},net={';'.join(net_errors[:4])}"
    if absent:
        # все ответы — 404 и/или несуществующие шарды → документа действительно нет
        return [], f"cert_json_no_docs:tried={tried},404={not_found},dns={host_absent}"
    # ничего внятного (только other_status) → помечаем neterror
    return [], f"cert_json_neterror:tried={tried},404={not_found},other={';'.join(other_status[:2])}"


# =============================================================================
# v27.9.x: БЫСТРОЕ определение плашки «Оригинал» через статический card.json.
# Найдено по HAR живой карточки WB: страница тянет card.json из той же basket-CDN,
# что и certificate.json (путь /info/ru/card.json). Это СТАТИЧЕСКИЙ файл — берётся
# обычным HTTP без браузера/токена/антибота, быстро и батчами (как certificate.json).
# В нём лежит полная карточка, включая признак оригинальности.
# =============================================================================

def card_json_urls(nm_id: int, max_hosts: int = 16) -> List[str]:
    """URL'ы статического card.json в той же basket-CDN (тот же шард, что и
    certificate.json). Путь /info/ru/card.json."""
    nm = int(nm_id)
    cert = certificate_json_urls(nm, max_hosts=max_hosts)
    return [u.replace(f"/{nm}/info/certificate.json", f"/{nm}/info/ru/card.json") for u in cert]


# Ключи/значения card.json, означающие «оригинальный товар».
# ВНИМАНИЕ: точный набор подтверждается по реальному телу card.json. Сканер
# намеренно широкий: и булевы флаги по ключу, и текст «оригинальный товар».
_ORIGINAL_KEY_RE = re.compile(
    r"(?:is_?original|has_?original(?:mark)?|original_?mark|original_?badge|"
    r"isoriginal|originalproduct|panelpromo)", re.I
)


def card_json_has_original(parsed: Any) -> bool:
    """Глубокий обход card.json: True, если найден признак «Оригинал»
    (булев флаг по ключу ИЛИ текст «оригинальный товар»)."""
    found = False

    def truthy(v: Any) -> bool:
        if v is True:
            return True
        if isinstance(v, (int, float)):
            return v != 0
        if isinstance(v, str):
            return v.strip().lower() in ("1", "true", "да", "yes", "y", "оригинал", "original")
        return False

    def walk(o: Any) -> None:
        nonlocal found
        if found:
            return
        if isinstance(o, dict):
            for k, v in o.items():
                if _ORIGINAL_KEY_RE.search(str(k)) and truthy(v):
                    found = True
                    return
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
        elif isinstance(o, str):
            if "оригинальный товар" in o.lower():
                found = True

    walk(parsed)
    return found


async def fetch_original_via_card_json(
    session: "aiohttp.ClientSession", nm_id: int, timeout_sec: float = 6.0,
    max_hosts: int = 16, concurrency: int = 0
) -> Tuple[Optional[bool], str]:
    """Тянет card.json по basket-шардам (конкурентно, первый 200 выигрывает) и
    определяет признак «Оригинал».

    Возвращает (is_original|None, detail). None = card.json не нашли (не знаем),
    тогда результат не должен перетирать другие методы.
    """
    headers = {
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    timeout = aiohttp.ClientTimeout(total=max(3.0, float(timeout_sec)))
    urls = card_json_urls(nm_id, max_hosts=max_hosts)
    probe_concurrency = max(1, int(concurrency) if concurrency else min(len(urls), 16))
    sem = asyncio.Semaphore(probe_concurrency)
    state = {"result": None, "host": ""}
    done = asyncio.Event()

    async def probe(url: str) -> None:
        if done.is_set():
            return
        async with sem:
            if done.is_set():
                return
            try:
                async with session.get(url, headers=headers, timeout=timeout) as r:
                    if int(r.status) != 200:
                        return
                    raw = (await r.text(errors="replace")).strip()
                    if not raw:
                        return
                    try:
                        parsed = json.loads(raw)
                    except Exception:
                        parsed = raw
                    state["result"] = card_json_has_original(parsed)
                    state["host"] = hostname(url)
                    done.set()
            except asyncio.CancelledError:
                raise
            except Exception:
                pass

    tasks = [asyncio.create_task(probe(u)) for u in urls]
    try:
        pending = set(tasks)
        while pending:
            _d, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
            if done.is_set():
                break
    finally:
        for t in tasks:
            if not t.done():
                t.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)

    if state["result"] is None:
        return None, "card_json_not_found"
    return bool(state["result"]), f"card_json_ok:{state['host']}:{'orig' if state['result'] else 'no'}"


# =============================================================================
# v46: БЕЙДЖ «Документ проверен WB» через статический card.json.
# Эмпирически (по живым карточкам, 7 «с кнопкой» / 7 «без»): признак наличия
# кнопки «Документы проверены» = card.json.certificate.verified == true.
# Карточки С кнопкой: {"certificate": {"verified": true}}; БЕЗ: {"certificate": {}}
# либо поля нет. Лежит в том же card.json (та же basket-CDN), что и «Оригинал».
# Бейджа НЕТ в viewFlags/supplierFlags/promotions динамического API (проверено).
# =============================================================================

def card_json_docs_verified(parsed: Any) -> bool:
    """True, если в card.json есть бейдж «Документ проверен WB»
    (certificate.verified == true)."""
    def _cert_verified(cert: Any) -> bool:
        return isinstance(cert, dict) and cert.get("verified") is True

    if isinstance(parsed, dict) and _cert_verified(parsed.get("certificate")):
        return True
    # защитный обход: certificate.verified==true в любом вложенном объекте
    found = False

    def walk(o: Any) -> None:
        nonlocal found
        if found:
            return
        if isinstance(o, dict):
            if _cert_verified(o.get("certificate")):
                found = True
                return
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)

    walk(parsed)
    return found


async def fetch_docs_verified_via_card_json(
    session: "aiohttp.ClientSession", nm_id: int, timeout_sec: float = 6.0,
    max_hosts: int = 16, concurrency: int = 0
) -> Tuple[Optional[bool], str]:
    """Тянет card.json по basket-шардам (первый 200 выигрывает) и определяет
    бейдж «Документ проверен WB» (certificate.verified). Возвращает
    (verified|None, detail). None = card.json не нашли (не знаем)."""
    headers = {
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
    }
    timeout = aiohttp.ClientTimeout(total=max(3.0, float(timeout_sec)))
    urls = card_json_urls(nm_id, max_hosts=max_hosts)
    probe_concurrency = max(1, int(concurrency) if concurrency else min(len(urls), 16))
    sem = asyncio.Semaphore(probe_concurrency)
    state = {"result": None, "host": ""}
    done = asyncio.Event()

    async def probe(url: str) -> None:
        if done.is_set():
            return
        async with sem:
            if done.is_set():
                return
            try:
                async with session.get(url, headers=headers, timeout=timeout) as r:
                    if int(r.status) != 200:
                        return
                    raw = (await r.text(errors="replace")).strip()
                    if not raw:
                        return
                    try:
                        parsed = json.loads(raw)
                    except Exception:
                        parsed = raw
                    state["result"] = card_json_docs_verified(parsed)
                    state["host"] = hostname(url)
                    done.set()
            except asyncio.CancelledError:
                raise
            except Exception:
                pass

    tasks = [asyncio.create_task(probe(u)) for u in urls]
    try:
        pending = set(tasks)
        while pending:
            _d, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
            if done.is_set():
                break
    finally:
        for t in tasks:
            if not t.done():
                t.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)

    if state["result"] is None:
        return None, "card_json_not_found"
    return bool(state["result"]), f"card_json_ok:{state['host']}:{'verified' if state['result'] else 'no'}"


# =============================================================================
# v40.3: дотягивание имени продавца (seller_name).
# WB API v18 в поисковой выдаче отдаёт supplierId, но НЕ имя продавца.
# Имя берём батчами через card.wb.ru/cards/v2/detail?nm=id1;id2;... (до 100 nm за раз).
# =============================================================================

# v27.9.x: бит маски viewFlags, означающий плашку «Оригинал». viewFlags WB
# отдаёт в поисковой выдаче по каждому товару, поэтому проверка БЕСПЛАТНА (без
# доп. запросов) и мгновенна.
# Значение подтверждено на реальных данных (3 товара с плашкой vs 4 без):
#   оригинал  863239425=135671833, 173642704=135028761, 519200637=202899465  → бит 3 есть
#   обычные   1008983494=811026, 267983778=135798913, 237643208=135028736, 663082548=8591261696 → бита 3 нет
# Бит 3 (значение 8) установлен у ВСЕХ оригинальных и НИ У ОДНОГО обычного.
WB_ORIGINAL_VIEWFLAG_BIT = 8


def _detect_wb_original(p: Dict[str, Any]) -> str:
    """v43: определяет признак оригинальности товара из объекта WB.

    Точного публичного булева поля «оригинал» в search-ответе WB нет — бейдж
    «Оригинальный товар» рендерится на странице карточки. Здесь распознаём явные
    сигналы из доступных полей (supplierFlags, promoTextCard, name и т.п.).
    Возвращает 'оригинал' / 'не указано'. Надёжнее проверять детальную страницу
    (см. enrich-режим), но это даёт быстрый сигнал по данным поиска.
    """
    try:
        # 0) viewFlags — битовая маска бейджей WB (есть прямо в поисковой выдаче).
        #    Если известен бит «Оригинал» — это самый быстрый и точный сигнал.
        if WB_ORIGINAL_VIEWFLAG_BIT:
            vf = p.get("viewFlags")
            if isinstance(vf, int) and (vf & WB_ORIGINAL_VIEWFLAG_BIT):
                return "оригинал"
        # 1) Явные строковые поля, где WB иногда помечает оригинальность
        for key in ("promoTextCard", "promoTextCat", "name"):
            v = str(p.get(key) or "")
            if v and re.search(r"\bоригинал", v, re.IGNORECASE):
                return "оригинал"
        # 2) supplierFlags — битовые/служебные флаги продавца. Некоторые значения
        #    соответствуют верифицированным/оригинальным продавцам. Точная битовая
        #    карта недокументирована, поэтому только отмечаем наличие флага.
        flags = p.get("supplierFlags")
        if isinstance(flags, int) and flags > 0:
            # не делаем ложного «оригинал» по одному флагу — оставляем нейтральным
            pass
        # 3) Иногда есть поле viewFlags / panelPromoId с признаком бренд-зоны
    except Exception:
        pass
    return "не указано"


def _seller_from_detail_product(p: Dict[str, Any]) -> str:
    """Достаёт имя продавца из объекта товара detail-API."""
    for key in ("supplierName", "sellerName", "supplier", "seller"):
        v = p.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, dict):
            for kk in ("name", "supplierName", "sellerName", "trademark", "title"):
                if isinstance(v.get(kk), str) and v[kk].strip():
                    return v[kk].strip()
    return ""


async def enrich_sellers_batch(cards: List[Card], args) -> None:
    """v40.3: массово заполняет card.seller_name через card.wb.ru detail API.

    Идёт батчами по 100 nm_id. Не критично если что-то не найдётся —
    supplier_id всё равно сохранён. Работает поверх обычного HTTP (не FSA),
    поэтому не блокируется антиботом.
    """
    if not cards:
        return
    need = [c for c in cards if not getattr(c, 'seller_name', '')]
    if not need:
        return
    workers = max(2, int(getattr(args, 'http_link_workers', 30)) // 3)
    batch_size = 100
    batches = [need[i:i + batch_size] for i in range(0, len(need), batch_size)]
    by_nm: Dict[int, Card] = {c.nm_id: c for c in need}

    headers = {
        "User-Agent": getattr(args, 'user_agent', '') or
                      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    timeout = aiohttp.ClientTimeout(total=20, connect=8)
    connector = aiohttp.TCPConnector(limit=workers * 2, ttl_dns_cache=300, ssl=False)
    sem = asyncio.Semaphore(workers)
    filled = {"n": 0}
    t0 = time.time()

    detail_templates = [
        "https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm={nms}",
        "https://card.wb.ru/cards/v4/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm={nms}",
    ]

    print(f"🏷  Дотягиваю имена продавцов для {len(need)} карточек ({len(batches)} батчей по {batch_size})...")

    async with aiohttp.ClientSession(timeout=timeout, headers=headers, connector=connector) as session:
        async def do_batch(batch: List[Card]):
            nms = ";".join(str(c.nm_id) for c in batch)
            for tmpl in detail_templates:
                try:
                    async with sem:
                        async with session.get(tmpl.format(nms=nms), timeout=timeout) as r:
                            if r.status != 200:
                                continue
                            data = json.loads(await r.text(errors="replace"))
                except Exception:
                    continue
                products = recursive_find_products(data)
                if not products:
                    continue
                for p in products:
                    nm = safe_int(p.get("id") or p.get("nmId") or p.get("nm_id"))
                    c = by_nm.get(nm)
                    if c is None:
                        continue
                    s = _seller_from_detail_product(p)
                    if s and not c.seller_name:
                        c.seller_name = s
                        filled["n"] += 1
                    # заодно дополним supplier_id если пуст
                    if not getattr(c, 'supplier_id', ''):
                        sid = str(p.get("supplierId") or p.get("supplierID") or "")
                        if sid:
                            c.supplier_id = sid
                    # v27.9.x: признак «Оригинал» из viewFlags detail-ответа
                    # (подстраховка к основному пути — viewFlags из поиска).
                    if WB_ORIGINAL_VIEWFLAG_BIT:
                        _vf = p.get("viewFlags")
                        if isinstance(_vf, int) and (_vf & WB_ORIGINAL_VIEWFLAG_BIT):
                            c.is_original = "оригинал"
                break  # один из шаблонов сработал
        tasks = [asyncio.create_task(do_batch(b)) for b in batches]
        await asyncio.gather(*tasks, return_exceptions=True)

    print(f"✓ Имена продавцов: заполнено {filled['n']}/{len(need)} за {time.time()-t0:.1f}с")


async def enrich_docs_verified_batch(cards: List[Card], args) -> None:
    """v46: массово проставляет card.docs_verified («Да»/«Нет») — бейдж
    «Документ проверен WB» из статического card.json (certificate.verified == true).

    Делается ОТДЕЛЬНЫМ проходом ДО сбора ссылок, чтобы признак попал в строку
    независимо от того, каким путём собрана карточка (HTTP fast-path ИЛИ браузер).
    Иначе карточки, собранные браузером, оставались с пустым значением.
    Тот же basket-CDN, что и certificate.json; обычный HTTP, без антибота.
    card.json не найден / бейджа нет → «Нет» (кнопки «Документы проверены» нет)."""
    if not cards:
        return
    need = [c for c in cards if not getattr(c, 'docs_verified', '')]
    if not need:
        return
    workers = max(2, int(getattr(args, 'http_link_workers', 30)))
    cert_timeout = float(getattr(args, 'cert_timeout_sec', 6.0))
    cert_max_hosts = int(getattr(args, 'cert_max_hosts', 30))
    headers = {
        "User-Agent": getattr(args, 'user_agent', '') or
                      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "Accept": "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    }
    timeout = aiohttp.ClientTimeout(total=30, connect=8)
    connector = aiohttp.TCPConnector(limit=max(32, workers * 2), ttl_dns_cache=300, ssl=False)
    sem = asyncio.Semaphore(workers)
    stats = {"yes": 0, "no": 0, "done": 0}
    t0 = time.time()
    print(f"🛡  Определяю бейдж «Документ проверен WB» для {len(need)} карточек (card.json)...")

    async with aiohttp.ClientSession(timeout=timeout, headers=headers, connector=connector) as session:
        async def work(card: Card):
            async with sem:
                try:
                    ver, _det = await fetch_docs_verified_via_card_json(
                        session, card.nm_id, timeout_sec=cert_timeout, max_hosts=cert_max_hosts
                    )
                except Exception:
                    ver = None
                card.docs_verified = "Да" if ver is True else "Нет"
                if ver is True:
                    stats["yes"] += 1
                else:
                    stats["no"] += 1
                stats["done"] += 1
        tasks = [asyncio.create_task(work(c)) for c in need]
        await asyncio.gather(*tasks, return_exceptions=True)

    print(f"✓ «Документ проверен WB»: Да={stats['yes']}, Нет={stats['no']} за {time.time()-t0:.1f}с")


async def run_http_link_prefetch(
    cards: List[Card], args, store: "ResultStore", processed: Set[int]
) -> Set[int]:
    """v39.12: массовый параллельный HTTP-сбор ссылок через certificate.json.

    ВАЖНО: запускается ДО браузерной очереди. Те карточки которым удалось
    HTTP-путём найти ссылку — сразу пишутся в store и НЕ идут в браузер.

    v39.14: НЕ помечаем NO_DOCS если не проверили все 30 basket-шардов.
    Раньше при cert_max_hosts=12 программа писала "НЕТ ДОКУМЕНТОВ" если
    первые 12 шардов дали 404. Но документ мог быть на шардах 13-30!
    Теперь:
      - если 404 на ВСЕХ 30 шардах → точно нет документов
      - если 404 < 30 → fallback в браузер (он умеет найти плашку через UI)
    """
    if not cards:
        return set()
    workers = max(1, int(getattr(args, 'http_link_workers', 30)))
    cert_timeout = float(getattr(args, 'cert_timeout_sec', 6.0))
    cert_max_hosts = int(getattr(args, 'cert_max_hosts', 30))
    # v40: на 1 этапе браузер НЕ нужен — наличие документа однозначно определяется
    # по certificate.json (HTTP 200 = документ есть, 404 на всех шардах = документа нет).
    # Браузерный fallback используется ТОЛЬКО для технических сетевых ошибок,
    # и только если режим это разрешает.
    link_mode = str(getattr(args, 'link_mode', 'http_only') or 'http_only').lower()
    allow_browser_fallback = (link_mode == 'http_first')

    print(f"🌐 HTTP fast-path: запрашиваю certificate.json для {len(cards)} карточек параллельностью {workers} (basket-шардов на карточку: {cert_max_hosts})...")
    print(f"   Логика: json со ссылкой → СОБРАНА; все 404 → НЕТ ДОКУМЕНТОВ; json без ссылки → НЕТ ССЫЛКИ; сетевая ошибка → {'браузер' if allow_browser_fallback else 'повтор (ERROR)'}.")

    timeout = aiohttp.ClientTimeout(total=30, connect=8)
    # v45.10: шарды пробуются по одному (см. fetch_certificate_json_for_nm), значит
    # одновременных соединений ~ числу воркеров. Пул workers*2 (как в старых версиях)
    # достаточно — никакого «залпа» больше нет.
    connector = aiohttp.TCPConnector(limit=max(32, workers * 2), ttl_dns_cache=300, ssl=False)
    headers = {
        "User-Agent": getattr(args, 'user_agent', '') or
                      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "Accept": "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    }
    sem = asyncio.Semaphore(workers)
    done_nm: Set[int] = set()
    stats = {"ok": 0, "no_docs": 0, "no_link": 0, "errors": 0, "fallback": 0, "done": 0}
    t0 = time.time()
    last_print = t0

    async def progress_loop():
        nonlocal last_print
        try:
            while stats['done'] < len(cards):
                await asyncio.sleep(3.0)
                el = max(0.1, time.time() - t0)
                speed = stats['done'] / el * 60.0
                eta = (len(cards) - stats['done']) / max(0.01, stats['done'] / el) if stats['done'] else 0
                print(
                    f"  [HTTP {stats['done']}/{len(cards)}] {speed:.0f}/мин, "
                    f"ссылок={stats['ok']}, нет_докум={stats['no_docs']}, "
                    f"нет_ссылки={stats['no_link']}, ошибки={stats['errors']}, "
                    f"fallback→браузер={stats['fallback']}, ETA≈{eta:.0f}с"
                )
                emit_progress("links", stats['done'], len(cards))
                last_print = time.time()
        except asyncio.CancelledError:
            return

    async with aiohttp.ClientSession(timeout=timeout, headers=headers, connector=connector) as session:
        async def work(card: Card):
            async with sem:
                if card.nm_id in processed:
                    stats['done'] += 1
                    return
                try:
                    urls, detail = await fetch_certificate_json_for_nm(
                        session, card.nm_id, timeout_sec=cert_timeout, max_hosts=cert_max_hosts
                    )
                    # v45.12: ОДИН повтор временной сетевой ошибки — но ТОЛЬКО пока
                    # ошибок мало. Если ошибок уже много, значит WB троттлит IP — тогда
                    # повторы лишь УСИЛИВАЮТ нагрузку и углубляют бан, поэтому НЕ повторяем
                    # (карточка пойдёт в ОШИБКУ, повторный запуск/resume её доберёт позже).
                    if (not urls and detail.startswith("cert_json_neterror")
                            and stats['errors'] < 15):
                        await asyncio.sleep(0.4 + random.uniform(0, 0.3))
                        try:
                            urls, detail = await fetch_certificate_json_for_nm(
                                session, card.nm_id, timeout_sec=cert_timeout, max_hosts=cert_max_hosts
                            )
                        except Exception as e:
                            urls, detail = [], f"cert_json_exception:{type(e).__name__}:{str(e)[:120]}"
                except Exception as e:
                    urls, detail = [], f"cert_json_exception:{type(e).__name__}:{str(e)[:120]}"

                # v27.9.x: card.json НЕ содержит признак «Оригинал» (подтверждено
                # реальным телом card.json) — плашка приходит из viewFlags в
                # динамическом API. Поэтому card.json-проба по умолчанию ОТКЛЮЧЕНА
                # (не тратим лишние запросы). Определение «Оригинал» — по биту
                # viewFlags из поисковой выдачи (см. _detect_wb_original).
                if getattr(args, 'check_original_cardjson', False):
                    try:
                        _is_orig, _odet = await fetch_original_via_card_json(
                            session, card.nm_id, timeout_sec=cert_timeout, max_hosts=cert_max_hosts
                        )
                        if _is_orig is True:
                            card.is_original = 'оригинал'
                    except Exception:
                        pass

                # v46: бейдж «Документ проверен WB» (card.docs_verified) проставляется
                # ОТДЕЛЬНЫМ проходом enrich_docs_verified_batch ДО сбора ссылок —
                # чтобы признак попал в строку независимо от пути сбора (HTTP/браузер).
                # Здесь, как защита, для карточек без документа фиксируем «Нет», если
                # проход почему-то не отработал (verified невозможен без сертификата).
                if not getattr(card, 'docs_verified', ''):
                    _has_doc = bool(urls) or detail.startswith("cert_json_ok") \
                        or detail.startswith("cert_json_has_doc_no_url")
                    if not _has_doc:
                        card.docs_verified = "Нет"

                base_fields = dict(
                    query=card.source_query or args.query,
                    nm_id=card.nm_id,
                    **_card_fields_for_result(card),
                    worker="http",
                    checked_at=now_iso(),
                )

                if urls:
                    # json есть + ссылка извлечена → СОБРАНА
                    url = urls[0]
                    row = ResultRow(
                        **base_fields,
                        status=STATUS_LINK_COLLECTED,
                        registry_url=url,
                        registry_host=hostname(url),
                        registry_record_id=extract_record_id(url),
                        details=f"http_fast_path; {detail}",
                    )
                    await store.add(row)
                    done_nm.add(card.nm_id)
                    stats['ok'] += 1
                elif detail.startswith("cert_json_no_docs"):
                    # ВСЕ шарды честно вернули 404 → документа НЕТ. Это надёжно по HTTP.
                    row = ResultRow(
                        **base_fields,
                        status=STATUS_NO_DOCS,
                        details=f"http_fast_path; {detail}",
                    )
                    await store.add(row)
                    done_nm.add(card.nm_id)
                    stats['no_docs'] += 1
                elif detail.startswith("cert_json_has_doc_no_url"):
                    # json существует (документ ЕСТЬ), но ссылку извлечь не смогли.
                    # Это НЕ «нет документов». Помечаем особым статусом —
                    # такие карточки видно в выгрузке, можно разобрать вручную/браузером на 2 этапе.
                    row = ResultRow(
                        **base_fields,
                        status=STATUS_NO_REGISTRY_LINK,
                        details=f"http_fast_path; документ есть, ссылка не извлеклась; {detail}",
                    )
                    await store.add(row)
                    done_nm.add(card.nm_id)
                    stats['no_link'] = stats.get('no_link', 0) + 1
                else:
                    # cert_json_neterror / exception — СЕТЕВАЯ ошибка, не 404.
                    # НЕ помечаем «нет документов». Если включён браузерный fallback —
                    # отдаём туда; иначе помечаем технической ошибкой (можно повторить запуском).
                    if allow_browser_fallback:
                        stats['fallback'] += 1
                    else:
                        row = ResultRow(
                            **base_fields,
                            status=STATUS_ERROR,
                            details=f"http_fast_path; сетевая ошибка (повторите запуск); {detail}",
                        )
                        await store.add(row)
                        done_nm.add(card.nm_id)
                        stats['errors'] = stats.get('errors', 0) + 1
                stats['done'] += 1

        prog = asyncio.create_task(progress_loop())
        try:
            tasks = [asyncio.create_task(work(c)) for c in cards]
            await asyncio.gather(*tasks, return_exceptions=True)
        finally:
            prog.cancel()
            try:
                await prog
            except Exception:
                pass

    elapsed = time.time() - t0
    print(
        f"✓ HTTP fast-path завершён за {elapsed:.1f}с: "
        f"ссылок собрано={stats['ok']}, нет документов={stats['no_docs']}, "
        f"нет ссылки (json без url)={stats['no_link']}, "
        f"тех.ошибки={stats['errors']}, в браузер на fallback={stats['fallback']}"
    )
    if stats['errors'] > 0:
        print(f"   ⚠️  {stats['errors']} карточек с сетевыми ошибками — это НЕ «нет документов». "
              f"Повтори запуск (resume пропустит уже собранные) либо добавь --link-mode http_first для браузерного fallback.")
    # Сохраним store в xlsx раз после prefetch, чтобы было видно
    try:
        await store.save()
    except Exception as e:
        print(f"⚠️  Сохранение после HTTP prefetch: {e}")
    return done_nm


def extract_urls_from_text(text: str, max_len: int = 200000) -> List[str]:
    if not text:
        return []
    t = str(text)[:max_len]
    out: List[str] = []
    for m in HTTP_URL_RE.finditer(t):
        out.append(clean_url(m.group(0)))
    # URL может быть query-параметром внутри другого URL
    expanded = list(out)
    for url in out:
        try:
            qs = parse_qs(urlparse(url).query)
            for vals in qs.values():
                for val in vals:
                    val = clean_url(val)
                    if val.startswith("http"):
                        expanded.append(val)
                        expanded.extend(extract_urls_from_text(val, 10000))
        except Exception:
            pass
    # dedupe
    seen = set()
    res = []
    for u in expanded:
        if u not in seen:
            seen.add(u)
            res.append(u)
    return res

def first_allowed_url(urls: Iterable[str]) -> str:
    for u in urls:
        u = clean_url(u)
        if is_allowed_registry_url(u):
            return u
        # иногда ссылка вложена в query параметр
        for nested in extract_urls_from_text(u, 10000):
            if is_allowed_registry_url(nested):
                return clean_url(nested)
    return ""

# -----------------------------
# Query generation and WB JSON collection
# -----------------------------

# v39.7: универсальные словари для генерации запросов и фильтрации по категориям WB.
# Цель: программа должна работать с ЛЮБОЙ товарной категорией, не только одеждой.

# Маппинг: domain (домен) → keywords для определения категории из запроса пользователя.
# Если запрос содержит ключевое слово — это направление считается активным.
QUERY_DOMAIN_KEYWORDS = {
    'clothing': [
        'одеж', 'куртк', 'комбинезон', 'костюм', 'платье', 'платья', 'футболк',
        'лонгслив', 'брюк', 'штан', 'шорт', 'кофт', 'толстовк', 'худи', 'свитшот',
        'боди', 'ползунк', 'пижам', 'халат', 'жилет', 'свитер', 'пальто', 'плащ',
        'юбк', 'сарафан', 'блуз', 'рубаш', 'белье', 'распашонк', 'песочник',
        'трикотаж', 'кардиган', 'водолазк',
    ],
    'shoes': [
        'обув', 'кроссовк', 'кеды', 'ботинк', 'сапог', 'туфли', 'сандали',
        'босоножк', 'тапочк', 'мокасин', 'валенк', 'угги', 'дутики',
    ],
    'toys': [
        'игрушк', 'кукл', 'конструктор', 'пазл', 'мягка игрушк', 'погремушк',
        'машинк', 'самолетик', 'настольн игр', 'лего', 'кубик', 'мяч',
        'неваляшк', 'юла', 'каталк', 'qb', 'pop it', 'попит', 'спиннер',
    ],
    'kids_accessories': [
        'шапк', 'панамк', 'кепк', 'варежк', 'перчатк', 'шарф', 'рюкзак',
        'пенал', 'портфел', 'ранец', 'сумочк', 'бант', 'заколк',
    ],
    'baby_gear': [
        'коляск', 'автокресл', 'манеж', 'переноск', 'кроватк', 'пеленальн',
        'стульчик для кормлен', 'ходунк', 'качел',
    ],
    'cosmetics': [
        'космет', 'крем', 'шампун', 'мыло', 'гель для душа', 'дезодорант',
        'духи', 'тушь', 'помад', 'лак для ногтей', 'маска для лица',
    ],
    'electronics': [
        'телефон', 'смартфон', 'наушник', 'колонк', 'планшет', 'ноутбук',
        'компьютер', 'монитор', 'клавиатур', 'мышк', 'роутер', 'провод', 'кабел',
        'зарядк', 'зарядное', 'powerbank', 'аккумулятор',
    ],
    'home': [
        'постельн', 'наволочк', 'подушк', 'одеял', 'плед', 'полотенц', 'покрывал',
        'простын', 'скатерт', 'штор', 'занавес', 'коврик', 'ковер',
    ],
    'kitchenware': [
        'кастрюл', 'сковород', 'нож кухонн', 'тарелк', 'кружк', 'чашк',
        'столов прибор', 'разделочн доск', 'венчик',
    ],
    'food': [
        'еда', 'продукт', 'смесь молочн', 'каша', 'пюре', 'напиток', 'сок',
        'печенье', 'конфет', 'шоколад', 'чай', 'кофе',
    ],
}

# Расширения каждой категории: типы товаров. Каждый тип будет одним из запросов.
DOMAIN_PRODUCT_TYPES = {
    'clothing': [
        'куртки', 'комбинезоны', 'костюмы', 'пижамы', 'платья', 'футболки',
        'лонгсливы', 'брюки', 'штаны', 'шорты', 'кофты', 'толстовки', 'худи',
        'свитшоты', 'боди', 'ползунки', 'носки', 'колготки', 'белье',
        'халаты', 'жилеты', 'свитеры', 'кардиганы', 'водолазки',
        'распашонки', 'песочники', 'юбки', 'сарафаны', 'блузки', 'рубашки',
    ],
    'shoes': [
        'кроссовки', 'кеды', 'ботинки', 'сапоги', 'туфли', 'сандалии',
        'босоножки', 'тапочки', 'мокасины', 'валенки', 'угги', 'дутики',
        'полусапоги',
    ],
    'toys': [
        'мягкие игрушки', 'куклы', 'конструкторы', 'пазлы', 'погремушки',
        'машинки', 'самолетики', 'настольные игры', 'кубики', 'мячи',
        'неваляшки', 'юла', 'каталки', 'игрушечная посуда', 'фигурки',
        'плюшевые игрушки', 'развивающие игрушки', 'игровые наборы',
        'спиннеры', 'pop it', 'мозаика', 'трещотки',
        'роботы', 'интерактивные игрушки', 'музыкальные игрушки', 'сортеры',
        'пирамидки', 'шнуровки', 'бизиборды', 'кинетический песок', 'слаймы',
        'наборы для творчества', 'пластилин', 'пальчиковые краски', 'раскраски',
        'игрушечная кухня', 'железная дорога', 'трек для машинок', 'набор доктора',
        'радиоуправляемые машинки', 'сюжетно-ролевые наборы',
    ],
    'kids_accessories': [
        'шапки', 'панамы', 'кепки', 'варежки', 'перчатки', 'шарфы',
        'рюкзаки', 'пеналы', 'портфели', 'ранцы', 'сумки', 'банты', 'заколки',
        'снуды', 'бейсболки', 'обручи', 'резинки для волос', 'ободки',
        'зонты', 'солнцезащитные очки', 'ремни', 'кошельки', 'часы',
    ],
    'baby_gear': [
        'коляски', 'автокресла', 'манежи', 'переноски', 'кроватки',
        'пеленальные столики', 'стульчики для кормления', 'ходунки', 'качели',
        'люльки', 'прыгунки', 'мобили', 'шезлонги', 'бортики в кроватку',
        'слинги', 'кенгуру', 'санки', 'беговелы', 'самокаты',
        'трёхколёсные велосипеды', 'дождевики для коляски', 'москитные сетки',
    ],
    'cosmetics': [
        'кремы', 'шампуни', 'мыло', 'гели для душа', 'дезодоранты', 'духи',
        'туши для ресниц', 'помады', 'лаки для ногтей', 'маски для лица',
        'бальзамы', 'кондиционеры для волос', 'пены для ванны', 'влажные салфетки',
        'зубные пасты', 'зубные щётки', 'ватные палочки', 'ватные диски',
        'присыпки', 'детское масло', 'пенки для умывания', 'тоники', 'скрабы',
        'сыворотки', 'патчи',
    ],
    'electronics': [
        'наушники', 'колонки', 'клавиатуры', 'мыши', 'роутеры', 'провода',
        'зарядные устройства', 'powerbank', 'usb кабели',
        'беспроводные наушники', 'гарнитуры', 'веб-камеры', 'микрофоны',
        'флешки', 'карты памяти', 'hdmi кабели', 'удлинители', 'сетевые фильтры',
        'адаптеры', 'переходники', 'блоки питания', 'держатели для телефона',
        'умные часы', 'фитнес-браслеты', 'смартфоны', 'планшеты', 'мониторы',
        'проекторы', 'приставки',
    ],
    'home': [
        'постельное белье', 'подушки', 'одеяла', 'пледы', 'полотенца',
        'покрывала', 'простыни', 'наволочки', 'скатерти', 'шторы', 'коврики',
        'матрасы', 'наматрасники', 'бортики', 'балдахины', 'пеленки',
        'конверты для новорожденных', 'спальные мешки', 'ночники',
        'корзины для игрушек', 'органайзеры', 'тканевые салфетки', 'полотенца с уголком',
    ],
    'kitchenware': [
        'кастрюли', 'сковороды', 'кухонные ножи', 'тарелки', 'кружки', 'чашки',
        'столовые приборы', 'разделочные доски',
        'сотейники', 'ковши', 'заварочные чайники', 'миски', 'салатники',
        'формы для выпечки', 'противни', 'дуршлаги', 'тёрки', 'лопатки',
        'половники', 'шумовки', 'контейнеры для еды', 'термосы',
        'бутылки для воды', 'ланч-боксы', 'детская посуда', 'поильники',
    ],
    'food': [
        'смеси молочные', 'каши', 'пюре', 'напитки детские', 'соки', 'печенье',
        'детская вода', 'детский чай', 'сухие завтраки', 'мюсли', 'батончики',
        'йогурты', 'творожки', 'снеки', 'фруктовое пюре', 'мясное пюре',
        'овощное пюре', 'кисели', 'макароны детские',
    ],
}

# Универсальные модификаторы — применяются к любым типам товаров.
UNIVERSAL_MODIFIERS = [
    'для мальчиков', 'для девочек', 'детские', 'подростковые', 'школьные',
    'для малышей', 'для дошкольников', 'для новорожденных',
]

# Цветовые/размерные модификаторы — для одежды и обуви.
APPAREL_MODIFIERS = [
    'летние', 'демисезонные', 'зимние', 'весна', 'осень',
    'черные', 'белые', 'синие', 'розовые', 'серые', 'бежевые',
    'красные', 'голубые', 'зеленые', 'желтые',
] + [f'рост {x}' for x in [62, 86, 104, 116, 128, 140, 152, 164]]


def detect_query_domain(base_query: str) -> str:
    """v39.7: определить домен запроса по ключевым словам.

    Стратегия: для каждого домена считаем суммарный «вес» совпавших ключевых
    слов, где вес = длина ключа. Длинные/специфичные ключи имеют приоритет.
    Пример: «постельное белье» содержит и 'белье' (clothing, 5 букв) и
    'постельн бель' (home, 13 букв) → побеждает home.
    """
    low = base_query.lower().replace('ё', 'е')
    scores = {}
    for domain, keys in QUERY_DOMAIN_KEYWORDS.items():
        weight = 0
        for k in keys:
            if k in low:
                weight += len(k)
        if weight > 0:
            scores[domain] = weight
    if not scores:
        return 'unknown'
    return max(scores.items(), key=lambda kv: kv[1])[0]


# v39.7: Whitelist subject-id WB для каждого домена.
# WB по любому запросу часто возвращает мусор смежных категорий — например по «детские игрушки серые»
# приходят платья, куртки, футболки. Этот whitelist отфильтровывает их на этапе сбора.
# Если subject не в whitelist'е домена — карточка отбрасывается (для strict-режима).
#
# Источник: реальные subject-id из выгрузок links_stage.xlsx (одежда) и links_toys.xlsx (игрушки).
DOMAIN_SUBJECT_WHITELIST = {
    'clothing': {
        # v39.7: построено на основе реальных subject WB из links_clothing.xlsx и
        # типичных категорий одежды (всё, что носят на теле).
        # Куртки/ветровки/пуховики/верх
        '168', '172', '174', '394', '1631', '2110', '4509', '4531', '4532', '4533',
        # Платья/сарафаны
        '69', '70', '162', '4000', '225',
        # Костюмы (детские/взрослые)
        '177', '3828', '4853',
        # Футболки/майки/топы/поло
        '192', '185', '219', '2230', '4601', '4568',
        # Брюки/штаны/шорты/велосипедки/джинсы/леггинсы
        '11', '38', '151', '148', '3835', '4608', '4617', '4640', '5204', '5267',
        # Лонгсливы/худи/толстовки/свитшоты/бомберы/водолазки/жилеты
        '153', '156', '159', '217', '233', '1635', '1724',
        # Юбки
        '38', '4001',
        # Боди/комбинезоны/слипы/ползунки/песочники/кофточки
        '157', '276', '1259', '1457', '1720', '3644', '3995', '4599',
        # Школьная одежда: блузки/рубашки
        '41', '184', '2572',
        # Кардиганы/болеро
        '225', '4642',
        # Пижамы / кигуруми
        '2649', '4615',
        # Юбки школьные
        '4001',
    },
    'shoes': set(),
    # v39.11: whitelist для shoes ОЧИЩЕН — раньше там были угаданные ID
    # ('527', '186' и т.п.) которые на практике оказались неправильными.
    # Кто-то запустил «кроссовки для детей» и получил 4 карточки из 100 — это был баг.
    # Теперь shoes фильтруется ТОЛЬКО по тексту (subjectName/название карточки)
    # через DOMAIN_SUBJECT_NAME_KEYWORDS. Это надёжнее чем угаданные ID.
    'toys': {
        # Игрушки: мягкие, интерактивные, антистресс, развивающие, фигурки, наборы
        '268', '1462', '2095', '2547',  # Мягкие/комфортеры/интерактивные/антистресс
        '125', '291', '227', '283',     # Игровые наборы, фигурки, куклы
        '120', '284',                    # Конструкторы, развивающие
        '310', '716',                    # Игрушки для ванной, карнавальные
        '2030', '4034',                  # Пазлы, настольные игры
        '5076', '305',                   # Радиоуправляемые, машинки
        # v39.7.1: добавлены по результатам теста на реальных данных
        '87',    # Игрушечные автомобили
        '122',   # Коллекционные игрушки/фигурки
        '1032',  # Конструкторы продвинутые
    },
    'kids_accessories': {
        # Аксессуары: рюкзаки, шапки, варежки, банты, заколки
        '138', '263', '197',
    },
    'baby_gear': set(),
    'cosmetics': set(),
    'electronics': set(),
    'home': set(),
    'kitchenware': set(),
    'food': set(),
}


# v39.9: текстовые ключевые слова для каждого домена.
# WB API в v18 возвращает поле `subjectName` (например "Кроссовки", "Платья", "Игрушки мягкие").
# Фильтрация по тексту НАМНОГО надёжнее чем по subject-id — ID можно угадать неправильно,
# а текст категории однозначен и не меняется.
#
# Логика проверки: «subjectName содержит хотя бы одно из этих ключевых слов» — карточка подходит.
DOMAIN_SUBJECT_NAME_KEYWORDS = {
    'clothing': [
        # одежда верх
        'куртк', 'ветровк', 'пуховик', 'пальто', 'плащ', 'парк', 'жилет',
        'кардиган', 'болеро', 'жакет',
        # платья и сарафаны
        'плать', 'сарафан',
        # костюмы / комплекты
        'костюм', 'комплект',
        # футболки, топы, майки
        'футболк', 'топ', 'майк',
        # лонгсливы, толстовки, худи, свитшоты, свитеры, водолазки
        'лонгслив', 'толстовк', 'худи', 'свитшот', 'свитер', 'водолазк', 'пуловер',
        'джемпер', 'кофт',
        # брюки, шорты, юбки, штаны, легинсы
        'брюк', 'джинс', 'штан', 'шорт', 'юбк', 'легинс', 'лосин', 'велосипедк', 'бридж',
        # боди, комбинезоны, песочники, ползунки, слипы, распашонки
        'боди', 'комбинезон', 'песочник', 'ползунк', 'распашонк', 'штанишк',
        # рубашки, блузки, сорочки
        'рубаш', 'блузк', 'сорочк',
        # бельё, пижамы, халаты, ночные сорочки
        'бель', 'пижам', 'халат', 'трус', 'бюстгальт', 'комбинаци',
        # бомберы
        'бомбер', 'олимпийк',
        # детская специфика
        'чепчик', 'пинетк', 'шапочк', 'кигуруми',
        # одежда общая
        'одежд', 'униформ', 'форм',
    ],
    'shoes': [
        # обувь
        'обув', 'кроссовк', 'кед', 'ботин', 'сапог', 'туфли', 'туфел',
        'сандали', 'босоножк', 'тапочк', 'тапки', 'мокасин', 'валенк',
        'угги', 'дутик', 'полусапог', 'слипон', 'эспадрил', 'балетк',
        'лоферы', 'лофер', 'челси', 'дерби',
        # v39.11: добавлены частые слова которые ищутся в названиях карточек WB
        'сникер',    # сникеры — синоним кроссовок
        'мюли',      # сабо/мюли
        'сабо',
        'галоши',
        'резиновая обув', 'резиновые сапоги',
        'каблук',    # туфли на каблуке
        'на шнурках',
    ],
    'toys': [
        # игрушки
        'игрушк', 'кукл', 'мишк',
        # конструкторы и наборы
        'конструктор', 'пазл', 'мозаик',
        # настольные игры
        'настольн', 'игр',  # "Игры" покрывает "Игры настольные"
        # машинки и техника
        'машинк', 'самолет', 'танк игр', 'трактор игр', 'верто',
        # антистресс
        'антистресс', 'попит', 'pop it', 'спиннер', 'сквиш',
        # фигурки и наборы
        'фигурк', 'трансформер',
        # для младенцев
        'погремушк', 'комфортер', 'неваляшк', 'юла', 'каталк',
        # для ванной
        'для ванн',
        # карнавальные / костюмы для игры
        'карнавал', 'маскарад',
        # развивающие
        'развивающ', 'обучающ',
        # радиоуправляемые
        'радиоуправляем', 'на радиоуправлении',
        # шарики, мячи
        'шарик', 'мяч игров',
    ],
    'kids_accessories': [
        'рюкзак', 'портфел', 'ранец',
        'шапк', 'панам', 'кепк', 'бейсболк',
        'варежк', 'перчатк', 'рукавиц',
        'шарф', 'снуд',
        'бант', 'заколк', 'резинк для волос', 'обруч',
        'сумк', 'кошелёк', 'кошелек', 'пенал',
    ],
    'baby_gear': [
        'коляск', 'автокресл', 'манеж', 'переноск', 'кроватк', 'пеленальн',
        'стульчик для кормлен', 'ходунк', 'качел детск', 'шезлонг детск',
        'кенгуру', 'слинг',
    ],
    'cosmetics': [
        'космет', 'крем', 'шампун', 'мыло', 'гель для душ', 'дезодорант',
        'духи', 'парфюм', 'туш', 'помад', 'лак для ногт', 'маска для лиц',
        'тоник', 'лосьон', 'скраб', 'пенка',
    ],
    'electronics': [
        'наушник', 'колонк', 'клавиатур', 'мыш компью', 'роутер', 'провод',
        'кабел', 'зарядк', 'зарядное', 'powerbank', 'аккумулятор',
        'usb', 'переходник', 'адаптер', 'смартфон', 'телефон',
    ],
    'home': [
        'постельн', 'наволочк', 'подушк', 'одеял', 'плед', 'полотенц',
        'покрывал', 'простын', 'скатерт', 'штор', 'занавес', 'коврик', 'ковер',
        'плед',
    ],
    'kitchenware': [
        'кастрюл', 'сковород', 'нож кухон', 'тарелк', 'кружк', 'чашк',
        'столов прибор', 'разделочн доск', 'венчик', 'половник',
    ],
    'food': [
        'смес молочн', 'каш', 'пюре дет', 'напиток дет', 'сок дет',
        'печень', 'питан',
    ],
}

# v27.9.x: домен «бытовая техника» (appliances) — для поиска по бренду (напр.
# indesit, bosch, electrolux), чтобы выдача ограничивалась техникой, а не
# попадали чехлы/аксессуары/одежда. Добавляем отдельно, чтобы не трогать большие
# литералы выше.
DOMAIN_PRODUCT_TYPES['appliances'] = [
    # крупная бытовая техника
    'стиральные машины', 'холодильники', 'посудомоечные машины', 'плиты',
    'духовые шкафы', 'варочные панели', 'микроволновые печи', 'вытяжки',
    'морозильные камеры', 'пылесосы', 'водонагреватели', 'кондиционеры',
    'сушильные машины', 'встраиваемые духовые шкафы', 'газовые плиты',
    'электроплиты', 'винные шкафы', 'сплит-системы',
    # мелкая кухонная техника
    'чайники электрические', 'мультиварки', 'блендеры', 'миксеры', 'тостеры',
    'кофемашины', 'кофеварки', 'мясорубки', 'соковыжималки', 'кухонные комбайны',
    'грили', 'аэрогрили', 'фритюрницы', 'хлебопечки', 'электрочайники',
    'пароварки', 'вафельницы', 'электрогрили', 'кулеры для воды',
    # климат, уход, мелкая бытовая
    'увлажнители воздуха', 'очистители воздуха', 'обогреватели', 'конвекторы',
    'тепловентиляторы', 'вентиляторы', 'фены', 'выпрямители для волос',
    'утюги', 'отпариватели', 'ирригаторы', 'электробритвы',
    'машинки для стрижки', 'триммеры', 'весы напольные', 'весы кухонные',
    'роботы-пылесосы', 'парогенераторы',
]
DOMAIN_SUBJECT_NAME_KEYWORDS['appliances'] = [
    # крупная
    'стиральн', 'холодильник', 'посудомоечн', 'плита', 'плиты', 'духов',
    'варочн', 'микроволнов', 'вытяжк', 'морозильник', 'морозильн', 'пылесос',
    'водонагреватель', 'кондиционер', 'сплит-систем', 'сушильн', 'винн', 'шкаф',
    'бытовая техник', 'духовой шкаф',
    # мелкая кухонная
    'чайник', 'мультиварк', 'блендер', 'миксер', 'тостер', 'кофемашин',
    'кофеварк', 'мясорубк', 'соковыжималк', 'комбайн', 'гриль', 'фритюр',
    'хлебопечк', 'пароварк', 'вафельниц', 'кулер', 'ростер', 'электропечь',
    # климат/уход/мелкая
    'увлажнитель', 'очиститель воздуха', 'обогреватель', 'конвектор',
    'тепловентилятор', 'вентилятор', 'фен ', 'фены', 'выпрямитель', 'плойк',
    'утюг', 'отпариватель', 'парогенератор', 'ирригатор', 'электробритв',
    'машинка для стрижки', 'триммер', 'весы', 'климат', 'обогрев',
]
DOMAIN_SUBJECT_WHITELIST['appliances'] = []  # фильтруем по ключевым словам названия


def is_card_relevant_for_domain(subject: str, domain: str, subject_name: str = '', product_name: str = '') -> bool:
    """v39.10: проверяет относится ли карточка к ожидаемому домену.

    БЕЗОПАСНОСТЬ ПРЕЖДЕ ВСЕГО: лучше пропустить мусор, чем потерять настоящий товар.
    Поэтому проверяем В ТРИ источника, в порядке надёжности:
      1. Текст subjectName (например «Кроссовки», «Платья детские») — самое надёжное.
      2. Текст product_name (название карточки) — тоже надёжное.
      3. subject ID через whitelist — fallback, может быть неточным.

    КАРТОЧКА ПРОПУСКАЕТСЯ если хотя бы один источник подтвердил релевантность.
    Это значит: даже если subjectName совсем не известен или whitelist неполный —
    подходящий товар всё равно попадёт по тексту названия.

    БЛОКИРУЕМ только если ВСЕ три источника явно показывают что это не наш домен,
    ИЛИ если для домена нет вообще никаких сигналов (старая логика).
    """
    if not domain or domain == 'unknown':
        return True

    # v27.9.x: поддержка НЕСКОЛЬКИХ доменов (мультикатегория бренда) — "clothing,shoes".
    # Карточка релевантна, если подходит ХОТЯ БЫ ПОД ОДИН из доменов.
    if ',' in str(domain):
        return any(is_card_relevant_for_domain(subject, d.strip(), subject_name, product_name)
                   for d in str(domain).split(',') if d.strip())

    keywords = DOMAIN_SUBJECT_NAME_KEYWORDS.get(domain, [])
    wl = DOMAIN_SUBJECT_WHITELIST.get(domain) or set()

    # v46: НЕГАТИВНЫЙ фильтр для «взрослых» категорий (техника/электроника/посуда/дом):
    # отсеиваем ИГРУШЕЧНЫЕ товары. WB по «стиральные машины» иногда возвращает
    # «детская стиральная машина игрушечная» — в названии есть «стиральн», поэтому
    # обычный фильтр её пропускал. Если subject/название = игрушка — это не техника.
    if domain in ('appliances', 'electronics', 'kitchenware', 'home'):
        _toy = ('игрушк', 'игрушеч', 'игрушечн')
        sn = norm_text(subject_name or '')
        pn = norm_text(product_name or '')
        if any(t in sn for t in _toy) or any(t in pn for t in _toy):
            return False

    # Если нет ни ключевых слов ни whitelist'а для этого домена — фильтр не применяем
    if not keywords and not wl:
        return True

    # 1. Проверка subjectName
    if subject_name and keywords:
        sn_low = norm_text(subject_name)
        if any(kw in sn_low for kw in keywords):
            return True

    # 2. Проверка product_name (название карточки) — главный источник!
    # «Кроссовки демисезонные для детей» содержит «кроссовк» → shoes.
    if product_name and keywords:
        pn_low = norm_text(product_name)
        if any(kw in pn_low for kw in keywords):
            return True

    # 3. Проверка subjectId через whitelist
    if wl and (subject or '').strip() in wl:
        return True

    # Ни один источник не подтвердил → отсеиваем
    return False


def generate_query_variants(base_query: str, profile: str = "auto", max_variants: int = 250) -> List[str]:
    """v39.7: универсальная генерация запросов под любую категорию.

    Логика:
      1) Определяем домен (clothing/toys/shoes/cosmetics/...) либо по profile, либо автоматически по словам запроса.
      2) Берём типы товаров этого домена (DOMAIN_PRODUCT_TYPES).
      3) Применяем универсальные модификаторы (UNIVERSAL_MODIFIERS) + APPAREL_MODIFIERS только для clothing/shoes.
      4) Базовый запрос пользователя — всегда в списке.

    Главное отличие от старой версии: для игрушек/обуви/косметики и т.п. не применяются «школьные/розовые» которые превращали поиск в одежду.
    """
    base_query = " ".join(base_query.split())
    profile = (profile or "auto").lower()

    # Маппинг профилей на домены. v46: ДОБАВЛЕНЫ все недостающие категории
    # (food/kitchenware/baby_gear/kids_accessories/appliances) + RU-алиасы. Раньше
    # их тут не было, поэтому для них домен «не распознавался» и вариантов запроса
    # генерировалось мало (только базовый + универсальные) — отсюда «у одежды много,
    # у продуктов/посуды мало».
    profile_to_domain = {
        'clothing': 'clothing', 'одежда': 'clothing',
        'shoes': 'shoes', 'обувь': 'shoes',
        'toys': 'toys', 'игрушки': 'toys',
        'cosmetics': 'cosmetics', 'косметика': 'cosmetics',
        'electronics': 'electronics', 'электроника': 'electronics',
        'home': 'home', 'дом': 'home', 'дом и текстиль': 'home', 'текстиль': 'home',
        'kitchenware': 'kitchenware', 'посуда': 'kitchenware',
        'food': 'food', 'продукты': 'food', 'питание': 'food',
        'baby_gear': 'baby_gear', 'детский транспорт': 'baby_gear',
        'kids_accessories': 'kids_accessories', 'детские аксессуары': 'kids_accessories',
        'appliances': 'appliances', 'бытовая техника': 'appliances', 'техника': 'appliances',
        'auto': None,
    }
    domain = profile_to_domain.get(profile)
    if domain is None:
        domain = detect_query_domain(base_query)

    types = DOMAIN_PRODUCT_TYPES.get(domain, [])
    is_kids = any(w in base_query.lower() for w in ('детск', 'для детей', 'для малыш', 'малышам'))
    is_apparel_like = domain in ('clothing', 'shoes', 'kids_accessories')
    # v46: «взрослые» категории — техника/электроника/посуда/дом. К ним НЕЛЬЗЯ
    # применять детские/возрастные модификаторы: «стиральные машины детские» тянет
    # ИГРУШЕЧНЫЕ стиральные машины, «бытовая техника для мальчиков» — детские товары.
    _appliance_like = domain in ('appliances', 'electronics', 'kitchenware', 'home')

    variants: List[str] = [base_query]

    # Стратегия 1: исходный запрос + универсальные (возрастные/гендерные) модификаторы.
    # Только если категория НЕ «взрослая» (иначе мусор: «бытовая техника детские»).
    if not _appliance_like:
        for m in UNIVERSAL_MODIFIERS:
            variants.append(f'{base_query} {m}')

    # Стратегия 2: для одежды/обуви — ещё цветовые/сезонные модификаторы
    if is_apparel_like:
        for m in APPAREL_MODIFIERS:
            variants.append(f'{base_query} {m}')

    # Стратегия 3: типы товаров домена в чистом виде ("куртки", "мягкие игрушки", "кроссовки")
    for t in types:
        if is_kids and not any(w in t.lower() for w in ('детск', 'для', 'малыш')):
            variants.append(f'детские {t}')
        else:
            variants.append(t)

    # Стратегия 4: ТИП + возрастной/гендерный модификатор. Для «детских» категорий
    # это кратно увеличивает охват. Для «взрослых» (техника/электроника/посуда/дом)
    # НЕ применяем — иначе «стиральные машины детские» = игрушки. Там охват берём
    # из расширенного списка ТИПОВ (стратегия 3).
    if domain and domain != 'unknown' and not _appliance_like:
        for t in types[:25]:
            for m in UNIVERSAL_MODIFIERS[:6]:
                variants.append(f'{t} {m}')

    # Если домен неизвестен — добавляем только базовый и универсальные комбинации.
    # Это не мусорит выдачу случайными цветами для категорий типа «канцелярия».
    if domain == 'unknown':
        variants = [base_query]
        for m in UNIVERSAL_MODIFIERS:
            variants.append(f'{base_query} {m}')

    # Дедупликация + ремонт «детские детские» / «детские детская»
    clean: List[str] = []
    seen = set()
    for q in variants:
        q = re.sub(r"\bдетские\s+детск(ие|ая|ое|ий)\b", "детские", q, flags=re.I)
        q = re.sub(r"\s+", " ", q).strip()
        if q and q.lower() not in seen:
            seen.add(q.lower())
            clean.append(q)
        if len(clean) >= max_variants:
            break
    return clean

def recursive_find_products(obj: Any) -> List[Dict[str, Any]]:
    res = []
    if isinstance(obj, dict):
        if isinstance(obj.get("products"), list):
            for p in obj["products"]:
                if isinstance(p, dict) and ("id" in p or "nmId" in p or "nm_id" in p):
                    res.append(p)
        for v in obj.values():
            if isinstance(v, (dict, list)):
                res.extend(recursive_find_products(v))
    elif isinstance(obj, list):
        for x in obj:
            if isinstance(x, (dict, list)):
                res.extend(recursive_find_products(x))
    return res


# =============================================================================
# v27.9.x: ПОЛНЫЙ каталог бренда. Текстовый поиск «reebok» обрезается WB на ~сотне
# карточек. Чтобы собрать ВСЕ товары бренда, находим brandId (через resultset=
# filters и подсказки WB), затем листаем прямой бренд-каталог
# catalog.wb.ru/brands/v2/catalog?brand=<id>. Логика портирована из main_brand.py.
# =============================================================================
def _recursive_find_brand_filter_ids(obj: Any, brand: str) -> List[str]:
    wanted = norm_key_brand(brand)
    ids: List[str] = []
    def add_id(v: Any):
        if v is None:
            return
        txt = str(v).strip()
        if txt and re.fullmatch(r"\d{1,10}", txt) and txt not in ids:
            ids.append(txt)
    def is_brand_like(x: Any) -> bool:
        k = norm_key_brand(str(x or ""))
        return bool(k) and (k == wanted or wanted in k or k in wanted)
    def walk(x: Any, in_brand_filter: bool = False):
        if isinstance(x, dict):
            name = x.get("name") or x.get("title") or x.get("value") or x.get("label")
            key = norm_key_brand(str(x.get("key") or x.get("type") or x.get("id") or ""))
            header = norm_key_brand(str(name or ""))
            brand_filter = in_brand_filter or header in {"бренд", "brand", "brands"} or key in {"fbrand", "brand", "brands"}
            if is_brand_like(name):
                for k in ("id", "value", "key"):
                    if k in x and str(x.get(k)).strip() != str(name).strip():
                        add_id(x.get(k))
                for k in ("ids", "values"):
                    if isinstance(x.get(k), list):
                        for v in x[k]:
                            add_id(v)
            if brand_filter:
                for arr_key in ("items", "values", "list", "filters"):
                    arr = x.get(arr_key)
                    if isinstance(arr, list):
                        for item in arr:
                            if isinstance(item, dict) and is_brand_like(item.get("name") or item.get("title") or item.get("value") or item.get("label")):
                                for k in ("id", "value", "key"):
                                    add_id(item.get(k))
            for v in x.values():
                if isinstance(v, (dict, list)):
                    walk(v, brand_filter)
        elif isinstance(x, list):
            for v in x:
                walk(v, in_brand_filter)
    walk(obj)
    return ids[:20]


def _find_brand_ids_in_suggestions(obj: Any, brand: str) -> List[str]:
    wanted = norm_key_brand(brand)
    ids: List[str] = []
    def walk(x: Any):
        if isinstance(x, dict):
            name = x.get("name") or x.get("text") or x.get("title") or x.get("value")
            nk = norm_key_brand(str(name or ""))
            type_hint = norm_key_brand(str(x.get("type") or x.get("entity") or ""))
            is_brand_entity = ("brand" in type_hint) or bool(x.get("brandId")) or bool(x.get("brand_id"))
            name_matches = nk and (nk == wanted or wanted in nk or nk in wanted)
            if is_brand_entity and (name_matches or not name):
                for k in ("brandId", "brand_id", "id"):
                    v = x.get(k)
                    if v is not None and re.fullmatch(r"\d{1,10}", str(v).strip()):
                        if str(v).strip() not in ids:
                            ids.append(str(v).strip())
            for v in x.values():
                if isinstance(v, (dict, list)):
                    walk(v)
        elif isinstance(x, list):
            for v in x:
                walk(v)
    walk(obj)
    return ids


async def discover_brand_filter_ids(session: aiohttp.ClientSession, brand: str, timeout: float = 10.0) -> List[str]:
    """Находит числовой brandId(ы) WB для прямого бренд-каталога."""
    q = urllib.parse.quote(brand)
    found: List[str] = []
    wanted = norm_key_brand(brand)

    # v27.9.x: САМЫЙ НАДЁЖНЫЙ источник — brandId прямо из ТОВАРОВ поисковой выдачи.
    # У каждого товара WB есть поля brand + brandId. Берём id у товаров, чей бренд
    # совпадает с искомым. Это работает даже когда resultset=filters/подсказки
    # отдают пусто или сменили структуру (из-за чего brandId «не находился»).
    product_urls = [
        f"https://search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page=1&query={q}&resultset=catalog&sort=popular&spp=30&suppressSpellcheck=false",
        f"https://u-search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page=1&query={q}&resultset=catalog&sort=popular&spp=30&suppressSpellcheck=false",
    ]
    _bid_counts: Dict[str, int] = {}
    for url in product_urls:
        data = await fetch_json(session, url, timeout=timeout)
        if not data:
            continue
        for p in recursive_find_products(data):
            bid = p.get("brandId") or p.get("brand_id") or p.get("brandID")
            pbrand = norm_key_brand(str(p.get("brand") or p.get("brandName") or ""))
            if not bid or not pbrand:
                continue
            if not re.fullmatch(r"\d{1,10}", str(bid).strip()):
                continue
            if pbrand == wanted or wanted in pbrand or pbrand in wanted:
                _bid_counts[str(bid).strip()] = _bid_counts.get(str(bid).strip(), 0) + 1
        if _bid_counts:
            break
    # самые частые brandId — впереди (основной бренд, а не случайные совпадения)
    for bid, _cnt in sorted(_bid_counts.items(), key=lambda kv: -kv[1]):
        if bid not in found:
            found.append(bid)

    filter_urls = [
        f"https://search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30&suppressSpellcheck=false",
        f"https://u-search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30&suppressSpellcheck=false",
        f"https://search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page=1&query={q}&resultset=filters&spp=30",
    ]
    for url in filter_urls:
        data = await fetch_json(session, url, timeout=timeout)
        if not data:
            continue
        for bid in _recursive_find_brand_filter_ids(data, brand):
            if bid not in found:
                found.append(bid)
    suggest_urls = [
        f"https://search.wb.ru/exactmatch/common/v5/search?query={q}&resultset=suggestions",
        f"https://suggests.wb.ru/api/v6/hint?query={q}&gender=common&locale=ru",
    ]
    for url in suggest_urls:
        data = await fetch_json(session, url, timeout=timeout)
        if not data:
            continue
        for bid in _find_brand_ids_in_suggestions(data, brand):
            if bid not in found:
                found.append(bid)
    return found[:10]

async def fetch_json(session: aiohttp.ClientSession, url: str, timeout: float = 12.0) -> Optional[Dict[str, Any]]:
    try:
        async with session.get(url, timeout=timeout) as r:
            if r.status != 200:
                return None
            txt = await r.text()
            return json.loads(txt)
    except Exception:
        return None


def _card_fields_for_result(card: Card) -> Dict[str, Any]:
    """v39.14: возвращает dict с WB-полями Card для подстановки в ResultRow(...).

    Использование: ResultRow(query=..., nm_id=..., **_card_fields_for_result(card), status=..., ...)
    """
    return {
        "product_name": card.product_name,
        "brand": card.brand,
        "subject": card.subject,
        "product_url": card.product_url or product_url(card.nm_id),
        "price_rub": getattr(card, "price_rub", 0.0),
        "sale_price_rub": getattr(card, "sale_price_rub", 0.0),
        "seller_name": getattr(card, "seller_name", ""),
        "supplier_id": getattr(card, "supplier_id", ""),
        "rating": getattr(card, "rating", 0.0),
        "feedbacks": getattr(card, "feedbacks", 0),
        "is_original": getattr(card, "is_original", ""),
        "docs_verified": getattr(card, "docs_verified", ""),
        "colors": getattr(card, "colors", ""),
        "wb_root": getattr(card, "wb_root", ""),
    }


# v39.14: глобальный кэш расширенных полей FSA-документа по registry_url.
# Когда HTTP-парсер успешно разобрал JSON документа, он кладёт сюда ВСЕ извлечённые
# поля (applicant, manufacturer, tnved, scheme, technical_regulation, date_start, date_end).
# При формировании ResultRow в _flush_url_to_store эти поля подставляются из кэша.
_FSA_EXTENDED_FIELDS_CACHE: Dict[str, Dict[str, str]] = {}

# v40.2: circuit breaker для HTTP-парсинга FSA через curl_cffi.
# Если в сети пользователя FSA блокирует curl_cffi (отдаёт 403), нет смысла
# пытаться HTTP на каждой карточке — это только тратит время. После N неудач
# подряд HTTP-попытки отключаются на весь прогон, парсинг идёт через браузер.
_FSA_HTTP_FAILS = 0
_FSA_HTTP_DISABLED = False
_FSA_HTTP_FAIL_LIMIT = 5

# v45.2: куки сессии ФСА, снятые с настоящего браузера (он проходит JS-антибот,
# который ставит сессионную куку). С этими куками curl_cffi обращается к API ФСА
# напрямую — быстро и одним лёгким запросом, без полной отрисовки SPA. Заполняется
# после первого УСПЕШНОГО браузерного парсинга и обновляется, если куки протухли.
_FSA_SESSION_COOKIES: Dict[str, str] = {}
_FSA_COOKIE_HTTP_OK = 0   # сколько документов добыто быстрым HTTP по кукам
_FSA_COOKIE_HTTP_FAIL = 0  # сколько раз HTTP по кукам не сработал (куки протухли/нет токена)

# v45.6: АВТО-ВОССТАНОВЛЕНИЕ при блокировке FSA. Когда FSA начинает массово отдавать
# сетевые ошибки/таймауты (rate-limit/временный бан), программа сама ставит FSA на
# паузу (cooldown с экспоненциальным ростом), сбрасывает «отравленную» сессию (куки)
# и потом продолжает. Это снимает временный rate-limit без участия пользователя.
# Полный сетевой бан IP пауза не лечит — тогда после N циклов FSA отпускается, а
# недобранное добивается кнопкой «Повторить упавшие FSA».
_FSA_CONSEC_FAILS = 0       # подряд идущих неудач FSA (сбрасывается успехом)
_FSA_COOLDOWN_UNTIL = 0.0   # время (epoch), до которого FSA на паузе
_FSA_COOLDOWN_CYCLES = 0    # сколько раз уже включали паузу за прогон
# v46: САМОНАСТРОЙКА темпа в медленном режиме. Множитель к базовой паузе: растёт при
# сбоях ФСА (тормозим), снижается при череде успехов (ускоряемся) — программа сама
# нащупывает самый быстрый безопасный темп вокруг 20-30 док/мин.
_FSA_SLOW_MULT = 1.0
_FSA_SLOW_OK = 0

# v27.9.x: один раз за прогон сохраняем сырой JSON-ответ API ФСА в файл —
# чтобы по реальной структуре доразобрать status/scheme/название (диагностика).
_FSA_SAMPLE_DUMPED = False

async def collect_one_query(session: aiohttp.ClientSession, query: str, per_query_limit: int = 250, sort: str = "popular", domain: str = '', stats: Optional[Dict[str, int]] = None, page: int = 1, fbrand_ids: Optional[List[str]] = None) -> List[Card]:
    """v39.9: subject отдельно от subjectName.

    WB API в v18 возвращает оба поля:
      - subject (subjectId) — число, например 168
      - subjectName — текст, например "Куртки детские"

    Раньше мы складывали их в одно поле через `or` — терялся subjectName.
    Теперь сохраняем subjectId в поле `subject` (для compare и xlsx),
    а subjectName используем для фильтрации (надёжнее чем угадывать ID).

    v39.11: добавлен опциональный dict stats для подсчёта сколько карточек
    WB вернул всего и сколько было отфильтровано. Это даёт пользователю
    видеть «WB вернул 0» vs «фильтр отбросил».

    v27.9.x: page — номер страницы WB-поиска. Нужно для ПОИСКА ПО БРЕНДУ одним
    «чистым» запросом (без расширения категориями): листаем страницы, чтобы
    собрать все товары бренда, а не только первую (~100 шт).
    """
    cards: List[Card] = []
    q = urllib.parse.quote(query)
    _pg = max(1, int(page or 1))
    # несколько актуальных хостов/версий, первый обычно быстрее
    urls = []
    # v27.9.x: ПОЛНЫЙ бренд-каталог по brandId — отдаёт ВСЕ товары бренда (а не
    # топ поисковой выдачи, обрезанный на ~сотне). Ставим ПЕРВЫМ источником.
    for bid in (fbrand_ids or []):
        _b = urllib.parse.quote(str(bid))
        urls.append(
            f"https://catalog.wb.ru/brands/v2/catalog?ab_testing=false&appType=1&brand={_b}"
            f"&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={_pg}&sort={sort}&spp=30")
        urls.append(
            f"https://catalog.wb.ru/brands/catalog?appType=1&brand={_b}"
            f"&curr=rub&dest=-1257786&lang=ru&page={_pg}&sort={sort}&spp=30")
    # поисковая выдача (+ опционально фильтр по бренду fbrand)
    _search_bases = [
        f"https://search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={_pg}&query={q}&resultset=catalog&sort={sort}&spp=30&suppressSpellcheck=false",
        f"https://u-search.wb.ru/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=13&lang=ru&page={_pg}&query={q}&resultset=catalog&sort={sort}&spp=30&suppressSpellcheck=false",
        f"https://search.wb.ru/exactmatch/ru/common/v13/search?appType=1&curr=rub&dest=-1257786&lang=ru&page={_pg}&query={q}&resultset=catalog&sort={sort}&spp=30",
    ]
    for _bu in _search_bases:
        for bid in (fbrand_ids or []):
            urls.append(_bu + "&fbrand=" + urllib.parse.quote(str(bid)))
        urls.append(_bu)
    for url in urls:
        data = await fetch_json(session, url)
        if not data:
            continue
        products = recursive_find_products(data)
        if products:
            if stats is not None:
                stats['wb_returned'] = stats.get('wb_returned', 0) + len(products)
            for p in products[:per_query_limit]:
                nm = safe_int(p.get("id") or p.get("nmId") or p.get("nm_id"))
                if nm <= 1000:
                    continue
                name = str(p.get("name") or p.get("productName") or "")
                brand = str(p.get("brand") or p.get("brandName") or "")
                # v39.9: subject (ID) и subjectName (текст) — отдельно
                subj_id = str(p.get("subjectId") or p.get("subject") or "")
                subj_name = str(p.get("subjectName") or "")
                # v39.10: фильтр учитывает И subjectName И название товара.
                # Если subjectName пустой (WB иногда не отдаёт его), название карточки
                # всё равно подскажет правильный домен.
                if domain and not is_card_relevant_for_domain(subj_id, domain, subj_name, name):
                    if stats is not None:
                        stats['filtered_out'] = stats.get('filtered_out', 0) + 1
                    continue
                # Для совместимости с compare_product_names и xlsx: subject = subjectName если есть, иначе ID
                # Это даёт более информативные данные в выгрузке.
                subject_for_card = subj_name or subj_id
                # v39.14: извлекаем дополнительные поля из WB API
                price_rub = 0.0
                sale_price_rub = 0.0
                # WB всегда отдаёт цены в КОПЕЙКАХ — и в классических полях
                # priceU/salePriceU, и в новом sizes[].price.{total,product,basic}.
                # v27.7: убрана ошибочная эвристика "/100 если >10000", из-за
                # которой дешёвые товары (<100 ₽ = <10000 коп) показывались
                # тысячами, а часть дорогих делилась неверно. Теперь всегда /100.
                try:
                    sv = p.get("salePriceU")
                    if isinstance(sv, (int, float)) and sv > 0:
                        sale_price_rub = round(float(sv) / 100.0, 2)
                    pv = p.get("priceU")
                    if isinstance(pv, (int, float)) and pv > 0:
                        price_rub = round(float(pv) / 100.0, 2)
                    if not price_rub or not sale_price_rub:
                        sizes = p.get("sizes") or []
                        for s in sizes:
                            pr = s.get("price") or {}
                            basic = pr.get("basic")
                            total = pr.get("total") or pr.get("product")
                            if not price_rub and isinstance(basic, (int, float)) and basic > 0:
                                price_rub = round(float(basic) / 100.0, 2)
                            if not sale_price_rub and isinstance(total, (int, float)) and total > 0:
                                sale_price_rub = round(float(total) / 100.0, 2)
                            if price_rub and sale_price_rub:
                                break
                    # Если базовая цена не найдена, но есть цена со скидкой —
                    # используем её как базовую (лучше, чем 0).
                    if not price_rub and sale_price_rub:
                        price_rub = sale_price_rub
                except Exception:
                    pass
                # v43: WB-поиск отдаёт продавца в поле supplier (ТЕКСТ), не supplierName.
                # Подтверждено реальным разбором network-ответа WB.
                seller_name = str(
                    p.get("supplier") or p.get("supplierName") or p.get("sellerName") or ""
                ).strip()
                supplier_id = str(p.get("supplierId") or p.get("supplierID") or p.get("sellerId") or "")
                # v43: оригинальность товара — по флагам/полям карточки WB.
                # supplierFlags и отдельные поля могут содержать признак оригинала.
                is_original = _detect_wb_original(p)
                # v27.9.x: дополнительные поля карточки (защитно, пусто если нет).
                color_names: List[str] = []
                for _c in (p.get("colors") or []):
                    if isinstance(_c, dict):
                        _cn = str(_c.get("name") or "").strip()
                        if _cn:
                            color_names.append(_cn)
                colors_str = ", ".join(dict.fromkeys(color_names))
                wb_root = str(p.get("root") or "")
                rating = 0.0
                try:
                    r_v = p.get("rating") or p.get("reviewRating") or p.get("supplierRating")
                    if isinstance(r_v, (int, float)):
                        rating = float(r_v)
                except Exception:
                    pass
                feedbacks = safe_int(p.get("feedbacks") or p.get("nmFeedbacks") or 0)
                pics_count = safe_int(p.get("pics") or 0)
                in_stock = safe_int(p.get("totalQuantity") or 0)
                cards.append(Card(
                    nm_id=nm, product_name=name, brand=brand, subject=subject_for_card,
                    source_query=query, product_url=product_url(nm),
                    price_rub=price_rub, sale_price_rub=sale_price_rub or price_rub,
                    seller_name=seller_name, supplier_id=supplier_id,
                    rating=rating, feedbacks=feedbacks, pics_count=pics_count, in_stock=in_stock,
                    is_original=is_original, colors=colors_str, wb_root=wb_root,
                    view_flags=(int(p.get("viewFlags")) if isinstance(p.get("viewFlags"), int) else 0),
                ))
            break
    return cards

async def collect_cards(args) -> List[Card]:
    if args.input_csv:
        cards = []
        with open(args.input_csv, "r", encoding="utf-8-sig", newline="") as f:
            rdr = csv.DictReader(f)
            for row in rdr:
                nm = safe_int(row.get("nm_id") or row.get("nmId") or row.get("id"))
                if nm:
                    cards.append(Card(
                        nm_id=nm,
                        product_name=row.get("product_name") or row.get("name") or "",
                        brand=row.get("brand") or "",
                        subject=row.get("subject") or "",
                        source_query=row.get("query") or args.query or "",
                        product_url=row.get("product_url") or product_url(nm)
                    ))
                if len(cards) >= args.limit:
                    break
        return cards

    _brand_search = bool((getattr(args, 'brand', '') or '').strip()) and \
        (getattr(args, 'brand_match', 'any') or 'any') != 'any'

    # v39.7 + v27.9.x: профиль → домен. Расширено: identity для всех доменных
    # ключей (clothing/shoes/.../appliances) и RU-алиасы. Нужно для ВЫБОРА
    # КАТЕГОРИИ при поиске по бренду (reebok→одежда, indesit→бытовая техника).
    profile_to_domain = {
        'clothing': 'clothing', 'одежда': 'clothing',
        'shoes': 'shoes', 'обувь': 'shoes',
        'toys': 'toys', 'игрушки': 'toys',
        'kids_accessories': 'kids_accessories', 'детские аксессуары': 'kids_accessories',
        'baby_gear': 'baby_gear', 'детский транспорт': 'baby_gear',
        'cosmetics': 'cosmetics', 'косметика': 'cosmetics',
        'electronics': 'electronics', 'электроника': 'electronics',
        'appliances': 'appliances', 'бытовая техника': 'appliances', 'техника': 'appliances',
        'home': 'home', 'дом': 'home', 'дом и текстиль': 'home', 'текстиль': 'home',
        'kitchenware': 'kitchenware', 'посуда': 'kitchenware',
        'food': 'food', 'продукты': 'food', 'питание': 'food',
    }
    # v27.9.x: КАТЕГОРИИ ТОВАРОВ для бренда — можно НЕСКОЛЬКО через запятую
    # (--query-profile clothing,shoes). Если выбраны — сужаем выдачу до них.
    _profile_raw = (getattr(args, 'query_profile', '') or 'auto').strip().lower()
    _profile_keys = [p.strip() for p in _profile_raw.split(',') if p.strip()]
    _brand_domains = []
    for _k in _profile_keys:
        if _k in ('auto', '', 'any', 'любая'):
            continue
        _d = profile_to_domain.get(_k)
        if _d and _d not in _brand_domains:
            _brand_domains.append(_d)
    _profile_key = _profile_keys[0] if _profile_keys else 'auto'

    variants = generate_query_variants(args.query, args.query_profile, args.max_expanded_queries)
    if _brand_search:
        # Бренд: чистый бренд + (если выбраны категории) «бренд + тип товара» из
        # КАЖДОЙ выбранной категории. Больше карточек, не выходя за категории.
        if _brand_domains:
            variants = [args.brand]
            for _d in _brand_domains:
                variants += [f"{args.brand} {t}" for t in DOMAIN_PRODUCT_TYPES.get(_d, [])]
        else:
            variants = [args.brand]
    elif args.limit <= args.per_query_limit and not args.auto_expand:
        variants = [args.query]

    strict_filter_enabled = getattr(args, 'strict_domain_filter', True)
    if _brand_search:
        # v45.3: ПОИСК ПО БРЕНДУ. Категория нужна НЕ чтобы резать выдачу, а чтобы
        # ДОБРАТЬ больше карточек бренда: по «голому» бренду WB отдаёт одну выдачу
        # (~сотню), а варианты «бренд + тип товара категории» (стиральные машины,
        # чайники, фены, …) вытаскивают карточки, которых в первой выдаче не было.
        # Сами карточки по категории НЕ отсеиваем — бренд-фильтр уже гарантирует, что
        # это товары нужного бренда, и терять их нельзя. Поэтому доменный фильтр ВЫКЛ.
        domain = ''
        use_filter = False
    elif _brand_domains:
        # Поиск ПО ЗАПРОСУ с явно выбранной категорией — здесь фильтр уместен
        # (уточняет выдачу запроса; бренда, который бы гарантировал релевантность, нет).
        domain = ','.join(_brand_domains)
        use_filter = strict_filter_enabled and any(
            DOMAIN_SUBJECT_NAME_KEYWORDS.get(d) or DOMAIN_SUBJECT_WHITELIST.get(d)
            for d in _brand_domains)
    else:
        # Обычный поиск без явной категории — авто-домен по тексту запроса.
        domain = profile_to_domain.get(_profile_key) or detect_query_domain(args.query)
        has_signals = bool(DOMAIN_SUBJECT_WHITELIST.get(domain)) or bool(DOMAIN_SUBJECT_NAME_KEYWORDS.get(domain))
        use_filter = strict_filter_enabled and domain and has_signals
    filter_label = f"domain={domain} subject_filter={'ON' if use_filter else 'OFF'}"
    print(f"Получаю список карточек через JSON-каталог WB: базовый query='{args.query}', вариантов={len(variants)}, общий limit={args.limit}, collect-workers={args.collect_workers}; {filter_label}")

    timeout = aiohttp.ClientTimeout(total=18)
    headers = {
        "User-Agent": args.user_agent,
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Origin": "https://www.wildberries.ru",
        "Referer": "https://www.wildberries.ru/",
    }
    cards_map: Dict[int, Card] = {}
    # v39.11: общий dict статистики для всех запросов — пользователь увидит
    # сколько карточек WB отдал всего и сколько отфильтровал наш домен-фильтр
    collect_stats = {'wb_returned': 0, 'filtered_out': 0}
    sem = asyncio.Semaphore(args.collect_workers)
    async with aiohttp.ClientSession(timeout=timeout, headers=headers) as session:
        # v27.9.x: для бренда листаем до нужного лимита (одна «чистая» выдача),
        # для обычного поиска — 1 страница (охват даёт расширение вариантов).
        _max_pages = max(1, min(60, (int(args.limit) // 90) + 2)) if _brand_search else 1

        # v27.9.x: для бренда находим brandId и тянем ПОЛНЫЙ бренд-каталог
        # (catalog.wb.ru/brands/...), а не обрезанную поисковую выдачу. Плюс
        # перебираем НЕСКОЛЬКО сортировок — это докидывает товары, которые в одной
        # сортировке не попали в первые страницы.
        _fbrand_ids: List[str] = []
        # v27.9.x: СКОРОСТЬ — для бренда хватает 2 сортировок (popular/newly) с
        # листанием; 5 сортировок почти не добавляли уникальных, но кратно
        # замедляли сбор (каждая — отдельная серия запросов). Если brandId найден —
        # бренд-каталог и так отдаёт всё, сортировки тем более не нужны.
        _brand_sorts = ["popular", "newly"]
        if _brand_search:
            try:
                _fbrand_ids = await discover_brand_filter_ids(session, args.brand)
            except Exception as _e:
                _fbrand_ids = []
            print(f"Бренд-каталог: brandId={_fbrand_ids or 'не найден (работаю по поиску)'}; "
                  f"сортировок={len(_brand_sorts)}, страниц до {_max_pages}")

        _brand_base_q = (args.query or args.brand or "").strip()

        async def work(q):
            async with sem:
                # Базовый запрос бренда: ПОЛНЫЙ каталог по brandId + все сортировки +
                # глубокое листание. Остальные варианты («reebok кроссовки» и т.п.):
                # только текстовый поиск, 1 сортировка, немного страниц — добор охвата.
                is_base = (not _brand_search) or (q == _brand_base_q)
                if _brand_search and is_base:
                    _sorts = _brand_sorts
                    _pages = _max_pages
                    _fb = _fbrand_ids
                elif _brand_search:
                    _sorts = ["popular"]
                    _pages = min(_max_pages, 4)
                    _fb = []
                else:
                    _sorts = [s.strip() or "popular" for s in args.search_sorts.split(",")]
                    _pages = _max_pages
                    _fb = []
                for sort in _sorts:
                    for _page in range(1, _pages + 1):
                        cards = await collect_one_query(
                            session, q, args.per_query_limit, sort,
                            domain=domain if use_filter else '', stats=collect_stats,
                            page=_page, fbrand_ids=_fb)
                        before = len(cards_map)
                        for c in cards:
                            # v27.9.x: для бренда ПРЕД-фильтруем по бренду прямо здесь —
                            # иначе варианты «reebok кроссовки» забивают лимит чужими
                            # товарами, а нужные карточки не попадают. Столбец «Запрос»
                            # оставляем = реальному варианту (видно, что вариаций много).
                            if _brand_search:
                                if not brand_matches_v39(getattr(c, 'brand', ''), args.brand,
                                                         getattr(args, 'brand_match', 'contains') or 'contains'):
                                    continue
                            cards_map.setdefault(c.nm_id, c)
                        # страница не дала ни одной НОВОЙ карточки — дальше листать смысла нет
                        if len(cards_map) == before or not cards:
                            break
                        if len(cards_map) >= args.limit:
                            break
                    if len(cards_map) >= args.limit:
                        break
        tasks = [asyncio.create_task(work(q)) for q in variants]
        for t in asyncio.as_completed(tasks):
            try:
                await t
            except Exception:
                pass
            if len(cards_map) >= args.limit:
                for x in tasks:
                    if not x.done():
                        x.cancel()
                break

        # v45.1: ДОБОР по фактическим категориям бренда. Раньше запускался ТОЛЬКО
        # когда категория не выбрана. Но при ВЫБРАННОЙ категории, если brandId не
        # нашёлся, базовый поиск «бренд» часто не листается дальше первой страницы
        # (WB отдаёт те же ~100), а варианты «бренд + тип» из моего списка отсеиваются
        # бренд-фильтром (это товары ЧУЖИХ брендов). Итог — сбор застревал на ~100,
        # хотя пользователь просил 1000. Поэтому добор теперь идёт и при выбранной
        # категории: subject'ы берём из УЖЕ собранных карточек (т.е. они реальные и
        # уже прошли доменный фильтр), а сам добор тоже фильтруется по домену —
        # это докидывает товары бренда той же категории, которых не было на 1-й стр.
        if _brand_search and len(cards_map) < args.limit:
            from collections import Counter as _Counter
            _subj_counter: "_Counter[str]" = _Counter()
            for c in list(cards_map.values()):
                sn = (getattr(c, 'subject', '') or '').strip()
                if sn and not sn.isdigit() and len(sn) <= 60:
                    _subj_counter[sn] += 1
            _top_subjects = [s for s, _ in _subj_counter.most_common(25)]

            async def _expand(subj: str):
                async with sem:
                    qx = f"{args.brand} {subj}".strip()
                    for sort in ("popular", "newly"):
                        for _page in range(1, min(_max_pages, 5) + 1):
                            if len(cards_map) >= args.limit:
                                return
                            cards = await collect_one_query(
                                session, qx, args.per_query_limit, sort,
                                domain=domain if use_filter else '', stats=collect_stats, page=_page)
                            before = len(cards_map)
                            for c in cards:
                                if not brand_matches_v39(getattr(c, 'brand', ''), args.brand,
                                                         getattr(args, 'brand_match', 'contains') or 'contains'):
                                    continue
                                cards_map.setdefault(c.nm_id, c)
                            if len(cards_map) == before or not cards:
                                break

            if _top_subjects:
                print(f"Добор по категориям бренда: {len(_top_subjects)} категорий "
                      f"({', '.join(_top_subjects[:6])}…)")
                _ex_tasks = [asyncio.create_task(_expand(s)) for s in _top_subjects]
                for t in asyncio.as_completed(_ex_tasks):
                    try:
                        await t
                    except Exception:
                        pass
                    if len(cards_map) >= args.limit:
                        for x in _ex_tasks:
                            if not x.done():
                                x.cancel()
                        break

    # v39.11: показать статистику сбора. Особенно важно когда найдено мало карточек
    # — пользователь поймёт причина в WB-выдаче или в нашем фильтре.
    n_kept = len(cards_map)
    n_total = collect_stats.get('wb_returned', 0)
    n_filtered = collect_stats.get('filtered_out', 0)
    print(f"📊 Статистика сбора: WB вернул всего {n_total} карточек (включая дубли), уникальных собрано {n_kept}, отфильтровано {n_filtered}.")
    # v27.9.x: WB вернул 0 на ВСЕ запросы — это не баг кода, а ВРЕМЕННЫЙ лимит
    # запросов WB (часто после серии частых прогонов WB начинает отдавать пусто).
    if n_total == 0:
        print("=" * 80)
        print("🔴 WB вернул 0 карточек на ВСЕ запросы. Это НЕ ошибка программы, а")
        print("   временный анти-бот лимит Wildberries (срабатывает после серии")
        print("   частых прогонов с одного IP). Что делать:")
        print("   • подожди 2–5 минут и запусти снова;")
        print("   • или смени сеть/IP (мобильный интернет, VPN);")
        print("   • проверь, что в обычном браузере открывается www.wildberries.ru.")
        print("=" * 80)
    if use_filter and n_filtered > n_kept * 5 and n_kept < args.limit / 2:
        print(f"   ⚠️  Фильтр отсеял много карточек. Если кажется что отсеиваются нужные товары:")
        print(f"      • попробуй --strict-domain-filter false (выключить фильтр)")
        print(f"      • или дай мне xlsx с примерами — расширю списки для домена '{domain}'")
    final_cards = list(cards_map.values())[:args.limit]

    # v46: ДИАГНОСТИКА для вывода бита «Документы проверены». viewFlags WB — битовая
    # маска бейджей (плашка «Оригинал» = бит 8). Бит «Документы проверены» неизвестен,
    # поэтому выводим сырой viewFlags по собранным карточкам в файл: пользователь
    # открывает несколько товаров, смотрит где есть/нет кнопки «Документы проверены»,
    # и по разнице viewFlags вычисляем бит. Лёгкий CSV, пишется один раз.
    if bool(getattr(args, 'dump_viewflags', True)) and final_cards:
        try:
            import csv as _csv
            with open("wb_viewflags.csv", "w", encoding="utf-8-sig", newline="") as _vf:
                _w = _csv.writer(_vf)
                _w.writerow(["nm_id", "viewFlags", "is_original(bit8)", "name", "product_url"])
                for _c in final_cards:
                    _vfv = int(getattr(_c, "view_flags", 0) or 0)
                    _w.writerow([_c.nm_id, _vfv, "да" if (_vfv & WB_ORIGINAL_VIEWFLAG_BIT) else "",
                                 (_c.product_name or "")[:80], product_url(_c.nm_id)])
            print("📝 [diag] viewFlags карточек сохранён в wb_viewflags.csv — открой 2-3 товара "
                  "С кнопкой «Документы проверены» и 2-3 БЕЗ неё, найди их nm_id в файле и пришли "
                  "мне их viewFlags: по разнице вычислю бит для колонки «Документы проверены».")
        except Exception:
            pass

    # v27.5: проверка плашки «Оригинальный товар» через wb_enhanced.
    # Поля .is_original проставляются внутри enrich_cards_batch в формате bool;
    # в ResultRow это попадает как "оригинал" / "не указано".
    # v27.9.x: СТАРЫЙ HTML-детектор «Оригинал» по умолчанию ОТКЛЮЧЁН — WB рендерит
    # плашку через JS, и в HTML её больше нет (всегда давал «не указано»). Его
    # заменил быстрый card.json-детектор в run_http_link_prefetch. Старый путь
    # оставлен только под явный флаг --check-original-html для отладки.
    if bool(getattr(args, 'check_original_html', False)) and final_cards:
        try:
            from wb_enhanced import WBEnhancedClient as _WBEnh
            _orig_domains = [d.strip() for d in str(getattr(args, 'check_original_domains', 'ru') or 'ru').split(',') if d.strip()]
            print(f"🔍 Проверяю плашку «Оригинальный товар» для {len(final_cards)} карточек через wb_enhanced (домены: {','.join(_orig_domains)})...")
            client = _WBEnh(html_domains=_orig_domains)
            # Сбросим старую "не указано" — иначе enrich_cards_batch пропустит их (truthy-check).
            for _c in final_cards:
                try:
                    _c.is_original = None  # явный флаг «ещё не проверяли»
                except Exception:
                    pass
            await client.enrich_cards_batch(
                final_cards,
                workers=int(getattr(args, 'check_original_workers', 10) or 10),
            )
            # Перевод bool → рус. строку для xlsx.
            orig_count = 0
            for c in final_cards:
                val = getattr(c, 'is_original', None)
                if val is True:
                    c.is_original = 'оригинал'
                    orig_count += 1
                elif val is False:
                    c.is_original = 'не указано'
                # если уже строка ("оригинал"/"не указано") — оставляем как есть
            print(f"   → плашка Оригинал: найдена у {orig_count} из {len(final_cards)} карточек.")
        except Exception as _e:
            print(f"   ⚠️ проверка оригинальности не выполнена: {type(_e).__name__}: {str(_e)[:200]}")
        # Чистка None → "не указано" (на случай сбоя).
        for _c in final_cards:
            v = getattr(_c, 'is_original', None)
            if v is None or v is False:
                try:
                    _c.is_original = 'не указано'
                except Exception:
                    pass
            elif v is True:
                try:
                    _c.is_original = 'оригинал'
                except Exception:
                    pass

    return final_cards


# -----------------------------
# Writer / autosave
# -----------------------------

# -----------------------------
# v25-reporting: отчётный слой для result.xlsx
# -----------------------------

DETAILS_HEADERS_RU_V39: Dict[str, str] = {
    "query": "Запрос",
    "nm_id": "Артикул WB",
    "product_name": "Название товара",
    "brand": "Бренд",
    "subject": "Категория WB",
    "product_url": "Ссылка на товар",
    "status": "Технический статус",
    "price_rub": "Цена, ₽",
    "sale_price_rub": "Цена со скидкой, ₽",
    "seller_name": "Продавец",
    "supplier_id": "ID продавца",
    "rating": "Рейтинг",
    "feedbacks": "Отзывы",
    "is_original": "Плашка 'Оригинал'",
    "docs_verified": "Документ проверен WB",
    "colors": "Цвет",
    "wb_root": "Корневой ID (WB)",
    "registry_url": "Ссылка на реестр",
    "registry_host": "Реестр (хост)",
    "registry_record_id": "ID записи реестра",
    "certificate_number": "Номер документа",
    "document_type": "Тип документа",
    "document_status": "Статус документа",
    "rf_status": "Статус на территории РФ",
    "certificate_product_name": "Название в реестре",
    "document_date_start": "Действует с",
    "document_date_end": "Действует до",
    "applicant_name": "Заявитель",
    "applicant_inn": "ИНН заявителя",
    "manufacturer_name": "Изготовитель",
    "tnved": "ТН ВЭД",
    "scheme": "Схема оценки",
    "technical_regulation": "Техрегламент",
    "score": "Оценка совпадения",
    "details": "Примечания",
    "worker": "Worker",
    "checked_at": "Проверено",
    # вычисляемые в отчёте
    "expiry_days_left": "Дней до окончания",
    "expiry_risk": "Риск по сроку",
}

EXPIRY_RISK_OK = "Действует"
EXPIRY_RISK_SOON = "Скоро истекает"
EXPIRY_RISK_EXPIRED = "Истёк"
EXPIRY_RISK_UNKNOWN = "Срок не известен"

_REPORT_DATE_FORMATS = (
    "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y",
    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
)


def _parse_date_loose_v39(s: Any) -> Optional[dt.date]:
    if s is None:
        return None
    text = str(s).strip()
    if not text:
        return None
    for fmt in _REPORT_DATE_FORMATS:
        try:
            return dt.datetime.strptime(text[:len(fmt) + 10], fmt).date()
        except Exception:
            pass
    try:
        return dt.datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except Exception:
        return None


def _compute_expiry_v39(row: "ResultRow", warning_days: int,
                        today: Optional[dt.date] = None) -> Tuple[Optional[int], str]:
    today = today or dt.date.today()
    end_date = _parse_date_loose_v39(getattr(row, "document_date_end", ""))
    if end_date is None:
        return None, EXPIRY_RISK_UNKNOWN
    days_left = (end_date - today).days
    if days_left < 0:
        return days_left, EXPIRY_RISK_EXPIRED
    if warning_days > 0 and days_left <= warning_days:
        return days_left, EXPIRY_RISK_SOON
    return days_left, EXPIRY_RISK_OK


def _build_summary_sheet_v39(wb_obj, rows: List["ResultRow"], warning_days: int) -> None:
    ws = wb_obj.create_sheet("Сводка", 0)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)

    def _style_table_header(row_idx: int) -> None:
        for col in ("A", "B", "C"):
            c = ws[f"{col}{row_idx}"]
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

    ws.append(["WB Registry Checker — расширенный отчёт"])
    ws.append(["Дата формирования", now_iso()])
    ws.append(["Версия движка", APP_VERSION])
    ws.append(["Порог 'Скоро истекает', дней", warning_days])
    ws.append(["Всего товаров", len(rows)])
    ws.append([])
    total = max(1, len(rows))

    # --- Таблица 1: технический статус ---
    ws.append(["Распределение по техническому статусу"])
    ws.append(["Статус", "Количество", "Доля, %"])
    status_hdr = ws.max_row
    status_counts: Dict[str, int] = {}
    for r in rows:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1
    for status_name, count in sorted(status_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([status_name, count, round(count * 100.0 / total, 1)])
    status_end = ws.max_row
    ws.append([])

    # --- Таблица 2: плашка «Оригинал» ---
    ws.append(["Плашка 'Оригинал'"])
    ws.append(["Значение", "Количество", "Доля, %"])
    orig_hdr = ws.max_row
    original_counts: Dict[str, int] = {}
    for r in rows:
        key = r.is_original or "—"
        original_counts[key] = original_counts.get(key, 0) + 1
    for key, count in sorted(original_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([key, count, round(count * 100.0 / total, 1)])
    orig_end = ws.max_row
    ws.append([])

    # --- Таблица 3: риски по сроку действия ---
    ws.append(["Риски по сроку действия документа"])
    ws.append(["Категория", "Количество", "Доля, %"])
    risk_hdr = ws.max_row
    risk_counts: Dict[str, int] = {EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
                                   EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0}
    for r in rows:
        _, risk = _compute_expiry_v39(r, warning_days)
        risk_counts[risk] = risk_counts.get(risk, 0) + 1
    for key in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
        ws.append([key, risk_counts[key], round(risk_counts[key] * 100.0 / total, 1)])
    risk_end = ws.max_row

    # --- Стили заголовков ---
    ws["A1"].font = Font(bold=True, size=14)
    for row_idx in (2, 3, 4, 5):
        ws[f"A{row_idx}"].font = Font(bold=True)
    _style_table_header(status_hdr)
    _style_table_header(orig_hdr)
    _style_table_header(risk_hdr)
    for sect_row in (status_hdr - 1, orig_hdr - 1, risk_hdr - 1):
        ws[f"A{sect_row}"].font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A7"

    # --- v27.7: нативные графики Excel + условное форматирование ---
    # Раньше отчёт был полностью статичным (ни графиков, ни data-bar).
    try:
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.formatting.rule import DataBarRule

        # Пончик по техническому статусу
        if status_end >= status_hdr + 1:
            pie = PieChart()
            pie.title = "Технический статус"
            pie.height, pie.width = 7.5, 13
            data = Reference(ws, min_col=2, min_row=status_hdr, max_row=status_end)
            cats = Reference(ws, min_col=1, min_row=status_hdr + 1, max_row=status_end)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            ws.add_chart(pie, "E7")

        # Пончик по плашке «Оригинал»
        if orig_end >= orig_hdr + 1:
            pie2 = PieChart()
            pie2.title = "Плашка «Оригинал»"
            pie2.height, pie2.width = 7.5, 13
            data2 = Reference(ws, min_col=2, min_row=orig_hdr, max_row=orig_end)
            cats2 = Reference(ws, min_col=1, min_row=orig_hdr + 1, max_row=orig_end)
            pie2.add_data(data2, titles_from_data=True)
            pie2.set_categories(cats2)
            ws.add_chart(pie2, "E24")

        # Столбчатый график по рискам срока действия
        if risk_end >= risk_hdr + 1:
            bar = BarChart()
            bar.type = "col"
            bar.title = "Риски по сроку действия"
            bar.height, bar.width = 7.5, 13
            bar.legend = None
            data3 = Reference(ws, min_col=2, min_row=risk_hdr, max_row=risk_end)
            cats3 = Reference(ws, min_col=1, min_row=risk_hdr + 1, max_row=risk_end)
            bar.add_data(data3, titles_from_data=True)
            bar.set_categories(cats3)
            ws.add_chart(bar, "E41")

        # Data-bar на колонку «Количество» по всем трём таблицам.
        # v27.9.x: ПРОПУСКАЕМ если таблица пустая (нет строк данных) — иначе
        # диапазон «B9:B8» падал ValueError и рушил сохранение всего файла.
        if status_end >= status_hdr + 1:
            ws.conditional_formatting.add(
                f"B{status_hdr + 1}:B{status_end}",
                DataBarRule(start_type="num", start_value=0, end_type="max",
                            color="4F81BD", showValue=True))
        if orig_end >= orig_hdr + 1:
            ws.conditional_formatting.add(
                f"B{orig_hdr + 1}:B{orig_end}",
                DataBarRule(start_type="num", start_value=0, end_type="max",
                            color="9BBB59", showValue=True))
        if risk_end >= risk_hdr + 1:
            ws.conditional_formatting.add(
                f"B{risk_hdr + 1}:B{risk_end}",
                DataBarRule(start_type="num", start_value=0, end_type="max",
                            color="C0504D", showValue=True))
    except Exception as _e:
        log.warning("Не удалось добавить графики/условное форматирование в Сводку: %s", _e)


def _write_run_log_v39(rows: List["ResultRow"], xlsx_path: Path, warning_days: int,
                       started_at: float, mode: str,
                       log_path: Optional[Path] = None) -> Optional[Path]:
    """v25-reporting: текстовый лог прогона для main_v39 (Stage 1 / Stage 2)."""
    try:
        if log_path is None:
            log_path = xlsx_path.with_name(xlsx_path.stem + "_run.log")
        finished = time.time()
        status_counts: Dict[str, int] = {}
        risk_counts: Dict[str, int] = {EXPIRY_RISK_OK: 0, EXPIRY_RISK_SOON: 0,
                                       EXPIRY_RISK_EXPIRED: 0, EXPIRY_RISK_UNKNOWN: 0}
        for r in rows:
            status_counts[r.status] = status_counts.get(r.status, 0) + 1
            _, risk = _compute_expiry_v39(r, warning_days)
            risk_counts[risk] = risk_counts.get(risk, 0) + 1
        lines: List[str] = []
        lines.append(f"WB Registry Checker — лог прогона ({mode})")
        lines.append(f"Версия движка: {APP_VERSION}")
        lines.append(f"Старт:     {dt.datetime.fromtimestamp(started_at).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Финиш:    {dt.datetime.fromtimestamp(finished).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Длительность: {int(finished - started_at)} сек")
        lines.append(f"Итоговый файл: {xlsx_path}")
        lines.append(f"Порог 'Скоро истекает', дней: {warning_days}")
        lines.append(f"Всего строк: {len(rows)}")
        lines.append("")
        lines.append("Распределение по техническому статусу:")
        for k, v in sorted(status_counts.items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"  {k:<40} {v}")
        lines.append("")
        lines.append("Распределение по риску по сроку:")
        for k in (EXPIRY_RISK_OK, EXPIRY_RISK_SOON, EXPIRY_RISK_EXPIRED, EXPIRY_RISK_UNKNOWN):
            lines.append(f"  {k:<20} {risk_counts.get(k, 0)}")
        log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return log_path
    except Exception as e:
        print(f"[run-log] Не удалось записать лог прогона: {e}")
        return None


# v27.9.x: openpyxl роняет ВЕСЬ файл (IllegalCharacterError) на управляющих
# символах, которые регулярно встречаются в спарсенном тексте (наименования и
# описания из карточек/реестров). Из-за этого отчёт мог сохраняться «криво» или
# не сохраняться целиком. _xlsx_safe гарантирует безопасное значение в ячейке.
_XLSX_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xlsx_safe(value: Any, max_len: int = 32000):
    """Безопасное значение для ячейки xlsx.

    Числа оставляем числами (для сортировки и числового формата Excel), всё
    остальное приводим к строке, вычищаем недопустимые управляющие символы,
    схлопываем переводы строк/табуляции в пробел (иначе колонки визуально
    «разъезжаются») и обрезаем по лимиту ячейки Excel (~32767 символов).
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else ""
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    if not s:
        return ""
    s = _XLSX_ILLEGAL_RE.sub("", s)
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    if "  " in s:
        s = re.sub(r"\s{2,}", " ", s)
    s = s.strip()
    if len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s


# v25-reporting: единый порядок ядра 'Подробностей' в обоих движках.
# Специфичные поля движка пойдут после ядра, expiry-флаги — в самом конце.
CORE_DETAILS_ORDER_V39: Tuple[str, ...] = (
    "query", "nm_id", "product_name", "brand", "subject", "product_url",
    "status", "price_rub", "sale_price_rub", "seller_name", "supplier_id",
    "is_original", "docs_verified", "registry_url", "document_type", "document_status",
    "rf_status", "certificate_number", "document_date_start", "document_date_end",
)


def _details_field_order_v39(base_fields: List[str]) -> List[str]:
    seen: Set[str] = set()
    ordered: List[str] = []
    for f in CORE_DETAILS_ORDER_V39:
        if f in base_fields and f not in seen:
            ordered.append(f); seen.add(f)
    for f in base_fields:
        if f not in seen:
            ordered.append(f); seen.add(f)
    return ordered


def _build_details_sheet_v39(wb_obj, rows: List["ResultRow"], warning_days: int) -> None:
    ws = wb_obj.create_sheet("Подробности")
    base_fields = list(ResultRow.__dataclass_fields__.keys())
    ordered = _details_field_order_v39(base_fields)
    fields = ordered + ["expiry_days_left", "expiry_risk"]
    ws.append([DETAILS_HEADERS_RU_V39.get(f, f) for f in fields])
    for r in rows:
        d = asdict(r)
        days_left, risk = _compute_expiry_v39(r, warning_days)
        d["expiry_days_left"] = "" if days_left is None else days_left
        d["expiry_risk"] = risk
        ws.append([_xlsx_safe(d.get(f, "")) for f in fields])
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions
    for idx, f in enumerate(fields, start=1):
        if f in {"product_name", "registry_url", "certificate_product_name", "details",
                 "product_url", "technical_regulation"}:
            width = 50
        elif f in {"nm_id", "price_rub", "sale_price_rub", "expiry_days_left", "score", "feedbacks"}:
            width = 14
        else:
            width = 22
        ws.column_dimensions[get_column_letter(idx)].width = width
    risk_col_idx = fields.index("expiry_risk") + 1
    fills = {
        EXPIRY_RISK_OK: PatternFill("solid", fgColor="D9EAD3"),
        EXPIRY_RISK_SOON: PatternFill("solid", fgColor="FFF2CC"),
        EXPIRY_RISK_EXPIRED: PatternFill("solid", fgColor="F4CCCC"),
        EXPIRY_RISK_UNKNOWN: PatternFill("solid", fgColor="EFEFEF"),
    }
    for row in ws.iter_rows(min_row=2):
        risk_val = str(row[risk_col_idx - 1].value or "")
        fill = fills.get(risk_val)
        if fill is not None:
            row[risk_col_idx - 1].fill = fill
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)


class ResultStore:
    def __init__(self, xlsx_path: Path, csv_path: Optional[Path] = None,
                 expiry_warning_days: int = 30, make_report_xlsx: bool = True):
        self.xlsx_path = xlsx_path
        self.csv_path = csv_path
        # v25-reporting: параметры отчётного слоя.
        self.expiry_warning_days = max(0, int(expiry_warning_days or 0))
        self.make_report_xlsx = bool(make_report_xlsx)
        self.rows: List[ResultRow] = []
        self.lock = asyncio.Lock()
        self.last_save_count = 0

    def processed_ids_from_csv(self, valid_ids: Optional[Set[int]] = None) -> Tuple[Set[int], int]:
        """
        Загружает resume-результаты только для nm_id текущего задания.
        Это важно для тестов с --limit 5/20: старый CSV не должен превращать прогресс в 14/5
        и не должен попадать в новый XLSX, если карточки не входят в текущую выборку.
        Возвращает (processed_ids_for_current_run, ignored_rows_count).
        """
        ids: Set[int] = set()
        ignored = 0
        if not self.csv_path or not self.csv_path.exists():
            return ids, ignored
        latest: Dict[int, ResultRow] = {}
        try:
            with self.csv_path.open("r", encoding="utf-8-sig", newline="") as f:
                rdr = csv.DictReader(f)
                for row in rdr:
                    nm = safe_int(row.get("nm_id"))
                    if not nm:
                        ignored += 1
                        continue
                    if valid_ids is not None and nm not in valid_ids:
                        ignored += 1
                        continue
                    # В v36 resume пропускает только уже найденные корректные ссылки.
                    # Ошибки, таймауты, НЕТ ДОКУМЕНТОВ и НЕТ ССЫЛКИ перепроверяются новой версией,
                    # чтобы старый плохой результат не «застревал» в CSV навсегда.
                    if (row.get("status") or "") != STATUS_LINK_COLLECTED:
                        ignored += 1
                        continue
                    data = {k: row.get(k, "") for k in ResultRow.__dataclass_fields__.keys()}
                    data["nm_id"] = nm
                    data["score"] = float(data.get("score") or 0)
                    latest[nm] = ResultRow(**data)
        except Exception:
            return ids, ignored
        self.rows = list(latest.values())
        ids = set(latest.keys())
        return ids, ignored

    async def add(self, row: ResultRow):
        async with self.lock:
            self.rows.append(row)

    async def save(self):
        async with self.lock:
            rows = list(self.rows)
        if self.csv_path:
            self._save_csv(rows, self.csv_path)
        self._save_xlsx(rows, self.xlsx_path)
        self.last_save_count = len(rows)

    def _save_csv(self, rows: List[ResultRow], path: Path):
        fields = list(ResultRow.__dataclass_fields__.keys())
        tmp = path.with_suffix(path.suffix + ".tmp")
        with tmp.open("w", encoding="utf-8-sig", newline="") as f:
            wr = csv.DictWriter(f, fieldnames=fields)
            wr.writeheader()
            for r in rows:
                wr.writerow(asdict(r))
        os.replace(tmp, path)

    def _save_xlsx(self, rows: List[ResultRow], path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        fields = list(ResultRow.__dataclass_fields__.keys())
        ws.append(fields)
        for r in rows:
            ws.append([_xlsx_safe(getattr(r, f)) for f in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="D9EAD3")
        # v27.7: ширины колонок вычисляются динамически по ИМЕНИ поля.
        # Раньше был жёсткий словарь A..P под старую схему из 18 колонок, а
        # в ResultRow теперь 33 поля (добавлены price/seller/rating и расширенные
        # поля реестра) — из-за рассинхрона колонки визуально съезжали.
        from openpyxl.utils import get_column_letter as _gcl
        _field_widths = {
            "query": 18, "nm_id": 12, "product_name": 42, "brand": 22, "subject": 20,
            "product_url": 48, "status": 28, "price_rub": 12, "sale_price_rub": 14,
            "seller_name": 26, "supplier_id": 14, "rating": 8, "feedbacks": 10,
            "is_original": 14, "colors": 20, "wb_root": 14,
            "registry_url": 60, "registry_host": 24,
            "registry_record_id": 20, "certificate_number": 30, "document_type": 16,
            "document_status": 16, "certificate_product_name": 55,
            "document_date_start": 16, "document_date_end": 16, "applicant_name": 40,
            "applicant_inn": 16, "manufacturer_name": 40, "tnved": 16, "scheme": 14,
            "technical_regulation": 30, "score": 10, "details": 60, "worker": 14,
            "checked_at": 20,
        }
        for _idx, _fname in enumerate(fields, start=1):
            ws.column_dimensions[_gcl(_idx)].width = _field_widths.get(
                _fname, max(12, min(40, len(_fname) + 4)))
        # review sheet
        review = wb.create_sheet("review")
        review.append(fields)
        # v39: исключаем оба «хороших» статуса: STATUS_LINK_COLLECTED (1 этап) и "OK" (2 этап).
        # До этого OK с 2 этапа ошибочно попадал в review.
        good_statuses = {STATUS_LINK_COLLECTED, "OK"}
        for r in rows:
            if r.status not in good_statuses:
                review.append([_xlsx_safe(getattr(r, f)) for f in fields])
        review.freeze_panes = "A2"
        review.auto_filter.ref = review.dimensions
        for cell in review[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FCE5CD")
        # summary
        summary = wb.create_sheet("summary")
        counts: Dict[str, int] = {}
        for r in rows:
            counts[r.status] = counts.get(r.status, 0) + 1
        summary.append(["status", "count"])
        for k, v in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
            summary.append([k, v])
        # v25-reporting: дополнительные русскоязычные листы «Сводка» и «Подробности».
        # Логику статусов НЕ меняем — только флаг expiry_risk + раскраска.
        if getattr(self, "make_report_xlsx", True):
            try:
                _build_summary_sheet_v39(wb, rows, int(getattr(self, "expiry_warning_days", 30) or 0))
            except Exception as _e:
                log.warning("build_summary_sheet failed: %s", _e)
            try:
                _build_details_sheet_v39(wb, rows, int(getattr(self, "expiry_warning_days", 30) or 0))
            except Exception as _e:
                log.warning("build_details_sheet failed: %s", _e)
        tmp = path.with_suffix(path.suffix + ".tmp")
        wb.save(tmp)
        # v40.2: os.replace падает PermissionError если файл открыт в Excel.
        # Делаем несколько попыток, потом fallback на файл с суффиксом, чтобы
        # данные НЕ терялись и прогон НЕ останавливался.
        import time as _t
        last_err = None
        for attempt in range(4):
            try:
                os.replace(tmp, path)
                return
            except PermissionError as e:
                last_err = e
                _t.sleep(0.6)
        # Все попытки не удались — файл занят. Пишем в запасной файл.
        try:
            alt = path.with_name(path.stem + "_live" + path.suffix)
            os.replace(tmp, alt)
            if not getattr(self, "_warned_locked", False):
                print(f"⚠️  Файл {path.name} открыт (вероятно в Excel) — пишу в {alt.name}. "
                      f"Закрой {path.name} в Excel, чтобы запись шла в него напрямую.")
                self._warned_locked = True
        except Exception:
            # Не удалось даже в запасной — оставляем .tmp, не падаем
            if not getattr(self, "_warned_locked", False):
                print(f"⚠️  Не могу записать {path.name} (открыт?). Данные во временном файле {tmp.name}.")
                self._warned_locked = True


# -----------------------------
# Playwright strict flow
# -----------------------------

CAPTURE_INIT_SCRIPT = r"""
(() => {
  try {
    if (!window.__wbCertCaptureInstalled) {
      window.__wbCertCaptureInstalled = true;
      window.__wbCertCaptured = [];
      window.__wbCertPush = (kind, url) => {
        try {
          if (!url) return;
          window.__wbCertCaptured.push({kind: kind || 'unknown', url: String(url), ts: Date.now()});
          if (window.__wbCertCaptured.length > 500) window.__wbCertCaptured.shift();
        } catch(e) {}
      };
      // WB/внешние страницы иногда открывают alert/confirm/beforeunload.
      // В long-run это могло уронить Playwright driver (ProtocolError Page.handleJavaScriptDialog).
      try {
        window.alert = function() {};
        window.confirm = function() { return true; };
        window.prompt = function() { return null; };
        window.onbeforeunload = null;
        Object.defineProperty(window, 'onbeforeunload', { configurable: true, get: () => null, set: () => {} });
      } catch(e) {}
      const oldOpen = window.open;
      window.open = function(url, name, specs) {
        try { window.__wbCertPush('window.open', url); } catch(e) {}
        return oldOpen.apply(window, arguments);
      };
      const oldAssign = window.location.assign.bind(window.location);
      window.location.assign = function(url) {
        try { window.__wbCertPush('location.assign', url); } catch(e) {}
        return oldAssign(url);
      };
      const oldFetch = window.fetch;
      window.fetch = function() {
        try {
          const u = arguments && arguments[0] && (arguments[0].url || arguments[0]);
          window.__wbCertPush('fetch', u);
        } catch(e) {}
        return oldFetch.apply(this, arguments);
      };
      const oldXHROpen = XMLHttpRequest.prototype.open;
      XMLHttpRequest.prototype.open = function(method, url) {
        try { this.__wbCertUrl = url; window.__wbCertPush('xhr.open', url); } catch(e) {}
        return oldXHROpen.apply(this, arguments);
      };
      const oldXHRSend = XMLHttpRequest.prototype.send;
      XMLHttpRequest.prototype.send = function() {
        try { window.__wbCertPush('xhr.send', this.__wbCertUrl); } catch(e) {}
        return oldXHRSend.apply(this, arguments);
      };
    }
  } catch(e) {}
})();
"""

FIND_TEXT_TARGET_JS = r"""
({wanted, exact, clickableOnly}) => {
  const norm = s => (s || '').toString().toLowerCase().replace(/ё/g, 'е').replace(/\\s+/g, ' ').trim();
  const wantedNorm = norm(wanted);
  const vw = window.innerWidth || 1920;
  const vh = window.innerHeight || 1080;
  const isVisible = (el) => {
    const st = window.getComputedStyle(el);
    if (!st || st.display === 'none' || st.visibility === 'hidden' || parseFloat(st.opacity || '1') < 0.05) return false;
    const r = el.getBoundingClientRect();
    return r.width > 2 && r.height > 2 && r.bottom > 0 && r.right > 0 && r.left < vw && r.top < vh;
  };
  const clickScore = (el) => {
    const tag = el.tagName || '';
    const role = el.getAttribute('role') || '';
    const cls = el.className || '';
    const st = window.getComputedStyle(el);
    let s = 0;
    if (['BUTTON','A'].includes(tag)) s += 8000;
    if (/button|link/i.test(role)) s += 6000;
    if (el.onclick) s += 3000;
    if (st.cursor === 'pointer') s += 2500;
    if (/btn|button|link|chip|badge|documents|certificate|modal|popup/i.test(String(cls))) s += 1200;
    return s;
  };
  const hasText = (txt) => exact ? norm(txt) === wantedNorm : norm(txt).includes(wantedNorm);
  const all = Array.from(document.querySelectorAll('body *'));
  let candidates = [];
  for (const el of all) {
    if (!isVisible(el)) continue;
    const own = norm(el.innerText || el.textContent || '');
    if (!hasText(own)) continue;
    let target = el;
    // Если нашли span, поднимаемся до ближайшего кликабельного родителя, но не выше 5 уровней.
    let p = el;
    for (let depth=0; depth<5 && p; depth++, p=p.parentElement) {
      if (!isVisible(p)) continue;
      if (clickScore(p) > clickScore(target)) target = p;
    }
    if (clickableOnly && clickScore(target) <= 0) {
      // всё равно оставим, если это текстовая плашка "Документы проверены": WB часто кликает по контейнеру
      if (!wantedNorm.includes('документы')) continue;
    }
    const r = target.getBoundingClientRect();
    const txt = norm(target.innerText || target.textContent || '');
    const area = r.width * r.height;
    let score = clickScore(target);
    if (norm(target.innerText || target.textContent || '') === wantedNorm) score += 12000;
    if (txt.includes(wantedNorm)) score += 4000;
    score -= Math.min(area / 20, 20000);
    if (txt.length > wantedNorm.length + 80) score -= 4000;
    candidates.push({
      tag: target.tagName,
      text: txt.slice(0, 180),
      score,
      x: Math.max(1, Math.min(vw - 1, r.left + r.width/2)),
      y: Math.max(1, Math.min(vh - 1, r.top + r.height/2)),
      w: r.width,
      h: r.height
    });
  }
  candidates.sort((a,b) => b.score - a.score);
  return candidates.slice(0, 12);
}
"""

PANEL_TEXT_JS = r"""
() => {
  const txt = (document.body && document.body.innerText || '').toLowerCase().replace(/ё/g,'е');
  return {
    hasSpecs: txt.includes('характеристики и описание') || txt.includes('основная информация'),
    hasDocs: txt.includes('документы проверены'),
    hasLook: txt.includes('смотреть на сайте'),
    textStart: txt.slice(0, 5000)
  };
}
"""

GET_CAPTURED_JS = r"""
() => {
  try { return (window.__wbCertCaptured || []).slice(); } catch(e) { return []; }
}
"""

CLEAR_CAPTURED_JS = r"""
() => {
  try { window.__wbCertCaptured = []; } catch(e) {}
}
"""

async def block_assets(route):
    try:
        req = route.request
        if req.resource_type in {"image", "media", "font"}:
            return await route.abort()
        # режем явную аналитику, но JS/CSS оставляем, иначе WB не рисует панель.
        url = req.url.lower()
        if any(x in url for x in ("google-analytics", "yandex/metrika", "doubleclick", "adsystem", "adservice")):
            return await route.abort()
        return await route.continue_()
    except Exception:
        try:
            await route.continue_()
        except Exception:
            pass

class PageCapture:
    """v39: handlers навешиваются ОДИН раз на page (через attach_once), а потом
    между карточками просто чистим буфер. Это убирает основную утечку, из-за
    которой после 1500 карточек на странице висели тысячи обработчиков.

    Использование:
        cap = PageCapture.attach_once(page)  # один раз для жизни page
        cap.clear_buffer()                    # перед каждой карточкой
        urls = await cap.snapshot_urls()      # после клика «Смотреть на сайте»
    """

    _ATTR = "_wb_cert_capture"

    def __init__(self, page):
        self.page = page
        self.urls: List[str] = []
        self._request_handler = None
        self._response_handler = None
        # ограничение размера буфера — на случай если страница долго висит
        self._max_buffer = 4000

    @classmethod
    def attach_once(cls, page) -> "PageCapture":
        existing = getattr(page, cls._ATTR, None)
        if existing is not None:
            return existing
        cap = cls(page)
        cap._attach()
        try:
            setattr(page, cls._ATTR, cap)
        except Exception:
            pass
        return cap

    def _attach(self):
        def on_request(req):
            try:
                if len(self.urls) < self._max_buffer:
                    self.urls.append(req.url)
            except Exception:
                pass

        def on_response(resp):
            try:
                if len(self.urls) < self._max_buffer:
                    self.urls.append(resp.url)
            except Exception:
                pass

        self._request_handler = on_request
        self._response_handler = on_response
        try:
            self.page.on("request", on_request)
            self.page.on("response", on_response)
        except Exception:
            pass

    def clear_buffer(self):
        """Быстрая синхронная очистка буфера URL. Вызывается между карточками."""
        self.urls.clear()

    async def clear(self):
        """Полная очистка: буфер + JS-storage. Вызывать в ключевых точках flow."""
        self.urls.clear()
        try:
            await self.page.evaluate(CLEAR_CAPTURED_JS)
        except Exception:
            pass

    async def snapshot_urls(self) -> List[str]:
        out = list(self.urls)
        try:
            captured = await self.page.evaluate(GET_CAPTURED_JS)
            for item in captured or []:
                if isinstance(item, dict) and item.get("url"):
                    out.append(str(item["url"]))
        except Exception:
            pass
        try:
            out.append(self.page.url)
        except Exception:
            pass
        seen = set(); res = []
        for u in out:
            u = clean_url(u)
            if u and u not in seen:
                seen.add(u); res.append(u)
        return res

    # Совместимость со старым кодом, который вызывал .start() / .stop()
    async def start(self):
        return

    async def stop(self):
        return

async def click_text_by_js(page, wanted: str, exact: bool = False, clickable_only: bool = False, trace: bool = False) -> bool:
    candidates = await page.evaluate(FIND_TEXT_TARGET_JS, {
        "wanted": wanted,
        "exact": exact,
        "clickableOnly": clickable_only
    })
    if trace:
        print(f"      candidates for '{wanted}': {candidates[:5]}")
    if not candidates:
        return False
    c = candidates[0]
    try:
        await page.mouse.move(c["x"], c["y"])
        await page.mouse.down()
        await page.mouse.up()
        if trace:
            print(f"    ✓ клик: {wanted}: {c.get('text')} / tag={c.get('tag')} / x={int(c['x'])},y={int(c['y'])}")
        return True
    except Exception:
        return False

async def wait_until_body_has(page, predicate, timeout_ms: int, poll_ms: int = 120) -> bool:
    start = time.monotonic()
    while (time.monotonic() - start) * 1000 < timeout_ms:
        try:
            state = await page.evaluate(PANEL_TEXT_JS)
            if predicate(state):
                return True
        except Exception:
            pass
        await asyncio.sleep(poll_ms / 1000)
    return False

async def strict_get_registry_url(page, context, card: Card, args, worker_name: str) -> Tuple[str, str]:
    """
    Возвращает (registry_url, details/status-detail).
    Строгий путь: карточка -> характеристики -> документы проверены -> смотреть на сайте.

    v39: PageCapture навешивается ОДИН раз на page (в make_page через attach_once),
    здесь только чистим буфер. Это устраняет утечку handlers, из-за которой
    программа замирала после ~1500 карточек.
    """
    cap = PageCapture.attach_once(page)
    cap.clear_buffer()
    # Чистим JS-storage перед карточкой
    try:
        await page.evaluate(CLEAR_CAPTURED_JS)
    except Exception:
        pass

    try:
        await page.goto(card.product_url, wait_until="domcontentloaded", timeout=args.goto_timeout_ms)
    except Exception as e:
        return "", f"goto_failed: {type(e).__name__}: {str(e)[:180]}"

    # Небольшое ожидание ключевого DOM. Не ждём полной загрузки.
    await asyncio.sleep(args.after_goto_ms / 1000)

    # 1. Характеристики и описание
    # Иногда характеристики уже видны на странице, но клик по кнопке нужен для боковой панели.
    specs_clicked = await click_text_by_js(page, "Характеристики и описание", exact=False, clickable_only=False, trace=args.trace)
    # v37: убран Playwright locator-fallback с коротким timeout.
    # На долгих запусках он давал "Future exception was never retrieved" при закрытии страницы.
    if specs_clicked:
        await asyncio.sleep(args.after_specs_click_ms / 1000)

    # Ждём боковую панель/характеристики.
    await wait_until_body_has(page, lambda s: s.get("hasSpecs") or s.get("hasDocs"), args.card_ready_timeout_ms)

    # 2. Документы проверены — ищем как плашку/чип. Это главный фикс по вашему скрину.
    docs_found = await wait_until_body_has(page, lambda s: s.get("hasDocs"), args.docs_timeout_ms)
    if not docs_found:
        # fallback: ещё раз открыть характеристики и подождать.
        if args.no_docs_fallback_ms > 0:
            await click_text_by_js(page, "Характеристики и описание", exact=False, clickable_only=False, trace=args.trace)
            await asyncio.sleep(args.after_specs_click_ms / 1000)
            docs_found = await wait_until_body_has(page, lambda s: s.get("hasDocs"), args.no_docs_fallback_ms)
    if not docs_found:
        return "", "NO_DOCS"

    cap.clear_buffer()
    try:
        await page.evaluate(CLEAR_CAPTURED_JS)
    except Exception:
        pass
    docs_clicked = await click_text_by_js(page, "Документы проверены", exact=False, clickable_only=False, trace=args.trace)
    if not docs_clicked:
        return "", "docs_badge_found_but_not_clicked"

    # После клика по плашке ждём появление модалки с Смотреть на сайте.
    await asyncio.sleep(args.after_docs_wait_ms / 1000)
    look_found = await wait_until_body_has(page, lambda s: s.get("hasLook"), args.look_button_timeout_ms)
    if not look_found:
        # Если после клика по "Документы проверены" сразу появился URL — не принимаем.
        # По вашей логике переход только после "Смотреть на сайте".
        return "", "docs_modal_opened_but_no_look_button"

    # 3. Смотреть на сайте — принимаем URL только после этого клика.
    cap.clear_buffer()
    try:
        await page.evaluate(CLEAR_CAPTURED_JS)
    except Exception:
        pass
    before_pages = set(context.pages)
    look_clicked = await click_text_by_js(page, "Смотреть на сайте", exact=False, clickable_only=False, trace=args.trace)
    if not look_clicked:
        return "", "look_button_found_but_not_clicked"

    # Ждём URL через сеть/window.open/new page/current page.
    deadline = time.monotonic() + args.after_look_wait_ms / 1000
    found_url = ""

    while time.monotonic() < deadline:
        # новые вкладки
        try:
            new_pages = [p for p in context.pages if p not in before_pages]
            for p in new_pages:
                try:
                    if p.url and p.url != "about:blank":
                        found_url = first_allowed_url([p.url])
                        if found_url:
                            break
                    await p.wait_for_load_state("domcontentloaded", timeout=400)
                    found_url = first_allowed_url([p.url])
                    if found_url:
                        break
                except Exception:
                    pass
            if found_url:
                break
        except Exception:
            pass

        urls = await cap.snapshot_urls()
        found_url = first_allowed_url(urls)
        if found_url:
            break
        await asyncio.sleep(0.08)

    # закрыть лишние вкладки, если открылись
    try:
        for p in list(context.pages):
            if p is not page and p not in before_pages:
                try:
                    await p.close()
                except Exception:
                    pass
    except Exception:
        pass

    if found_url:
        return found_url, "strict_flow_ok"
    return "", "look_clicked_but_registry_url_not_captured"

async def process_card_with_timeout(page, context, card: Card, args, worker_name: str) -> ResultRow:
    async def _inner():
        url, detail = await strict_get_registry_url(page, context, card, args, worker_name)
        if detail == "NO_DOCS":
            status = STATUS_NO_DOCS
        elif url:
            status = STATUS_LINK_COLLECTED
        else:
            status = STATUS_NO_REGISTRY_LINK
        return ResultRow(
            query=card.source_query or args.query,
            nm_id=card.nm_id,
            **_card_fields_for_result(card),
            status=status,
            registry_url=url,
            registry_host=hostname(url) if url else "",
            registry_record_id=extract_record_id(url) if url else "",
            details=detail,
            worker=worker_name,
            checked_at=now_iso(),
        )
    try:
        return await asyncio.wait_for(_inner(), timeout=args.card_hard_timeout_ms / 1000)
    except asyncio.TimeoutError:
        return ResultRow(
            query=card.source_query or args.query,
            nm_id=card.nm_id,
            **_card_fields_for_result(card),
            status=STATUS_TIMEOUT,
            details=f"hard_timeout_{args.card_hard_timeout_ms}ms",
            worker=worker_name,
            checked_at=now_iso(),
        )
    except Exception as e:
        return ResultRow(
            query=card.source_query or args.query,
            nm_id=card.nm_id,
            **_card_fields_for_result(card),
            status=STATUS_ERROR,
            details=f"{type(e).__name__}: {str(e)[:250]}",
            worker=worker_name,
            checked_at=now_iso(),
        )


async def launch_browser(p, args):
    return await p.chromium.launch(
        headless=args.headless,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-background-timer-throttling",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-extensions",
        ],
        timeout=60000,
    )

class BrowserPool:
    """Реально использует --browser-count: N процессов Chromium и много контекстов внутри них."""
    def __init__(self, p, args):
        self.p = p
        self.args = args
        self.count = max(1, int(args.browser_count or 1))
        self.browsers = [None] * self.count
        self.locks = [asyncio.Lock() for _ in range(self.count)]

    async def get_browser(self, idx: int):
        idx = idx % self.count
        async with self.locks[idx]:
            br = self.browsers[idx]
            try:
                if br is not None and br.is_connected():
                    return br
            except Exception:
                pass
            try:
                if br is not None:
                    await br.close()
            except Exception:
                pass
            br = await launch_browser(self.p, self.args)
            self.browsers[idx] = br
            return br

    async def new_context(self, idx: int):
        br = await self.get_browser(idx)
        try:
            ctx = await br.new_context(
                user_agent=self.args.user_agent,
                viewport={"width": self.args.viewport_width, "height": self.args.viewport_height},
                locale="ru-RU",
                ignore_https_errors=True,
            )
        except Exception:
            # Если процесс браузера умер, перезапускаем только его слот.
            async with self.locks[idx % self.count]:
                try:
                    old = self.browsers[idx % self.count]
                    if old:
                        await old.close()
                except Exception:
                    pass
                self.browsers[idx % self.count] = await launch_browser(self.p, self.args)
                br = self.browsers[idx % self.count]
            ctx = await br.new_context(
                user_agent=self.args.user_agent,
                viewport={"width": self.args.viewport_width, "height": self.args.viewport_height},
                locale="ru-RU",
                ignore_https_errors=True,
            )
        if self.args.block_assets:
            await ctx.route("**/*", block_assets)
        try:
            await ctx.add_init_script(CAPTURE_INIT_SCRIPT)
        except Exception:
            pass
        return ctx

    async def close_all(self):
        for br in list(self.browsers):
            try:
                if br:
                    await br.close()
            except Exception:
                pass

    async def restart_browser(self, idx: int) -> None:
        """v39: полностью убить процесс Chromium в слоте idx и пересоздать его.
        Лечит утечку памяти на длинных прогонах (после ~1500 карточек)."""
        idx = idx % self.count
        async with self.locks[idx]:
            old = self.browsers[idx]
            try:
                if old is not None:
                    await old.close()
            except Exception:
                pass
            self.browsers[idx] = None
            # Следующий get_browser/new_context создаст новый процесс.

async def make_page(pool: BrowserPool, pool_idx: int, args):
    ctx = await pool.new_context(pool_idx)
    page = await ctx.new_page()
    # Безопасно гасим любые JS-диалоги, чтобы они не блокировали страницу и не роняли driver.
    async def _safe_dialog_close(dialog):
        try:
            await dialog.dismiss()
        except Exception:
            pass
    try:
        page.on("dialog", lambda d: asyncio.create_task(_safe_dialog_close(d)))
    except Exception:
        pass
    page.set_default_timeout(args.default_timeout_ms)
    # v39: CAPTURE_INIT_SCRIPT уже навешан на context в BrowserPool.new_context.
    # Повторный page.add_init_script тут добавлял дубликат скрипта на каждом
    # пересоздании контекста — убрано.
    # Сразу навешиваем capture handlers ОДИН РАЗ на эту страницу.
    PageCapture.attach_once(page)
    return ctx, page

async def close_context_quiet(context):
    try:
        if context:
            await context.close()
    except Exception:
        pass

async def link_worker(worker_id: int, pool: BrowserPool, queue: asyncio.Queue, store: ResultStore, args, progress: Dict[str, Any]):
    pool_idx = (worker_id - 1) % max(1, int(args.browser_count or 1))
    worker_name = f"w{worker_id}"
    context = None
    page = None
    # v39: локальные счётчики этого воркера. Пересоздание контекста/браузера
    # теперь привязано к НИМ, а не к глобальному progress["done"], потому что
    # глобальный имеет race condition и срабатывал у случайного воркера.
    local_done = 0
    cards_since_browser_restart = 0
    ctx_refresh_every = max(20, int(args.context_refresh_every or 80))
    browser_restart_every = max(0, int(getattr(args, "browser_restart_every", 0) or 0))
    try:
        context, page = await make_page(pool, pool_idx, args)
        while True:
            try:
                card = await queue.get()
            except asyncio.CancelledError:
                break
            progress["active"][worker_name] = {"nm_id": card.nm_id, "started": time.time(), "attempt": 0}
            if args.trace:
                print(f"[{worker_name}] карточка nm_id={card.nm_id}: открываю {card.product_url or product_url(card.nm_id)}")
            final_row = None
            try:
                attempt = 0
                while True:
                    progress["active"][worker_name] = {"nm_id": card.nm_id, "started": time.time(), "attempt": attempt + 1}
                    row = await process_card_with_timeout(page, context, card, args, worker_name)
                    retryable_tech = row.status in {STATUS_TIMEOUT, STATUS_ERROR}
                    retryable_missing = row.status in {STATUS_NO_DOCS, STATUS_NO_REGISTRY_LINK}
                    max_retry = args.max_card_retries if retryable_tech else args.retry_missing
                    if (retryable_tech or retryable_missing) and attempt < max_retry:
                        attempt += 1
                        if args.verbose_each or args.trace:
                            print(f"[{worker_name}] nm_id={card.nm_id}: повтор {attempt}/{max_retry} после {row.status}: {row.details[:120]}")
                        # После любой неудачи пересоздаём контекст: это очищает модалки, crash-состояние и память страницы.
                        await close_context_quiet(context)
                        context, page = await make_page(pool, pool_idx, args)
                        await asyncio.sleep(min(0.8, 0.2 * attempt))
                        continue
                    final_row = row
                    break

                row = final_row
                await store.add(row)
                progress["done"] += 1
                local_done += 1
                cards_since_browser_restart += 1
                if row.status == STATUS_LINK_COLLECTED:
                    progress["links"] += 1
                elif row.status == STATUS_NO_DOCS:
                    progress["no_docs"] += 1
                elif row.status == STATUS_NO_REGISTRY_LINK:
                    progress["no_link"] += 1
                elif row.status in {STATUS_TIMEOUT, STATUS_ERROR}:
                    progress["tech"] += 1

                if args.verbose_each or (progress["done"] <= 10) or (row.status == STATUS_LINK_COLLECTED and args.print_links):
                    print(f"[{worker_name}] #{progress['done']} nm_id={card.nm_id}: {row.status} — {row.details[:160]}")

                # v39: ПОЛНЫЙ рестарт браузера каждые browser_restart_every карточек —
                # лечит утечку памяти Chromium. context_refresh не помогает, потому что
                # процесс Chromium живёт и копит память между context.close().
                if browser_restart_every > 0 and cards_since_browser_restart >= browser_restart_every:
                    print(f"[{worker_name}] плановый рестарт браузера (обработано {cards_since_browser_restart} карточек этим воркером)")
                    await close_context_quiet(context)
                    await pool.restart_browser(pool_idx)
                    context, page = await make_page(pool, pool_idx, args)
                    cards_since_browser_restart = 0
                # Обычный refresh контекста — дешевле, чищет страницу/модалки.
                elif local_done % ctx_refresh_every == 0:
                    await close_context_quiet(context)
                    context, page = await make_page(pool, pool_idx, args)

                if args.autosave_every > 0 and progress["done"] % args.autosave_every == 0:
                    await store.save()
            except Exception as e:
                progress["tech"] += 1
                err_row = ResultRow(
                    query=card.source_query or args.query,
                    nm_id=card.nm_id,
                    product_name=card.product_name,
                    brand=card.brand,
                    subject=card.subject,
                    product_url=card.product_url or product_url(card.nm_id),
                    status=STATUS_ERROR,
                    details=f"worker_exception: {type(e).__name__}: {str(e)[:250]}",
                    worker=worker_name,
                    checked_at=now_iso(),
                )
                await store.add(err_row)
                progress["done"] += 1
                local_done += 1
                cards_since_browser_restart += 1
                if args.verbose_each or progress["done"] <= 10:
                    print(f"[{worker_name}] #{progress['done']} nm_id={card.nm_id}: ОШИБКА — {err_row.details[:160]}")
                await close_context_quiet(context)
                context, page = await make_page(pool, pool_idx, args)
            finally:
                progress["active"].pop(worker_name, None)
                queue.task_done()
    finally:
        await close_context_quiet(context)



def _queue_ids(queue: asyncio.Queue) -> Set[int]:
    """Лучший-effort доступ к ids, которые ещё лежат в asyncio.Queue."""
    try:
        return {getattr(c, "nm_id", 0) for c in list(queue._queue)}  # type: ignore[attr-defined]
    except Exception:
        return set()

async def recover_missing_cards(queue: asyncio.Queue, store: ResultStore, all_cards: Dict[int, Card], progress: Dict[str, Any], reason: str) -> int:
    """
    Если worker/driver умер между queue.get() и записью результата, карточка могла выпасть из очереди.
    Эта функция находит такие nm_id и возвращает их обратно в очередь.
    """
    async with store.lock:
        done_ids = {r.nm_id for r in store.rows}
    queued = _queue_ids(queue)
    active = {safe_int(v.get("nm_id")) for v in progress.get("active", {}).values() if isinstance(v, dict)}
    missing = [nm for nm in all_cards.keys() if nm not in done_ids and nm not in queued and nm not in active]
    for nm in missing:
        await queue.put(all_cards[nm])
    if missing:
        print(f"⚠️  Восстановлены потерянные карточки после {reason}: {len(missing)} шт. Первые: {missing[:8]}")
    return len(missing)

async def supervisor_loop(worker_tasks: Dict[int, asyncio.Task], pool: BrowserPool, queue: asyncio.Queue, store: ResultStore, args, progress: Dict[str, Any], all_cards: Dict[int, Card], total: int):
    """Перезапускает умершие worker-задачи и возвращает потерянные карточки в очередь."""
    while progress.get("done", 0) < total:
        await asyncio.sleep(2.0)
        # Если все worker-задачи умерли, а очередь ещё не пуста — это не должно останавливать прогон.
        for wid, task in list(worker_tasks.items()):
            if task.done():
                try:
                    exc = task.exception()
                    if exc:
                        print(f"⚠️  worker w{wid} упал: {type(exc).__name__}: {str(exc)[:180]}")
                except asyncio.CancelledError:
                    continue
                except Exception:
                    pass
                await recover_missing_cards(queue, store, all_cards, progress, f"падения worker w{wid}")
                if progress.get("done", 0) < total:
                    worker_tasks[wid] = asyncio.create_task(link_worker(wid, pool, queue, store, args, progress))
                    print(f"↻ worker w{wid} перезапущен")
        if queue.empty() and not progress.get("active") and progress.get("done", 0) < total:
            added = await recover_missing_cards(queue, store, all_cards, progress, "пустой очереди при незавершённом прогоне")
            if added == 0:
                # Не даём бесконечно висеть: фиксируем недостающие карточки как тех.ошибки.
                async with store.lock:
                    done_ids = {r.nm_id for r in store.rows}
                missing = [nm for nm in all_cards.keys() if nm not in done_ids]
                for nm in missing:
                    c = all_cards[nm]
                    await store.add(ResultRow(
                        query=c.source_query or args.query,
                        nm_id=c.nm_id,
                        product_name=c.product_name,
                        brand=c.brand,
                        subject=c.subject,
                        product_url=c.product_url or product_url(c.nm_id),
                        status=STATUS_ERROR,
                        details="supervisor_marked_missing_after_worker_crash",
                        checked_at=now_iso(),
                    ))
                    progress["done"] += 1
                    progress["tech"] += 1
                if missing:
                    print(f"⚠️  {len(missing)} карточек помечены как ОШИБКА, чтобы прогон не завис.")
                break
async def progress_loop(queue: asyncio.Queue, store: ResultStore, args, progress: Dict[str, Any], total: int):
    start = progress["start_time"]
    last_done = progress.get("done", 0)
    last_change = time.time()
    try:
        while progress["done"] < total:
            await asyncio.sleep(args.progress_interval_sec)
            done = progress["done"]
            if done != last_done:
                last_done = done
                last_change = time.time()
            elapsed_min = max((time.time() - start) / 60, 0.001)
            speed = done / elapsed_min
            link_speed = progress["links"] / elapsed_min
            active_desc = []
            now = time.time()
            for w, info in list(progress["active"].items())[:8]:
                age = int(now - info.get("started", now))
                active_desc.append(f"{w}:{info.get('nm_id')}:{age}s/a{info.get('attempt',1)}")
            # v39: если done не меняется — явно показать сколько секунд уже застой,
            # чтобы было видно тенденцию к зависанию ДО того как сработает watchdog
            stall_age = int(now - last_change)
            stall_str = f", БЕЗ ИЗМЕНЕНИЙ {stall_age}с" if stall_age >= args.progress_interval_sec * 2 else ""
            print(
                f"Прогресс: обработано={done}/{total}, скорость≈{speed:.1f} карточек/мин, "
                f"ссылки≈{link_speed:.1f}/мин, найдено ссылок={progress['links']}, "
                f"нет документов={progress['no_docs']}, нет ссылки={progress['no_link']}, "
                f"очередь={queue.qsize()}, тех={progress['tech']}, активные=[{'; '.join(active_desc)}]{stall_str}"
            )
            emit_progress("links", done, total)
            if now - last_change > args.stuck_report_sec:
                stuck = []
                for w, info in list(progress["active"].items()):
                    age = now - info["started"]
                    if age > args.stuck_report_sec:
                        stuck.append(f"{w}: nm_id={info['nm_id']} висит {int(age)}с попытка={info.get('attempt',1)}")
                if stuck:
                    print("⚠️  Зависшие воркеры: " + "; ".join(stuck[:8]))
                if args.stall_autosave_sec and now - last_change > args.stall_autosave_sec:
                    await store.save()
                    last_change = now
    except asyncio.CancelledError:
        return

async def run_link_collection(args):
    if async_playwright is None:
        raise RuntimeError("Playwright не установлен. Выполните: python -m pip install playwright && python -m playwright install chromium")

    _run_started_at = time.time()
    cards = await collect_cards(args)
    # v25-reporting: строгий бренд-фильтр (опциональный). Для режима 'any' функция всегда True — без эффекта.
    _brand_wanted = getattr(args, "brand", "") or ""
    _brand_mode = getattr(args, "brand_match", "any") or "any"
    if _brand_wanted and _brand_mode != "any":
        before = len(cards)
        cards = [c for c in cards if brand_matches_v39(getattr(c, "brand", ""), _brand_wanted, _brand_mode)]
        print(f"Бренд-фильтр ({_brand_mode}, '{_brand_wanted}'): отфильтровано {before - len(cards)} карточек, осталось {len(cards)}")
    print(f"К проверке подготовлено карточек: {len(cards)}")
    if cards[:10]:
        print("Первые nm_id:", ", ".join(str(c.nm_id) for c in cards[:10]))

    store = ResultStore(
        Path(args.output),
        Path(args.output_links_csv) if args.output_links_csv else None,
        expiry_warning_days=getattr(args, "expiry_warning_days", 30),
        make_report_xlsx=getattr(args, "make_report_xlsx", True),
    )
    valid_ids = {c.nm_id for c in cards}
    processed: Set[int] = set()
    ignored_resume_rows = 0
    if args.reset_output:
        for path in (store.csv_path, store.xlsx_path):
            try:
                if path and path.exists():
                    path.unlink()
            except Exception:
                pass
    if args.resume and store.csv_path and store.csv_path.exists():
        processed, ignored_resume_rows = store.processed_ids_from_csv(valid_ids=valid_ids)
        if processed or ignored_resume_rows:
            print(
                f"Resume: принято для текущей выборки={len(processed)}, "
                f"проигнорировано старых строк вне текущей выборки={ignored_resume_rows}"
            )

    remaining = [c for c in cards if c.nm_id not in processed]

    # v40.3: дотягиваем имена продавцов (WB поиск отдаёт только supplierId).
    # Делаем для всех карточек текущей выборки (не только remaining), чтобы seller_name
    # попал и в строки которые запишет prefetch.
    if getattr(args, 'fetch_sellers', True) and remaining:
        try:
            await enrich_sellers_batch(remaining, args)
        except Exception as e:
            print(f"⚠️  Обогащение продавцами не удалось (не критично): {type(e).__name__}: {e}")

    # v46: бейдж «Документ проверен WB» — отдельным проходом для ВСЕХ карточек
    # текущей выборки, чтобы значение попало в строку независимо от пути сбора
    # (HTTP fast-path или браузер). Раньше браузерные карточки оставались пустыми.
    if getattr(args, 'check_docs_verified', True) and remaining:
        try:
            await enrich_docs_verified_batch(remaining, args)
        except Exception as e:
            print(f"⚠️  Определение «Документ проверен WB» не удалось (не критично): {type(e).__name__}: {e}")

    # =============================================================================
    # v40: HTTP сбор ссылок через certificate.json. БЕЗ браузера на 1 этапе.
    # Наличие документа однозначно определяется по json: 200=есть, 404 везде=нет.
    # =============================================================================
    link_mode = str(getattr(args, 'link_mode', 'http_only') or 'http_only').lower()
    # v40.1: http_only ТОЖЕ запускает prefetch (это его основной режим!).
    # Раньше http_only отсутствовал в этом условии — из-за чего prefetch не запускался
    # и ВСЕ карточки ошибочно помечались «НЕТ ССЫЛКИ». Критический баг, исправлен.
    if link_mode in ('http_first', 'http_only', 'http', 'auto') and remaining:
        try:
            http_done = await run_http_link_prefetch(remaining, args, store, processed)
            if http_done:
                processed = processed | http_done
                # Перебираем remaining заново — те кого нет в processed (только сетевые ошибки в http_first)
                remaining = [c for c in cards if c.nm_id not in processed]
        except Exception as e:
            import traceback as _tb
            print(f"⚠️  HTTP fast-path упал: {type(e).__name__}: {e}")
            print(_tb.format_exc()[:800])

    q: asyncio.Queue = asyncio.Queue()
    for c in remaining:
        await q.put(c)

    progress = {
        "done": len(processed),
        "links": sum(1 for r in store.rows if r.status == STATUS_LINK_COLLECTED),
        "no_docs": sum(1 for r in store.rows if r.status == STATUS_NO_DOCS),
        "no_link": sum(1 for r in store.rows if r.status == STATUS_NO_REGISTRY_LINK),
        "tech": sum(1 for r in store.rows if r.status in {STATUS_TIMEOUT, STATUS_ERROR}),
        "active": {},
        "start_time": time.time(),
    }
    print(f"К обработке в этом запуске: осталось={len(remaining)}, уже есть={len(processed)}")
    if not remaining:
        await store.save()
        print(f"Все карточки текущей выборки уже есть в CSV/XLSX: {len(processed)}/{len(cards)}")
        return

    if link_mode in ('http_only',):
        # v40: в http_only prefetch уже классифицировал все карточки (СОБРАНА/НЕТ ДОКУМЕНТОВ/
        # НЕТ ССЫЛКИ/ERROR) и записал их — сюда попадают только если что-то осталось
        # необработанным (крайне редко). Помечаем НЕТ ССЫЛКИ как защиту.
        if remaining:
            print(f"--link-mode http_only: {len(remaining)} карточек не получили ответа от certificate.json, помечаю НЕТ ССЫЛКИ НА РЕЕСТР")
        for c in remaining:
            row = ResultRow(
                query=c.source_query or args.query,
                nm_id=c.nm_id, product_name=c.product_name, brand=c.brand, subject=c.subject,
                product_url=c.product_url or product_url(c.nm_id),
                status=STATUS_NO_REGISTRY_LINK,
                details="http_only_mode; certificate.json не ответил",
                worker="http", checked_at=now_iso(),
            )
            await store.add(row)
        await store.save()
        return

    print(f"Параллельность: browsers={args.browser_count}, workers={args.workers}, headless={args.headless}, registry-mode=strict")
    async with async_playwright() as p:
        pool = BrowserPool(p, args)
        worker_tasks: Dict[int, asyncio.Task] = {}
        supervisor: Optional[asyncio.Task] = None
        prog: Optional[asyncio.Task] = None
        all_cards = {c.nm_id: c for c in cards}

        async def _cancel_workers(reason: str):
            nonlocal worker_tasks, supervisor
            if worker_tasks:
                print(f"⚠️  Перезапуск воркеров: {reason}. Текущие активные карточки будут возвращены в очередь.")
            for t in list(worker_tasks.values()):
                try:
                    t.cancel()
                except Exception:
                    pass
            if worker_tasks:
                try:
                    await asyncio.wait_for(asyncio.gather(*list(worker_tasks.values()), return_exceptions=True), timeout=8.0)
                except Exception:
                    pass
            worker_tasks = {}
            if supervisor:
                try:
                    supervisor.cancel()
                    await asyncio.gather(supervisor, return_exceptions=True)
                except Exception:
                    pass
                supervisor = None

        async def _start_workers(reason: str):
            nonlocal pool, worker_tasks, supervisor
            await pool.close_all()
            pool = BrowserPool(p, args)
            for i in range(max(1, args.browser_count)):
                await pool.get_browser(i)
            worker_tasks = {}
            for i in range(args.workers):
                worker_tasks[i + 1] = asyncio.create_task(link_worker(i + 1, pool, q, store, args, progress))
            supervisor = asyncio.create_task(supervisor_loop(worker_tasks, pool, q, store, args, progress, all_cards, len(cards)))
            if reason:
                print(f"↻ Воркеры запущены: {reason}")

        # v39: счётчик watchdog-рестартов поднят до finally, чтобы в финальный лог попал
        restart_count = 0

        try:
            await _start_workers("старт")
            prog = asyncio.create_task(progress_loop(q, store, args, progress, len(cards)))

            last_done = progress["done"]
            last_progress_ts = time.time()
            # v39: явный анонс watchdog'a при старте — чтобы было видно, что он включён
            stall_restart_sec_init = int(getattr(args, 'stall_restart_sec', 0) or 0)
            if stall_restart_sec_init > 0:
                print(f"🛡  Watchdog активен: рестарт воркеров/браузеров при отсутствии прогресса > {stall_restart_sec_init}с")
            else:
                print("🛡  Watchdog ВЫКЛЮЧЕН (stall-restart-sec=0). Включи флагом --stall-restart-sec 120.")
            # Не используем только queue.join(): если worker умер между get() и task_done(), q.join() зависнет навсегда.
            while progress["done"] < len(cards):
                await asyncio.sleep(2.0)
                current_done = progress["done"]
                if current_done != last_done:
                    last_done = current_done
                    last_progress_ts = time.time()

                # Если прогресс не меняется слишком долго, значит Playwright/Chromium или воркеры зависли.
                # Мы принудительно пересоздаём браузеры и возвращаем потерянные active-карточки обратно в очередь.
                stall_restart_sec = int(getattr(args, 'stall_restart_sec', 120) or 0)
                if stall_restart_sec > 0 and time.time() - last_progress_ts > stall_restart_sec:
                    restart_count += 1
                    stuck = []
                    now = time.time()
                    for w, info in list(progress.get("active", {}).items()):
                        try:
                            stuck.append(f"{w}:{info.get('nm_id')}:{int(now-info.get('started', now))}s/a{info.get('attempt',1)}")
                        except Exception:
                            pass
                    print("=" * 80)
                    print(f"⚠️  WATCHDOG СРАБОТАЛ #{restart_count}: нет прогресса {int(time.time()-last_progress_ts)}с (порог {stall_restart_sec}с)")
                    print(f"   Прогресс на момент срабатывания: {progress['done']}/{len(cards)}")
                    print(f"   Зависшие воркеры: {('; '.join(stuck[:10])) if stuck else '(нет)'}")
                    print(f"   Действия: cancel воркеров -> close all browsers -> reload -> restart")
                    print("=" * 80)
                    await _cancel_workers(f"нет прогресса {stall_restart_sec}с")
                    # После cancel active должен очиститься в finally воркеров. На всякий случай очищаем вручную:
                    progress.get("active", {}).clear()
                    await recover_missing_cards(q, store, all_cards, progress, "stall-restart")
                    await store.save()
                    await asyncio.sleep(min(20, 2 + restart_count * 2))
                    await _start_workers(f"stall-restart #{restart_count}")
                    print(f"✓  WATCHDOG: воркеры перезапущены, продолжаем (всего рестартов: {restart_count})")
                    last_done = progress["done"]
                    last_progress_ts = time.time()
                    continue

                if q.empty() and not progress["active"]:
                    await recover_missing_cards(q, store, all_cards, progress, "основной монитор пустой очереди")
                    if q.empty() and progress["done"] >= len(cards):
                        break

            await _cancel_workers("завершение")
            if prog:
                prog.cancel()
                await asyncio.gather(prog, return_exceptions=True)
        finally:
            await pool.close_all()

    await store.save()
    print(
        f"Финальный прогресс: обработано={progress['done']}/{len(cards)}, ссылок={progress['links']}, "
        f"нет документов={progress['no_docs']}, нет ссылки={progress['no_link']}, тех.статусы={progress['tech']}"
    )
    # v39: показать сколько раз watchdog сработал — полезно для оценки здоровья прогона
    if restart_count > 0:
        print(f"🛡  Watchdog сработал {restart_count} раз. Если число большое — попробуй уменьшить --browser-restart-every и --workers.")
    else:
        stall_restart_sec_init = int(getattr(args, 'stall_restart_sec', 0) or 0)
        if stall_restart_sec_init > 0:
            print("🛡  Watchdog ни разу не сработал — прогон прошёл без зависаний.")
    if args.output_links_csv:
        print(f"CSV со ссылками сохранён: {Path(args.output_links_csv).resolve()}")
    print(f"Готово. Excel сохранён: {Path(args.output).resolve()}")
    # v25-reporting: финальный текстовый лог прогона
    try:
        _log_arg = getattr(args, "run_log", "") or ""
        _log_path = Path(_log_arg) if _log_arg else None
        _wd = int(getattr(args, "expiry_warning_days", 30) or 0)
        _written = _write_run_log_v39(list(store.rows), Path(args.output), _wd,
                                      _run_started_at, mode="Stage 1 (ссылки)",
                                      log_path=_log_path)
        if _written:
            print(f"Лог прогона записан: {_written}")
    except Exception as _e:
        print(f"[run-log] ошибка: {_e}")


# -----------------------------
# Registry parsing / comparison
# -----------------------------

def flatten_json(obj: Any, prefix: str = "") -> Iterable[Tuple[str, Any]]:
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            yield from flatten_json(v, key)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from flatten_json(v, f"{prefix}[{i}]")
    else:
        yield prefix, obj

def clean_registry_value(value: Any) -> str:
    """Нормализует значение поля реестра, не ломая юридически важный текст."""
    s = html.unescape(str(value or "")).replace("\xa0", " ")
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    s = re.sub(r"\n\s*\n+", "\n", s)
    s = re.sub(r"\s+", " ", s).strip(" \t\n:-—")
    # Иногда BeautifulSoup склеивает повторяющиеся заголовки в начало значения.
    garbage_prefixes = [
        "товар 1",
        "сведения о продукции, на которую выдан сертификат соответствия",
        "сведения о продукции",
    ]
    changed = True
    while changed:
        changed = False
        low = norm_text(s)
        for pref in garbage_prefixes:
            if low.startswith(pref):
                s = s[len(pref):].strip(" :-—\n\t")
                changed = True
    return s[:4000]

PRODUCT_LABEL_PRIORITY = [
    # Киргизский SWIS/TULPAR и ФСА-сертификаты часто используют именно это поле.
    (1000, ("полное", "наименование", "продукц", "идентификац")),
    (950, ("полное", "наименование", "продукц")),
    (900, ("наименование", "продукц", "идентификац")),
    (860, ("сведения", "продукц", "идентификац")),
    (820, ("однородное", "наименование", "продукц")),
    (760, ("наименование", "продукц")),
    (700, ("описание", "продукц")),
    (650, ("product", "name")),
    (620, ("product", "description")),
    (600, ("goods", "name")),
]

NUMBER_LABEL_PRIORITY = [
    (1000, ("регистрационный", "номер", "документ")),
    (950, ("регистрационный", "номер", "сертифик")),
    (930, ("регистрационный", "номер", "деклара")),
    (900, ("registry", "number")),
    (880, ("reg", "number")),
    (860, ("certificate", "number")),
    (840, ("declaration", "number")),
    (700, ("номер", "документ")),
]

def _label_has(label_norm: str, words: Tuple[str, ...]) -> bool:
    return all(w in label_norm for w in words)

def _priority_for_label(label: str, rules: List[Tuple[int, Tuple[str, ...]]]) -> int:
    ln = norm_text(label)
    # Частые поля, которые НЕ являются названием продукции.
    bad_parts = (
        "изготовител", "заявител", "продавец", "адрес", "страна", "орган",
        "лаборатор", "эксперт", "схема", "статус", "дата", "срок", "основани",
        "норматив", "протокол", "код тн", "тн вэд", "место", "телефон", "email",
        "brand", "manufacturer", "applicant", "address", "country", "status", "date",
    )
    if any(b in ln for b in bad_parts):
        return -1000
    for score, words in rules:
        if _label_has(ln, words):
            return score
    return 0

def _looks_like_product_value(value: str) -> bool:
    v = clean_registry_value(value)
    low = norm_text(v)
    if len(v) < 18:
        return False
    if low.startswith("http"):
        return False
    if any(x in low for x in ("сведения о сертификате", "сведения о декларации", "регистрационный номер документа")) and len(v) < 120:
        return False
    # Значение должно содержать хотя бы буквы и не быть номером/датой.
    if len(re.findall(r"[а-яa-z]", low)) < 10:
        return False
    return True

def _extract_pairs_from_soup(soup) -> List[Tuple[str, str]]:
    """Возвращает пары label/value из таблиц и definition-list похожей верстки."""
    pairs: List[Tuple[str, str]] = []
    if not soup:
        return pairs
    # Таблицы: SWIS/TULPAR и многие страницы ФСА делают именно так.
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if len(cells) >= 2:
            vals = [clean_registry_value(c.get_text(" ", strip=True)) for c in cells]
            vals = [v for v in vals if v]
            if len(vals) >= 2:
                label = vals[0]
                value = " ".join(vals[1:])
                pairs.append((label, value))
    # Иногда label/value находятся в соседних div/span/li.
    for parent in soup.find_all(["div", "li", "section"]):
        kids = [clean_registry_value(x.get_text(" ", strip=True)) for x in parent.find_all(recursive=False)]
        kids = [k for k in kids if k]
        if 2 <= len(kids) <= 8:
            for i in range(len(kids) - 1):
                pairs.append((kids[i], kids[i+1]))
    return pairs

def _extract_pairs_from_visible_text(visible: str) -> List[Tuple[str, str]]:
    lines = [clean_registry_value(x) for x in visible.split("\n")]
    lines = [x for x in lines if x]
    pairs: List[Tuple[str, str]] = []
    for i, line in enumerate(lines):
        # Случай: label: value на одной строке.
        if ":" in line and len(line) < 300:
            left, right = line.split(":", 1)
            if left.strip() and right.strip():
                pairs.append((left.strip(), right.strip()))
        # Случай: label в одной строке, value в следующей.
        pr = _priority_for_label(line, PRODUCT_LABEL_PRIORITY)
        nr = _priority_for_label(line, NUMBER_LABEL_PRIORITY)
        if pr > 0 or nr > 0:
            buf = []
            for j in range(i + 1, min(i + 8, len(lines))):
                nxt = lines[j]
                # Остановиться на следующем label.
                if j > i + 1 and (_priority_for_label(nxt, PRODUCT_LABEL_PRIORITY) > 0 or _priority_for_label(nxt, NUMBER_LABEL_PRIORITY) > 0):
                    break
                # Заголовки секций пропускаем.
                if norm_text(nxt) in {"товар 1", "товар 2", "товар 3"}:
                    continue
                buf.append(nxt)
                # Для номера достаточно одной строки.
                if nr > 0 and buf:
                    break
                # Для названия продукции обычно достаточно 1-3 строк, но длинный текст может идти дальше.
                if pr > 0 and sum(len(x) for x in buf) > 120:
                    break
            if buf:
                pairs.append((line, " ".join(buf)))
    return pairs

def _pick_best_from_pairs(pairs: List[Tuple[str, str]]) -> Tuple[str, str]:
    cert_candidates: List[Tuple[int, str]] = []
    product_candidates: List[Tuple[int, str]] = []
    for label, value in pairs:
        val = clean_registry_value(value)
        if not val:
            continue
        nscore = _priority_for_label(label, NUMBER_LABEL_PRIORITY)
        if nscore > 0 and 4 <= len(val) <= 250:
            cert_candidates.append((nscore, val))
        pscore = _priority_for_label(label, PRODUCT_LABEL_PRIORITY)
        if pscore > 0 and _looks_like_product_value(val):
            # Чуть повышаем приоритет длинного полного поля, но не даём описанию на 4000 символов доминировать над точным полем.
            product_candidates.append((pscore + min(len(val), 500) // 10, val))
    cert = sorted(cert_candidates, key=lambda x: x[0], reverse=True)[0][1] if cert_candidates else ""
    product = sorted(product_candidates, key=lambda x: x[0], reverse=True)[0][1] if product_candidates else ""
    return cert, product

def pick_from_json(data: Any) -> Tuple[str, str, str]:
    cert_num = ""
    product_name = ""
    doc_type = ""
    product_candidates: List[Tuple[int, str]] = []
    cert_candidates: List[Tuple[int, str]] = []

    for k, v in flatten_json(data):
        if not isinstance(v, (str, int, float)):
            continue
        val = clean_registry_value(v)
        if not val:
            continue
        kl = k.lower()
        kln = norm_text(k)

        nscore = _priority_for_label(kln, NUMBER_LABEL_PRIORITY)
        if nscore > 0 and 4 <= len(val) <= 250:
            cert_candidates.append((nscore, val))

        pscore = _priority_for_label(kln, PRODUCT_LABEL_PRIORITY)
        # Дополнительные англоязычные ключи ФСА API.
        if any(x in kl for x in ("productname", "productfullname", "productdescription", "goodsdescription", "goodsname")):
            pscore = max(pscore, 850)
        if pscore > 0 and _looks_like_product_value(val):
            product_candidates.append((pscore + min(len(val), 500) // 10, val))

        if any(x in kl for x in ("certificate", "cert")) or "сертифик" in kln:
            doc_type = doc_type or "сертификат"
        if "declaration" in kl or "деклара" in kln:
            doc_type = doc_type or "декларация"

    if cert_candidates:
        cert_num = sorted(cert_candidates, key=lambda x: x[0], reverse=True)[0][1]
    if product_candidates:
        product_name = sorted(product_candidates, key=lambda x: x[0], reverse=True)[0][1]
    return cert_num, product_name, doc_type

async def http_get(session, url: str, timeout: float = 20.0) -> Tuple[int, str, str]:
    try:
        async with session.get(url, timeout=timeout, allow_redirects=True) as r:
            ct = r.headers.get("content-type", "")
            txt = await r.text(errors="ignore")
            return r.status, txt, ct
    except Exception as e:
        return 0, "", str(e)


def api_candidates_for_registry(url: str) -> List[str]:
    """HTTP/API/SPA candidates for registry pages.

    v38 link collection is preserved. This function only improves stage 2.
    For FSA the visible registry page is a SPA; the product text can be on
    /product route or in JSON fetched by the page. We therefore try the original
    URL, product route, and known API-like routes without treating any single
    endpoint as mandatory.
    """
    rec = extract_record_id(url)
    h = hostname(url)
    p = urlparse(url).path.lower()
    out = [url]
    if h == "pub.fsa.gov.ru" and rec:
        # The UI has separate tabs/routes. Product is often not present on baseInfo/common.
        if "/view/" in p:
            parts = [x for x in p.split('/') if x]
            try:
                vi = parts.index('view')
                base = '/' + '/'.join(parts[:vi+2])
                for suffix in ("/product", "/common", "/baseInfo"):
                    out.append(f"https://pub.fsa.gov.ru{base}{suffix}")
            except Exception:
                pass
        if "/rds/declaration/" in p:
            out.extend([
                f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{rec}",
                f"https://pub.fsa.gov.ru/api/v1/rds/declarations/{rec}",
                f"https://pub.fsa.gov.ru/api/v1/rds/declarations/{rec}/product",
                f"https://pub.fsa.gov.ru/api/v1/rds/common/declarations/{rec}/product",
            ])
        if "/rss/certificate/" in p:
            out.extend([
                f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{rec}",
                f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{rec}/baseInfo",
                f"https://pub.fsa.gov.ru/api/v1/rss/common/certificates/{rec}/product",
                f"https://pub.fsa.gov.ru/api/v1/rss/certificates/{rec}",
                f"https://pub.fsa.gov.ru/api/v1/rss/certificates/{rec}/product",
                f"https://pub.fsa.gov.ru/api/v1/rss/certificate/{rec}",
            ])
    return list(dict.fromkeys(out))

# Более широкий набор label-слов для ФСА и SWIS/TULPAR.
FSA_PRODUCT_LABELS = [
    "Продукция",
    "Наименование продукции",
    "Полное наименование продукции",
    "Сведения о продукции",
    "Описание продукции",
    "Полное наименование продукции и сведения, обеспечивающие её идентификацию",
    "Полное наименование продукции и сведения, обеспечивающие ее идентификацию",
    "Однородное наименование продукции",
]

FSA_NUMBER_LABELS = [
    "Регистрационный номер документа",
    "Регистрационный номер сертификата соответствия",
    "Регистрационный номер сертификата",
    "Регистрационный номер декларации",
    "Номер сертификата",
    "Номер декларации",
    "Номер документа",
]

STOP_NEXT_LABELS = [
    "код тн", "тн вэд", "количество", "дополнительная информация", "товар ",
    "заявитель", "изготовитель", "производитель", "сведения об", "сведения о",
    "адрес", "страна", "дата", "срок", "статус", "номер", "регистрационный",
    "документ", "сертификат", "декларация", "основание", "протокол", "орган",
    "соответствует требованиям", "схема", "инспекционный", "обозначение", "реквизиты",
]


def _trim_product_value(val: str) -> str:
    val = clean_registry_value(val)
    # If a shorter label matched inside the long SWIS label, remove the tail of the label itself.
    val = re.sub(r"(?i)^и\s+сведения,?\s+обеспечивающие\s+е[её]\s+идентификацию\s*\([^)]*\)\s*", "", val).strip()
    val = re.sub(r"(?i)^и\s+сведения,?\s+обеспечивающие\s+е[её]\s+идентификацию\s*", "", val).strip()
    # Stop when another registry field begins. Do not stop on punctuation inside product lists.
    stop_patterns = [
        r"\bЗаявитель\b", r"\bИзготовитель\b", r"\bПроизводитель\b", r"\bСоответствует требованиям\b",
        r"\bКод\s+ТН\b", r"\bТН\s+ВЭД\b", r"\bДата\b", r"\bСрок\b", r"\bСтатус\b",
        r"\bРегистрационный\s+номер\b", r"\bНомер\s+документа\b", r"\bОрган\s+по\s+сертификации\b",
        r"\bСведения\s+о\s+заявител", r"\bСведения\s+об\s+изготовител", r"\bДополнительная\s+информация\b",
        r"\bКоличество\b", r"\bТовар\s+\d+\b",
    ]
    cut = len(val)
    for pat in stop_patterns:
        m = re.search(pat, val, flags=re.I)
        if m and m.start() >= 20:
            cut = min(cut, m.start())
    return clean_registry_value(val[:cut])

def _extract_after_label_same_or_next(text: str, labels: List[str], max_chars: int = 3500) -> List[str]:
    """Extract value after labels like 'Продукция:' even when value is on the same line.

    This is the key fix for FSA and SWIS/TULPAR. The old parser mostly expected
    table-like label/value pairs or label on one line and value on the next. FSA
    often renders 'Продукция: <long value>' in one text block.
    """
    if not text:
        return []
    # Preserve line breaks but also make an additional collapsed string for SPA text.
    visible = clean_registry_value(text).replace(" :", ":")
    # Re-expand known labels onto new lines to make stop detection more reliable.
    expanded = visible
    for lab in sorted(labels + FSA_NUMBER_LABELS, key=len, reverse=True):
        expanded = re.sub(rf"(?i)(?<!^)(\s)({re.escape(lab)}\s*:)", r"\n\2", expanded)
    # Important SWIS pattern has no colon sometimes: label + value.
    for lab in sorted(labels, key=len, reverse=True):
        expanded = re.sub(rf"(?i)(?<!^)(\s)({re.escape(lab)})(\s+)", r"\n\2 ", expanded)
    candidates: List[str] = []
    lines = [clean_registry_value(x) for x in expanded.split("\n") if clean_registry_value(x)]

    for i, line in enumerate(lines):
        low_line = norm_text(line)
        for lab in sorted(labels, key=len, reverse=True):
            low_lab = norm_text(lab)
            if low_lab not in low_line:
                continue
            # Same-line value after label / label:
            val = ""
            m = re.search(re.escape(lab) + r"\s*:?\s*(.*)$", line, flags=re.I)
            if m:
                val = _trim_product_value(m.group(1))
            if not val or norm_text(val) == low_lab:
                parts = []
                for nxt in lines[i+1:i+10]:
                    nl = norm_text(nxt)
                    if any(nl.startswith(x) for x in STOP_NEXT_LABELS):
                        break
                    # stop on another product/number label after first value line
                    if parts and any(norm_text(l) in nl[:160] for l in labels + FSA_NUMBER_LABELS):
                        break
                    parts.append(nxt)
                    if sum(len(x) for x in parts) >= max_chars or (parts and sum(len(x) for x in parts) >= 250):
                        # Usually enough, unless this is a list; keep it bounded for speed/readability.
                        break
                val = _trim_product_value(" ".join(parts))
            # Remove label accidentally repeated in value.
            for lab2 in labels:
                val = re.sub(rf"(?i)^\s*{re.escape(lab2)}\s*:?\s*", "", val).strip()
            if _looks_like_product_value(val):
                candidates.append(val[:max_chars])
    return candidates


def _extract_cert_number_by_labels(text: str) -> str:
    visible = clean_registry_value(text).replace(" :", ":")
    # Same-line and next-line variants.
    for lab in FSA_NUMBER_LABELS:
        patterns = [
            rf"(?i){re.escape(lab)}\s*:?\s*([^\n;|]{{4,250}})",
            rf"(?i){re.escape(lab)}\s*\n\s*([^\n]{{4,250}})",
        ]
        for pat in patterns:
            m = re.search(pat, visible)
            if m:
                val = clean_registry_value(m.group(1))
                # Avoid swallowing product text after very generic labels.
                val = re.split(r"\s+(?:Дата|Статус|Заявитель|Изготовитель|Продукция)\b", val, flags=re.I)[0]
                if 4 <= len(val) <= 250:
                    return val
    return ""


def parse_html_registry(url: str, text: str) -> Tuple[str, str, str]:
    """Robust HTML/SPA text parser for FSA and SWIS/TULPAR.

    v38 link collection is untouched. This parser is deliberately conservative:
    it prefers explicit fields with labels, especially FSA 'Продукция:' and
    SWIS 'Полное наименование продукции и сведения...'.
    """
    low = norm_text(text)
    doc_type = ""
    if "/rds/declaration/" in url.lower() or "сведения о декларации" in low or "деклараци" in low:
        doc_type = "декларация"
    if "/rss/certificate/" in url.lower() or "сведения о сертификате" in low or "сертификат" in low:
        doc_type = "сертификат"

    if BeautifulSoup:
        soup = BeautifulSoup(text or "", "html.parser")
        for tag in soup(["script", "style", "noscript", "svg"]):
            tag.decompose()
        visible = soup.get_text("\n", strip=True)
    else:
        soup = None
        visible = re.sub(r"<[^>]+>", "\n", text or "")

    pairs: List[Tuple[str, str]] = []
    if soup is not None:
        pairs.extend(_extract_pairs_from_soup(soup))
    pairs.extend(_extract_pairs_from_visible_text(visible))

    cert_num, product_name = _pick_best_from_pairs(pairs)

    # Priority fix: FSA label 'Продукция:' and SWIS full product field.
    label_candidates = _extract_after_label_same_or_next(visible, FSA_PRODUCT_LABELS)
    if label_candidates:
        # Prefer the richest field, but don't let factory/address appendices dominate.
        label_candidates = sorted(set(label_candidates), key=lambda x: (len(x), x), reverse=True)
        product_name = label_candidates[0]

    if not cert_num:
        cert_num = _extract_cert_number_by_labels(visible)

    # Regex fallback for SWIS/FSA fields when line breaks are odd.
    if not product_name:
        collapsed = clean_registry_value(visible)
        product_patterns = [
            r"Продукция\s*:?\s*([\s\S]{20,2500}?)(?=\s+(?:Заявитель|Изготовитель|Производитель|Соответствует требованиям|Код\s+ТН|ТН\s+ВЭД|Дата|Срок|Статус|Регистрационный\s+номер|Номер\s+документа|$))",
            r"Полное\s+наименование\s+продукции\s+и\s+сведения,?\s+обеспечивающие\s+е[её]\s+идентификацию[^а-яa-z0-9]+([\s\S]{20,2500}?)(?=\s+(?:Код\s+ТН|ТН\s+ВЭД|Количество|Дополнительная|Товар\s+\d+|Сведения\s+об|Инспекционный|$))",
            r"Однородное\s+наименование\s+продукции[^а-яa-z0-9]+([\s\S]{20,1800}?)(?=\s+(?:Полное\s+наименование|Код\s+ТН|ТН\s+ВЭД|Количество|Дополнительная|Товар\s+\d+|$))",
            r"Наименование\s+продукции[^а-яa-z0-9]+([\s\S]{20,1800}?)(?=\s+(?:Код\s+ТН|ТН\s+ВЭД|Количество|Дополнительная|Товар\s+\d+|$))",
        ]
        for pat in product_patterns:
            m = re.search(pat, collapsed, flags=re.I)
            if m:
                val = clean_registry_value(m.group(1))
                if _looks_like_product_value(val):
                    product_name = _trim_product_value(val)
                    break

    return cert_num, product_name, doc_type





# -----------------------------
# v38.2 overrides: registry parsers without touching WB link collection
# -----------------------------

_LEGACY_PARSE_HTML_REGISTRY = parse_html_registry


def _unique_keep_order(values: Iterable[str]) -> List[str]:
    out: List[str] = []
    seen: Set[str] = set()
    for v in values:
        vv = clean_registry_value(v)
        key = norm_text(vv)
        if vv and key and key not in seen:
            out.append(vv)
            seen.add(key)
    return out



def _table_pairs_exact(soup) -> List[Tuple[str, str]]:
    """Extract exact label/value pairs from registry HTML.

    v38.9: FSA exact tab/number fix. In pub.fsa.gov.ru the `title` attribute with
    «Регистрационный номер сертификата» / «Регистрационный номер декларации о
    соответствии» is often placed on the header element, while the value is in a
    sibling inside a parent `figis-card-info-row`. The previous parser read only
    the element with `title`, therefore it saw the label but missed the value.
    This version climbs to the nearest row/container and reads the sibling value.
    """
    pairs: List[Tuple[str, str]] = []
    if not soup:
        return pairs

    def add_pair(label: str, value: str):
        label = clean_registry_value(label)
        value = clean_registry_value(value)
        if not label or not value:
            return
        nl = norm_text(label)
        nv = norm_text(value)
        if nv == nl:
            return
        if nv.startswith(nl):
            value = clean_registry_value(value[len(label):])
        if value:
            pairs.append((label, value))

    def cls_has(tag, needle: str) -> bool:
        try:
            classes = tag.get('class') or []
        except Exception:
            return False
        return any(needle in str(c) for c in classes)

    def attr_label(tag) -> str:
        for attr in ('title', 'ng-reflect-title', 'data-title', 'aria-label'):
            try:
                v = tag.get(attr)
            except Exception:
                v = None
            if v:
                return clean_registry_value(v)
        return ''

    def is_value_node(t) -> bool:
        return bool(getattr(t, 'name', None)) and (
            cls_has(t, 'info-row__text') or cls_has(t, 'info-row-text') or
            cls_has(t, 'row__text') or cls_has(t, 'value') or cls_has(t, 'description') or
            cls_has(t, 'field-value')
        )

    def is_header_node(t) -> bool:
        return bool(getattr(t, 'name', None)) and (
            cls_has(t, 'info-row__header') or cls_has(t, 'info-row-header') or
            cls_has(t, 'row__header') or cls_has(t, 'label') or cls_has(t, 'field-title') or
            t.name in ('th',)
        )

    def nearest_row_candidates(tag):
        cur = tag
        out = []
        for _ in range(7):
            if not cur:
                break
            out.append(cur)
            try:
                cur = cur.parent
            except Exception:
                break
        return out

    def value_from_container(container, label: str) -> str:
        # Prefer explicit value/text nodes.
        vals = []
        for n in container.find_all(is_value_node):
            txt = clean_registry_value(n.get_text(' ', strip=True))
            if txt and norm_text(txt) != norm_text(label):
                vals.append(txt)
        if vals:
            return clean_registry_value(' '.join(vals))
        # Classic cells inside selected container.
        cells = container.find_all(['td', 'th'])
        if len(cells) >= 2:
            vals = [clean_registry_value(c.get_text(' ', strip=True)) for c in cells]
            if vals and norm_text(vals[0]) == norm_text(label):
                return clean_registry_value(' '.join(vals[1:]))
        # Last fallback: clone and remove obvious header/label nodes, then read rest.
        try:
            clone = BeautifulSoup(str(container), 'html.parser')
            for h in clone.find_all(is_header_node):
                h.decompose()
            # Also remove nodes whose own attr title exactly equals the label.
            for h in clone.find_all(attrs={'title': True}):
                if norm_text(h.get('title')) == norm_text(label):
                    h.decompose()
            txt = clean_registry_value(clone.get_text(' ', strip=True))
            if norm_text(txt).startswith(norm_text(label)):
                txt = clean_registry_value(txt[len(label):])
            return txt
        except Exception:
            txt = clean_registry_value(container.get_text(' ', strip=True))
            if norm_text(txt).startswith(norm_text(label)):
                txt = clean_registry_value(txt[len(label):])
            return txt

    # 1) Classic table layout.
    for tr in soup.find_all('tr'):
        cells = tr.find_all(['th', 'td'])
        if len(cells) >= 2:
            label = clean_registry_value(cells[0].get_text(' ', strip=True))
            value = clean_registry_value(' '.join(c.get_text(' ', strip=True) for c in cells[1:]))
            add_pair(label, value)

    # 2) FSA Angular/custom components by title-like attributes. Important:
    # label may be on header child, value in parent/sibling.
    for tag in soup.find_all(lambda t: getattr(t, 'name', None) and (attr_label(t) != '')):
        label = attr_label(tag)
        if not label:
            continue
        best = ''
        for container in nearest_row_candidates(tag):
            v = value_from_container(container, label)
            if v and norm_text(v) != norm_text(label):
                best = v
                # Stop once we got a plausible value from a reasonably small row.
                if len(best) < 5000:
                    break
        add_pair(label, best)

    # 3) FSA header/text class layout.
    for row in soup.find_all(lambda t: getattr(t, 'name', None) and (cls_has(t, 'figis-card-info-row') or cls_has(t, 'card-info-row') or cls_has(t, 'info-row') or t.name in ('figis-card-info-row',))):
        header = row.find(is_header_node)
        text_node = row.find(is_value_node)
        if header and text_node:
            add_pair(header.get_text(' ', strip=True), text_node.get_text(' ', strip=True))

    # 4) Generic adjacent label/value blocks: useful for minified Angular markup.
    for lab_tag in soup.find_all(is_header_node):
        label = clean_registry_value(lab_tag.get_text(' ', strip=True)) or attr_label(lab_tag)
        if not label:
            continue
        parent = lab_tag.parent
        if parent:
            v = value_from_container(parent, label)
            add_pair(label, v)

    out: List[Tuple[str, str]] = []
    seen: Set[Tuple[str, str]] = set()
    for label, value in pairs:
        key = (norm_text(label), norm_text(value))
        if key not in seen:
            out.append((label, value))
            seen.add(key)
    return out


def _extract_same_line_or_next_exact(visible: str, label: str, max_chars: int = 5000) -> List[str]:
    """Extract values for exact Russian labels in table-like visible text.

    Works when HTML is flattened as:
      Label\nValue
    and when it is flattened as:
      Label Value
    """
    out: List[str] = []
    if not visible:
        return out
    text = visible.replace('\xa0', ' ')
    label_re = re.escape(label).replace('е\\ё', '[её]')
    # Line-based variant.
    lines = [clean_registry_value(x) for x in re.split(r'[\r\n]+', text) if clean_registry_value(x)]
    for i, line in enumerate(lines):
        ln = norm_text(line)
        if norm_text(label) == ln or ln.startswith(norm_text(label) + ' '):
            val = ''
            # same line after label
            m = re.match(rf'(?is)^{label_re}\s*:?[\s\-–—]*(.+)$', line)
            if m:
                val = clean_registry_value(m.group(1))
            if not val or norm_text(val) == norm_text(label):
                buf: List[str] = []
                for nxt in lines[i + 1:i + 12]:
                    nln = norm_text(nxt)
                    if nln.startswith(('код тн', 'тн вэд', 'количество', 'дополнительная информация', 'товар ', 'сведения об', 'сведения о ', 'инспекционный контроль', 'регистрационный номер документа')):
                        break
                    if any(nln.startswith(norm_text(x)) for x in SWIS_PRODUCT_LABELS + FSA_EXACT_PRODUCT_LABELS + FSA_NUMBER_LABELS):
                        if buf:
                            break
                    buf.append(nxt)
                    if sum(len(x) for x in buf) >= max_chars:
                        break
                val = clean_registry_value(' '.join(buf))
            if _looks_like_product_value(val) or ('регистрационный номер' in norm_text(label) and val):
                out.append(_trim_product_value(val) if 'продукц' in norm_text(label) else clean_registry_value(val))
    # Collapsed variant for rows flattened into one string.
    collapsed = clean_registry_value(text)
    stop = r'(?=\s+(?:Код\s+ТН|ТН\s+ВЭД|Количество|Дополнительная\s+информация|Товар\s+\d+|Сведения\s+об|Сведения\s+о\s+[А-ЯA-Z]|Инспекционный\s+контроль|Регистрационный\s+номер\s+документа|Заявитель|Изготовитель|Производитель|Документы|$))'
    for m in re.finditer(rf'(?is){label_re}\s*:?[\s\-–—]*(.{{10,{max_chars}}}?){stop}', collapsed):
        val = clean_registry_value(m.group(1))
        if 'продукц' in norm_text(label):
            val = _trim_product_value(val)
        if val and ((_looks_like_product_value(val)) or ('регистрационный номер' in norm_text(label) and len(val) >= 4)):
            out.append(val)
    return _unique_keep_order(out)


SWIS_PRODUCT_LABELS = [
    'Однородное наименование продукции',
    'Полное наименование продукции и сведения, обеспечивающие её идентификацию (тип, марка, модель, артикул продукции и др.)',
    'Полное наименование продукции и сведения, обеспечивающие ее идентификацию (тип, марка, модель, артикул продукции и др.)',
]

FSA_EXACT_PRODUCT_LABELS = ['Наименование (обозначение) продукции', 'Общее наименование продукции']
FSA_EXACT_NUMBER_LABELS = ['Регистрационный номер сертификата', 'Регистрационный номер декларации о соответствии']


def _swis_value_is_not_product(value: str) -> bool:
    low = norm_text(value)
    # SWIS may repeat the same long label for manufacturer branches. Do not mix
    # factory/address rows into product name. Keep only product-like rows.
    bad = ('филиал', 'филиалы', 'завод', 'изготовител', 'адрес', 'координат', 'manufacturing', 'co. ltd', 'co., ltd', 'limited', 'ltd')
    if any(x in low for x in bad):
        # If it also contains clear apparel/product words, keep it; otherwise it is likely a factory/address appendice.
        product_words = ('издел', 'одежд', 'бель', 'куртк', 'пальто', 'пижам', 'футбол', 'брюк', 'плать', 'трикотаж', 'швейн', 'обув', 'игруш')
        return not any(w in low for w in product_words)
    return False


def parse_swis_registry_strict(text: str) -> Tuple[str, str, str]:
    """Parse Kyrgyz SWIS/TULPAR.

    Product data is intentionally taken ONLY from:
      1) Однородное наименование продукции
      2) Полное наименование продукции и сведения, обеспечивающие её идентификацию (...)
    as confirmed by the user and by public SWIS pages.
    """
    if BeautifulSoup:
        soup = BeautifulSoup(text or '', 'html.parser')
        for tag in soup(['script', 'style', 'noscript', 'svg']):
            tag.decompose()
        visible = soup.get_text('\n', strip=True)
        pairs = _table_pairs_exact(soup)
    else:
        visible = re.sub(r'<[^>]+>', '\n', text or '')
        pairs = []

    cert_numbers: List[str] = []
    homogeneous: List[str] = []
    full_products: List[str] = []

    for label, value in pairs:
        ln = norm_text(label)
        val = clean_registry_value(value)
        if not val:
            continue
        if 'регистрационный номер документа' in ln:
            cert_numbers.append(val)
        elif 'однородное наименование продукции' in ln:
            if _looks_like_product_value(val):
                homogeneous.append(_trim_product_value(val))
        elif 'полное наименование продукции' in ln and 'идентификац' in ln:
            val = _trim_product_value(val)
            if _looks_like_product_value(val) and not _swis_value_is_not_product(val):
                full_products.append(val)

    if not cert_numbers:
        cert_numbers.extend(_extract_same_line_or_next_exact(visible, 'Регистрационный номер документа', max_chars=250))
    if not homogeneous:
        homogeneous.extend(_extract_same_line_or_next_exact(visible, 'Однородное наименование продукции', max_chars=1800))
    if not full_products:
        for lab in SWIS_PRODUCT_LABELS[1:]:
            full_products.extend([v for v in _extract_same_line_or_next_exact(visible, lab, max_chars=5000) if not _swis_value_is_not_product(v)])

    cert_num = _unique_keep_order(cert_numbers)[0] if _unique_keep_order(cert_numbers) else ''
    full_products = _unique_keep_order(full_products)
    homogeneous = _unique_keep_order(homogeneous)

    # Prefer detailed 'Полное...' rows. If there are several product rows, join them;
    # this is safer for matching WB card against a broad certificate.
    if full_products:
        product = '; '.join(full_products[:20])
    elif homogeneous:
        product = '; '.join(homogeneous[:5])
    else:
        product = ''
    doc_type = 'сертификат' if 'сведения о сертификате' in norm_text(visible) or 'сертификат соответствия' in norm_text(visible) else ''
    return cert_num, product, doc_type


def parse_fsa_registry_strict(text: str, url: str = '') -> Tuple[str, str, str]:
    """Parse FSA registry by exact UI fields confirmed by user screenshots.

    Product name is taken ONLY from:
      - «Наименование (обозначение) продукции»

    Document number is taken ONLY from:
      - «Регистрационный номер сертификата»
      - «Регистрационный номер декларации о соответствии»

    FSA pages are SPA/Angular pages, so the function supports both table-like
    text and figis-card-info-row/title-based DOM captured by browser fallback.
    """
    if BeautifulSoup:
        soup = BeautifulSoup(text or '', 'html.parser')
        for tag in soup(['script', 'style', 'noscript', 'svg']):
            tag.decompose()
        visible = soup.get_text('\n', strip=True)
        pairs = _table_pairs_exact(soup)
    else:
        visible = re.sub(r'<[^>]+>', '\n', text or '')
        pairs = []

    product_candidates: List[str] = []
    cert_candidates: List[str] = []

    product_label_norms = {norm_text(x) for x in FSA_EXACT_PRODUCT_LABELS}
    number_label_norms = {norm_text(x) for x in FSA_EXACT_NUMBER_LABELS}

    for label, value in pairs:
        ln = norm_text(label).strip(': .')
        val = clean_registry_value(value)
        if not val:
            continue
        if ln in product_label_norms:
            val = _trim_product_value(val)
            if _looks_like_product_value(val):
                product_candidates.append(val)
        elif ln in number_label_norms:
            if 4 <= len(val) <= 250:
                cert_candidates.append(val)

    # Text fallback for rendered/flattened pages.
    if not product_candidates:
        for lab in FSA_EXACT_PRODUCT_LABELS:
            product_candidates.extend(_extract_same_line_or_next_exact(visible, lab, max_chars=7000))

    if not cert_candidates:
        for lab in FSA_EXACT_NUMBER_LABELS:
            vals = _extract_same_line_or_next_exact(visible, lab, max_chars=250)
            cert_candidates.extend([v for v in vals if 4 <= len(v) <= 250])

    product_candidates = _unique_keep_order([_trim_product_value(v) for v in product_candidates if _looks_like_product_value(v)])
    cert_candidates = _unique_keep_order([clean_registry_value(v) for v in cert_candidates if 4 <= len(clean_registry_value(v)) <= 250])

    low = norm_text(visible + ' ' + url)
    if '/rds/declaration/' in url.lower() or 'деклараци' in low:
        doc_type = 'декларация'
    elif '/rss/certificate/' in url.lower() or 'сертификат' in low:
        doc_type = 'сертификат'
    else:
        doc_type = ''

    return (cert_candidates[0] if cert_candidates else ''), ('; '.join(product_candidates[:10]) if product_candidates else ''), doc_type


def parse_html_registry(url: str, text: str) -> Tuple[str, str, str]:
    h = hostname(url)
    if h in {'swis.trade.kg', 'trade.kg'}:
        cert, prod, typ = parse_swis_registry_strict(text)
        if prod or cert:
            return cert, prod, typ
    if h == 'pub.fsa.gov.ru':
        cert, prod, typ = parse_fsa_registry_strict(text, url)
        if prod or cert:
            return cert, prod, typ
    return _LEGACY_PARSE_HTML_REGISTRY(url, text)


# -----------------------------
# v38.4 exact FSA registry parser override
# -----------------------------
# This block intentionally DOES NOT touch WB link collection. It replaces only
# registry-stage parsing for pub.fsa.gov.ru and keeps SWIS/TULPAR strict logic.

FSA_EXACT_PRODUCT_LABELS = ['Наименование (обозначение) продукции', 'Общее наименование продукции']
FSA_EXACT_CERT_LABELS = ['Регистрационный номер сертификата']
FSA_EXACT_DECL_LABELS = ['Регистрационный номер декларации о соответствии']
FSA_EXACT_NUMBER_LABELS = FSA_EXACT_CERT_LABELS + FSA_EXACT_DECL_LABELS


def fsa_exact_routes(url: str) -> Tuple[List[str], List[str]]:
    """Return (number_routes, product_routes) for FSA exact UI fields.

    Number is collected ONLY from the certificate/declaration tab:
      - «СЕРТИФИКАТ» / «ДЕКЛАРАЦИЯ О СООТВЕТСТВИИ»
      - exact row title «Регистрационный номер сертификата» or
        «Регистрационный номер декларации о соответствии»

    Product is collected ONLY from:
      - «СВЕДЕНИЯ О ПРОДУКЦИИ»
      - exact row title «Наименование (обозначение) продукции» or «Общее наименование продукции»
    """
    if hostname(url) != 'pub.fsa.gov.ru':
        return [url], [url]
    parsed = urlparse(url)
    path = parsed.path.lower()
    parts = [x for x in path.split('/') if x]
    base = ''
    try:
        vi = parts.index('view')
        base = '/' + '/'.join(parts[:vi + 2])
    except Exception:
        # fallback from record id and known registry type
        rec = extract_record_id(url)
        if rec and '/rss/certificate/' in path:
            base = f'/rss/certificate/view/{rec}'
        elif rec and '/rds/declaration/' in path:
            base = f'/rds/declaration/view/{rec}'
    if not base:
        return [url], [url]

    root = 'https://pub.fsa.gov.ru'
    if '/rss/certificate/' in path:
        number_routes = [f'{root}{base}/baseInfo', f'{root}{base}/common', url]
        product_routes = [f'{root}{base}/product', url]
    elif '/rds/declaration/' in path:
        # In the public UI declarations normally keep the document number on
        # /common, while product data is on /product. Include /baseInfo as a
        # harmless fallback because the UI naming changed several times.
        number_routes = [f'{root}{base}/common', f'{root}{base}/baseInfo', url]
        product_routes = [f'{root}{base}/product', url]
    else:
        number_routes = [url, f'{root}{base}/baseInfo', f'{root}{base}/common']
        product_routes = [f'{root}{base}/product', url]
    return list(dict.fromkeys(number_routes)), list(dict.fromkeys(product_routes))


def fsa_extended_routes(url: str) -> List[str]:
    """v44: дополнительные вкладки FSA, где лежат изготовитель / ТР ТС / заявитель.

    На pub.fsa.gov.ru данные документа разнесены по вкладкам:
      /baseInfo   — номер, статус, даты, схема (number page)
      /applicant  — заявитель + ИНН
      /manufacturer — изготовитель
      /product    — продукция + ТН ВЭД (product page)
      /document   — технические регламенты (ТР ТС)
    number_routes/product_routes покрывают baseInfo и product. Эта функция
    возвращает applicant/manufacturer/document, чтобы добрать остальные поля.
    """
    if hostname(url) != 'pub.fsa.gov.ru':
        return []
    parsed = urlparse(url)
    parts = [x for x in parsed.path.lower().split('/') if x]
    base = ''
    try:
        vi = parts.index('view')
        base = '/' + '/'.join(parts[:vi + 2])
    except Exception:
        rec = extract_record_id(url)
        low = parsed.path.lower()
        if rec and '/rss/certificate/' in low:
            base = f'/rss/certificate/view/{rec}'
        elif rec and '/rds/declaration/' in low:
            base = f'/rds/declaration/view/{rec}'
    if not base:
        return []
    root = 'https://pub.fsa.gov.ru'
    return [f'{root}{base}/applicant', f'{root}{base}/manufacturer', f'{root}{base}/document']


def _normalize_fsa_label(label: str) -> str:
    return norm_text(label).strip(' :.;-–—')


def _fsa_value_clean(label: str, value: str) -> str:
    value = clean_registry_value(value)
    if not value:
        return ''
    nl = _normalize_fsa_label(label)
    nv = norm_text(value)
    if nv.startswith(nl):
        # Remove duplicated label if DOM text was flattened.
        value = clean_registry_value(value[len(label):])
    # Remove common UI noise.
    value = re.sub(r'(?is)^[:\s\-–—]+', '', value)
    value = re.split(r'(?is)\s+(?:Статус сертификата|Статус декларации|Дата регистрации|Дата окончания|Номер бланка|Свободное распространение|Лицо, подписавшее|Происхождение продукции|Общие условия хранения|Код\s+ТН|ТН\s+ВЭД|Сведения об|Сведения о таможенном|Заявитель|Изготовитель)\b', value)[0]
    return clean_registry_value(value)


def _extract_exact_label_values_from_html(text: str, labels: List[str]) -> Dict[str, List[str]]:
    """Extract exact label values from rendered HTML/text, not by generic guesses."""
    out: Dict[str, List[str]] = {lab: [] for lab in labels}
    if not text:
        return out
    if BeautifulSoup:
        soup = BeautifulSoup(text or '', 'html.parser')
        for tag in soup(['script', 'style', 'noscript', 'svg']):
            tag.decompose()
        visible = soup.get_text('\n', strip=True)
        pairs = _table_pairs_exact(soup)
        wanted = {_normalize_fsa_label(lab): lab for lab in labels}
        for label, value in pairs:
            lab_key = wanted.get(_normalize_fsa_label(label))
            if lab_key:
                val = _fsa_value_clean(lab_key, value)
                if val:
                    out[lab_key].append(val)
    else:
        visible = re.sub(r'<[^>]+>', '\n', text or '')

    # Exact visible-text fallback. It is still exact-label based.
    for lab in labels:
        vals = _extract_same_line_or_next_exact(visible, lab, max_chars=8000 if 'продукц' in norm_text(lab) else 350)
        for v in vals:
            v = _fsa_value_clean(lab, v)
            if v:
                out[lab].append(v)

    for lab in list(out):
        out[lab] = _unique_keep_order(out[lab])
    return out


def parse_fsa_registry_strict(text: str, url: str = '') -> Tuple[str, str, str]:
    """FSA parser by exact fields only.

    Product is taken ONLY from:
      «Наименование (обозначение) продукции»
      or «Общее наименование продукции»
    Number is taken ONLY from:
      «Регистрационный номер сертификата» or
      «Регистрационный номер декларации о соответствии»
    """
    product_vals_map = _extract_exact_label_values_from_html(text, FSA_EXACT_PRODUCT_LABELS)
    product_vals = []
    for lab in FSA_EXACT_PRODUCT_LABELS:
        product_vals.extend(product_vals_map.get(lab, []))
    number_labels = _fsa_number_labels_for_url(url)
    number_vals_map = _extract_exact_label_values_from_html(text, number_labels)
    cert_candidates: List[str] = []
    for lab in number_labels:
        cert_candidates.extend(number_vals_map.get(lab, []))

    product_vals = _unique_keep_order([_trim_product_value(v) for v in product_vals if _looks_like_product_value(v)])
    cert_candidates = _unique_keep_order([v for v in cert_candidates if 4 <= len(v) <= 250])

    low = norm_text(url + ' ' + text[:2000])
    if '/rds/declaration/' in url.lower() or 'деклараци' in low:
        doc_type = 'декларация'
    elif '/rss/certificate/' in url.lower() or 'сертификат' in low:
        doc_type = 'сертификат'
    else:
        doc_type = ''

    return (cert_candidates[0] if cert_candidates else ''), ('; '.join(product_vals[:10]) if product_vals else ''), doc_type


FSA_EXTRACT_JS = r"""
(arg) => {
  const productLabels = arg.productLabels || [];
  const numberLabels = arg.numberLabels || [];
  const allLabels = productLabels.concat(numberLabels);
  const norm = s => (s || '').toString().toLowerCase().replace(/ё/g, 'е').replace(/\s+/g, ' ').trim().replace(/[ :.;\-–—]+$/g, '');
  const clean = s => (s || '').toString().replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const wanted = new Map(allLabels.map(l => [norm(l), l]));
  const out = {};
  allLabels.forEach(l => out[l] = []);

  function push(label, value) {
    value = clean(value);
    if (!value) return;
    const nl = norm(label);
    if (norm(value).startsWith(nl)) {
      // try to remove repeated label from flattened text
      const raw = clean(value);
      const idx = raw.toLowerCase().replace(/ё/g,'е').indexOf(label.toLowerCase().replace(/ё/g,'е'));
      if (idx === 0) value = clean(raw.slice(label.length));
    }
    value = clean(value.replace(/^[:\s\-–—]+/g, ''));
    if (value) out[label].push(value);
  }

  function extractFromRow(row, label) {
    let values = [];
    const textNodes = Array.from(row.querySelectorAll('[class*="info-row__text"], .info-row__text, [class*="row__text"], [class*="field-value"], [class*="value"]'));
    for (const n of textNodes) values.push(n.innerText || n.textContent || '');
    if (!values.join(' ').trim()) {
      const cells = Array.from(row.querySelectorAll('td, th'));
      if (cells.length >= 2 && norm(cells[0].innerText || cells[0].textContent) === norm(label)) {
        values.push(cells.slice(1).map(c => c.innerText || c.textContent || '').join(' '));
      }
    }
    if (!values.join(' ').trim()) {
      const clone = row.cloneNode(true);
      clone.querySelectorAll('[class*="info-row__header"], .info-row__header, [class*="row__header"], [class*="field-title"], th').forEach(x => x.remove());
      Array.from(clone.querySelectorAll('[title]')).forEach(x => { if (norm(x.getAttribute('title')) === norm(label)) x.remove(); });
      values.push(clone.innerText || clone.textContent || '');
    }
    return clean(values.join(' '));
  }

  function candidateContainers(el) {
    const out = [];
    let cur = el;
    for (let i = 0; i < 7 && cur; i++, cur = cur.parentElement) out.push(cur);
    return out;
  }

  function attrLabel(el) {
    for (const a of ['title','ng-reflect-title','data-title','aria-label']) {
      const v = el.getAttribute && el.getAttribute(a);
      if (v && wanted.has(norm(v))) return wanted.get(norm(v));
    }
    return null;
  }

  // 1) FSA Angular rows by title-like attribute. The attribute can be placed on
  // the header element, while the actual value is in a sibling in a parent row.
  for (const el of Array.from(document.querySelectorAll('[title], [ng-reflect-title], [data-title], [aria-label]'))) {
    const lab = attrLabel(el);
    if (!lab) continue;
    for (const row of candidateContainers(el)) {
      const val = extractFromRow(row, lab);
      if (val && norm(val) !== norm(lab)) { push(lab, val); break; }
    }
  }

  // 2) FSA header/text class layout.
  for (const row of Array.from(document.querySelectorAll('figis-card-info-row, [class*="info-row"], [class*="card-info-row"], tr'))) {
    let header = row.querySelector('[class*="info-row__header"], .info-row__header, th, td:first-child');
    if (!header) continue;
    const lab = wanted.get(norm(header.innerText || header.textContent));
    if (lab) push(lab, extractFromRow(row, lab));
  }

  // 3) Exact line fallback from visible text.
  const lines = (document.body ? (document.body.innerText || '') : '').split(/\n+/).map(clean).filter(Boolean);
  for (let i = 0; i < lines.length; i++) {
    const lab = wanted.get(norm(lines[i]));
    if (!lab) continue;
    let buf = [];
    for (let j = i + 1; j < Math.min(lines.length, i + 30); j++) {
      const n = norm(lines[j]);
      if (wanted.has(n)) break;
      if (/^(статус|дата|номер бланка|свободное распространение|лицо, подписавшее|происхождение продукции|общие условия|код тн|тн вэд|сведения об|заявитель|изготовитель)/i.test(lines[j])) break;
      buf.push(lines[j]);
      if (buf.join(' ').length > (productLabels.includes(lab) ? 8000 : 400)) break;
    }
    push(lab, buf.join(' '));
  }

  for (const k of Object.keys(out)) {
    const seen = new Set();
    out[k] = out[k].map(clean).filter(v => {
      const key = norm(v);
      if (!key || seen.has(key)) return false;
      seen.add(key); return true;
    });
  }
  return out;
}
"""


async def _fsa_wait_and_extract_exact_from_page(page, product_labels: List[str], number_labels: List[str], wait_ms: int) -> Dict[str, List[str]]:
    labels = product_labels + number_labels
    # Wait specifically for one of the exact labels, but do not fail hard: some
    # FSA pages render slowly and the extraction below may still catch content.
    try:
        await page.wait_for_function(
            "(labels) => { const n=s=>(s||'').toString().toLowerCase().replace(/ё/g,'е').replace(/\\s+/g,' ').trim(); const txt=document.body?document.body.innerText:''; return labels.some(l => n(txt).includes(n(l))); }",
            labels,
            timeout=max(1000, wait_ms),
        )
    except Exception:
        pass
    try:
        await page.wait_for_timeout(max(0, min(wait_ms, 3000)))
    except Exception:
        pass
    try:
        data = await page.evaluate(FSA_EXTRACT_JS, {'productLabels': product_labels, 'numberLabels': number_labels})
        return {k: [clean_registry_value(v) for v in vals if clean_registry_value(v)] for k, vals in (data or {}).items()}
    except Exception:
        try:
            content = await page.content()
            combined: Dict[str, List[str]] = {}
            for lab, vals in _extract_exact_label_values_from_html(content, labels).items():
                combined[lab] = vals
            return combined
        except Exception:
            return {lab: [] for lab in labels}




def parse_html_registry(url: str, text: str) -> Tuple[str, str, str]:
    h = hostname(url)
    if h in {'swis.trade.kg', 'trade.kg'}:
        cert, prod, typ = parse_swis_registry_strict(text)
        if prod or cert:
            return cert, prod, typ
    if h == 'pub.fsa.gov.ru':
        # For FSA do not use generic fields. Only the exact fields specified by
        # the user are valid.
        return parse_fsa_registry_strict(text, url)
    return _LEGACY_PARSE_HTML_REGISTRY(url, text)




# -----------------------------
# Professional comparison
# -----------------------------

GENERIC_STOPWORDS = {
    "для", "и", "или", "из", "с", "со", "в", "во", "на", "по", "при", "без", "под", "над", "от", "до", "а", "также",
    "товар", "товары", "продукция", "изделие", "изделия", "предмет", "предметы", "комплект", "комплекты", "торговой", "марки", "марка",
    "детский", "детская", "детские", "детское", "детей", "ребенок", "мальчиков", "девочек", "мальчика", "девочки",
    "размер", "рост", "цвет", "артикул", "модель", "тип", "включая", "том", "числе", "иные", "прочие",
}

CATEGORY_TERMS = {
    "clothing": {
        "одежда", "швейные", "швейная", "трикотажные", "трикотажная", "бельевые",
        "верхние", "верхняя", "легкая", "костюмные",
        "куртка", "куртки", "пальто", "полупальто", "ветровка", "ветровки", "плащ",
        "плащи", "комбинезон", "комбинезоны",
        "футболка", "футболки", "брюки", "штаны", "штанишки", "шорты", "джинсы",
        "легинсы", "леггинсы", "лосины", "бриджи", "велосипедки",
        "платье", "платья", "сарафан", "сарафаны", "юбка", "юбки",
        "толстовка", "толстовки", "худи", "свитшот", "свитшоты", "джемпер",
        "джемперы", "кофта", "кофты", "рубашка", "рубашки", "блузка", "блузки",
        "пижама", "пижамы", "боди", "лонгслив", "водолазка", "майка",
        "майки", "трусы", "белье", "носки", "колготки", "жилет", "жилеты",
        "пиджак", "пиджаки", "блейзер", "блейзеры", "поло", "топ", "топы",
        "сорочка", "сорочки", "распашонка", "распашонки", "ползунки", "слипы",
        "кардиган", "кардиганы", "болеро", "бомбер", "бомберы", "анорак",
        "пуховик", "пуховики", "шуба", "шубы", "дублёнка", "дубленка",
        # v41: расширение
        "комбинезончик", "песочник", "боди-комбинезон", "распашонки", "чепчик", "чепчики",
        "свитер", "свитера", "пуловер", "пуловеры", "жакет", "жакеты", "тренч",
        "халат", "халаты", "ночнушка", "сорочка ночная", "лиф", "корсет",
        "гольфы", "подследники", "леггинсы", "капри", "комбидрес",
        "купальник", "купальники", "плавки", "пляжные", "термобелье", "термобельё",
        "рейтузы", "штанишки", "юбка-шорты", "юбка-брюки", "жилетка", "безрукавка",
        "толстовки", "костюм", "костюмы", "спортивный костюм", "тренировочный",
    },
    "footwear": {
        "обувь", "кроссовки", "кеды", "ботинки", "сапоги", "туфли", "сандалии",
        "босоножки", "тапочки", "сланцы", "чешки",
        "сникеры", "мокасины", "лоферы", "балетки", "угги", "валенки", "дутики",
        # v41: расширение
        "полуботинки", "полусапоги", "ботильоны", "сапожки", "туфельки",
        "пинетки", "сабо", "шлепанцы", "шлёпанцы", "вьетнамки", "галоши",
        "резиновые сапоги", "слипоны", "эспадрильи", "топсайдеры", "челси",
        "дерби", "оксфорды", "монки", "берцы", "унты", "ботфорты", "мюли",
        "тапки", "пантолеты", "босоножки", "тимберленды",
    },
    "accessories": {
        # v41: НОВАЯ категория — аксессуары (раньше их не было совсем)
        "шапка", "шапки", "шапочка", "панама", "панамка", "кепка", "кепки",
        "бейсболка", "шляпа", "шляпы", "берет", "косынка", "бандана",
        "шарф", "шарфы", "снуд", "платок", "палантин", "варежки", "перчатки",
        "митенки", "рукавицы", "ремень", "ремни", "пояс", "галстук", "бабочка",
        "сумка", "сумки", "сумочка", "рюкзак", "рюкзаки", "ранец", "портфель",
        "кошелёк", "кошелек", "клатч", "барсетка", "несессер", "очки",
        "зонт", "зонты", "повязка", "ободок", "заколка", "резинка для волос",
        # v41.2: обобщающие формулировки из сертификатов
        "головные уборы", "головной убор", "галантерея", "аксессуар", "аксессуары",
    },
    "home_textile": {
        # v41: НОВАЯ категория — домашний текстиль
        "постельное", "постельное белье", "постельные принадлежности", "простыня", "простынь", "пододеяльник",
        "наволочка", "наволочки", "одеяло", "одеяла", "подушка", "подушки",
        "плед", "пледы", "покрывало", "покрывала", "полотенце", "полотенца",
        "скатерть", "салфетка", "штора", "шторы", "занавеска", "тюль",
        "матрас", "матрац", "наматрасник", "конверт для новорожденного",
        "спальный мешок", "бортик", "балдахин", "пеленка", "пелёнка", "пеленки",
        # v41.2
        "постельн", "домашний текстиль", "текстиль постельный",
    },
    "toys": {"игрушка", "игрушки", "кукла", "куклы", "конструктор", "конструкторы",
             "машинка", "машинки", "пазл", "пазлы", "мяч", "мячи", "погремушка",
             "погремушки", "фигурка", "фигурки",
             # v41
             "мягкая игрушка", "плюшевая", "неваляшка", "юла", "пирамидка",
             "сортер", "кубики", "мозаика", "лего", "робот", "вертолет",
             "самолет", "грузовик", "трактор", "паровозик", "железная дорога",
             "настольная игра", "лото", "домино", "развивающая", "каталка",
             "качалка", "интерактивная игрушка", "слайм", "антистресс", "спиннер"},
    "cosmetics": {"косметика", "крем", "кремы", "шампунь", "шампуни", "гель",
                  "гели", "мыло", "лосьон", "лосьоны", "зубная паста",
                  # v41
                  "пенка", "пена для", "бальзам", "масло детское", "присыпка",
                  "влажные салфетки", "подгузник", "подгузники", "крем под подгузник",
                  "пудра", "молочко", "тоник", "маска", "скраб", "дезодорант"},
    "food": {"пюре", "смесь", "смеси", "каша", "каши", "напиток", "напитки",
             "печенье", "конфеты", "шоколад", "молоко", "чай",
             # v41
             "сок", "соки", "вода детская", "пюрешка", "снек", "батончик",
             "йогурт", "творожок", "кисель", "компот", "морс"},
    "feeding": {
        # v41: НОВАЯ категория — товары для кормления
        "бутылочка", "соска", "пустышка", "ниблер", "поильник", "слюнявчик",
        "нагрудник", "тарелка детская", "ложка детская", "контейнер для",
        "молокоотсос", "стерилизатор", "подогреватель",
        # v41.2
        "для кормления", "посуда детская", "детская посуда",
    },
    "nursery": {
        # v41: НОВАЯ категория — детские товары/мебель/транспорт
        "коляска", "коляски", "автокресло", "автолюлька", "кроватка", "манеж",
        "ходунки", "прыгунки", "шезлонг", "качели", "стульчик для кормления",
        "пеленальный", "горшок", "комод пеленальный", "матрасик", "мобиль",
        "радионяня", "видеоняня", "переноска", "слинг", "кенгуру", "эрго-рюкзак",
    },
    # v27.7: НОВЫЕ категории — расширение на любые товарные группы (не только лёгкая промышленность).
    "electronics": {
        "электроника", "наушники", "наушник", "гарнитура", "колонка", "колонки",
        "акустика", "акустическая система", "акустическ", "саундбар", "смартфон", "смартфоны", "телефон",
        "телефоны", "телефоны мобильные", "планшет", "планшеты", "ноутбук",
        "ноутбуки", "монитор", "мониторы", "клавиатура", "мышь компьютерная",
        "мышь", "мышка", "мышки", "устройство ввода", "устройства ввода",
        "манипулятор", "манипулятор компьютерный",
        "телевизор", "телевизоры", "повербанк", "повербанки", "пауэрбанк",
        "power bank", "powerbank", "внешний аккумулятор", "батареи аккумуляторные",
        "аккумуляторная батарея",
        "зарядное устройство", "зарядные устройства", "зарядное", "зарядка", "зарядки",
        "сетевое зарядное", "блок питания", "адаптер питания",
        "кабель", "кабели", "адаптер", "адаптеры", "переходник", "переходники",
        "штекер", "штекеры", "коннектор", "разъём", "разъем", "usb", "type-c", "hdmi", "otg",
        "роутер", "маршрутизатор", "флешка", "флеш-накопитель", "карта памяти",
        "фотоаппарат", "видеокамера", "веб-камера", "магнитола", "радиоприёмник",
        "проектор", "приставка игровая", "геймпад", "джойстик", "умные часы",
        "фитнес-браслет", "электронная книга", "модем", "микрофон",
        "аппаратура радиоэлектронная", "радиоэлектронная бытовая", "устройства связи",
        "средства связи", "техника электронная",
    },
    "appliances": {
        "бытовая техника", "приборы бытовые", "прибор бытовой", "электроприбор",
        "электроприборы", "приборы бытовые электрические", "машины электрические бытовые",
        "фен", "фены", "пылесос", "пылесосы", "утюг", "утюги", "миксер", "блендер",
        "тостер", "кофеварка", "кофемашина", "кофемолка", "мультиварка", "пароварка",
        "микроволновая печь", "микроволновка", "свч-печь", "холодильник", "морозильник",
        "стиральная машина", "посудомоечная машина", "сушильная машина",
        "электроплита", "варочная панель", "духовой шкаф", "вытяжка",
        "вентилятор", "обогреватель", "конвектор", "тепловентилятор", "увлажнитель",
        "очиститель воздуха", "водонагреватель", "бойлер", "мясорубка электрическая",
        "соковыжималка", "электрочайник", "термопот", "весы кухонные", "весы напольные",
        "эпилятор", "машинка для стрижки", "выпрямитель для волос", "плойка",
    },
    "kitchenware": {
        "посуда", "посуда кухонная", "посуда столовая", "сковорода", "сковородка",
        "кастрюля", "кастрюли", "ковш", "сотейник", "казан", "противень",
        "чайник заварочный", "заварочный чайник", "турка", "нож кухонный",
        "ножи кухонные", "ножи", "изделия ножевые", "вилка", "вилки", "ложка",
        "ложки", "тарелка", "тарелки", "кружка", "кружки", "стакан", "стаканы",
        "чашка", "чашки", "бокал", "бокалы", "миска", "миски", "салатник",
        "форма для выпечки", "разделочная доска", "доска разделочная", "половник",
        "дуршлаг", "тёрка", "терка", "шумовка", "лопатка кухонная", "столовые приборы",
        "приборы столовые", "контейнер пищевой", "термос", "термокружка", "графин",
    },
    "tools": {
        "инструмент", "инструменты", "инструмент электрический", "инструмент ручной",
        "инструмент слесарно-монтажный", "электроинструмент", "дрель", "дрели",
        "шуруповёрт", "шуруповерт", "отвёртка", "отвертка", "отвёртки", "молоток",
        "молотки", "пила", "ножовка", "плоскогубцы", "пассатижи", "кусачки",
        "гаечный ключ", "ключ гаечный", "болгарка", "углошлифовальная машина",
        "перфоратор", "лобзик", "рубанок", "стамеска", "напильник", "тиски",
        "струбцина", "шлифмашина", "степлер строительный", "уровень строительный",
        "рулетка измерительная", "набор инструментов", "сверло", "свёрла", "бур",
        "паяльник", "точило", "стусло", "клещи",
    },
    "auto": {
        "автотовары", "автомобильные принадлежности", "автомобильный", "автозапчасти",
        "запчасти", "масло моторное", "моторное масло", "трансмиссионное масло",
        "антифриз", "тосол", "омыватель", "стеклоомыватель", "щётки стеклоочистителя",
        "стеклоочистители", "дворники", "тормозные колодки", "колодки тормозные",
        "свечи зажигания", "фильтр масляный", "фильтр воздушный", "фильтр салонный",
        "шины", "покрышки", "автошины", "диски колёсные", "автохимия",
        "автомобильный аккумулятор", "автокосметика", "видеорегистратор",
        "компрессор автомобильный", "домкрат", "буксировочный трос",
    },
    "sport_equipment": {
        "спортивный инвентарь", "инвентарь спортивный", "спорттовары", "тренажёр",
        "тренажер", "велотренажёр", "гантели", "гантель", "гиря", "гири", "штанга",
        "блины для штанги", "скакалка", "эспандер", "коврик для йоги", "фитбол",
        "ракетка", "ракетки", "клюшка", "шлем спортивный", "защита спортивная",
        "ролики", "роликовые коньки", "коньки", "скейтборд", "самокат",
        "велосипед", "велосипеды", "палатка", "спальный мешок туристический",
        "турник", "брусья", "степ-платформа", "боксёрские перчатки", "груша боксёрская",
    },
    "stationery": {
        "канцелярские товары", "канцелярские принадлежности", "принадлежности канцелярские",
        "канцтовары", "ручка шариковая", "ручка", "ручки", "карандаш", "карандаши",
        "тетрадь", "тетради", "блокнот", "блокноты", "маркер", "маркеры", "фломастер",
        "фломастеры", "ластик", "линейка", "степлер", "скрепки", "клей канцелярский",
        "ножницы", "папка", "папки", "бумага офисная", "бумага для печати", "пенал",
        "дневник", "альбом для рисования", "краски", "пластилин", "точилка",
        "корректор", "кнопки канцелярские", "скотч", "обложка для книг",
    },
    "pet": {
        "зоотовары", "товары для животных", "для животных", "для кошек", "для собак",
        "корм для", "корма для", "корм для кошек", "корм для собак",
        "корма для непродуктивных животных", "наполнитель для туалета", "наполнитель",
        "поводок", "ошейник", "намордник", "лоток для животных", "когтеточка",
        "аквариум", "клетка для", "переноска для животных", "лежанка для",
        "игрушка для собак", "лакомство для", "шлейка",
    },
    "household_chemistry": {
        "бытовая химия", "средства моющие", "моющее средство", "средства чистящие",
        "чистящее средство", "стиральный порошок", "порошок стиральный", "гель для стирки",
        "капсулы для стирки", "кондиционер для белья", "ополаскиватель для белья",
        "отбеливатель", "пятновыводитель", "средство для мытья посуды",
        "средство для мытья", "средство для уборки", "освежитель воздуха",
        "средство чистящее", "антинакипин", "средство для стёкол", "дезинфицирующее средство",
        "мыло хозяйственное", "сода кальцинированная",
    },
    "furniture": {
        "мебель", "мебель бытовая", "мебель корпусная", "стол", "столы", "стул",
        "стулья", "кресло", "кресла", "диван", "диваны", "шкаф", "шкафы", "комод",
        "комоды", "полка", "полки", "стеллаж", "стеллажи", "тумба", "тумбочка",
        "табурет", "табуретка", "вешалка напольная", "столешница", "кухонный гарнитур",
        "прихожая", "пуф", "банкетка", "софа", "кушетка", "этажерка",
    },
    "plumbing": {
        "сантехника", "сантехнический", "смеситель", "смесители", "кран водопроводный",
        "раковина", "умывальник", "унитаз", "ванна", "душевая кабина", "душевая лейка",
        "шланг для душа", "сифон", "мойка кухонная", "инсталляция", "полотенцесушитель",
        "сливной бачок", "биде",
    },
    "lighting": {
        "освещение", "светильник", "светильники", "лампа", "лампы", "лампочка",
        "светодиодная лампа", "led-лампа", "люстра", "люстры", "бра", "торшер",
        "ночник", "прожектор", "гирлянда", "лента светодиодная", "настольная лампа",
        "фонарь", "фонарик",
    },
    "jewelry": {
        "ювелирные изделия", "украшение", "украшения", "бижутерия", "кольцо",
        "кольца", "серьги", "серёжки", "цепочка", "цепочки", "браслет", "браслеты",
        "кулон", "подвеска", "брошь", "колье", "пирсинг", "запонки",
    },
}

# v41: пол — отдельный маркер (раньше только child/adult)
GENDER_MALE_WORDS = ("мужск", "мужчин", "для него", "для мужчин", "для мальчик", "мальчиков", "мальчику")
GENDER_FEMALE_WORDS = ("женск", "женщин", "для нее", "для неё", "для женщин", "для девоч", "девочек", "девочке")

SUBTYPE_SYNONYMS = {
    "куртка": {"куртка", "куртки", "ветровка", "ветровки", "парка", "парки", "штормовка", "анорак"},
    "пальто": {"пальто", "полупальто", "плащ", "плащи", "накидка", "накидки", "тренч"},
    "комбинезон": {"комбинезон", "комбинезоны", "полукомбинезон", "полукомбинезоны", "комбинезончик", "песочник"},
    "футболка": {"футболка", "футболки", "майка", "майки", "топ", "топы", "поло"},
    "брюки": {"брюки", "штаны", "штанишки", "рейтузы", "капри"},  # v41: джинсы вынесены отдельно
    "джинсы": {"джинсы", "джинсовые"},  # v41: отдельный подтип
    "леггинсы": {"легинсы", "леггинсы", "лосины", "тайтсы"},  # v41: отдельный подтип
    "шорты": {"шорты", "бриджи", "велосипедки"},
    "платье": {"платье", "платья", "сарафан", "сарафаны", "платьице"},
    "юбка": {"юбка", "юбки", "юбочка"},
    "толстовка": {"толстовка", "толстовки", "худи", "свитшот", "свитшоты", "кофта", "кофты", "джемпер", "джемперы", "лонгслив", "лонгсливы", "свитер", "свитера", "водолазка", "водолазки", "пуловер", "пуловеры", "кардиган", "кардиганы"},
    "рубашка": {"рубашка", "рубашки", "блузка", "блузки", "сорочка", "сорочки", "блуза"},
    "пижама": {"пижама", "пижамы", "сорочка ночная", "ночнушка"},
    "белье": {"белье", "бельё", "трусы", "трусики", "майка", "майки", "бюстгальтер", "лиф", "плавки"},
    "носки": {"носки", "гольфы", "подследники", "следки"},
    "колготки": {"колготки", "колготы"},
    "купальник": {"купальник", "купальники", "плавки", "купальный"},  # v41
    "кроссовки": {"кроссовки", "кеды", "сникеры", "слипоны", "кроссовочки"},
    "ботинки": {"ботинки", "сапоги", "полусапоги", "полуботинки", "ботильоны", "сапожки", "берцы", "челси", "дерби", "ботиночки"},
    "сандалии": {"сандалии", "босоножки", "сандалики"},
    "туфли": {"туфли", "туфельки", "балетки", "лоферы", "мокасины", "оксфорды"},  # v41
    "тапочки": {"тапочки", "тапки", "сабо", "шлепанцы", "шлёпанцы", "пантолеты"},  # v41
    "костюм": {"костюм", "костюмы", "костюмные", "комплект", "комплекты"},
    # v41: подтипы аксессуаров
    "шапка": {"шапка", "шапки", "шапочка", "берет", "колпак"},
    "панама": {"панама", "панамка", "кепка", "кепки", "бейсболка", "шляпа", "шляпы"},
    "шарф": {"шарф", "шарфы", "снуд", "платок", "палантин", "косынка", "бандана"},
    "перчатки": {"перчатки", "варежки", "митенки", "рукавицы"},
    "сумка": {"сумка", "сумки", "сумочка", "клатч", "барсетка"},
    "рюкзак": {"рюкзак", "рюкзаки", "ранец", "портфель"},
    "ремень": {"ремень", "ремни", "пояс"},
    # v41: подтипы текстиля
    "постельное": {"постельное", "простыня", "простынь", "пододеяльник", "наволочка", "наволочки", "кпб"},
    "одеяло": {"одеяло", "одеяла", "плед", "пледы", "покрывало", "покрывала"},
    "подушка": {"подушка", "подушки"},
    "полотенце": {"полотенце", "полотенца"},
    "пеленка": {"пеленка", "пелёнка", "пеленки", "пелёнки"},
}

# v41: маппинг подтип → категория (для проверки конфликтов внутри категории)
SUBTYPE_TO_CATEGORY = {
    "куртка": "clothing", "пальто": "clothing", "комбинезон": "clothing",
    "футболка": "clothing", "брюки": "clothing", "джинсы": "clothing",
    "леггинсы": "clothing", "шорты": "clothing", "платье": "clothing",
    "юбка": "clothing", "толстовка": "clothing", "рубашка": "clothing",
    "пижама": "clothing", "белье": "clothing", "носки": "clothing",
    "колготки": "clothing", "купальник": "clothing", "костюм": "clothing",
    "кроссовки": "footwear", "ботинки": "footwear", "сандалии": "footwear",
    "туфли": "footwear", "тапочки": "footwear",
    "шапка": "accessories", "панама": "accessories", "шарф": "accessories",
    "перчатки": "accessories", "сумка": "accessories", "рюкзак": "accessories",
    "ремень": "accessories",
    "постельное": "home_textile", "одеяло": "home_textile", "подушка": "home_textile",
    "полотенце": "home_textile", "пеленка": "home_textile",
}

BROAD_CLOTHING_PHRASES = (
    "изделия швейные", "одежда", "верхние швейные", "изделия верхние", "изделия трикотажные",
    "бельевые изделия", "изделия бельевые", "одежда верхняя", "одежда трикотажная", "изделия костюмные",
)


def _tokens_for_compare(text: str) -> Set[str]:
    t = norm_text(text)
    toks = re.findall(r"[а-яa-z0-9]{2,}", t)
    out = set()
    for tok in toks:
        if tok in GENERIC_STOPWORDS:
            continue
        # very light stemming for Russian plurals/adjectives, enough for category matching.
        stem = re.sub(r"(ами|ями|ого|ему|ими|ыми|ой|ая|ое|ые|ий|ый|ых|их|ам|ям|ах|ях|ов|ев|ом|ем|ою|ею|а|я|ы|и|е|о|у|ю)$", "", tok)
        out.add(tok)
        if len(stem) >= 4:
            out.add(stem)
    return out


def _stem_token(tok: str) -> str:
    """Лёгкий стемминг русских окончаний (как в _tokens_for_compare).
    v27.7: добавлено отсечение голой 'о' (масло→масл, чтобы совпадало с масла→масл)."""
    return re.sub(r"(ами|ями|ого|ему|ими|ыми|ой|ая|ое|ые|ий|ый|ых|их|ам|ям|ах|ях|ов|ев|ом|ем|ою|ею|а|я|ы|и|е|о|у|ю)$", "", tok)


def _stem_set(text_low: str) -> Set[str]:
    """Множество стем-токенов текста (для сопоставления многословных терминов
    без учёта порядка слов и окончаний)."""
    out: Set[str] = set()
    for tok in re.findall(r"[а-яa-z0-9]{2,}", text_low):
        out.add(tok)
        st = _stem_token(tok)
        if len(st) >= 3:
            out.add(st)
    return out


def _detect_categories(text: str) -> Set[str]:
    low = norm_text(text)
    text_stems = _stem_set(low)
    cats = set()
    for cat, terms in CATEGORY_TERMS.items():
        for term in terms:
            tl = norm_text(term)
            if " " in tl or "-" in tl:
                # v27.7: многословный термин — совпадение по стемам в ЛЮБОМ
                # порядке (cert «Масла моторные» матчит термин «масло моторное»,
                # «Приборы электрические бытовые» — «приборы бытовые электрические»).
                term_stems = {_stem_token(t) for t in re.findall(r"[а-яa-z0-9]{2,}", tl)}
                term_stems = {t for t in term_stems if len(t) >= 3}
                if term_stems and term_stems <= text_stems:
                    cats.add(cat)
                    break
            else:
                if _contains_word_token(low, tl):
                    cats.add(cat)
                    break
    return cats


# v46: «родственные» товарные группы — категории, которые в реестрах ФСА часто
# смешаны и НЕ должны считаться конфликтом. Бытовая электроника и «приборы
# бытового назначения» (appliances) — одна предметная область: реестровый
# сертификат «Аппараты электрические бытового назначения: сетевые зарядные
# устройства/акустические системы» (appliances по префиксу) покрывает WB-карточку
# «зарядка/колонка» (electronics). Раньше это давало ложное «НЕСООТВЕТСТВИЕ».
RELATED_CATEGORY_GROUPS: List[frozenset] = [
    frozenset({"electronics", "appliances"}),
]


def _categories_related(a: Set[str], b: Set[str]) -> bool:
    """True, если категории из a и b лежат в одной «родственной» группе
    (например electronics и appliances) — тогда несовпадение категорий НЕ конфликт."""
    if not a or not b:
        return False
    for g in RELATED_CATEGORY_GROUPS:
        if (a & g) and (b & g):
            return True
    return False


def _contains_word_token(text_low: str, term_low: str) -> bool:
    """v39.3: проверяет содержит ли text term'om по границе слова.
    «следа» НЕ содержит «еда» (граница слов нарушена).
    «электронная коммерция» НЕ содержит «электр» если требуется word boundary,
    но содержит — потому что начинается со слова. Этот случай лечится через
    очистку названий фирм/брендов в _strip_brand_marks.

    v39.8: regex pattern кэшируется через _word_boundary_pattern (раньше
    re.compile вызывался на каждом из тысяч вызовов).
    """
    if not term_low or not text_low:
        return False
    return bool(_word_boundary_pattern(term_low).search(text_low))


# v39.8: предкомпилированные паттерны для _strip_brand_marks (вызывается на КАЖДОМ сравнении)
_BRAND_CUT_PATTERNS = [
    re.compile(p, re.IGNORECASE) for p in (
        r'\bс\s+маркировк[аиоей]+\b',
        r'\bпод\s+товарны[мх]+\s+знак[оа]м\b',
        r'\bс\s+товарны[мх]+\s+знак[аоа]м[иа]?\b',
        r'\bторгов[аойых]+\s+марк[аиойх]+\b',
        r'\bООО\s+["«]',
        r'\bИП\s+[А-ЯЁ]',
    )
]


def _strip_brand_marks(text: str) -> str:
    """v39.3: вырезает из cert названия фирм/брендов, чтобы они не влияли на категоризацию.
    Примеры что режется:
      «с маркировкой ООО "МИМЭН ЭЛЕКТРОННАЯ КОММЕРЦИЯ"» → ничего после «с маркировкой»
      «с товарными знаками "ART-CHARGE"» → ничего после «знаками»
    Это убирает false positive: слово «ЭЛЕКТРОННАЯ» в названии фирмы попадает в electronics.
    v39.8: используются предкомпилированные паттерны.
    """
    if not text:
        return text
    earliest = len(text)
    for pat in _BRAND_CUT_PATTERNS:
        m = pat.search(text)
        if m and m.start() < earliest:
            earliest = m.start()
    return text[:earliest] if earliest < len(text) else text


def _detect_subtypes(text: str) -> Set[str]:
    low = norm_text(text)
    out = set()
    for subtype, terms in SUBTYPE_SYNONYMS.items():
        for term in terms:
            if _contains_word_token(low, norm_text(term)):
                out.add(subtype)
                break
    return out



# Extend old subtype/age dictionaries without touching link collection.
# v27.9.x: убрали синоним «слип» — токен-матчер сопоставляет по началу слова
# (стем), поэтому «слип» совпадал со «слипоны»/«слипон» (ОБУВЬ) и тапочки-слипоны
# ложно классифицировались как бельё → ложный OK против бельевых сертификатов.
# Детский слип-комбинезон ловится по «боди/ползунки/распашонка/комбинезон».
SUBTYPE_SYNONYMS.setdefault("белье", set()).update({"боди", "распашонка", "распашонки", "ползунки", "ползунок", "фуфайка", "фуфайки", "лонгслив"})

def _age_marker(text: str) -> str:
    low = norm_text(text)
    child_words = ("детск", "для детей", "для мальчик", "для девоч", "ясель", "дошколь", "школь", "подрост", "новорожден", "малыш", "младен", "baby", "kids", "children")
    adult_words = ("для взрослых", "женск", "мужск")
    has_child = any(w in low for w in child_words)
    has_adult = any(w in low for w in adult_words)
    if has_child:
        return "child"
    if has_adult:
        return "adult"
    return "unknown"


def _gender_marker(text: str) -> str:
    """v41: определяет пол товара — male / female / unisex / unknown.
    Помогает ловить конфликты «мужское» vs «женское» и не мешает когда пол не указан."""
    low = norm_text(text)
    has_m = any(w in low for w in GENDER_MALE_WORDS)
    has_f = any(w in low for w in GENDER_FEMALE_WORDS)
    if has_m and has_f:
        return "unisex"
    if has_m:
        return "male"
    if has_f:
        return "female"
    return "unknown"


# v41: подтипы которые НЕ должны путаться внутри одной категории (явный конфликт вида).
# Если карточка про «платье», а сертификат строго про «брюки» (и ничего общего),
# это несоответствие даже при совпадении категории clothing.
_INCOMPATIBLE_SUBTYPES = [
    ({"платье"}, {"брюки", "джинсы", "шорты", "куртка", "пальто"}),
    ({"юбка"}, {"брюки", "джинсы", "куртка", "пальто"}),
    ({"купальник"}, {"куртка", "пальто", "брюки", "джинсы", "толстовка"}),
    ({"белье"}, {"куртка", "пальто", "джинсы"}),
    # обувь vs обувь обычно совместима (общий сертификат), не конфликтуем
]

# Broader but conservative ontology for children's apparel certificates.
CLOTHING_LAYER_TERMS = {
    'first_layer': {
        'бельев', 'первого слоя', '1 слоя', 'нательное', 'боди', 'пижам', 'футбол',
        'фуфайк', 'майк', 'трус', 'сорочк', 'комбинезон бельев', 'ползунк',
        'распашонк', 'лонгслив',
    },
    'second_layer': {
        'второго слоя', '2 слоя', 'плать', 'сарафан', 'юбк', 'брюк', 'штан',
        'джинс', 'шорт', 'блуз', 'рубаш', 'джемпер', 'свитер', 'толстовк',
        'худи', 'костюм', 'жакет', 'свитшот', 'кардиган', 'водолазк', 'лосин',
        'леггинс', 'бриджи',
    },
    'third_layer': {
        'третьего слоя', '3 слоя', 'верхн', 'куртк', 'пальто', 'полупальто',
        'ветровк', 'плащ', 'комбинезон утепл', 'полукомбинезон', 'жилет утепл',
        'парка', 'пуховик', 'шуба', 'дублёнк', 'дубленк', 'анорак',
    },
    'hosiery': {'носк', 'гольф', 'колгот', 'чулоч'},
    'headwear': {'шапк', 'панам', 'кепк', 'шляп', 'головн'},
}

PRODUCT_CONFLICT_TERMS = {
    'toys': {'игруш', 'кукл', 'конструктор', 'мяч', 'пазл'},
    'cosmetics': {'космет', 'крем', 'шампун', 'мыло', 'гель душ', 'дезодорант'},
    # v39.3: удлинены стемы — «еда» матчилась в «без следа», «крем» в «кремовый».
    # Теперь требуют конкретные пищевые слова.
    'food': {'пищев', 'продукт пита', 'детское пита', 'пюре', 'смесь молоч', 'каша', 'напиток', 'печенье'},
    # v39.3: «электр» матчился в названиях фирм («ЭЛЕКТРОННАЯ КОММЕРЦИЯ»).
    # Удлинены до конкретных электротоваров.
    'electronics': {'электроприбор', 'электротехнич', 'электронн прибор', 'аппарат для', 'прибор бытов', 'кабель', 'зарядное устройств', 'наушник', 'смартфон', 'радиоэлектронная'},
    'furniture': {'мебель', 'кроват', 'стул', 'шкаф', 'комод', 'диван', 'кресло'},
    # v27.7: новые конфликтные домены — чтобы карточка одежды/обуви явно
    # конфликтовала с сертификатом из чужой группы (и наоборот).
    'tools': {'инструмент электрич', 'инструмент ручн', 'инструмент слесарн', 'дрель', 'шуруповёрт', 'шуруповерт', 'перфоратор', 'болгарк'},
    'auto': {'моторное масло', 'масло моторн', 'антифриз', 'стеклоочистител', 'тормозные колодк', 'свечи зажиган', 'автошин'},
    'kitchenware': {'посуда кухонн', 'посуда столов', 'изделия ножев', 'сковород', 'кастрюл'},
    'household_chemistry': {'средства моющие', 'моющее средств', 'чистящее средств', 'стиральный порош', 'порошок стиральн', 'гель для стирк'},
    'plumbing': {'смесител', 'сантехник', 'унитаз', 'умывальник'},
    'sport_equipment': {'спортивный инвентар', 'инвентарь спортивн', 'тренажёр', 'тренажер', 'гантел'},
    'stationery': {'канцелярск', 'канцтовар'},
    'pet': {'зоотовар', 'корм для', 'корма для', 'наполнитель для туалет'},
    'lighting': {'светильник', 'светодиодная лампа', 'люстр'},
    'jewelry': {'ювелирные издел', 'бижутер'},
}


def _detect_layers(text: str) -> Set[str]:
    low = norm_text(text)
    out: Set[str] = set()
    for layer, terms in CLOTHING_LAYER_TERMS.items():
        if any(_contains_word_token(low, norm_text(t)) for t in terms):
            out.add(layer)
    return out


def _detect_hard_conflict_domain(text: str) -> Set[str]:
    """v39.3: используем word-boundary matching + предварительно вырезаем
    названия фирм/брендов из текста — иначе false positives."""
    cleaned = _strip_brand_marks(text)
    low = norm_text(cleaned)
    out: Set[str] = set()
    for name, terms in PRODUCT_CONFLICT_TERMS.items():
        if any(_contains_word_token(low, norm_text(t)) for t in terms):
            out.add(name)
    return out


def _seq_ratio(a: str, b: str) -> float:
    return 100.0 * difflib.SequenceMatcher(None, norm_text(a), norm_text(b)).ratio()


def _contains_broad_child_clothing(cert_text: str) -> bool:
    """True только для ДЕЙСТВИТЕЛЬНО широких сертификатов на детскую одежду
    («Изделия швейные/трикотажные для детей», «Одежда детская»).

    v27.9.x: сертификат со СПЕЦИФИЧНЫМ слоем/типом («первого слоя», «бельевые»,
    «верхняя одежда» и т.п.) — НЕ широкий: он покрывает только свой слой, и
    приклеивать к нему другие виды одежды (костюм, платье) нельзя. Раньше слово
    «бельевые» считалось «широким» и давало ложный OK для костюма против
    бельевого сертификата. По решению пользователя такие случаи → ПРОВЕРИТЬ ВРУЧНУЮ.
    """
    low = norm_text(cert_text)
    specific = ('первого слоя', 'второго слоя', 'третьего слоя', 'верхнего слоя',
                'бельев', 'белье', 'верхняя одежда', 'одежда верхняя', 'изделия верхние')
    if any(x in low for x in specific):
        return False
    broad = ('изделия швейные', 'изделия трикотажные', 'одежда')
    child = ('дет', 'подрост', 'дошколь', 'школь', 'новорожден', 'ясель')
    return any(x in low for x in broad) and any(x in low for x in child)


def compare_product_names(product_name: str, cert_product_name: str, brand: str = '', subject: str = '', doc_status: str = '') -> Tuple[str, float, str]:
    """Conservative professional comparison.

    The goal is not to force a verdict for every row, but to separate:
      OK when certificate clearly covers the WB card;
      НЕСООТВЕТСТВИЕ when there is a clear product/category conflict;
      ПРОВЕРИТЬ ВРУЧНУЮ when the certificate is too broad or ambiguous.

    v39.5: дополнительно учитываем статус документа из реестра. Если документ
    не «Действует» (Прекращён, Приостановлен, Архивный, Аннулирован) — это
    самостоятельный вердикт «НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ» с приоритетом выше товарного
    сравнения. Карточке WB нельзя приклеивать к недействующему сертификату.
    """
    # v39.5: первичная проверка статуса документа. Документ должен быть «Действует».
    status_norm = _normalize_doc_status(doc_status) if doc_status else ''
    if status_norm and status_norm.lower().replace('ё','е') not in ('действует', ''):
        # Документ есть и его статус известен и НЕ «Действует» — это серьёзный сигнал.
        return 'НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ', 0.0, f'doc_status={status_norm}; документ не действует, проверка соответствия товара не выполняется'

    if not cert_product_name:
        return 'НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА', 0.0, 'В реестре не извлечено поле продукции'

    card_text = clean_registry_value(' '.join(x for x in [product_name, subject] if x))
    cert_text = clean_registry_value(cert_product_name)
    card_low = norm_text(card_text)
    cert_low = norm_text(cert_text)

    card_cats = _detect_categories(card_text)
    cert_cats = _detect_categories(cert_text)
    card_sub = _detect_subtypes(card_text)
    cert_sub = _detect_subtypes(cert_text)
    card_layers = _detect_layers(card_text)
    cert_layers = _detect_layers(cert_text)
    card_age = _age_marker(card_text)
    cert_age = _age_marker(cert_text)
    card_gender = _gender_marker(card_text)  # v41
    cert_gender = _gender_marker(cert_text)  # v41
    card_conflict_domains = _detect_hard_conflict_domain(card_text)
    cert_conflict_domains = _detect_hard_conflict_domain(cert_text)

    conflicts: List[str] = []
    soft_conflicts: List[str] = []  # v27.9.x: «мягкие» (несовпадение слоя) -> ПРОВЕРИТЬ ВРУЧНУЮ
    # Hard domain conflicts: e.g. WB clothing card vs toy/electronics certificate.
    if 'clothing' in card_cats and cert_conflict_domains:
        conflicts.append(f'сертификат относится к другой группе: {sorted(cert_conflict_domains)}')
    if (card_cats and cert_cats and card_cats.isdisjoint(cert_cats)
            and not _categories_related(card_cats, cert_cats)):
        conflicts.append(f'категория карточки {sorted(card_cats)} не совпадает с категорией сертификата {sorted(cert_cats)}')
    if card_age == 'child' and cert_age == 'adult':
        conflicts.append('карточка детская, а сертификат явно для взрослых')
    # v41.1: конфликт пола — только когда товарный вид совпадает/пересекается.
    # Иначе «костюм для мальчика» vs «бельё для девочек» — это конфликт вида/слоя,
    # а не пола, и пол лишь добавляет шум. Конфликт пола осмыслен для ОДНОГО вида:
    # «трусы для мальчиков» vs «трусы для девочек».
    if (card_gender in ('male', 'female') and cert_gender in ('male', 'female')
            and card_gender != cert_gender and (card_sub & cert_sub)):
        conflicts.append(f'пол не совпадает при совпадающем виде: карточка {card_gender}, сертификат {cert_gender}')
    # v41.1: конфликт несовместимых подтипов — ТОЛЬКО когда ОБЕ стороны узкие
    # (не более 2 подтипов каждая). Если сертификат перечисляет много видов (широкий) —
    # это покрывающий сертификат, конфликт вида не применяем (сработает конфликт слоя если нужно).
    if (card_sub and cert_sub and not (card_sub & cert_sub)
            and len(cert_sub) <= 2 and len(card_sub) <= 2
            and not _contains_broad_child_clothing(cert_text)):
        for set_a, set_b in _INCOMPATIBLE_SUBTYPES:
            if (card_sub & set_a and cert_sub & set_b and not (cert_sub & set_a)) or \
               (card_sub & set_b and cert_sub & set_a and not (cert_sub & set_b)):
                conflicts.append(f'несовместимый вид: карточка {sorted(card_sub)}, сертификат {sorted(cert_sub)}')
                break
    # Layer conflict is meaningful only when both sides are apparel and both layers known.
    if 'clothing' in card_cats and 'clothing' in cert_cats and card_layers and cert_layers and card_layers.isdisjoint(cert_layers):
        # Allow if exact subtype still overlaps (e.g. комбинезон can be different layer depending on context).
        if not (card_sub & cert_sub):
            # v27.9.x: несовпадение слоя — «мягкий» сигнал. Узкий сертификат (напр.
            # «бельё первого слоя») vs другой вид одежды (костюм/свитшот/платье) —
            # это НЕ доказанное несоответствие, а повод ПРОВЕРИТЬ ВРУЧНУЮ.
            soft_conflicts.append(f'слой/вид одежды не совпадает: карточка {sorted(card_layers)}, сертификат {sorted(cert_layers)}')

    card_tokens = _tokens_for_compare(card_text)
    cert_tokens = _tokens_for_compare(cert_text)
    inter = card_tokens & cert_tokens
    union = card_tokens | cert_tokens
    token_score = 100.0 * len(inter) / len(union) if union else 0.0
    seq_score = _seq_ratio(card_text, cert_text)

    score = token_score * 0.55 + seq_score * 0.25
    if card_cats and cert_cats and not card_cats.isdisjoint(cert_cats):
        score += 18
    if card_age == 'child' and cert_age == 'child':
        score += 12
    if card_sub & cert_sub:
        score += 30
    if card_layers & cert_layers:
        score += 20
    if _contains_broad_child_clothing(cert_text) and 'clothing' in card_cats:
        score += 15
    score = round(min(100.0, score), 1)

    # v39: специальное правило для «костюм» в карточке.
    # WB-карточка «Костюм спортивный для девочки» имеет card_sub={'костюм'},
    # а сертификат на «верхние второго слоя ... брюки, толстовка» имеет cert_sub={'брюки','толстовка'}.
    # Прямого пересечения по subtype нет, но фактически сертификат ПОКРЫВАЕТ карточку,
    # потому что костюм = толстовка+брюки. Это правило фиксит ложноотрицательный «ПРОВЕРИТЬ ВРУЧНУЮ».
    costume_implied_match = False
    if 'костюм' in card_sub and 'клонинг' not in cert_low:
        upper_parts = cert_sub & {'толстовка', 'футболка', 'рубашка'}
        lower_parts = cert_sub & {'брюки', 'шорты', 'юбка'}
        if upper_parts and lower_parts:
            costume_implied_match = True
        elif 'костюмные' in cert_low or 'изделия костюмные' in cert_low:
            costume_implied_match = True

    matched = sorted(list(inter))[:50]
    detail_base = (
        f'matched={matched}; card_sub={sorted(card_sub)}; cert_sub={sorted(cert_sub)}; '
        f'card_layers={sorted(card_layers)}; cert_layers={sorted(cert_layers)}; '
        f'card_age={card_age}; cert_age={cert_age}; card_gender={card_gender}; cert_gender={cert_gender}; '
        f'token={round(token_score,1)}; seq={round(seq_score,1)}'
    )

    if conflicts:
        return 'НЕСООТВЕТСТВИЕ', score, 'conflicts=' + '; '.join(conflicts) + '; ' + detail_base

    # Strong OK: exact subtype or apparel layer match.
    if card_sub & cert_sub:
        return 'OK', max(score, 82.0), 'Совпал конкретный вид продукции; ' + detail_base
    # v39: костюм покрыт компонентами в сертификате
    if costume_implied_match:
        return 'OK', max(score, 78.0), 'Костюм покрыт компонентами сертификата (верх + низ); ' + detail_base
    if 'clothing' in card_cats and 'clothing' in cert_cats and card_layers & cert_layers and (cert_age in {'child', 'unknown'} or card_age != 'child'):
        return 'OK', max(score, 78.0), 'Совпал слой/класс детской одежды; ' + detail_base
    # v27.9.x: «мягкий» конфликт слоя (без жёстких конфликтов и без точного
    # совпадения вида/слоя) — не авто-OK и не доказанное несоответствие.
    # Решение пользователя: узкий сертификат vs другой вид одежды → ПРОВЕРИТЬ ВРУЧНУЮ.
    if soft_conflicts:
        return 'ПРОВЕРИТЬ ВРУЧНУЮ', score, 'soft_conflicts=' + '; '.join(soft_conflicts) + '; ' + detail_base
    if 'clothing' in card_cats and _contains_broad_child_clothing(cert_text) and not card_layers:
        return 'OK', max(score, 72.0), 'Сертификат содержит широкую группу детской одежды, явных конфликтов нет; ' + detail_base

    # v40.3: ПРАВИЛА ДЛЯ ОБУВИ (раньше их не было — отсюда тысячи ложных ПРОВЕРИТЬ ВРУЧНУЮ).
    # Сертификаты на обувь почти всегда общие: «Обувь детская повседневная (ясельная,
    # малодетская, дошкольная, школьная)» — без конкретного вида. А в карточке —
    # «кроссовки/кеды/сандалии/сникеры». Это корректное совпадение: конкретный вид обуви
    # покрыт общим обувным сертификатом.
    if 'footwear' in card_cats and 'footwear' in cert_cats and not conflicts:
        # возраст совместим (оба детские / неизвестные, либо карточка не детская)
        age_ok = (cert_age in {'child', 'unknown'}) or (card_age == cert_age) or (card_age != 'child')
        if age_ok:
            return 'OK', max(score, 76.0), 'Обувь: вид из карточки покрыт обувным сертификатом, конфликтов нет; ' + detail_base

    # v40.3: общий случай — обе категории совпадают (одна и та же предметная область),
    # возраст совместим, явных конфликтов нет. Это покрывает обувь и прочие категории
    # где сертификат сформулирован обобщённо.
    if (card_cats and cert_cats and (card_cats & cert_cats) and not conflicts
            and card_age in ('child', 'unknown') and cert_age in ('child', 'unknown')):
        return 'OK', max(score, 70.0), 'Совпадает товарная категория, возраст совместим, конфликтов нет; ' + detail_base

    # v41.1: правила OK для НОВЫХ категорий (аксессуары, текстиль, кормление, детские товары, игрушки,
    # косметика, питание). Логика: если карточка и сертификат в одной такой категории и нет конфликтов —
    # это покрытие. Сертификаты в этих группах часто обобщённые («Головные уборы трикотажные детские»,
    # «Комплект постельного белья», «Изделия для кормления»).
    _coverable_cats = {'accessories', 'home_textile', 'feeding', 'nursery', 'toys', 'cosmetics', 'food'}
    shared_cats = (card_cats & cert_cats) & _coverable_cats
    if shared_cats and not conflicts:
        # совпал конкретный подтип — уверенный OK; иначе OK по категории
        if card_sub & cert_sub:
            return 'OK', max(score, 80.0), f'Совпал вид в категории {sorted(shared_cats)}; ' + detail_base
        return 'OK', max(score, 72.0), f'Совпадает категория {sorted(shared_cats)}, конфликтов нет; ' + detail_base

    # v41.1: если у карточки определён подтип и он пересекается с подтипом сертификата —
    # это сильный сигнал совпадения для ЛЮБОЙ категории (не только одежда/обувь).
    if card_sub & cert_sub and not conflicts:
        return 'OK', max(score, 80.0), 'Совпал конкретный вид продукции (любая категория); ' + detail_base

    # v39.3: если есть пересечения по ключевым словам (matched непустой) + одежда + age совместим,
    # это OK с пониженным score. Раньше такие случаи попадали в «ПРОВЕРИТЬ ВРУЧНУЮ» (~50 кейсов)
    # — например «Штанишки для новорожденных» + cert «для новорожденных до 3-х лет», где matched
    # = ['новорожденн'], но subtypes не пересекаются точно.
    if (matched and 'clothing' in card_cats and 'clothing' in cert_cats
            and card_age in ('child', 'unknown') and cert_age in ('child', 'unknown')
            and not conflicts):
        return 'OK', max(score, 70.0), f'Есть совпадения по ключевым словам ({len(matched)}) + одежда + age совместим; ' + detail_base

    # v27.7: ОБЩЕЕ правило для ЛЮБОЙ товарной категории.
    # Если карточка и сертификат относятся к одной и той же категории (electronics,
    # appliances, kitchenware, tools, auto, sport, furniture, household_chemistry,
    # stationery, pet, lighting, jewelry, и т.д.) и явных конфликтов нет — сертификат
    # покрывает товар. Возрастные/видовые/доменные конфликты уже отсеяны выше.
    # Это убирает массовое ложное «ПРОВЕРИТЬ ВРУЧНУЮ» для всех неодёжных категорий.
    shared_any = card_cats & cert_cats
    if shared_any and not conflicts:
        if card_sub & cert_sub:
            return 'OK', max(score, 82.0), f'Совпал вид + категория {sorted(shared_any)}; ' + detail_base
        return 'OK', max(score, 72.0), f'Совпадает товарная категория {sorted(shared_any)}, конфликтов нет; ' + detail_base

    # v46: «родственные» категории (electronics ↔ appliances) — одна предметная
    # область. Реестр часто пишет общий префикс «приборы бытового назначения»
    # (appliances), а WB — конкретику «зарядка/колонка/мышь» (electronics).
    if _categories_related(card_cats, cert_cats) and not conflicts:
        if card_sub & cert_sub:
            return 'OK', max(score, 80.0), f'Совпал вид; родственные категории {sorted(card_cats)}/{sorted(cert_cats)}; ' + detail_base
        return 'OK', max(score, 70.0), f'Родственные товарные категории {sorted(card_cats)}/{sorted(cert_cats)}, конфликтов нет; ' + detail_base

    if score >= 78:
        return 'OK', score, 'Высокое суммарное совпадение без конфликтов; ' + detail_base
    if score >= 40:
        return 'ПРОВЕРИТЬ ВРУЧНУЮ', score, 'Частичное совпадение, но недостаточно для автоматического OK; ' + detail_base
    return 'ПРОВЕРИТЬ ВРУЧНУЮ', score, 'Низкое совпадение без доказанного конфликта; ' + detail_base



# -----------------------------
# v38.6 registry-stage rewrite: browser-first visible parser with progress
# -----------------------------
# This block intentionally overrides only stage 2 (input-links-csv / registry parsing).
# WB link collection from v38.x is not touched.

FSA_PRODUCT_LABEL = "Наименование (обозначение) продукции"
FSA_COMMON_PRODUCT_LABEL = "Общее наименование продукции"
FSA_PRODUCT_LABELS_V386 = [FSA_PRODUCT_LABEL, FSA_COMMON_PRODUCT_LABEL]
FSA_CERT_NUMBER_LABEL = "Регистрационный номер сертификата"
FSA_DECL_NUMBER_LABEL = "Регистрационный номер декларации о соответствии"
SWIS_NUMBER_LABEL = "Регистрационный номер документа"

# v40.3: расширенные лейблы FSA — для извлечения через БРАУЗЕР (когда curl_cffi
# заблокирован антиботом). Каждое поле сопоставлено с ключом ResultRow.
# Лейблы — это точные подписи на странице pub.fsa.gov.ru.
FSA_EXTENDED_LABELS = {
    "applicant_name": [
        "Полное наименование заявителя", "Наименование заявителя", "Заявитель",
    ],
    "applicant_inn": [
        "Идентификационный номер налогоплательщика заявителя",
        "ИНН заявителя",
    ],
    "manufacturer_name": [
        "Полное наименование изготовителя", "Наименование изготовителя", "Изготовитель",
    ],
    "tnved": [
        "Код ТН ВЭД ЕАЭС", "Коды ТН ВЭД ЕАЭС", "Код ТН ВЭД", "ТН ВЭД ЕАЭС",
    ],
    "scheme": [
        "Схема сертификации", "Схема декларирования", "Схема подтверждения соответствия",
    ],
    "technical_regulation": [
        "Технические регламенты", "Технический регламент", "Наименование технического регламента",
    ],
    "document_date_start": [
        "Дата регистрации сертификата", "Дата регистрации декларации о соответствии",
        "Дата регистрации", "Дата начала действия",
    ],
    "document_date_end": [
        "Дата окончания действия сертификата",
        "Дата окончания действия декларации о соответствии",
        "Срок действия до", "Дата окончания действия",
    ],
}
# Плоский список всех расширенных лейблов
FSA_EXTENDED_LABELS_FLAT = [lab for labs in FSA_EXTENDED_LABELS.values() for lab in labs]

# v39.5: лейблы для статуса документа.
# На странице FSA сертификата: «Статус сертификата» → «Действует»/«Прекращён»/«Приостановлен».
# На странице FSA декларации: бывает «Статус декларации» либо «Статус действия декларации о соответствии».
# На странице SWIS: «Признак действия» → «Действует».
FSA_CERT_STATUS_LABEL = "Статус сертификата"
FSA_DECL_STATUS_LABELS = ["Статус декларации", "Статус действия декларации о соответствии", "Статус декларации о соответствии"]
SWIS_STATUS_LABEL = "Признак действия"


def _fsa_status_labels_for_url(url: str) -> List[str]:
    """v39.5: какие label искать для статуса документа FSA."""
    kind = _fsa_doc_kind_from_url(url)
    if kind == 'certificate':
        return [FSA_CERT_STATUS_LABEL]
    if kind == 'declaration':
        return list(FSA_DECL_STATUS_LABELS)
    return [FSA_CERT_STATUS_LABEL] + list(FSA_DECL_STATUS_LABELS)


def _normalize_doc_status(value: str) -> str:
    """v39.5: нормализация значения статуса документа. FSA и SWIS возвращают
    немного разные тексты — приводим к коротким стандартным:
        Действует, Прекращён, Приостановлен, Аннулирован, Архивный, Возобновлён.
    Если значение не известное — возвращаем исходник без изменений (укороченный).
    """
    if not value:
        return ''
    v = str(value).strip()
    if not v:
        return ''
    low = v.lower().replace('ё', 'е')
    mapping = [
        ('действуе', 'Действует'),
        ('прекраще', 'Прекращён'),
        ('прекращ', 'Прекращён'),
        ('приостанов', 'Приостановлен'),
        ('возобнов', 'Возобновлён'),
        ('аннулирован', 'Аннулирован'),
        ('архивн', 'Архивный'),
        ('недейств', 'Недействителен'),
    ]
    for needle, canonical in mapping:
        if needle in low:
            return canonical
    # Неизвестный — возвращаем как есть, обрезанным
    return v[:60]


def _fsa_doc_kind_from_url(url: str) -> str:
    u = (url or '').lower()
    if '/rss/certificate/' in u:
        return 'certificate'
    if '/rds/declaration/' in u:
        return 'declaration'
    return ''


def _fsa_number_labels_for_url(url: str) -> List[str]:
    """Return the ONLY allowed number label for an FSA URL.

    Per requirement:
      - certificate tab: «Регистрационный номер сертификата»
      - declaration tab: «Регистрационный номер декларации о соответствии»
    """
    kind = _fsa_doc_kind_from_url(url)
    if kind == 'certificate':
        return [FSA_CERT_NUMBER_LABEL]
    if kind == 'declaration':
        return [FSA_DECL_NUMBER_LABEL]
    return [FSA_CERT_NUMBER_LABEL, FSA_DECL_NUMBER_LABEL]


async def _click_fsa_section_tab_if_exists(page, url: str, purpose: str, timeout_ms: int = 1200) -> None:
    """Click the visible FSA tab before extracting fields, if it exists.

    This does not replace URL routing; it mirrors the manual process and helps
    when pub.fsa.gov.ru keeps state inside the SPA.
    """
    try:
        import re as _re
        kind = _fsa_doc_kind_from_url(url)
        if purpose == 'number':
            patterns = []
            if kind == 'certificate':
                patterns = [r'^\s*СЕРТИФИКАТ\s*$', r'^\s*Сертификат\s*$']
            elif kind == 'declaration':
                patterns = [r'^\s*ДЕКЛАРАЦИЯ\s+О\s+СООТВЕТСТВИИ\s*$', r'^\s*Декларация\s+о\s+соответствии\s*$', r'^\s*ДЕКЛАРАЦИЯ\s*$']
            else:
                patterns = [r'^\s*СЕРТИФИКАТ\s*$', r'^\s*ДЕКЛАРАЦИЯ\s+О\s+СООТВЕТСТВИИ\s*$']
        else:
            patterns = [r'^\s*СВЕДЕНИЯ\s+О\s+ПРОДУКЦИИ\s*$', r'^\s*Сведения\s+о\s+продукции\s*$']
        for pat in patterns:
            try:
                loc = page.get_by_text(_re.compile(pat, _re.I)).first
                await loc.click(timeout=timeout_ms)
                await page.wait_for_timeout(350)
                return
            except Exception:
                continue
    except Exception:
        return
SWIS_PRODUCT_LABELS = [
    "Однородное наименование продукции",
    "Полное наименование продукции и сведения, обеспечивающие её идентификацию (тип, марка, модель, артикул продукции и др.)",
]

REGISTRY_EXTRACT_JS = r"""
(arg) => {
  const labels = arg.labels || [];
  const stopLabels = arg.stopLabels || [];
  const norm = s => (s || '').toString().toLowerCase().replace(/ё/g,'е').replace(/\u00a0/g,' ').replace(/\s+/g,' ').trim().replace(/[ :.;\-–—]+$/g,'');
  const clean = s => (s || '').toString().replace(/\u00a0/g,' ').replace(/\s+/g,' ').trim().replace(/^[:\s\-–—]+/g,'');
  const wanted = new Map(labels.map(l => [norm(l), l]));
  const stop = new Set(stopLabels.map(norm));
  const out = {};
  labels.forEach(l => out[l] = []);
  function push(label, value) {
    value = clean(value);
    if (!value) return;
    const nl = norm(label);
    const nv = norm(value);
    if (nv.startsWith(nl)) {
      const raw = clean(value);
      value = clean(raw.slice(label.length));
    }
    value = clean(value);
    if (value) out[label].push(value);
  }
  function rowValue(row, label) {
    const valNodes = Array.from(row.querySelectorAll('[class*="info-row__text"], .info-row__text, [class*="_text"], [class*="value"], [class*="field-value"], td:nth-child(n+2)'));
    let vals = valNodes.map(n => n.innerText || n.textContent || '').map(clean).filter(Boolean);
    vals = vals.filter(v => norm(v) !== norm(label));
    if (vals.length) return clean(vals.join(' '));
    const cells = Array.from(row.querySelectorAll('td, th'));
    if (cells.length >= 2) {
      const first = clean(cells[0].innerText || cells[0].textContent || '');
      if (norm(first) === norm(label)) return clean(cells.slice(1).map(c => c.innerText || c.textContent || '').join(' '));
    }
    const clone = row.cloneNode(true);
    clone.querySelectorAll('[class*="info-row__header"], .info-row__header, th, td:first-child, [class*="header"], [class*="title"], [class*="field-title"]').forEach(x => x.remove());
    Array.from(clone.querySelectorAll('[title], [ng-reflect-title], [data-title], [aria-label]')).forEach(x => {
      for (const a of ['title','ng-reflect-title','data-title','aria-label']) {
        const v = x.getAttribute && x.getAttribute(a);
        if (v && norm(v) === norm(label)) { x.remove(); break; }
      }
    });
    return clean(clone.innerText || clone.textContent || '');
  }

  function titleLabel(el) {
    for (const attr of ['title', 'ng-reflect-title', 'data-title', 'aria-label']) {
      const v = el.getAttribute && el.getAttribute(attr);
      const lab = wanted.get(norm(v || ''));
      if (lab) return lab;
    }
    return null;
  }
  function containers(el) {
    const out = [];
    let cur = el;
    for (let i = 0; i < 7 && cur; i++, cur = cur.parentElement) out.push(cur);
    return out;
  }

  // 1) Exact attribute title, critical for pub.fsa.gov.ru figis-card-info-row.
  // In FSA the title can be on a header child, and the value is in a sibling in the parent row.
  for (const el of Array.from(document.querySelectorAll('[title], [ng-reflect-title], [data-title], [aria-label]'))) {
    const lab = titleLabel(el);
    if (!lab) continue;
    for (const row of containers(el)) {
      const v = rowValue(row, lab);
      if (v && norm(v) !== norm(lab)) { push(lab, v); break; }
    }
  }
  // 2) Tables: SWIS and old FSA layouts.
  for (const tr of Array.from(document.querySelectorAll('tr'))) {
    const cells = Array.from(tr.querySelectorAll('th,td'));
    if (cells.length >= 2) {
      const lab = wanted.get(norm(cells[0].innerText || cells[0].textContent));
      if (lab) push(lab, cells.slice(1).map(c => c.innerText || c.textContent || '').join(' '));
    }
  }
  // 3) Card row header/text layouts.
  for (const row of Array.from(document.querySelectorAll('figis-card-info-row, [class*="info-row"], [class*="card-info-row"], [class*="row"]'))) {
    const header = row.querySelector('[class*="info-row__header"], .info-row__header, [class*="header"], [class*="title"], th, td:first-child');
    if (!header) continue;
    const lab = wanted.get(norm(header.innerText || header.textContent));
    if (lab) push(lab, rowValue(row, lab));
  }
  // 4) Exact visible-text fallback. This is still strict: only exact label names can start extraction.
  const lines = (document.body ? (document.body.innerText || '') : '').split(/\n+/).map(clean).filter(Boolean);
  function escRe(s) { return (s || '').replace(/[.*+?^${}()|[\]\\]/g, '\\$&'); }
  for (let i = 0; i < lines.length; i++) {
    let lab = wanted.get(norm(lines[i]));
    let sameLineValue = '';
    if (!lab) {
      // FSA sometimes renders the row as one visible line:
      // «Регистрационный номер сертификата ЕАЭС ...»
      // or «Регистрационный номер декларации о соответствии ЕАЭС ...».
      for (const [nl, original] of wanted.entries()) {
        const li = norm(lines[i]);
        if (li.startsWith(nl + ' ')) {
          lab = original;
          sameLineValue = clean(lines[i].replace(new RegExp('^\\s*' + escRe(original) + '\\s*:?\\s*', 'i'), ''));
          if (norm(sameLineValue) === norm(lines[i])) {
            // fallback when regexp failed because of NBSP/case/ё
            sameLineValue = clean(lines[i].slice(original.length));
          }
          break;
        }
      }
    }
    if (!lab) continue;
    if (sameLineValue) {
      push(lab, sameLineValue);
      continue;
    }
    const isProduct = norm(lab).includes('продукц');
    let buf = [];
    for (let j = i + 1; j < Math.min(lines.length, i + (isProduct ? 80 : 12)); j++) {
      const nj = norm(lines[j]);
      if (wanted.has(nj) || stop.has(nj)) break;
      if (/^(статус|дата|номер бланка|свободное распространение|лицо, подписавшее|происхождение продукции|общие условия|код тн|тн вэд|сведения об|сведения о |заявитель|изготовитель|адрес|телефон|факс|email|e-mail|схема)/i.test(lines[j])) break;
      buf.push(lines[j]);
      if (buf.join(' ').length > (isProduct ? 12000 : 500)) break;
    }
    push(lab, buf.join(' '));
  }
  for (const k of Object.keys(out)) {
    const seen = new Set();
    out[k] = out[k].map(clean).filter(v => {
      const key = norm(v);
      if (!key || seen.has(key)) return false;
      seen.add(key); return true;
    });
  }
  return out;
}
"""


# v27.5: Лейблы, которые часто прилипают в начало значения applicant/manufacturer на SWIS
# (swis.trade.kg) и FSA: HTML отдаёт ячейки без явных разделителей, и заголовок попадает в значение.
_ORG_HEAD_LABEL_RES = [
    # SWIS (Киргизия) — всё до наименования.
    r'^(?:Тип\s+организации\s+[^\n]*?\s+)?Полное\s+наименование\s+организации[—\-—]?изготовителя\s+продукции\s+',
    r'^(?:Тип\s+организации\s+[^\n]*?\s+)?Полное\s+наименование\s+изготовителя\s+',
    r'^(?:Тип\s+организации\s+[^\n]*?\s+)?Полное\s+наименование\s+',
    r'^Тип\s+организации\s+(?:Изготовитель|Уполномоченное|Продавец|Заявитель|Юридическое|Физическое)[^\n]*?\s+',
    # BelGISS / ЕАЭС — всё до наименования (на всякий случай).
    r'^Страна\s*\([A-ZА-Я]{2,3}\)\s+[А-Яа-я\w]+\s+(?:Краткое|Наименование)\s+наименование\s+хозяйствующего\s+субъекта\s+',
    r'^Краткое\s+наименование\s+хозяйствующего\s+субъекта\s+',
    r'^Наименование\s+хозяйствующего\s+субъекта\s+',
    r'^Страна\s*\([A-ZА-Я]{2,3}\)\s+',  # «Страна (BY) Беларусь» как префикс
]

# v27.5: Лейблы, которые встречаются в середине/конце значения applicant/manufacturer
# и являются началом соседнего поля. Отрезаем всё, что идёт после них.
_ORG_TAIL_LABEL_RE = re.compile(
    r'(?is)\s*(?:,\s+)?(?:'
    r'ОГРН(?:ИП)?[\s:]+\d|'
    r'ИНН(?:/КПП)?[\s:]+\d|'
    r'КПП[\s:]+\d|'
    r'ФИО\s+руководителя|'
    r'Юридический\s+адрес|'
    r'Почтовый\s+адрес|'
    r'Адрес\s+места|'
    r'Номер\s+телефона|'
    r'Адрес\s+электронной\s+почты|'
    r'Адрес\s+мест\s+осуществления|'
    r'Продукция\s+Обувь|'  # “Продукция Обувь детская...” — форма FSA-лейла
    r'СВЕДЕНИЯ\s+О\s+|'  # заголовки разделов
    r'Идентификатор\s+хозяйствующего|'
    r'Сведения\s+о\s+регистрации'
    r')'
)

# v27.5: Лейблы-префиксы, которые повтораются внутри product_name BelGISS:
# "оценки соответствия обувь детская; оценки соответствия сапоги...". Оставлено для
# совместимости со старыми вызовами (BelGISS теперь вообще не извлекает product).
_PRODUCT_REPEATED_PREFIX_RE = re.compile(r'(?i)(?:^|;\s*)(?:оценки\s+соответствия\s+|в составе\s+)')


def _clean_org_name(value: str) -> str:
    """v27.5: Очищает имя организации (applicant/manufacturer) от прилипших лейблов
    HTML-таблицы SWIS/FSA: "Тип организации", "Полное наименование", "ОГРН", "ИНН"
    и т.д.
    """
    if not value:
        return ''
    s = clean_registry_value(value)
    if not s:
        return ''
    # 1) Срезаем головные лейблы (несколько проходов, пока что-то срезается).
    for _pass in range(3):
        before = s
        for pat in _ORG_HEAD_LABEL_RES:
            new_s = re.sub(pat, '', s, count=1, flags=re.IGNORECASE)
            if new_s != s and new_s.strip():
                s = new_s
                break
        if s == before:
            break
    # 2) Обрезаем хвосты по лейблам соседних полей.
    parts = _ORG_TAIL_LABEL_RE.split(s, maxsplit=1)
    s = parts[0] if parts else s
    # 3) Последняя запятая без организации («..., ОГРН ...») — уже убралось выше,
    # но всё равно на всякий случай чистим хвост от служебных знаков.
    s = clean_registry_value(s)
    # 4) Если остался только заголовок-секция или строка < 4 символов — вернём пусто.
    if not s or len(s) < 3:
        return ''
    low_s = s.lower().replace('ё', 'е')
    bad_only = (
        'сведения о таможенном декларировании',
        'сведения об изготовителе',
        'сведения о заявителе',
        'сведения о сертификате',
        'идентификатор хозяйствующего субъекта',
    )
    for bad in bad_only:
        if low_s.startswith(bad):
            return ''
    return s[:300]


def _clean_tnved_code(value: str) -> str:
    """v27.5: Нормализует Код ТН ВЭД. Оставляет только цифры/точки/пробелы
    первым блоком, вырезая хвостовое описание.
    Пример: «6404 Обувь с подошвой...» → «6404».
    «ТС 6404199000» → «ТС 6404199000» (оставляем префикс ТС).
    """
    if not value:
        return ''
    s = clean_registry_value(value).strip()
    # Разрешён префикс "ТС" или "ЕАЭС".
    m = re.match(r'^(?:ТС|TS|ЕАЭС|EAEU)?\s*([\d.\s]+)', s)
    if not m:
        return s[:60]
    digits_part = m.group(1).strip()
    prefix = m.group(0)[:m.start(1)].strip()
    # Длина полного кода ТН ВЭД — 10 цифр (с возм. точками). Пробелы оставляем
    # в префиксе, но внутри цифр пробелы сжимаем.
    digits_clean = re.sub(r'\s+', ' ', digits_part).strip()
    # Если после цифр в исходнике начинались буквы — это описание, выбрасываем.
    if not digits_clean:
        return s[:60]
    if prefix:
        return f'{prefix} {digits_clean}'.strip()
    return digits_clean


def _registry_value_postclean(label: str, value: str) -> str:
    value = clean_registry_value(value)
    if not value:
        return ""
    ln = norm_text(label)
    # Strict stop markers: fields after the target value must not leak into the product/number.
    if 'продукц' in ln:
        value = re.split(
            r'(?is)\s+(?:Код\s+ТН\s+ВЭД|ТН\s+ВЭД|Артикул|Документ,\s*в\s*соответствии|Иная\s+информация|Стандарты|Происхождение\s+продукции|Общие\s+условия|Срок\s+хранения|Изготовитель|Заявитель|Сведения\s+об|Регистрационный\s+номер)\b',
            value,
        )[0]
        return clean_registry_value(value)
    # v27.5: орг-поля — applicant_name / manufacturer_name / Заявитель / Изготовитель.
    if any(k in ln for k in ('applicant', 'manufacturer', 'заявител', 'изготовител')):
        return _clean_org_name(value)
    # v27.5: Код ТН ВЭД — обрезать описание после цифр.
    if ('тн' in ln and 'вэд' in ln) or 'tnved' in ln:
        return _clean_tnved_code(value)
    value = re.split(
        r'(?is)\s+(?:Дата\s+регистрации|Дата\s+окончания|Статус|Номер\s+бланка|Свободное\s+распространение|Лицо,\s*подписавшее)\b',
        value,
    )[0]
    return clean_registry_value(value)


async def _extract_labels_from_page_v386(page, labels: List[str], wait_ms: int) -> Dict[str, List[str]]:
    stop_labels = [
        'Статус сертификата', 'Регистрационный номер сертификата', 'Регистрационный номер декларации о соответствии',
        'Дата регистрации сертификата', 'Дата регистрации декларации о соответствии', 'Дата окончания действия сертификата',
        'Дата окончания действия декларации о соответствии', 'Номер бланка', 'Код ТН ВЭД ЕАЭС', 'Документ, в соответствии с которым изготовлена продукция',
        'Иная информация о продукции', 'Происхождение продукции', 'Общие условия хранения продукции',
        'Общее наименование продукции', 'Однородное наименование продукции', 'Полное наименование продукции и сведения, обеспечивающие её идентификацию (тип, марка, модель, артикул продукции и др.)',
    ]
    # v39: ждём НЕ ТОЛЬКО появления лейбла, но и появления НЕПУСТОГО значения рядом с ним.
    # Старая логика заканчивала ожидание, как только лейбл появлялся в DOM — но в Angular SPA
    # значение появляется позже, и extract возвращал пустоту. Это была главная причина
    # того, что 60% FSA-реестров получали статус "НЕ УДАЛОСЬ ИЗВЛЕЧЬ".
    try:
        await page.wait_for_function(
            """(args) => {
                const labels = args.labels;
                const norm = s => (s||'').toString().toLowerCase().replace(/ё/g,'е').replace(/\\u00a0/g,' ').replace(/\\s+/g,' ').trim();
                const clean = s => (s||'').toString().replace(/\\u00a0/g,' ').replace(/\\s+/g,' ').trim();
                // 1) Проверяем title-атрибуты (figis-card-info-row).
                const titleEls = Array.from(document.querySelectorAll('[title], [ng-reflect-title], [data-title]'));
                for (const el of titleEls) {
                    for (const attr of ['title','ng-reflect-title','data-title']) {
                        const t = el.getAttribute && el.getAttribute(attr);
                        if (!t) continue;
                        const tn = norm(t);
                        if (!labels.some(l => norm(l) === tn)) continue;
                        // Нашли элемент с правильным title. Ищем значение в его родителе.
                        let parent = el;
                        for (let i = 0; i < 6 && parent; i++, parent = parent.parentElement) {
                            const txt = clean(parent.innerText || parent.textContent || '');
                            // Считаем что значение есть, если общая длина текста > длины лейбла + 8 символов.
                            if (txt.length > t.length + 8) return true;
                        }
                    }
                }
                // 2) Проверяем видимый текст: лейбл + хотя бы несколько символов после.
                const body = document.body ? (document.body.innerText || '') : '';
                const lines = body.split(/\\n+/).map(clean).filter(Boolean);
                for (const lab of labels) {
                    const nl = norm(lab);
                    for (let i = 0; i < lines.length; i++) {
                        const ln = norm(lines[i]);
                        // Лейбл и значение на одной строке: "Регистрационный номер сертификата ЕАЭС N..."
                        if (ln.startsWith(nl + ' ') && (lines[i].length > lab.length + 3)) return true;
                        // Лейбл на отдельной строке, значение — на следующей.
                        if (ln === nl && i + 1 < lines.length) {
                            const next = clean(lines[i + 1]);
                            if (next && next.length > 3 && norm(next) !== nl) return true;
                        }
                    }
                }
                return false;
            }""",
            {"labels": labels},
            timeout=max(2000, wait_ms),
        )
    except Exception:
        # fallback на старую проверку — только лейбл
        try:
            await page.wait_for_function(
                """(labels) => {
                    const norm=s=>(s||'').toString().toLowerCase().replace(/ё/g,'е').replace(/\\s+/g,' ').trim();
                    const txt=document.body ? document.body.innerText : '';
                    return labels.some(l => norm(txt).includes(norm(l)));
                }""",
                labels,
                timeout=max(1000, wait_ms // 2),
            )
        except Exception:
            pass
    # Дополнительная пауза, чтобы Angular точно дорендерил соседние строки.
    try:
        await page.wait_for_timeout(max(0, min(wait_ms, 2000)))
    except Exception:
        pass
    data: Dict[str, List[str]] = {l: [] for l in labels}
    try:
        raw = await page.evaluate(REGISTRY_EXTRACT_JS, {'labels': labels, 'stopLabels': stop_labels})
        if raw:
            for lab in labels:
                vals = []
                for v in raw.get(lab, []) or []:
                    vv = _registry_value_postclean(lab, v)
                    if vv:
                        vals.append(vv)
                data[lab] = _unique_keep_order(vals)
    except Exception:
        pass
    if not any(data.values()):
        try:
            html_text = await page.content()
            extracted = _extract_exact_label_values_from_html(html_text, labels)
            for lab in labels:
                data[lab] = [_registry_value_postclean(lab, v) for v in extracted.get(lab, []) if _registry_value_postclean(lab, v)]
                data[lab] = _unique_keep_order(data[lab])
        except Exception:
            pass
    return data


async def _wait_until_fsa_rendered(page, timeout_ms: int = 20000) -> bool:
    """v39.6: ждать пока Angular SPA ФСА реально отрендерит контент страницы.

    v40.2: УБРАН долгий networkidle (до 15с). FSA имеет фоновый polling, из-за
    которого networkidle почти никогда не наступает и всегда ждёт полный таймаут.
    Теперь только wait_for_function с коротким таймаутом — данные появляются за 2-5с.
    Это ускоряет парсинг каждого документа в несколько раз.
    """
    # Ждём пока body станет содержательным (данные подгрузились через AJAX).
    # Короткий таймаут: на практике FSA рендерит за 2-5с, 10с — с запасом.
    fast_timeout = min(timeout_ms, 10000)
    try:
        await page.wait_for_function(
            """() => {
                const body = document.body;
                if (!body) return false;
                const len = (body.innerText || '').length;
                if (len > 500) return true;
                if (document.querySelector('figis-card-info-row')) return true;
                if (document.querySelector('[ng-reflect-title]')) return true;
                if (document.querySelector('[class*="info-row"]')) return true;
                if (document.querySelector('app-info')) return true;
                return false;
            }""",
            timeout=fast_timeout,
        )
        return True
    except Exception:
        return False


async def _fetch_fsa_via_browser_page(page, url: str) -> Optional[Dict[str, str]]:
    """v27.9.x: тянет JSON-API ФСА ЧЕРЕЗ САМ БРАУЗЕР (same-origin fetch со страницы
    pub.fsa.gov.ru).

    Зачем: когда антибот ФСА блокирует curl_cffi (HTTP 403 → срабатывает circuit
    breaker), весь парсинг уходит в скрейп рендера, который НЕ добирает часть
    полей (ИНН заявителя, изготовитель, схема, техрегламент — они на других
    вкладках/в JSON). Но страницы ФСА в браузере открываются нормально, значит
    same-origin fetch к /api/v1/... несёт настоящие cookie/TLS браузера и
    проходит там, где curl_cffi падает. Ответ — структурированный JSON со ВСЕМИ
    полями, парсим его тем же parse_fsa_json.

    Требование: страница уже должна быть на домене pub.fsa.gov.ru (иначе fetch
    будет cross-origin). Возвращает dict полей (как parse_fsa_json) или None.
    """
    kind, doc_id = extract_fsa_kind_id(url)
    if not kind or not doc_id:
        return None
    for _label, api_url in _fsa_candidates(kind, doc_id, aggressive=True):
        try:
            data = await page.evaluate(
                """async (u) => {
                    try {
                        const r = await fetch(u, {headers: {'Accept': 'application/json'}, credentials: 'include'});
                        if (!r.ok) return {__status: r.status};
                        return await r.json();
                    } catch (e) { return {__error: String(e)}; }
                }""",
                api_url,
            )
        except Exception:
            continue
        if isinstance(data, dict) and "__status" not in data and "__error" not in data:
            try:
                parsed = parse_fsa_json(data, url, kind, doc_id)
            except Exception:
                parsed = None
            if parsed and parsed.get("doc_number"):
                parsed["source"] = f"browser_page_fetch:{api_url}"
                return parsed
    return None


async def _harvest_fsa_cookies(page) -> None:
    """v45.2: снимает куки сессии pub.fsa.gov.ru с браузерного контекста в общий
    кэш _FSA_SESSION_COOKIES. Вызывается после успешной загрузки FSA-документа —
    браузер к этому моменту уже прошёл JS-антибот и получил валидную сессию.
    С этими куками curl_cffi обращается к API ФСА напрямую (быстрый путь)."""
    global _FSA_SESSION_COOKIES
    try:
        ck = await page.context.cookies("https://pub.fsa.gov.ru")
    except Exception:
        return
    jar = {}
    for c in ck or []:
        name = c.get("name")
        val = c.get("value")
        if name and val is not None:
            jar[str(name)] = str(val)
    if jar:
        _FSA_SESSION_COOKIES = jar


async def _parse_fsa_with_existing_page_v386(page, url: str, args) -> Tuple[str, str, str, str, str]:
    """v39.5: возвращает (cert_number, product_name, doc_type, doc_status, detail).

    v39.13: ПЕРЕД браузерной обработкой пробуем HTTP fast-path через curl_cffi.
    v40.2: добавлен circuit breaker — если curl_cffi даёт 403/неудачу несколько раз
    подряд (антибот FSA блокирует HTTP в этой сети), HTTP-попытки ОТКЛЮЧАЮТСЯ на весь
    прогон, и парсинг идёт сразу через браузер. Это убирает ~10-20 сек бесполезных
    HTTP-запросов перед каждым документом.
    """
    global _FSA_HTTP_FAILS, _FSA_HTTP_DISABLED, _FSA_COOKIE_HTTP_OK, _FSA_COOKIE_HTTP_FAIL

    def _apply_http_result(result) -> Optional[Tuple[str, str, str, str, str]]:
        """Раскладывает успешный HTTP-результат в кортеж + кэш расширенных полей."""
        if not (result and result.get('doc_number')):
            return None
        _FSA_EXTENDED_FIELDS_CACHE[url] = {
            'document_date_start': result.get('date_start', ''),
            'document_date_end': result.get('date_end', ''),
            'applicant_name': _clean_org_name(result.get('applicant', '')),
            'applicant_inn': result.get('applicant_inn', ''),
            'manufacturer_name': _clean_org_name(result.get('manufacturer', '')),
            'tnved': _clean_tnved_code(result.get('tnved', '')),
            'scheme': result.get('scheme', ''),
            'technical_regulation': result.get('technical_regulation', ''),
        }
        return (result.get('doc_number', ''), result.get('product_full', ''),
                result.get('doc_type', ''), result.get('status', ''),
                f"fsa_cookie_http_ok; source={result.get('source','')[:120]}")

    # v45.2: БЫСТРЫЙ путь по КУКАМ браузера. Если у нас уже есть сессионные куки ФСА
    # (сняты с браузера, который прошёл JS-антибот), тянем JSON API напрямую через
    # curl_cffi — это один лёгкий запрос вместо полной загрузки SPA. Так первый
    # документ идёт через браузер (и отдаёт куки), а все следующие — мгновенно по HTTP.
    if (bool(getattr(args, 'fsa_cookie_http', True)) and _FSA_SESSION_COOKIES
            and is_curl_cffi_available()):
        try:
            impersonate = str(getattr(args, 'fsa_curl_cffi_impersonate', 'chrome') or 'chrome')
            http_timeout = min(8.0, max(4.0, int(getattr(args, 'registry_browser_timeout_ms', 30000)) / 1000.0))
            result = await asyncio.to_thread(
                fetch_fsa_via_http, url,
                timeout_sec=http_timeout, impersonate=impersonate,
                user_agent=getattr(args, 'user_agent', None),
                skip_warmup=True, cookies=dict(_FSA_SESSION_COOKIES),
            )
            tup = _apply_http_result(result)
            if tup is not None:
                _FSA_COOKIE_HTTP_OK += 1
                return tup
            # куки не сработали (протухли / API сменил формат) — сбрасываем, дальше
            # пойдём через браузер, который заодно добудет свежие куки.
            _FSA_COOKIE_HTTP_FAIL += 1
            if _FSA_COOKIE_HTTP_FAIL >= 3:
                _FSA_SESSION_COOKIES.clear()
                _FSA_COOKIE_HTTP_FAIL = 0
        except Exception:
            _FSA_COOKIE_HTTP_FAIL += 1

    # v40.2: «голый» HTTP fast-path без кук (по умолчанию ВЫКЛ — FSA режет по TLS/IP).
    if bool(getattr(args, 'fsa_http_fast_path', False)) and not _FSA_HTTP_DISABLED:
        try:
            if is_curl_cffi_available():
                impersonate = str(getattr(args, 'fsa_curl_cffi_impersonate', 'chrome') or 'chrome')
                http_timeout = min(8.0, max(4.0, int(getattr(args, 'registry_browser_timeout_ms', 30000)) / 1000.0))
                result = await asyncio.to_thread(
                    fetch_fsa_via_http, url,
                    timeout_sec=http_timeout, impersonate=impersonate,
                    user_agent=getattr(args, 'user_agent', None),
                )
                tup = _apply_http_result(result)
                if tup is not None:
                    _FSA_HTTP_FAILS = 0
                    return tup
                _FSA_HTTP_FAILS += 1
                if _FSA_HTTP_FAILS >= _FSA_HTTP_FAIL_LIMIT and not _FSA_HTTP_DISABLED:
                    _FSA_HTTP_DISABLED = True
                    print(f"⚡ «Голый» HTTP-парсинг FSA отключён после {_FSA_HTTP_FAILS} неудач подряд.")
        except Exception:
            _FSA_HTTP_FAILS += 1
            if _FSA_HTTP_FAILS >= _FSA_HTTP_FAIL_LIMIT and not _FSA_HTTP_DISABLED:
                _FSA_HTTP_DISABLED = True

    number_routes, product_routes = fsa_exact_routes(url)
    wait_ms = int(getattr(args, 'registry_browser_wait_ms', 12000))
    timeout_ms = int(getattr(args, 'registry_browser_timeout_ms', 30000))
    cert_vals: List[str] = []
    prod_vals: List[str] = []
    status_vals: List[str] = []  # v39.5
    ext_vals: Dict[str, List[str]] = {}  # v44: расширенные поля (схема/изготовитель/ТР ТС/даты) — со всех вкладок
    details: List[str] = []
    number_labels = _fsa_number_labels_for_url(url)
    status_labels = _fsa_status_labels_for_url(url)  # v39.5

    # Шаг 1: номер документа.
    any_goto_succeeded = False  # v39.2: трекаем удалось ли вообще достучаться
    number_route_used = ''
    last_page_url_after_number = ''

    # v27.9.x: ГЛАВНЫЙ путь ФСА — ПЕРЕХВАТ ответа JSON-API, который Angular-SPA
    # ФСА делает САМА при загрузке страницы (с настоящим Bearer-токеном). Это
    # даёт ПОЛНЫЙ набор полей (ИНН заявителя, изготовитель, схема, техрегламент)
    # там, где curl_cffi и «голый» fetch получают 403. page.expect_response
    # авто-очищается (не течёт между документами). Не вышло — ниже обычный
    # браузерный парсинг (полный откат).
    _kind_fsa, _doc_id_fsa = extract_fsa_kind_id(url)
    if _kind_fsa and _doc_id_fsa and number_routes:
        def _fsa_api_match(r):
            u = r.url
            return ("/api/v1/" in u and str(_doc_id_fsa) in u
                    and ("certificate" in u or "declaration" in u))
        try:
            async with page.expect_response(_fsa_api_match, timeout=min(timeout_ms, 15000)) as _info:
                await page.goto(number_routes[0], wait_until='domcontentloaded', timeout=timeout_ms)
                await _wait_until_fsa_rendered(page, timeout_ms=min(timeout_ms, 12000))
            _resp = await _info.value
            any_goto_succeeded = True
            number_route_used = number_routes[0]
            if _resp.ok:
                # v45.2: страница прошла JS-антибот ФСА и API ответил 200 — снимаем
                # сессионные куки браузера в общий кэш. Следующие документы пойдут
                # быстрым HTTP-путём по этим кукам (без полной загрузки SPA).
                if bool(getattr(args, 'fsa_cookie_http', True)):
                    try:
                        await _harvest_fsa_cookies(page)
                    except Exception:
                        pass
                _cap = await _resp.json()
                # v27.9.x: один раз за прогон сохраняем сырой JSON ФСА для разбора
                # структуры (status/scheme/название приходят кодами/в др. полях).
                global _FSA_SAMPLE_DUMPED
                if not _FSA_SAMPLE_DUMPED:
                    try:
                        Path("fsa_api_sample.json").write_text(
                            json.dumps(_cap, ensure_ascii=False, indent=2), encoding="utf-8")
                        _FSA_SAMPLE_DUMPED = True
                        print("📝 [diag] Структура ответа API ФСА сохранена в fsa_api_sample.json "
                              "— пришлите этот файл для полного быстрого разбора ФСА.")
                    except Exception:
                        pass
                _parsed = parse_fsa_json(_cap, url, _kind_fsa, _doc_id_fsa)
                if _parsed.get('doc_number'):
                    # v27.9.x: разбор по реальной структуре JSON ФСА. В кэш кладём
                    # все надёжные поля (ИНН, заявитель, изготовитель, схема,
                    # техрегламент, даты).
                    _tech_reg = _parsed.get('technical_regulation', '')
                    # v27.9.x: ТР ТС и ТН ВЭД в JSON ФСА заданы только внутренними
                    # числовыми id (idTechnicalReglaments / idTnveds), а не текстом.
                    # Берём настоящие значения из текста уже загруженной страницы.
                    _tnved = ''
                    _ptxt = ''
                    try:
                        _ptxt = await page.evaluate(
                            "() => (document.body ? (document.body.innerText || '') : '')")
                    except Exception:
                        _ptxt = ''
                    if not _tech_reg and _ptxt:
                        _trm = re.findall(r'ТР\s+(?:ТС|ЕАЭС)\s+\d{3}/\d{4}', _ptxt)
                        if _trm:
                            _tech_reg = '; '.join(dict.fromkeys(
                                re.sub(r'\s+', ' ', x).strip() for x in _trm))
                    if _ptxt:
                        # после «ТН ВЭД» идёт 4–10-значный код (возможно с пробелами)
                        _tnm = re.findall(r'ТН\s?ВЭД[^\d]{0,40}?(\d[\d\s]{2,13}\d)', _ptxt)
                        _codes = []
                        for _c in _tnm:
                            _cc = _clean_tnved_code(_c)
                            if _cc and _cc not in _codes:
                                _codes.append(_cc)
                        if _codes:
                            _tnved = '; '.join(_codes[:10])
                    # v46: СХЕМА — берём из ТЕКСТА страницы («Схема декларирования: 1д» /
                    # «Схема сертификации: 1с»). Для деклараций JSON-поле idDeclScheme —
                    # внутренний id (3581), не номер схемы, поэтому текст надёжнее.
                    _scheme_val = _parsed.get('scheme', '')
                    if _ptxt:
                        _scm = re.search(
                            r'Схема\s+(?:сертификации|декларирования|подтверждения\s+соответствия)'
                            r'\D{0,25}?(\d{1,2})\s*([сдСД])', _ptxt)
                        if _scm:
                            _scheme_val = f"{_scm.group(1)}{_scm.group(2).lower()}"
                    _api_ext = {
                        'document_date_start': _parsed.get('date_start', ''),
                        'document_date_end': _parsed.get('date_end', ''),
                        'applicant_name': _clean_org_name(_parsed.get('applicant', '')),
                        'applicant_inn': _parsed.get('applicant_inn', ''),
                        'manufacturer_name': _clean_org_name(_parsed.get('manufacturer', '')),
                        'scheme': _scheme_val,
                        'technical_regulation': _tech_reg,
                        'tnved': _tnved,
                    }
                    _FSA_EXTENDED_FIELDS_CACHE[url] = {k: v for k, v in _api_ext.items() if v}
                    _st = _parsed.get('status', '')
                    _known = bool(_st) and not _st.startswith('Статус ')
                    if _known and _parsed.get('product_full'):
                        # Статус ПОДТВЕРЖДЁН (idStatus в словаре, напр. 6=Действует)
                        # и есть название из product.fullName — отдаём всё из API,
                        # БЕЗ браузера: быстро и консистентно. Название берём из
                        # JSON (а не со скачущей вёрстки) — это чинит и сравнение.
                        details.append('fsa_api_full_ok; via=spa_response')
                        return (_parsed.get('doc_number', ''), _parsed.get('product_full', ''),
                                _parsed.get('doc_type', ''), _st, 'fsa_api_capture_ok')
                    # Статус неизвестен (не код 6) — добираем браузером надёжный
                    # текст статуса и название (редкий случай: недействующие).
                    details.append('fsa_api_ext_ok')
        except Exception as _e:
            details.append(f'fsa_api_capture_miss={type(_e).__name__}')

    for route in number_routes:
        try:
            await page.goto(route, wait_until='domcontentloaded', timeout=timeout_ms)
            any_goto_succeeded = True
            number_route_used = route
            # v39.6: КЛЮЧЕВОЙ ФИКС — FSA это Angular SPA. После domcontentloaded
            # body может содержать пустой каркас (~120 символов). Нужно явно ждать
            # пока Angular подгрузит данные через AJAX и страница станет «настоящей».
            rendered = await _wait_until_fsa_rendered(page, timeout_ms=min(timeout_ms, 12000))
            # v39.6.1: лог длины текста — диагностика если страница пуста
            try:
                _len_now = await page.evaluate("() => (document.body ? (document.body.innerText || '').length : 0)")
            except Exception:
                _len_now = -1
            details.append(f'number_page_render={"ok" if rendered else "timeout"};number_page_text_len={_len_now}')

            await _click_fsa_section_tab_if_exists(page, url, 'number')

            # v40.2: до 2 попыток extract (было 3) с короткой паузой — ускоряет парсинг.
            for attempt in range(2):
                # v44: на baseInfo берём номер, статус И расширенные поля этой вкладки
                # (схема, даты регистрации/окончания) — раньше они терялись.
                base_ext_labels = (FSA_EXTENDED_LABELS["scheme"] +
                                   FSA_EXTENDED_LABELS["document_date_start"] +
                                   FSA_EXTENDED_LABELS["document_date_end"])
                fields = await _extract_labels_from_page_v386(
                    page, number_labels + status_labels + base_ext_labels, wait_ms)
                for lab in number_labels:
                    cert_vals.extend(fields.get(lab, []))
                for lab in status_labels:
                    status_vals.extend(fields.get(lab, []))
                # расширенные поля baseInfo
                for fk in ("scheme", "document_date_start", "document_date_end"):
                    if fk not in ext_vals:
                        for lab in FSA_EXTENDED_LABELS[fk]:
                            if fields.get(lab):
                                ext_vals[fk] = fields[lab]
                                break
                # Дедуп
                cert_vals = list(dict.fromkeys(cert_vals))
                status_vals = list(dict.fromkeys(status_vals))
                if cert_vals and status_vals:
                    break
                # Не хватает данных — короткая пауза и ещё попытка
                try:
                    await page.wait_for_timeout(800)
                except Exception:
                    pass

            if cert_vals:
                details.append(f'number_route={route};number_label={number_labels[0]};attempts={attempt+1}')
                if status_vals:
                    details.append(f'status_label_ok')
                break
        except Exception as e:
            # v39.2: вытащить точную причину (net::ERR_NAME_NOT_RESOLVED, net::ERR_CONNECTION_TIMED_OUT и т.п.)
            err_msg = str(e)[:200].replace(';', ',').replace('\n', ' ')
            details.append(f'number_route_error={type(e).__name__}:{err_msg}')
            # v27.9.x: если это СЕТЕВОЙ сбой (host недоступен/таймаут соединения),
            # остальные маршруты ТОГО ЖЕ хоста тоже не достучатся — не перебираем их
            # по 30с каждый. Раньше один недоступный FSA-документ занимал воркер
            # ~114с (перебор number+product+manufacturer вкладок) и UI «висел».
            if any(mk in err_msg for mk in (
                    'ERR_CONNECTION', 'ERR_TIMED_OUT', 'ERR_NAME_NOT_RESOLVED',
                    'ERR_ADDRESS_UNREACHABLE', 'ERR_INTERNET_DISCONNECTED',
                    'ERR_NETWORK', 'ERR_ABORTED', 'ERR_SOCKET')):
                details.append('number_routes_network_abort')
                break
            continue

    # v39.4: если поиск по точному label НЕ нашёл номер — пробуем regex-fallback
    # по всему тексту страницы. Стандартные форматы:
    #   ЕАЭС N RU Д-...          (декларация)
    #   ЕАЭС RU С-...            (сертификат)
    #   ROSS RU.AB12.B12345       (старый)
    # Это лечит ситуацию когда FSA меняет верстку лейблов — формат самого номера стабилен.
    # v39.5: одновременно ищем статус документа.
    if any_goto_succeeded:
        try:
            page_text = await page.evaluate("() => (document.body ? (document.body.innerText || '') : '')")
            import re as _re
            # Regex для номера документа
            if not cert_vals:
                number_patterns = [
                    # ЕАЭС N RU Д-RU.РА03.В.12345/24
                    r'ЕАЭС\s+N\s+RU\s+Д[-\s]+[A-Z]{2}\.[А-Я0-9]{2,6}\.[А-Я]\.\d{4,6}/\d{2}',
                    # ЕАЭС RU С-UZ.НЕ24.В.02715/24
                    r'ЕАЭС\s+RU\s+С[-\s]+[A-Z]{2}\.[А-Я0-9]{2,6}\.[А-Я]\.\d{4,6}/\d{2}',
                    # ROSS RU.НВ19.В12345 (старый формат)
                    r'ROSS\s+[A-Z]{2}\.[А-Я0-9]{4,6}\.[A-Z]\d{4,6}',
                    # ЕАЭС KG 417/021.UZ.02.04606 (КГ — нечасто, но бывает)
                    r'ЕАЭС\s+KG\s*\d{3}/\d{3}\.[A-Z]{2}\.\d{2}\.\d{4,6}',
                ]
                for pat in number_patterns:
                    matches = _re.findall(pat, page_text)
                    if matches:
                        candidate = _re.sub(r'\s+', ' ', matches[0]).strip()
                        cert_vals.append(candidate)
                        details.append(f'number_regex_fallback_ok=pattern_idx_{number_patterns.index(pat)}')
                        break
                if not cert_vals:
                    txt_len = len(page_text) if page_text else 0
                    details.append(f'number_not_found;number_text_len={txt_len}')

            # v39.5: regex-fallback для статуса. Ищем «Статус <слово> <значение>».
            # Возможные значения: «Действует», «Прекращён», «Приостановлен», «Архивный», «Аннулирован».
            if not status_vals:
                # Ищем последовательно: после «Статус сертификата» / «Статус декларации» / «Статус действия»
                status_patterns = [
                    r'Статус\s+сертификата[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                    r'Статус\s+декларации(?:\s+о\s+соответствии)?[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                    r'Статус\s+действия\s+декларации[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                ]
                for pat in status_patterns:
                    m = _re.search(pat, page_text)
                    if m:
                        cand = _re.sub(r'\s+', ' ', m.group(1)).strip()
                        # Нормализация — типовые значения короткие
                        if 1 <= len(cand) <= 50:
                            status_vals.append(cand)
                            details.append('status_regex_fallback_ok')
                            break
                # Дополнительный fallback: ищем сразу значение если оно стоит отдельно как известный статус
                if not status_vals:
                    known_status_pattern = r'\b(Действует|Прекращ[её]н|Приостановлен[а-я]*|Архивн[а-я]+|Аннулирован[а-я]*|Возобновлен[а-я]*)\b'
                    m = _re.search(known_status_pattern, page_text)
                    if m:
                        status_vals.append(m.group(1))
                        details.append('status_keyword_fallback_ok')
        except Exception as e:
            details.append(f'number_regex_error={type(e).__name__}')

    # v27.9.x: если НИ ОДИН number-route не достучался — это сетевой сбой host'а
    # (pub.fsa.gov.ru недоступен/таймаутит). Вкладки /product и /manufacturer того
    # же хоста тоже не загрузятся, поэтому НЕ тратим ещё 60-90с на заведомо
    # провальные goto — сразу выходим с пометкой network failure (её ловит
    # счётчик fsa_network_failures и второй проход FSA). Это убирает «зависание»
    # этапа 2, когда FSA недоступен: запись падает за ~30-60с, а не за ~114с.
    if not any_goto_succeeded:
        details.append('NETWORK_FAILURE_all_goto_failed')
        doc_type = ('декларация' if '/rds/declaration/' in url.lower()
                    else ('сертификат' if '/rss/certificate/' in url.lower() else ''))
        return '', '', doc_type, '', ';'.join(details)

    # Шаг 2: название продукции.
    for route in product_routes:
        try:
            await page.goto(route, wait_until='domcontentloaded', timeout=timeout_ms)
            any_goto_succeeded = True
            # v40.2: быстрый рендер (было 25с, стало 12с)
            await _wait_until_fsa_rendered(page, timeout_ms=min(timeout_ms, 12000))
            await _click_fsa_section_tab_if_exists(page, url, 'product')
            # v39.6.1: на product-странице ищем И продукт, И статус (если ещё не нашли).
            # v40.3: + расширенные поля (заявитель, изготовитель, ТНВЭД, схема, ТР ТС, даты).
            extra_labels = [] if status_vals else status_labels
            fields = await _extract_labels_from_page_v386(
                page, FSA_PRODUCT_LABELS_V386 + extra_labels + FSA_EXTENDED_LABELS_FLAT, wait_ms
            )
            for lab in FSA_PRODUCT_LABELS_V386:
                prod_vals.extend(fields.get(lab, []))
            if not status_vals:
                for lab in status_labels:
                    status_vals.extend(fields.get(lab, []))
                if status_vals:
                    details.append('status_from_product_page')
            # v40.3: собираем расширенные поля (берём первое непустое значение каждого)
            for field_key, labs in FSA_EXTENDED_LABELS.items():
                if field_key in ext_vals:
                    continue
                for lab in labs:
                    vals = fields.get(lab, [])
                    if vals:
                        ext_vals[field_key] = vals
                        break
            if prod_vals:
                details.append(f'product_route={route}')
                break
        except Exception as e:
            err_msg = str(e)[:200].replace(';', ',').replace('\n', ' ')
            details.append(f'product_route_error={type(e).__name__}:{err_msg}')
            continue

    # v44: добираем поля с дополнительных вкладок FSA (изготовитель, ТР ТС, заявитель/ИНН).
    # Эти данные НЕ на /baseInfo и /product — поэтому раньше manufacturer/tech_reg/date_end были 0%.
    # Заходим только за теми полями, которых ещё нет.
    # v46: по требованию НЕ собираем заявителя/изготовителя/ИНН/ТН ВЭД. Это убирает
    # ЗАХОД НА ДОПОЛНИТЕЛЬНЫЕ ВКЛАДКИ ФСА (изготовитель/заявитель лежат не на
    # /baseInfo и /product) — меньше запросов к ФСА, ниже риск блокировки. Номер,
    # статус, даты, схема, техрегламент и название продукции остаются.
    _fsa_skip_org = bool(getattr(args, 'fsa_skip_org_fields', True))
    _ext_want = (("technical_regulation", "document_date_end") if _fsa_skip_org
                 else ("manufacturer_name", "technical_regulation", "applicant_name", "document_date_end"))
    # v46: при skip-org НЕ ходим на доп. вкладки ФСА (/manufacturer, /applicant и т.п.)
    # ВООБЩЕ. Изготовитель/заявитель не нужны, а техрегламент/даты берутся из JSON и
    # baseInfo. Раньше движок всё равно лез на /manufacturer искать недостающие поля —
    # именно там воркеры подвисали по 100с и шли ошибки/блокировки на больших прогонах.
    need_more = (not _fsa_skip_org) and any(k not in ext_vals for k in _ext_want)
    if need_more:
        for ext_route in fsa_extended_routes(url):
            # какие поля ищем на этой вкладке
            still_missing = [k for k in FSA_EXTENDED_LABELS if k not in ext_vals and k in _ext_want]
            if not still_missing:
                break
            try:
                await page.goto(ext_route, wait_until='domcontentloaded', timeout=timeout_ms)
                any_goto_succeeded = True
                await _wait_until_fsa_rendered(page, timeout_ms=min(timeout_ms, 10000))
                labels_to_get = [lab for k in still_missing for lab in FSA_EXTENDED_LABELS[k]]
                fields = await _extract_labels_from_page_v386(page, labels_to_get, wait_ms)
                for field_key in still_missing:
                    for lab in FSA_EXTENDED_LABELS[field_key]:
                        vals = fields.get(lab, [])
                        if vals:
                            ext_vals[field_key] = vals
                            break
                tail = ext_route.rsplit('/', 1)[-1]
                details.append(f'ext_tab_{tail}_fields={sum(1 for k in still_missing if k in ext_vals)}')
            except Exception as e:
                details.append(f'ext_tab_error={type(e).__name__}:{str(e)[:60]}')
                continue

    # v39.2: если НИ ОДИН goto не прошёл — нет смысла скроллить и делать диагностику,
    # это всё равно chrome-error://chromewebdata/. Сразу возвращаем с явной пометкой.
    if not any_goto_succeeded:
        details.append('NETWORK_FAILURE_all_goto_failed')
        cert_vals = []
        prod_vals = []
        doc_type = 'декларация' if '/rds/declaration/' in url.lower() else ('сертификат' if '/rss/certificate/' in url.lower() else '')
        return '', '', doc_type, '', ';'.join(details)

    # v39.4: если на номер не нашли — попробовать regex ещё раз на текущей странице
    # (после product navigation; иногда номер виден на /baseInfo вместе с продукцией).
    # v39.6.1: одновременно делаем regex-fallback для статуса.
    if not cert_vals or not status_vals:
        try:
            page_text2 = await page.evaluate("() => (document.body ? (document.body.innerText || '') : '')")
            import re as _re2
            # Номер
            if not cert_vals:
                for pat in [
                    r'ЕАЭС\s+N\s+RU\s+Д[-\s]+[A-Z]{2}\.[А-Я0-9]{2,6}\.[А-Я]\.\d{4,6}/\d{2}',
                    r'ЕАЭС\s+RU\s+С[-\s]+[A-Z]{2}\.[А-Я0-9]{2,6}\.[А-Я]\.\d{4,6}/\d{2}',
                    r'ROSS\s+[A-Z]{2}\.[А-Я0-9]{4,6}\.[A-Z]\d{4,6}',
                    r'ЕАЭС\s+KG\s*\d{3}/\d{3}\.[A-Z]{2}\.\d{2}\.\d{4,6}',
                ]:
                    matches = _re2.findall(pat, page_text2)
                    if matches:
                        cand = _re2.sub(r'\s+', ' ', matches[0]).strip()
                        cert_vals.append(cand)
                        details.append('number_regex_from_product_page_ok')
                        break
            # v39.6.1: статус — на product-странице, через regex и keyword fallback
            if not status_vals:
                for pat in [
                    r'Статус\s+сертификата[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                    r'Статус\s+декларации(?:\s+о\s+соответствии)?[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                    r'Статус\s+действия\s+декларации[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)',
                ]:
                    m = _re2.search(pat, page_text2)
                    if m:
                        cand = _re2.sub(r'\s+', ' ', m.group(1)).strip()
                        if 1 <= len(cand) <= 50:
                            status_vals.append(cand)
                            details.append('status_regex_from_product_page_ok')
                            break
                # Keyword fallback
                if not status_vals:
                    m2 = _re2.search(
                        r'\b(Действует|Прекращ[её]н[а-я]*|Приостановлен[а-я]*|Архивн[а-я]+|Аннулирован[а-я]*|Возобновлен[а-я]*)\b',
                        page_text2,
                    )
                    if m2:
                        status_vals.append(m2.group(1))
                        details.append('status_keyword_from_product_page_ok')
        except Exception:
            pass

    # v39.1: единственный лёгкий retry — проскроллить страницу и попробовать ещё раз
    # БЕЗ повторной навигации. Это лечит случаи когда Angular рендерит блок только
    # при попадании в viewport. НЕ делаем повторных goto — это вызывало троттлинг.
    if not prod_vals:
        try:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight); window.scrollTo(0, 0);")
            await page.wait_for_timeout(800)
            fields = await _extract_labels_from_page_v386(page, FSA_PRODUCT_LABELS_V386, min(wait_ms, 6000))
            for lab in FSA_PRODUCT_LABELS_V386:
                prod_vals.extend(fields.get(lab, []))
            if prod_vals:
                details.append('product_scroll_ok')
        except Exception as e:
            details.append(f'product_scroll_error={type(e).__name__}')

    # Диагностика если пусто — поможет понять причину
    if not prod_vals:
        try:
            txt_len = await page.evaluate("() => (document.body ? (document.body.innerText || '') : '').length")
            cur_url = await page.evaluate("() => location.href")
            details.append(f'final_page_text_len={txt_len};final_url={cur_url[:120]}')
        except Exception:
            pass

    cert_vals = _unique_keep_order([_registry_value_postclean(FSA_CERT_NUMBER_LABEL, v) for v in cert_vals if 4 <= len(clean_registry_value(v)) <= 300])
    prod_vals = _unique_keep_order([_trim_product_value(_registry_value_postclean(FSA_PRODUCT_LABEL, v)) for v in prod_vals if _looks_like_product_value(v)])
    # v39.5: чистим и нормализуем статус
    status_vals = _unique_keep_order([_normalize_doc_status(v) for v in status_vals if v and len(str(v).strip()) <= 80])
    status_vals = [v for v in status_vals if v]
    doc_type = 'декларация' if '/rds/declaration/' in url.lower() else ('сертификат' if '/rss/certificate/' in url.lower() else '')

    # v40.3: нормализуем и кладём расширенные поля в глобальный кэш, чтобы
    # _flush_url_to_store подставил их в ResultRow (даже когда парсинг шёл браузером,
    # а не curl_cffi). Раньше эти поля заполнял только HTTP-парсер.
    if ext_vals:
        def _first_clean(key: str) -> str:
            vals = ext_vals.get(key) or []
            for v in vals:
                vv = clean_registry_value(v)
                if vv and len(vv) <= 400:
                    return vv
            return ''
        def _norm_date(s: str) -> str:
            s = (s or '').strip()
            mm = re.search(r'(\d{2})[.\-/](\d{2})[.\-/](\d{4})', s)
            if mm:
                return f"{mm.group(1)}.{mm.group(2)}.{mm.group(3)}"
            mm2 = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
            if mm2:
                return f"{mm2.group(3)}.{mm2.group(2)}.{mm2.group(1)}"
            return s[:10]
        prev = _FSA_EXTENDED_FIELDS_CACHE.get(url, {})
        merged = dict(prev)
        # v27.5: для орг-полей и ТН ВЭД применяем специальную чистку.
        for k in ("applicant_name", "manufacturer_name"):
            val = _clean_org_name(_first_clean(k))
            if val and not merged.get(k):
                merged[k] = val
        tnved_val = _clean_tnved_code(_first_clean("tnved"))
        if tnved_val and not merged.get("tnved"):
            merged["tnved"] = tnved_val
        for k in ("applicant_inn", "scheme", "technical_regulation"):
            val = _first_clean(k)
            if val and not merged.get(k):
                merged[k] = val
        for k in ("document_date_start", "document_date_end"):
            val = _norm_date(_first_clean(k))
            if val and not merged.get(k):
                merged[k] = val
        if merged:
            _FSA_EXTENDED_FIELDS_CACHE[url] = merged
            details.append('ext_fields_browser_ok')

    return (
        (cert_vals[0] if cert_vals else ''),
        ('; '.join(prod_vals[:10]) if prod_vals else ''),
        doc_type,
        (status_vals[0] if status_vals else ''),
        ';'.join(details),
    )


async def _parse_swis_with_existing_page_v386(page, url: str, args) -> Tuple[str, str, str, str, str]:
    """v39.5: возвращает (cert, prod, doc_type, doc_status, detail).
    Статус из поля «Признак действия» (SWIS), значения «Действует» и т.п."""
    wait_ms = int(getattr(args, 'registry_browser_wait_ms', 12000))
    timeout_ms = int(getattr(args, 'registry_browser_timeout_ms', 30000))
    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=timeout_ms)
        # v44: расширенные поля SWIS — заявитель/изготовитель/ТН ВЭД/ТР ТС/даты.
        swis_ext_map = {
            "applicant_name": ["Заявитель", "Наименование заявителя"],
            "manufacturer_name": ["Изготовитель", "Наименование изготовителя", "Производитель"],
            "tnved": ["Код ТН ВЭД", "ТН ВЭД ЕАЭС", "Код ТН ВЭД ЕАЭС", "Коды ТН ВЭД"],
            "technical_regulation": ["Технический регламент", "Технические регламенты", "Наименование технического регламента"],
            "document_date_start": ["Дата регистрации", "Действует с", "Дата начала действия"],
            "document_date_end": ["Срок действия", "Действует по", "Дата окончания действия", "Действителен до"],
        }
        swis_ext_flat = [l for ls in swis_ext_map.values() for l in ls]
        labels = [SWIS_NUMBER_LABEL, SWIS_STATUS_LABEL] + SWIS_PRODUCT_LABELS + swis_ext_flat
        fields = await _extract_labels_from_page_v386(page, labels, wait_ms)
        cert_vals = fields.get(SWIS_NUMBER_LABEL, [])
        status_vals = fields.get(SWIS_STATUS_LABEL, [])
        prod_vals = []
        for lab in SWIS_PRODUCT_LABELS:
            prod_vals.extend(fields.get(lab, []))
        # v44: собираем расширенные поля
        swis_ext_vals: Dict[str, str] = {}
        for key, labs in swis_ext_map.items():
            for lab in labs:
                vals = fields.get(lab, [])
                if vals:
                    vv = clean_registry_value(vals[0])
                    if vv and len(vv) <= 400:
                        swis_ext_vals[key] = vv
                        break
        cert_vals = _unique_keep_order([_registry_value_postclean(SWIS_NUMBER_LABEL, v) for v in cert_vals if 4 <= len(clean_registry_value(v)) <= 300])
        prod_vals = _unique_keep_order([_trim_product_value(_registry_value_postclean('продукция', v)) for v in prod_vals if _looks_like_product_value(v)])
        status_vals = _unique_keep_order([_normalize_doc_status(v) for v in status_vals if v and len(str(v).strip()) <= 80])
        status_vals = [v for v in status_vals if v]

        # v39.5: regex-fallback для статуса SWIS — если по лейблу не нашли
        if not status_vals:
            try:
                page_text = await page.evaluate("() => (document.body ? (document.body.innerText || '') : '')")
                import re as _re
                m = _re.search(r'Признак\s+действия[\s:]+([А-ЯЁа-яё]+[А-ЯЁа-яё\s]{0,40}?)(?=[\n\r]|Дата|Орган|Заявитель|Изготовитель|Регистрационный|$)', page_text)
                if m:
                    cand = _normalize_doc_status(_re.sub(r'\s+', ' ', m.group(1)).strip())
                    if cand:
                        status_vals.append(cand)
                if not status_vals:
                    m2 = _re.search(r'\b(Действует|Прекращ[её]н|Приостановлен[а-я]*|Архивн[а-я]+|Аннулирован[а-я]*|Возобновлен[а-я]*)\b', page_text)
                    if m2:
                        status_vals.append(_normalize_doc_status(m2.group(1)))
            except Exception:
                pass

        # v44: расширенные поля SWIS → в общий кэш (как у FSA/Belgiss)
        # v27.5: орг-поля и ТН ВЭД чистим специальными функциями.
        if swis_ext_vals:
            def _nd(s):
                s = (s or '').strip()
                mm = re.search(r'(\d{2})[.\-/](\d{2})[.\-/](\d{4})', s)
                return f"{mm.group(1)}.{mm.group(2)}.{mm.group(3)}" if mm else s[:10]
            merged = dict(_FSA_EXTENDED_FIELDS_CACHE.get(url, {}))
            for k in ("applicant_name", "manufacturer_name"):
                cleaned = _clean_org_name(swis_ext_vals.get(k, ''))
                if cleaned and not merged.get(k):
                    merged[k] = cleaned
            cleaned_tnved = _clean_tnved_code(swis_ext_vals.get("tnved", ''))
            if cleaned_tnved and not merged.get("tnved"):
                merged["tnved"] = cleaned_tnved
            if swis_ext_vals.get("technical_regulation") and not merged.get("technical_regulation"):
                merged["technical_regulation"] = swis_ext_vals["technical_regulation"]
            for k in ("document_date_start", "document_date_end"):
                if swis_ext_vals.get(k) and not merged.get(k):
                    merged[k] = _nd(swis_ext_vals[k])
            if merged:
                _FSA_EXTENDED_FIELDS_CACHE[url] = merged

        return (
            (cert_vals[0] if cert_vals else ''),
            ('; '.join(prod_vals[:10]) if prod_vals else ''),
            'сертификат',
            (status_vals[0] if status_vals else ''),
            'swis_browser_exact' + (';swis_ext_ok' if swis_ext_vals else ''),
        )
    except Exception as e:
        return '', '', '', '', f'swis_browser_error={type(e).__name__}: {str(e)[:120]}'


# ---------------------------------------------------------------------------
# v27.7: HTTP-парсер киргизского реестра SWIS (требование: только HTTP-путь).
# Страница swis.trade.kg отдаётся сервером готовым HTML (НЕ SPA), поэтому
# браузер не нужен. Данные лежат в <table> внутри
# <div class="ComplianceDeclaration_public_table_view"> строками вида
# <tr><td>Лейбл</td><td>Значение</td></tr>, разделы — <th colspan="2">.
# Разбор ведётся С УЧЁТОМ текущего раздела, потому что лейбл «Полное
# наименование» встречается и у Заявителя, и у Изготовителя.
# ---------------------------------------------------------------------------

def _swis_cell_text(td) -> str:
    """Текст ячейки SWIS с сохранением кусочков <span class="Value"> через пробел."""
    spans = td.find_all("span", class_="Value")
    if spans:
        parts = [s.get_text(" ", strip=True) for s in spans]
        return clean_registry_value(" ".join(p for p in parts if p))
    return clean_registry_value(td.get_text(" ", strip=True))


def parse_swis_http_full(text: str, url: str = "") -> Dict[str, str]:
    """Разбирает HTML-страницу SWIS (сертификат/декларация ТС или KGZ-National).

    Возвращает dict со всеми извлечёнными полями. Раздел отслеживается по
    строкам-заголовкам <th>, чтобы корректно различать «Полное наименование»
    Заявителя и Изготовителя.
    """
    out: Dict[str, str] = {}
    homogeneous: List[str] = []
    full_products: List[str] = []
    generic_products: List[str] = []  # v27.9.x: запасной захват наименования продукции
    tnved_codes: List[str] = []

    if not BeautifulSoup:
        return out
    soup = BeautifulSoup(text or "", "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    section = ""  # текущий раздел (Заявитель / Изготовитель / ...)
    for tr in soup.find_all("tr"):
        th = tr.find("th")
        if th is not None:
            section = norm_text(th.get_text(" ", strip=True))
            continue
        tds = tr.find_all("td", recursive=False) or tr.find_all("td")
        if len(tds) < 2:
            continue
        label = norm_text(tds[0].get_text(" ", strip=True))
        value = _swis_cell_text(tds[1])
        if not label or not value:
            continue

        # --- Общие сведения ---
        if "регистрационный номер документа" in label and not out.get("cert_number"):
            out["cert_number"] = value
        elif label.startswith("учетный номер") and not out.get("blank_number"):
            out["blank_number"] = value
        elif "дата начала действия" in label and not out.get("date_start"):
            out["date_start"] = value
        elif "дата окончания действия" in label and not out.get("date_end"):
            out["date_end"] = value
        elif "признак действия" in label and not out.get("status"):
            out["status"] = _normalize_doc_status(value)

        # --- Заявитель ---
        elif "заявител" in section:
            if label == "полное наименование" and not out.get("applicant_name"):
                out["applicant_name"] = _clean_org_name(value)
            elif label == "инн" and not out.get("applicant_inn"):
                out["applicant_inn"] = value

        # --- Изготовитель ---
        elif "изготовител" in section:
            if "полное наименование" in label and not out.get("manufacturer_name"):
                out["manufacturer_name"] = _clean_org_name(value)

        # --- Сведения о продукции / Товар N (раздел может быть любым) ---
        if "однородное наименование продукции" in label:
            v = _trim_product_value(value)
            if _looks_like_product_value(v):
                homogeneous.append(v)
        elif "полное наименование продукции" in label and "идентификац" in label:
            v = _trim_product_value(value)
            if _looks_like_product_value(v) and not _swis_value_is_not_product(v):
                full_products.append(v)
        # v27.9.x: ЗАПАСНОЙ захват наименования продукции для других вёрсток SWIS
        # (KGZ-National / упрощённые декларации), где нет слов «однородное» или
        # «идентификац». Срабатывает только если выше точные метки не подошли —
        # это чинит пустое «название из реестра» по части киргизских документов.
        elif (("наименование продукции" in label or "наименование товара" in label
               or "наименование объекта" in label)
              and "однородн" not in label):
            v = _trim_product_value(value)
            if _looks_like_product_value(v) and not _swis_value_is_not_product(v):
                generic_products.append(v)
        elif "схема сертификации" in label and not out.get("scheme"):
            out["scheme"] = value[:40]
        elif ("обозначение тр" in label or "технического регламента" in label or
              "технический регламент" in label) and not out.get("technical_regulation"):
            out["technical_regulation"] = value[:400]
        elif ("тн вэд" in label) and value not in tnved_codes:
            code = _clean_tnved_code(value)
            if code:
                tnved_codes.append(code)

    # Итоговое наименование продукции: однородное (чистое для сверки) + детальные
    homogeneous = _unique_keep_order(homogeneous)
    full_products = _unique_keep_order(full_products)
    product_parts: List[str] = []
    product_parts.extend(homogeneous[:5])
    product_parts.extend(full_products[:20])
    # v27.9.x: если точные метки ничего не дали — берём запасные наименования.
    if not product_parts and generic_products:
        product_parts.extend(_unique_keep_order(generic_products)[:10])
    if product_parts:
        out["product"] = "; ".join(_unique_keep_order(product_parts))
    if tnved_codes:
        out["tnved"] = "; ".join(tnved_codes[:10])

    # Тип документа — ТОЛЬКО по заголовку страницы (в теле есть навигация с
    # обоими словами). Заголовок SWIS иногда содержит латинскую 'c' в слове
    # «cертификате» — нормализуем латиницу в кириллицу перед проверкой.
    head_txt = norm_text(soup.title.get_text() if soup.title else "")
    head_txt = head_txt.replace("c", "с").replace("a", "а").replace("o", "о").replace("e", "е")
    if "деклараци" in head_txt:
        out["doc_type"] = "декларация"
    else:
        out["doc_type"] = "сертификат"
    return out


async def _parse_swis_http_v277(session, url: str, args) -> Tuple[str, str, str, str, str]:
    """SWIS только по HTTP (без браузера). Возвращает 5-tuple
    (cert, product, doc_type, doc_status, detail) и кладёт расширенные поля
    (заявитель/изготовитель/ТН ВЭД/ТР/даты/схема) в общий кэш, как у FSA."""
    timeout = float(getattr(args, "registry_http_timeout", 30.0) or 30.0)
    status_code, text, ct = await http_get(session, url, timeout=timeout)
    if not text:
        return "", "", "", "", f"swis_http_no_body(status={status_code})"
    try:
        data = parse_swis_http_full(text, url)
    except Exception as e:
        return "", "", "", "", f"swis_http_parse_error={type(e).__name__}: {str(e)[:120]}"

    # Расширенные поля → общий кэш (подставляются в ResultRow на этапе flush)
    merged = dict(_FSA_EXTENDED_FIELDS_CACHE.get(url, {}))
    for k in ("applicant_name", "applicant_inn", "manufacturer_name", "tnved",
              "technical_regulation", "scheme", "date_start", "date_end"):
        v = data.get(k)
        if not v:
            continue
        target = {"date_start": "document_date_start",
                  "date_end": "document_date_end"}.get(k, k)
        if not merged.get(target):
            merged[target] = v
    if merged:
        _FSA_EXTENDED_FIELDS_CACHE[url] = merged

    detail = "swis_http_ok" + (";ext_ok" if merged else "")
    return (
        data.get("cert_number", ""),
        data.get("product", ""),
        data.get("doc_type", "сертификат"),
        data.get("status", ""),
        detail,
    )


async def _parse_belgiss_with_existing_page_v42(page, url: str, args) -> Tuple[str, str, str, str, str]:
    """v27.5: УПРОЩЁННЫЙ парсер реестра ЕАЭС Belgiss (tsouz.belgiss.by).

    По требованию пользователя — с белорусского реестра берём ТОЛЬКО номер
    сертификата/декларации. Все остальные поля (applicant/manufacturer/product/
    status/dates) НЕ извлекаем, потому что DOM Belgiss отдаёт их склеенными с
    заголовками таблиц («Страна (BY)», «Краткое наименование хозяйствующего
    субъекта», «оценки соответствия», «действия сертификата (декларации)»),
    и их корректная разборка нерентабельна.

    Возвращает (cert_number, '', doc_type, '', detail) — только номер и тип.
    """
    timeout_ms = int(getattr(args, 'registry_browser_timeout_ms', 30000))
    details: List[str] = []
    cert_vals: List[str] = []

    try:
        await page.goto(url, wait_until='domcontentloaded', timeout=timeout_ms)
    except Exception as e:
        return '', '', '', '', f'belgiss_goto_error={type(e).__name__}:{str(e)[:120]}'

    # Ждём пока SPA наполнит DOM.
    try:
        await page.wait_for_function(
            "() => document.body && (document.body.innerText||'').length > 400",
            timeout=min(timeout_ms, 12000),
        )
    except Exception:
        pass
    try:
        text_len = await page.evaluate("() => (document.body ? (document.body.innerText||'').length : 0)")
    except Exception:
        text_len = -1
    details.append(f'belgiss_render_text_len={text_len}')

    # Сначала пробуем взять номер из URL (для /certifs/<id>/view это не работает,
    # но дальше у нас regex по тексту страницы).
    # Главное — regex по тексту страницы для номера документа ЕАЭС.
    try:
        body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
    except Exception:
        body_text = ''
    for pat in (
        r'ЕАЭС\s+BY/[0-9.\s]+\s*ТР\d+\s+[0-9.\s]+',  # ЕАЭС BY/112 02.01. ТР007 118.01 02748
        r'ЕАЭС\s+[NN№]?\s*RU\s*[ДД]-[A-ZА-Я0-9./\-]+',
        r'ЕАЭС\s+RU\s*[СC]-[A-ZА-Я0-9./\-]+',
        r'[NN№]\s*BY/[0-9./\-]+',
        r'ROSS\s+[A-ZА-Я0-9./\-]+',
    ):
        mm = re.search(pat, body_text or '')
        if mm:
            cert_vals.append(re.sub(r'\s+', ' ', mm.group(0)).strip())
            details.append('belgiss_number_from_regex')
            break

    # Если regex не сработал — пробуем по подписям (только номер).
    if not cert_vals:
        number_labels = [
            "Регистрационный номер", "Номер документа", "Номер сертификата",
            "Номер декларации", "Рег. номер",
        ]
        try:
            fields = await _extract_labels_from_page_v386(page, number_labels, 4000)
        except Exception:
            fields = {}
        for lab in number_labels:
            cert_vals.extend(fields.get(lab, []))
        cert_vals = _unique_keep_order([v for v in cert_vals if 4 <= len(clean_registry_value(v)) <= 300])
        if cert_vals:
            details.append('belgiss_number_from_labels')

    low = url.lower()
    doc_type = 'декларация' if 'declaration' in low or '/decl' in low else ('сертификат' if 'cert' in low else '')

    # ВАЖНО: НЕ заполняем _FSA_EXTENDED_FIELDS_CACHE — у Belgiss поля грязные,
    # пользователь явно сказал не фиксить, брать только номер.
    details.append('belgiss_minimal_mode')

    return (
        (cert_vals[0] if cert_vals else ''),
        '',  # product_name — пусто
        doc_type,
        '',  # doc_status — пусто
        'belgiss_browser; ' + ';'.join(details),
    )


async def _parse_other_registry_http_v386(session, url: str, args) -> Tuple[str, str, str, str, str]:
    # Keep legacy support for non-FSA/SWIS registries. This does not affect the two main registries.
    # v39.5: возвращает 5-tuple (cert, prod, type, status, detail). status у legacy парсера пока пустой.
    best_cert = best_product = best_type = ''
    for u in api_candidates_for_registry(url):
        status, text, ct = await http_get(session, u, timeout=getattr(args, 'registry_http_timeout', 30.0))
        if not text:
            continue
        cert_num = product_name = doc_type = ''
        if 'json' in ct.lower() or text.strip().startswith(('{', '[')):
            try:
                cert_num, product_name, doc_type = pick_from_json(json.loads(text))
            except Exception:
                pass
        if not cert_num or not product_name:
            h_cert, h_prod, h_type = parse_html_registry(u, text)
            cert_num = cert_num or h_cert
            product_name = product_name or h_prod
            doc_type = doc_type or h_type
        best_cert = best_cert or cert_num
        best_product = best_product or product_name
        best_type = best_type or doc_type
        if best_cert and best_product:
            break
    return best_cert, best_product, best_type, '', 'other_http_legacy'


# =============================================================================
# v27.9.x: БРАУЗЕРНЫЙ сбор FSA БЕЗ блокировок (без прокси). Маскируем headless под
# обычный Chrome (stealth), прогреваем сессию (cookies) и держим человеческий темп
# (случайные паузы) — так антибот FSA не банит IP, и данные собираются стабильно.
# =============================================================================
_FSA_STEALTH_JS = r"""
(() => {
  try { Object.defineProperty(navigator, 'webdriver', {get: () => undefined}); } catch(e){}
  try { Object.defineProperty(navigator, 'languages', {get: () => ['ru-RU','ru','en-US','en']}); } catch(e){}
  try { Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]}); } catch(e){}
  try { window.chrome = window.chrome || {runtime: {}}; } catch(e){}
  try {
    const q = window.navigator.permissions && window.navigator.permissions.query;
    if (q) window.navigator.permissions.query = (p) =>
      (p && p.name === 'notifications') ? Promise.resolve({state: Notification.permission}) : q(p);
  } catch(e){}
  try {
    const gp = WebGLRenderingContext.prototype.getParameter;
    WebGLRenderingContext.prototype.getParameter = function(p){
      if (p === 37445) return 'Intel Inc.';
      if (p === 37446) return 'Intel Iris OpenGL Engine';
      return gp.call(this, p);
    };
  } catch(e){}
  try { Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8}); } catch(e){}
  try { Object.defineProperty(navigator, 'deviceMemory', {get: () => 8}); } catch(e){}
})();
"""


async def _apply_stealth(ctx) -> None:
    """Навешивает stealth-init-script на контекст — прячет признаки автоматизации."""
    try:
        await ctx.add_init_script(_FSA_STEALTH_JS)
    except Exception:
        pass


async def _fsa_warmup_context(page, timeout_ms: int = 15000) -> None:
    """Прогрев сессии: заходим на главную FSA (cookies/токены), чтобы дальнейшие
    запросы к документам выглядели продолжением обычного визита."""
    try:
        await page.goto("https://pub.fsa.gov.ru/", wait_until='domcontentloaded', timeout=timeout_ms)
        try:
            await page.wait_for_timeout(random.randint(700, 1600))
        except Exception:
            pass
    except Exception:
        pass


def _fsa_human_delay_range_ms(raw: str, default=(300, 1400)) -> Tuple[float, float]:
    """Разбирает строку «min,max» (мс) в диапазон секунд (lo, hi)."""
    try:
        parts = [int(x) for x in re.split(r'[,\s]+', str(raw).strip()) if x]
        lo = parts[0] if parts else default[0]
        hi = parts[1] if len(parts) > 1 else lo
        return max(0, lo) / 1000.0, max(lo, hi) / 1000.0
    except Exception:
        return default[0] / 1000.0, default[1] / 1000.0


def _fsa_human_delay_range(args) -> Tuple[float, float]:
    """Диапазон случайной паузы между FSA-документами (сек). --fsa-human-delay-ms «min,max»."""
    return _fsa_human_delay_range_ms(getattr(args, 'fsa_human_delay_ms', '300,1400') or '300,1400')


def _load_prior_registry_parsed(xlsx_path: Path) -> Dict[str, Tuple[str, str, str, str, str]]:
    """v46: читает предыдущий result.xlsx (лист «Подробности») и возвращает
    {registry_url: (cert, prod, typ, doc_status, detail)} — чтобы «Повторить
    упавшие FSA» НЕ пере-парсил уже успешно собранные реестры, а переносил их
    как есть и трогал только упавшие FSA-ссылки."""
    out: Dict[str, Tuple[str, str, str, str, str]] = {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return out
    if not xlsx_path.exists():
        return out
    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return out
    try:
        ws = wb['Подробности'] if 'Подробности' in wb.sheetnames else wb[wb.sheetnames[-1]]
        rows_iter = ws.iter_rows(values_only=True)
        hdr = list(next(rows_iter))
        ru2field = {v: k for k, v in DETAILS_HEADERS_RU_V39.items()}
        idx: Dict[str, int] = {}
        for n, h in enumerate(hdr):
            f = ru2field.get(h)
            if f is not None:
                idx[f] = n
        if 'registry_url' not in idx:
            return out
        def _cell(r, field):
            j = idx.get(field)
            return '' if j is None or j >= len(r) or r[j] is None else str(r[j])
        for r in rows_iter:
            url = clean_url(_cell(r, 'registry_url'))
            if not url or url in out:
                continue
            out[url] = (
                _cell(r, 'certificate_number'), _cell(r, 'certificate_product_name'),
                _cell(r, 'document_type'), _cell(r, 'document_status'),
                _cell(r, 'details') or 'prior_result',
            )
    except Exception:
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


async def run_registry_stage(args):
    """v38.6: visible-browser registry stage with first-stage-like progress.

    FSA and SWIS are parsed from the rendered page by exact user-confirmed fields.
    HTTP is used only for other allowed registries.
    """
    _run_started_at = time.time()

    # v46: загружаем таблицу статусов киргизских документов на территории РФ.
    # Явный путь из --kg-rf-status-file, иначе kg_rf_status.xlsx рядом с программой.
    _kg_file = (getattr(args, 'kg_rf_status_file', '') or '').strip()
    if not _kg_file:
        for _cand in ('kg_rf_status.xlsx', 'kg_rf_status.csv'):
            if Path(_cand).exists():
                _kg_file = _cand
                break
    if _kg_file:
        try:
            _kg_n = load_kg_rf_status(_kg_file)
            if _kg_n:
                print(f"🇰🇬 Таблица статусов КГ-документов в РФ загружена: {_kg_n} записей "
                      f"(совпавшие получат «Статус на территории РФ» и вердикт «{STATUS_INVALID_IN_RF}»).")
        except Exception as _e:
            print(f"⚠️  Таблица статусов КГ-РФ не загружена: {type(_e).__name__}: {_e}")

    # v27.9.x: глушим КОСМЕТИЧЕСКИЕ ошибки event-loop'а вида «Future exception was
    # never retrieved» с net::ERR_ABORTED / «frame was detached». Они возникают,
    # когда жёсткий per-record timeout отменяет навигацию Playwright в момент
    # перехода между вкладками FSA (например .../manufacturer): навигация затем
    # завершается с ERR_ABORTED, но её результат уже никто не ждёт. На итог это
    # не влияет, но в логе выглядит как «программа выдала ошибку». Прочие ошибки
    # пропускаем дальше без изменений.
    try:
        _loop = asyncio.get_running_loop()
        _prev_exc_handler = _loop.get_exception_handler()

        def _quiet_nav_exc_handler(loop, context):
            exc = context.get('exception')
            text = f"{context.get('message', '')} {type(exc).__name__ if exc else ''}: {exc if exc else ''}"
            for marker in ('ERR_ABORTED', 'frame was detached',
                           'Target page, context or browser has been closed',
                           'ERR_CONNECTION_TIMED_OUT', 'ERR_TIMED_OUT', 'ERR_NETWORK_CHANGED'):
                if marker in text:
                    return  # косметика отменённой/упавшей навигации — не шумим
            if _prev_exc_handler is not None:
                _prev_exc_handler(loop, context)
            else:
                loop.default_exception_handler(context)

        _loop.set_exception_handler(_quiet_nav_exc_handler)
    except Exception:
        pass

    input_csv = Path(args.input_links_csv)
    if not input_csv.exists():
        raise FileNotFoundError(input_csv)

    rows: List[Dict[str, str]] = []
    with input_csv.open('r', encoding='utf-8-sig', newline='') as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            url = clean_url(row.get('registry_url', ''))
            if url and is_allowed_registry_url(url):
                row['registry_url'] = url
                rows.append(row)
            if len(rows) >= args.limit:
                break
    # v25-reporting: строгий бренд-фильтр (см. Stage 1).
    _brand_wanted = getattr(args, "brand", "") or ""
    _brand_mode = getattr(args, "brand_match", "any") or "any"
    if _brand_wanted and _brand_mode != "any":
        before = len(rows)
        rows = [r for r in rows if brand_matches_v39(r.get("brand", ""), _brand_wanted, _brand_mode)]
        print(f"Бренд-фильтр ({_brand_mode}, '{_brand_wanted}'): отфильтровано {before - len(rows)} строк, осталось {len(rows)}")
    print(f"К проверке реестров подготовлено строк: {len(rows)}")
    unique_urls = list(dict.fromkeys(clean_url(r.get('registry_url', '')) for r in rows if r.get('registry_url')))
    print(f"Уникальных ссылок на реестры: {len(unique_urls)}")
    if not rows:
        print('Нет строк с registry_url для второго этапа.')
        return

    # v46: режим «Повторить упавшие FSA». Раньше второй этап ПОЛНОСТЬЮ
    # перезапускался — пере-парсил ВСЕ реестры (включая уже успешные). Теперь в
    # этом режиме из предыдущего result.xlsx переносятся все успешные документы,
    # а пере-проверяются ТОЛЬКО упавшие FSA-ссылки (нет номера/названия/статуса).
    _retry_seed: Dict[str, Tuple[str, str, str, str, str]] = {}
    if bool(getattr(args, 'registry_fsa_retry', False)):
        prior = _load_prior_registry_parsed(Path(args.output))
        if not prior:
            print("🔁 Повтор FSA: предыдущий result.xlsx не найден/не прочитан — выполняю обычный полный второй этап.")
        else:
            def _fsa_failed_prior(u: str) -> bool:
                if hostname(u) != 'pub.fsa.gov.ru':
                    return False
                cert, prod, typ, dst, det = prior.get(u, ('', '', '', '', ''))
                return not (cert or '').strip() or not (prod or '').strip() or not (dst or '').strip()
            failed = [u for u in unique_urls if _fsa_failed_prior(u)]
            if not failed:
                print(f"🔁 Повтор FSA: упавших FSA-ссылок в предыдущем result.xlsx не найдено — "
                      f"перепроверять нечего. Файл будет пересобран из прежних данных.")
            else:
                print(f"🔁 Повтор FSA: из {len(unique_urls)} реестров будут пере-проверены только "
                      f"{len(failed)} упавших FSA-ссылок; остальные перенесены из предыдущего result.xlsx.")
            # Сидируем прежними значениями ВСЁ, кроме упавших (их соберём заново).
            _failed_set = set(failed)
            for u, val in prior.items():
                if u not in _failed_set:
                    _retry_seed[u] = val
            # Первый проход трогает только упавшие FSA-ссылки.
            unique_urls = failed

    # Default second stage to browser-visible parsing for main registries. It is slower, but substantially more reliable for FSA SPA.
    browser_workers = max(1, int(getattr(args, 'registry_browser_workers', 2)))
    args.registry_headless = getattr(args, 'registry_headless', True)
    args.registry_browser_wait_ms = int(getattr(args, 'registry_browser_wait_ms', 12000))
    args.registry_browser_timeout_ms = int(getattr(args, 'registry_browser_timeout_ms', 30000))

    parsed: Dict[str, Tuple[str, str, str, str]] = {}
    # v46: переносим успешные документы из предыдущего прогона (режим повтора FSA),
    # чтобы их строки записались без повторного парсинга реестра.
    for _u, _val in _retry_seed.items():
        parsed.setdefault(_u, _val)
    q: asyncio.Queue[str] = asyncio.Queue()
    for u in unique_urls:
        q.put_nowait(u)

    # v39.1: группируем rows по registry_url, чтобы сразу как только реестр распарсен
    # записать в out_store все ResultRow для всех товаров с этим реестром.
    # Это даёт промежуточный xlsx во время работы, а не только в конце.
    rows_by_url: Dict[str, List[Dict[str, str]]] = {}
    for r in rows:
        u = clean_url(r.get('registry_url', ''))
        rows_by_url.setdefault(u, []).append(r)

    stats = {
        'started_at': time.time(), 'done': 0, 'ok': 0, 'empty': 0, 'errors': 0,
        'fsa': sum(1 for u in unique_urls if hostname(u) == 'pub.fsa.gov.ru'),
        'swis': sum(1 for u in unique_urls if hostname(u) in {'swis.trade.kg', 'trade.kg'}),
        # v27.9.x: счётчик BelGISS/ЕАЭС — чтобы он попадал в живой график «Реестры».
        'belgiss': sum(1 for u in unique_urls if hostname(u) in _BELGISS_EAEU_HOSTS),
        'rows_written': 0,
        # v39.2: счётчики для раннего детекта сетевой недоступности FSA
        'fsa_done': 0,
        'fsa_network_failures': 0,
        'fsa_warning_shown': False,
    }
    active: Dict[str, Tuple[str, float]] = {}
    out_store = ResultStore(
        Path(args.output),
        None,
        expiry_warning_days=getattr(args, "expiry_warning_days", 30),
        make_report_xlsx=getattr(args, "make_report_xlsx", True),
    )

    async def _flush_url_to_store(url: str, cert: str, prod: str, typ: str, doc_status: str, parse_detail: str):
        """v39.1: сразу записать ResultRow для всех товаров с этим registry_url.
        Это даёт промежуточный xlsx во время парсинга.
        v39.5: добавлен doc_status.
        v39.14: подставляем расширенные поля из _FSA_EXTENDED_FIELDS_CACHE + WB-поля из row."""
        # v39.14: достаём расширенные поля документа из глобального кэша
        ext = _FSA_EXTENDED_FIELDS_CACHE.get(url, {})
        _host = hostname(url)
        _skip_org_fields = bool(getattr(args, 'fsa_skip_org_fields', True))
        for row in rows_by_url.get(url, []):
            # v27.9.x: BelGISS/ЕАЭС по требованию НЕ парсим — просто оставляем
            # ссылку на реестр (статус «собрана»), без вердикта «не удалось».
            if _host in _BELGISS_EAEU_HOSTS:
                verdict, score, cmp_details = STATUS_LINK_COLLECTED, 0.0, 'belgiss_link_only'
            else:
                verdict, score, cmp_details = compare_product_names(
                    row.get('product_name', ''), prod,
                    brand=row.get('brand', ''), subject=row.get('subject', ''),
                    doc_status=doc_status,
                )
                if not prod and verdict != 'НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ':
                    verdict = 'НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА'
                    score = 0.0
            # v46: киргизский документ из таблицы статусов РФ — отдельный вердикт
            rf_status = kg_rf_status_text(cert)
            if rf_status:
                verdict = STATUS_INVALID_IN_RF
            rr = ResultRow(
                query=row.get('query', ''),
                nm_id=safe_int(row.get('nm_id')),
                product_name=row.get('product_name', ''),
                brand=row.get('brand', ''),
                subject=row.get('subject', ''),
                product_url=row.get('product_url', ''),
                status=verdict,
                # v39.14: WB-поля из CSV (если они там были)
                price_rub=safe_float(row.get('price_rub')),
                sale_price_rub=safe_float(row.get('sale_price_rub')),
                seller_name=row.get('seller_name', ''),
                is_original=row.get('is_original', ''),
                # пусто (старый CSV без признака) → «Нет»: бейдж не зафиксирован
                docs_verified=row.get('docs_verified') or 'Нет',
                supplier_id=row.get('supplier_id', ''),
                rating=safe_float(row.get('rating')),
                feedbacks=safe_int(row.get('feedbacks')),
                registry_url=url,
                registry_host=hostname(url),
                registry_record_id=extract_record_id(url),
                certificate_number=cert,
                document_type=typ,
                document_status=doc_status,
                rf_status=rf_status,
                certificate_product_name=prod,
                # v39.14: расширенные поля FSA-документа
                # v27.5: двойная защита — дожимаем орг-поля и ТН ВЭД, если в кэше остался мусор.
                document_date_start=ext.get('document_date_start', ''),
                document_date_end=ext.get('document_date_end', ''),
                applicant_name=('' if _skip_org_fields else _clean_org_name(ext.get('applicant_name', ''))),
                applicant_inn=('' if _skip_org_fields else ext.get('applicant_inn', '')),
                manufacturer_name=('' if _skip_org_fields else _clean_org_name(ext.get('manufacturer_name', ''))),
                tnved=('' if _skip_org_fields else _clean_tnved_code(ext.get('tnved', ''))),
                scheme=ext.get('scheme', ''),
                technical_regulation=ext.get('technical_regulation', ''),
                score=score,
                details=(f'registry_browser_stage; {parse_detail}; {cmp_details}' if prod else f'empty_registry_product_name; {parse_detail}'),
                checked_at=now_iso(),
            )
            await out_store.add(rr)
            stats['rows_written'] += 1

    async def progress_loop():
        try:
            while stats['done'] < len(unique_urls):
                elapsed = max(1.0, time.time() - stats['started_at'])
                speed = stats['done'] / elapsed * 60.0
                ok_speed = stats['ok'] / elapsed * 60.0
                act = '; '.join(f"{k}:{extract_record_id(v[0]) or hostname(v[0])}:{int(time.time()-v[1])}s" for k, v in list(active.items())[:browser_workers+3])
                print(
                    f"Реестры: {stats['done']}/{len(unique_urls)}, скорость≈{speed:.1f}/мин, "
                    f"извлечено={stats['ok']}, пусто={stats['empty']}, ошибки={stats['errors']}, "
                    f"строк_в_xlsx={stats['rows_written']}/{len(rows)}, "
                    f"очередь={q.qsize()}, FSA={stats['fsa']}, SWIS={stats['swis']}, "
                    f"BELGISS={stats['belgiss']}, активные=[{act}]"
                )
                emit_progress("registry", stats['done'], len(unique_urls))
                await asyncio.sleep(max(5, min(30, int(getattr(args, 'progress_interval_sec', 15)))))
        except asyncio.CancelledError:
            return

    # v39.1: отдельная фоновая задача — периодически сохраняет xlsx.
    # Это критично: без него файл появляется только в конце прогона.
    async def saver_loop():
        save_interval = max(30, min(120, int(getattr(args, 'progress_interval_sec', 15)) * 2))
        last_saved = 0
        try:
            while stats['done'] < len(unique_urls):
                await asyncio.sleep(save_interval)
                if stats['rows_written'] > last_saved:
                    try:
                        await out_store.save()
                        last_saved = stats['rows_written']
                        print(f"💾 Промежуточное сохранение: {last_saved}/{len(rows)} строк в {Path(args.output).name}")
                    except Exception as e:
                        print(f"⚠️  Ошибка сохранения xlsx: {type(e).__name__}: {e}")
        except asyncio.CancelledError:
            return

    progress_task = asyncio.create_task(progress_loop())
    saver_task = asyncio.create_task(saver_loop())

    async with aiohttp.ClientSession(
        timeout=aiohttp.ClientTimeout(total=max(60, int(getattr(args, 'registry_http_timeout', 30)) + 20)),
        connector=aiohttp.TCPConnector(limit=max(10, int(getattr(args, 'http_connector_limit', 300))), ssl=False, ttl_dns_cache=300),
        headers={'User-Agent': args.user_agent, 'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.7'},
    ) as session:
        if async_playwright is None:
            raise RuntimeError('playwright не установлен. Выполните: python -m pip install playwright && python -m playwright install chromium')
        async with async_playwright() as p:
            # v45: FSA парсится ТОЛЬКО через браузер (без прокси). Чтобы не нарваться
            # на блокировку — маскируем headless-Chromium под обычный браузер (stealth),
            # прогреваем сессию на главной FSA (cookies) и делаем человекоподобные
            # случайные паузы между документами.
            _has_fsa = any(hostname(u) == 'pub.fsa.gov.ru' for u in unique_urls)
            # v45.4: замок бутстрапа кук FSA — пока кук нет, браузерные FSA-документы
            # идут по одному (а не залпом из 5 сессий), чтобы первый прошёл антибот и
            # отдал куки; дальше всё летит быстрым HTTP параллельно.
            _fsa_bootstrap_sem = asyncio.Semaphore(1)
            # v46: МЕДЛЕННЫЙ РЕЖИМ ФСА — для больших прогонов (до 10k) без блокировок.
            _fsa_slow_mode = bool(getattr(args, 'fsa_slow_mode', False))
            _fsa_serial_sem = asyncio.Semaphore(1)  # ФСА строго по одному при slow-mode
            _fsa_last_req = [0.0]  # время последнего запроса к ФСА (для РЕАЛЬНОЙ паузы между запросами)
            _fsa_slow_lo, _fsa_slow_hi = _fsa_human_delay_range_ms(
                getattr(args, 'fsa_slow_delay_ms', '2000,3500') or '2000,3500')
            if _fsa_slow_mode and _has_fsa:
                print(f"🐢 МЕДЛЕННЫЙ режим ФСА: документы по одному, пауза "
                      f"{_fsa_slow_lo:.1f}-{_fsa_slow_hi:.1f}с + адаптивный бэкофф. "
                      f"Без блокировок, но небыстро (≈{60.0/max(0.1,(_fsa_slow_lo+_fsa_slow_hi)/2):.0f} док/мин). "
                      f"SWIS/прочие реестры идут параллельно.")
            _launch_kwargs = dict(
                headless=getattr(args, 'registry_headless', True),
                args=['--disable-dev-shm-usage', '--no-sandbox', '--disable-blink-features=AutomationControlled'],
            )
            browser = await p.chromium.launch(**_launch_kwargs)
            _browser_restart = [0.0]  # v46: метка последнего перезапуска браузера (анти-дребезг)

            async def worker(wid: int):
                global _FSA_CONSEC_FAILS, _FSA_COOLDOWN_UNTIL, _FSA_COOLDOWN_CYCLES, _FSA_SESSION_COOKIES
                global _FSA_SLOW_MULT, _FSA_SLOW_OK
                # v39: единая функция пересоздания контекста + страницы. Если что-то падает —
                # вернёт хотя бы пустой page, чтобы worker не умер.
                async def _fresh_context_and_page(old_ctx):
                    nonlocal browser
                    try:
                        if old_ctx is not None:
                            await old_ctx.close()
                    except Exception:
                        pass
                    try:
                        new_ctx = await browser.new_context(
                            user_agent=args.user_agent,
                            viewport={'width': 1440, 'height': 1000},
                            locale='ru-RU',
                        )
                    except Exception:
                        # v46: Chromium мог упасть (OOM на длинных FSA-прогонах). Раньше это
                        # валило всех воркеров и прогон завершался кодом 1. Перезапускаем
                        # браузер и продолжаем — частичный результат уже сохранён.
                        if time.time() - _browser_restart[0] > 8:
                            _browser_restart[0] = time.time()
                            try:
                                await browser.close()
                            except Exception:
                                pass
                            try:
                                browser = await p.chromium.launch(**_launch_kwargs)
                                print("⚠️  Браузер этапа 2 перезапущен (Chromium упал) — прогон продолжается.")
                            except Exception:
                                pass
                        else:
                            await asyncio.sleep(1.0)
                        new_ctx = await browser.new_context(
                            user_agent=args.user_agent,
                            viewport={'width': 1440, 'height': 1000},
                            locale='ru-RU',
                        )
                    # v45: stealth — прячем признаки headless/автоматизации, чтобы FSA
                    # не отдавал капчу/блок. Навешиваем ДО первой навигации.
                    await _apply_stealth(new_ctx)
                    new_pg = await new_ctx.new_page()
                    # v45.1: прогрев сессии (заход на главную FSA) — ПО УМОЛЧАНИЮ ВЫКЛЮЧЕН.
                    # Главная FSA — тяжёлый Angular-SPA, который при загрузке делает десятки
                    # запросов к API. Когда 5 воркеров стартуют ОДНОВРЕМЕННО и каждый грузит
                    # главную — это залп из ~5 тяжёлых загрузок в первую же секунду, и FSA
                    # включает rate-limit СРАЗУ (блок «с порога»). Проверенное рабочее
                    # поведение (как в старых версиях) — идти СРАЗУ на страницу документа,
                    # без захода на главную. Прогрев можно включить флагом --fsa-warmup.
                    if _has_fsa and bool(getattr(args, 'fsa_warmup', False)):
                        await _fsa_warmup_context(
                            new_pg,
                            timeout_ms=int(getattr(args, 'registry_browser_timeout_ms', 30000)),
                        )
                    return new_ctx, new_pg

                context, page = await _fresh_context_and_page(None)
                processed_by_worker = 0
                _fsa_delay_lo, _fsa_delay_hi = _fsa_human_delay_range(args)
                # v39: ctx_refresh_every на 2 этапе — лечит постепенное «толстеющее» SPA-состояние
                ctx_refresh_every_stage2 = max(20, int(getattr(args, 'registry_ctx_refresh_every', 50)))
                try:
                    while True:
                        try:
                            url = await asyncio.wait_for(q.get(), timeout=1.0)
                        except asyncio.TimeoutError:
                            if q.empty():
                                break
                            continue
                        # v45.6: АВТО-ПАУЗА FSA. Если включён cooldown (FSA массово
                        # блокировал) — возвращаем FSA-ссылку в очередь и ждём, не
                        # трогая её, пока пауза не кончится. Не-FSA ссылки идут как обычно.
                        if (hostname(url) == 'pub.fsa.gov.ru'
                                and _FSA_COOLDOWN_UNTIL > time.time()):
                            await q.put(url)
                            q.task_done()
                            await asyncio.sleep(min(3.0, max(0.5, _FSA_COOLDOWN_UNTIL - time.time())))
                            continue
                        # v46: МЕДЛЕННЫЙ режим — ФСА реально ПО ОДНОМУ. Если ФСА сейчас
                        # занят другим воркером, НЕ ждём в очереди (иначе все 5 воркеров
                        # встанут на ФСА залпом) — возвращаем ссылку и берём НЕ-ФСА (SWIS
                        # и пр. идут параллельно). Так ФСА строго последовательный, без залпа.
                        if (_fsa_slow_mode and hostname(url) == 'pub.fsa.gov.ru'
                                and _fsa_serial_sem.locked()):
                            await q.put(url)
                            q.task_done()
                            await asyncio.sleep(0.4)
                            continue
                        active[f'w{wid}'] = (url, time.time())
                        cert = prod = typ = detail = ''
                        doc_status = ''  # v39.5
                        # v39: жёсткий timeout на один реестр. До этого если playwright/FSA
                        # подвисал на bizarre input — worker мог стоять бесконечно.
                        # v46: для ФСА таймаут КОРОЧЕ (≤60с): без захода на доп. вкладки
                        # один документ — это ~1 загрузка + перехват API (≤30с). 100-секундные
                        # зависания на больших прогонах душили прогресс и валили watchdog.
                        _h0 = hostname(url)
                        if _h0 == 'pub.fsa.gov.ru':
                            per_registry_timeout = max(
                                30, min(60, int(getattr(args, 'registry_browser_timeout_ms', 30000) / 1000) * 2 + 10))
                        else:
                            per_registry_timeout = max(
                                30,
                                int(getattr(args, 'registry_browser_timeout_ms', 30000) / 1000) * 3 + 30,
                            )
                        try:
                            h = hostname(url)
                            if h == 'pub.fsa.gov.ru':
                                # v46: МЕДЛЕННЫЙ РЕЖИМ — держим темп ниже лимита ФСА.
                                # Пауза между документами + АДАПТИВНЫЙ бэкофф (после каждой
                                # авто-паузы темп снижается ×(1+циклы)). Парсинг ФСА —
                                # СЕРИЙНО (Semaphore 1), чтобы не было всплеска из 4 сессий.
                                if _fsa_slow_mode:
                                    async with _fsa_serial_sem:
                                        # РЕАЛЬНАЯ пауза МЕЖДУ запросами к ФСА (а не параллельно
                                        # в каждом воркере). Множитель самонастраивается:
                                        # 1.0 при чистом темпе, растёт при сбоях, спадает при успехах.
                                        _gap = random.uniform(_fsa_slow_lo, _fsa_slow_hi) * _FSA_SLOW_MULT
                                        _wait = (_fsa_last_req[0] + _gap) - time.time()
                                        if _wait > 0:
                                            await asyncio.sleep(_wait)
                                        _fsa_last_req[0] = time.time()
                                        cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                            _parse_fsa_with_existing_page_v386(page, url, args),
                                            timeout=per_registry_timeout,
                                        )
                                else:
                                    # обычный режим: человекоподобная пауза + бутстрап-замок кук
                                    if _fsa_delay_hi > 0:
                                        await asyncio.sleep(random.uniform(_fsa_delay_lo, _fsa_delay_hi))
                                    if (bool(getattr(args, 'fsa_cookie_http', True))
                                            and not _FSA_SESSION_COOKIES):
                                        async with _fsa_bootstrap_sem:
                                            cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                                _parse_fsa_with_existing_page_v386(page, url, args),
                                                timeout=per_registry_timeout,
                                            )
                                    else:
                                        cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                            _parse_fsa_with_existing_page_v386(page, url, args),
                                            timeout=per_registry_timeout,
                                        )
                            elif h in {'swis.trade.kg', 'trade.kg'}:
                                # v27.7: киргизский SWIS — ТОЛЬКО HTTP (требование).
                                # Страница серверного рендеринга, браузер не нужен.
                                cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                    _parse_swis_http_v277(session, url, args),
                                    timeout=per_registry_timeout,
                                )
                            elif h in _BELGISS_EAEU_HOSTS:
                                # v27.9.x: BelGISS/ЕАЭС по требованию НЕ парсим —
                                # оставляем только ссылку. Экономит ~2-4с на документ.
                                cert, prod, typ, doc_status, detail = '', '', '', '', 'belgiss_link_only'
                            else:
                                cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                    _parse_other_registry_http_v386(session, url, args),
                                    timeout=per_registry_timeout,
                                )
                        except asyncio.TimeoutError:
                            detail = f'registry_hard_timeout_{per_registry_timeout}s'
                            stats['errors'] += 1
                            # Контекст может быть в неконсистентном состоянии — пересоздаём
                            context, page = await _fresh_context_and_page(context)
                        except Exception as e:
                            detail = f'registry_worker_error={type(e).__name__}: {str(e)[:180]}'
                            stats['errors'] += 1
                            context, page = await _fresh_context_and_page(context)
                        finally:
                            # v39.5: добавили doc_status в parsed
                            parsed[url] = (cert or '', prod or '', typ or '', doc_status or '', detail or '')
                            stats['done'] += 1
                            if prod:
                                stats['ok'] += 1
                            else:
                                stats['empty'] += 1
                            # v39.2: трекаем сетевые сбои FSA отдельно
                            if hostname(url) == 'pub.fsa.gov.ru':
                                stats['fsa_done'] += 1
                                if 'NETWORK_FAILURE_all_goto_failed' in (detail or ''):
                                    stats['fsa_network_failures'] += 1
                                # v45.6: АВТО-ВОССТАНОВЛЕНИЕ. Считаем неудачи подряд (пусто/
                                # сетевая ошибка/таймаут). Успех — сбрасываем счётчик. Когда
                                # неудач подряд накопилось много — ставим FSA на паузу
                                # (cooldown), чистим отравленную сессию (куки), и после паузы
                                # FSA пробуется заново. Это снимает временный rate-limit сам.
                                _cd_base = float(getattr(args, 'fsa_cooldown_sec', 90.0) or 0)
                                _cd_fails = max(2, int(getattr(args, 'fsa_cooldown_fails', 8)))
                                _cd_max = max(0, int(getattr(args, 'fsa_max_cooldowns', 3)))
                                if prod:
                                    _FSA_CONSEC_FAILS = 0
                                    # v46: самонастройка медленного режима — череда успехов
                                    # => осторожно ускоряемся (множитель паузы к базовому).
                                    if _fsa_slow_mode:
                                        _FSA_SLOW_OK += 1
                                        if _FSA_SLOW_OK >= 20 and _FSA_SLOW_MULT > 1.0:
                                            _FSA_SLOW_MULT = max(1.0, _FSA_SLOW_MULT - 0.3)
                                            _FSA_SLOW_OK = 0
                                else:
                                    _FSA_CONSEC_FAILS += 1
                                    # v46: сбой ФСА в медл. режиме => тормозим заранее (ещё до
                                    # полной авто-паузы), увеличивая паузу между запросами.
                                    if _fsa_slow_mode:
                                        _FSA_SLOW_MULT = min(5.0, _FSA_SLOW_MULT + 0.4)
                                        _FSA_SLOW_OK = 0
                                    if (_cd_base > 0 and _FSA_CONSEC_FAILS >= _cd_fails
                                            and _FSA_COOLDOWN_UNTIL <= time.time()
                                            and _FSA_COOLDOWN_CYCLES < _cd_max):
                                        _FSA_COOLDOWN_CYCLES += 1
                                        _dur = min(900.0, _cd_base * (2 ** (_FSA_COOLDOWN_CYCLES - 1)))
                                        _FSA_COOLDOWN_UNTIL = time.time() + _dur
                                        _FSA_CONSEC_FAILS = 0
                                        _FSA_SESSION_COOKIES = {}  # сбросить отравленную сессию
                                        print("=" * 80)
                                        print(f"⏸  FSA массово блокирует ({_cd_fails} неудач подряд). "
                                              f"АВТО-ПАУЗА {int(_dur)}с — попытка {_FSA_COOLDOWN_CYCLES}/{_cd_max} "
                                              f"снять rate-limit. Остальные реестры (SWIS и др.) продолжают идти.")
                                        print(f"   После паузы FSA пробуется заново со свежей сессией.")
                                        print("=" * 80)
                                    elif (_cd_base > 0 and _FSA_CONSEC_FAILS >= _cd_fails
                                          and _FSA_COOLDOWN_CYCLES >= _cd_max
                                          and not stats.get('fsa_gaveup_shown')):
                                        stats['fsa_gaveup_shown'] = True
                                        print(f"🔴 FSA не восстановился после {_cd_max} авто-пауз — похоже на сетевой "
                                              f"бан IP. Недобранные FSA добей кнопкой «Повторить упавшие FSA» "
                                              f"(после смены сети / паузы).")
                                # Раннее предупреждение: если первые 10 FSA — все network failure,
                                # FSA недоступен в принципе и тратить время дальше бессмысленно
                                if (not stats['fsa_warning_shown']
                                        and stats['fsa_done'] >= 10
                                        and stats['fsa_network_failures'] >= 10):
                                    stats['fsa_warning_shown'] = True
                                    print("=" * 80)
                                    print("🔴 ВНИМАНИЕ: ВСЕ первые 10 FSA-реестров вернули СЕТЕВУЮ ОШИБКУ.")
                                    print("   Это значит pub.fsa.gov.ru недоступен с твоего ПК.")
                                    print("   Проверь: открывается ли в обычном Chrome:")
                                    print("   https://pub.fsa.gov.ru/rss/certificate/view/3418716/baseInfo")
                                    print("   Если нет — попробуй: VPN с РФ-IP / мобильный интернет / DNS 8.8.8.8 /")
                                    print("   позвони провайдеру (может блокировать gov.ru).")
                                    print("   Скрипт продолжит работу с SWIS, FSA пропустит.")
                                    print("=" * 80)
                            # v39.1: СРАЗУ записать строки в out_store, чтобы файл рос
                            # по ходу прогона, а не появлялся только в самом конце
                            try:
                                await _flush_url_to_store(url, cert or '', prod or '', typ or '', doc_status or '', detail or '')
                            except Exception as e:
                                print(f"⚠️  flush_url_to_store ошибка: {type(e).__name__}: {e}")
                            active.pop(f'w{wid}', None)
                            processed_by_worker += 1
                            # v39: чаще пересоздаём context на 2 этапе (50 вместо 80).
                            # FSA-страницы тяжелее WB-карточек, память растёт быстрее.
                            if processed_by_worker % ctx_refresh_every_stage2 == 0:
                                if getattr(args, 'verbose_each', False) or processed_by_worker >= ctx_refresh_every_stage2:
                                    print(f"[reg-w{wid}] context refresh (обработано {processed_by_worker} реестров)")
                                context, page = await _fresh_context_and_page(context)
                            q.task_done()
                finally:
                    try:
                        await context.close()
                    except Exception:
                        pass

            # v39.8: watchdog 2 этапа — мониторит прогресс и перезапускает воркеры если зависли.
            # На 1 этапе watchdog есть, на 2 этапе раньше его не было: если все воркеры зависнут
            # на одном медленном реестре, прогон стоит. Теперь не стоит.
            stall_restart_sec_stage2 = max(120, int(getattr(args, 'registry_stall_restart_sec', 180) or 0))
            restart_count = [0]  # mutable

            async def watchdog_loop():
                """Раз в 10 секунд проверяет: меняется ли stats['done']. Если не меняется
                дольше stall_restart_sec_stage2 — отменяет воркеры, чистит active,
                и они стартуют заново."""
                last_done = stats['done']
                last_change = time.time()
                nonlocal workers
                try:
                    while stats['done'] < len(unique_urls):
                        await asyncio.sleep(10)
                        # v46: во время АВТО-ПАУЗЫ FSA (cooldown) прогресс намеренно стоит —
                        # это не зависание, поэтому watchdog не должен дёргать рестарт.
                        if _FSA_COOLDOWN_UNTIL > time.time():
                            last_change = time.time()
                            continue
                        current = stats['done']
                        if current != last_done:
                            last_done = current
                            last_change = time.time()
                            continue
                        stuck_for = time.time() - last_change
                        if stuck_for > stall_restart_sec_stage2:
                            restart_count[0] += 1
                            stuck_workers = list(active.items())[:10]
                            print("=" * 80)
                            print(f"⚠️  WATCHDOG-2 #{restart_count[0]}: нет прогресса {int(stuck_for)}с (порог {stall_restart_sec_stage2}с)")
                            print(f"   Прогресс: {stats['done']}/{len(unique_urls)} реестров")
                            emit_progress("registry", stats['done'], len(unique_urls))
                            print(f"   Зависшие: {stuck_workers}")
                            print(f"   Действие: cancel воркеров + перезапуск")
                            print("=" * 80)
                            # Отменяем все воркеры
                            for w in workers:
                                if not w.done():
                                    w.cancel()
                            # Ждём завершения отмены
                            await asyncio.gather(*workers, return_exceptions=True)
                            active.clear()
                            # Стартуем новые воркеры
                            workers = [asyncio.create_task(worker(i + 1)) for i in range(browser_workers)]
                            last_done = stats['done']
                            last_change = time.time()
                            print(f"✓ WATCHDOG-2: {browser_workers} воркеров перезапущены (всего рестартов: {restart_count[0]})")
                except asyncio.CancelledError:
                    return

            if stall_restart_sec_stage2 > 0:
                print(f"🛡  Watchdog 2 этапа активен: рестарт воркеров при отсутствии прогресса > {stall_restart_sec_stage2}с")
            watchdog_task = asyncio.create_task(watchdog_loop())

            workers = [asyncio.create_task(worker(i + 1)) for i in range(browser_workers)]

            # Ждём пока все реестры обработаны ИЛИ все воркеры умерли
            while stats['done'] < len(unique_urls):
                await asyncio.sleep(1.0)
                if all(w.done() for w in workers):
                    # Воркеры могли быть перезапущены watchdog'ом — проверим через 1 сек
                    await asyncio.sleep(1.0)
                    if all(w.done() for w in workers):
                        # Реально все умерли. Если очередь не пуста — лог и выход
                        if not q.empty():
                            print(f"⚠️  Все воркеры завершились, но в очереди ещё {q.qsize()} реестров")
                        break

            # Отменяем watchdog
            watchdog_task.cancel()
            try:
                await watchdog_task
            except asyncio.CancelledError:
                pass
            # Ждём завершения воркеров (они должны сами выйти когда очередь пуста)
            await asyncio.gather(*workers, return_exceptions=True)

            # v27.9.x: ВТОРОЙ ПРОХОД для FSA-ссылок, упавших по сетевой ошибке.
            # По умолчанию ВЫКЛючен — запускается ТОЛЬКО по кнопке в окне
            # (--registry-fsa-retry true). Раньше он шёл автоматически и «висел»
            # после 100%, когда FSA недоступен. Теперь повтор — осознанное действие
            # пользователя (когда FSA снова заработает).
            if not bool(getattr(args, 'registry_fsa_retry', False)):
                # считаем и совсем не извлечённые, и ЧАСТИЧНЫЕ (есть название, но нет
                # номера/статуса) — кнопка «Повторить упавшие FSA» дозаберёт и те, и те.
                _fsa_fail_n = sum(1 for u, v in parsed.items()
                                  if hostname(u) == 'pub.fsa.gov.ru'
                                  and (not (v[1] or '').strip()
                                       or not (v[0] or '').strip()
                                       or not (v[3] or '').strip()))
                if _fsa_fail_n:
                    print(f"ℹ️  {_fsa_fail_n} FSA-документ(ов) извлеклись не полностью (нет названия/номера/статуса). "
                          f"Когда FSA снова заработает — нажми в окне «🔁 Повторить упавшие FSA» (дозаберёт).")
            try:
                if not bool(getattr(args, 'registry_fsa_retry', False)):
                    raise _SkipSecondPass()
                def _is_transient_fsa_fail(detail_str: str) -> bool:
                    d = detail_str or ''
                    return any(k in d for k in (
                        'ERR_CONNECTION_TIMED_OUT', 'ERR_TIMED_OUT', 'ERR_CONNECTION',
                        'ERR_NETWORK', 'ERR_ABORTED', 'registry_hard_timeout',
                        'NETWORK_FAILURE_all_goto_failed', 'TimeoutError',
                    ))

                def _fsa_partial(val) -> bool:
                    # v46: ЧАСТИЧНО извлечённый документ — название продукции есть,
                    # но не добрались номер ИЛИ статус (медленная загрузка/таймаут).
                    # Такие тоже дозабираем повторным заходом.
                    return bool((val[1] or '').strip()) and (
                        not (val[0] or '').strip() or not (val[3] or '').strip())

                retry_urls = [
                    u for u, val in list(parsed.items())
                    if hostname(u) == 'pub.fsa.gov.ru'
                    and (
                        (not (val[1] or '').strip() and _is_transient_fsa_fail(val[4]))  # совсем не извлеклось
                        or _fsa_partial(val)  # частично: есть название, но нет номера/статуса
                    )
                ]
                # v27.9.x: РАЗЛИЧАЕМ «FSA недоступен» и «FSA нестабилен».
                #  • если НИ ОДНА FSA-ссылка не извлеклась (fsa_ok==0) — host реально
                #    недоступен в этой сети, повтор бесполезен → пропускаем;
                #  • если что-то ИЗВЛЕКЛОСЬ (fsa_ok>0) — соединение просто нестабильное
                #    (флапает), и повтор упавших обычно частично спасает → делаем его,
                #    но в рамках жёсткого таймбюджета и с ранним обрывом.
                fsa_total = sum(1 for u in parsed if hostname(u) == 'pub.fsa.gov.ru')
                fsa_ok = sum(1 for u, v in parsed.items()
                             if hostname(u) == 'pub.fsa.gov.ru' and (v[1] or '').strip())
                if fsa_total and fsa_ok == 0:
                    print(f"🔁 Второй проход FSA ПРОПУЩЕН: ни одна из {fsa_total} FSA-ссылок не извлеклась — "
                          f"похоже pub.fsa.gov.ru недоступен в этой сети (VPN с РФ-IP / мобильный интернет). "
                          f"Повтор не поможет; ссылки сохранены — перезапустите позже.")
                    retry_urls = []
                max_retry = int(getattr(args, 'registry_fsa_retry_max', 80) or 0)
                if retry_urls and max_retry > 0:
                    retry_urls = retry_urls[:max_retry]
                    print("=" * 80)
                    print(f"🔁 Второй проход FSA: повтор {len(retry_urls)} ссылок (упавшие + частично извлечённые: нет номера/статуса)")
                    print("=" * 80)
                    per_registry_timeout = max(
                        30, int(getattr(args, 'registry_browser_timeout_ms', 30000) / 1000) * 3 + 30)
                    rctx = await browser.new_context(
                        user_agent=args.user_agent, viewport={'width': 1440, 'height': 1000}, locale='ru-RU')
                    rpage = await rctx.new_page()
                    recovered = 0
                    # v27.9.x: жёсткий ТАЙМБЮДЖЕТ на весь второй проход + ранний обрыв,
                    # чтобы после 100% окно НЕ «висело» минутами и графики прогрузились.
                    retry_budget_sec = float(getattr(args, 'registry_fsa_retry_budget_sec', 150) or 150)
                    retry_started = time.time()
                    consecutive_fail = 0
                    try:
                        for _i, ru in enumerate(retry_urls, 1):
                            if time.time() - retry_started > retry_budget_sec:
                                print(f"🔁 Второй проход FSA остановлен по таймбюджету "
                                      f"{retry_budget_sec:.0f}с ({_i-1}/{len(retry_urls)} обработано, "
                                      f"восстановлено {recovered})")
                                break
                            if consecutive_fail >= 8:
                                print("🔁 Второй проход FSA прерван: 8 повторов подряд без результата "
                                      "(FSA всё ещё недоступен). Ссылки сохранены — перезапустите позже.")
                                break
                            cert = prod = typ = doc_status = detail = ''
                            try:
                                cert, prod, typ, doc_status, detail = await asyncio.wait_for(
                                    _parse_fsa_with_existing_page_v386(rpage, ru, args),
                                    timeout=per_registry_timeout,
                                )
                            except asyncio.TimeoutError:
                                detail = 'fsa_retry_hard_timeout'
                                try:
                                    await rctx.close()
                                except Exception:
                                    pass
                                rctx = await browser.new_context(
                                    user_agent=args.user_agent, viewport={'width': 1440, 'height': 1000}, locale='ru-RU')
                                rpage = await rctx.new_page()
                            except Exception as e:
                                detail = f'fsa_retry_error={type(e).__name__}'
                            if (prod or '').strip():
                                # заменяем устаревшие (пустые) строки этого url на свежие
                                cu = clean_url(ru)
                                async with out_store.lock:
                                    out_store.rows = [r for r in out_store.rows
                                                      if clean_url(r.registry_url) != cu]
                                parsed[ru] = (cert or '', prod or '', typ or '',
                                              doc_status or '', (detail or '') + ';fsa_retry_ok')
                                await _flush_url_to_store(
                                    ru, cert or '', prod or '', typ or '', doc_status or '',
                                    (detail or '') + ';fsa_retry_ok')
                                recovered += 1
                                consecutive_fail = 0
                            else:
                                consecutive_fail += 1
                            # держим живым прогресс/лог, чтобы окно не казалось зависшим
                            print(f"🔁 Второй проход FSA: {_i}/{len(retry_urls)}, восстановлено {recovered}")
                            emit_progress("registry", stats['done'], len(unique_urls))
                    finally:
                        try:
                            await rctx.close()
                        except Exception:
                            pass
                    print(f"🔁 Второй проход FSA: восстановлено {recovered}/{len(retry_urls)}")
                    try:
                        await out_store.save()
                    except Exception:
                        pass
            except _SkipSecondPass:
                pass
            except Exception as _e:
                print(f"⚠️  Второй проход FSA пропущен: {type(_e).__name__}: {_e}")

            await browser.close()
            if restart_count[0] > 0:
                print(f"🛡  Watchdog 2 этапа сработал {restart_count[0]} раз за прогон")

    progress_task.cancel()
    saver_task.cancel()
    for t in (progress_task, saver_task):
        try:
            await t
        except asyncio.CancelledError:
            pass

    # v39.1: данные уже в out_store (flush сразу после парсинга каждого реестра).
    # Здесь только страхуем: если по каким-то url не было flush (например, не дошло
    # из-за внезапного выхода) — допишем их сейчас, и сохраним финал.
    written_urls = {clean_url(r.registry_url) for r in out_store.rows}
    missed_count = 0
    for row in rows:
        url = clean_url(row.get('registry_url', ''))
        if url in written_urls:
            continue
        # Этот url не был записан — допишем как «not_parsed»
        # v39.5: unpack 5-tuple (cert, prod, typ, doc_status, detail)
        cert, prod, typ, doc_status, parse_detail = parsed.get(url, ('', '', '', '', 'not_parsed'))
        _host = hostname(url)
        if _host in _BELGISS_EAEU_HOSTS:
            # v27.9.x: BelGISS/ЕАЭС не парсим — только ссылка.
            verdict, score, cmp_details = STATUS_LINK_COLLECTED, 0.0, 'belgiss_link_only'
        else:
            verdict, score, cmp_details = compare_product_names(
                row.get('product_name', ''), prod, brand=row.get('brand', ''), subject=row.get('subject', ''),
                doc_status=doc_status,
            )
            if not prod and verdict != 'НЕДЕЙСТВУЮЩИЙ ДОКУМЕНТ':
                verdict = 'НЕ УДАЛОСЬ ИЗВЛЕЧЬ НАЗВАНИЕ ИЗ РЕЕСТРА'
                score = 0.0
        # v46: киргизский документ из таблицы статусов РФ — отдельный вердикт
        rf_status = kg_rf_status_text(cert)
        if rf_status:
            verdict = STATUS_INVALID_IN_RF
        # v39.14: достаём расширенные поля FSA из глобального кэша
        ext = _FSA_EXTENDED_FIELDS_CACHE.get(url, {})
        rr = ResultRow(
            query=row.get('query', ''),
            nm_id=safe_int(row.get('nm_id')),
            product_name=row.get('product_name', ''),
            brand=row.get('brand', ''),
            subject=row.get('subject', ''),
            product_url=row.get('product_url', ''),
            status=verdict,
            price_rub=safe_float(row.get('price_rub')),
            sale_price_rub=safe_float(row.get('sale_price_rub')),
            seller_name=row.get('seller_name', ''),
            is_original=row.get('is_original', ''),
            # пусто (старый CSV без признака) → «Нет»: бейдж не зафиксирован
            docs_verified=row.get('docs_verified') or 'Нет',
            supplier_id=row.get('supplier_id', ''),
            rating=safe_float(row.get('rating')),
            feedbacks=safe_int(row.get('feedbacks')),
            registry_url=url,
            registry_host=hostname(url),
            registry_record_id=extract_record_id(url),
            certificate_number=cert,
            document_type=typ,
            document_status=doc_status,
            rf_status=rf_status,
            certificate_product_name=prod,
            document_date_start=ext.get('document_date_start', ''),
            document_date_end=ext.get('document_date_end', ''),
            # v46: по требованию орг-поля и ТН ВЭД можно не собирать (по умолчанию не собираем).
            applicant_name=('' if bool(getattr(args, 'fsa_skip_org_fields', True)) else _clean_org_name(ext.get('applicant_name', ''))),
            applicant_inn=('' if bool(getattr(args, 'fsa_skip_org_fields', True)) else ext.get('applicant_inn', '')),
            manufacturer_name=('' if bool(getattr(args, 'fsa_skip_org_fields', True)) else _clean_org_name(ext.get('manufacturer_name', ''))),
            tnved=('' if bool(getattr(args, 'fsa_skip_org_fields', True)) else _clean_tnved_code(ext.get('tnved', ''))),
            scheme=ext.get('scheme', ''),
            technical_regulation=ext.get('technical_regulation', ''),
            score=score,
            details=(f'registry_browser_stage; {parse_detail}; {cmp_details}' if prod else f'empty_registry_product_name; {parse_detail}'),
            checked_at=now_iso(),
        )
        await out_store.add(rr)
        missed_count += 1
    if missed_count:
        print(f"Допишу {missed_count} строк, которые не успели сохраниться при парсинге.")
    await out_store.save()
    print(
        f"Готово. Excel сохранён: {Path(args.output).resolve()}. "
        f"Извлечено названий по уникальным реестрам: {stats['ok']}/{len(unique_urls)}; пусто={stats['empty']}; ошибки={stats['errors']}"
    )
    # v45.2: сводка по быстрому HTTP-пути (куки браузера). Показывает, сколько
    # FSA-документов добыто лёгким HTTP вместо полной загрузки SPA.
    if _FSA_COOKIE_HTTP_OK > 0:
        print(f"⚡ FSA: {_FSA_COOKIE_HTTP_OK} документ(ов) добыто быстрым HTTP по кукам браузера "
              f"(без полной загрузки страницы — кратно быстрее и меньше запросов к FSA).")
    # v39.2: явная сводка по FSA, чтобы было видно — это сетевая проблема или парсинга
    if stats['fsa_done'] > 0:
        fsa_net_pct = 100.0 * stats['fsa_network_failures'] / stats['fsa_done']
        if stats['fsa_network_failures'] == stats['fsa_done']:
            print(f"🔴 FSA полностью недоступен с твоего ПК: {stats['fsa_network_failures']}/{stats['fsa_done']} реестров вернули сетевую ошибку (chrome-error://).")
            print(f"   Это НЕ ошибка кода — твой ПК не может подключиться к pub.fsa.gov.ru.")
            print(f"   Попробуй: VPN с РФ-IP / мобильный интернет / сменить DNS на 8.8.8.8.")
        elif fsa_net_pct > 30:
            print(f"⚠️  FSA сетевые сбои: {stats['fsa_network_failures']}/{stats['fsa_done']} ({fsa_net_pct:.0f}%). Возможно нестабильное соединение к pub.fsa.gov.ru.")

    # v25-reporting: финальный текстовый лог прогона
    try:
        _log_arg = getattr(args, "run_log", "") or ""
        _log_path = Path(_log_arg) if _log_arg else None
        _wd = int(getattr(args, "expiry_warning_days", 30) or 0)
        _written = _write_run_log_v39(list(out_store.rows), Path(args.output), _wd,
                                      _run_started_at, mode="Stage 2 (реестры)",
                                      log_path=_log_path)
        if _written:
            print(f"Лог прогона записан: {_written}")
    except Exception as _e:
        print(f"[run-log] ошибка: {_e}")


# -----------------------------
# CLI
# -----------------------------

def apply_speed_profile(args):
    if args.max_speed:
        args.headless = True
        args.block_assets = True
        # ВАЖНО: не делаем слишком ранний commit — иначе WB не рисует плашку "Документы проверены".
        args.goto_timeout_ms = min(args.goto_timeout_ms, 18000)
        args.default_timeout_ms = min(args.default_timeout_ms, 5000)
        args.after_goto_ms = max(args.after_goto_ms, 500)
        args.after_specs_click_ms = max(args.after_specs_click_ms, 650)
        args.after_docs_wait_ms = max(args.after_docs_wait_ms, 550)
        args.after_look_wait_ms = max(args.after_look_wait_ms, 1400)
        args.card_ready_timeout_ms = max(args.card_ready_timeout_ms, 3600)
        args.docs_timeout_ms = max(args.docs_timeout_ms, 4800)
        args.look_button_timeout_ms = max(args.look_button_timeout_ms, 3600)
        args.no_docs_fallback_ms = max(args.no_docs_fallback_ms, 2200)
        args.card_hard_timeout_ms = max(args.card_hard_timeout_ms, 30000)
        args.autosave_every = args.autosave_every or 25
        args.stall_restart_sec = min(max(getattr(args, "stall_restart_sec", 120), 60), 180)
        args.progress_interval_sec = min(args.progress_interval_sec, 20)
        args.max_card_retries = max(args.max_card_retries, 2)
        args.retry_missing = max(args.retry_missing, 1)
        args.context_refresh_every = min(max(args.context_refresh_every, 40), 120)
        # v39: при max-speed гарантируем что browser_restart_every разумный
        if not getattr(args, 'browser_restart_every', None):
            args.browser_restart_every = 600
        args.browser_restart_every = min(max(args.browser_restart_every, 200), 1200)
    return args

def build_parser():
    ap = argparse.ArgumentParser()
    ap.add_argument("--query", default="", help="Поисковый запрос WB")
    ap.add_argument("--query-profile", default="auto",
                    help="v39.7: домен товаров для умной генерации запросов. auto = определяется по запросу. "
                         "Можно несколько через запятую (напр. clothing,shoes). Поддержка: clothing/одежда, "
                         "shoes/обувь, toys/игрушки, kids_accessories, baby_gear, cosmetics/косметика, "
                         "electronics/электроника, appliances/бытовая техника, home/дом, kitchenware/посуда, food/продукты.")
    ap.add_argument("--strict-domain-filter", type=lambda s: s.lower() in ('1','true','yes','y','on'),
                    default=True,
                    help="v39.7: фильтровать карточки по subject-id WB строго в рамках домена. Лечит ситуацию когда по запросу «детские игрушки» приходят платья. true/false (default true).")
    ap.add_argument("--limit", type=int, default=1000)
    ap.add_argument("--input-csv", default="")
    ap.add_argument("--input-links-csv", default="")
    ap.add_argument("--output", default="result.xlsx")
    ap.add_argument("--output-links-csv", default="registry_links.csv")

    ap.add_argument("--link-only", type=str_to_bool, default=False)
    ap.add_argument("--max-speed", type=str_to_bool, default=False)
    ap.add_argument("--auto-expand", type=str_to_bool, default=True)
    ap.add_argument("--max-expanded-queries", type=int, default=250)
    ap.add_argument("--per-query-limit", type=int, default=250)
    ap.add_argument("--collect-workers", type=int, default=24)
    ap.add_argument("--search-sorts", default="popular,rate,newly,benefit")
    # v27.5: проверка плашки «Оригинальный товар» через wb_enhanced. По умолчанию включено.
    ap.add_argument("--check-original", type=str_to_bool, default=True,
                    help="Проверять плашку «Оригинальный товар» через HTML-страницу WB и basket card.json (wb_enhanced).")
    ap.add_argument("--dump-viewflags", type=str_to_bool, default=True,
                    help="v46: диагностика — сохранять viewFlags собранных карточек в wb_viewflags.csv "
                         "(для вычисления бита «Документы проверены»). Лёгкий CSV, не влияет на скорость.")
    ap.add_argument("--check-docs-verified", type=str_to_bool, default=True,
                    help="v46: собирать бейдж «Документ проверен WB» (Да/Нет) из card.json "
                         "(certificate.verified). Отдельный лёгкий HTTP-проход по basket-CDN.")
    ap.add_argument("--check-original-workers", type=int, default=20,
                    help="Сколько параллельных воркеров для проверки оригинальности. v27.7: поднято 10→20 — после сведения проверки к одному домену (ru) нагрузка на карточку упала ~4×, можно больше параллелизма.")
    ap.add_argument("--check-original-domains", default="ru",
                    help="Доменные зоны WB для HTML-проверки плашки «Оригинал», через запятую. По умолчанию только 'ru' (быстро). Можно 'ru,by,kg,ge' для максимального покрытия (медленнее ~4×).")

    ap.add_argument("--browser-count", type=int, default=2, help="Сколько процессов Chromium держать одновременно. Воркеры распределяются между ними как контексты/страницы")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--headless", type=str_to_bool, default=True)
    ap.add_argument("--block-assets", type=str_to_bool, default=True)
    ap.add_argument("--viewport-width", type=int, default=1536)
    ap.add_argument("--viewport-height", type=int, default=900)
    ap.add_argument("--user-agent", default="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

    ap.add_argument("--goto-timeout-ms", type=int, default=20000)
    ap.add_argument("--default-timeout-ms", type=int, default=6000)
    ap.add_argument("--card-hard-timeout-ms", type=int, default=35000)
    ap.add_argument("--max-card-retries", type=int, default=2, help="Повторить карточку при Page crashed/timeout/error до N раз перед финальным статусом")
    ap.add_argument("--retry-missing", type=int, default=1, help="Повторить карточку при НЕТ ДОКУМЕНТОВ/НЕТ ССЫЛКИ до N раз в строгом режиме")
    ap.add_argument("--context-refresh-every", type=int, default=80, help="Пересоздавать контекст каждого воркера каждые N обработанных карточек, чтобы не падал Chromium")
    ap.add_argument("--browser-restart-every", type=int, default=600, help="v39: полностью убить и пересоздать процесс Chromium каждые N карточек на воркера. 0 = выключено. Главное лекарство от утечки памяти, из-за которой программа замирает после ~1500 карточек.")
    ap.add_argument("--after-goto-ms", type=int, default=600)
    ap.add_argument("--after-specs-click-ms", type=int, default=700)
    ap.add_argument("--after-docs-wait-ms", type=int, default=700)
    ap.add_argument("--after-look-wait-ms", type=int, default=1600)
    ap.add_argument("--card-ready-timeout-ms", type=int, default=3800)
    ap.add_argument("--docs-timeout-ms", type=int, default=5200)
    ap.add_argument("--look-button-timeout-ms", type=int, default=3800)
    ap.add_argument("--no-docs-fallback-ms", type=int, default=2400)

    ap.add_argument("--autosave-every", type=int, default=25)
    ap.add_argument("--stall-autosave-sec", type=int, default=90)
    ap.add_argument("--stall-restart-sec", type=int, default=120, help="Если прогресс не меняется N секунд, перезапустить браузеры/воркеры и вернуть active-карточки в очередь")
    ap.add_argument("--stuck-report-sec", type=int, default=45)
    ap.add_argument("--progress-interval-sec", type=int, default=15)
    ap.add_argument("--resume", type=str_to_bool, default=True)
    ap.add_argument("--reset-output", type=str_to_bool, default=False, help="Удалить старые output/output-links-csv перед запуском. Удобно для чистых тестов.")
    ap.add_argument("--verbose-each", type=str_to_bool, default=False)
    ap.add_argument("--print-links", type=str_to_bool, default=False)
    ap.add_argument("--trace", type=str_to_bool, default=False)

    ap.add_argument("--http-workers", type=int, default=120)
    ap.add_argument("--http-connector-limit", type=int, default=400)
    ap.add_argument("--save-every", type=int, default=1000)
    ap.add_argument("--ok-score", type=float, default=45.0)
    ap.add_argument("--registry-http-timeout", type=float, default=30.0, help="HTTP timeout на один запрос к реестру, секунд")
    ap.add_argument("--registry-browser-fallback", type=str_to_bool, default=False, help="Дополнительный общий browser fallback для реестров")
    ap.add_argument("--registry-browser-only", type=str_to_bool, default=False, help="Открывать поддерживаемые реестры браузером и извлекать данные только из видимой страницы; для ФСА и SWIS это самый точный режим, но медленнее HTTP")
    ap.add_argument("--fsa-exact-browser-fallback", type=str_to_bool, default=True, help="Для ФСА по умолчанию открыть /baseInfo|/common и /product браузером, если HTTP не достал точные поля")
    ap.add_argument("--registry-browser-workers", type=int, default=6, help="Сколько Chromium одновременно можно использовать для второго этапа/ФСА browser fallback")
    ap.add_argument("--registry-headless", type=str_to_bool, default=True)
    ap.add_argument("--registry-fsa-retry", type=str_to_bool, default=False,
                    help="v27.9.x: второй проход по упавшим FSA-ссылкам. По умолчанию FALSE — "
                         "запускается ТОЛЬКО по кнопке в окне (когда FSA снова доступен).")
    ap.add_argument("--fsa-human-delay-ms", default="300,1400",
                    help="v45: случайная человекоподобная пауза между документами FSA, мс, в формате "
                         "«min,max» (по умолчанию 300,1400). Снижает риск блокировки за слишком ровный "
                         "автоматический темп. 0,0 — без паузы (быстрее, но рискованнее).")
    ap.add_argument("--kg-rf-status-file", default="",
                    help="v46: путь к таблице (xlsx/csv) статусов КИРГИЗСКИХ документов на территории "
                         "РФ (колонки number + id_status_in_rf: 14=прекращён, 15=приостановлен). Если не "
                         "задан, ищется kg_rf_status.xlsx рядом с программой. Совпавшие по номеру "
                         "документы получают колонку «Статус на территории РФ» и вердикт «НЕДЕЙСТВУЕТ В РФ».")
    ap.add_argument("--fsa-slow-mode", type=str_to_bool, default=False,
                    help="v46: МЕДЛЕННЫЙ режим ФСА для больших прогонов (до 10k) без блокировок. "
                         "ФСА парсится строго ПО ОДНОМУ документу с паузой (--fsa-slow-delay-ms) и "
                         "адаптивным бэкоффом. SWIS/прочие реестры идут параллельно. Медленно (часы для "
                         "10k), но ФСА не банит IP.")
    ap.add_argument("--fsa-slow-delay-ms", default="2000,3500",
                    help="v46: пауза между документами ФСА в медленном режиме, мс, «min,max» "
                         "(по умолчанию 2000,3500 ≈ 22 док/мин). Меньше — быстрее, но выше риск бана; "
                         "при блокировках пауза сама растёт (адаптивный бэкофф).")
    ap.add_argument("--fsa-skip-org-fields", type=str_to_bool, default=True,
                    help="v46: НЕ собирать заявителя/изготовителя/ИНН/ТН ВЭД из ФСА. По умолчанию TRUE: "
                         "это убирает заход на ДОПОЛНИТЕЛЬНЫЕ вкладки ФСА (где лежат изготовитель/заявитель) "
                         "— меньше запросов к ФСА и ниже риск блокировки. Номер, статус, даты, схема, "
                         "техрегламент и название продукции собираются как обычно.")
    ap.add_argument("--fsa-cookie-http", type=str_to_bool, default=False,
                    help="v45.8: ПО УМОЛЧАНИЮ FALSE — ФСА идёт ТОЛЬКО через браузер (по требованию: "
                         "другие способы пока не помогают). Браузерные документы парсятся ПАРАЛЛЕЛЬНО, "
                         "как раньше. true — включить быстрый HTTP-путь по кукам браузера (первый "
                         "документ через браузер отдаёт куки, остальные тянутся лёгким HTTP); при "
                         "включении первые документы идут по одному (бутстрап кук).")
    ap.add_argument("--fsa-warmup", type=str_to_bool, default=False,
                    help="v45.1: прогрев сессии заходом на главную pub.fsa.gov.ru перед документами. "
                         "ПО УМОЛЧАНИЮ FALSE: при старте нескольких воркеров одновременная загрузка "
                         "тяжёлой главной = залп запросов, и FSA блокирует сразу. Без прогрева (как в "
                         "старых рабочих версиях) браузер идёт прямо на документ — надёжнее.")
    ap.add_argument("--fsa-cooldown-sec", type=float, default=90.0,
                    help="v45.6: АВТО-ВОССТАНОВЛЕНИЕ FSA. Базовая пауза (сек), когда FSA начал "
                         "массово блокировать; с каждым разом удваивается (90→180→360…). 0 — выключить.")
    ap.add_argument("--fsa-cooldown-fails", type=int, default=8,
                    help="v45.6: сколько неудач FSA подряд включают авто-паузу.")
    ap.add_argument("--fsa-max-cooldowns", type=int, default=3,
                    help="v45.6: сколько раз максимум авто-паузить FSA за прогон. После — FSA "
                         "отпускается (недобранное добивается кнопкой «Повторить упавшие FSA»).")
    # v27.6: уменьшен с 45000 до 28000 — раньше FSA-карточка занимала 165с
    ap.add_argument("--registry-browser-timeout-ms", type=int, default=28000)
    ap.add_argument("--registry-browser-wait-ms", type=int, default=8000,
                    help="v27.6: уменьшен до 8000 — раньше 165с/карточку, хотим быстрее.")
    ap.add_argument("--registry-ctx-refresh-every", type=int, default=50, help="v39: пересоздавать контекст у каждого 2-этапного воркера каждые N реестров (лечит память на SPA-страницах FSA)")
    ap.add_argument("--registry-stall-restart-sec", type=int, default=180,
                    help="v39.8: watchdog 2 этапа. Если все воркеры не делают прогресса N сек — перезапуск. 0=выключено.")
    # v39.12: HTTP fast-path параметры
    ap.add_argument("--link-mode", default="http_only",
                    choices=["http_first", "http_only", "browser_only"],
                    help="v40: режим сбора ссылок. http_only (по умолчанию)=ТОЛЬКО certificate.json по HTTP, без браузера (наличие документа однозначно по json: есть json→документ есть, 404 везде→документа нет). http_first=HTTP + браузер на сетевые ошибки. browser_only=только браузер (старая логика).")
    ap.add_argument("--http-link-workers", type=int, default=30,
                    help="параллельность HTTP-запросов к certificate.json. Можно 30-100 — это лёгкие HTTP-запросы к CDN.")
    ap.add_argument("--fetch-sellers", type=lambda s: str(s).lower() in ('1','true','yes','y','on'),
                    default=True,
                    help="v40.3: дотягивать имена продавцов (seller_name) через card.wb.ru detail API. WB-поиск отдаёт только supplierId. true/false (default true).")
    ap.add_argument("--cert-timeout-sec", type=float, default=6.0,
                    help="таймаут одного запроса к basket-NN.wbbasket.ru/certificate.json")
    ap.add_argument("--cert-max-hosts", type=int, default=16,
                    help="сколько basket-шардов пробовать на карточку. v27.9.x: перебор стал ПАРАЛЛЕЛЬНЫМ, поэтому большее число шардов почти не стоит времени, но заметно повышает покрытие — особенно для товаров с высоким vol (новые nm_id), где границы шардов известны хуже. Дефолт поднят 8→16. Можно повышать до 30+ для максимальной надёжности.")
    ap.add_argument("--no-docs-confirm-404", type=int, default=0,
                    help="v40: устарел (логика теперь автоматическая: все честные 404 = нет документов, сетевая ошибка = повтор). Оставлен для совместимости команд.")
    # v39.13: HTTP fast-path для FSA-парсинга на 2 этапе
    ap.add_argument("--fsa-http-fast-path", type=lambda s: s.lower() in ('1','true','yes','y','on'),
                    default=False,
                    help="v27.7: по умолчанию FALSE — ФСА парсится ТОЛЬКО браузером (требование: для ФСА только браузерный путь). HTTP-fast-path через curl_cffi был быстрее, но менее надёжен/режется TLS-fingerprint'ом. Включи true, если хочешь рискнуть скоростью.")
    ap.add_argument("--fsa-curl-cffi-impersonate", default="chrome",
                    help="v39.13: профиль TLS-имперсонации curl_cffi: chrome / chrome120 / chrome110. Если основной даёт 403 — пробуются дополнительные.")
    # --- v25-reporting: отчётный слой ---
    ap.add_argument("--expiry-warning-days", type=int, default=30,
                    help="Сколько дней до окончания действия документа считать риском 'Скоро истекает'. По умолчанию 30. Не меняет технический статус, только подсветку и флаг в отчёте.")
    ap.add_argument("--make-report-xlsx", type=str_to_bool, default=True,
                    help="Добавлять в итоговый XLSX листы 'Сводка' и 'Подробности' с русскими заголовками и подсветкой по сроку. По умолчанию true.")
    # v25-reporting: strict brand filter (портирован из main_brand.py)
    ap.add_argument("--brand", default="",
                    help="Опциональный строгий бренд-фильтр. Если задан, оставляет только карточки/строки того же бренда.")
    ap.add_argument("--brand-match", default="any", choices=["any", "exact", "contains"],
                    help="Режим бренд-фильтра: any (выкл., по умолчанию) / exact / contains.")
    ap.add_argument("--run-log", default="",
                    help="Путь к текстовому логу прогона (run-summary). Если пусто — положится рядом с output как *_run.log.")
    return ap

async def main_async():
    parser = build_parser()
    args = parser.parse_args()
    args = apply_speed_profile(args)

    if args.input_links_csv:
        # v39.13: проверяем curl_cffi перед 2 этапом — без него FSA не парсится через HTTP
        # v39.15: функция встроена в этот файл, внешний модуль не нужен
        if bool(getattr(args, 'fsa_http_fast_path', True)):
            try:
                if is_curl_cffi_available():
                    print("✓ curl_cffi доступен — FSA-парсинг 2 этапа будет идти через HTTP (без браузера для FSA-документов).")
                else:
                    print("=" * 80)
                    print("⚠️  curl_cffi НЕ установлен. FSA-парсинг будет идти через медленный браузер.")
                    print("    Для ускорения в 10-30 раз: pip install curl_cffi")
                    print("=" * 80)
            except Exception:
                pass
        await run_registry_stage(args)
        return

    if not args.query and not args.input_csv:
        args.query = input("Введите поисковый запрос: ").strip()
    if not args.query and not args.input_csv:
        raise SystemExit("Не указан query или input-csv")

    await run_link_collection(args)

def main():
    try:
        asyncio.run(main_async())
    except KeyboardInterrupt:
        print("\nОстановлено пользователем. Уже сохранённые CSV/XLSX остаются на диске.")
    except Exception as e:
        print(f"Критическая ошибка: {type(e).__name__}: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
