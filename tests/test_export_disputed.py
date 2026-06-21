"""
Тест (улучшение №10, v53): выгрузка спорных строк отдельным Excel со ссылками.

Вместо листания 50 000 строк — файл «на проверку» только со спорными вердиктами
(ПРОВЕРИТЬ ВРУЧНУЮ + НЕСООТВЕТСТВИЕ), с кликабельными ссылками на карточку WB и
реестр + причиной. Проверяем фильтрацию, гиперссылки и пустой случай.

Запуск:  python3 tests/test_export_disputed.py
"""
import sys
import tempfile
import types
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.modules.setdefault("webview", types.ModuleType("webview"))

import wb_checker as w  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

RESULTS = []


def check(name, cond):
    RESULTS.append(bool(cond))
    print(f"{'PASS' if cond else 'FAIL'}  {name}")


def _make(path, statuses):
    wb = Workbook()
    ws = wb.active
    ws.title = "Подробности"
    ws.append(["Артикул", "Название товара", "Бренд", "Технический статус",
               "Название в реестре", "Примечания", "Ссылка на товар (WB)",
               "Ссылка на реестр"])
    for i, st in enumerate(statuses, 1):
        ws.append([100 + i, f"Товар {i}", "BR", st, "Рег",
                   "детали", f"https://wb.ru/catalog/{100+i}/detail.aspx",
                   f"https://pub.fsa.gov.ru/x{i}"])
    wb.save(str(path))


def main():
    b = w.Bridge()
    out_files = []
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "result.xlsx"
        _make(p, ["OK", "ПРОВЕРИТЬ ВРУЧНУЮ", "НЕСООТВЕТСТВИЕ", "OK",
                  "ДОКУМЕНТ НЕ ПРОВЕРЕН", "ПРОВЕРИТЬ ВРУЧНУЮ"])
        res = b.export_disputed(str(p))
        check("экспорт успешен", res.get("ok"))
        check("отобраны только спорные (3 из 6)", res.get("rows") == 3)
        if res.get("ok"):
            out_files.append(res["path"])
            ws = load_workbook(res["path"]).active
            statuses = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
            check("в файле только спорные вердикты",
                  set(statuses) <= {"ПРОВЕРИТЬ ВРУЧНУЮ", "НЕСООТВЕТСТВИЕ"})
            check("ссылка на карточку WB кликабельна (hyperlink)",
                  ws.cell(row=2, column=7).hyperlink is not None
                  and "wb.ru" in ws.cell(row=2, column=7).hyperlink.target)
            check("ссылка на реестр кликабельна (hyperlink)",
                  ws.cell(row=2, column=8).hyperlink is not None
                  and "fsa.gov.ru" in ws.cell(row=2, column=8).hyperlink.target)
            check("есть колонка «Причина»",
                  "Причина" in [c.value for c in ws[1]])
            check("шапка заморожена (freeze_panes)", ws.freeze_panes == "A2")

        # пустой случай: нет спорных
        p2 = Path(td) / "clean.xlsx"
        _make(p2, ["OK", "OK", "ДОКУМЕНТ НЕ ПРОВЕРЕН"])
        res2 = b.export_disputed(str(p2))
        check("нет спорных → ok=False с понятным сообщением",
              not res2.get("ok") and "спорных" in (res2.get("error") or "").lower())

    for f in out_files:
        try:
            Path(f).unlink()
        except Exception:
            pass

    # GUI-обвязка присутствует
    gui = (Path(__file__).resolve().parents[1] / "wb_checker.py").read_text(encoding="utf-8")
    check("кнопка «На проверку» есть в окне", "btn-export-disputed" in gui)
    check("обработчик вызывает export_disputed", "api.export_disputed(null)" in gui)


if __name__ == "__main__":
    main()
    p, t = sum(RESULTS), len(RESULTS)
    print("-" * 60)
    print(f"ИТОГО: {p}/{t} прошло")
    sys.exit(0 if p == t else 1)
